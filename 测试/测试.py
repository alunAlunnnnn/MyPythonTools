import openpyxl



data = r"D:\工作项目\项目_南京管线\111\DL_GGX_POINT.xlsx"

wb = openpyxl.load_workbook(data)
sht = wb.active

for eachRow in sht.rows:
    dataList = [eachCell.value for eachCell in eachRow]
    print(dataList)

maxRow = sht.max_row

for i in range(1, maxRow + 1):
    cell = sht.cell(i, 3).value
    print(cell)