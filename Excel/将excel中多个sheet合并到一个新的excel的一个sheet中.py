import openpyxl
import os


# 通过excel赋值到属性表
def _mergeExcel(xls):
    """
    :usage: 将excel文件中的多个sheet合并到一个新的excel中
    :param xls: 输入excel文件路径，string
    :return: 返回一个新的excel文件
    """

    sheetNameList = ["环境监测", "设备监控", "视频监控", "防入侵"]
    wb = openpyxl.load_workbook(xls)

    # 创建一个新的excel用来合并所有数据
    wb_new = openpyxl.Workbook()
    sht_new = wb_new.create_sheet(title="total")
    sht_new.append(["所属分区", "所属舱室", "桩号", "设备类型", "唯一ID"])

    for eachSheet in sheetNameList:
        sht = wb[eachSheet]
        for eachRow in sht.rows:
            eachRow = [eachCell.value for eachCell in eachRow]
            sht_new.append(eachRow)

    path = os.path.dirname(xls)
    print(os.path.join(path, "data_total.xlsx"))

    del wb_new["Sheet"]
    wb_new.save(os.path.join(path, "data_total.xlsx"))
