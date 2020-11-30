import functools
import math
import arcpy
import re
import sys
import pandas as pd
import os
import openpyxl
import sqlite3

arcpy.env.overwriteOutput = True


class dataTypeNotAvailable(Exception):
    pass


class fileExtionNotAvailable(Exception):
    pass


class coordTupleNotAvailable(Exception):
    pass


class notEnoughPntInputed(Exception):
    pass


# todo line实例在计算临近点的时候，需要增加验证垂点是否在线上
class myXYLine:
    """
    usage: 平面线段类，用以实例化仅有首尾两点的线段
    """

    def __init__(self, startPnt, endPnt):
        """
        usage: 该类仅支持平面的 x, y 坐标，不支持带有 z 值线
        :param startPnt: tuple or list, like (x1, y1)
        :param endPnt: tuple or list, like (x2, y2)

        该类的对象具有如下属性：
            * x1, y1, x2, y2 —— 数据的坐标值
            * length —— 线路长度
            * center —— 由起终点计算的中间点
            * centroid —— 由extent计算点中心点（在线段中与center相同）
            * directionRadian —— 线段的方向角（0~2pi）
            * directionDegree —— 线段的方向角（0~360）
            * normal —— 法线的坐标二元组，((xc, yc), (xn, yn))
                xc, yc: 中心点的 x, y;
                xn, yn: 法线长度为1的点坐标;
            * normalDegree —— 法线角度（其中一个，另外一个为 180 + 该度数）
        """

        # 验证传入的元组为数值型元组，若非尝试将其中数据转为数值，无法转换则抛出 “dataTypeNotAvailable” 异常
        # 并且将坐标值限定到小数点后9位，进行四舍五入
        startPnt = myXYLine.coordTupleTest(startPnt)
        endPnt = myXYLine.coordTupleTest(endPnt)

        # 初始话坐标属性
        self.x1, self.y1 = startPnt
        self.x2, self.y2 = endPnt
        self.startPnt = startPnt
        self.endPnt = endPnt
        # 计算长度，并添加 .length 属性
        self._calGeometry()

    @staticmethod
    def coordTupleTest(coordTup):
        # 确保数据是元组或字典，目前不支持其他类型
        if not isinstance(coordTup, (tuple, list)):
            print("Error --- init myxyline failed, the type of coordTup inputed not in (tuple, list)")
            raise dataTypeNotAvailable

        # 仅保留前两个数值，过滤掉z值及其他数据
        if len(coordTup) > 2:
            coordTup = coordTup[:2]
        elif len(coordTup) < 2:
            print("Error --- the coord tuple inputed only have one or less coord")
            raise coordTupleNotAvailable

        newCoordTup = []
        for each in coordTup:
            if not isinstance(each, (int, float)):
                try:
                    newCoordTup.append(round(float(each), 9))
                except:
                    print("Error --- ")
                    raise dataTypeNotAvailable
            else:
                newCoordTup.append(round(each, 9))
        return newCoordTup

    def _calGeometry(self):
        # 长度
        self.length = round(math.sqrt((self.y2 - self.y1) ** 2 + (self.x2 - self.x1) ** 2), 6)
        # 中间点
        self.center = ((self.x1 + self.x2) / 2, (self.y1 + self.y2) / 2)
        # 中心点
        self.centroid = ((self.x1 + self.x2) / 2, (self.y1 + self.y2) / 2)
        # 方向弧度
        self.directionRadian = round(math.atan2((self.y1 - self.y2), (self.x1 - self.x2)) + math.pi, 6)
        # 方向角度
        self.directionDegree = round(math.degrees(math.atan2((self.y1 - self.y2), (self.x1 - self.x2))) + 180, 6)
        # 直线的k
        if self.directionDegree == 90 or self.directionDegree == -90 or \
                self.directionDegree == 270 or self.directionDegree == -270:
            self.k = "infinity"
        elif self.directionDegree == 180 or self.directionDegree == 180 or self.directionDegree == 360:
            self.k = 0
        else:
            self.k = round(math.tan(self.directionRadian), 6)
        # 直线的b
        if self.k != "infinity":
            self.b = round(self.y2 - self.k * self.x2, 6)
        else:
            self.b = 0
        # 线段法线（长度为1单位，由终点向外指向），用起终点元组表组，避免直接实例化后的无限调用（法线再求法线....）
        self.calNormalLine()
        return None

    def calNormalLine(self):
        # 以中点作为一点，求其中垂线
        xc, yc = self.center

        # 求法线参数
        self.normalDegree = self.directionDegree + 90
        self.normalRadian = math.radians(self.normalDegree)
        yNormal = math.sin(self.normalRadian) + yc
        xNormal = math.sqrt(1 - (yNormal - yc) ** 2) + xc
        self.normal = ((xc, yc), (xNormal, yNormal))
        return None

    def calDisAndInterPnt(self, tarPnt):
        """
        usage: 用于计算某点到线的最近距离
        :param tarPnt: tuple, (x, y)
        :return: (交点坐标元组， 距离) --- ((xInter, yInter), distance)
        """
        xt, yt = tarPnt
        # 法线的k 与 垂线相同
        norLineObj = myXYLine(*self.normal)

        # 线段本身是否为竖直线
        if self.k == "infinity":
            dis = abs(xt - self.x1)
            pntInter = (self.x1, yt)
        # 线段本身为水平线，则法线为竖直线
        elif norLineObj.k == "infinity":
            dis = abs(yt - self.y1)
            pntInter = (xt, self.y1)
        else:
            xInter = (norLineObj.b - self.b) / (self.k - norLineObj.k)
            yInter = self.k * xInter + self.b
            pntInter = (xInter, yInter)
            dis = math.sqrt((yInter - yt) ** 2 + (xInter - xt) ** 2)
        return (pntInter, dis)

    def calXYWithLength(self, length):
        """
        usage: 用于通过给定一个长度来计算，此长度在线上的点的坐标
        :param length:
        :return:
        计算原理：
            1、给定的长度已知，线长已知
            2、起终点坐标已知
            3、通过构造相似三角形，二者长度之比 等于 (X求解点 - X起点) / (X终点 - X起点)，Y同理
        """
        # 获取相似三角形的长度比例
        scale = length / self.length
        # 通过长度比例计算坐标
        xTar = scale * (self.x2 - self.x1) + self.x1
        yTar = scale * (self.y2 - self.y1) + self.y1
        return (xTar, yTar)


class myXYPolyline:
    """
    usage:

    """

    def __init__(self, coordsList, mileage=None):
        """
        usage: 用以将单段的线对象组织为多段线
        :param coordsList: tuple、list， 所有线段起止点的坐标集合
        将多段线的每一段实例化为线对象，线对象以如下格式组织
               {"ID": , "LINE": , "STARTMILEAGE": , "ENDMILEAGE":}，
        分别代表 线段的序号、线对象、  当前线对象的起始里程、 当前线对象的终点里程
        """
        self.lineCoords = coordsList

        #
        lines = []
        if not mileage:
            mileage = 0
        for i, eachPnt in enumerate(coordsList):
            # 跳过第一个点
            if i == 0:
                continue

            lineDict = {}
            lineObj = myXYLine(coordsList[i - 1], eachPnt)
            lineDict["ID"] = i
            lineDict["LINE"] = lineObj
            lineDict["STARTMILEAGE"] = mileage

            mileage += lineObj.length
            lineDict["ENDMILEAGE"] = mileage
            lineDict["STARTPNT"] = tuple(coordsList[i - 1])
            lineDict["ENDPNT"] = tuple(eachPnt)

            lines.append(lineDict)

        self.lines = lines
        self.length = mileage

    @staticmethod
    def coordsListTest(coordsList):
        # 确保数据是元组或字典，目前不支持其他类型
        if not isinstance(coordsList, (tuple, list)):
            print("Error --- init myxyply failed, the type of coordList inputed not in (tuple, list)")
            raise dataTypeNotAvailable

        # 仅有一个及以下个点，无法使用
        if len(coordsList) < 2:
            print("Error --- init myxlply failed, not enough pnts have inputed")
            raise notEnoughPntInputed


def getLineWKT(inFC, upOrDownFieldName):
    """
    usage: 获取输入的线要素类中每条线要素的坐标点集
    :param inFC: str, 输入不带z值的线要素类
    :param upOrDownFieldName: str, 字段名，用以标识哪个字段区分上下行
    :return: list, [[”线1上行或下行“, [x1, y1], [x2, y2], .... ], [”线2上行或下行“, [x1, y1], [x2, y2], .... ], ....]
    """
    plyCoordList = []
    # 获取除了上下行字段、shape_area、shape_length、SHAPE字段外的其他字段
    fieldNames = [each.name for each in arcpy.ListFields(inFC) if each.name.lower() != "shape_length" and
                  each.name.lower() != "shape_area" and each.name.lower() != "shape" and each.name != upOrDownFieldName]
    if len(fieldNames) == 0:
        attrList = ["SHAPE@WKT", upOrDownFieldName]
    else:
        attrList = ["SHAPE@WKT", upOrDownFieldName, *fieldNames]
    # print(attrList)
    with arcpy.da.SearchCursor(inFC, attrList) as cur:
        for row in cur:
            data = {}
            attr = {}
            # 获取wkt
            wkt = row[0]
            # print("wkt", wkt)
            wkt = re.findall(r"-?\d+\.?\d+\s?-?\d+\.?\d+", wkt)

            # 清除wkt中的非数值型字符，并且如果数据有z值，则z值会被过滤掉
            coord = [list(map(float, eachPnt.split(" ")[:2])) for eachPnt in wkt]

            # 获取线路的上下行
            direction = row[1]
            data["GEOMETRY"] = coord
            data["DIRECTION"] = direction

            # 获取线路的其他属性
            if len(fieldNames) > 0:
                for j, eachAttr in enumerate(fieldNames):
                    attr[eachAttr] = row[j + 2]
                data["ATTRIBUTES"] = attr
            else:
                data["ATTRIBUTES"] = {}
            plyCoordList.append(data)
            # print("coord", coord)
    # print("plyCoordList", plyCoordList)
    return plyCoordList


def readSplitPntFromXLSX(xls, startRow):
    startRow = int(startRow)
    # 确保文件类型为 .xlsx
    data, ext = os.path.splitext(xls)
    if ext == ".xls":
        # todo 目前需要将一个xls的多个sheet各自读取并写入到新的xlsx中
        df = pd.read_excel(xls)
        xlsx = data + ".xlsx"

        if os.path.exists(xlsx):
            os.remove(xlsx)

        df.to_excel(xlsx, index=False)
    elif ext != ".xlsx":
        print("Error --- the extension of input file not in (.xls, .xlsx)")
        raise fileExtionNotAvailable
    else:
        xlsx = xls

    wb = openpyxl.load_workbook(xlsx)
    shts = wb.sheetnames
    # print(shts)

    xlsxDataList = []
    for eachSht in shts:
        shtData = {}
        shtDataList = {}
        # 每个单独的sheet
        sht = wb[eachSht]
        rowMax = sht.max_row
        # 有效数据从第三行开始
        oDict = {}
        for i in range(startRow, rowMax + 1):
            dataDict = {}
            oId, oStart, oEnd, oDir = (sht.cell(i, 1).value, sht.cell(i, 2).value,
                                       sht.cell(i, 3).value, sht.cell(i, 4).value)
            if oId is None or oStart is None or oEnd is None or oDir is None:
                continue
            oDict.setdefault(oDir, [])
            (dataDict["id"], dataDict["start"],
             dataDict["end"], dataDict["direc"]) = (oId, oStart, oEnd, oDir)
            oDict[oDir].append(dataDict)
            # print(oDict)
            # shtDataList[oDir] = oDict
        shtData["NAME"] = eachSht
        shtData["DATA"] = oDict
        xlsxDataList.append(shtData)
    print(xlsxDataList)
    return xlsxDataList


# todo 现在已经获取了excel中需要打断的数据，明天按打断的数据 先做复制 6 个图层，每个图层打断一次
# todo 图层集合拆分掉

def _splitLine(singleLineDict, singleXlsxDict):
    """
    usage: 用于按照每个sheet字典中的数据，分割某条线要素
    :param singleLineDict: dict, 从每条线要素类中获取的 某条线的经过处理后的WKT，用于存储线的坐标值及其属性
    :param singleXlsxDict: dict, 从 xlsx 中获取的某个 某个sheet的有效数据，用于分割线
    :return:
    """
    # 拆分线数据（字典）
    lineGeo, lineDir, attrDict = singleLineDict["GEOMETRY"], singleLineDict["DIRECTION"], singleLineDict["ATTRIBUTES"]

    # 实例化多段线
    oPly = myXYPolyline(lineGeo)

    oLines = [oPly.lines]
    # print(oLines)

    shtName, shtData = singleXlsxDict["NAME"], singleXlsxDict["DATA"]

    resPlyFC = []
    # 从xlsx中获取与lineDir相同的（上行或下行）数据
    splitDataList = shtData[lineDir]
    for eachData in splitDataList:
        dId, dStart, dEnd, dDir = eachData["id"], eachData["start"], eachData["end"], eachData["direc"]
        plyFC = []

        # 各多段线
        for l, eachLine in enumerate(oLines):
            # 按里程插入点，并将点集提取为单独列表
            ply = []

            # 多段线中的各段
            for k, each in enumerate(eachLine):
                (pID, pStartMile, pLine,
                 pEndMile, pStartPnt, pEndPnt) = (each["ID"], each["STARTMILEAGE"], each["LINE"],
                                                  each["ENDMILEAGE"], each["STARTPNT"], each["ENDPNT"])
                if l == 0 and k == 0:
                    startMile = pStartMile
                elif l != 0 and k == 0:
                    startMile = pStartMile
                # 情况一、线段的终止里程 小于 要拆分的数据的起点里程
                # 将起终两点加入到ply中
                if pEndMile <= dStart:
                    if pStartPnt not in ply:
                        ply.append(pStartPnt)

                    if pEndPnt not in ply:
                        ply.append(pEndPnt)
                # 情况二、线段的起终点 夹着 excel 分割的起点，但是线段的终点 仍有小于/等于/大于 分割终点，三种情况
                elif pEndMile >= dStart > pStartMile:
                    # step1、将起点加入到点集中
                    if pStartPnt not in ply:
                        ply.append(pStartPnt)

                    # step2、按里程差（分割起点）计算分割起点的坐标
                    length = dStart - pStartMile
                    # 计算分割起点坐标
                    splitPnt = pLine.calXYWithLength(length)
                    # 将分割起点坐标加入到 上一段线的点集中
                    ply.append(splitPnt)
                    # 将上一段线加入到线要素类中
                    plyFC.append((ply, startMile, dStart))

                    # 设置下一段的 startMile 为上一段的 pEndMile
                    startMile = dStart

                    # step3、上一段线结束，开启下一段线
                    # 将分割起点加入到新线段的ply中
                    ply = [splitPnt]

                    # 此处开始有三种情况，
                    #   * 线段终点 小于/等于 分割终点
                    if pEndMile <= dEnd:
                        # 将未被分割的线段的终点直接加入到新线段的点集中
                        # 目前有两个点，分割点及未分割线的终点，在后面的段中 将分割点纳入，并分割
                        ply.append(pEndPnt)
                    #   * 线段终点 大于 分割终点，两个分割点 将线段拆为三份
                    else:
                        length = dEnd - pStartMile
                        # 计算分割终点坐标
                        splitPnt = pLine.calXYWithLength(length)
                        # 将分割终点加入到 线的点集中
                        ply.append(splitPnt)
                        # 将线要素加入到线要素类中，该段结束（分割的起点、终点）
                        plyFC.append((ply, startMile, dEnd))

                        startMile = dEnd

                        # 此处开启新的线段，并以分割点的终点作为 第一个点， 线段的终点作为 第二个点
                        ply = [splitPnt, pEndPnt]
                # 情况三、线段起点大于等于分割起点
                elif dStart <= pStartMile < dEnd:
                    # 线段起点与分割点起点重合
                    if pStartPnt not in ply:
                        ply.append(pStartPnt)

                    # 终点的三种情况 在分割终点 左侧/上/右侧
                    # 终点在分割终点的  左侧 / 上
                    if pEndMile <= dEnd:
                        if pEndPnt not in ply:
                            ply.append(pEndPnt)
                    # 线段终点在分割终点的右侧，以分割终点 将线段分割开
                    else:
                        # 分割点前的段
                        length = dEnd - pStartMile
                        splitPnt = pLine.calXYWithLength(length)
                        ply.append(splitPnt)
                        plyFC.append((ply, pStartMile, dEnd))

                        startMile = dEnd

                        # 分割点后段
                        ply = [splitPnt, pEndPnt]
                # 线段起点在分割终点后
                elif pStartMile >= dEnd:
                    if pStartPnt not in ply:
                        ply.append(pStartPnt)

                    if pEndPnt not in ply:
                        ply.append(pEndPnt)
                else:
                    print("注意！！！ 出现没想到的情况了！！！！")

            # 每条线(eachLine)分割结束后
            plyFC.append((ply, startMile, pEndMile))

        # 所有的线要素分割结束后，
        # 修改olines， 将其改为本次循环结束后的所有线要素类 实例化后的对象
        oLines = []
        for eachPly in plyFC:
            oStartMile = eachPly[1]
            eachPly = eachPly[0]
            plyObj = myXYPolyline(eachPly, oStartMile)
            oLines.append(plyObj.lines)
        n = len(plyFC)
        m = len(oLines)
    print(shtName, n)
    print(shtName, m)
    return plyFC


def createFeatureClass(plyList, outputPath, outputName, wkid=None, wkt=None):
    if wkid:
        sr = arcpy.SpatialReference(wkid)
    elif wkt:
        sr = arcpy.SpatialReference()
        sr.loadFromString(wkt)
    else:
        sr = None

    # plyfc = []
    # for eachPly in plyList:
    #     plyfc.append(arcpy.Polyline(arcpy.Array([arcpy.Point(*eachPnt) for eachPnt in eachPly])))
    # print(plyfc)
    # outputData = os.path.join(outputPath, outputName)
    # arcpy.CopyFeatures_management(plyfc, outputData)

    data = arcpy.CreateFeatureclass_management(outputPath, outputName, "POLYLINE", spatial_reference=sr)
    arcpy.AddField_management(data, "startMile", "DOUBLE")
    arcpy.AddField_management(data, "endMile", "DOUBLE")

    with arcpy.da.InsertCursor(data, ["SHAPE@", "startMile", "endMile"]) as cur:
        for eachPly in plyList:
            # print(eachPly)
            ply = arcpy.Polyline(arcpy.Array([arcpy.Point(*eachPnt) for eachPnt in eachPly[0]]))
            cur.insertRow([ply, eachPly[1], eachPly[2]])


def spilitLineWithXlsxData(xlsxData, plyCoord, wkt):
    for i, eachPly in enumerate(plyCoord):
        for j, eachData in enumerate(xlsxData):
            shtName = eachData["NAME"]
            plyFC = _splitLine(eachPly, eachData)
            name = f"ply_{shtName}_{i}_{j}"
            print(plyFC)
            createFeatureClass(plyFC, r"D:\codeProjcet\ArcGISProPycharm\myScript\自用工具_github\上海申通业务\data\res.gdb", name,
                               wkt=wkt)
            with open(rf"./data/{name}.txt", "w", encoding="utf-8") as f:
                for each in plyFC:
                    f.write(str(each))
                    f.write("\n")
                    f.write("\n")
                    f.write("\n")
                    f.write("\n")


def createTable(dbFile, tableName):
    db = sqlite3.connect(dbFile)
    cur = db.cursor()
    cur.execute(f"drop table if exists {tableName};")
    cur.execute(f"create table if not exists {tableName}(ID integer primary key autoincrement, "
                f"X real, Y real, MILE real, LINE integer, SHT_LINE_ID integer);")
    db.commit()
    cur.close()
    db.close()


def openConnectDB(dbFile):
    db = sqlite3.connect(dbFile)
    cur = db.cursor()
    return db, cur


def insertValue(dbObj, curObj, tableName, values):
    curObj.execute(f"insert into {tableName}(ID, X, Y, MILE, LINE, SHT_LINE_ID)"
                f" values({values[0]}, {values[1]}, {values[2]}, {values[3]}, {values[4]}, {values[5]});")
    dbObj.commit()


def insertValueXlsxData(dbObj, curObj, tableName, values):
    # 查看当前的里程点 是否已经存在于表中（已经是节点了）
    res = curObj.execute(f"select id, x, y from {tableName} where MILE = {values[3]};").fetchall()
    # 表中不存在当前里程点
    if len(res) == 0:
        curObj.execute(f"insert into {tableName}(ID, X, Y, MILE, LINE, SHT_LINE_ID)"
                    f" values({values[0]}, {values[1]}, {values[2]}, {values[3]}, {values[4]}, {values[5]});")
    # 表中存在点，且 x, y 非空
    elif res[0][1] is not None and res[0][2] is not None:
        curObj.execute(f"update {tableName} set LINE = 2 where ID = {res[0][0]};")

    dbObj.commit()



def copyOrderdTable(dbObj, curObj, oriTable, tarTable):
    resData = curObj.execute(f"select * from {oriTable} order by MILE;").fetchall()
    for each in resData:
        # print(each)
        insertData = []
        for eachSingle in each:
            if eachSingle == None:
                eachSingle = "null"
            insertData.append(eachSingle)
        curObj.execute(f"insert into {tarTable}(X, Y, MILE, LINE, SHT_LINE_ID) "
                       f"values({insertData[1]}, {insertData[2]}, {insertData[3]}, {insertData[4]}, {insertData[5]});")
        # curObj.execute(f"insert into {tarTable} select * from {oriTable} order by MILE;")
    dbObj.commit()


def closeConnectDB(dbObj, curObj):
    curObj.close()
    dbObj.close()


def calculateMiles(plyCoord):
    for eachPly in plyCoord:
        geoList = eachPly["GEOMETRY"]
        mile = 0
        for i, eachPnt in enumerate(geoList):
            # 第一个点，直接添加里程为0，并且保留下其坐标值
            if i == 0:
                eachPnt.append(0)
                eachPnt.insert(0, i + 1)
                prePnt = eachPnt
                continue
            # 后续的点
            # 保留当前点的坐标
            proPnt = eachPnt

            # 计算两点距离
            dis = round(math.sqrt((proPnt[1] - prePnt[2]) ** 2 + (proPnt[0] - prePnt[1]) ** 2), 3)
            mile += dis

            # 添加里程值
            eachPnt.append(mile)
            eachPnt.insert(0, i + 1)

            # 将当前点坐标保留为前一点
            prePnt = proPnt

        eachPly["GEOMETRY"] = geoList
    print(plyCoord)
    return plyCoord


def calXYInDB(dbFile, tarTableNameList):
    db, cur = openConnectDB(dbFile)
    # 读取表中所有数据
    for eachTable in tarTableNameList:
        resData = cur.execute(f"select * from {eachTable} where LINE = 1;").fetchall()

        print(resData)
        print(len(resData))
        # sys.exit()

        #
        for eachLine in resData:
            (oId, oX, oY, oMile,
             oLine, oShtLineId) = (eachLine[0], eachLine[1], eachLine[2],
                                   eachLine[3], eachLine[4], eachLine[5])

            calX, calY, preX, preY, nextX, nextY = None, None, None, None, None, None
            calId = oId
            calMile, preMile, nextMile = None, None, None
            # 寻找无 x, y 的行
            if oX is None and oY is None:
                # 向前寻找具有坐标值的点
                while calX is None and calY is None:
                    calId -= 1
                    coord = cur.execute(f"select X, Y, MILE from {eachTable} where ID = {calId}").fetchall()
                    calX, calY, calMile = coord[0]

                # 前向坐标点
                preX, preY, preMile = calX, calY, calMile

                # 向后搜索坐标值
                calX, calY, calMile = None, None, None
                calId = oId
                while calX is None and calY is None:
                    calId += 1
                    coord = cur.execute(f"select X, Y, MILE from {eachTable} where ID = {calId}").fetchall()
                    calX, calY, calMile = coord[0]

                # 后向坐标点
                nextX, nextY, nextMile = calX, calY, calMile

                # 求当前点到前向点的 里程差
                preDis = float(oMile) - float(preMile)

                # 实例化线
                lineObj = myXYLine((preX, preY), (nextX, nextY))
                # 求当前点的坐标值
                tarX, tarY = lineObj.calXYWithLength(preDis)

                # 将坐标值写入到表中
                cur.execute(f"update {eachTable} set X = {tarX}, Y = {tarY} where ID = {oId}")
                db.commit()

    closeConnectDB(db, cur)


def insertPlyToDB(dbFile, xlsxData, plyCoord):
    for eachPly in plyCoord:
        plyGeoList, plyDir = eachPly["GEOMETRY"], eachPly["DIRECTION"]
        for eachSheet in xlsxData:
            shtName, shtData = eachSheet["NAME"], eachSheet["DATA"]

            # 每个sheet创建一张表，表名为 sheet名_方向
            tableName = shtName + "_" + plyDir
            createTable(dbFile, tableName)
            print(f"正在处理{tableName}")
            # 连接数据库
            db, cur = openConnectDB(dbFile)

            # 先插入线的原始数据
            for eachPnt in plyGeoList:
                values = eachPnt
                values.append(0)
                values.append("null")
                # print(values)
                insertValue(db, cur, tableName, values)

            closeConnectDB(db, cur)


def insertXlsxToDB(dbFile, xlsxData, plyCoord):
    tableNameList = []
    # 连接数据库
    db, cur = openConnectDB(dbFile)

    for eachPly in plyCoord:
        plyGeoList, plyDir = eachPly["GEOMETRY"], eachPly["DIRECTION"]
        for eachSheet in xlsxData:
            shtName, shtData = eachSheet["NAME"], eachSheet["DATA"]
            shtDirData = shtData[plyDir]
            tableName = shtName + "_" + plyDir

            # 获取该表中几何图形已存在的最大里程
            geoMaxMile = cur.execute(f"select max(MILE) from {tableName};").fetchone()[0]

            # 插入所有值
            for eachData in shtDirData:
                shtLineId, sPnt, ePnt = eachData["id"], eachData["start"], eachData["end"]
                valuesStart = ("null", "null", "null", sPnt, "1", shtLineId)
                valuesEnd = ("null", "null", "null", ePnt, "1", shtLineId)

                # 当分割点的起始里程 大于线长时，则不插入起点 和 终点
                if sPnt > geoMaxMile:
                    continue

                insertValueXlsxData(db, cur, tableName, valuesStart)

                if ePnt <= geoMaxMile:
                    insertValueXlsxData(db, cur, tableName, valuesEnd)

                tableNameList.append(tableName)

    closeConnectDB(db, cur)
    return tableNameList



def createOrderdTable(dbFile, tableNameList):
    """
    usage: 创建排过序的表
    :param dbFile:  str, 数据库文件地址
    :param tableNameList:  list, 所有的表名列表
    :return:
    """
    newTableNames = []
    for eachTable in tableNameList:
        newTable = eachTable + "_ORDERD"
        # 创建完全相同的表，sqlite3 没有用 like 复制表结构的语句
        createTable(dbFile, newTable)

        # 连接数据库
        db, cur = openConnectDB(dbFile)

        # 复制数据
        copyOrderdTable(db, cur, eachTable, newTable)

        # 关闭数据库连接
        closeConnectDB(db, cur)

        newTableNames.append(newTable)
    return newTableNames



# ************************* 生成桩号点 *************************


def readPntXlsx(xls, startRow, colTarList):
    colId, colStart, colDir = colTarList

    # 确保文件类型为 .xlsx
    data, ext = os.path.splitext(xls)
    if ext == ".xls":
        # todo 目前需要将一个xls的多个sheet各自读取并写入到新的xlsx中
        df = pd.read_excel(xls)
        xlsx = data + ".xlsx"

        if os.path.exists(xlsx):
            os.remove(xlsx)

        df.to_excel(xlsx, index=False)
    elif ext != ".xlsx":
        print("Error --- the extension of input file not in (.xls, .xlsx)")
        raise fileExtionNotAvailable
    else:
        xlsx = xls

    wb = openpyxl.load_workbook(xlsx)
    shts = wb.sheetnames
    # print(shts)

    xlsxDataList = []
    for eachSht in shts:
        shtData = {}
        # 每个单独的sheet
        sht = wb[eachSht]
        rowMax = sht.max_row
        # 有效数据从第三行开始
        oDict = {}
        for i in range(startRow, rowMax + 1):
            dataDict = {}

            oId, oStart, oDir = (sht.cell(i, colId).value, sht.cell(i, colStart).value,
                                       sht.cell(i, colDir).value)
            if oId is None or oStart is None or oDir is None:
                continue
            oDict.setdefault(oDir, [])
            (dataDict["id"], dataDict["start"],
              dataDict["direc"]) = (oId, oStart, oDir)
            oDict[oDir].append(dataDict)
            # print(oDict)
            # shtDataList[oDir] = oDict
        shtData["NAME"] = eachSht
        shtData["DATA"] = oDict
        xlsxDataList.append(shtData)
    print(xlsxDataList)
    return xlsxDataList


def insertPntXlsxToDB(dbFile, xlsxData, plyCoord):
    tableNameList = set()
    # 连接数据库
    db, cur = openConnectDB(dbFile)

    for eachPly in plyCoord:
        plyGeoList, plyDir = eachPly["GEOMETRY"], eachPly["DIRECTION"]
        for eachSheet in xlsxData:
            shtName, shtData = eachSheet["NAME"], eachSheet["DATA"]
            shtDirData = shtData[plyDir]
            tableName = shtName + "_" + plyDir

            # 获取该表中几何图形已存在的最大里程
            geoMaxMile = cur.execute(f"select max(MILE) from {tableName};").fetchone()[0]

            # 插入所有值
            for eachData in shtDirData:
                shtLineId, sPnt = eachData["id"], eachData["start"]
                valuesStart = ("null", "null", "null", sPnt, "1", shtLineId)

                # 当分割点的起始里程 大于线长时，则不插入起点 和 终点
                if sPnt > geoMaxMile:
                    continue

                insertValueXlsxData(db, cur, tableName, valuesStart)

                tableNameList.add(tableName)

    closeConnectDB(db, cur)
    return tableNameList


def createOrderdPntTable(dbFile, tableNameList):
    """
    usage: 创建排过序的表
    :param dbFile:  str, 数据库文件地址
    :param tableNameList:  list, 所有的表名列表
    :return:
    """
    newTableNames = set()
    for eachTable in tableNameList:
        newTable = eachTable + "_ORDERD"
        # 创建完全相同的表，sqlite3 没有用 like 复制表结构的语句
        createTable(dbFile, newTable)

        # 连接数据库
        db, cur = openConnectDB(dbFile)

        # 复制数据
        copyOrderdTable(db, cur, eachTable, newTable)

        # 关闭数据库连接
        closeConnectDB(db, cur)

        newTableNames.add(newTable)
    return newTableNames


def calXYInPntDB(dbFile, tarTableNameList):
    db, cur = openConnectDB(dbFile)
    # 读取表中所有数据
    for eachTable in tarTableNameList:
        resData = cur.execute(f"select * from {eachTable} where LINE = 1;").fetchall()

        print(resData)
        print(len(resData))
        # sys.exit()

        #
        for eachLine in resData:
            (oId, oX, oY, oMile,
             oLine, oShtLineId) = (eachLine[0], eachLine[1], eachLine[2],
                                   eachLine[3], eachLine[4], eachLine[5])

            calX, calY, preX, preY, nextX, nextY = None, None, None, None, None, None
            calId = oId
            calMile, preMile, nextMile = None, None, None
            # 寻找无 x, y 的行
            if oX is None and oY is None:
                # 向前寻找具有坐标值的点
                while calX is None and calY is None:
                    calId -= 1
                    coord = cur.execute(f"select X, Y, MILE from {eachTable} where ID = {calId}").fetchall()
                    calX, calY, calMile = coord[0]

                # 前向坐标点
                preX, preY, preMile = calX, calY, calMile

                # 向后搜索坐标值
                calX, calY, calMile = None, None, None
                calId = oId
                while calX is None and calY is None:
                    calId += 1
                    coord = cur.execute(f"select X, Y, MILE from {eachTable} where ID = {calId}").fetchall()
                    calX, calY, calMile = coord[0]

                # 后向坐标点
                nextX, nextY, nextMile = calX, calY, calMile

                # 求当前点到前向点的 里程差
                preDis = float(oMile) - float(preMile)

                # 实例化线
                lineObj = myXYLine((preX, preY), (nextX, nextY))
                # 求当前点的坐标值
                tarX, tarY = lineObj.calXYWithLength(preDis)

                # 将坐标值写入到表中
                cur.execute(f"update {eachTable} set X = {tarX}, Y = {tarY} where ID = {oId}")
                db.commit()

    closeConnectDB(db, cur)


def createPntFCWithDB(dbFile, tableName, outputPath):
    db, cur = openConnectDB(dbFile)

    # 筛选桩号点
    # resData = cur.execute(f"select X, Y, MILE, LINE, SHT_LINE_ID from {tableName} where LINE = 1;").fetchall()

    #
    querySql = f"select X, Y, MILE, LINE, SHT_LINE_ID from {tableName} where LINE = 1;"
    resData = pd.read_sql(querySql, db)
    print(resData)

    resData.to_csv(os.path.join(outputPath, tableName + ".txt"), index=False, encoding="utf-8")



# ***********************************************************



# # inFC = r"F:\工作项目\项目_上海申通\数据_cass打断线\ditie\zhengxian.gdb\zhengxian"
# inFC = r"F:\工作项目\项目_上海申通\数据_cass打断线\简化后\jianhua.shp"
# upOrDownFieldName = "行别"
# xls = r"F:\工作项目\项目_上海申通\数据_cass打断线\地铁正线分段表序号.xlsx"
# wkt = 'PROJCS["shanghaicity",GEOGCS["GCS_Beijing_1954",DATUM["D_Beijing_1954",SPHEROID["Krasovsky_1940",6378245.0,298.3]],PRIMEM["Greenwich",0.0],UNIT["Degree",0.0174532925199433]],PROJECTION["Transverse_Mercator"],PARAMETER["False_Easting",-3457147.81],PARAMETER["False_Northing",0.0],PARAMETER["Central_Meridian",121.2751921],PARAMETER["Scale_Factor",1.0],PARAMETER["Latitude_Of_Origin",0.0],UNIT["Meter",1.0]]'
# # wkt = r"‪C:\Users\lyce\Desktop\shanghaicity.prj"
# # dbFile = r"D:\codeProjcet\ArcGISProPycharm\myScript\自用工具_github\上海申通业务\data\st_main.db"
# dbFile = r"D:\codeProjcet\ArcGISProPycharm\myScript\自用工具_github\上海申通业务\data\st_main_pnt.db"


# =================== 共用组件 =======================
#
# # 从矢量数据中读取所有坐标点信息
# plyCoord = getLineWKT(inFC, upOrDownFieldName)
#
#
# # 从xlsx中读取数据 —— 读取 前四列，第三行开始的所有数据
# xlsxData = readSplitPntFromXLSX(xls)
#
# print(plyCoord)
#
#
# ==================================================





# =================== 暂停开发 =======================

# # 第一种方法 —— 给线建模，以模型进行运算（目前存在bug）
# spilitLineWithXlsxData(xlsxData, plyCoord, wkt)

# =================================================


# # =================== 开发 =======================
#
#
# # 第二种方法 —— 以数据库插点的形式计算
#
# # 给多段线中所有点单独计算出一个自己的里程值
# plyCoord = calculateMiles(plyCoord)
#
# # 将多段线中所有点插入表中
# insertPlyToDB(dbFile, xlsxData, plyCoord)
#
# # 将 xlsx 的分割线，按照起始里程、终止里程当作点插入到表中
# oriTableNameList = insertXlsxToDB(dbFile, xlsxData, plyCoord)
#
#
# # oriTableNameList = ["VERT_ID_上行", "CUR_ID_上行", "LOT_ID_上行", "ACO_ID_上行", "SLO_ID_上行",
# #                     "GUA_ID_上行", "VERT_ID_下行", "CUR_ID_下行", "LOT_ID_下行", "ACO_ID_下行",
# #                     "SLO_ID_下行", "GUA_ID_下行"]
#
# # 将插入xlsx数据后的表格，重排序
# newTableNameList = createOrderdTable(dbFile, oriTableNameList)
#
# newTableNameList = ['VERT_ID_上行_ORDERD', 'CUR_ID_上行_ORDERD', 'LOT_ID_上行_ORDERD',
#                     'ACO_ID_上行_ORDERD', 'SLO_ID_上行_ORDERD', 'GUA_ID_上行_ORDERD',
#                     'VERT_ID_下行_ORDERD', 'CUR_ID_下行_ORDERD', 'LOT_ID_下行_ORDERD',
#                     'ACO_ID_下行_ORDERD', 'SLO_ID_下行_ORDERD', 'GUA_ID_下行_ORDERD']
#
# # newTableNameList = ['LOT_ID_上行_ORDERD']
#
# # print(newTableNameList)
# # 计算分割点的 x, y 坐标
# calXYInDB(dbFile, newTableNameList)
#
#
#
#
# # ===============================================



# =================== 生成桩号点 --- 输入数据 =======================

# inFC = r"F:\工作项目\项目_上海申通\数据_cass打断线\简化后\jianhua.shp"
inFC = r"F:\工作项目\项目_上海申通\数据_加点新_20201130\处理_数据生成\数据_中间数据\检测数据\zhengxian_增密.shp"
upOrDownFieldName = "行别"
xls = r"F:\工作项目\项目_上海申通\数据_加点新_20201130\处理_数据生成\数据_输入数据\标志标牌修改.xlsx"
dbFile = r"F:\工作项目\项目_上海申通\数据_加点新_20201130\处理_数据生成\数据_中间数据\st_main_pnt_增密后.db"
startRow = 3
# 输入excel中id、里程值、上下行所在的列索引，从1开始
colTarList = [1, 11, 5]
outputPath = r"F:\工作项目\项目_上海申通\数据_加点新_20201130\处理_数据生成\数据_中间数据"

# =================== 生成桩号点 --- 输入数据 =======================



# =================== 共用组件 =======================

# 从矢量数据中读取所有坐标点信息
plyCoord = getLineWKT(inFC, upOrDownFieldName)

# ==================================================



# =================== 生成桩号点 =======================

# 读取 xlsx 中点数据，并生成格式化的json
xlsxPntData = readPntXlsx(xls, startRow, colTarList)

plyCoord = calculateMiles(plyCoord)

insertPlyToDB(dbFile, xlsxPntData, plyCoord)

tableNameList = insertPntXlsxToDB(dbFile, xlsxPntData, plyCoord)
print(tableNameList)

newTableNameList = createOrderdPntTable(dbFile, tableNameList)
print(newTableNameList)

calXYInPntDB(dbFile, newTableNameList)

# newTableNameList = {'Sheet1_上行_ORDERD', 'Sheet1_下行_ORDERD'}
for tableName in newTableNameList:
    createPntFCWithDB(dbFile, tableName, outputPath)
# ===============================================