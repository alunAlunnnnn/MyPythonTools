import sys
import os
import logging
import datetime

# get arcgis pro toolbox directory
tbxDir = os.path.dirname(sys.argv[0])

# create log directory
logDir = os.path.join(tbxDir, "tbx_log")
try:
    if not os.path.exists(logDir):
        os.makedirs(logDir)
except:
    pass

# make sure this script have rights to create file here
createFileRights = False
fileRightTest = os.path.join(logDir, "t_t_.txt")
logFile = os.path.join(logDir, "tool3_gxstst_log.txt")
try:
    with open(fileRightTest, "w", encoding="utf-8") as f:
        f.write("create file rights test")
    createFileRights = True
except:
    createFileRights = False

# init log set config
if createFileRights:
    logging.basicConfig(filename=logFile, filemode="w", level=logging.INFO,
                        format="\n\n *** \n %(asctime)s    %(levelname)s ==== %(message)s \n *** \n\n")
else:
    logging.basicConfig(level=logging.DEBUG,
                        format="\n\n *** \n %(asctime)s    %(levelname)s ==== %(message)s \n *** \n\n")

try:
    import arcpy
    import functools
    import inspect
    import openpyxl
    import math
    import json

    logging.info("Module import success")
except BaseException as e:
    logging.error(str(e))

arcpy.env.overwriteOutput = True

logging.info(" ========== Progress Start Running ========== ")


def logIt(func):
    @functools.wraps(func)
    def _wrapper(*args, **kwargs):
        logging.info(f" ======================== Start Function {func.__name__} ========================")
        parameters = [each.strip() for each in str(inspect.signature(func))[1:-1].split(",")]
        keyPara = kwargs
        for eachKey in keyPara:
            eachKey = eachKey.strip()
            if eachKey in parameters:
                parameters.remove(eachKey)
        new_args = list(args)
        for eachValue in kwargs.values():
            new_args.append(eachValue)

        mes = ""
        for i, eachPara in enumerate(parameters):
            mes += f"\n {eachPara}: {new_args[i]} "

        for eachKey, eachValue in keyPara.items():
            mes += f"\n {eachKey}: {eachValue} "

        logging.info(mes)
        res = func(*args, **kwargs)

        return res

    return _wrapper


class NotPointStyleField(Exception):
    pass


# ---------- show message in toolbox ----------

def _addMessage(mes):
    print(mes)
    arcpy.AddMessage(mes)


def _addWarning(mes):
    print(mes)
    arcpy.AddWarning(mes)


def _addError(mes):
    print(mes)
    arcpy.AddError(mes)


def getRunTime(func):
    @functools.wraps(func)
    def _wrapper(*args, **kwargs):
        start = datetime.datetime.now()
        print("start at: {} ".format(start))
        res = func(*args, **kwargs)
        finish = datetime.datetime.now()
        print("start at: {} ".format(start))
        print("finish at: {}".format(finish))
        cost = finish - start
        print("total cost: {}".format(cost))
        return res

    return _wrapper


# ---------- data fields process ----------

def _addField(inFeaShp, fieldName, fieldType):
    # arcpy.AddField_management(inFeaShp, fieldName, fieldType)
    try:
        arcpy.AddField_management(inFeaShp, fieldName, fieldType)
    except:
        arcpy.DeleteField_management(inFeaShp, fieldName)
        arcpy.AddField_management(inFeaShp, fieldName, fieldType)


# add a filed named "unic_id_" and calculate the unic id
def _addUnicIDField(inFeaShp):
    _addField(inFeaShp, "unic_id_", "LONG")
    codes = """a = 0
def calUnicField():
    global a
    a += 1
    return a"""
    arcpy.CalculateField_management(inFeaShp, "unic_id_", "calUnicField()", "PYTHON3", codes)


# add coord fields
def _addCoordFields(inFC, feaType):
    if feaType == "point":
        fieldList = ["x_cen_", "y_cen_", "z_cen_"]
        expList = ["!shape.centroid.X!", "!shape.centroid.Y!", "!shape.centroid.Z!"]
        for i, each in enumerate(fieldList):
            _addField(inFC, each, "DOUBLE")
            arcpy.CalculateField_management(inFC, each, expList[i], "PYTHON3")
    elif feaType == "line":
        fieldList = ["x_f_", "y_f_", "z_f_", "x_l_", "y_l_", "z_l_"]
        expList = ["!shape.firstPoint.X!", "!shape.firstPoint.Y!", "!shape.firstPoint.Z!",
                   "!shape.lastPoint.X!", "!shape.lastPoint.Y!", "!shape.lastPoint.Z!"]
        for i, each in enumerate(fieldList):
            _addField(inFC, each, "DOUBLE")
            arcpy.CalculateField_management(inFC, each, expList[i], "PYTHON3")


# ---------- line class ----------

# points type is not tuple
class pointError(Exception):
    pass


class calKError(Exception):
    pass


class GenerateSpatialIndexError(Exception):
    pass


class lineEquation:
    """
    usage: generate a line object
    :param args: accept as
     --- (fp_x, fp_y, fp_z), (lp_x, lp_y, lp_z), (ex_xin, ex_ymin, ex_xmax, ex_ymax)
    firstPoint --- fp
    lastPoint --- lp
    extent --- ex
    """

    def __init__(self, *args):
        # save all points
        self.points = []

        for each in args[:2]:
            if not isinstance(each, tuple):
                _addMessage('Point coord is not tuple type')
                raise pointError

            self.points.append((float(each[0]), float(each[1]), float(each[2])))

        self.extent_xmin = args[2][0]
        self.extent_ymin = args[2][1]
        self.extent_xmax = args[2][2]
        self.extent_ymax = args[2][3]
        self.extent = args[2]

        # get point number, start with 1
        self.pntNum = len(args)

        # init pipe size
        self.pipeSize = None

        # set coord of start point and finish point
        self.x1 = self.points[0][0]
        self.y1 = self.points[0][1]
        self.z1 = self.points[0][2]
        self.x2 = self.points[-1][0]
        self.y2 = self.points[-1][1]
        self.z2 = self.points[-1][2]

        # extent available detect
        self._extentAvailableDetect()

        self.spaindex = None
        self.spaindex_row = None
        self.spaindex_col = None
        self.spaind_totalext = None

        self.calculateK_xy()
        self.calculateB_xy()

        self.calculateK_yz()
        self.calculateB_yz()

        self.calculateK_xz()
        self.calculateB_xz()

        self.generateEquation()

    # calculate k --- ( y2 - y1 ) / ( x2 - x1 )
    def calculateK_xy(self):
        if self.x1 == self.x2:
            _addMessage('ERROR --- calculate k_xy faild,'
                        ' x1 is equal to x2, x1 is {}. x2 is {}'.format(self.x1, self.x2))
            self.k_xy = -999
            return self
            # raise calKError
        k = (self.y2 - self.y1) / (self.x2 - self.x1)
        self.k_xy = k
        return self

    # calculate b --- y1 - k * x1
    def calculateB_xy(self):
        if self.k_xy == -999:
            self.b_xy = -999
        else:
            b = self.y1 - self.k_xy * self.x1
            self.b_xy = b
        return self

    # calculate k --- ( z2 - z1 ) / ( y2 - y1 )
    def calculateK_yz(self):
        if self.y1 == self.y2:
            _addMessage('ERROR --- calculate k_yz faild, y1 is equal to y2. y1 is %s' % self.y1)
            self.k_yz = -999
            return self
            # raise calKError
        k = (self.z2 - self.z1) / (self.y2 - self.y1)
        self.k_yz = k
        return self

    # calculate b --- z1 - k * y1
    def calculateB_yz(self):
        if self.k_yz == -999:
            self.b_yz = -999
        else:
            b = self.z1 - self.k_yz * self.y1
            self.b_yz = b
        return self

    # calculate k --- ( z2 - z1 ) / ( y2 - y1 )
    def calculateK_xz(self):
        if self.x1 == self.x2:
            _addMessage('ERROR --- calculate k_xz faild, x1 is equal to x2. x1 is %s' % self.x1)
            self.k_xz = -999
            return self
            # raise calKError
        k = (self.z2 - self.z1) / (self.x2 - self.x1)
        self.k_xz = k
        return self

    # calculate b --- z1 - k * y1
    def calculateB_xz(self):
        if self.k_xz == -999:
            self.b_xz = -999
        else:
            b = self.z1 - self.k_xz * self.x1
            self.b_xz = b
        return self

    # generate function equation
    def generateEquation(self):
        self.euqation_xy = '%s * x + %s' % (self.k_xy, self.b_xy)
        self.euqation_yz = '%s * x + %s' % (self.k_yz, self.b_yz)
        self.euqation_xz = '%s * x + %s' % (self.k_xz, self.b_xz)
        return self

    # calculate the intersect point
    def calculateIntersect(self, otherLineObj):
        if self.k_xy == otherLineObj.k_xy:
            self.intersect = 'false'
            otherLineObj.intersect = 'false'
            return None

        if self.b_xy == otherLineObj.b_xy:
            x = 0
            y = self.b_xy
        else:
            x = (otherLineObj.b_xy - self.b_xy) / (self.k_xy - otherLineObj.k_xy)
            y = self.k_xy * x + self.b_xy

        # detect whether the point in the rectangle of self line
        if x > self.extent_xmin and x < self.extent_xmax:
            if y > self.extent_ymin and y < self.extent_ymax:
                self.intersect = 'true'
            else:
                self.intersect = 'false'
        else:
            self.intersect = 'false'

        # detect whether the point in the rectangle of another line
        if x > otherLineObj.extent_xmin and x < otherLineObj.extent_xmax:
            if y > otherLineObj.extent_ymin and y < otherLineObj.extent_ymax:
                otherLineObj.intersect = 'true'
            else:
                otherLineObj.intersect = 'false'
        else:
            otherLineObj.intersect = 'false'

        return x, y

    # calculate z value in intersect x,y
    def calculateZCoord_yz(self, y):
        z = self.k_yz * y + self.b_yz
        return z

    # calculate z value in intersect x,y
    def calculateZCoord_xz(self, x):
        z = self.k_xz * x + self.b_xz
        return z

    # make sure the x_min, y_min is less or equal than x_max, y_max in extent
    def _extentAvailableDetect(self):
        assert int(self.extent_xmin * 10 ** 8) <= int(
            self.extent_xmax * 10 ** 8), "Error --- Extent of line object is not available"
        assert int(self.extent_ymin * 10 ** 8) <= int(
            self.extent_ymax * 10 ** 8), "Error --- Extent of line object is not available"

    # detect the point can touch the line with a tolerance or not
    def pointTouchDet(self, pnt, tolerance, totalExtent):
        """
        usage: used to detect that, the point is in the line with a tolerance num.
                before use this, please call the function "generateSpatialIndex()" yet
        :param pnt:(x, y)
        :param tolerance:
        :return:
        """
        pnt_x, pnt_y = pnt[0], pnt[1]
        pnt_index = None
        spatialIndex_x = 1
        spatialIndex_y = 1
        # generate the spatial index for point
        spaindex_step_x = round(float((totalExtent[2] - totalExtent[0]) / spatialIndex_x), 6) + 0.0001
        spaindex_step_y = round(float((totalExtent[3] - totalExtent[1]) / spatialIndex_y), 6) + 0.0001
        total_ext_ymax = totalExtent[3]
        total_ext_xmax = totalExtent[2]

        # detect whether is out of the ply's extent
        if (self.extent_xmax > totalExtent[2] + 0.0001 or self.extent_ymax > totalExtent[3] + 0.0001
                or self.extent_xmin < totalExtent[0] - 0.0001 or self.extent_ymin < totalExtent[1] - 0.0001):
            if self.extent_xmax > totalExtent[2] + 0.0001:
                print("pnt xxxx1xxxx pnt")
            if self.extent_ymax > totalExtent[3] + 0.0001:
                print("pnt xxxx2xxxx pnt")
            if self.extent_xmin < totalExtent[0] - 0.0001:
                print("pnt xxxx3xxxx pnt")
            if self.extent_ymin < totalExtent[1] - 0.0001:
                print("pnt xxxx4xxxx pnt")
            _addError("Error --- generate spatial index for point failed, "
                      "the point is ({}, {})".format(pnt_x, pnt_y))
            _addError("total extent is {}".format(totalExtent))
            raise GenerateSpatialIndexError

        find_key = False
        # for i in range(1, 11):
        for i in range(1, spatialIndex_y + 1):
            # if spatial index has finded, break the loop
            if find_key:
                break
            if pnt_y >= total_ext_ymax - i * spaindex_step_y:
                pnt_index_y = str(i)
                # for j in range(1, 11):
                for j in range(1, spatialIndex_x + 1):
                    if pnt_x >= total_ext_xmax - j * spaindex_step_x:
                        pnt_index_x = str(j)
                        pnt_index = (str(i) + "," + str(j))
                        find_key = True
                        break

        # no spatial index, no calculate
        assert pnt_index, "there are no spatial index in point"

        # match spatial index
        if pnt_index == self.spaindex:
            # point is in the extent of polyline
            if (self.extent_xmin - tolerance <= pnt_x <= self.extent_xmax + tolerance
                    and self.extent_ymin - tolerance <= pnt_y <= self.extent_ymax + tolerance):
                # point is in the extent of polyline
                # detect whether the point is on the line

                # vertical line
                if self.k_xy == -999:
                    if min(self.x1, self.x2) - tolerance <= pnt_x <= max(self.x1, self.x2) + tolerance:
                        if min(self.y1, self.y2) - tolerance <= pnt_y <= max(self.y1, self.y2) + tolerance:
                            return True
                        else:
                            return False
                    else:
                        return False
                # horizon line
                elif self.k_xy == 0:
                    if min(self.y1, self.y2) - tolerance <= pnt_y <= max(self.y1, self.y2) + tolerance:
                        if min(self.x1, self.x2) - tolerance <= pnt_x <= max(self.x1, self.x2) + tolerance:
                            return True
                        else:
                            return False
                    else:
                        return False
                # calculate the point in line with k_xy
                else:
                    det_y = self.k_xy * pnt_x + self.b_xy
                    if det_y - tolerance <= pnt_y <= det_y + tolerance:
                        return True
                    else:
                        return False
            # point is out of ply's extent
            else:
                return False
        # the spatial index between point and ply is not equal
        else:
            return False

    def generateSpatialIndex(self, totalExtent):
        """
        usage: generate spatial index, now is generate 10*10 grid index.

        index number as this:
         ----------
        |1,10|...|1,2|1,1|
         ----------
        |2,10|...|2,2|2,1|
         ----------
        |...........|
        |.          |
        |.          |
         -----------
        |10,10|...|10,2|10,1|

        :param totalExtent: (xmin, ymin, xmax, ymax)
        :return:
        """
        spatialIndex_x = 1
        spatialIndex_y = 1
        spaindex_step_x = round(float((totalExtent[2] - totalExtent[0]) / spatialIndex_x), 6) + 0.0001
        spaindex_step_y = round(float((totalExtent[3] - totalExtent[1]) / spatialIndex_y), 6) + 0.0001
        total_ext_ymax = totalExtent[3]
        total_ext_xmax = totalExtent[2]

        self.spaind_totalext = totalExtent

        if (self.extent_xmax > totalExtent[2] + 0.0001 or self.extent_ymax > totalExtent[3] + 0.0001
                or self.extent_xmin < totalExtent[0] - 0.0001 or self.extent_ymin < totalExtent[1] - 0.0001):
            if self.extent_xmax > totalExtent[2] + 0.0001:
                print("xxxx1xxxx")
            if self.extent_ymax > totalExtent[3] + 0.0001:
                print("xxxx2xxxx")
            if self.extent_xmin < totalExtent[0] - 0.0001:
                print("xxxx3xxxx")
            if self.extent_ymin < totalExtent[1] - 0.0001:
                print("xxxx4xxxx")
            _addError("Error --- generate spatial index failed, "
                      "line object's extent is not in total extent. "
                      "the line's first point is ({}, {})".format(self.x1, self.y1))
            _addError("total extent is {}".format(totalExtent))
            _addError("ply extent is {}".format(self.extent))
            raise GenerateSpatialIndexError

        find_key = False
        # for i in range(1, 11):
        for i in range(1, spatialIndex_y + 1):
            # if spatial index has finded, break the loop
            if find_key:
                break
            if self.extent_ymax >= total_ext_ymax - i * spaindex_step_y:
                self.spaindex_row = str(i)
                # for j in range(1, 11):
                for j in range(1, spatialIndex_y + 1):
                    if self.extent_xmax >= total_ext_xmax - j * spaindex_step_x:
                        self.spaindex_col = str(j)
                        self.spaindex = (str(i) + "," + str(j))
                        find_key = True
                        break
        return self

    def setPipeSize(self, pipesize):
        if isinstance(pipesize, int) or isinstance(pipesize, float):
            self.pipeSize = pipesize
        else:
            _addMessage("Warning --- pipe size is not a number type, pipe size init failed")
            self.pipeSize = None


# convert feature class to excel
@logIt
def _convertToXlsx(FC, outputPath, feaType, convertAll=True):
    """
    usage: used to convert feature class's attributes to xlsx with the same name
           based on module --- arcpy， openpyxl
    :param FC: type --- String.  input the feature class in esri data type
    :param outputPath: type --- String.  define where the xlsx data will be saved
    :param convertAll: type --- Boolean.  True --> convert all attribute,
           False --> add a field named 'unic_id_' and only convert geometry coordinate and 'unic_id_'
    :return: type --- String.  the directory of output xlsx
    """
    global pntFieldList, plyFieldList, plyTotalExtent
    # create a xlsx file
    wb = openpyxl.Workbook()
    sht = wb.active

    # make sure that the environment of workspace is not None
    arcpy.env.workspace = os.path.dirname(FC) if not arcpy.env.workspace else arcpy.env.workspace

    # add coord field, feaType -- "line" or "point"
    _addCoordFields(FC, feaType)
    # add 'unic_id_' field and calculate the unic id for each row,
    # this must under other add field function, such as _addCoordFields
    _addUnicIDField(FC)

    # convert all attributes to xlsx
    if convertAll:
        fieldNameList = [each.name for each in arcpy.ListFields(FC)
                         if each.type != "OID" and each.type != "Geometry"
                         and each.name != "Shape_Length" and each.name != "Shape_Area"]
        fieldTypeList = [each.type for each in arcpy.ListFields(FC)
                         if each.type != "OID" and each.type != "Geometry"
                         and each.name != "Shape_Length" and each.name != "Shape_Area"]
        if feaType == "point":
            pntFieldList = fieldNameList
        elif feaType == "line":
            plyFieldList = fieldNameList
            ext = arcpy.Describe(FC).extent
            plyTotalExtent = (ext.XMin, ext.YMin, ext.XMax, ext.YMax)
        _addMessage("pntFieldList")
        _addMessage(pntFieldList)
        # add data title to xlsx
        sht["A1"] = "unic_id_"
        for i, eachFieldName in enumerate(fieldNameList[:-1]):
            i += 2
            sht.cell(1, i).value = eachFieldName

        # statistic the number of all fields without unic_id_, it will be add to the field column
        fieldLength = len(fieldNameList) - 1
        with arcpy.da.SearchCursor(FC, fieldNameList) as cur:
            for rowNum, row in enumerate(cur):
                unic_id_ = row[fieldLength]
                sht.cell(rowNum + 2, 1).value = unic_id_
                for i in range(fieldLength):
                    cellData = row[i]
                    sht.cell(rowNum + 2, i + 2).value = cellData

    else:
        pass

    wb.save(os.path.join(outputPath, os.path.basename(FC) + ".xlsx"))
    return os.path.join(outputPath, os.path.basename(FC) + ".xlsx")


# delete the field will be added later, make sure the field "unic_id_" is the last in fields
@logIt
def _deleteFields(inFC, fcType):
    fieldList = []
    if fcType == "point":
        fieldList = [each.name for each in arcpy.ListFields(inFC) if each.name == "unic_id_" or
                     each.name == "x_cen_" or each.name == "y_cen_" or each.name == "z_cen_"]
    elif fcType == "line":
        fieldList = [each.name for each in arcpy.ListFields(inFC) if each.name == "unic_id_" or
                     each.name == "x_f_" or each.name == "y_f_" or each.name == "z_f_" or
                     each.name == "x_l_" or each.name == "y_l_" or each.name == "z_l_" or
                     each.name == "gj_num_"]
    if len(fieldList) > 0:
        arcpy.DeleteField_management(inFC, fieldList)
    else:
        pass


@logIt
def _generateLineObj(lineXlsx, plyFieldList, plyTotalExtent):
    wb = openpyxl.load_workbook(lineXlsx)
    sht = wb.active

    # line obj
    lineObjList = []

    # get the index of each point in xlsx
    (x_f_index, y_f_index, z_f_index, x_l_index, y_l_index, z_l_index) = (
        plyFieldList.index("x_f_") + 1, plyFieldList.index("y_f_") + 1, plyFieldList.index("z_f_") + 1,
        plyFieldList.index("x_l_") + 1, plyFieldList.index("y_l_") + 1, plyFieldList.index("z_l_") + 1)
    gl_size_index = plyFieldList.index("gj_num_") + 1

    # read xlsx, start at row 2
    for i, eachRowTup in enumerate(sht.rows):
        # skip the title row
        if i == 0:
            continue

        # get coord
        x_f, y_f, z_f, x_l, y_l, z_l = (eachRowTup[x_f_index].value, eachRowTup[y_f_index].value,
                                        eachRowTup[z_f_index].value, eachRowTup[x_l_index].value,
                                        eachRowTup[y_l_index].value, eachRowTup[z_l_index].value)

        # get the attribute of the pipe size, and save it into line object
        gl_size = eachRowTup[gl_size_index].value

        # generate line object
        firstPoint = (x_f, y_f, z_f)
        lastPoint = (x_l, y_l, z_l)
        extent = (min(x_f, x_l), min(y_f, y_l), max(x_f, x_l), max(y_f, y_l))
        lineObj = lineEquation(firstPoint, lastPoint, extent)
        lineObj.generateSpatialIndex(plyTotalExtent)
        lineObj.setPipeSize(gl_size)

        lineObjList.append(lineObj)

    return lineObjList


@logIt
def _getDirectory(pntXlsx, plyObjList, plyTotalExtent):
    global pntFieldList, tolerance
    # read each row in xlsx, get point type, x, y, z
    wb = openpyxl.load_workbook(pntXlsx)
    sht = wb.active

    # "STYLE_TYPE" is the field within santong, sitong
    maxCol = sht.max_column
    maxRow = sht.max_row
    pntStyleIndex = pntFieldList.index("STYLE_TYPE") + 2
    xCenIndex = pntFieldList.index("x_cen_") + 2
    yCenIndex = pntFieldList.index("y_cen_") + 2
    zCenIndex = pntFieldList.index("z_cen_") + 2
    # _addMessage("STYLE_TYPE")
    # _addMessage(pntStyleIndex)
    # _addMessage("x_cen_")
    # _addMessage(xCenIndex)
    # _addMessage("y_cen_")
    # _addMessage(yCenIndex)
    # _addMessage("z_cen_")
    # _addMessage(zCenIndex)


    if sht.cell(1, pntStyleIndex).value == "STYLE_TYPE":

        # set directory fields
        fieldTitle = ["z1_line_", "dir_1_", "gj_line1_", "roll_z_adjust1_", "roll_z_adjust1_y_",
                      "z2_line_", "dir_2_", "gj_line2_", "roll_z_adjust2_", "roll_z_adjust2_y_",
                      "z3_line_", "dir_3_", "gj_line3_", "roll_z_adjust3_", "roll_z_adjust3_y_",
                      "z4_line_", "dir_4_", "gj_line4_", "roll_z_adjust4_", "roll_z_adjust4_y_"]
        for i, eachTitle in enumerate(fieldTitle):
            i += 1
            sht.cell(1, maxCol + i).value = eachTitle
        for i, eachRow in enumerate(sht.rows):
            step = 1
            i += 1
            # skip title row
            if i == 1:
                continue

            pntType = eachRow[pntStyleIndex].value.strip()
            x_cord = float(sht.cell(i, xCenIndex).value)
            y_cord = float(sht.cell(i, yCenIndex).value)
            z_cord = float(sht.cell(i, zCenIndex).value)

            # whether the point is on the line
            for eachLineObj in plyObjList:

                if eachLineObj.pointTouchDet((x_cord, y_cord), tolerance, plyTotalExtent):
                    writeCol = maxCol + (5 * step - 3)
                    gjSize = maxCol + (5 * step - 2)
                    newzCol = maxCol + (5 * step - 4)
                    rollZAdjust = maxCol + (5 * step - 1)
                    rollZAdjust_y = maxCol + (5 * step)

                    # point is on the ply
                    dis_1 = math.sqrt((x_cord - eachLineObj.x1) ** 2 + (y_cord - eachLineObj.y1) ** 2)
                    dis_2 = math.sqrt((x_cord - eachLineObj.x2) ** 2 + (y_cord - eachLineObj.y2) ** 2)
                    # end point in ply is more far from point
                    if dis_1 < dis_2:
                        assert dis_2 > 0, "Error --- the max distance is not bigger 0"
                        # dir_pnt = math.degrees(math.asin((y_cord - eachLineObj.y2) / dis_2))
                        dir_pnt = math.degrees(
                            math.atan2((eachLineObj.y2 - eachLineObj.y1), (eachLineObj.x2 - eachLineObj.x1)))

                        roll_z_adjust = math.degrees(
                            math.atan2((eachLineObj.z2 - eachLineObj.z1), (eachLineObj.x2 - eachLineObj.x1)))

                        roll_z_adjust_y = math.degrees(
                            math.atan2((eachLineObj.z2 - eachLineObj.z1), (eachLineObj.y2 - eachLineObj.y1)))

                        # get the line z in point x, y as the new z of point
                        pnt_line_z = eachLineObj.z1

                    else:
                        assert dis_1 > 0, "Error --- the max distance is not bigger 0"
                        # dir_pnt = math.degrees(math.asin((y_cord - eachLineObj.y1) / dis_1))
                        dir_pnt = math.degrees(
                            math.atan2((eachLineObj.y1 - eachLineObj.y2), (eachLineObj.x1 - eachLineObj.x2)))

                        roll_z_adjust = math.degrees(
                            math.atan2((eachLineObj.z1 - eachLineObj.z2), (eachLineObj.x1 - eachLineObj.x2)))

                        roll_z_adjust_y = math.degrees(
                            math.atan2((eachLineObj.z1 - eachLineObj.z2), (eachLineObj.y1 - eachLineObj.y2)))
                        # get the line z in point x, y as the new z of point
                        pnt_line_z = eachLineObj.z2

                    # pnt size
                    pnt_gj = eachLineObj.pipeSize
                    sht.cell(i, newzCol).value = pnt_line_z
                    sht.cell(i, writeCol).value = dir_pnt
                    sht.cell(i, gjSize).value = pnt_gj
                    sht.cell(i, rollZAdjust).value = roll_z_adjust
                    sht.cell(i, rollZAdjust_y).value = roll_z_adjust_y
                    step += 1

    else:
        _addError("Error --- the index of field 'STYLE_TYPE' from fiedl list, is not same"
                  "in poing xlsx file. check it!")
        raise NotPointStyleField

    resxlsx = os.path.join(os.path.dirname(pntXlsx), os.path.basename(pntXlsx).split(".xlsx")[0] + "_new.xlsx")
    wb.save(resxlsx)
    return resxlsx


# calculate gj to a number field
@logIt
def generateGJNumField(ply):
    _addField(ply, "gj_num_", "SHORT")
    codes = """def f(a):
    a = a.strip()
    try:
        if "X" in a:
            dataList = a.split("X")
            x, y = int(dataList[0]), int(dataList[1])
            res = max(x, y)
        else:
            a = int(a)
            if a < 20:
                res = 20
            else:
                res = a
        return res
    except:
        return 20"""
    arcpy.CalculateField_management(ply, "gj_num_", "f(!GJ!)", "PYTHON3", codes)


# copy points records depend on the field "STYLE_TYPE"
@logIt
def copyPointsInXlsx(pntxlsx):
    """
    usage: copy the records in pnt xlsx file depend on the field pnt_style(santong, sitong,
    guaidian, wantou).
    :param pntxlsx:
    :return:
    """
    wb = openpyxl.load_workbook(pntxlsx)
    sht = wb.active

    # get max row number and copy all points records on the end
    maxRow = sht.max_row
    maxCol = sht.max_column

    # set type field
    sht.cell(1, maxCol + 1).value = "PNT_TYPE_"
    sht.cell(1, maxCol + 2).value = "ROLL_Z_"
    sht.cell(1, maxCol + 3).value = "ROLL_X_"
    sht.cell(1, maxCol + 4).value = "new_z_fin_"
    sht.cell(1, maxCol + 5).value = "gj_gj_gj_"
    sht.cell(1, maxCol + 6).value = "roll_z_adjust_"
    sht.cell(1, maxCol + 7).value = "roll_z_adjust_y"

    styleFiledIndex = 0
    dir_1_index = 0
    dir_2_index = 0
    dir_3_index = 0
    dir_4_index = 0
    # z1_line_
    z_1_index = 0
    z_2_index = 0
    z_3_index = 0
    z_4_index = 0

    gj_line1_index = 0
    gj_line2_index = 0
    gj_line3_index = 0
    gj_line4_index = 0

    roll_z_adjust1_ = 0
    roll_z_adjust2_ = 0
    roll_z_adjust3_ = 0
    roll_z_adjust4_ = 0

    roll_z_adjust1_y = 0
    roll_z_adjust2_y = 0
    roll_z_adjust3_y = 0
    roll_z_adjust4_y = 0

    z_origin = 0
    # get the field of santong, sitong
    for colNum in range(1, maxCol + 1):
        if sht.cell(1, colNum).value == "STYLE_TYPE":
            styleFiledIndex = colNum
        elif sht.cell(1, colNum).value == "dir_1_":
            dir_1_index = colNum
        elif sht.cell(1, colNum).value == "dir_2_":
            dir_2_index = colNum
        elif sht.cell(1, colNum).value == "dir_3_":
            dir_3_index = colNum
        elif sht.cell(1, colNum).value == "dir_4_":
            dir_4_index = colNum
        elif sht.cell(1, colNum).value == "z1_line_":
            z_1_index = colNum
        elif sht.cell(1, colNum).value == "z2_line_":
            z_2_index = colNum
        elif sht.cell(1, colNum).value == "z3_line_":
            z_3_index = colNum
        elif sht.cell(1, colNum).value == "z4_line_":
            z_4_index = colNum
        elif sht.cell(1, colNum).value == "z_cen_":
            z_origin = colNum
        elif sht.cell(1, colNum).value == "gj_line1_":
            gj_line1_index = colNum
        elif sht.cell(1, colNum).value == "gj_line2_":
            gj_line2_index = colNum
        elif sht.cell(1, colNum).value == "gj_line3_":
            gj_line3_index = colNum
        elif sht.cell(1, colNum).value == "gj_line4_":
            gj_line4_index = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust1_":
            roll_z_adjust1_ = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust2_":
            roll_z_adjust2_ = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust3_":
            roll_z_adjust3_ = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust4_":
            roll_z_adjust4_ = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust1_y_":
            roll_z_adjust1_y = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust2_y_":
            roll_z_adjust2_y = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust3_y_":
            roll_z_adjust3_y = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust4_y_":
            roll_z_adjust4_y = colNum

    #
    for i, eachRow in enumerate(sht.rows):
        i += 1
        # pass the title row
        if i == 1:
            continue

        # set common z to ball model
        z1, z2, z3, z4 = (sht.cell(i, z_1_index).value, sht.cell(i, z_2_index).value,
                          sht.cell(i, z_3_index).value, sht.cell(i, z_4_index).value)
        gj1, gj2, gj3, gj4 = (sht.cell(i, gj_line1_index).value, sht.cell(i, gj_line2_index).value,
                              sht.cell(i, gj_line3_index).value, sht.cell(i, gj_line4_index).value)
        z_a1, z_a2, z_a3, z_a4 = (sht.cell(i, roll_z_adjust1_).value, sht.cell(i, roll_z_adjust2_).value,
                                  sht.cell(i, roll_z_adjust3_).value, sht.cell(i, roll_z_adjust4_).value)
        if z4 is not None:
            z1, z2, z3, z4 = (float(z1), float(z2), float(z3), float(z4))
            gj1, gj2, gj3, gj4 = (float(gj1), float(gj2), float(gj3), float(gj4))
            sht.cell(i, maxCol + 4).value = max(z1, z2, z3, z4)
            sht.cell(i, maxCol + 5).value = max(gj1, gj2, gj3, gj4)
            sht.cell(i, maxCol + 6).value = 0
            sht.cell(i, maxCol + 7).value = 0
        else:
            if z3 is not None:
                z1, z2, z3 = (float(z1), float(z2), float(z3))
                gj1, gj2, gj3 = (float(gj1), float(gj2), float(gj3))
                sht.cell(i, maxCol + 4).value = max(z1, z2, z3)
                sht.cell(i, maxCol + 5).value = max(gj1, gj2, gj3)
                sht.cell(i, maxCol + 6).value = 0
                sht.cell(i, maxCol + 7).value = 0
            else:
                if z2 is not None:
                    z1, z2 = (float(z1), float(z2))
                    gj1, gj2 = (float(gj1), float(gj2))
                    sht.cell(i, maxCol + 4).value = max(z1, z2)
                    sht.cell(i, maxCol + 5).value = max(gj1, gj2)
                    sht.cell(i, maxCol + 6).value = 0
                    sht.cell(i, maxCol + 7).value = 0
                else:
                    if z1 is not None:
                        z1 = float(z1)
                        gj1 = float(gj1)
                        sht.cell(i, maxCol + 4).value = z1
                        sht.cell(i, maxCol + 5).value = gj1
                        sht.cell(i, maxCol + 6).value = 0
                        sht.cell(i, maxCol + 7).value = 0
                    else:
                        sht.cell(i, maxCol + 4).value = float(z_origin)
                        sht.cell(i, maxCol + 4).value = 100

        # add point type to exlce
        sht.cell(i, maxCol + 1).value = 0
        sht.cell(i, maxCol + 2).value = 0
        sht.cell(i, maxCol + 3).value = 0

        # get the santong, sitong field value
        pointStyle = sht.cell(i, styleFiledIndex).value.strip()

        # copy two new rows, total 3 rows
        if pointStyle == "弯头" or pointStyle == "拐点":
            dir_1_data = sht.cell(i, dir_1_index).value
            dir_2_data = sht.cell(i, dir_2_index).value
            z_1_data = sht.cell(i, z_1_index).value
            z_2_data = sht.cell(i, z_2_index).value
            gj_1_data = sht.cell(i, gj_line1_index).value
            gj_2_data = sht.cell(i, gj_line2_index).value
            roll_z_1_data = sht.cell(i, roll_z_adjust1_).value
            roll_z_2_data = sht.cell(i, roll_z_adjust2_).value
            roll_z_1_data_y = sht.cell(i, roll_z_adjust1_y).value
            roll_z_2_data_y = sht.cell(i, roll_z_adjust2_y).value

            # set the end field first
            if dir_2_data is not None:
                new_row1 = maxRow + 1
                new_row2 = maxRow + 2
                maxRow += 2
                # set new point type to excel
                sht.cell(new_row1, maxCol + 1).value = 1
                sht.cell(new_row2, maxCol + 1).value = 1

                # set point roll-z
                sht.cell(new_row1, maxCol + 2).value = dir_1_data
                sht.cell(new_row2, maxCol + 2).value = dir_2_data

                # set point roll-x
                sht.cell(new_row1, maxCol + 3).value = 270
                sht.cell(new_row2, maxCol + 3).value = 270

                # set new z of point
                sht.cell(new_row1, maxCol + 4).value = z_1_data
                sht.cell(new_row2, maxCol + 4).value = z_2_data

                # set gj field
                sht.cell(new_row1, maxCol + 5).value = gj_1_data
                sht.cell(new_row2, maxCol + 5).value = gj_2_data

                # set roll_z_adjust field
                sht.cell(new_row1, maxCol + 6).value = roll_z_1_data
                sht.cell(new_row2, maxCol + 6).value = roll_z_2_data

                sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                sht.cell(new_row2, maxCol + 7).value = roll_z_2_data_y
                copyControl = 2
            else:
                if dir_1_data is not None:
                    new_row1 = maxRow + 1
                    maxRow += 1
                    # set new point type to excel
                    sht.cell(new_row1, maxCol + 1).value = 1

                    # set point roll-z
                    sht.cell(new_row1, maxCol + 2).value = dir_1_data

                    # set point roll-x
                    sht.cell(new_row1, maxCol + 3).value = 270

                    # set new z of point
                    sht.cell(new_row1, maxCol + 4).value = z_1_data

                    # set gj field
                    sht.cell(new_row1, maxCol + 5).value = gj_1_data

                    sht.cell(new_row1, maxCol + 6).value = roll_z_1_data

                    sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                    copyControl = 1
                else:
                    copyControl = 0

            # set new row's value
            if copyControl > 0:
                for eachCol in range(1, maxCol + 1):
                    data = sht.cell(i, eachCol).value

                    # get origin attributes from xlsx
                    if copyControl == 2:
                        sht.cell(new_row1, eachCol).value = data
                        sht.cell(new_row2, eachCol).value = data
                    elif copyControl == 1:
                        sht.cell(new_row1, eachCol).value = data
                    else:
                        pass
            else:
                pass


        # copy three new rows, total 4 rows
        elif pointStyle == "三通":
            dir_1_data = sht.cell(i, dir_1_index).value
            dir_2_data = sht.cell(i, dir_2_index).value
            dir_3_data = sht.cell(i, dir_3_index).value
            z_1_data = sht.cell(i, z_1_index).value
            z_2_data = sht.cell(i, z_2_index).value
            z_3_data = sht.cell(i, z_3_index).value
            gj_1_data = sht.cell(i, gj_line1_index).value
            gj_2_data = sht.cell(i, gj_line2_index).value
            gj_3_data = sht.cell(i, gj_line3_index).value
            roll_z_1_data = sht.cell(i, roll_z_adjust1_).value
            roll_z_2_data = sht.cell(i, roll_z_adjust2_).value
            roll_z_3_data = sht.cell(i, roll_z_adjust3_).value

            roll_z_1_data_y = sht.cell(i, roll_z_adjust1_y).value
            roll_z_2_data_y = sht.cell(i, roll_z_adjust2_y).value
            roll_z_3_data_y = sht.cell(i, roll_z_adjust3_y).value

            # set the end field first
            if dir_3_data is not None:
                new_row1 = maxRow + 1
                new_row2 = maxRow + 2
                new_row3 = maxRow + 3
                maxRow += 3
                # set new point type to excel
                sht.cell(new_row1, maxCol + 1).value = 1
                sht.cell(new_row2, maxCol + 1).value = 1
                sht.cell(new_row3, maxCol + 1).value = 1

                # set point roll-z
                sht.cell(new_row1, maxCol + 2).value = dir_1_data
                sht.cell(new_row2, maxCol + 2).value = dir_2_data
                sht.cell(new_row3, maxCol + 2).value = dir_3_data

                # set point roll-x
                sht.cell(new_row1, maxCol + 3).value = 270
                sht.cell(new_row2, maxCol + 3).value = 270
                sht.cell(new_row3, maxCol + 3).value = 270

                # set new z of point
                sht.cell(new_row1, maxCol + 4).value = z_1_data
                sht.cell(new_row2, maxCol + 4).value = z_2_data
                sht.cell(new_row3, maxCol + 4).value = z_3_data

                # set gj field
                sht.cell(new_row1, maxCol + 5).value = gj_1_data
                sht.cell(new_row2, maxCol + 5).value = gj_2_data
                sht.cell(new_row3, maxCol + 5).value = gj_3_data

                sht.cell(new_row1, maxCol + 6).value = roll_z_1_data
                sht.cell(new_row2, maxCol + 6).value = roll_z_2_data
                sht.cell(new_row3, maxCol + 6).value = roll_z_3_data

                sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                sht.cell(new_row2, maxCol + 7).value = roll_z_2_data_y
                sht.cell(new_row3, maxCol + 7).value = roll_z_3_data_y
                copyControl = 3
            else:
                if dir_2_data is not None:
                    new_row1 = maxRow + 1
                    new_row2 = maxRow + 2
                    maxRow += 2
                    # set new point type to excel
                    sht.cell(new_row1, maxCol + 1).value = 1
                    sht.cell(new_row2, maxCol + 1).value = 1

                    # set point roll-z
                    sht.cell(new_row1, maxCol + 2).value = dir_1_data
                    sht.cell(new_row2, maxCol + 2).value = dir_2_data

                    # set point roll-x
                    sht.cell(new_row1, maxCol + 3).value = 270
                    sht.cell(new_row2, maxCol + 3).value = 270

                    # set new z of point
                    sht.cell(new_row1, maxCol + 4).value = z_1_data
                    sht.cell(new_row2, maxCol + 4).value = z_2_data

                    # set gj field
                    sht.cell(new_row1, maxCol + 5).value = gj_1_data
                    sht.cell(new_row2, maxCol + 5).value = gj_2_data

                    sht.cell(new_row1, maxCol + 6).value = roll_z_1_data
                    sht.cell(new_row2, maxCol + 6).value = roll_z_2_data

                    sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                    sht.cell(new_row2, maxCol + 7).value = roll_z_2_data_y
                    copyControl = 2
                else:
                    if dir_3_data is not None:
                        new_row1 = maxRow + 1
                        maxRow += 1
                        # set new point type to excel
                        sht.cell(new_row1, maxCol + 1).value = 1
                        # set point roll-z
                        sht.cell(new_row1, maxCol + 2).value = dir_1_data
                        # set point roll-x
                        sht.cell(new_row1, maxCol + 3).value = 270
                        # set new z of point
                        sht.cell(new_row1, maxCol + 4).value = z_1_data
                        # set gj field
                        sht.cell(new_row1, maxCol + 5).value = gj_1_data

                        sht.cell(new_row1, maxCol + 6).value = roll_z_1_data

                        sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                        copyControl = 1
                    else:
                        copyControl = 0

            if copyControl > 0:
                # set new row's value
                for eachCol in range(1, maxCol + 1):
                    data = sht.cell(i, eachCol).value

                    # get origin attributes from xlsx
                    if copyControl == 3:
                        sht.cell(new_row1, eachCol).value = data
                        sht.cell(new_row2, eachCol).value = data
                        sht.cell(new_row3, eachCol).value = data
                    elif copyControl == 2:
                        sht.cell(new_row1, eachCol).value = data
                        sht.cell(new_row2, eachCol).value = data
                    elif copyControl == 1:
                        sht.cell(new_row1, eachCol).value = data

        # copy four new rows, total 5 rows
        elif pointStyle == "四通":
            dir_1_data = sht.cell(i, dir_1_index).value
            dir_2_data = sht.cell(i, dir_2_index).value
            dir_3_data = sht.cell(i, dir_3_index).value
            dir_4_data = sht.cell(i, dir_4_index).value

            z_1_data = sht.cell(i, z_1_index).value
            z_2_data = sht.cell(i, z_2_index).value
            z_3_data = sht.cell(i, z_3_index).value
            z_4_data = sht.cell(i, z_4_index).value

            gj_1_data = sht.cell(i, gj_line1_index).value
            gj_2_data = sht.cell(i, gj_line2_index).value
            gj_3_data = sht.cell(i, gj_line3_index).value
            gj_4_data = sht.cell(i, gj_line4_index).value

            roll_z_1_data = sht.cell(i, roll_z_adjust1_).value
            roll_z_2_data = sht.cell(i, roll_z_adjust2_).value
            roll_z_3_data = sht.cell(i, roll_z_adjust3_).value
            roll_z_4_data = sht.cell(i, roll_z_adjust4_).value

            roll_z_1_data_y = sht.cell(i, roll_z_adjust1_y).value
            roll_z_2_data_y = sht.cell(i, roll_z_adjust2_y).value
            roll_z_3_data_y = sht.cell(i, roll_z_adjust3_y).value
            roll_z_4_data_y = sht.cell(i, roll_z_adjust4_y).value

            # set the end field first
            if dir_4_data is not None:
                new_row1 = maxRow + 1
                new_row2 = maxRow + 2
                new_row3 = maxRow + 3
                new_row4 = maxRow + 4
                maxRow += 4

                # set new point type to excel
                sht.cell(new_row1, maxCol + 1).value = 1
                sht.cell(new_row2, maxCol + 1).value = 1
                sht.cell(new_row3, maxCol + 1).value = 1
                sht.cell(new_row4, maxCol + 1).value = 1

                # set point roll-z
                sht.cell(new_row1, maxCol + 2).value = dir_1_data
                sht.cell(new_row2, maxCol + 2).value = dir_2_data
                sht.cell(new_row3, maxCol + 2).value = dir_3_data
                sht.cell(new_row4, maxCol + 2).value = dir_4_data

                # set point roll-x
                sht.cell(new_row1, maxCol + 3).value = 270
                sht.cell(new_row2, maxCol + 3).value = 270
                sht.cell(new_row3, maxCol + 3).value = 270
                sht.cell(new_row4, maxCol + 3).value = 270

                # set new z of point
                sht.cell(new_row1, maxCol + 4).value = z_1_data
                sht.cell(new_row2, maxCol + 4).value = z_2_data
                sht.cell(new_row3, maxCol + 4).value = z_3_data
                sht.cell(new_row4, maxCol + 4).value = z_4_data

                # set gj field
                sht.cell(new_row1, maxCol + 5).value = gj_1_data
                sht.cell(new_row2, maxCol + 5).value = gj_2_data
                sht.cell(new_row3, maxCol + 5).value = gj_3_data
                sht.cell(new_row4, maxCol + 5).value = gj_4_data

                sht.cell(new_row1, maxCol + 6).value = roll_z_1_data
                sht.cell(new_row2, maxCol + 6).value = roll_z_2_data
                sht.cell(new_row3, maxCol + 6).value = roll_z_3_data
                sht.cell(new_row4, maxCol + 6).value = roll_z_4_data

                sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                sht.cell(new_row2, maxCol + 7).value = roll_z_2_data_y
                sht.cell(new_row3, maxCol + 7).value = roll_z_3_data_y
                sht.cell(new_row4, maxCol + 7).value = roll_z_4_data_y
                copyControl = 4
            else:
                # set the end field first
                if dir_3_data is not None:
                    new_row1 = maxRow + 1
                    new_row2 = maxRow + 2
                    new_row3 = maxRow + 3
                    maxRow += 3
                    # set new point type to excel
                    sht.cell(new_row1, maxCol + 1).value = 1
                    sht.cell(new_row2, maxCol + 1).value = 1
                    sht.cell(new_row3, maxCol + 1).value = 1

                    # set point roll-z
                    sht.cell(new_row1, maxCol + 2).value = dir_1_data
                    sht.cell(new_row2, maxCol + 2).value = dir_2_data
                    sht.cell(new_row3, maxCol + 2).value = dir_3_data

                    # set point roll-x
                    sht.cell(new_row1, maxCol + 3).value = 270
                    sht.cell(new_row2, maxCol + 3).value = 270
                    sht.cell(new_row3, maxCol + 3).value = 270

                    # set new z of point
                    sht.cell(new_row1, maxCol + 4).value = z_1_data
                    sht.cell(new_row2, maxCol + 4).value = z_2_data
                    sht.cell(new_row3, maxCol + 4).value = z_3_data

                    # set gj field
                    sht.cell(new_row1, maxCol + 5).value = gj_1_data
                    sht.cell(new_row2, maxCol + 5).value = gj_2_data
                    sht.cell(new_row3, maxCol + 5).value = gj_3_data

                    sht.cell(new_row1, maxCol + 6).value = roll_z_1_data
                    sht.cell(new_row2, maxCol + 6).value = roll_z_2_data
                    sht.cell(new_row3, maxCol + 6).value = roll_z_3_data

                    sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                    sht.cell(new_row2, maxCol + 7).value = roll_z_2_data_y
                    sht.cell(new_row3, maxCol + 7).value = roll_z_3_data_y
                    copyControl = 3
                else:
                    if dir_2_data is not None:
                        new_row1 = maxRow + 1
                        new_row2 = maxRow + 2
                        maxRow += 2
                        # set new point type to excel
                        sht.cell(new_row1, maxCol + 1).value = 1
                        sht.cell(new_row2, maxCol + 1).value = 1

                        # set point roll-z
                        sht.cell(new_row1, maxCol + 2).value = dir_1_data
                        sht.cell(new_row2, maxCol + 2).value = dir_2_data

                        # set point roll-x
                        sht.cell(new_row1, maxCol + 3).value = 270
                        sht.cell(new_row2, maxCol + 3).value = 270

                        # set new z of point
                        sht.cell(new_row1, maxCol + 4).value = z_1_data
                        sht.cell(new_row2, maxCol + 4).value = z_2_data

                        # set gj field
                        sht.cell(new_row1, maxCol + 5).value = gj_1_data
                        sht.cell(new_row2, maxCol + 5).value = gj_2_data

                        sht.cell(new_row1, maxCol + 6).value = roll_z_1_data
                        sht.cell(new_row2, maxCol + 6).value = roll_z_2_data

                        sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                        sht.cell(new_row2, maxCol + 7).value = roll_z_2_data_y
                        copyControl = 2
                    else:
                        if dir_1_data is not None:
                            new_row1 = maxRow + 1
                            maxRow += 1
                            # set new point type to excel
                            sht.cell(new_row1, maxCol + 1).value = 1
                            # set point roll-z
                            sht.cell(new_row1, maxCol + 2).value = dir_1_data
                            # set point roll-x
                            sht.cell(new_row1, maxCol + 3).value = 270
                            # set new z of point
                            sht.cell(new_row1, maxCol + 4).value = z_1_data
                            # set gj field
                            sht.cell(new_row1, maxCol + 5).value = gj_1_data

                            sht.cell(new_row1, maxCol + 6).value = roll_z_1_data

                            sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                            copyControl = 1
                        else:
                            copyControl = 0

            if copyControl > 0:
                # set new row's value
                for eachCol in range(1, maxCol + 1):
                    data = sht.cell(i, eachCol).value

                    # get origin attributes from xlsx
                    if copyControl == 4:
                        sht.cell(new_row1, eachCol).value = data
                        sht.cell(new_row2, eachCol).value = data
                        sht.cell(new_row3, eachCol).value = data
                        sht.cell(new_row4, eachCol).value = data
                    elif copyControl == 3:
                        sht.cell(new_row1, eachCol).value = data
                        sht.cell(new_row2, eachCol).value = data
                        sht.cell(new_row3, eachCol).value = data
                    elif copyControl == 2:
                        sht.cell(new_row1, eachCol).value = data
                        sht.cell(new_row2, eachCol).value = data
                    elif copyControl == 1:
                        sht.cell(new_row1, eachCol).value = data

    delZCen, delZ1Line, delPntType, delRollZ, delRollZ1 = None, None, None, None, None
    # delete temp fields
    maxCol = sht.max_column
    for i in range(1, maxCol + 1):
        value = sht.cell(1, i).value
        if value == "z1_line_":
            delZ1Line = i
        elif value == "PNT_TYPE_":
            delPntType = i

    if delZ1Line and delPntType:
        sht.delete_cols(delZ1Line, (delPntType - delZ1Line))

    maxCol = sht.max_column
    for i in range(1, maxCol + 1):
        value = sht.cell(1, i).value
        if value == "z_cen_":
            delZCen = i

    if delZCen:
        sht.delete_cols(delZCen, 1)

    # maxCol = sht.max_column
    # for i in range(1, maxCol + 1):
    #     value = sht.cell(1, i).value
    #     if value == "roll_z_adjust_":
    #         delRollZ1 = i
    #
    # if delRollZ1:
    #     sht.delete_cols(delRollZ1, 1)
    #
    # maxCol = sht.max_column
    # for i in range(1, maxCol + 1):
    #     value = sht.cell(1, i).value
    #     if value == "roll_z_adjust_2":
    #         delRollZ1 = i
    #
    # if delRollZ1:
    #     sht.delete_cols(delRollZ1, 1)

    basedir = os.path.dirname(pntxlsx)
    basename = os.path.basename(pntxlsx)
    newdir = os.path.join(basedir, "res_xlsx")
    if not os.path.exists(newdir):
        os.makedirs(newdir)
    wb.save(os.path.join(newdir, basename))
    return os.path.join(newdir, basename)


# get all point data with the attribute of santong, sitong, wantou
def selectTargetData(inGDB, outputGDB):
    arcpy.env.workspace = inGDB

    outputdir = os.path.dirname(outputGDB)
    outputname = os.path.basename(outputGDB)
    # make sure the target GDB is exists
    if not arcpy.Exists(outputGDB):
        arcpy.CreateFileGDB_management(outputdir,
                                       outputname)

    dataList = arcpy.ListFeatureClasses()
    for each in dataList:
        # select data
        if each.split("_")[-1] == "POINT":
            outFC = arcpy.Select_analysis(each, os.path.join(outputGDB, each),
                                          "TZD LIKE '%三通%' Or TZD LIKE '%四通%' Or TZD LIKE '%弯头%' Or TZD LIKE '%拐点%'")
        elif each.split("_")[-1] == "LINE":
            arcpy.CopyFeatures_management(each, os.path.join(outputGDB, each))

    # reset out put gdb
    arcpy.env.workspace = outputGDB

    pntList = arcpy.ListFeatureClasses("*_POINT")
    plyList = arcpy.ListFeatureClasses("*_LINE")

    resList = []
    for eachpnt in pntList:
        basename = eachpnt[:-6]
        dataindex = plyList.index(basename + "_LINE")
        dataTup = (eachpnt, plyList[dataindex])
        resList.append(dataTup)

    return resList


def f(rollz, rollx, zx, zy):
    # dir - x
    if -45 <= rollz <= 45 or 135 <= rollz <= 180 or -180 <= rollz <= -135:
        if 0 <= zx <= 90:
            res = rollx + zx
        elif 90 < zx <= 180:
            res = 180 - zx + rollx
        elif 0 > zx >= -90:
            res = rollx + zx
        elif -180 <= zx < -90:
            res = rollx - (180 + zx)
    else:
        if 0 <= zy <= 90:
            res = rollx + zy
        elif 90 < zy <= 180:
            res = 180 - zy + rollx
        elif 0 > zy >= -90:
            res = rollx + zy
        elif -180 <= zy < -90:
            res = rollx - (180 + zy)
    return res


def generateNewRoll_X(pntxlsx):
    wb = openpyxl.load_workbook(pntxlsx)
    sht = wb.active

    roll_z_index = 0
    roll_x_index = 0
    roll_x_adju_x_index = 0
    roll_x_adju_y_index = 0

    maxCol = sht.max_column
    for i, eachCol in enumerate(sht.columns):
        i += 1
        if sht.cell(1, i).value == "ROLL_Z_":
            roll_z_index = i
        elif sht.cell(1, i).value == "ROLL_X_":
            roll_x_index = i
        elif sht.cell(1, i).value == "roll_z_adjust_":
            roll_x_adju_x_index = i
        elif sht.cell(1, i).value == "roll_z_adjust_y":
            roll_x_adju_y_index = i

    for i, eachRow in enumerate(sht.rows):
        i += 1
        if i == 1:
            pass

        roll_z_data = sht.cell(i, roll_z_index).value
        roll_x_data = sht.cell(i, roll_x_index).value
        roll_x_adju_x_data = sht.cell(i, roll_x_adju_x_index).value
        roll_x_adju_y_data = sht.cell(i, roll_x_adju_y_index).value
        # try:
        roll_z_data, roll_x_data, roll_x_adju_x_data, roll_x_adju_y_data = (
            float(roll_z_data), float(roll_x_data), float(roll_x_adju_x_data),
            float(roll_x_adju_y_data))
        res = f(roll_z_data, roll_x_data, roll_x_adju_x_data, roll_x_adju_y_data)
        sht.cell(i, maxCol + 1).value = res
        # except:
        #     sht.cell(i, maxCol + 1).value = "None"

    wb.save(os.path.join(os.path.dirname(pntxlsx),
                         os.path.basename(pntxlsx).strip(".xlsx") + "_res.xlsx"))

    return os.path.join(os.path.dirname(pntxlsx),
                        os.path.basename(pntxlsx).strip(".xlsx") + "_res.xlsx")


@logIt
def xlsxToPointFC(xlsx, sr, outputPath):
    resData = os.path.join(outputPath, os.path.basename(xlsx).strip(".xlsx"))
    tv = arcpy.MakeTableView_management(xlsx + "\Sheet$", "tv_tmp")
    xyLyr = arcpy.MakeXYEventLayer_management(tv, "x_cen_", "y_cen_",
                                              os.path.basename(xlsx).strip(".xlsx") + "_temp",
                                              sr, "new_z_fin_")
    arcpy.CopyFeatures_management(xyLyr, resData)
    addNewRollXField(resData)
    return resData


@logIt
def addNewRollXField(pntFC):
    _addField(pntFC, "ROLL_X_N_", "DOUBLE")
    codes = """def f(rollz, rollx, zx, zy):
    # dir - x
    if -45 <= rollz <= 45 or 135 <= rollz <= 180 or -180 <= rollz <= -135:
        if 0 <= zx <= 90:
            res = rollx + zx
        elif 90 < zx <= 180:
            res = 180 - zx + rollx
        elif 0 > zx >= -90:
            res = rollx + zx
        elif -180 <= zx < -90:
            res = rollx - (180 + zx)
    else:
        if 0 <= zy <= 90:
            res = rollx + zy
        elif 90 < zy <= 180:
            res = 180 - zy + rollx
        elif 0 > zy >= -90:
            res = rollx + zy
        elif -180 <= zy < -90:
            res = rollx - (180 + zy)
    return res"""
    arcpy.CalculateField_management(pntFC, "ROLL_X_N_", "f(!ROLL_Z_!, !ROLL_X_!, !roll_z_adjust_!, !roll_z_adjust_y!)",
                                    "PYTHON3", codes)


def createSymbolFile():
    data = """{
  "type" : "CIMLayerDocument",
  "version" : "2.6.0",
  "build" : 24783,
  "layers" : [
    "CIMPATH=scene/xx_yxx_point_new.xml"
  ],
  "layerDefinitions" : [
    {
      "type" : "CIMFeatureLayer",
      "name" : "XX_YXX_POINT_new",
      "uRI" : "CIMPATH=scene/xx_yxx_point_new.xml",
      "sourceModifiedTime" : {
        "type" : "TimeInstant"
      },
      "useSourceMetadata" : true,
      "description" : "XX_YXX_POINT_new",
      "layerElevation" : {
        "type" : "CIMLayerElevationSurface"
      },
      "expanded" : true,
      "layer3DProperties" : {
        "type" : "CIM3DLayerProperties",
        "castShadows" : true,
        "isLayerLit" : true,
        "layerFaceCulling" : "None",
        "maxDistance" : -1,
        "minDistance" : -1,
        "preloadTextureCutoffHigh" : 0,
        "preloadTextureCutoffLow" : 0.25,
        "textureCutoffHigh" : 0.25,
        "textureCutoffLow" : 1,
        "useCompressedTextures" : true,
        "verticalExaggeration" : 1,
        "exaggerationMode" : "ScaleZ",
        "verticalUnit" : {
          "uwkid" : 9001
        },
        "lighting" : "OneSideDataNormal"
      },
      "layerType" : "Operational",
      "showLegends" : true,
      "visibility" : true,
      "displayCacheType" : "Permanent",
      "maxDisplayCacheAge" : 5,
      "showPopups" : true,
      "serviceLayerID" : -1,
      "refreshRate" : -1,
      "refreshRateUnit" : "esriTimeUnitsSeconds",
      "autoGenerateFeatureTemplates" : true,
      "featureElevationExpression" : "Shape.Z",
      "featureTable" : {
        "type" : "CIMFeatureTable",
        "displayField" : "UL_NAME",
        "editable" : true,
        "fieldDescriptions" : [
          {
            "type" : "CIMFieldDescription",
            "alias" : "ObjectID_1",
            "fieldName" : "OBJECTID_1",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "Shape",
            "fieldName" : "Shape",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "unic_id_",
            "fieldName" : "unic_id_",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "OBJECTID",
            "fieldName" : "OBJECTID",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "FEATURECOD",
            "fieldName" : "FEATURECOD",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "GXDM",
            "fieldName" : "GXDM",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "WTDH",
            "fieldName" : "WTDH",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "TSDH",
            "fieldName" : "TSDH",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "YSDM",
            "fieldName" : "YSDM",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "TFH",
            "fieldName" : "TFH",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "XZB",
            "fieldName" : "XZB",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "YZB",
            "fieldName" : "YZB",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "DMGC",
            "fieldName" : "DMGC",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "TZD",
            "fieldName" : "TZD",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "FSW",
            "fieldName" : "FSW",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "SZDL",
            "fieldName" : "SZDL",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "JSDM",
            "fieldName" : "JSDM",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "XZJ",
            "fieldName" : "XZJ",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "JS",
            "fieldName" : "JS",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "JGXZ",
            "fieldName" : "JGXZ",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "JGCZ",
            "fieldName" : "JGCZ",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "JGZJ",
            "fieldName" : "JGZJ",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "JGC",
            "fieldName" : "JGC",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "JGK",
            "fieldName" : "JGK",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "JXJCZ",
            "fieldName" : "JXJCZ",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "JBS",
            "fieldName" : "JBS",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "JSZJ",
            "fieldName" : "JSZJ",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "QSDW",
            "fieldName" : "QSDW",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "TCDW",
            "fieldName" : "TCDW",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "TCRQ",
            "fieldName" : "TCRQ",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "SJLY",
            "fieldName" : "SJLY",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "BZ",
            "fieldName" : "BZ",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "PRJNO",
            "fieldName" : "PRJNO",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "DLPRJNO",
            "fieldName" : "DLPRJNO",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "DDATE",
            "fieldName" : "DDATE",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "DUSER",
            "fieldName" : "DUSER",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "RKDATE",
            "fieldName" : "RKDATE",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "ISLOCK",
            "fieldName" : "ISLOCK",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "NODEID",
            "fieldName" : "NODEID",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "ULPRJNO",
            "fieldName" : "ULPRJNO",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "SSQY",
            "fieldName" : "SSQY",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "YLQY",
            "fieldName" : "YLQY",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "CZ",
            "fieldName" : "CZ",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "GTLX",
            "fieldName" : "GTLX",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "BHGD",
            "fieldName" : "BHGD",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "GG_FGG",
            "fieldName" : "GG_FGG",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "UL_NAME",
            "fieldName" : "UL_NAME",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "DSDX",
            "fieldName" : "DSDX",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "SSQY_TMP",
            "fieldName" : "SSQY_TMP",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "STYLE_TYPE",
            "fieldName" : "STYLE_TYPE",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "JTZJ",
            "fieldName" : "JTZJ",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "JS_1",
            "fieldName" : "JS_1",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "x_cen_",
            "fieldName" : "x_cen_",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "y_cen_",
            "fieldName" : "y_cen_",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "PNT_TYPE_",
            "fieldName" : "PNT_TYPE_",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "ROLL_Z_",
            "fieldName" : "ROLL_Z_",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "ROLL_X_",
            "fieldName" : "ROLL_X_",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "new_z_fin_",
            "fieldName" : "new_z_fin_",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "gj_gj_gj_",
            "fieldName" : "gj_gj_gj_",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "roll_z_adjust_",
            "fieldName" : "roll_z_adjust_",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "roll_z_adjust_y",
            "fieldName" : "roll_z_adjust_y",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "ROLL_X_N_",
            "fieldName" : "ROLL_X_N_",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          },
          {
            "type" : "CIMFieldDescription",
            "alias" : "GJ_NEW",
            "fieldName" : "GJ_NEW",
            "numberFormat" : {
              "type" : "CIMNumericFormat",
              "alignmentOption" : "esriAlignRight",
              "alignmentWidth" : 0,
              "roundingOption" : "esriRoundNumberOfDecimals",
              "roundingValue" : 6
            },
            "visible" : true,
            "searchMode" : "Exact"
          }
        ],
        "dataConnection" : {
          "type" : "CIMStandardDataConnection",
          "workspaceConnectionString" : "DATABASE=.\\\\新测试数据.gdb",
          "workspaceFactory" : "FileGDB",
          "dataset" : "XX_YXX_POINT_new",
          "datasetType" : "esriDTFeatureClass"
        },
        "studyAreaSpatialRel" : "esriSpatialRelUndefined",
        "searchOrder" : "esriSearchOrderSpatial"
      },
      "htmlPopupEnabled" : true,
      "selectable" : true,
      "featureCacheType" : "Session",
      "displayFiltersType" : "ByScale",
      "labelClasses" : [
        {
          "type" : "CIMLabelClass",
          "expression" : "$feature.UL_NAME",
          "expressionEngine" : "Arcade",
          "featuresToLabel" : "AllVisibleFeatures",
          "maplexLabelPlacementProperties" : {
            "type" : "CIMMaplexLabelPlacementProperties",
            "featureType" : "Point",
            "avoidPolygonHoles" : true,
            "canOverrunFeature" : true,
            "canPlaceLabelOutsidePolygon" : true,
            "canRemoveOverlappingLabel" : true,
            "canStackLabel" : true,
            "connectionType" : "Unambiguous",
            "constrainOffset" : "NoConstraint",
            "contourAlignmentType" : "Page",
            "contourLadderType" : "Straight",
            "contourMaximumAngle" : 90,
            "enableConnection" : true,
            "enablePointPlacementPriorities" : true,
            "featureWeight" : 0,
            "fontHeightReductionLimit" : 4,
            "fontHeightReductionStep" : 0.5,
            "fontWidthReductionLimit" : 90,
            "fontWidthReductionStep" : 5,
            "graticuleAlignmentType" : "Straight",
            "keyNumberGroupName" : "Default",
            "labelBuffer" : 15,
            "labelLargestPolygon" : true,
            "labelPriority" : -1,
            "labelStackingProperties" : {
              "type" : "CIMMaplexLabelStackingProperties",
              "stackAlignment" : "ChooseBest",
              "maximumNumberOfLines" : 3,
              "minimumNumberOfCharsPerLine" : 3,
              "maximumNumberOfCharsPerLine" : 24,
              "separators" : [
                {
                  "type" : "CIMMaplexStackingSeparator",
                  "separator" : " ",
                  "splitAfter" : true
                },
                {
                  "type" : "CIMMaplexStackingSeparator",
                  "separator" : ",",
                  "visible" : true,
                  "splitAfter" : true
                }
              ]
            },
            "lineFeatureType" : "General",
            "linePlacementMethod" : "OffsetCurvedFromLine",
            "maximumLabelOverrun" : 36,
            "maximumLabelOverrunUnit" : "Point",
            "minimumFeatureSizeUnit" : "Map",
            "multiPartOption" : "OneLabelPerPart",
            "offsetAlongLineProperties" : {
              "type" : "CIMMaplexOffsetAlongLineProperties",
              "placementMethod" : "BestPositionAlongLine",
              "labelAnchorPoint" : "CenterOfLabel",
              "distanceUnit" : "Percentage",
              "useLineDirection" : true
            },
            "pointExternalZonePriorities" : {
              "type" : "CIMMaplexExternalZonePriorities",
              "aboveLeft" : 4,
              "aboveCenter" : 2,
              "aboveRight" : 1,
              "centerRight" : 3,
              "belowRight" : 5,
              "belowCenter" : 7,
              "belowLeft" : 8,
              "centerLeft" : 6
            },
            "pointPlacementMethod" : "AroundPoint",
            "polygonAnchorPointType" : "GeometricCenter",
            "polygonBoundaryWeight" : 0,
            "polygonExternalZones" : {
              "type" : "CIMMaplexExternalZonePriorities",
              "aboveLeft" : 4,
              "aboveCenter" : 2,
              "aboveRight" : 1,
              "centerRight" : 3,
              "belowRight" : 5,
              "belowCenter" : 7,
              "belowLeft" : 8,
              "centerLeft" : 6
            },
            "polygonFeatureType" : "General",
            "polygonInternalZones" : {
              "type" : "CIMMaplexInternalZonePriorities",
              "center" : 1
            },
            "polygonPlacementMethod" : "CurvedInPolygon",
            "primaryOffset" : 1,
            "primaryOffsetUnit" : "Point",
            "removeExtraWhiteSpace" : true,
            "repetitionIntervalUnit" : "Point",
            "rotationProperties" : {
              "type" : "CIMMaplexRotationProperties",
              "rotationType" : "Arithmetic",
              "alignmentType" : "Straight"
            },
            "secondaryOffset" : 100,
            "strategyPriorities" : {
              "type" : "CIMMaplexStrategyPriorities",
              "stacking" : 1,
              "overrun" : 2,
              "fontCompression" : 3,
              "fontReduction" : 4,
              "abbreviation" : 5
            },
            "thinningDistanceUnit" : "Point",
            "truncationMarkerCharacter" : ".",
            "truncationMinimumLength" : 1,
            "truncationPreferredCharacters" : "aeiou",
            "truncationExcludedCharacters" : "0123456789"
          },
          "name" : "Class 1",
          "priority" : -1,
          "standardLabelPlacementProperties" : {
            "type" : "CIMStandardLabelPlacementProperties",
            "featureType" : "Line",
            "featureWeight" : "Low",
            "labelWeight" : "High",
            "numLabelsOption" : "OneLabelPerName",
            "lineLabelPosition" : {
              "type" : "CIMStandardLineLabelPosition",
              "above" : true,
              "inLine" : true,
              "parallel" : true
            },
            "lineLabelPriorities" : {
              "type" : "CIMStandardLineLabelPriorities",
              "aboveStart" : 3,
              "aboveAlong" : 3,
              "aboveEnd" : 3,
              "centerStart" : 3,
              "centerAlong" : 3,
              "centerEnd" : 3,
              "belowStart" : 3,
              "belowAlong" : 3,
              "belowEnd" : 3
            },
            "pointPlacementMethod" : "AroundPoint",
            "pointPlacementPriorities" : {
              "type" : "CIMStandardPointPlacementPriorities",
              "aboveLeft" : 2,
              "aboveCenter" : 2,
              "aboveRight" : 1,
              "centerLeft" : 3,
              "centerRight" : 2,
              "belowLeft" : 3,
              "belowCenter" : 3,
              "belowRight" : 2
            },
            "rotationType" : "Arithmetic",
            "polygonPlacementMethod" : "AlwaysHorizontal"
          },
          "textSymbol" : {
            "type" : "CIMSymbolReference",
            "symbol" : {
              "type" : "CIMTextSymbol",
              "blockProgression" : "TTB",
              "depth3D" : 1,
              "extrapolateBaselines" : true,
              "fontEffects" : "Normal",
              "fontEncoding" : "Unicode",
              "fontFamilyName" : "SimSun",
              "fontStyleName" : "Regular",
              "fontType" : "Unspecified",
              "haloSize" : 1,
              "height" : 10,
              "hinting" : "Default",
              "horizontalAlignment" : "Left",
              "kerning" : true,
              "letterWidth" : 100,
              "ligatures" : true,
              "lineGapType" : "ExtraLeading",
              "symbol" : {
                "type" : "CIMPolygonSymbol",
                "symbolLayers" : [
                  {
                    "type" : "CIMSolidFill",
                    "enable" : true,
                    "color" : {
                      "type" : "CIMRGBColor",
                      "values" : [
                        0,
                        0,
                        0,
                        100
                      ]
                    }
                  }
                ]
              },
              "textCase" : "Normal",
              "textDirection" : "LTR",
              "verticalAlignment" : "Bottom",
              "verticalGlyphOrientation" : "Right",
              "wordSpacing" : 100,
              "billboardMode3D" : "FaceNearPlane"
            }
          },
          "useCodedValue" : true,
          "visibility" : true,
          "iD" : -1
        }
      ],
      "renderer" : {
        "type" : "CIMUniqueValueRenderer",
        "visualVariables" : [
          {
            "type" : "CIMRotationVisualVariable",
            "visualVariableInfoX" : {
              "type" : "CIMVisualVariableInfo",
              "randomMax" : 360,
              "visualVariableInfoType" : "Expression",
              "valueExpressionInfo" : {
                "type" : "CIMExpressionInfo",
                "title" : "Custom",
                "expression" : "$feature.ROLL_X_N_",
                "returnType" : "Default"
              }
            },
            "visualVariableInfoY" : {
              "type" : "CIMVisualVariableInfo",
              "randomMax" : 360,
              "visualVariableInfoType" : "None",
              "valueExpressionInfo" : {
                "type" : "CIMExpressionInfo",
                "returnType" : "Default"
              }
            },
            "visualVariableInfoZ" : {
              "type" : "CIMVisualVariableInfo",
              "randomMax" : 360,
              "visualVariableInfoType" : "Expression",
              "valueExpressionInfo" : {
                "type" : "CIMExpressionInfo",
                "title" : "Custom",
                "expression" : "$feature.ROLL_Z_",
                "returnType" : "Default"
              }
            },
            "rotationTypeZ" : "Arithmetic"
          },
          {
            "type" : "CIMSizeVisualVariable",
            "authoringInfo" : {
              "type" : "CIMVisualVariableAuthoringInfo",
              "minSliderValue" : 0.029999999999999999,
              "maxSliderValue" : 1.5,
              "heading" : "GJ_NEW"
            },
            "randomMax" : 1,
            "maxSize" : 0,
            "minValue" : 0.029999999999999999,
            "maxValue" : 1.5,
            "valueRepresentation" : "Radius",
            "variableType" : "Expression",
            "valueShape" : "Unknown",
            "axis" : "HeightAxis",
            "normalizationType" : "Nothing",
            "valueExpressionInfo" : {
              "type" : "CIMExpressionInfo",
              "title" : "Custom",
              "expression" : "$feature.GJ_NEW",
              "returnType" : "Default"
            }
          }
        ],
        "colorRamp" : {
          "type" : "CIMRandomHSVColorRamp",
          "colorSpace" : {
            "type" : "CIMICCColorSpace",
            "url" : "Default RGB"
          },
          "maxH" : 360,
          "minS" : 15,
          "maxS" : 30,
          "minV" : 99,
          "maxV" : 100,
          "minAlpha" : 100,
          "maxAlpha" : 100
        },
        "defaultLabel" : "<all other values>",
        "defaultSymbol" : {
          "type" : "CIMSymbolReference",
          "symbol" : {
            "type" : "CIMPointSymbol",
            "symbolLayers" : [
              {
                "type" : "CIMVectorMarker",
                "enable" : true,
                "anchorPointUnits" : "Relative",
                "dominantSizeAxis3D" : "Z",
                "size" : 0.43747609286026362,
                "billboardMode3D" : "FaceNearPlane",
                "frame" : {
                  "xmin" : -2,
                  "ymin" : -2,
                  "xmax" : 2,
                  "ymax" : 2
                },
                "markerGraphics" : [
                  {
                    "type" : "CIMMarkerGraphic",
                    "geometry" : {
                      "curveRings" : [
                        [
                          [
                            1.2246467991473532e-16,
                            2
                          ],
                          {
                            "a" : [
                              [
                                1.2246467991473532e-16,
                                2
                              ],
                              [
                                2.8808137021634957e-15,
                                0
                              ],
                              0,
                              1
                            ]
                          }
                        ]
                      ]
                    },
                    "symbol" : {
                      "type" : "CIMPolygonSymbol",
                      "symbolLayers" : [
                        {
                          "type" : "CIMSolidStroke",
                          "enable" : true,
                          "capStyle" : "Round",
                          "joinStyle" : "Round",
                          "lineStyle3D" : "Strip",
                          "miterLimit" : 10,
                          "width" : 0.076558316250546127,
                          "color" : {
                            "type" : "CIMRGBColor",
                            "values" : [
                              0,
                              0,
                              0,
                              100
                            ]
                          }
                        },
                        {
                          "type" : "CIMSolidFill",
                          "enable" : true,
                          "color" : {
                            "type" : "CIMRGBColor",
                            "values" : [
                              130,
                              130,
                              130,
                              100
                            ]
                          }
                        }
                      ],
                      "useRealWorldSymbolSizes" : true
                    }
                  }
                ],
                "respectFrame" : true
              }
            ],
            "useRealWorldSymbolSizes" : true,
            "haloSize" : 0.10936902321506591,
            "scaleX" : 1,
            "angleAlignment" : "Display"
          }
        },
        "defaultSymbolPatch" : "Default",
        "fields" : [
          "PNT_TYPE_"
        ],
        "groups" : [
          {
            "type" : "CIMUniqueValueGroup",
            "classes" : [
              {
                "type" : "CIMUniqueValueClass",
                "label" : "0",
                "patch" : "Default",
                "symbol" : {
                  "type" : "CIMSymbolReference",
                  "symbol" : {
                    "type" : "CIMPointSymbol",
                    "symbolLayers" : [
                      {
                        "type" : "CIMObjectMarker3D",
                        "enable" : true,
                        "anchorPointUnits" : "Relative",
                        "dominantSizeAxis3D" : "Z",
                        "size" : 5,
                        "billboardMode3D" : "None",
                        "modelURI" : "CIMPATH=ObjectMarker3D/08bf87a6862b549c21337e1a054a9ca5.dat",
                        "width" : 5,
                        "depth" : 5,
                        "thumbnail" : "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAJ0AAACdCAYAAACuJnrWAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAAAB3SURBVHhe7cEBDQAAAMKg909tDjcgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD4VwOB0AABB9w5uQAAAABJRU5ErkJggg==",
                        "useAnchorPoint" : true,
                        "lODs" : [
                          {
                            "type" : "CIMObjectMarker3DLOD",
                            "faceCount" : 12
                          }
                        ]
                      }
                    ],
                    "useRealWorldSymbolSizes" : true,
                    "haloSize" : 0.10936902321506591,
                    "scaleX" : 1,
                    "angleAlignment" : "Display"
                  }
                },
                "values" : [
                  {
                    "type" : "CIMUniqueValue",
                    "fieldValues" : [
                      "0"
                    ]
                  }
                ],
                "visible" : true
              },
              {
                "type" : "CIMUniqueValueClass",
                "label" : "1",
                "patch" : "Default",
                "symbol" : {
                  "type" : "CIMSymbolReference",
                  "symbol" : {
                    "type" : "CIMPointSymbol",
                    "symbolLayers" : [
                      {
                        "type" : "CIMObjectMarker3D",
                        "enable" : true,
                        "anchorPoint" : {
                          "x" : 0,
                          "y" : 0,
                          "z" : -0.20000000000000001
                        },
                        "anchorPointUnits" : "Relative",
                        "dominantSizeAxis3D" : "Z",
                        "size" : 3.2232010364532471,
                        "billboardMode3D" : "None",
                        "modelURI" : "CIMPATH=ObjectMarker3D/74713a22678849d97fa38d013785ecfb.dat",
                        "width" : 2.4000000953674316,
                        "depth" : 2.3635380268096924,
                        "thumbnail" : "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAJ0AAACdCAYAAACuJnrWAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAAAB3SURBVHhe7cEBDQAAAMKg909tDjcgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD4VwOB0AABB9w5uQAAAABJRU5ErkJggg==",
                        "useAnchorPoint" : true
                      }
                    ],
                    "useRealWorldSymbolSizes" : true,
                    "haloSize" : 1,
                    "scaleX" : 1,
                    "angleAlignment" : "Display"
                  }
                },
                "values" : [
                  {
                    "type" : "CIMUniqueValue",
                    "fieldValues" : [
                      "1"
                    ]
                  }
                ],
                "visible" : true
              }
            ],
            "heading" : "PNT_TYPE_"
          }
        ],
        "useDefaultSymbol" : true,
        "polygonSymbolColorTarget" : "Fill"
      },
      "scaleSymbols" : true,
      "snappable" : true,
      "useRealWorldSymbolSizes" : true
    }
  ],
  "binaryReferences" : [
    {
      "type" : "CIMBinaryReference",
      "uRI" : "CIMPATH=ObjectMarker3D/08bf87a6862b549c21337e1a054a9ca5.dat",
      "data" : "{\\"version\\":\\"1.2\\",\\"authoringApp\\":\\"ArcGIS Pro\\",\\"authoringAppVersion\\":\\"2.6.0\\",\\"lods\\":[{\\"metrics\\":{\\"faceCount\\":960}}],\\"model\\":{\\"geometries\\":[{\\"type\\":\\"Embedded\\",\\"transformation\\":[1.0,0.0,0.0,0.0,0.0,1.0,0.0,0.0,0.0,0.0,1.0,0.0,0.0,0.0,0.0,1.0],\\"params\\":{\\"topology\\":\\"PerAttributeArray\\",\\"material\\":\\"/materialDefinitions/0\\",\\"texture\\":\\"/textureDefinitions/20fd5397feabb33012b0b3b27f683e3e.dat\\",\\"vertexAttributes\\":{\\"position\\":{\\"valueType\\":\\"Float32\\",\\"valuesPerElement\\":3,\\"values\\":[0.0380598,0.1913416,-0.9807853,-0.0,0.1950901,-0.9807853,0.0,0.0,-1.0,0.0746574,0.1802399,-0.9807853,0.0380598,0.1913416,-0.9807853,0.0,0.0,-1.0,0.1083859,0.1622117,-0.9807853,0.0746574,0.1802399,-0.9807853,0.0,0.0,-1.0,0.1379493,0.1379498,-0.9807853,0.1083859,0.1622117,-0.9807853,0.0,0.0,-1.0,0.1622113,0.1083865,-0.9807853,0.1379493,0.1379498,-0.9807853,0.0,0.0,-1.0,0.1802396,0.0746581,-0.9807853,0.1622113,0.1083865,-0.9807853,0.0,0.0,-1.0,0.1913414,0.0380605,-0.9807853,0.1802396,0.0746581,-0.9807853,0.0,0.0,-1.0,0.1950901,0.0000003,-0.9807853,0.1913414,0.0380605,-0.9807853,0.0,0.0,-1.0,0.1913415,-0.0380599,-0.9807853,0.1950901,0.0000003,-0.9807853,0.0,0.0,-1.0,0.1802398,-0.0746575,-0.9807853,0.1913415,-0.0380599,-0.9807853,0.0,0.0,-1.0,0.1622116,-0.108386,-0.9807853,0.1802398,-0.0746575,-0.9807853,0.0,0.0,-1.0,0.1379497,-0.1379493,-0.9807853,0.1622116,-0.108386,-0.9807853,0.0,0.0,-1.0,0.1083864,-0.1622113,-0.9807853,0.1379497,-0.1379493,-0.9807853,0.0,0.0,-1.0,0.0746579,-0.1802396,-0.9807853,0.1083864,-0.1622113,-0.9807853,0.0,0.0,-1.0,0.0380604,-0.1913414,-0.9807853,0.0746579,-0.1802396,-0.9807853,0.0,0.0,-1.0,0.0000002,-0.1950901,-0.9807853,0.0380604,-0.1913414,-0.9807853,0.0,0.0,-1.0,-0.03806,-0.1913415,-0.9807853,0.0000002,-0.1950901,-0.9807853,0.0,0.0,-1.0,-0.0746576,-0.1802398,-0.9807853,-0.03806,-0.1913415,-0.9807853,0.0,0.0,-1.0,-0.1083861,-0.1622116,-0.9807853,-0.0746576,-0.1802398,-0.9807853,0.0,0.0,-1.0,-0.1379494,-0.1379496,-0.9807853,-0.1083861,-0.1622116,-0.9807853,0.0,0.0,-1.0,-0.1622114,-0.1083863,-0.9807853,-0.1379494,-0.1379496,-0.9807853,0.0,0.0,-1.0,-0.1802397,-0.0746578,-0.9807853,-0.1622114,-0.1083863,-0.9807853,0.0,0.0,-1.0,-0.1913415,-0.0380603,-0.9807853,-0.1802397,-0.0746578,-0.9807853,0.0,0.0,-1.0,-0.1950901,-0.0000001,-0.9807853,-0.1913415,-0.0380603,-0.9807853,0.0,0.0,-1.0,-0.1913415,0.0380601,-0.9807853,-0.1950901,-0.0000001,-0.9807853,0.0,0.0,-1.0,-0.1802398,0.0746577,-0.9807853,-0.1913415,0.0380601,-0.9807853,0.0,0.0,-1.0,-0.1622115,0.1083862,-0.9807853,-0.1802398,0.0746577,-0.9807853,0.0,0.0,-1.0,-0.1379495,0.1379495,-0.9807853,-0.1622115,0.1083862,-0.9807853,0.0,0.0,-1.0,-0.1083862,0.1622115,-0.9807853,-0.1379495,0.1379495,-0.9807853,0.0,0.0,-1.0,-0.0746577,0.1802397,-0.9807853,-0.1083862,0.1622115,-0.9807853,0.0,0.0,-1.0,-0.0380602,0.1913415,-0.9807853,-0.0746577,0.1802397,-0.9807853,0.0,0.0,-1.0,-0.0,0.1950901,-0.9807853,-0.0380602,0.1913415,-0.9807853,0.0,0.0,-1.0,-0.0,0.3826833,-0.9238796,-0.0,0.1950901,-0.9807853,0.074657,0.3753303,-0.9238796,-0.0,0.1950901,-0.9807853,0.0380598,0.1913416,-0.9807853,0.074657,0.3753303,-0.9238796,0.074657,0.3753303,-0.9238796,0.0380598,0.1913416,-0.9807853,0.1464458,0.3535536,-0.9238796,0.0380598,0.1913416,-0.9807853,0.0746574,0.1802399,-0.9807853,0.1464458,0.3535536,-0.9238796,0.1464458,0.3535536,-0.9238796,0.0746574,0.1802399,-0.9807853,0.2126068,0.3181899,-0.9238796,0.0746574,0.1802399,-0.9807853,0.1083859,0.1622117,-0.9807853,0.2126068,0.3181899,-0.9238796,0.2126068,0.3181899,-0.9238796,0.1083859,0.1622117,-0.9807853,0.2705974,0.2705985,-0.9238796,0.1083859,0.1622117,-0.9807853,0.1379493,0.1379498,-0.9807853,0.2705974,0.2705985,-0.9238796,0.2705974,0.2705985,-0.9238796,0.1379493,0.1379498,-0.9807853,0.3181891,0.212608,-0.9238796,0.1379493,0.1379498,-0.9807853,0.1622113,0.1083865,-0.9807853,0.3181891,0.212608,-0.9238796,0.3181891,0.212608,-0.9238796,0.1622113,0.1083865,-0.9807853,0.353553,0.1464472,-0.9238796,0.1622113,0.1083865,-0.9807853,0.1802396,0.0746581,-0.9807853,0.353553,0.1464472,-0.9238796,0.353553,0.1464472,-0.9238796,0.1802396,0.0746581,-0.9807853,0.37533,0.0746584,-0.9238796,0.1802396,0.0746581,-0.9807853,0.1913414,0.0380605,-0.9807853,0.37533,0.0746584,-0.9238796,0.37533,0.0746584,-0.9238796,0.1913414,0.0380605,-0.9807853,0.3826833,0.0000006,-0.9238796,0.1913414,0.0380605,-0.9807853,0.1950901,0.0000003,-0.9807853,0.3826833,0.0000006,-0.9238796,0.3826833,0.0000006,-0.9238796,0.1950901,0.0000003,-0.9807853,0.3753302,-0.0746572,-0.9238796,0.1950901,0.0000003,-0.9807853,0.1913415,-0.0380599,-0.9807853,0.3753302,-0.0746572,-0.9238796,0.3753302,-0.0746572,-0.9238796,0.1913415,-0.0380599,-0.9807853,0.3535535,-0.146446,-0.9238796,0.1913415,-0.0380599,-0.9807853,0.1802398,-0.0746575,-0.9807853,0.3535535,-0.146446,-0.9238796,0.3535535,-0.146446,-0.9238796,0.1802398,-0.0746575,-0.9807853,0.3181898,-0.212607,-0.9238796,0.1802398,-0.0746575,-0.9807853,0.1622116,-0.108386,-0.9807853,0.3181898,-0.212607,-0.9238796,0.3181898,-0.212607,-0.9238796,0.1622116,-0.108386,-0.9807853,0.2705983,-0.2705976,-0.9238796,0.1622116,-0.108386,-0.9807853,0.1379497,-0.1379493,-0.9807853,0.2705983,-0.2705976,-0.9238796,0.2705983,-0.2705976,-0.9238796,0.1379497,-0.1379493,-0.9807853,0.2126078,-0.3181893,-0.9238796,0.1379497,-0.1379493,-0.9807853,0.1083864,-0.1622113,-0.9807853,0.2126078,-0.3181893,-0.9238796,0.2126078,-0.3181893,-0.9238796,0.1083864,-0.1622113,-0.9807853,0.1464469,-0.3535531,-0.9238796,0.1083864,-0.1622113,-0.9807853,0.0746579,-0.1802396,-0.9807853,0.1464469,-0.3535531,-0.9238796,0.1464469,-0.3535531,-0.9238796,0.0746579,-0.1802396,-0.9807853,0.0746582,-0.37533,-0.9238796,0.0746579,-0.1802396,-0.9807853,0.0380604,-0.1913414,-0.9807853,0.0746582,-0.37533,-0.9238796,0.0746582,-0.37533,-0.9238796,0.0380604,-0.1913414,-0.9807853,0.0000004,-0.3826833,-0.9238796,0.0380604,-0.1913414,-0.9807853,0.0000002,-0.1950901,-0.9807853,0.0000004,-0.3826833,-0.9238796,0.0000004,-0.3826833,-0.9238796,0.0000002,-0.1950901,-0.9807853,-0.0746575,-0.3753302,-0.9238796,0.0000002,-0.1950901,-0.9807853,-0.03806,-0.1913415,-0.9807853,-0.0746575,-0.3753302,-0.9238796,-0.0746575,-0.3753302,-0.9238796,-0.03806,-0.1913415,-0.9807853,-0.1464463,-0.3535534,-0.9238796,-0.03806,-0.1913415,-0.9807853,-0.0746576,-0.1802398,-0.9807853,-0.1464463,-0.3535534,-0.9238796,-0.1464463,-0.3535534,-0.9238796,-0.0746576,-0.1802398,-0.9807853,-0.2126072,-0.3181897,-0.9238796,-0.0746576,-0.1802398,-0.9807853,-0.1083861,-0.1622116,-0.9807853,-0.2126072,-0.3181897,-0.9238796,-0.2126072,-0.3181897,-0.9238796,-0.1083861,-0.1622116,-0.9807853,-0.2705978,-0.2705981,-0.9238796,-0.1083861,-0.1622116,-0.9807853,-0.1379494,-0.1379496,-0.9807853,-0.2705978,-0.2705981,-0.9238796,-0.2705978,-0.2705981,-0.9238796,-0.1379494,-0.1379496,-0.9807853,-0.3181894,-0.2126076,-0.9238796,-0.1379494,-0.1379496,-0.9807853,-0.1622114,-0.1083863,-0.9807853,-0.3181894,-0.2126076,-0.9238796,-0.3181894,-0.2126076,-0.9238796,-0.1622114,-0.1083863,-0.9807853,-0.3535532,-0.1464467,-0.9238796,-0.1622114,-0.1083863,-0.9807853,-0.1802397,-0.0746578,-0.9807853,-0.3535532,-0.1464467,-0.9238796,-0.3535532,-0.1464467,-0.9238796,-0.1802397,-0.0746578,-0.9807853,-0.3753301,-0.074658,-0.9238796,-0.1802397,-0.0746578,-0.9807853,-0.1913415,-0.0380603,-0.9807853,-0.3753301,-0.074658,-0.9238796,-0.3753301,-0.074658,-0.9238796,-0.1913415,-0.0380603,-0.9807853,-0.3826833,-0.0000001,-0.9238796,-0.1913415,-0.0380603,-0.9807853,-0.1950901,-0.0000001,-0.9807853,-0.3826833,-0.0000001,-0.9238796,-0.3826833,-0.0000001,-0.9238796,-0.1950901,-0.0000001,-0.9807853,-0.3753302,0.0746577,-0.9238796,-0.1950901,-0.0000001,-0.9807853,-0.1913415,0.0380601,-0.9807853,-0.3753302,0.0746577,-0.9238796,-0.3753302,0.0746577,-0.9238796,-0.1913415,0.0380601,-0.9807853,-0.3535533,0.1464465,-0.9238796,-0.1913415,0.0380601,-0.9807853,-0.1802398,0.0746577,-0.9807853,-0.3535533,0.1464465,-0.9238796,-0.3535533,0.1464465,-0.9238796,-0.1802398,0.0746577,-0.9807853,-0.3181895,0.2126074,-0.9238796,-0.1802398,0.0746577,-0.9807853,-0.1622115,0.1083862,-0.9807853,-0.3181895,0.2126074,-0.9238796,-0.3181895,0.2126074,-0.9238796,-0.1622115,0.1083862,-0.9807853,-0.2705979,0.2705979,-0.9238796,-0.1622115,0.1083862,-0.9807853,-0.1379495,0.1379495,-0.9807853,-0.2705979,0.2705979,-0.9238796,-0.2705979,0.2705979,-0.9238796,-0.1379495,0.1379495,-0.9807853,-0.2126074,0.3181895,-0.9238796,-0.1379495,0.1379495,-0.9807853,-0.1083862,0.1622115,-0.9807853,-0.2126074,0.3181895,-0.9238796,-0.2126074,0.3181895,-0.9238796,-0.1083862,0.1622115,-0.9807853,-0.1464465,0.3535533,-0.9238796,-0.1083862,0.1622115,-0.9807853,-0.0746577,0.1802397,-0.9807853,-0.1464465,0.3535533,-0.9238796,-0.1464465,0.3535533,-0.9238796,-0.0746577,0.1802397,-0.9807853,-0.0746578,0.3753301,-0.9238796,-0.0746577,0.1802397,-0.9807853,-0.0380602,0.1913415,-0.9807853,-0.0746578,0.3753301,-0.9238796,-0.0746578,0.3753301,-0.9238796,-0.0380602,0.1913415,-0.9807853,-0.0,0.3826833,-0.9238796,-0.0380602,0.1913415,-0.9807853,-0.0,0.1950901,-0.9807853,-0.0,0.3826833,-0.9238796,-0.0,0.5555702,-0.8314697,-0.0,0.3826833,-0.9238796,0.1083852,0.5448953,-0.8314697,-0.0,0.3826833,-0.9238796,0.074657,0.3753303,-0.9238796,0.1083852,0.5448953,-0.8314697,0.1083852,0.5448953,-0.8314697,0.074657,0.3753303,-0.9238796,0.2126064,0.5132804,-0.8314697,0.074657,0.3753303,-0.9238796,0.1464458,0.3535536,-0.9238796,0.2126064,0.5132804,-0.8314697,0.2126064,0.5132804,-0.8314697,0.1464458,0.3535536,-0.9238796,0.3086573,0.4619403,-0.8314697,0.1464458,0.3535536,-0.9238796,0.2126068,0.3181899,-0.9238796,0.3086573,0.4619403,-0.8314697,0.3086573,0.4619403,-0.8314697,0.2126068,0.3181899,-0.9238796,0.3928467,0.3928482,-0.8314697,0.2126068,0.3181899,-0.9238796,0.2705974,0.2705985,-0.9238796,0.3928467,0.3928482,-0.8314697,0.3928467,0.3928482,-0.8314697,0.2705974,0.2705985,-0.9238796,0.4619392,0.3086591,-0.8314697,0.2705974,0.2705985,-0.9238796,0.3181891,0.212608,-0.9238796,0.4619392,0.3086591,-0.8314697,0.4619392,0.3086591,-0.8314697,0.3181891,0.212608,-0.9238796,0.5132796,0.2126084,-0.8314697,0.3181891,0.212608,-0.9238796,0.353553,0.1464472,-0.9238796,0.5132796,0.2126084,-0.8314697,0.5132796,0.2126084,-0.8314697,0.353553,0.1464472,-0.9238796,0.5448949,0.1083873,-0.8314697,0.353553,0.1464472,-0.9238796,0.37533,0.0746584,-0.9238796,0.5448949,0.1083873,-0.8314697,0.5448949,0.1083873,-0.8314697,0.37533,0.0746584,-0.9238796,0.5555702,0.0000009,-0.8314697,0.37533,0.0746584,-0.9238796,0.3826833,0.0000006,-0.9238796,0.5555702,0.0000009,-0.8314697,0.5555702,0.0000009,-0.8314697,0.3826833,0.0000006,-0.9238796,0.5448952,-0.1083855,-0.8314697,0.3826833,0.0000006,-0.9238796,0.3753302,-0.0746572,-0.9238796,0.5448952,-0.1083855,-0.8314697,0.5448952,-0.1083855,-0.8314697,0.3753302,-0.0746572,-0.9238796,0.5132802,-0.2126068,-0.8314697,0.3753302,-0.0746572,-0.9238796,0.3535535,-0.146446,-0.9238796,0.5132802,-0.2126068,-0.8314697,0.5132802,-0.2126068,-0.8314697,0.3535535,-0.146446,-0.9238796,0.4619401,-0.3086576,-0.8314697,0.3535535,-0.146446,-0.9238796,0.3181898,-0.212607,-0.9238796,0.4619401,-0.3086576,-0.8314697,0.4619401,-0.3086576,-0.8314697,0.3181898,-0.212607,-0.9238796,0.3928479,-0.3928469,-0.8314697,0.3181898,-0.212607,-0.9238796,0.2705983,-0.2705976,-0.9238796,0.3928479,-0.3928469,-0.8314697,0.3928479,-0.3928469,-0.8314697,0.2705983,-0.2705976,-0.9238796,0.3086588,-0.4619394,-0.8314697,0.2705983,-0.2705976,-0.9238796,0.2126078,-0.3181893,-0.9238796,0.3086588,-0.4619394,-0.8314697,0.3086588,-0.4619394,-0.8314697,0.2126078,-0.3181893,-0.9238796,0.2126081,-0.5132797,-0.8314697,0.2126078,-0.3181893,-0.9238796,0.1464469,-0.3535531,-0.9238796,0.2126081,-0.5132797,-0.8314697,0.2126081,-0.5132797,-0.8314697,0.1464469,-0.3535531,-0.9238796,0.1083869,-0.5448949,-0.8314697,0.1464469,-0.3535531,-0.9238796,0.0746582,-0.37533,-0.9238796,0.1083869,-0.5448949,-0.8314697,0.1083869,-0.5448949,-0.8314697,0.0746582,-0.37533,-0.9238796,0.0000005,-0.5555702,-0.8314697,0.0746582,-0.37533,-0.9238796,0.0000004,-0.3826833,-0.9238796,0.0000005,-0.5555702,-0.8314697,0.0000005,-0.5555702,-0.8314697,0.0000004,-0.3826833,-0.9238796,-0.1083859,-0.5448952,-0.8314697,0.0000004,-0.3826833,-0.9238796,-0.0746575,-0.3753302,-0.9238796,-0.1083859,-0.5448952,-0.8314697,-0.1083859,-0.5448952,-0.8314697,-0.0746575,-0.3753302,-0.9238796,-0.2126071,-0.5132801,-0.8314697,-0.0746575,-0.3753302,-0.9238796,-0.1464463,-0.3535534,-0.9238796,-0.2126071,-0.5132801,-0.8314697,-0.2126071,-0.5132801,-0.8314697,-0.1464463,-0.3535534,-0.9238796,-0.3086579,-0.46194,-0.8314697,-0.1464463,-0.3535534,-0.9238796,-0.2126072,-0.3181897,-0.9238796,-0.3086579,-0.46194,-0.8314697,-0.3086579,-0.46194,-0.8314697,-0.2126072,-0.3181897,-0.9238796,-0.3928472,-0.3928477,-0.8314697,-0.2126072,-0.3181897,-0.9238796,-0.2705978,-0.2705981,-0.9238796,-0.3928472,-0.3928477,-0.8314697,-0.3928472,-0.3928477,-0.8314697,-0.2705978,-0.2705981,-0.9238796,-0.4619395,-0.3086585,-0.8314697,-0.2705978,-0.2705981,-0.9238796,-0.3181894,-0.2126076,-0.9238796,-0.4619395,-0.3086585,-0.8314697,-0.4619395,-0.3086585,-0.8314697,-0.3181894,-0.2126076,-0.9238796,-0.5132798,-0.2126078,-0.8314697,-0.3181894,-0.2126076,-0.9238796,-0.3535532,-0.1464467,-0.9238796,-0.5132798,-0.2126078,-0.8314697,-0.5132798,-0.2126078,-0.8314697,-0.3535532,-0.1464467,-0.9238796,-0.544895,-0.1083866,-0.8314697,-0.3535532,-0.1464467,-0.9238796,-0.3753301,-0.074658,-0.9238796,-0.544895,-0.1083866,-0.8314697,-0.544895,-0.1083866,-0.8314697,-0.3753301,-0.074658,-0.9238796,-0.5555702,-0.0000002,-0.8314697,-0.3753301,-0.074658,-0.9238796,-0.3826833,-0.0000001,-0.9238796,-0.5555702,-0.0000002,-0.8314697,-0.5555702,-0.0000002,-0.8314697,-0.3826833,-0.0000001,-0.9238796,-0.5448951,0.1083862,-0.8314697,-0.3826833,-0.0000001,-0.9238796,-0.3753302,0.0746577,-0.9238796,-0.5448951,0.1083862,-0.8314697,-0.5448951,0.1083862,-0.8314697,-0.3753302,0.0746577,-0.9238796,-0.51328,0.2126074,-0.8314697,-0.3753302,0.0746577,-0.9238796,-0.3535533,0.1464465,-0.9238796,-0.51328,0.2126074,-0.8314697,-0.51328,0.2126074,-0.8314697,-0.3535533,0.1464465,-0.9238796,-0.4619398,0.3086582,-0.8314697,-0.3535533,0.1464465,-0.9238796,-0.3181895,0.2126074,-0.9238796,-0.4619398,0.3086582,-0.8314697,-0.4619398,0.3086582,-0.8314697,-0.3181895,0.2126074,-0.9238796,-0.3928474,0.3928474,-0.8314697,-0.3181895,0.2126074,-0.9238796,-0.2705979,0.2705979,-0.9238796,-0.3928474,0.3928474,-0.8314697,-0.3928474,0.3928474,-0.8314697,-0.2705979,0.2705979,-0.9238796,-0.3086582,0.4619398,-0.8314697,-0.2705979,0.2705979,-0.9238796,-0.2126074,0.3181895,-0.9238796,-0.3086582,0.4619398,-0.8314697,-0.3086582,0.4619398,-0.8314697,-0.2126074,0.3181895,-0.9238796,-0.2126075,0.5132799,-0.8314697,-0.2126074,0.3181895,-0.9238796,-0.1464465,0.3535533,-0.9238796,-0.2126075,0.5132799,-0.8314697,-0.2126075,0.5132799,-0.8314697,-0.1464465,0.3535533,-0.9238796,-0.1083864,0.5448951,-0.8314697,-0.1464465,0.3535533,-0.9238796,-0.0746578,0.3753301,-0.9238796,-0.1083864,0.5448951,-0.8314697,-0.1083864,0.5448951,-0.8314697,-0.0746578,0.3753301,-0.9238796,-0.0,0.5555702,-0.8314697,-0.0746578,0.3753301,-0.9238796,-0.0,0.3826833,-0.9238796,-0.0,0.5555702,-0.8314697,-0.0,0.7071068,-0.7071068,-0.0,0.5555702,-0.8314697,0.1379482,0.6935202,-0.7071068,-0.0,0.5555702,-0.8314697,0.1083852,0.5448953,-0.8314697,0.1379482,0.6935202,-0.7071068,0.1379482,0.6935202,-0.7071068,0.1083852,0.5448953,-0.8314697,0.2705967,0.653282,-0.7071068,0.1083852,0.5448953,-0.8314697,0.2126064,0.5132804,-0.8314697,0.2705967,0.653282,-0.7071068,0.2705967,0.653282,-0.7071068,0.2126064,0.5132804,-0.8314697,0.3928463,0.5879386,-0.7071068,0.2126064,0.5132804,-0.8314697,0.3086573,0.4619403,-0.8314697,0.3928463,0.5879386,-0.7071068,0.3928463,0.5879386,-0.7071068,0.3086573,0.4619403,-0.8314697,0.499999,0.500001,-0.7071068,0.3086573,0.4619403,-0.8314697,0.3928467,0.3928482,-0.8314697,0.499999,0.500001,-0.7071068,0.499999,0.500001,-0.7071068,0.3928467,0.3928482,-0.8314697,0.5879371,0.3928486,-0.7071068,0.3928467,0.3928482,-0.8314697,0.4619392,0.3086591,-0.8314697,0.5879371,0.3928486,-0.7071068,0.5879371,0.3928486,-0.7071068,0.4619392,0.3086591,-0.8314697,0.653281,0.2705992,-0.7071068,0.4619392,0.3086591,-0.8314697,0.5132796,0.2126084,-0.8314697,0.653281,0.2705992,-0.7071068,0.653281,0.2705992,-0.7071068,0.5132796,0.2126084,-0.8314697,0.6935197,0.1379509,-0.7071068,0.5132796,0.2126084,-0.8314697,0.5448949,0.1083873,-0.8314697,0.6935197,0.1379509,-0.7071068,0.6935197,0.1379509,-0.7071068,0.5448949,0.1083873,-0.8314697,0.7071068,0.0000011,-0.7071068,0.5448949,0.1083873,-0.8314697,0.5555702,0.0000009,-0.8314697,0.7071068,0.0000011,-0.7071068,0.7071068,0.0000011,-0.7071068,0.5555702,0.0000009,-0.8314697,0.6935201,-0.1379486,-0.7071068,0.5555702,0.0000009,-0.8314697,0.5448952,-0.1083855,-0.8314697,0.6935201,-0.1379486,-0.7071068,0.6935201,-0.1379486,-0.7071068,0.5448952,-0.1083855,-0.8314697,0.6532819,-0.2705971,-0.7071068,0.5448952,-0.1083855,-0.8314697,0.5132802,-0.2126068,-0.8314697,0.6532819,-0.2705971,-0.7071068,0.6532819,-0.2705971,-0.7071068,0.5132802,-0.2126068,-0.8314697,0.5879383,-0.3928467,-0.7071068,0.5132802,-0.2126068,-0.8314697,0.4619401,-0.3086576,-0.8314697,0.5879383,-0.3928467,-0.7071068,0.5879383,-0.3928467,-0.7071068,0.4619401,-0.3086576,-0.8314697,0.5000006,-0.4999993,-0.7071068,0.4619401,-0.3086576,-0.8314697,0.3928479,-0.3928469,-0.8314697,0.5000006,-0.4999993,-0.7071068,0.5000006,-0.4999993,-0.7071068,0.3928479,-0.3928469,-0.8314697,0.3928482,-0.5879373,-0.7071068,0.3928479,-0.3928469,-0.8314697,0.3086588,-0.4619394,-0.8314697,0.3928482,-0.5879373,-0.7071068,0.3928482,-0.5879373,-0.7071068,0.3086588,-0.4619394,-0.8314697,0.2705988,-0.6532812,-0.7071068,0.3086588,-0.4619394,-0.8314697,0.2126081,-0.5132797,-0.8314697,0.2705988,-0.6532812,-0.7071068,0.2705988,-0.6532812,-0.7071068,0.2126081,-0.5132797,-0.8314697,0.1379504,-0.6935198,-0.7071068,0.2126081,-0.5132797,-0.8314697,0.1083869,-0.5448949,-0.8314697,0.1379504,-0.6935198,-0.7071068,0.1379504,-0.6935198,-0.7071068,0.1083869,-0.5448949,-0.8314697,0.0000007,-0.7071068,-0.7071068,0.1083869,-0.5448949,-0.8314697,0.0000005,-0.5555702,-0.8314697,0.0000007,-0.7071068,-0.7071068,0.0000007,-0.7071068,-0.7071068,0.0000005,-0.5555702,-0.8314697,-0.1379491,-0.69352,-0.7071068,0.0000005,-0.5555702,-0.8314697,-0.1083859,-0.5448952,-0.8314697,-0.1379491,-0.69352,-0.7071068,-0.1379491,-0.69352,-0.7071068,-0.1083859,-0.5448952,-0.8314697,-0.2705975,-0.6532817,-0.7071068,-0.1083859,-0.5448952,-0.8314697,-0.2126071,-0.5132801,-0.8314697,-0.2705975,-0.6532817,-0.7071068,-0.2705975,-0.6532817,-0.7071068,-0.2126071,-0.5132801,-0.8314697,-0.3928471,-0.5879381,-0.7071068,-0.2126071,-0.5132801,-0.8314697,-0.3086579,-0.46194,-0.8314697,-0.3928471,-0.5879381,-0.7071068,-0.3928471,-0.5879381,-0.7071068,-0.3086579,-0.46194,-0.8314697,-0.4999997,-0.5000003,-0.7071068,-0.3086579,-0.46194,-0.8314697,-0.3928472,-0.3928477,-0.8314697,-0.4999997,-0.5000003,-0.7071068,-0.4999997,-0.5000003,-0.7071068,-0.3928472,-0.3928477,-0.8314697,-0.5879376,-0.3928478,-0.7071068,-0.3928472,-0.3928477,-0.8314697,-0.4619395,-0.3086585,-0.8314697,-0.5879376,-0.3928478,-0.7071068,-0.5879376,-0.3928478,-0.7071068,-0.4619395,-0.3086585,-0.8314697,-0.6532813,-0.2705984,-0.7071068,-0.4619395,-0.3086585,-0.8314697,-0.5132798,-0.2126078,-0.8314697,-0.6532813,-0.2705984,-0.7071068,-0.6532813,-0.2705984,-0.7071068,-0.5132798,-0.2126078,-0.8314697,-0.6935198,-0.13795,-0.7071068,-0.5132798,-0.2126078,-0.8314697,-0.544895,-0.1083866,-0.8314697,-0.6935198,-0.13795,-0.7071068,-0.6935198,-0.13795,-0.7071068,-0.544895,-0.1083866,-0.8314697,-0.7071068,-0.0000002,-0.7071068,-0.544895,-0.1083866,-0.8314697,-0.5555702,-0.0000002,-0.8314697,-0.7071068,-0.0000002,-0.7071068,-0.7071068,-0.0000002,-0.7071068,-0.5555702,-0.0000002,-0.8314697,-0.6935199,0.1379495,-0.7071068,-0.5555702,-0.0000002,-0.8314697,-0.5448951,0.1083862,-0.8314697,-0.6935199,0.1379495,-0.7071068,-0.6935199,0.1379495,-0.7071068,-0.5448951,0.1083862,-0.8314697,-0.6532815,0.2705979,-0.7071068,-0.5448951,0.1083862,-0.8314697,-0.51328,0.2126074,-0.8314697,-0.6532815,0.2705979,-0.7071068,-0.6532815,0.2705979,-0.7071068,-0.51328,0.2126074,-0.8314697,-0.5879378,0.3928474,-0.7071068,-0.51328,0.2126074,-0.8314697,-0.4619398,0.3086582,-0.8314697,-0.5879378,0.3928474,-0.7071068,-0.5879378,0.3928474,-0.7071068,-0.4619398,0.3086582,-0.8314697,-0.5,0.5,-0.7071068,-0.4619398,0.3086582,-0.8314697,-0.3928474,0.3928474,-0.8314697,-0.5,0.5,-0.7071068,-0.5,0.5,-0.7071068,-0.3928474,0.3928474,-0.8314697,-0.3928474,0.5879378,-0.7071068,-0.3928474,0.3928474,-0.8314697,-0.3086582,0.4619398,-0.8314697,-0.3928474,0.5879378,-0.7071068,-0.3928474,0.5879378,-0.7071068,-0.3086582,0.4619398,-0.8314697,-0.270598,0.6532815,-0.7071068,-0.3086582,0.4619398,-0.8314697,-0.2126075,0.5132799,-0.8314697,-0.270598,0.6532815,-0.7071068,-0.270598,0.6532815,-0.7071068,-0.2126075,0.5132799,-0.8314697,-0.1379497,0.6935199,-0.7071068,-0.2126075,0.5132799,-0.8314697,-0.1083864,0.5448951,-0.8314697,-0.1379497,0.6935199,-0.7071068,-0.1379497,0.6935199,-0.7071068,-0.1083864,0.5448951,-0.8314697,-0.0,0.7071068,-0.7071068,-0.1083864,0.5448951,-0.8314697,-0.0,0.5555702,-0.8314697,-0.0,0.7071068,-0.7071068,-0.0,0.8314697,-0.5555702,-0.0,0.7071068,-0.7071068,0.1622099,0.8154936,-0.5555702,-0.0,0.7071068,-0.7071068,0.1379482,0.6935202,-0.7071068,0.1622099,0.8154936,-0.5555702,0.1622099,0.8154936,-0.5555702,0.1379482,0.6935202,-0.7071068,0.3181881,0.7681785,-0.5555702,0.1379482,0.6935202,-0.7071068,0.2705967,0.653282,-0.7071068,0.3181881,0.7681785,-0.5555702,0.3181881,0.7681785,-0.5555702,0.2705967,0.653282,-0.7071068,0.4619384,0.6913427,-0.5555702,0.2705967,0.653282,-0.7071068,0.3928463,0.5879386,-0.7071068,0.4619384,0.6913427,-0.5555702,0.4619384,0.6913427,-0.5555702,0.3928463,0.5879386,-0.7071068,0.5879367,0.587939,-0.5555702,0.3928463,0.5879386,-0.7071068,0.499999,0.500001,-0.7071068,0.5879367,0.587939,-0.5555702,0.5879367,0.587939,-0.5555702,0.499999,0.500001,-0.7071068,0.6913409,0.4619411,-0.5555702,0.499999,0.500001,-0.7071068,0.5879371,0.3928486,-0.7071068,0.6913409,0.4619411,-0.5555702,0.6913409,0.4619411,-0.5555702,0.5879371,0.3928486,-0.7071068,0.7681772,0.318191,-0.5555702,0.5879371,0.3928486,-0.7071068,0.653281,0.2705992,-0.7071068,0.7681772,0.318191,-0.5555702,0.7681772,0.318191,-0.5555702,0.653281,0.2705992,-0.7071068,0.8154929,0.1622131,-0.5555702,0.653281,0.2705992,-0.7071068,0.6935197,0.1379509,-0.7071068,0.8154929,0.1622131,-0.5555702,0.8154929,0.1622131,-0.5555702,0.6935197,0.1379509,-0.7071068,0.8314697,0.0000013,-0.5555702,0.6935197,0.1379509,-0.7071068,0.7071068,0.0000011,-0.7071068,0.8314697,0.0000013,-0.5555702,0.8314697,0.0000013,-0.5555702,0.7071068,0.0000011,-0.7071068,0.8154934,-0.1622104,-0.5555702,0.7071068,0.0000011,-0.7071068,0.6935201,-0.1379486,-0.7071068,0.8154934,-0.1622104,-0.5555702,0.8154934,-0.1622104,-0.5555702,0.6935201,-0.1379486,-0.7071068,0.7681783,-0.3181885,-0.5555702,0.6935201,-0.1379486,-0.7071068,0.6532819,-0.2705971,-0.7071068,0.7681783,-0.3181885,-0.5555702,0.7681783,-0.3181885,-0.5555702,0.6532819,-0.2705971,-0.7071068,0.6913424,-0.4619389,-0.5555702,0.6532819,-0.2705971,-0.7071068,0.5879383,-0.3928467,-0.7071068,0.6913424,-0.4619389,-0.5555702,0.6913424,-0.4619389,-0.5555702,0.5879383,-0.3928467,-0.7071068,0.5879385,-0.5879371,-0.5555702,0.5879383,-0.3928467,-0.7071068,0.5000006,-0.4999993,-0.7071068,0.5879385,-0.5879371,-0.5555702,0.5879385,-0.5879371,-0.5555702,0.5000006,-0.4999993,-0.7071068,0.4619406,-0.6913412,-0.5555702,0.5000006,-0.4999993,-0.7071068,0.3928482,-0.5879373,-0.7071068,0.4619406,-0.6913412,-0.5555702,0.4619406,-0.6913412,-0.5555702,0.3928482,-0.5879373,-0.7071068,0.3181905,-0.7681774,-0.5555702,0.3928482,-0.5879373,-0.7071068,0.2705988,-0.6532812,-0.7071068,0.3181905,-0.7681774,-0.5555702,0.3181905,-0.7681774,-0.5555702,0.2705988,-0.6532812,-0.7071068,0.1622125,-0.815493,-0.5555702,0.2705988,-0.6532812,-0.7071068,0.1379504,-0.6935198,-0.7071068,0.1622125,-0.815493,-0.5555702,0.1622125,-0.815493,-0.5555702,0.1379504,-0.6935198,-0.7071068,0.0000008,-0.8314697,-0.5555702,0.1379504,-0.6935198,-0.7071068,0.0000007,-0.7071068,-0.7071068,0.0000008,-0.8314697,-0.5555702,0.0000008,-0.8314697,-0.5555702,0.0000007,-0.7071068,-0.7071068,-0.162211,-0.8154933,-0.5555702,0.0000007,-0.7071068,-0.7071068,-0.1379491,-0.69352,-0.7071068,-0.162211,-0.8154933,-0.5555702,-0.162211,-0.8154933,-0.5555702,-0.1379491,-0.69352,-0.7071068,-0.318189,-0.768178,-0.5555702,-0.1379491,-0.69352,-0.7071068,-0.2705975,-0.6532817,-0.7071068,-0.318189,-0.768178,-0.5555702,-0.318189,-0.768178,-0.5555702,-0.2705975,-0.6532817,-0.7071068,-0.4619393,-0.6913421,-0.5555702,-0.2705975,-0.6532817,-0.7071068,-0.3928471,-0.5879381,-0.7071068,-0.4619393,-0.6913421,-0.5555702,-0.4619393,-0.6913421,-0.5555702,-0.3928471,-0.5879381,-0.7071068,-0.5879375,-0.5879382,-0.5555702,-0.3928471,-0.5879381,-0.7071068,-0.4999997,-0.5000003,-0.7071068,-0.5879375,-0.5879382,-0.5555702,-0.5879375,-0.5879382,-0.5555702,-0.4999997,-0.5000003,-0.7071068,-0.6913415,-0.4619402,-0.5555702,-0.4999997,-0.5000003,-0.7071068,-0.5879376,-0.3928478,-0.7071068,-0.6913415,-0.4619402,-0.5555702,-0.6913415,-0.4619402,-0.5555702,-0.5879376,-0.3928478,-0.7071068,-0.7681776,-0.31819,-0.5555702,-0.5879376,-0.3928478,-0.7071068,-0.6532813,-0.2705984,-0.7071068,-0.7681776,-0.31819,-0.5555702,-0.7681776,-0.31819,-0.5555702,-0.6532813,-0.2705984,-0.7071068,-0.8154931,-0.162212,-0.5555702,-0.6532813,-0.2705984,-0.7071068,-0.6935198,-0.13795,-0.7071068,-0.8154931,-0.162212,-0.5555702,-0.8154931,-0.162212,-0.5555702,-0.6935198,-0.13795,-0.7071068,-0.8314697,-0.0000003,-0.5555702,-0.6935198,-0.13795,-0.7071068,-0.7071068,-0.0000002,-0.7071068,-0.8314697,-0.0000003,-0.5555702,-0.8314697,-0.0000003,-0.5555702,-0.7071068,-0.0000002,-0.7071068,-0.8154932,0.1622115,-0.5555702,-0.7071068,-0.0000002,-0.7071068,-0.6935199,0.1379495,-0.7071068,-0.8154932,0.1622115,-0.5555702,-0.8154932,0.1622115,-0.5555702,-0.6935199,0.1379495,-0.7071068,-0.7681779,0.3181895,-0.5555702,-0.6935199,0.1379495,-0.7071068,-0.6532815,0.2705979,-0.7071068,-0.7681779,0.3181895,-0.5555702,-0.7681779,0.3181895,-0.5555702,-0.6532815,0.2705979,-0.7071068,-0.6913418,0.4619398,-0.5555702,-0.6532815,0.2705979,-0.7071068,-0.5879378,0.3928474,-0.7071068,-0.6913418,0.4619398,-0.5555702,-0.6913418,0.4619398,-0.5555702,-0.5879378,0.3928474,-0.7071068,-0.5879378,0.5879378,-0.5555702,-0.5879378,0.3928474,-0.7071068,-0.5,0.5,-0.7071068,-0.5879378,0.5879378,-0.5555702,-0.5879378,0.5879378,-0.5555702,-0.5,0.5,-0.7071068,-0.4619398,0.6913418,-0.5555702,-0.5,0.5,-0.7071068,-0.3928474,0.5879378,-0.7071068,-0.4619398,0.6913418,-0.5555702,-0.4619398,0.6913418,-0.5555702,-0.3928474,0.5879378,-0.7071068,-0.3181896,0.7681778,-0.5555702,-0.3928474,0.5879378,-0.7071068,-0.270598,0.6532815,-0.7071068,-0.3181896,0.7681778,-0.5555702,-0.3181896,0.7681778,-0.5555702,-0.270598,0.6532815,-0.7071068,-0.1622117,0.8154932,-0.5555702,-0.270598,0.6532815,-0.7071068,-0.1379497,0.6935199,-0.7071068,-0.1622117,0.8154932,-0.5555702,-0.1622117,0.8154932,-0.5555702,-0.1379497,0.6935199,-0.7071068,-0.0,0.8314697,-0.5555702,-0.1379497,0.6935199,-0.7071068,-0.0,0.7071068,-0.7071068,-0.0,0.8314697,-0.5555702,-0.0,0.9238796,-0.3826833,-0.0,0.8314697,-0.5555702,0.180238,0.9061279,-0.3826833,-0.0,0.8314697,-0.5555702,0.1622099,0.8154936,-0.5555702,0.180238,0.9061279,-0.3826833,0.180238,0.9061279,-0.3826833,0.1622099,0.8154936,-0.5555702,0.3535516,0.8535542,-0.3826833,0.1622099,0.8154936,-0.5555702,0.3181881,0.7681785,-0.5555702,0.3535516,0.8535542,-0.3826833,0.3535516,0.8535542,-0.3826833,0.3181881,0.7681785,-0.5555702,0.5132784,0.7681788,-0.3826833,0.3181881,0.7681785,-0.5555702,0.4619384,0.6913427,-0.5555702,0.5132784,0.7681788,-0.3826833,0.5132784,0.7681788,-0.3826833,0.4619384,0.6913427,-0.5555702,0.6532802,0.6532828,-0.3826833,0.4619384,0.6913427,-0.5555702,0.5879367,0.587939,-0.5555702,0.6532802,0.6532828,-0.3826833,0.6532802,0.6532828,-0.3826833,0.5879367,0.587939,-0.5555702,0.7681769,0.5132814,-0.3826833,0.5879367,0.587939,-0.5555702,0.6913409,0.4619411,-0.5555702,0.7681769,0.5132814,-0.3826833,0.7681769,0.5132814,-0.3826833,0.6913409,0.4619411,-0.5555702,0.8535528,0.3535549,-0.3826833,0.6913409,0.4619411,-0.5555702,0.7681772,0.318191,-0.5555702,0.8535528,0.3535549,-0.3826833,0.8535528,0.3535549,-0.3826833,0.7681772,0.318191,-0.5555702,0.9061272,0.1802415,-0.3826833,0.7681772,0.318191,-0.5555702,0.8154929,0.1622131,-0.5555702,0.9061272,0.1802415,-0.3826833,0.9061272,0.1802415,-0.3826833,0.8154929,0.1622131,-0.5555702,0.9238796,0.0000015,-0.3826833,0.8154929,0.1622131,-0.5555702,0.8314697,0.0000013,-0.5555702,0.9238796,0.0000015,-0.3826833,0.9238796,0.0000015,-0.3826833,0.8314697,0.0000013,-0.5555702,0.9061278,-0.1802386,-0.3826833,0.8314697,0.0000013,-0.5555702,0.8154934,-0.1622104,-0.5555702,0.9061278,-0.1802386,-0.3826833,0.9061278,-0.1802386,-0.3826833,0.8154934,-0.1622104,-0.5555702,0.853554,-0.3535522,-0.3826833,0.8154934,-0.1622104,-0.5555702,0.7681783,-0.3181885,-0.5555702,0.853554,-0.3535522,-0.3826833,0.853554,-0.3535522,-0.3826833,0.7681783,-0.3181885,-0.5555702,0.7681785,-0.513279,-0.3826833,0.7681783,-0.3181885,-0.5555702,0.6913424,-0.4619389,-0.5555702,0.7681785,-0.513279,-0.3826833,0.7681785,-0.513279,-0.3826833,0.6913424,-0.4619389,-0.5555702,0.6532823,-0.6532807,-0.3826833,0.6913424,-0.4619389,-0.5555702,0.5879385,-0.5879371,-0.5555702,0.6532823,-0.6532807,-0.3826833,0.6532823,-0.6532807,-0.3826833,0.5879385,-0.5879371,-0.5555702,0.5132809,-0.7681772,-0.3826833,0.5879385,-0.5879371,-0.5555702,0.4619406,-0.6913412,-0.5555702,0.5132809,-0.7681772,-0.3826833,0.5132809,-0.7681772,-0.3826833,0.4619406,-0.6913412,-0.5555702,0.3535544,-0.853553,-0.3826833,0.4619406,-0.6913412,-0.5555702,0.3181905,-0.7681774,-0.5555702,0.3535544,-0.853553,-0.3826833,0.3535544,-0.853553,-0.3826833,0.3181905,-0.7681774,-0.5555702,0.1802409,-0.9061273,-0.3826833,0.3181905,-0.7681774,-0.5555702,0.1622125,-0.815493,-0.5555702,0.1802409,-0.9061273,-0.3826833,0.1802409,-0.9061273,-0.3826833,0.1622125,-0.815493,-0.5555702,0.0000009,-0.9238796,-0.3826833,0.1622125,-0.815493,-0.5555702,0.0000008,-0.8314697,-0.5555702,0.0000009,-0.9238796,-0.3826833,0.0000009,-0.9238796,-0.3826833,0.0000008,-0.8314697,-0.5555702,-0.1802392,-0.9061276,-0.3826833,0.0000008,-0.8314697,-0.5555702,-0.162211,-0.8154933,-0.5555702,-0.1802392,-0.9061276,-0.3826833,-0.1802392,-0.9061276,-0.3826833,-0.162211,-0.8154933,-0.5555702,-0.3535527,-0.8535537,-0.3826833,-0.162211,-0.8154933,-0.5555702,-0.318189,-0.768178,-0.5555702,-0.3535527,-0.8535537,-0.3826833,-0.3535527,-0.8535537,-0.3826833,-0.318189,-0.768178,-0.5555702,-0.5132794,-0.7681782,-0.3826833,-0.318189,-0.768178,-0.5555702,-0.4619393,-0.6913421,-0.5555702,-0.5132794,-0.7681782,-0.3826833,-0.5132794,-0.7681782,-0.3826833,-0.4619393,-0.6913421,-0.5555702,-0.6532811,-0.6532819,-0.3826833,-0.4619393,-0.6913421,-0.5555702,-0.5879375,-0.5879382,-0.5555702,-0.6532811,-0.6532819,-0.3826833,-0.6532811,-0.6532819,-0.3826833,-0.5879375,-0.5879382,-0.5555702,-0.7681775,-0.5132805,-0.3826833,-0.5879375,-0.5879382,-0.5555702,-0.6913415,-0.4619402,-0.5555702,-0.7681775,-0.5132805,-0.3826833,-0.7681775,-0.5132805,-0.3826833,-0.6913415,-0.4619402,-0.5555702,-0.8535532,-0.3535538,-0.3826833,-0.6913415,-0.4619402,-0.5555702,-0.7681776,-0.31819,-0.5555702,-0.8535532,-0.3535538,-0.3826833,-0.8535532,-0.3535538,-0.3826833,-0.7681776,-0.31819,-0.5555702,-0.9061274,-0.1802403,-0.3826833,-0.7681776,-0.31819,-0.5555702,-0.8154931,-0.162212,-0.5555702,-0.9061274,-0.1802403,-0.3826833,-0.9061274,-0.1802403,-0.3826833,-0.8154931,-0.162212,-0.5555702,-0.9238796,-0.0000003,-0.3826833,-0.8154931,-0.162212,-0.5555702,-0.8314697,-0.0000003,-0.5555702,-0.9238796,-0.0000003,-0.3826833,-0.9238796,-0.0000003,-0.3826833,-0.8314697,-0.0000003,-0.5555702,-0.9061275,0.1802397,-0.3826833,-0.8314697,-0.0000003,-0.5555702,-0.8154932,0.1622115,-0.5555702,-0.9061275,0.1802397,-0.3826833,-0.9061275,0.1802397,-0.3826833,-0.8154932,0.1622115,-0.5555702,-0.8535535,0.3535533,-0.3826833,-0.8154932,0.1622115,-0.5555702,-0.7681779,0.3181895,-0.5555702,-0.8535535,0.3535533,-0.3826833,-0.8535535,0.3535533,-0.3826833,-0.7681779,0.3181895,-0.5555702,-0.7681778,0.5132799,-0.3826833,-0.7681779,0.3181895,-0.5555702,-0.6913418,0.4619398,-0.5555702,-0.7681778,0.5132799,-0.3826833,-0.7681778,0.5132799,-0.3826833,-0.6913418,0.4619398,-0.5555702,-0.6532815,0.6532815,-0.3826833,-0.6913418,0.4619398,-0.5555702,-0.5879378,0.5879378,-0.5555702,-0.6532815,0.6532815,-0.3826833,-0.6532815,0.6532815,-0.3826833,-0.5879378,0.5879378,-0.5555702,-0.5132799,0.7681778,-0.3826833,-0.5879378,0.5879378,-0.5555702,-0.4619398,0.6913418,-0.5555702,-0.5132799,0.7681778,-0.3826833,-0.5132799,0.7681778,-0.3826833,-0.4619398,0.6913418,-0.5555702,-0.3535534,0.8535535,-0.3826833,-0.4619398,0.6913418,-0.5555702,-0.3181896,0.7681778,-0.5555702,-0.3535534,0.8535535,-0.3826833,-0.3535534,0.8535535,-0.3826833,-0.3181896,0.7681778,-0.5555702,-0.18024,0.9061275,-0.3826833,-0.3181896,0.7681778,-0.5555702,-0.1622117,0.8154932,-0.5555702,-0.18024,0.9061275,-0.3826833,-0.18024,0.9061275,-0.3826833,-0.1622117,0.8154932,-0.5555702,-0.0,0.9238796,-0.3826833,-0.1622117,0.8154932,-0.5555702,-0.0,0.8314697,-0.5555702,-0.0,0.9238796,-0.3826833,-0.0,0.9807853,-0.1950902,-0.0,0.9238796,-0.3826833,0.1913396,0.9619402,-0.1950902,-0.0,0.9238796,-0.3826833,0.180238,0.9061279,-0.3826833,0.1913396,0.9619402,-0.1950902,0.1913396,0.9619402,-0.1950902,0.180238,0.9061279,-0.3826833,0.3753284,0.9061283,-0.1950902,0.180238,0.9061279,-0.3826833,0.3535516,0.8535542,-0.3826833,0.3753284,0.9061283,-0.1950902,0.3753284,0.9061283,-0.1950902,0.3535516,0.8535542,-0.3826833,0.5448935,0.8154943,-0.1950902,0.3535516,0.8535542,-0.3826833,0.5132784,0.7681788,-0.3826833,0.5448935,0.8154943,-0.1950902,0.5448935,0.8154943,-0.1950902,0.5132784,0.7681788,-0.3826833,0.6935186,0.6935213,-0.1950902,0.5132784,0.7681788,-0.3826833,0.6532802,0.6532828,-0.3826833,0.6935186,0.6935213,-0.1950902,0.6935186,0.6935213,-0.1950902,0.6532802,0.6532828,-0.3826833,0.8154922,0.5448967,-0.1950902,0.6532802,0.6532828,-0.3826833,0.7681769,0.5132814,-0.3826833,0.8154922,0.5448967,-0.1950902,0.8154922,0.5448967,-0.1950902,0.7681769,0.5132814,-0.3826833,0.9061268,0.3753319,-0.1950902,0.7681769,0.5132814,-0.3826833,0.8535528,0.3535549,-0.3826833,0.9061268,0.3753319,-0.1950902,0.9061268,0.3753319,-0.1950902,0.8535528,0.3535549,-0.3826833,0.9619395,0.1913434,-0.1950902,0.8535528,0.3535549,-0.3826833,0.9061272,0.1802415,-0.3826833,0.9619395,0.1913434,-0.1950902,0.9619395,0.1913434,-0.1950902,0.9061272,0.1802415,-0.3826833,0.9807853,0.0000016,-0.1950902,0.9061272,0.1802415,-0.3826833,0.9238796,0.0000015,-0.3826833,0.9807853,0.0000016,-0.1950902,0.9807853,0.0000016,-0.1950902,0.9238796,0.0000015,-0.3826833,0.9619401,-0.1913403,-0.1950902,0.9238796,0.0000015,-0.3826833,0.9061278,-0.1802386,-0.3826833,0.9619401,-0.1913403,-0.1950902,0.9619401,-0.1913403,-0.1950902,0.9061278,-0.1802386,-0.3826833,0.906128,-0.375329,-0.1950902,0.9061278,-0.1802386,-0.3826833,0.853554,-0.3535522,-0.3826833,0.906128,-0.375329,-0.1950902,0.906128,-0.375329,-0.1950902,0.853554,-0.3535522,-0.3826833,0.8154939,-0.544894,-0.1950902,0.853554,-0.3535522,-0.3826833,0.7681785,-0.513279,-0.3826833,0.8154939,-0.544894,-0.1950902,0.8154939,-0.544894,-0.1950902,0.7681785,-0.513279,-0.3826833,0.6935208,-0.6935191,-0.1950902,0.7681785,-0.513279,-0.3826833,0.6532823,-0.6532807,-0.3826833,0.6935208,-0.6935191,-0.1950902,0.6935208,-0.6935191,-0.1950902,0.6532823,-0.6532807,-0.3826833,0.5448961,-0.8154925,-0.1950902,0.6532823,-0.6532807,-0.3826833,0.5132809,-0.7681772,-0.3826833,0.5448961,-0.8154925,-0.1950902,0.5448961,-0.8154925,-0.1950902,0.5132809,-0.7681772,-0.3826833,0.3753313,-0.906127,-0.1950902,0.5132809,-0.7681772,-0.3826833,0.3535544,-0.853553,-0.3826833,0.3753313,-0.906127,-0.1950902,0.3753313,-0.906127,-0.1950902,0.3535544,-0.853553,-0.3826833,0.1913427,-0.9619396,-0.1950902,0.3535544,-0.853553,-0.3826833,0.1802409,-0.9061273,-0.3826833,0.1913427,-0.9619396,-0.1950902,0.1913427,-0.9619396,-0.1950902,0.1802409,-0.9061273,-0.3826833,0.0000009,-0.9807853,-0.1950902,0.1802409,-0.9061273,-0.3826833,0.0000009,-0.9238796,-0.3826833,0.0000009,-0.9807853,-0.1950902,0.0000009,-0.9807853,-0.1950902,0.0000009,-0.9238796,-0.3826833,-0.1913409,-0.9619399,-0.1950902,0.0000009,-0.9238796,-0.3826833,-0.1802392,-0.9061276,-0.3826833,-0.1913409,-0.9619399,-0.1950902,-0.1913409,-0.9619399,-0.1950902,-0.1802392,-0.9061276,-0.3826833,-0.3753296,-0.9061278,-0.1950902,-0.1802392,-0.9061276,-0.3826833,-0.3535527,-0.8535537,-0.3826833,-0.3753296,-0.9061278,-0.1950902,-0.3753296,-0.9061278,-0.1950902,-0.3535527,-0.8535537,-0.3826833,-0.5448946,-0.8154936,-0.1950902,-0.3535527,-0.8535537,-0.3826833,-0.5132794,-0.7681782,-0.3826833,-0.5448946,-0.8154936,-0.1950902,-0.5448946,-0.8154936,-0.1950902,-0.5132794,-0.7681782,-0.3826833,-0.6935195,-0.6935204,-0.1950902,-0.5132794,-0.7681782,-0.3826833,-0.6532811,-0.6532819,-0.3826833,-0.6935195,-0.6935204,-0.1950902,-0.6935195,-0.6935204,-0.1950902,-0.6532811,-0.6532819,-0.3826833,-0.8154929,-0.5448956,-0.1950902,-0.6532811,-0.6532819,-0.3826833,-0.7681775,-0.5132805,-0.3826833,-0.8154929,-0.5448956,-0.1950902,-0.8154929,-0.5448956,-0.1950902,-0.7681775,-0.5132805,-0.3826833,-0.9061273,-0.3753307,-0.1950902,-0.7681775,-0.5132805,-0.3826833,-0.8535532,-0.3535538,-0.3826833,-0.9061273,-0.3753307,-0.1950902,-0.9061273,-0.3753307,-0.1950902,-0.8535532,-0.3535538,-0.3826833,-0.9619397,-0.1913421,-0.1950902,-0.8535532,-0.3535538,-0.3826833,-0.9061274,-0.1802403,-0.3826833,-0.9619397,-0.1913421,-0.1950902,-0.9619397,-0.1913421,-0.1950902,-0.9061274,-0.1802403,-0.3826833,-0.9807853,-0.0000003,-0.1950902,-0.9061274,-0.1802403,-0.3826833,-0.9238796,-0.0000003,-0.3826833,-0.9807853,-0.0000003,-0.1950902,-0.9807853,-0.0000003,-0.1950902,-0.9238796,-0.0000003,-0.3826833,-0.9619398,0.1913415,-0.1950902,-0.9238796,-0.0000003,-0.3826833,-0.9061275,0.1802397,-0.3826833,-0.9619398,0.1913415,-0.1950902,-0.9619398,0.1913415,-0.1950902,-0.9061275,0.1802397,-0.3826833,-0.9061276,0.3753302,-0.1950902,-0.9061275,0.1802397,-0.3826833,-0.8535535,0.3535533,-0.3826833,-0.9061276,0.3753302,-0.1950902,-0.9061276,0.3753302,-0.1950902,-0.8535535,0.3535533,-0.3826833,-0.8154932,0.5448951,-0.1950902,-0.8535535,0.3535533,-0.3826833,-0.7681778,0.5132799,-0.3826833,-0.8154932,0.5448951,-0.1950902,-0.8154932,0.5448951,-0.1950902,-0.7681778,0.5132799,-0.3826833,-0.6935199,0.6935199,-0.1950902,-0.7681778,0.5132799,-0.3826833,-0.6532815,0.6532815,-0.3826833,-0.6935199,0.6935199,-0.1950902,-0.6935199,0.6935199,-0.1950902,-0.6532815,0.6532815,-0.3826833,-0.5448951,0.8154932,-0.1950902,-0.6532815,0.6532815,-0.3826833,-0.5132799,0.7681778,-0.3826833,-0.5448951,0.8154932,-0.1950902,-0.5448951,0.8154932,-0.1950902,-0.5132799,0.7681778,-0.3826833,-0.3753302,0.9061275,-0.1950902,-0.5132799,0.7681778,-0.3826833,-0.3535534,0.8535535,-0.3826833,-0.3753302,0.9061275,-0.1950902,-0.3753302,0.9061275,-0.1950902,-0.3535534,0.8535535,-0.3826833,-0.1913417,0.9619398,-0.1950902,-0.3535534,0.8535535,-0.3826833,-0.18024,0.9061275,-0.3826833,-0.1913417,0.9619398,-0.1950902,-0.1913417,0.9619398,-0.1950902,-0.18024,0.9061275,-0.3826833,-0.0,0.9807853,-0.1950902,-0.18024,0.9061275,-0.3826833,-0.0,0.9238796,-0.3826833,-0.0,0.9807853,-0.1950902,-0.0,1.0,0.0000001,-0.0,0.9807853,-0.1950902,0.1950882,0.9807857,0.0000001,-0.0,0.9807853,-0.1950902,0.1913396,0.9619402,-0.1950902,0.1950882,0.9807857,0.0000001,0.1950882,0.9807857,0.0000001,0.1913396,0.9619402,-0.1950902,0.3826815,0.9238803,0.0000001,0.1913396,0.9619402,-0.1950902,0.3753284,0.9061283,-0.1950902,0.3826815,0.9238803,0.0000001,0.3826815,0.9238803,0.0000001,0.3753284,0.9061283,-0.1950902,0.5555686,0.8314707,0.0000001,0.3753284,0.9061283,-0.1950902,0.5448935,0.8154943,-0.1950902,0.5555686,0.8314707,0.0000001,0.5555686,0.8314707,0.0000001,0.5448935,0.8154943,-0.1950902,0.7071054,0.7071081,0.0000001,0.5448935,0.8154943,-0.1950902,0.6935186,0.6935213,-0.1950902,0.7071054,0.7071081,0.0000001,0.7071054,0.7071081,0.0000001,0.6935186,0.6935213,-0.1950902,0.8314686,0.5555718,0.0000001,0.6935186,0.6935213,-0.1950902,0.8154922,0.5448967,-0.1950902,0.8314686,0.5555718,0.0000001,0.8314686,0.5555718,0.0000001,0.8154922,0.5448967,-0.1950902,0.9238788,0.3826851,0.0000001,0.8154922,0.5448967,-0.1950902,0.9061268,0.3753319,-0.1950902,0.9238788,0.3826851,0.0000001,0.9238788,0.3826851,0.0000001,0.9061268,0.3753319,-0.1950902,0.980785,0.195092,0.0000001,0.9061268,0.3753319,-0.1950902,0.9619395,0.1913434,-0.1950902,0.980785,0.195092,0.0000001,0.980785,0.195092,0.0000001,0.9619395,0.1913434,-0.1950902,1.0,0.0000016,0.0000001,0.9619395,0.1913434,-0.1950902,0.9807853,0.0000016,-0.1950902,1.0,0.0000016,0.0000001,1.0,0.0000016,0.0000001,0.9807853,0.0000016,-0.1950902,0.9807855,-0.1950888,0.0000001,0.9807853,0.0000016,-0.1950902,0.9619401,-0.1913403,-0.1950902,0.9807855,-0.1950888,0.0000001,0.9807855,-0.1950888,0.0000001,0.9619401,-0.1913403,-0.1950902,0.9238801,-0.3826821,0.0000001,0.9619401,-0.1913403,-0.1950902,0.906128,-0.375329,-0.1950902,0.9238801,-0.3826821,0.0000001,0.9238801,-0.3826821,0.0000001,0.906128,-0.375329,-0.1950902,0.8314704,-0.5555691,0.0000001,0.906128,-0.375329,-0.1950902,0.8154939,-0.544894,-0.1950902,0.8314704,-0.5555691,0.0000001,0.8314704,-0.5555691,0.0000001,0.8154939,-0.544894,-0.1950902,0.7071077,-0.7071059,0.0000001,0.8154939,-0.544894,-0.1950902,0.6935208,-0.6935191,-0.1950902,0.7071077,-0.7071059,0.0000001,0.7071077,-0.7071059,0.0000001,0.6935208,-0.6935191,-0.1950902,0.5555713,-0.8314689,0.0000001,0.6935208,-0.6935191,-0.1950902,0.5448961,-0.8154925,-0.1950902,0.5555713,-0.8314689,0.0000001,0.5555713,-0.8314689,0.0000001,0.5448961,-0.8154925,-0.1950902,0.3826845,-0.9238791,0.0000001,0.5448961,-0.8154925,-0.1950902,0.3753313,-0.906127,-0.1950902,0.3826845,-0.9238791,0.0000001,0.3826845,-0.9238791,0.0000001,0.3753313,-0.906127,-0.1950902,0.1950914,-0.9807851,0.0000001,0.3753313,-0.906127,-0.1950902,0.1913427,-0.9619396,-0.1950902,0.1950914,-0.9807851,0.0000001,0.1950914,-0.9807851,0.0000001,0.1913427,-0.9619396,-0.1950902,0.000001,-1.0,0.0000001,0.1913427,-0.9619396,-0.1950902,0.0000009,-0.9807853,-0.1950902,0.000001,-1.0,0.0000001,0.000001,-1.0,0.0000001,0.0000009,-0.9807853,-0.1950902,-0.1950895,-0.9807854,0.0000001,0.0000009,-0.9807853,-0.1950902,-0.1913409,-0.9619399,-0.1950902,-0.1950895,-0.9807854,0.0000001,-0.1950895,-0.9807854,0.0000001,-0.1913409,-0.9619399,-0.1950902,-0.3826827,-0.9238799,0.0000001,-0.1913409,-0.9619399,-0.1950902,-0.3753296,-0.9061278,-0.1950902,-0.3826827,-0.9238799,0.0000001,-0.3826827,-0.9238799,0.0000001,-0.3753296,-0.9061278,-0.1950902,-0.5555696,-0.83147,0.0000001,-0.3753296,-0.9061278,-0.1950902,-0.5448946,-0.8154936,-0.1950902,-0.5555696,-0.83147,0.0000001,-0.5555696,-0.83147,0.0000001,-0.5448946,-0.8154936,-0.1950902,-0.7071064,-0.7071072,0.0000001,-0.5448946,-0.8154936,-0.1950902,-0.6935195,-0.6935204,-0.1950902,-0.7071064,-0.7071072,0.0000001,-0.7071064,-0.7071072,0.0000001,-0.6935195,-0.6935204,-0.1950902,-0.8314693,-0.5555707,0.0000001,-0.6935195,-0.6935204,-0.1950902,-0.8154929,-0.5448956,-0.1950902,-0.8314693,-0.5555707,0.0000001,-0.8314693,-0.5555707,0.0000001,-0.8154929,-0.5448956,-0.1950902,-0.9238793,-0.3826839,0.0000001,-0.8154929,-0.5448956,-0.1950902,-0.9061273,-0.3753307,-0.1950902,-0.9238793,-0.3826839,0.0000001,-0.9238793,-0.3826839,0.0000001,-0.9061273,-0.3753307,-0.1950902,-0.9807852,-0.1950907,0.0000001,-0.9061273,-0.3753307,-0.1950902,-0.9619397,-0.1913421,-0.1950902,-0.9807852,-0.1950907,0.0000001,-0.9807852,-0.1950907,0.0000001,-0.9619397,-0.1913421,-0.1950902,-1.0,-0.0000003,0.0000001,-0.9619397,-0.1913421,-0.1950902,-0.9807853,-0.0000003,-0.1950902,-1.0,-0.0000003,0.0000001,-1.0,-0.0000003,0.0000001,-0.9807853,-0.0000003,-0.1950902,-0.9807853,0.1950901,0.0000001,-0.9807853,-0.0000003,-0.1950902,-0.9619398,0.1913415,-0.1950902,-0.9807853,0.1950901,0.0000001,-0.9807853,0.1950901,0.0000001,-0.9619398,0.1913415,-0.1950902,-0.9238796,0.3826833,0.0000001,-0.9619398,0.1913415,-0.1950902,-0.9061276,0.3753302,-0.1950902,-0.9238796,0.3826833,0.0000001,-0.9238796,0.3826833,0.0000001,-0.9061276,0.3753302,-0.1950902,-0.8314697,0.5555702,0.0000001,-0.9061276,0.3753302,-0.1950902,-0.8154932,0.5448951,-0.1950902,-0.8314697,0.5555702,0.0000001,-0.8314697,0.5555702,0.0000001,-0.8154932,0.5448951,-0.1950902,-0.7071068,0.7071068,0.0000001,-0.8154932,0.5448951,-0.1950902,-0.6935199,0.6935199,-0.1950902,-0.7071068,0.7071068,0.0000001,-0.7071068,0.7071068,0.0000001,-0.6935199,0.6935199,-0.1950902,-0.5555702,0.8314697,0.0000001,-0.6935199,0.6935199,-0.1950902,-0.5448951,0.8154932,-0.1950902,-0.5555702,0.8314697,0.0000001,-0.5555702,0.8314697,0.0000001,-0.5448951,0.8154932,-0.1950902,-0.3826834,0.9238796,0.0000001,-0.5448951,0.8154932,-0.1950902,-0.3753302,0.9061275,-0.1950902,-0.3826834,0.9238796,0.0000001,-0.3826834,0.9238796,0.0000001,-0.3753302,0.9061275,-0.1950902,-0.1950903,0.9807853,0.0000001,-0.3753302,0.9061275,-0.1950902,-0.1913417,0.9619398,-0.1950902,-0.1950903,0.9807853,0.0000001,-0.1950903,0.9807853,0.0000001,-0.1913417,0.9619398,-0.1950902,-0.0,1.0,0.0000001,-0.1913417,0.9619398,-0.1950902,-0.0,0.9807853,-0.1950902,-0.0,1.0,0.0000001,-0.0,0.9807853,0.1950904,-0.0,1.0,0.0000001,0.1913396,0.9619402,0.1950904,-0.0,1.0,0.0000001,0.1950882,0.9807857,0.0000001,0.1913396,0.9619402,0.1950904,0.1913396,0.9619402,0.1950904,0.1950882,0.9807857,0.0000001,0.3753284,0.9061282,0.1950904,0.1950882,0.9807857,0.0000001,0.3826815,0.9238803,0.0000001,0.3753284,0.9061282,0.1950904,0.3753284,0.9061282,0.1950904,0.3826815,0.9238803,0.0000001,0.5448934,0.8154942,0.1950904,0.3826815,0.9238803,0.0000001,0.5555686,0.8314707,0.0000001,0.5448934,0.8154942,0.1950904,0.5448934,0.8154942,0.1950904,0.5555686,0.8314707,0.0000001,0.6935185,0.6935213,0.1950904,0.5555686,0.8314707,0.0000001,0.7071054,0.7071081,0.0000001,0.6935185,0.6935213,0.1950904,0.6935185,0.6935213,0.1950904,0.7071054,0.7071081,0.0000001,0.8154921,0.5448966,0.1950904,0.7071054,0.7071081,0.0000001,0.8314686,0.5555718,0.0000001,0.8154921,0.5448966,0.1950904,0.8154921,0.5448966,0.1950904,0.8314686,0.5555718,0.0000001,0.9061267,0.3753319,0.1950904,0.8314686,0.5555718,0.0000001,0.9238788,0.3826851,0.0000001,0.9061267,0.3753319,0.1950904,0.9061267,0.3753319,0.1950904,0.9238788,0.3826851,0.0000001,0.9619394,0.1913433,0.1950904,0.9238788,0.3826851,0.0000001,0.980785,0.195092,0.0000001,0.9619394,0.1913433,0.1950904,0.9619394,0.1913433,0.1950904,0.980785,0.195092,0.0000001,0.9807853,0.0000016,0.1950904,0.980785,0.195092,0.0000001,1.0,0.0000016,0.0000001,0.9807853,0.0000016,0.1950904,0.9807853,0.0000016,0.1950904,1.0,0.0000016,0.0000001,0.96194,-0.1913402,0.1950904,1.0,0.0000016,0.0000001,0.9807855,-0.1950888,0.0000001,0.96194,-0.1913402,0.1950904,0.96194,-0.1913402,0.1950904,0.9807855,-0.1950888,0.0000001,0.906128,-0.375329,0.1950904,0.9807855,-0.1950888,0.0000001,0.9238801,-0.3826821,0.0000001,0.906128,-0.375329,0.1950904,0.906128,-0.375329,0.1950904,0.9238801,-0.3826821,0.0000001,0.8154939,-0.544894,0.1950904,0.9238801,-0.3826821,0.0000001,0.8314704,-0.5555691,0.0000001,0.8154939,-0.544894,0.1950904,0.8154939,-0.544894,0.1950904,0.8314704,-0.5555691,0.0000001,0.6935208,-0.693519,0.1950904,0.8314704,-0.5555691,0.0000001,0.7071077,-0.7071059,0.0000001,0.6935208,-0.693519,0.1950904,0.6935208,-0.693519,0.1950904,0.7071077,-0.7071059,0.0000001,0.5448961,-0.8154925,0.1950904,0.7071077,-0.7071059,0.0000001,0.5555713,-0.8314689,0.0000001,0.5448961,-0.8154925,0.1950904,0.5448961,-0.8154925,0.1950904,0.5555713,-0.8314689,0.0000001,0.3753313,-0.906127,0.1950904,0.5555713,-0.8314689,0.0000001,0.3826845,-0.9238791,0.0000001,0.3753313,-0.906127,0.1950904,0.3753313,-0.906127,0.1950904,0.3826845,-0.9238791,0.0000001,0.1913427,-0.9619395,0.1950904,0.3826845,-0.9238791,0.0000001,0.1950914,-0.9807851,0.0000001,0.1913427,-0.9619395,0.1950904,0.1913427,-0.9619395,0.1950904,0.1950914,-0.9807851,0.0000001,0.0000009,-0.9807853,0.1950904,0.1950914,-0.9807851,0.0000001,0.000001,-1.0,0.0000001,0.0000009,-0.9807853,0.1950904,0.0000009,-0.9807853,0.1950904,0.000001,-1.0,0.0000001,-0.1913409,-0.9619399,0.1950904,0.000001,-1.0,0.0000001,-0.1950895,-0.9807854,0.0000001,-0.1913409,-0.9619399,0.1950904,-0.1913409,-0.9619399,0.1950904,-0.1950895,-0.9807854,0.0000001,-0.3753295,-0.9061278,0.1950904,-0.1950895,-0.9807854,0.0000001,-0.3826827,-0.9238799,0.0000001,-0.3753295,-0.9061278,0.1950904,-0.3753295,-0.9061278,0.1950904,-0.3826827,-0.9238799,0.0000001,-0.5448945,-0.8154935,0.1950904,-0.3826827,-0.9238799,0.0000001,-0.5555696,-0.83147,0.0000001,-0.5448945,-0.8154935,0.1950904,-0.5448945,-0.8154935,0.1950904,-0.5555696,-0.83147,0.0000001,-0.6935195,-0.6935204,0.1950904,-0.5555696,-0.83147,0.0000001,-0.7071064,-0.7071072,0.0000001,-0.6935195,-0.6935204,0.1950904,-0.6935195,-0.6935204,0.1950904,-0.7071064,-0.7071072,0.0000001,-0.8154928,-0.5448956,0.1950904,-0.7071064,-0.7071072,0.0000001,-0.8314693,-0.5555707,0.0000001,-0.8154928,-0.5448956,0.1950904,-0.8154928,-0.5448956,0.1950904,-0.8314693,-0.5555707,0.0000001,-0.9061272,-0.3753307,0.1950904,-0.8314693,-0.5555707,0.0000001,-0.9238793,-0.3826839,0.0000001,-0.9061272,-0.3753307,0.1950904,-0.9061272,-0.3753307,0.1950904,-0.9238793,-0.3826839,0.0000001,-0.9619396,-0.1913421,0.1950904,-0.9238793,-0.3826839,0.0000001,-0.9807852,-0.1950907,0.0000001,-0.9619396,-0.1913421,0.1950904,-0.9619396,-0.1913421,0.1950904,-0.9807852,-0.1950907,0.0000001,-0.9807853,-0.0000003,0.1950904,-0.9807852,-0.1950907,0.0000001,-1.0,-0.0000003,0.0000001,-0.9807853,-0.0000003,0.1950904,-0.9807853,-0.0000003,0.1950904,-1.0,-0.0000003,0.0000001,-0.9619398,0.1913415,0.1950904,-1.0,-0.0000003,0.0000001,-0.9807853,0.1950901,0.0000001,-0.9619398,0.1913415,0.1950904,-0.9619398,0.1913415,0.1950904,-0.9807853,0.1950901,0.0000001,-0.9061275,0.3753301,0.1950904,-0.9807853,0.1950901,0.0000001,-0.9238796,0.3826833,0.0000001,-0.9061275,0.3753301,0.1950904,-0.9061275,0.3753301,0.1950904,-0.9238796,0.3826833,0.0000001,-0.8154932,0.5448951,0.1950904,-0.9238796,0.3826833,0.0000001,-0.8314697,0.5555702,0.0000001,-0.8154932,0.5448951,0.1950904,-0.8154932,0.5448951,0.1950904,-0.8314697,0.5555702,0.0000001,-0.6935199,0.6935199,0.1950904,-0.8314697,0.5555702,0.0000001,-0.7071068,0.7071068,0.0000001,-0.6935199,0.6935199,0.1950904,-0.6935199,0.6935199,0.1950904,-0.7071068,0.7071068,0.0000001,-0.5448951,0.8154932,0.1950904,-0.7071068,0.7071068,0.0000001,-0.5555702,0.8314697,0.0000001,-0.5448951,0.8154932,0.1950904,-0.5448951,0.8154932,0.1950904,-0.5555702,0.8314697,0.0000001,-0.3753302,0.9061275,0.1950904,-0.5555702,0.8314697,0.0000001,-0.3826834,0.9238796,0.0000001,-0.3753302,0.9061275,0.1950904,-0.3753302,0.9061275,0.1950904,-0.3826834,0.9238796,0.0000001,-0.1913417,0.9619397,0.1950904,-0.3826834,0.9238796,0.0000001,-0.1950903,0.9807853,0.0000001,-0.1913417,0.9619397,0.1950904,-0.1913417,0.9619397,0.1950904,-0.1950903,0.9807853,0.0000001,-0.0,0.9807853,0.1950904,-0.1950903,0.9807853,0.0000001,-0.0,1.0,0.0000001,-0.0,0.9807853,0.1950904,-0.0,0.9238795,0.3826834,-0.0,0.9807853,0.1950904,0.180238,0.9061278,0.3826834,-0.0,0.9807853,0.1950904,0.1913396,0.9619402,0.1950904,0.180238,0.9061278,0.3826834,0.180238,0.9061278,0.3826834,0.1913396,0.9619402,0.1950904,0.3535516,0.8535541,0.3826834,0.1913396,0.9619402,0.1950904,0.3753284,0.9061282,0.1950904,0.3535516,0.8535541,0.3826834,0.3535516,0.8535541,0.3826834,0.3753284,0.9061282,0.1950904,0.5132784,0.7681788,0.3826834,0.3753284,0.9061282,0.1950904,0.5448934,0.8154942,0.1950904,0.5132784,0.7681788,0.3826834,0.5132784,0.7681788,0.3826834,0.5448934,0.8154942,0.1950904,0.6532802,0.6532827,0.3826834,0.5448934,0.8154942,0.1950904,0.6935185,0.6935213,0.1950904,0.6532802,0.6532827,0.3826834,0.6532802,0.6532827,0.3826834,0.6935185,0.6935213,0.1950904,0.7681768,0.5132814,0.3826834,0.6935185,0.6935213,0.1950904,0.8154921,0.5448966,0.1950904,0.7681768,0.5132814,0.3826834,0.7681768,0.5132814,0.3826834,0.8154921,0.5448966,0.1950904,0.8535528,0.3535549,0.3826834,0.8154921,0.5448966,0.1950904,0.9061267,0.3753319,0.1950904,0.8535528,0.3535549,0.3826834,0.8535528,0.3535549,0.3826834,0.9061267,0.3753319,0.1950904,0.9061271,0.1802415,0.3826834,0.9061267,0.3753319,0.1950904,0.9619394,0.1913433,0.1950904,0.9061271,0.1802415,0.3826834,0.9061271,0.1802415,0.3826834,0.9619394,0.1913433,0.1950904,0.9238795,0.0000015,0.3826834,0.9619394,0.1913433,0.1950904,0.9807853,0.0000016,0.1950904,0.9238795,0.0000015,0.3826834,0.9238795,0.0000015,0.3826834,0.9807853,0.0000016,0.1950904,0.9061277,-0.1802386,0.3826834,0.9807853,0.0000016,0.1950904,0.96194,-0.1913402,0.1950904,0.9061277,-0.1802386,0.3826834,0.9061277,-0.1802386,0.3826834,0.96194,-0.1913402,0.1950904,0.8535539,-0.3535521,0.3826834,0.96194,-0.1913402,0.1950904,0.906128,-0.375329,0.1950904,0.8535539,-0.3535521,0.3826834,0.8535539,-0.3535521,0.3826834,0.906128,-0.375329,0.1950904,0.7681785,-0.5132789,0.3826834,0.906128,-0.375329,0.1950904,0.8154939,-0.544894,0.1950904,0.7681785,-0.5132789,0.3826834,0.7681785,-0.5132789,0.3826834,0.8154939,-0.544894,0.1950904,0.6532823,-0.6532806,0.3826834,0.8154939,-0.544894,0.1950904,0.6935208,-0.693519,0.1950904,0.6532823,-0.6532806,0.3826834,0.6532823,-0.6532806,0.3826834,0.6935208,-0.693519,0.1950904,0.5132809,-0.7681771,0.3826834,0.6935208,-0.693519,0.1950904,0.5448961,-0.8154925,0.1950904,0.5132809,-0.7681771,0.3826834,0.5132809,-0.7681771,0.3826834,0.5448961,-0.8154925,0.1950904,0.3535543,-0.8535529,0.3826834,0.5448961,-0.8154925,0.1950904,0.3753313,-0.906127,0.1950904,0.3535543,-0.8535529,0.3826834,0.3535543,-0.8535529,0.3826834,0.3753313,-0.906127,0.1950904,0.1802409,-0.9061272,0.3826834,0.3753313,-0.906127,0.1950904,0.1913427,-0.9619395,0.1950904,0.1802409,-0.9061272,0.3826834,0.1802409,-0.9061272,0.3826834,0.1913427,-0.9619395,0.1950904,0.0000009,-0.9238795,0.3826834,0.1913427,-0.9619395,0.1950904,0.0000009,-0.9807853,0.1950904,0.0000009,-0.9238795,0.3826834,0.0000009,-0.9238795,0.3826834,0.0000009,-0.9807853,0.1950904,-0.1802392,-0.9061276,0.3826834,0.0000009,-0.9807853,0.1950904,-0.1913409,-0.9619399,0.1950904,-0.1802392,-0.9061276,0.3826834,-0.1802392,-0.9061276,0.3826834,-0.1913409,-0.9619399,0.1950904,-0.3535527,-0.8535537,0.3826834,-0.1913409,-0.9619399,0.1950904,-0.3753295,-0.9061278,0.1950904,-0.3535527,-0.8535537,0.3826834,-0.3535527,-0.8535537,0.3826834,-0.3753295,-0.9061278,0.1950904,-0.5132794,-0.7681781,0.3826834,-0.3753295,-0.9061278,0.1950904,-0.5448945,-0.8154935,0.1950904,-0.5132794,-0.7681781,0.3826834,-0.5132794,-0.7681781,0.3826834,-0.5448945,-0.8154935,0.1950904,-0.6532811,-0.6532819,0.3826834,-0.5448945,-0.8154935,0.1950904,-0.6935195,-0.6935204,0.1950904,-0.6532811,-0.6532819,0.3826834,-0.6532811,-0.6532819,0.3826834,-0.6935195,-0.6935204,0.1950904,-0.7681774,-0.5132804,0.3826834,-0.6935195,-0.6935204,0.1950904,-0.8154928,-0.5448956,0.1950904,-0.7681774,-0.5132804,0.3826834,-0.7681774,-0.5132804,0.3826834,-0.8154928,-0.5448956,0.1950904,-0.8535532,-0.3535538,0.3826834,-0.8154928,-0.5448956,0.1950904,-0.9061272,-0.3753307,0.1950904,-0.8535532,-0.3535538,0.3826834,-0.8535532,-0.3535538,0.3826834,-0.9061272,-0.3753307,0.1950904,-0.9061273,-0.1802403,0.3826834,-0.9061272,-0.3753307,0.1950904,-0.9619396,-0.1913421,0.1950904,-0.9061273,-0.1802403,0.3826834,-0.9061273,-0.1802403,0.3826834,-0.9619396,-0.1913421,0.1950904,-0.9238795,-0.0000003,0.3826834,-0.9619396,-0.1913421,0.1950904,-0.9807853,-0.0000003,0.1950904,-0.9238795,-0.0000003,0.3826834,-0.9238795,-0.0000003,0.3826834,-0.9807853,-0.0000003,0.1950904,-0.9061275,0.1802397,0.3826834,-0.9807853,-0.0000003,0.1950904,-0.9619398,0.1913415,0.1950904,-0.9061275,0.1802397,0.3826834,-0.9061275,0.1802397,0.3826834,-0.9619398,0.1913415,0.1950904,-0.8535535,0.3535532,0.3826834,-0.9619398,0.1913415,0.1950904,-0.9061275,0.3753301,0.1950904,-0.8535535,0.3535532,0.3826834,-0.8535535,0.3535532,0.3826834,-0.9061275,0.3753301,0.1950904,-0.7681777,0.5132799,0.3826834,-0.9061275,0.3753301,0.1950904,-0.8154932,0.5448951,0.1950904,-0.7681777,0.5132799,0.3826834,-0.7681777,0.5132799,0.3826834,-0.8154932,0.5448951,0.1950904,-0.6532815,0.6532815,0.3826834,-0.8154932,0.5448951,0.1950904,-0.6935199,0.6935199,0.1950904,-0.6532815,0.6532815,0.3826834,-0.6532815,0.6532815,0.3826834,-0.6935199,0.6935199,0.1950904,-0.5132799,0.7681777,0.3826834,-0.6935199,0.6935199,0.1950904,-0.5448951,0.8154932,0.1950904,-0.5132799,0.7681777,0.3826834,-0.5132799,0.7681777,0.3826834,-0.5448951,0.8154932,0.1950904,-0.3535534,0.8535534,0.3826834,-0.5448951,0.8154932,0.1950904,-0.3753302,0.9061275,0.1950904,-0.3535534,0.8535534,0.3826834,-0.3535534,0.8535534,0.3826834,-0.3753302,0.9061275,0.1950904,-0.1802399,0.9061274,0.3826834,-0.3753302,0.9061275,0.1950904,-0.1913417,0.9619397,0.1950904,-0.1802399,0.9061274,0.3826834,-0.1802399,0.9061274,0.3826834,-0.1913417,0.9619397,0.1950904,-0.0,0.9238795,0.3826834,-0.1913417,0.9619397,0.1950904,-0.0,0.9807853,0.1950904,-0.0,0.9238795,0.3826834,-0.0,0.8314697,0.5555702,-0.0,0.9238795,0.3826834,0.1622099,0.8154936,0.5555702,-0.0,0.9238795,0.3826834,0.180238,0.9061278,0.3826834,0.1622099,0.8154936,0.5555702,0.1622099,0.8154936,0.5555702,0.180238,0.9061278,0.3826834,0.3181881,0.7681785,0.5555702,0.180238,0.9061278,0.3826834,0.3535516,0.8535541,0.3826834,0.3181881,0.7681785,0.5555702,0.3181881,0.7681785,0.5555702,0.3535516,0.8535541,0.3826834,0.4619384,0.6913427,0.5555702,0.3535516,0.8535541,0.3826834,0.5132784,0.7681788,0.3826834,0.4619384,0.6913427,0.5555702,0.4619384,0.6913427,0.5555702,0.5132784,0.7681788,0.3826834,0.5879367,0.587939,0.5555702,0.5132784,0.7681788,0.3826834,0.6532802,0.6532827,0.3826834,0.5879367,0.587939,0.5555702,0.5879367,0.587939,0.5555702,0.6532802,0.6532827,0.3826834,0.6913409,0.4619411,0.5555702,0.6532802,0.6532827,0.3826834,0.7681768,0.5132814,0.3826834,0.6913409,0.4619411,0.5555702,0.6913409,0.4619411,0.5555702,0.7681768,0.5132814,0.3826834,0.7681772,0.318191,0.5555702,0.7681768,0.5132814,0.3826834,0.8535528,0.3535549,0.3826834,0.7681772,0.318191,0.5555702,0.7681772,0.318191,0.5555702,0.8535528,0.3535549,0.3826834,0.8154929,0.1622131,0.5555702,0.8535528,0.3535549,0.3826834,0.9061271,0.1802415,0.3826834,0.8154929,0.1622131,0.5555702,0.8154929,0.1622131,0.5555702,0.9061271,0.1802415,0.3826834,0.8314697,0.0000013,0.5555702,0.9061271,0.1802415,0.3826834,0.9238795,0.0000015,0.3826834,0.8314697,0.0000013,0.5555702,0.8314697,0.0000013,0.5555702,0.9238795,0.0000015,0.3826834,0.8154934,-0.1622104,0.5555702,0.9238795,0.0000015,0.3826834,0.9061277,-0.1802386,0.3826834,0.8154934,-0.1622104,0.5555702,0.8154934,-0.1622104,0.5555702,0.9061277,-0.1802386,0.3826834,0.7681783,-0.3181885,0.5555702,0.9061277,-0.1802386,0.3826834,0.8535539,-0.3535521,0.3826834,0.7681783,-0.3181885,0.5555702,0.7681783,-0.3181885,0.5555702,0.8535539,-0.3535521,0.3826834,0.6913424,-0.4619389,0.5555702,0.8535539,-0.3535521,0.3826834,0.7681785,-0.5132789,0.3826834,0.6913424,-0.4619389,0.5555702,0.6913424,-0.4619389,0.5555702,0.7681785,-0.5132789,0.3826834,0.5879385,-0.5879371,0.5555702,0.7681785,-0.5132789,0.3826834,0.6532823,-0.6532806,0.3826834,0.5879385,-0.5879371,0.5555702,0.5879385,-0.5879371,0.5555702,0.6532823,-0.6532806,0.3826834,0.4619406,-0.6913412,0.5555702,0.6532823,-0.6532806,0.3826834,0.5132809,-0.7681771,0.3826834,0.4619406,-0.6913412,0.5555702,0.4619406,-0.6913412,0.5555702,0.5132809,-0.7681771,0.3826834,0.3181905,-0.7681774,0.5555702,0.5132809,-0.7681771,0.3826834,0.3535543,-0.8535529,0.3826834,0.3181905,-0.7681774,0.5555702,0.3181905,-0.7681774,0.5555702,0.3535543,-0.8535529,0.3826834,0.1622125,-0.815493,0.5555702,0.3535543,-0.8535529,0.3826834,0.1802409,-0.9061272,0.3826834,0.1622125,-0.815493,0.5555702,0.1622125,-0.815493,0.5555702,0.1802409,-0.9061272,0.3826834,0.0000008,-0.8314697,0.5555702,0.1802409,-0.9061272,0.3826834,0.0000009,-0.9238795,0.3826834,0.0000008,-0.8314697,0.5555702,0.0000008,-0.8314697,0.5555702,0.0000009,-0.9238795,0.3826834,-0.162211,-0.8154933,0.5555702,0.0000009,-0.9238795,0.3826834,-0.1802392,-0.9061276,0.3826834,-0.162211,-0.8154933,0.5555702,-0.162211,-0.8154933,0.5555702,-0.1802392,-0.9061276,0.3826834,-0.318189,-0.768178,0.5555702,-0.1802392,-0.9061276,0.3826834,-0.3535527,-0.8535537,0.3826834,-0.318189,-0.768178,0.5555702,-0.318189,-0.768178,0.5555702,-0.3535527,-0.8535537,0.3826834,-0.4619393,-0.6913421,0.5555702,-0.3535527,-0.8535537,0.3826834,-0.5132794,-0.7681781,0.3826834,-0.4619393,-0.6913421,0.5555702,-0.4619393,-0.6913421,0.5555702,-0.5132794,-0.7681781,0.3826834,-0.5879375,-0.5879382,0.5555702,-0.5132794,-0.7681781,0.3826834,-0.6532811,-0.6532819,0.3826834,-0.5879375,-0.5879382,0.5555702,-0.5879375,-0.5879382,0.5555702,-0.6532811,-0.6532819,0.3826834,-0.6913415,-0.4619402,0.5555702,-0.6532811,-0.6532819,0.3826834,-0.7681774,-0.5132804,0.3826834,-0.6913415,-0.4619402,0.5555702,-0.6913415,-0.4619402,0.5555702,-0.7681774,-0.5132804,0.3826834,-0.7681776,-0.31819,0.5555702,-0.7681774,-0.5132804,0.3826834,-0.8535532,-0.3535538,0.3826834,-0.7681776,-0.31819,0.5555702,-0.7681776,-0.31819,0.5555702,-0.8535532,-0.3535538,0.3826834,-0.8154931,-0.162212,0.5555702,-0.8535532,-0.3535538,0.3826834,-0.9061273,-0.1802403,0.3826834,-0.8154931,-0.162212,0.5555702,-0.8154931,-0.162212,0.5555702,-0.9061273,-0.1802403,0.3826834,-0.8314697,-0.0000003,0.5555702,-0.9061273,-0.1802403,0.3826834,-0.9238795,-0.0000003,0.3826834,-0.8314697,-0.0000003,0.5555702,-0.8314697,-0.0000003,0.5555702,-0.9238795,-0.0000003,0.3826834,-0.8154932,0.1622115,0.5555702,-0.9238795,-0.0000003,0.3826834,-0.9061275,0.1802397,0.3826834,-0.8154932,0.1622115,0.5555702,-0.8154932,0.1622115,0.5555702,-0.9061275,0.1802397,0.3826834,-0.7681779,0.3181895,0.5555702,-0.9061275,0.1802397,0.3826834,-0.8535535,0.3535532,0.3826834,-0.7681779,0.3181895,0.5555702,-0.7681779,0.3181895,0.5555702,-0.8535535,0.3535532,0.3826834,-0.6913418,0.4619398,0.5555702,-0.8535535,0.3535532,0.3826834,-0.7681777,0.5132799,0.3826834,-0.6913418,0.4619398,0.5555702,-0.6913418,0.4619398,0.5555702,-0.7681777,0.5132799,0.3826834,-0.5879378,0.5879378,0.5555702,-0.7681777,0.5132799,0.3826834,-0.6532815,0.6532815,0.3826834,-0.5879378,0.5879378,0.5555702,-0.5879378,0.5879378,0.5555702,-0.6532815,0.6532815,0.3826834,-0.4619398,0.6913418,0.5555702,-0.6532815,0.6532815,0.3826834,-0.5132799,0.7681777,0.3826834,-0.4619398,0.6913418,0.5555702,-0.4619398,0.6913418,0.5555702,-0.5132799,0.7681777,0.3826834,-0.3181896,0.7681778,0.5555702,-0.5132799,0.7681777,0.3826834,-0.3535534,0.8535534,0.3826834,-0.3181896,0.7681778,0.5555702,-0.3181896,0.7681778,0.5555702,-0.3535534,0.8535534,0.3826834,-0.1622117,0.8154932,0.5555702,-0.3535534,0.8535534,0.3826834,-0.1802399,0.9061274,0.3826834,-0.1622117,0.8154932,0.5555702,-0.1622117,0.8154932,0.5555702,-0.1802399,0.9061274,0.3826834,-0.0,0.8314697,0.5555702,-0.1802399,0.9061274,0.3826834,-0.0,0.9238795,0.3826834,-0.0,0.8314697,0.5555702,-0.0,0.7071068,0.7071068,-0.0,0.8314697,0.5555702,0.1379482,0.6935202,0.7071068,-0.0,0.8314697,0.5555702,0.1622099,0.8154936,0.5555702,0.1379482,0.6935202,0.7071068,0.1379482,0.6935202,0.7071068,0.1622099,0.8154936,0.5555702,0.2705967,0.653282,0.7071068,0.1622099,0.8154936,0.5555702,0.3181881,0.7681785,0.5555702,0.2705967,0.653282,0.7071068,0.2705967,0.653282,0.7071068,0.3181881,0.7681785,0.5555702,0.3928463,0.5879386,0.7071068,0.3181881,0.7681785,0.5555702,0.4619384,0.6913427,0.5555702,0.3928463,0.5879386,0.7071068,0.3928463,0.5879386,0.7071068,0.4619384,0.6913427,0.5555702,0.499999,0.500001,0.7071068,0.4619384,0.6913427,0.5555702,0.5879367,0.587939,0.5555702,0.499999,0.500001,0.7071068,0.499999,0.500001,0.7071068,0.5879367,0.587939,0.5555702,0.5879371,0.3928486,0.7071068,0.5879367,0.587939,0.5555702,0.6913409,0.4619411,0.5555702,0.5879371,0.3928486,0.7071068,0.5879371,0.3928486,0.7071068,0.6913409,0.4619411,0.5555702,0.653281,0.2705992,0.7071068,0.6913409,0.4619411,0.5555702,0.7681772,0.318191,0.5555702,0.653281,0.2705992,0.7071068,0.653281,0.2705992,0.7071068,0.7681772,0.318191,0.5555702,0.6935197,0.1379509,0.7071068,0.7681772,0.318191,0.5555702,0.8154929,0.1622131,0.5555702,0.6935197,0.1379509,0.7071068,0.6935197,0.1379509,0.7071068,0.8154929,0.1622131,0.5555702,0.7071068,0.0000011,0.7071068,0.8154929,0.1622131,0.5555702,0.8314697,0.0000013,0.5555702,0.7071068,0.0000011,0.7071068,0.7071068,0.0000011,0.7071068,0.8314697,0.0000013,0.5555702,0.6935201,-0.1379486,0.7071068,0.8314697,0.0000013,0.5555702,0.8154934,-0.1622104,0.5555702,0.6935201,-0.1379486,0.7071068,0.6935201,-0.1379486,0.7071068,0.8154934,-0.1622104,0.5555702,0.6532819,-0.2705971,0.7071068,0.8154934,-0.1622104,0.5555702,0.7681783,-0.3181885,0.5555702,0.6532819,-0.2705971,0.7071068,0.6532819,-0.2705971,0.7071068,0.7681783,-0.3181885,0.5555702,0.5879383,-0.3928467,0.7071068,0.7681783,-0.3181885,0.5555702,0.6913424,-0.4619389,0.5555702,0.5879383,-0.3928467,0.7071068,0.5879383,-0.3928467,0.7071068,0.6913424,-0.4619389,0.5555702,0.5000006,-0.4999993,0.7071068,0.6913424,-0.4619389,0.5555702,0.5879385,-0.5879371,0.5555702,0.5000006,-0.4999993,0.7071068,0.5000006,-0.4999993,0.7071068,0.5879385,-0.5879371,0.5555702,0.3928482,-0.5879373,0.7071068,0.5879385,-0.5879371,0.5555702,0.4619406,-0.6913412,0.5555702,0.3928482,-0.5879373,0.7071068,0.3928482,-0.5879373,0.7071068,0.4619406,-0.6913412,0.5555702,0.2705988,-0.6532812,0.7071068,0.4619406,-0.6913412,0.5555702,0.3181905,-0.7681774,0.5555702,0.2705988,-0.6532812,0.7071068,0.2705988,-0.6532812,0.7071068,0.3181905,-0.7681774,0.5555702,0.1379504,-0.6935198,0.7071068,0.3181905,-0.7681774,0.5555702,0.1622125,-0.815493,0.5555702,0.1379504,-0.6935198,0.7071068,0.1379504,-0.6935198,0.7071068,0.1622125,-0.815493,0.5555702,0.0000007,-0.7071068,0.7071068,0.1622125,-0.815493,0.5555702,0.0000008,-0.8314697,0.5555702,0.0000007,-0.7071068,0.7071068,0.0000007,-0.7071068,0.7071068,0.0000008,-0.8314697,0.5555702,-0.1379491,-0.69352,0.7071068,0.0000008,-0.8314697,0.5555702,-0.162211,-0.8154933,0.5555702,-0.1379491,-0.69352,0.7071068,-0.1379491,-0.69352,0.7071068,-0.162211,-0.8154933,0.5555702,-0.2705975,-0.6532817,0.7071068,-0.162211,-0.8154933,0.5555702,-0.318189,-0.768178,0.5555702,-0.2705975,-0.6532817,0.7071068,-0.2705975,-0.6532817,0.7071068,-0.318189,-0.768178,0.5555702,-0.3928471,-0.5879381,0.7071068,-0.318189,-0.768178,0.5555702,-0.4619393,-0.6913421,0.5555702,-0.3928471,-0.5879381,0.7071068,-0.3928471,-0.5879381,0.7071068,-0.4619393,-0.6913421,0.5555702,-0.4999997,-0.5000003,0.7071068,-0.4619393,-0.6913421,0.5555702,-0.5879375,-0.5879382,0.5555702,-0.4999997,-0.5000003,0.7071068,-0.4999997,-0.5000003,0.7071068,-0.5879375,-0.5879382,0.5555702,-0.5879376,-0.3928478,0.7071068,-0.5879375,-0.5879382,0.5555702,-0.6913415,-0.4619402,0.5555702,-0.5879376,-0.3928478,0.7071068,-0.5879376,-0.3928478,0.7071068,-0.6913415,-0.4619402,0.5555702,-0.6532813,-0.2705984,0.7071068,-0.6913415,-0.4619402,0.5555702,-0.7681776,-0.31819,0.5555702,-0.6532813,-0.2705984,0.7071068,-0.6532813,-0.2705984,0.7071068,-0.7681776,-0.31819,0.5555702,-0.6935198,-0.13795,0.7071068,-0.7681776,-0.31819,0.5555702,-0.8154931,-0.162212,0.5555702,-0.6935198,-0.13795,0.7071068,-0.6935198,-0.13795,0.7071068,-0.8154931,-0.162212,0.5555702,-0.7071068,-0.0000002,0.7071068,-0.8154931,-0.162212,0.5555702,-0.8314697,-0.0000003,0.5555702,-0.7071068,-0.0000002,0.7071068,-0.7071068,-0.0000002,0.7071068,-0.8314697,-0.0000003,0.5555702,-0.6935199,0.1379495,0.7071068,-0.8314697,-0.0000003,0.5555702,-0.8154932,0.1622115,0.5555702,-0.6935199,0.1379495,0.7071068,-0.6935199,0.1379495,0.7071068,-0.8154932,0.1622115,0.5555702,-0.6532815,0.2705979,0.7071068,-0.8154932,0.1622115,0.5555702,-0.7681779,0.3181895,0.5555702,-0.6532815,0.2705979,0.7071068,-0.6532815,0.2705979,0.7071068,-0.7681779,0.3181895,0.5555702,-0.5879378,0.3928474,0.7071068,-0.7681779,0.3181895,0.5555702,-0.6913418,0.4619398,0.5555702,-0.5879378,0.3928474,0.7071068,-0.5879378,0.3928474,0.7071068,-0.6913418,0.4619398,0.5555702,-0.5,0.5,0.7071068,-0.6913418,0.4619398,0.5555702,-0.5879378,0.5879378,0.5555702,-0.5,0.5,0.7071068,-0.5,0.5,0.7071068,-0.5879378,0.5879378,0.5555702,-0.3928474,0.5879378,0.7071068,-0.5879378,0.5879378,0.5555702,-0.4619398,0.6913418,0.5555702,-0.3928474,0.5879378,0.7071068,-0.3928474,0.5879378,0.7071068,-0.4619398,0.6913418,0.5555702,-0.270598,0.6532815,0.7071068,-0.4619398,0.6913418,0.5555702,-0.3181896,0.7681778,0.5555702,-0.270598,0.6532815,0.7071068,-0.270598,0.6532815,0.7071068,-0.3181896,0.7681778,0.5555702,-0.1379497,0.6935199,0.7071068,-0.3181896,0.7681778,0.5555702,-0.1622117,0.8154932,0.5555702,-0.1379497,0.6935199,0.7071068,-0.1379497,0.6935199,0.7071068,-0.1622117,0.8154932,0.5555702,-0.0,0.7071068,0.7071068,-0.1622117,0.8154932,0.5555702,-0.0,0.8314697,0.5555702,-0.0,0.7071068,0.7071068,-0.0,0.5555702,0.8314696,-0.0,0.7071068,0.7071068,0.1083852,0.5448954,0.8314696,-0.0,0.7071068,0.7071068,0.1379482,0.6935202,0.7071068,0.1083852,0.5448954,0.8314696,0.1083852,0.5448954,0.8314696,0.1379482,0.6935202,0.7071068,0.2126065,0.5132805,0.8314696,0.1379482,0.6935202,0.7071068,0.2705967,0.653282,0.7071068,0.2126065,0.5132805,0.8314696,0.2126065,0.5132805,0.8314696,0.2705967,0.653282,0.7071068,0.3086574,0.4619404,0.8314696,0.2705967,0.653282,0.7071068,0.3928463,0.5879386,0.7071068,0.3086574,0.4619404,0.8314696,0.3086574,0.4619404,0.8314696,0.3928463,0.5879386,0.7071068,0.3928467,0.3928483,0.8314696,0.3928463,0.5879386,0.7071068,0.499999,0.500001,0.7071068,0.3928467,0.3928483,0.8314696,0.3928467,0.3928483,0.8314696,0.499999,0.500001,0.7071068,0.4619392,0.3086592,0.8314696,0.499999,0.500001,0.7071068,0.5879371,0.3928486,0.7071068,0.4619392,0.3086592,0.8314696,0.4619392,0.3086592,0.8314696,0.5879371,0.3928486,0.7071068,0.5132796,0.2126084,0.8314696,0.5879371,0.3928486,0.7071068,0.653281,0.2705992,0.7071068,0.5132796,0.2126084,0.8314696,0.5132796,0.2126084,0.8314696,0.653281,0.2705992,0.7071068,0.5448949,0.1083873,0.8314696,0.653281,0.2705992,0.7071068,0.6935197,0.1379509,0.7071068,0.5448949,0.1083873,0.8314696,0.5448949,0.1083873,0.8314696,0.6935197,0.1379509,0.7071068,0.5555702,0.0000009,0.8314696,0.6935197,0.1379509,0.7071068,0.7071068,0.0000011,0.7071068,0.5555702,0.0000009,0.8314696,0.5555702,0.0000009,0.8314696,0.7071068,0.0000011,0.7071068,0.5448953,-0.1083855,0.8314696,0.7071068,0.0000011,0.7071068,0.6935201,-0.1379486,0.7071068,0.5448953,-0.1083855,0.8314696,0.5448953,-0.1083855,0.8314696,0.6935201,-0.1379486,0.7071068,0.5132803,-0.2126068,0.8314696,0.6935201,-0.1379486,0.7071068,0.6532819,-0.2705971,0.7071068,0.5132803,-0.2126068,0.8314696,0.5132803,-0.2126068,0.8314696,0.6532819,-0.2705971,0.7071068,0.4619402,-0.3086577,0.8314696,0.6532819,-0.2705971,0.7071068,0.5879383,-0.3928467,0.7071068,0.4619402,-0.3086577,0.8314696,0.4619402,-0.3086577,0.8314696,0.5879383,-0.3928467,0.7071068,0.392848,-0.392847,0.8314696,0.5879383,-0.3928467,0.7071068,0.5000006,-0.4999993,0.7071068,0.392848,-0.392847,0.8314696,0.392848,-0.392847,0.8314696,0.5000006,-0.4999993,0.7071068,0.3086589,-0.4619394,0.8314696,0.5000006,-0.4999993,0.7071068,0.3928482,-0.5879373,0.7071068,0.3086589,-0.4619394,0.8314696,0.3086589,-0.4619394,0.8314696,0.3928482,-0.5879373,0.7071068,0.2126081,-0.5132797,0.8314696,0.3928482,-0.5879373,0.7071068,0.2705988,-0.6532812,0.7071068,0.2126081,-0.5132797,0.8314696,0.2126081,-0.5132797,0.8314696,0.2705988,-0.6532812,0.7071068,0.1083869,-0.544895,0.8314696,0.2705988,-0.6532812,0.7071068,0.1379504,-0.6935198,0.7071068,0.1083869,-0.544895,0.8314696,0.1083869,-0.544895,0.8314696,0.1379504,-0.6935198,0.7071068,0.0000005,-0.5555702,0.8314696,0.1379504,-0.6935198,0.7071068,0.0000007,-0.7071068,0.7071068,0.0000005,-0.5555702,0.8314696,0.0000005,-0.5555702,0.8314696,0.0000007,-0.7071068,0.7071068,-0.1083859,-0.5448952,0.8314696,0.0000007,-0.7071068,0.7071068,-0.1379491,-0.69352,0.7071068,-0.1083859,-0.5448952,0.8314696,-0.1083859,-0.5448952,0.8314696,-0.1379491,-0.69352,0.7071068,-0.2126071,-0.5132802,0.8314696,-0.1379491,-0.69352,0.7071068,-0.2705975,-0.6532817,0.7071068,-0.2126071,-0.5132802,0.8314696,-0.2126071,-0.5132802,0.8314696,-0.2705975,-0.6532817,0.7071068,-0.308658,-0.46194,0.8314696,-0.2705975,-0.6532817,0.7071068,-0.3928471,-0.5879381,0.7071068,-0.308658,-0.46194,0.8314696,-0.308658,-0.46194,0.8314696,-0.3928471,-0.5879381,0.7071068,-0.3928472,-0.3928477,0.8314696,-0.3928471,-0.5879381,0.7071068,-0.4999997,-0.5000003,0.7071068,-0.3928472,-0.3928477,0.8314696,-0.3928472,-0.3928477,0.8314696,-0.4999997,-0.5000003,0.7071068,-0.4619396,-0.3086586,0.8314696,-0.4999997,-0.5000003,0.7071068,-0.5879376,-0.3928478,0.7071068,-0.4619396,-0.3086586,0.8314696,-0.4619396,-0.3086586,0.8314696,-0.5879376,-0.3928478,0.7071068,-0.5132799,-0.2126078,0.8314696,-0.5879376,-0.3928478,0.7071068,-0.6532813,-0.2705984,0.7071068,-0.5132799,-0.2126078,0.8314696,-0.5132799,-0.2126078,0.8314696,-0.6532813,-0.2705984,0.7071068,-0.5448951,-0.1083866,0.8314696,-0.6532813,-0.2705984,0.7071068,-0.6935198,-0.13795,0.7071068,-0.5448951,-0.1083866,0.8314696,-0.5448951,-0.1083866,0.8314696,-0.6935198,-0.13795,0.7071068,-0.5555702,-0.0000002,0.8314696,-0.6935198,-0.13795,0.7071068,-0.7071068,-0.0000002,0.7071068,-0.5555702,-0.0000002,0.8314696,-0.5555702,-0.0000002,0.8314696,-0.7071068,-0.0000002,0.7071068,-0.5448951,0.1083862,0.8314696,-0.7071068,-0.0000002,0.7071068,-0.6935199,0.1379495,0.7071068,-0.5448951,0.1083862,0.8314696,-0.5448951,0.1083862,0.8314696,-0.6935199,0.1379495,0.7071068,-0.51328,0.2126074,0.8314696,-0.6935199,0.1379495,0.7071068,-0.6532815,0.2705979,0.7071068,-0.51328,0.2126074,0.8314696,-0.51328,0.2126074,0.8314696,-0.6532815,0.2705979,0.7071068,-0.4619398,0.3086583,0.8314696,-0.6532815,0.2705979,0.7071068,-0.5879378,0.3928474,0.7071068,-0.4619398,0.3086583,0.8314696,-0.4619398,0.3086583,0.8314696,-0.5879378,0.3928474,0.7071068,-0.3928475,0.3928475,0.8314696,-0.5879378,0.3928474,0.7071068,-0.5,0.5,0.7071068,-0.3928475,0.3928475,0.8314696,-0.3928475,0.3928475,0.8314696,-0.5,0.5,0.7071068,-0.3086583,0.4619398,0.8314696,-0.5,0.5,0.7071068,-0.3928474,0.5879378,0.7071068,-0.3086583,0.4619398,0.8314696,-0.3086583,0.4619398,0.8314696,-0.3928474,0.5879378,0.7071068,-0.2126075,0.51328,0.8314696,-0.3928474,0.5879378,0.7071068,-0.270598,0.6532815,0.7071068,-0.2126075,0.51328,0.8314696,-0.2126075,0.51328,0.8314696,-0.270598,0.6532815,0.7071068,-0.1083864,0.5448951,0.8314696,-0.270598,0.6532815,0.7071068,-0.1379497,0.6935199,0.7071068,-0.1083864,0.5448951,0.8314696,-0.1083864,0.5448951,0.8314696,-0.1379497,0.6935199,0.7071068,-0.0,0.5555702,0.8314696,-0.1379497,0.6935199,0.7071068,-0.0,0.7071068,0.7071068,-0.0,0.5555702,0.8314696,-0.0,0.3826835,0.9238795,-0.0,0.5555702,0.8314696,0.074657,0.3753305,0.9238795,-0.0,0.5555702,0.8314696,0.1083852,0.5448954,0.8314696,0.074657,0.3753305,0.9238795,0.074657,0.3753305,0.9238795,0.1083852,0.5448954,0.8314696,0.1464459,0.3535537,0.9238795,0.1083852,0.5448954,0.8314696,0.2126065,0.5132805,0.8314696,0.1464459,0.3535537,0.9238795,0.1464459,0.3535537,0.9238795,0.2126065,0.5132805,0.8314696,0.2126069,0.3181901,0.9238795,0.2126065,0.5132805,0.8314696,0.3086574,0.4619404,0.8314696,0.2126069,0.3181901,0.9238795,0.2126069,0.3181901,0.9238795,0.3086574,0.4619404,0.8314696,0.2705975,0.2705986,0.9238795,0.3086574,0.4619404,0.8314696,0.3928467,0.3928483,0.8314696,0.2705975,0.2705986,0.9238795,0.2705975,0.2705986,0.9238795,0.3928467,0.3928483,0.8314696,0.3181893,0.2126081,0.9238795,0.3928467,0.3928483,0.8314696,0.4619392,0.3086592,0.8314696,0.3181893,0.2126081,0.9238795,0.3181893,0.2126081,0.9238795,0.4619392,0.3086592,0.8314696,0.3535531,0.1464472,0.9238795,0.4619392,0.3086592,0.8314696,0.5132796,0.2126084,0.8314696,0.3535531,0.1464472,0.9238795,0.3535531,0.1464472,0.9238795,0.5132796,0.2126084,0.8314696,0.3753302,0.0746585,0.9238795,0.5132796,0.2126084,0.8314696,0.5448949,0.1083873,0.8314696,0.3753302,0.0746585,0.9238795,0.3753302,0.0746585,0.9238795,0.5448949,0.1083873,0.8314696,0.3826835,0.0000006,0.9238795,0.5448949,0.1083873,0.8314696,0.5555702,0.0000009,0.8314696,0.3826835,0.0000006,0.9238795,0.3826835,0.0000006,0.9238795,0.5555702,0.0000009,0.8314696,0.3753304,-0.0746573,0.9238795,0.5555702,0.0000009,0.8314696,0.5448953,-0.1083855,0.8314696,0.3753304,-0.0746573,0.9238795,0.3753304,-0.0746573,0.9238795,0.5448953,-0.1083855,0.8314696,0.3535536,-0.1464461,0.9238795,0.5448953,-0.1083855,0.8314696,0.5132803,-0.2126068,0.8314696,0.3535536,-0.1464461,0.9238795,0.3535536,-0.1464461,0.9238795,0.5132803,-0.2126068,0.8314696,0.3181899,-0.2126071,0.9238795,0.5132803,-0.2126068,0.8314696,0.4619402,-0.3086577,0.8314696,0.3181899,-0.2126071,0.9238795,0.3181899,-0.2126071,0.9238795,0.4619402,-0.3086577,0.8314696,0.2705984,-0.2705977,0.9238795,0.4619402,-0.3086577,0.8314696,0.392848,-0.392847,0.8314696,0.2705984,-0.2705977,0.9238795,0.2705984,-0.2705977,0.9238795,0.392848,-0.392847,0.8314696,0.2126079,-0.3181894,0.9238795,0.392848,-0.392847,0.8314696,0.3086589,-0.4619394,0.8314696,0.2126079,-0.3181894,0.9238795,0.2126079,-0.3181894,0.9238795,0.3086589,-0.4619394,0.8314696,0.146447,-0.3535532,0.9238795,0.3086589,-0.4619394,0.8314696,0.2126081,-0.5132797,0.8314696,0.146447,-0.3535532,0.9238795,0.146447,-0.3535532,0.9238795,0.2126081,-0.5132797,0.8314696,0.0746582,-0.3753302,0.9238795,0.2126081,-0.5132797,0.8314696,0.1083869,-0.544895,0.8314696,0.0746582,-0.3753302,0.9238795,0.0746582,-0.3753302,0.9238795,0.1083869,-0.544895,0.8314696,0.0000004,-0.3826835,0.9238795,0.1083869,-0.544895,0.8314696,0.0000005,-0.5555702,0.8314696,0.0000004,-0.3826835,0.9238795,0.0000004,-0.3826835,0.9238795,0.0000005,-0.5555702,0.8314696,-0.0746575,-0.3753304,0.9238795,0.0000005,-0.5555702,0.8314696,-0.1083859,-0.5448952,0.8314696,-0.0746575,-0.3753304,0.9238795,-0.0746575,-0.3753304,0.9238795,-0.1083859,-0.5448952,0.8314696,-0.1464463,-0.3535535,0.9238795,-0.1083859,-0.5448952,0.8314696,-0.2126071,-0.5132802,0.8314696,-0.1464463,-0.3535535,0.9238795,-0.1464463,-0.3535535,0.9238795,-0.2126071,-0.5132802,0.8314696,-0.2126073,-0.3181898,0.9238795,-0.2126071,-0.5132802,0.8314696,-0.308658,-0.46194,0.8314696,-0.2126073,-0.3181898,0.9238795,-0.2126073,-0.3181898,0.9238795,-0.308658,-0.46194,0.8314696,-0.2705979,-0.2705982,0.9238795,-0.308658,-0.46194,0.8314696,-0.3928472,-0.3928477,0.8314696,-0.2705979,-0.2705982,0.9238795,-0.2705979,-0.2705982,0.9238795,-0.3928472,-0.3928477,0.8314696,-0.3181895,-0.2126077,0.9238795,-0.3928472,-0.3928477,0.8314696,-0.4619396,-0.3086586,0.8314696,-0.3181895,-0.2126077,0.9238795,-0.3181895,-0.2126077,0.9238795,-0.4619396,-0.3086586,0.8314696,-0.3535533,-0.1464468,0.9238795,-0.4619396,-0.3086586,0.8314696,-0.5132799,-0.2126078,0.8314696,-0.3535533,-0.1464468,0.9238795,-0.3535533,-0.1464468,0.9238795,-0.5132799,-0.2126078,0.8314696,-0.3753303,-0.074658,0.9238795,-0.5132799,-0.2126078,0.8314696,-0.5448951,-0.1083866,0.8314696,-0.3753303,-0.074658,0.9238795,-0.3753303,-0.074658,0.9238795,-0.5448951,-0.1083866,0.8314696,-0.3826835,-0.0000001,0.9238795,-0.5448951,-0.1083866,0.8314696,-0.5555702,-0.0000002,0.8314696,-0.3826835,-0.0000001,0.9238795,-0.3826835,-0.0000001,0.9238795,-0.5555702,-0.0000002,0.8314696,-0.3753303,0.0746578,0.9238795,-0.5555702,-0.0000002,0.8314696,-0.5448951,0.1083862,0.8314696,-0.3753303,0.0746578,0.9238795,-0.3753303,0.0746578,0.9238795,-0.5448951,0.1083862,0.8314696,-0.3535534,0.1464466,0.9238795,-0.5448951,0.1083862,0.8314696,-0.51328,0.2126074,0.8314696,-0.3535534,0.1464466,0.9238795,-0.3535534,0.1464466,0.9238795,-0.51328,0.2126074,0.8314696,-0.3181897,0.2126075,0.9238795,-0.51328,0.2126074,0.8314696,-0.4619398,0.3086583,0.8314696,-0.3181897,0.2126075,0.9238795,-0.3181897,0.2126075,0.9238795,-0.4619398,0.3086583,0.8314696,-0.2705981,0.2705981,0.9238795,-0.4619398,0.3086583,0.8314696,-0.3928475,0.3928475,0.8314696,-0.2705981,0.2705981,0.9238795,-0.2705981,0.2705981,0.9238795,-0.3928475,0.3928475,0.8314696,-0.2126075,0.3181897,0.9238795,-0.3928475,0.3928475,0.8314696,-0.3086583,0.4619398,0.8314696,-0.2126075,0.3181897,0.9238795,-0.2126075,0.3181897,0.9238795,-0.3086583,0.4619398,0.8314696,-0.1464466,0.3535534,0.9238795,-0.3086583,0.4619398,0.8314696,-0.2126075,0.51328,0.8314696,-0.1464466,0.3535534,0.9238795,-0.1464466,0.3535534,0.9238795,-0.2126075,0.51328,0.8314696,-0.0746578,0.3753303,0.9238795,-0.2126075,0.51328,0.8314696,-0.1083864,0.5448951,0.8314696,-0.0746578,0.3753303,0.9238795,-0.0746578,0.3753303,0.9238795,-0.1083864,0.5448951,0.8314696,-0.0,0.3826835,0.9238795,-0.1083864,0.5448951,0.8314696,-0.0,0.5555702,0.8314696,-0.0,0.3826835,0.9238795,-0.0,0.1950903,0.9807853,-0.0,0.3826835,0.9238795,0.0380598,0.1913418,0.9807853,-0.0,0.3826835,0.9238795,0.074657,0.3753305,0.9238795,0.0380598,0.1913418,0.9807853,0.0380598,0.1913418,0.9807853,0.074657,0.3753305,0.9238795,0.0746575,0.1802401,0.9807853,0.074657,0.3753305,0.9238795,0.1464459,0.3535537,0.9238795,0.0746575,0.1802401,0.9807853,0.0746575,0.1802401,0.9807853,0.1464459,0.3535537,0.9238795,0.1083861,0.1622119,0.9807853,0.1464459,0.3535537,0.9238795,0.2126069,0.3181901,0.9238795,0.1083861,0.1622119,0.9807853,0.1083861,0.1622119,0.9807853,0.2126069,0.3181901,0.9238795,0.1379494,0.13795,0.9807853,0.2126069,0.3181901,0.9238795,0.2705975,0.2705986,0.9238795,0.1379494,0.13795,0.9807853,0.1379494,0.13795,0.9807853,0.2705975,0.2705986,0.9238795,0.1622115,0.1083867,0.9807853,0.2705975,0.2705986,0.9238795,0.3181893,0.2126081,0.9238795,0.1622115,0.1083867,0.9807853,0.1622115,0.1083867,0.9807853,0.3181893,0.2126081,0.9238795,0.1802398,0.0746582,0.9807853,0.3181893,0.2126081,0.9238795,0.3535531,0.1464472,0.9238795,0.1802398,0.0746582,0.9807853,0.1802398,0.0746582,0.9807853,0.3535531,0.1464472,0.9238795,0.1913417,0.0380606,0.9807853,0.3535531,0.1464472,0.9238795,0.3753302,0.0746585,0.9238795,0.1913417,0.0380606,0.9807853,0.1913417,0.0380606,0.9807853,0.3753302,0.0746585,0.9238795,0.1950903,0.0000003,0.9807853,0.3753302,0.0746585,0.9238795,0.3826835,0.0000006,0.9238795,0.1950903,0.0000003,0.9807853,0.1950903,0.0000003,0.9807853,0.3826835,0.0000006,0.9238795,0.1913418,-0.0380599,0.9807853,0.3826835,0.0000006,0.9238795,0.3753304,-0.0746573,0.9238795,0.1913418,-0.0380599,0.9807853,0.1913418,-0.0380599,0.9807853,0.3753304,-0.0746573,0.9238795,0.1802401,-0.0746576,0.9807853,0.3753304,-0.0746573,0.9238795,0.3535536,-0.1464461,0.9238795,0.1802401,-0.0746576,0.9807853,0.1802401,-0.0746576,0.9807853,0.3535536,-0.1464461,0.9238795,0.1622118,-0.1083862,0.9807853,0.3535536,-0.1464461,0.9238795,0.3181899,-0.2126071,0.9238795,0.1622118,-0.1083862,0.9807853,0.1622118,-0.1083862,0.9807853,0.3181899,-0.2126071,0.9238795,0.1379499,-0.1379495,0.9807853,0.3181899,-0.2126071,0.9238795,0.2705984,-0.2705977,0.9238795,0.1379499,-0.1379495,0.9807853,0.1379499,-0.1379495,0.9807853,0.2705984,-0.2705977,0.9238795,0.1083866,-0.1622115,0.9807853,0.2705984,-0.2705977,0.9238795,0.2126079,-0.3181894,0.9238795,0.1083866,-0.1622115,0.9807853,0.1083866,-0.1622115,0.9807853,0.2126079,-0.3181894,0.9238795,0.074658,-0.1802399,0.9807853,0.2126079,-0.3181894,0.9238795,0.146447,-0.3535532,0.9238795,0.074658,-0.1802399,0.9807853,0.074658,-0.1802399,0.9807853,0.146447,-0.3535532,0.9238795,0.0380604,-0.1913417,0.9807853,0.146447,-0.3535532,0.9238795,0.0746582,-0.3753302,0.9238795,0.0380604,-0.1913417,0.9807853,0.0380604,-0.1913417,0.9807853,0.0746582,-0.3753302,0.9238795,0.0000002,-0.1950903,0.9807853,0.0746582,-0.3753302,0.9238795,0.0000004,-0.3826835,0.9238795,0.0000002,-0.1950903,0.9807853,0.0000002,-0.1950903,0.9807853,0.0000004,-0.3826835,0.9238795,-0.0380601,-0.1913417,0.9807853,0.0000004,-0.3826835,0.9238795,-0.0746575,-0.3753304,0.9238795,-0.0380601,-0.1913417,0.9807853,-0.0380601,-0.1913417,0.9807853,-0.0746575,-0.3753304,0.9238795,-0.0746577,-0.18024,0.9807853,-0.0746575,-0.3753304,0.9238795,-0.1464463,-0.3535535,0.9238795,-0.0746577,-0.18024,0.9807853,-0.0746577,-0.18024,0.9807853,-0.1464463,-0.3535535,0.9238795,-0.1083863,-0.1622118,0.9807853,-0.1464463,-0.3535535,0.9238795,-0.2126073,-0.3181898,0.9238795,-0.1083863,-0.1622118,0.9807853,-0.1083863,-0.1622118,0.9807853,-0.2126073,-0.3181898,0.9238795,-0.1379496,-0.1379498,0.9807853,-0.2126073,-0.3181898,0.9238795,-0.2705979,-0.2705982,0.9238795,-0.1379496,-0.1379498,0.9807853,-0.1379496,-0.1379498,0.9807853,-0.2705979,-0.2705982,0.9238795,-0.1622116,-0.1083865,0.9807853,-0.2705979,-0.2705982,0.9238795,-0.3181895,-0.2126077,0.9238795,-0.1622116,-0.1083865,0.9807853,-0.1622116,-0.1083865,0.9807853,-0.3181895,-0.2126077,0.9238795,-0.1802399,-0.0746579,0.9807853,-0.3181895,-0.2126077,0.9238795,-0.3535533,-0.1464468,0.9238795,-0.1802399,-0.0746579,0.9807853,-0.1802399,-0.0746579,0.9807853,-0.3535533,-0.1464468,0.9238795,-0.1913417,-0.0380603,0.9807853,-0.3535533,-0.1464468,0.9238795,-0.3753303,-0.074658,0.9238795,-0.1913417,-0.0380603,0.9807853,-0.1913417,-0.0380603,0.9807853,-0.3753303,-0.074658,0.9238795,-0.1950903,-0.0000001,0.9807853,-0.3753303,-0.074658,0.9238795,-0.3826835,-0.0000001,0.9238795,-0.1950903,-0.0000001,0.9807853,-0.1950903,-0.0000001,0.9807853,-0.3826835,-0.0000001,0.9238795,-0.1913417,0.0380602,0.9807853,-0.3826835,-0.0000001,0.9238795,-0.3753303,0.0746578,0.9238795,-0.1913417,0.0380602,0.9807853,-0.1913417,0.0380602,0.9807853,-0.3753303,0.0746578,0.9238795,-0.18024,0.0746578,0.9807853,-0.3753303,0.0746578,0.9238795,-0.3535534,0.1464466,0.9238795,-0.18024,0.0746578,0.9807853,-0.18024,0.0746578,0.9807853,-0.3535534,0.1464466,0.9238795,-0.1622117,0.1083864,0.9807853,-0.3535534,0.1464466,0.9238795,-0.3181897,0.2126075,0.9238795,-0.1622117,0.1083864,0.9807853,-0.1622117,0.1083864,0.9807853,-0.3181897,0.2126075,0.9238795,-0.1379497,0.1379497,0.9807853,-0.3181897,0.2126075,0.9238795,-0.2705981,0.2705981,0.9238795,-0.1379497,0.1379497,0.9807853,-0.1379497,0.1379497,0.9807853,-0.2705981,0.2705981,0.9238795,-0.1083864,0.1622117,0.9807853,-0.2705981,0.2705981,0.9238795,-0.2126075,0.3181897,0.9238795,-0.1083864,0.1622117,0.9807853,-0.1083864,0.1622117,0.9807853,-0.2126075,0.3181897,0.9238795,-0.0746578,0.18024,0.9807853,-0.2126075,0.3181897,0.9238795,-0.1464466,0.3535534,0.9238795,-0.0746578,0.18024,0.9807853,-0.0746578,0.18024,0.9807853,-0.1464466,0.3535534,0.9238795,-0.0380602,0.1913417,0.9807853,-0.1464466,0.3535534,0.9238795,-0.0746578,0.3753303,0.9238795,-0.0380602,0.1913417,0.9807853,-0.0380602,0.1913417,0.9807853,-0.0746578,0.3753303,0.9238795,-0.0,0.1950903,0.9807853,-0.0746578,0.3753303,0.9238795,-0.0,0.3826835,0.9238795,-0.0,0.1950903,0.9807853,-0.0,0.1950903,0.9807853,0.0380598,0.1913418,0.9807853,0.0,0.0,1.0,0.0380598,0.1913418,0.9807853,0.0746575,0.1802401,0.9807853,0.0,0.0,1.0,0.0746575,0.1802401,0.9807853,0.1083861,0.1622119,0.9807853,0.0,0.0,1.0,0.1083861,0.1622119,0.9807853,0.1379494,0.13795,0.9807853,0.0,0.0,1.0,0.1379494,0.13795,0.9807853,0.1622115,0.1083867,0.9807853,0.0,0.0,1.0,0.1622115,0.1083867,0.9807853,0.1802398,0.0746582,0.9807853,0.0,0.0,1.0,0.1802398,0.0746582,0.9807853,0.1913417,0.0380606,0.9807853,0.0,0.0,1.0,0.1913417,0.0380606,0.9807853,0.1950903,0.0000003,0.9807853,0.0,0.0,1.0,0.1950903,0.0000003,0.9807853,0.1913418,-0.0380599,0.9807853,0.0,0.0,1.0,0.1913418,-0.0380599,0.9807853,0.1802401,-0.0746576,0.9807853,0.0,0.0,1.0,0.1802401,-0.0746576,0.9807853,0.1622118,-0.1083862,0.9807853,0.0,0.0,1.0,0.1622118,-0.1083862,0.9807853,0.1379499,-0.1379495,0.9807853,0.0,0.0,1.0,0.1379499,-0.1379495,0.9807853,0.1083866,-0.1622115,0.9807853,0.0,0.0,1.0,0.1083866,-0.1622115,0.9807853,0.074658,-0.1802399,0.9807853,0.0,0.0,1.0,0.074658,-0.1802399,0.9807853,0.0380604,-0.1913417,0.9807853,0.0,0.0,1.0,0.0380604,-0.1913417,0.9807853,0.0000002,-0.1950903,0.9807853,0.0,0.0,1.0,0.0000002,-0.1950903,0.9807853,-0.0380601,-0.1913417,0.9807853,0.0,0.0,1.0,-0.0380601,-0.1913417,0.9807853,-0.0746577,-0.18024,0.9807853,0.0,0.0,1.0,-0.0746577,-0.18024,0.9807853,-0.1083863,-0.1622118,0.9807853,0.0,0.0,1.0,-0.1083863,-0.1622118,0.9807853,-0.1379496,-0.1379498,0.9807853,0.0,0.0,1.0,-0.1379496,-0.1379498,0.9807853,-0.1622116,-0.1083865,0.9807853,0.0,0.0,1.0,-0.1622116,-0.1083865,0.9807853,-0.1802399,-0.0746579,0.9807853,0.0,0.0,1.0,-0.1802399,-0.0746579,0.9807853,-0.1913417,-0.0380603,0.9807853,0.0,0.0,1.0,-0.1913417,-0.0380603,0.9807853,-0.1950903,-0.0000001,0.9807853,0.0,0.0,1.0,-0.1950903,-0.0000001,0.9807853,-0.1913417,0.0380602,0.9807853,0.0,0.0,1.0,-0.1913417,0.0380602,0.9807853,-0.18024,0.0746578,0.9807853,0.0,0.0,1.0,-0.18024,0.0746578,0.9807853,-0.1622117,0.1083864,0.9807853,0.0,0.0,1.0,-0.1622117,0.1083864,0.9807853,-0.1379497,0.1379497,0.9807853,0.0,0.0,1.0,-0.1379497,0.1379497,0.9807853,-0.1083864,0.1622117,0.9807853,0.0,0.0,1.0,-0.1083864,0.1622117,0.9807853,-0.0746578,0.18024,0.9807853,0.0,0.0,1.0,-0.0746578,0.18024,0.9807853,-0.0380602,0.1913417,0.9807853,0.0,0.0,1.0,-0.0380602,0.1913417,0.9807853,-0.0,0.1950903,0.9807853,0.0,0.0,1.0]},\\"normal\\":{\\"valueType\\":\\"Float32\\",\\"valuesPerElement\\":3,\\"values\\":[0.039,0.197,-0.98,-0.0,0.201,-0.98,0.0,-0.0,-1.0,0.077,0.186,-0.98,0.039,0.197,-0.98,0.0,-0.0,-1.0,0.112,0.167,-0.98,0.077,0.186,-0.98,0.0,-0.0,-1.0,0.142,0.142,-0.98,0.112,0.167,-0.98,0.0,-0.0,-1.0,0.167,0.112,-0.98,0.142,0.142,-0.98,0.0,-0.0,-1.0,0.186,0.077,-0.98,0.167,0.112,-0.98,0.0,-0.0,-1.0,0.197,0.039,-0.98,0.186,0.077,-0.98,0.0,-0.0,-1.0,0.201,0.0,-0.98,0.197,0.039,-0.98,0.0,-0.0,-1.0,0.197,-0.039,-0.98,0.201,0.0,-0.98,0.0,-0.0,-1.0,0.186,-0.077,-0.98,0.197,-0.039,-0.98,0.0,-0.0,-1.0,0.167,-0.112,-0.98,0.186,-0.077,-0.98,0.0,-0.0,-1.0,0.142,-0.142,-0.98,0.167,-0.112,-0.98,0.0,-0.0,-1.0,0.112,-0.167,-0.98,0.142,-0.142,-0.98,0.0,-0.0,-1.0,0.077,-0.186,-0.98,0.112,-0.167,-0.98,0.0,-0.0,-1.0,0.039,-0.197,-0.98,0.077,-0.186,-0.98,0.0,-0.0,-1.0,0.0,-0.201,-0.98,0.039,-0.197,-0.98,0.0,-0.0,-1.0,-0.039,-0.197,-0.98,0.0,-0.201,-0.98,0.0,-0.0,-1.0,-0.077,-0.186,-0.98,-0.039,-0.197,-0.98,0.0,-0.0,-1.0,-0.112,-0.167,-0.98,-0.077,-0.186,-0.98,0.0,-0.0,-1.0,-0.142,-0.142,-0.98,-0.112,-0.167,-0.98,0.0,-0.0,-1.0,-0.167,-0.112,-0.98,-0.142,-0.142,-0.98,0.0,-0.0,-1.0,-0.186,-0.077,-0.98,-0.167,-0.112,-0.98,0.0,-0.0,-1.0,-0.197,-0.039,-0.98,-0.186,-0.077,-0.98,0.0,-0.0,-1.0,-0.201,-0.0,-0.98,-0.197,-0.039,-0.98,0.0,-0.0,-1.0,-0.197,0.039,-0.98,-0.201,-0.0,-0.98,0.0,-0.0,-1.0,-0.186,0.077,-0.98,-0.197,0.039,-0.98,0.0,-0.0,-1.0,-0.167,0.112,-0.98,-0.186,0.077,-0.98,0.0,-0.0,-1.0,-0.142,0.142,-0.98,-0.167,0.112,-0.98,0.0,-0.0,-1.0,-0.112,0.167,-0.98,-0.142,0.142,-0.98,0.0,-0.0,-1.0,-0.077,0.186,-0.98,-0.112,0.167,-0.98,0.0,-0.0,-1.0,-0.039,0.197,-0.98,-0.077,0.186,-0.98,0.0,-0.0,-1.0,-0.0,0.201,-0.98,-0.039,0.197,-0.98,0.0,-0.0,-1.0,-0.0,0.388,-0.922,-0.0,0.201,-0.98,0.076,0.38,-0.922,-0.0,0.201,-0.98,0.039,0.197,-0.98,0.076,0.38,-0.922,0.076,0.38,-0.922,0.039,0.197,-0.98,0.148,0.358,-0.922,0.039,0.197,-0.98,0.077,0.186,-0.98,0.148,0.358,-0.922,0.148,0.358,-0.922,0.077,0.186,-0.98,0.215,0.323,-0.922,0.077,0.186,-0.98,0.112,0.167,-0.98,0.215,0.323,-0.922,0.215,0.323,-0.922,0.112,0.167,-0.98,0.274,0.274,-0.922,0.112,0.167,-0.98,0.142,0.142,-0.98,0.274,0.274,-0.922,0.274,0.274,-0.922,0.142,0.142,-0.98,0.323,0.215,-0.922,0.142,0.142,-0.98,0.167,0.112,-0.98,0.323,0.215,-0.922,0.323,0.215,-0.922,0.167,0.112,-0.98,0.358,0.148,-0.922,0.167,0.112,-0.98,0.186,0.077,-0.98,0.358,0.148,-0.922,0.358,0.148,-0.922,0.186,0.077,-0.98,0.38,0.076,-0.922,0.186,0.077,-0.98,0.197,0.039,-0.98,0.38,0.076,-0.922,0.38,0.076,-0.922,0.197,0.039,-0.98,0.388,0.0,-0.922,0.197,0.039,-0.98,0.201,0.0,-0.98,0.388,0.0,-0.922,0.388,0.0,-0.922,0.201,0.0,-0.98,0.38,-0.076,-0.922,0.201,0.0,-0.98,0.197,-0.039,-0.98,0.38,-0.076,-0.922,0.38,-0.076,-0.922,0.197,-0.039,-0.98,0.358,-0.148,-0.922,0.197,-0.039,-0.98,0.186,-0.077,-0.98,0.358,-0.148,-0.922,0.358,-0.148,-0.922,0.186,-0.077,-0.98,0.323,-0.215,-0.922,0.186,-0.077,-0.98,0.167,-0.112,-0.98,0.323,-0.215,-0.922,0.323,-0.215,-0.922,0.167,-0.112,-0.98,0.274,-0.274,-0.922,0.167,-0.112,-0.98,0.142,-0.142,-0.98,0.274,-0.274,-0.922,0.274,-0.274,-0.922,0.142,-0.142,-0.98,0.215,-0.323,-0.922,0.142,-0.142,-0.98,0.112,-0.167,-0.98,0.215,-0.323,-0.922,0.215,-0.323,-0.922,0.112,-0.167,-0.98,0.148,-0.358,-0.922,0.112,-0.167,-0.98,0.077,-0.186,-0.98,0.148,-0.358,-0.922,0.148,-0.358,-0.922,0.077,-0.186,-0.98,0.076,-0.38,-0.922,0.077,-0.186,-0.98,0.039,-0.197,-0.98,0.076,-0.38,-0.922,0.076,-0.38,-0.922,0.039,-0.197,-0.98,0.0,-0.388,-0.922,0.039,-0.197,-0.98,0.0,-0.201,-0.98,0.0,-0.388,-0.922,0.0,-0.388,-0.922,0.0,-0.201,-0.98,-0.076,-0.38,-0.922,0.0,-0.201,-0.98,-0.039,-0.197,-0.98,-0.076,-0.38,-0.922,-0.076,-0.38,-0.922,-0.039,-0.197,-0.98,-0.148,-0.358,-0.922,-0.039,-0.197,-0.98,-0.077,-0.186,-0.98,-0.148,-0.358,-0.922,-0.148,-0.358,-0.922,-0.077,-0.186,-0.98,-0.215,-0.323,-0.922,-0.077,-0.186,-0.98,-0.112,-0.167,-0.98,-0.215,-0.323,-0.922,-0.215,-0.323,-0.922,-0.112,-0.167,-0.98,-0.274,-0.274,-0.922,-0.112,-0.167,-0.98,-0.142,-0.142,-0.98,-0.274,-0.274,-0.922,-0.274,-0.274,-0.922,-0.142,-0.142,-0.98,-0.323,-0.215,-0.922,-0.142,-0.142,-0.98,-0.167,-0.112,-0.98,-0.323,-0.215,-0.922,-0.323,-0.215,-0.922,-0.167,-0.112,-0.98,-0.358,-0.148,-0.922,-0.167,-0.112,-0.98,-0.186,-0.077,-0.98,-0.358,-0.148,-0.922,-0.358,-0.148,-0.922,-0.186,-0.077,-0.98,-0.38,-0.076,-0.922,-0.186,-0.077,-0.98,-0.197,-0.039,-0.98,-0.38,-0.076,-0.922,-0.38,-0.076,-0.922,-0.197,-0.039,-0.98,-0.388,-0.0,-0.922,-0.197,-0.039,-0.98,-0.201,-0.0,-0.98,-0.388,-0.0,-0.922,-0.388,-0.0,-0.922,-0.201,-0.0,-0.98,-0.38,0.076,-0.922,-0.201,-0.0,-0.98,-0.197,0.039,-0.98,-0.38,0.076,-0.922,-0.38,0.076,-0.922,-0.197,0.039,-0.98,-0.358,0.148,-0.922,-0.197,0.039,-0.98,-0.186,0.077,-0.98,-0.358,0.148,-0.922,-0.358,0.148,-0.922,-0.186,0.077,-0.98,-0.323,0.215,-0.922,-0.186,0.077,-0.98,-0.167,0.112,-0.98,-0.323,0.215,-0.922,-0.323,0.215,-0.922,-0.167,0.112,-0.98,-0.274,0.274,-0.922,-0.167,0.112,-0.98,-0.142,0.142,-0.98,-0.274,0.274,-0.922,-0.274,0.274,-0.922,-0.142,0.142,-0.98,-0.215,0.323,-0.922,-0.142,0.142,-0.98,-0.112,0.167,-0.98,-0.215,0.323,-0.922,-0.215,0.323,-0.922,-0.112,0.167,-0.98,-0.148,0.358,-0.922,-0.112,0.167,-0.98,-0.077,0.186,-0.98,-0.148,0.358,-0.922,-0.148,0.358,-0.922,-0.077,0.186,-0.98,-0.076,0.38,-0.922,-0.077,0.186,-0.98,-0.039,0.197,-0.98,-0.076,0.38,-0.922,-0.076,0.38,-0.922,-0.039,0.197,-0.98,-0.0,0.388,-0.922,-0.039,0.197,-0.98,-0.0,0.201,-0.98,-0.0,0.388,-0.922,-0.0,0.56,-0.829,-0.0,0.388,-0.922,0.109,0.549,-0.829,-0.0,0.388,-0.922,0.076,0.38,-0.922,0.109,0.549,-0.829,0.109,0.549,-0.829,0.076,0.38,-0.922,0.214,0.517,-0.829,0.076,0.38,-0.922,0.148,0.358,-0.922,0.214,0.517,-0.829,0.214,0.517,-0.829,0.148,0.358,-0.922,0.311,0.465,-0.829,0.148,0.358,-0.922,0.215,0.323,-0.922,0.311,0.465,-0.829,0.311,0.465,-0.829,0.215,0.323,-0.922,0.396,0.396,-0.829,0.215,0.323,-0.922,0.274,0.274,-0.922,0.396,0.396,-0.829,0.396,0.396,-0.829,0.274,0.274,-0.922,0.465,0.311,-0.829,0.274,0.274,-0.922,0.323,0.215,-0.922,0.465,0.311,-0.829,0.465,0.311,-0.829,0.323,0.215,-0.922,0.517,0.214,-0.829,0.323,0.215,-0.922,0.358,0.148,-0.922,0.517,0.214,-0.829,0.517,0.214,-0.829,0.358,0.148,-0.922,0.549,0.109,-0.829,0.358,0.148,-0.922,0.38,0.076,-0.922,0.549,0.109,-0.829,0.549,0.109,-0.829,0.38,0.076,-0.922,0.56,0.0,-0.829,0.38,0.076,-0.922,0.388,0.0,-0.922,0.56,0.0,-0.829,0.56,0.0,-0.829,0.388,0.0,-0.922,0.549,-0.109,-0.829,0.388,0.0,-0.922,0.38,-0.076,-0.922,0.549,-0.109,-0.829,0.549,-0.109,-0.829,0.38,-0.076,-0.922,0.517,-0.214,-0.829,0.38,-0.076,-0.922,0.358,-0.148,-0.922,0.517,-0.214,-0.829,0.517,-0.214,-0.829,0.358,-0.148,-0.922,0.465,-0.311,-0.829,0.358,-0.148,-0.922,0.323,-0.215,-0.922,0.465,-0.311,-0.829,0.465,-0.311,-0.829,0.323,-0.215,-0.922,0.396,-0.396,-0.829,0.323,-0.215,-0.922,0.274,-0.274,-0.922,0.396,-0.396,-0.829,0.396,-0.396,-0.829,0.274,-0.274,-0.922,0.311,-0.465,-0.829,0.274,-0.274,-0.922,0.215,-0.323,-0.922,0.311,-0.465,-0.829,0.311,-0.465,-0.829,0.215,-0.323,-0.922,0.214,-0.517,-0.829,0.215,-0.323,-0.922,0.148,-0.358,-0.922,0.214,-0.517,-0.829,0.214,-0.517,-0.829,0.148,-0.358,-0.922,0.109,-0.549,-0.829,0.148,-0.358,-0.922,0.076,-0.38,-0.922,0.109,-0.549,-0.829,0.109,-0.549,-0.829,0.076,-0.38,-0.922,0.0,-0.56,-0.829,0.076,-0.38,-0.922,0.0,-0.388,-0.922,0.0,-0.56,-0.829,0.0,-0.56,-0.829,0.0,-0.388,-0.922,-0.109,-0.549,-0.829,0.0,-0.388,-0.922,-0.076,-0.38,-0.922,-0.109,-0.549,-0.829,-0.109,-0.549,-0.829,-0.076,-0.38,-0.922,-0.214,-0.517,-0.829,-0.076,-0.38,-0.922,-0.148,-0.358,-0.922,-0.214,-0.517,-0.829,-0.214,-0.517,-0.829,-0.148,-0.358,-0.922,-0.311,-0.465,-0.829,-0.148,-0.358,-0.922,-0.215,-0.323,-0.922,-0.311,-0.465,-0.829,-0.311,-0.465,-0.829,-0.215,-0.323,-0.922,-0.396,-0.396,-0.829,-0.215,-0.323,-0.922,-0.274,-0.274,-0.922,-0.396,-0.396,-0.829,-0.396,-0.396,-0.829,-0.274,-0.274,-0.922,-0.465,-0.311,-0.829,-0.274,-0.274,-0.922,-0.323,-0.215,-0.922,-0.465,-0.311,-0.829,-0.465,-0.311,-0.829,-0.323,-0.215,-0.922,-0.517,-0.214,-0.829,-0.323,-0.215,-0.922,-0.358,-0.148,-0.922,-0.517,-0.214,-0.829,-0.517,-0.214,-0.829,-0.358,-0.148,-0.922,-0.549,-0.109,-0.829,-0.358,-0.148,-0.922,-0.38,-0.076,-0.922,-0.549,-0.109,-0.829,-0.549,-0.109,-0.829,-0.38,-0.076,-0.922,-0.56,-0.0,-0.829,-0.38,-0.076,-0.922,-0.388,-0.0,-0.922,-0.56,-0.0,-0.829,-0.56,-0.0,-0.829,-0.388,-0.0,-0.922,-0.549,0.109,-0.829,-0.388,-0.0,-0.922,-0.38,0.076,-0.922,-0.549,0.109,-0.829,-0.549,0.109,-0.829,-0.38,0.076,-0.922,-0.517,0.214,-0.829,-0.38,0.076,-0.922,-0.358,0.148,-0.922,-0.517,0.214,-0.829,-0.517,0.214,-0.829,-0.358,0.148,-0.922,-0.465,0.311,-0.829,-0.358,0.148,-0.922,-0.323,0.215,-0.922,-0.465,0.311,-0.829,-0.465,0.311,-0.829,-0.323,0.215,-0.922,-0.396,0.396,-0.829,-0.323,0.215,-0.922,-0.274,0.274,-0.922,-0.396,0.396,-0.829,-0.396,0.396,-0.829,-0.274,0.274,-0.922,-0.311,0.465,-0.829,-0.274,0.274,-0.922,-0.215,0.323,-0.922,-0.311,0.465,-0.829,-0.311,0.465,-0.829,-0.215,0.323,-0.922,-0.214,0.517,-0.829,-0.215,0.323,-0.922,-0.148,0.358,-0.922,-0.214,0.517,-0.829,-0.214,0.517,-0.829,-0.148,0.358,-0.922,-0.109,0.549,-0.829,-0.148,0.358,-0.922,-0.076,0.38,-0.922,-0.109,0.549,-0.829,-0.109,0.549,-0.829,-0.076,0.38,-0.922,-0.0,0.56,-0.829,-0.076,0.38,-0.922,-0.0,0.388,-0.922,-0.0,0.56,-0.829,-0.0,0.71,-0.704,-0.0,0.56,-0.829,0.139,0.696,-0.704,-0.0,0.56,-0.829,0.109,0.549,-0.829,0.139,0.696,-0.704,0.139,0.696,-0.704,0.109,0.549,-0.829,0.272,0.656,-0.704,0.109,0.549,-0.829,0.214,0.517,-0.829,0.272,0.656,-0.704,0.272,0.656,-0.704,0.214,0.517,-0.829,0.395,0.59,-0.704,0.214,0.517,-0.829,0.311,0.465,-0.829,0.395,0.59,-0.704,0.395,0.59,-0.704,0.311,0.465,-0.829,0.502,0.502,-0.704,0.311,0.465,-0.829,0.396,0.396,-0.829,0.502,0.502,-0.704,0.502,0.502,-0.704,0.396,0.396,-0.829,0.59,0.395,-0.704,0.396,0.396,-0.829,0.465,0.311,-0.829,0.59,0.395,-0.704,0.59,0.395,-0.704,0.465,0.311,-0.829,0.656,0.272,-0.704,0.465,0.311,-0.829,0.517,0.214,-0.829,0.656,0.272,-0.704,0.656,0.272,-0.704,0.517,0.214,-0.829,0.696,0.139,-0.704,0.517,0.214,-0.829,0.549,0.109,-0.829,0.696,0.139,-0.704,0.696,0.139,-0.704,0.549,0.109,-0.829,0.71,0.0,-0.704,0.549,0.109,-0.829,0.56,0.0,-0.829,0.71,0.0,-0.704,0.71,0.0,-0.704,0.56,0.0,-0.829,0.696,-0.139,-0.704,0.56,0.0,-0.829,0.549,-0.109,-0.829,0.696,-0.139,-0.704,0.696,-0.139,-0.704,0.549,-0.109,-0.829,0.656,-0.272,-0.704,0.549,-0.109,-0.829,0.517,-0.214,-0.829,0.656,-0.272,-0.704,0.656,-0.272,-0.704,0.517,-0.214,-0.829,0.59,-0.395,-0.704,0.517,-0.214,-0.829,0.465,-0.311,-0.829,0.59,-0.395,-0.704,0.59,-0.395,-0.704,0.465,-0.311,-0.829,0.502,-0.502,-0.704,0.465,-0.311,-0.829,0.396,-0.396,-0.829,0.502,-0.502,-0.704,0.502,-0.502,-0.704,0.396,-0.396,-0.829,0.395,-0.59,-0.704,0.396,-0.396,-0.829,0.311,-0.465,-0.829,0.395,-0.59,-0.704,0.395,-0.59,-0.704,0.311,-0.465,-0.829,0.272,-0.656,-0.704,0.311,-0.465,-0.829,0.214,-0.517,-0.829,0.272,-0.656,-0.704,0.272,-0.656,-0.704,0.214,-0.517,-0.829,0.139,-0.696,-0.704,0.214,-0.517,-0.829,0.109,-0.549,-0.829,0.139,-0.696,-0.704,0.139,-0.696,-0.704,0.109,-0.549,-0.829,0.0,-0.71,-0.704,0.109,-0.549,-0.829,0.0,-0.56,-0.829,0.0,-0.71,-0.704,0.0,-0.71,-0.704,0.0,-0.56,-0.829,-0.139,-0.696,-0.704,0.0,-0.56,-0.829,-0.109,-0.549,-0.829,-0.139,-0.696,-0.704,-0.139,-0.696,-0.704,-0.109,-0.549,-0.829,-0.272,-0.656,-0.704,-0.109,-0.549,-0.829,-0.214,-0.517,-0.829,-0.272,-0.656,-0.704,-0.272,-0.656,-0.704,-0.214,-0.517,-0.829,-0.395,-0.59,-0.704,-0.214,-0.517,-0.829,-0.311,-0.465,-0.829,-0.395,-0.59,-0.704,-0.395,-0.59,-0.704,-0.311,-0.465,-0.829,-0.502,-0.502,-0.704,-0.311,-0.465,-0.829,-0.396,-0.396,-0.829,-0.502,-0.502,-0.704,-0.502,-0.502,-0.704,-0.396,-0.396,-0.829,-0.59,-0.395,-0.704,-0.396,-0.396,-0.829,-0.465,-0.311,-0.829,-0.59,-0.395,-0.704,-0.59,-0.395,-0.704,-0.465,-0.311,-0.829,-0.656,-0.272,-0.704,-0.465,-0.311,-0.829,-0.517,-0.214,-0.829,-0.656,-0.272,-0.704,-0.656,-0.272,-0.704,-0.517,-0.214,-0.829,-0.696,-0.139,-0.704,-0.517,-0.214,-0.829,-0.549,-0.109,-0.829,-0.696,-0.139,-0.704,-0.696,-0.139,-0.704,-0.549,-0.109,-0.829,-0.71,-0.0,-0.704,-0.549,-0.109,-0.829,-0.56,-0.0,-0.829,-0.71,-0.0,-0.704,-0.71,-0.0,-0.704,-0.56,-0.0,-0.829,-0.696,0.139,-0.704,-0.56,-0.0,-0.829,-0.549,0.109,-0.829,-0.696,0.139,-0.704,-0.696,0.139,-0.704,-0.549,0.109,-0.829,-0.656,0.272,-0.704,-0.549,0.109,-0.829,-0.517,0.214,-0.829,-0.656,0.272,-0.704,-0.656,0.272,-0.704,-0.517,0.214,-0.829,-0.59,0.395,-0.704,-0.517,0.214,-0.829,-0.465,0.311,-0.829,-0.59,0.395,-0.704,-0.59,0.395,-0.704,-0.465,0.311,-0.829,-0.502,0.502,-0.704,-0.465,0.311,-0.829,-0.396,0.396,-0.829,-0.502,0.502,-0.704,-0.502,0.502,-0.704,-0.396,0.396,-0.829,-0.395,0.59,-0.704,-0.396,0.396,-0.829,-0.311,0.465,-0.829,-0.395,0.59,-0.704,-0.395,0.59,-0.704,-0.311,0.465,-0.829,-0.272,0.656,-0.704,-0.311,0.465,-0.829,-0.214,0.517,-0.829,-0.272,0.656,-0.704,-0.272,0.656,-0.704,-0.214,0.517,-0.829,-0.139,0.696,-0.704,-0.214,0.517,-0.829,-0.109,0.549,-0.829,-0.139,0.696,-0.704,-0.139,0.696,-0.704,-0.109,0.549,-0.829,-0.0,0.71,-0.704,-0.109,0.549,-0.829,-0.0,0.56,-0.829,-0.0,0.71,-0.704,-0.0,0.833,-0.553,-0.0,0.71,-0.704,0.163,0.817,-0.553,-0.0,0.71,-0.704,0.139,0.696,-0.704,0.163,0.817,-0.553,0.163,0.817,-0.553,0.139,0.696,-0.704,0.319,0.77,-0.553,0.139,0.696,-0.704,0.272,0.656,-0.704,0.319,0.77,-0.553,0.319,0.77,-0.553,0.272,0.656,-0.704,0.463,0.693,-0.553,0.272,0.656,-0.704,0.395,0.59,-0.704,0.463,0.693,-0.553,0.463,0.693,-0.553,0.395,0.59,-0.704,0.589,0.589,-0.553,0.395,0.59,-0.704,0.502,0.502,-0.704,0.589,0.589,-0.553,0.589,0.589,-0.553,0.502,0.502,-0.704,0.693,0.463,-0.553,0.502,0.502,-0.704,0.59,0.395,-0.704,0.693,0.463,-0.553,0.693,0.463,-0.553,0.59,0.395,-0.704,0.77,0.319,-0.553,0.59,0.395,-0.704,0.656,0.272,-0.704,0.77,0.319,-0.553,0.77,0.319,-0.553,0.656,0.272,-0.704,0.817,0.163,-0.553,0.656,0.272,-0.704,0.696,0.139,-0.704,0.817,0.163,-0.553,0.817,0.163,-0.553,0.696,0.139,-0.704,0.833,0.0,-0.553,0.696,0.139,-0.704,0.71,0.0,-0.704,0.833,0.0,-0.553,0.833,0.0,-0.553,0.71,0.0,-0.704,0.817,-0.163,-0.553,0.71,0.0,-0.704,0.696,-0.139,-0.704,0.817,-0.163,-0.553,0.817,-0.163,-0.553,0.696,-0.139,-0.704,0.77,-0.319,-0.553,0.696,-0.139,-0.704,0.656,-0.272,-0.704,0.77,-0.319,-0.553,0.77,-0.319,-0.553,0.656,-0.272,-0.704,0.693,-0.463,-0.553,0.656,-0.272,-0.704,0.59,-0.395,-0.704,0.693,-0.463,-0.553,0.693,-0.463,-0.553,0.59,-0.395,-0.704,0.589,-0.589,-0.553,0.59,-0.395,-0.704,0.502,-0.502,-0.704,0.589,-0.589,-0.553,0.589,-0.589,-0.553,0.502,-0.502,-0.704,0.463,-0.693,-0.553,0.502,-0.502,-0.704,0.395,-0.59,-0.704,0.463,-0.693,-0.553,0.463,-0.693,-0.553,0.395,-0.59,-0.704,0.319,-0.77,-0.553,0.395,-0.59,-0.704,0.272,-0.656,-0.704,0.319,-0.77,-0.553,0.319,-0.77,-0.553,0.272,-0.656,-0.704,0.163,-0.817,-0.553,0.272,-0.656,-0.704,0.139,-0.696,-0.704,0.163,-0.817,-0.553,0.163,-0.817,-0.553,0.139,-0.696,-0.704,0.0,-0.833,-0.553,0.139,-0.696,-0.704,0.0,-0.71,-0.704,0.0,-0.833,-0.553,0.0,-0.833,-0.553,0.0,-0.71,-0.704,-0.163,-0.817,-0.553,0.0,-0.71,-0.704,-0.139,-0.696,-0.704,-0.163,-0.817,-0.553,-0.163,-0.817,-0.553,-0.139,-0.696,-0.704,-0.319,-0.77,-0.553,-0.139,-0.696,-0.704,-0.272,-0.656,-0.704,-0.319,-0.77,-0.553,-0.319,-0.77,-0.553,-0.272,-0.656,-0.704,-0.463,-0.693,-0.553,-0.272,-0.656,-0.704,-0.395,-0.59,-0.704,-0.463,-0.693,-0.553,-0.463,-0.693,-0.553,-0.395,-0.59,-0.704,-0.589,-0.589,-0.553,-0.395,-0.59,-0.704,-0.502,-0.502,-0.704,-0.589,-0.589,-0.553,-0.589,-0.589,-0.553,-0.502,-0.502,-0.704,-0.693,-0.463,-0.553,-0.502,-0.502,-0.704,-0.59,-0.395,-0.704,-0.693,-0.463,-0.553,-0.693,-0.463,-0.553,-0.59,-0.395,-0.704,-0.77,-0.319,-0.553,-0.59,-0.395,-0.704,-0.656,-0.272,-0.704,-0.77,-0.319,-0.553,-0.77,-0.319,-0.553,-0.656,-0.272,-0.704,-0.817,-0.163,-0.553,-0.656,-0.272,-0.704,-0.696,-0.139,-0.704,-0.817,-0.163,-0.553,-0.817,-0.163,-0.553,-0.696,-0.139,-0.704,-0.833,-0.0,-0.553,-0.696,-0.139,-0.704,-0.71,-0.0,-0.704,-0.833,-0.0,-0.553,-0.833,-0.0,-0.553,-0.71,-0.0,-0.704,-0.817,0.163,-0.553,-0.71,-0.0,-0.704,-0.696,0.139,-0.704,-0.817,0.163,-0.553,-0.817,0.163,-0.553,-0.696,0.139,-0.704,-0.77,0.319,-0.553,-0.696,0.139,-0.704,-0.656,0.272,-0.704,-0.77,0.319,-0.553,-0.77,0.319,-0.553,-0.656,0.272,-0.704,-0.693,0.463,-0.553,-0.656,0.272,-0.704,-0.59,0.395,-0.704,-0.693,0.463,-0.553,-0.693,0.463,-0.553,-0.59,0.395,-0.704,-0.589,0.589,-0.553,-0.59,0.395,-0.704,-0.502,0.502,-0.704,-0.589,0.589,-0.553,-0.589,0.589,-0.553,-0.502,0.502,-0.704,-0.463,0.693,-0.553,-0.502,0.502,-0.704,-0.395,0.59,-0.704,-0.463,0.693,-0.553,-0.463,0.693,-0.553,-0.395,0.59,-0.704,-0.319,0.77,-0.553,-0.395,0.59,-0.704,-0.272,0.656,-0.704,-0.319,0.77,-0.553,-0.319,0.77,-0.553,-0.272,0.656,-0.704,-0.163,0.817,-0.553,-0.272,0.656,-0.704,-0.139,0.696,-0.704,-0.163,0.817,-0.553,-0.163,0.817,-0.553,-0.139,0.696,-0.704,-0.0,0.833,-0.553,-0.139,0.696,-0.704,-0.0,0.71,-0.704,-0.0,0.833,-0.553,-0.0,0.925,-0.381,-0.0,0.833,-0.553,0.18,0.907,-0.381,-0.0,0.833,-0.553,0.163,0.817,-0.553,0.18,0.907,-0.381,0.18,0.907,-0.381,0.163,0.817,-0.553,0.354,0.854,-0.381,0.163,0.817,-0.553,0.319,0.77,-0.553,0.354,0.854,-0.381,0.354,0.854,-0.381,0.319,0.77,-0.553,0.514,0.769,-0.381,0.319,0.77,-0.553,0.463,0.693,-0.553,0.514,0.769,-0.381,0.514,0.769,-0.381,0.463,0.693,-0.553,0.654,0.654,-0.381,0.463,0.693,-0.553,0.589,0.589,-0.553,0.654,0.654,-0.381,0.654,0.654,-0.381,0.589,0.589,-0.553,0.769,0.514,-0.381,0.589,0.589,-0.553,0.693,0.463,-0.553,0.769,0.514,-0.381,0.769,0.514,-0.381,0.693,0.463,-0.553,0.854,0.354,-0.381,0.693,0.463,-0.553,0.77,0.319,-0.553,0.854,0.354,-0.381,0.854,0.354,-0.381,0.77,0.319,-0.553,0.907,0.18,-0.381,0.77,0.319,-0.553,0.817,0.163,-0.553,0.907,0.18,-0.381,0.907,0.18,-0.381,0.817,0.163,-0.553,0.925,0.0,-0.381,0.817,0.163,-0.553,0.833,0.0,-0.553,0.925,0.0,-0.381,0.925,0.0,-0.381,0.833,0.0,-0.553,0.907,-0.18,-0.381,0.833,0.0,-0.553,0.817,-0.163,-0.553,0.907,-0.18,-0.381,0.907,-0.18,-0.381,0.817,-0.163,-0.553,0.854,-0.354,-0.381,0.817,-0.163,-0.553,0.77,-0.319,-0.553,0.854,-0.354,-0.381,0.854,-0.354,-0.381,0.77,-0.319,-0.553,0.769,-0.514,-0.381,0.77,-0.319,-0.553,0.693,-0.463,-0.553,0.769,-0.514,-0.381,0.769,-0.514,-0.381,0.693,-0.463,-0.553,0.654,-0.654,-0.381,0.693,-0.463,-0.553,0.589,-0.589,-0.553,0.654,-0.654,-0.381,0.654,-0.654,-0.381,0.589,-0.589,-0.553,0.514,-0.769,-0.381,0.589,-0.589,-0.553,0.463,-0.693,-0.553,0.514,-0.769,-0.381,0.514,-0.769,-0.381,0.463,-0.693,-0.553,0.354,-0.854,-0.381,0.463,-0.693,-0.553,0.319,-0.77,-0.553,0.354,-0.854,-0.381,0.354,-0.854,-0.381,0.319,-0.77,-0.553,0.18,-0.907,-0.381,0.319,-0.77,-0.553,0.163,-0.817,-0.553,0.18,-0.907,-0.381,0.18,-0.907,-0.381,0.163,-0.817,-0.553,0.0,-0.925,-0.381,0.163,-0.817,-0.553,0.0,-0.833,-0.553,0.0,-0.925,-0.381,0.0,-0.925,-0.381,0.0,-0.833,-0.553,-0.18,-0.907,-0.381,0.0,-0.833,-0.553,-0.163,-0.817,-0.553,-0.18,-0.907,-0.381,-0.18,-0.907,-0.381,-0.163,-0.817,-0.553,-0.354,-0.854,-0.381,-0.163,-0.817,-0.553,-0.319,-0.77,-0.553,-0.354,-0.854,-0.381,-0.354,-0.854,-0.381,-0.319,-0.77,-0.553,-0.514,-0.769,-0.381,-0.319,-0.77,-0.553,-0.463,-0.693,-0.553,-0.514,-0.769,-0.381,-0.514,-0.769,-0.381,-0.463,-0.693,-0.553,-0.654,-0.654,-0.381,-0.463,-0.693,-0.553,-0.589,-0.589,-0.553,-0.654,-0.654,-0.381,-0.654,-0.654,-0.381,-0.589,-0.589,-0.553,-0.769,-0.514,-0.381,-0.589,-0.589,-0.553,-0.693,-0.463,-0.553,-0.769,-0.514,-0.381,-0.769,-0.514,-0.381,-0.693,-0.463,-0.553,-0.854,-0.354,-0.381,-0.693,-0.463,-0.553,-0.77,-0.319,-0.553,-0.854,-0.354,-0.381,-0.854,-0.354,-0.381,-0.77,-0.319,-0.553,-0.907,-0.18,-0.381,-0.77,-0.319,-0.553,-0.817,-0.163,-0.553,-0.907,-0.18,-0.381,-0.907,-0.18,-0.381,-0.817,-0.163,-0.553,-0.925,-0.0,-0.381,-0.817,-0.163,-0.553,-0.833,-0.0,-0.553,-0.925,-0.0,-0.381,-0.925,-0.0,-0.381,-0.833,-0.0,-0.553,-0.907,0.18,-0.381,-0.833,-0.0,-0.553,-0.817,0.163,-0.553,-0.907,0.18,-0.381,-0.907,0.18,-0.381,-0.817,0.163,-0.553,-0.854,0.354,-0.381,-0.817,0.163,-0.553,-0.77,0.319,-0.553,-0.854,0.354,-0.381,-0.854,0.354,-0.381,-0.77,0.319,-0.553,-0.769,0.514,-0.381,-0.77,0.319,-0.553,-0.693,0.463,-0.553,-0.769,0.514,-0.381,-0.769,0.514,-0.381,-0.693,0.463,-0.553,-0.654,0.654,-0.381,-0.693,0.463,-0.553,-0.589,0.589,-0.553,-0.654,0.654,-0.381,-0.654,0.654,-0.381,-0.589,0.589,-0.553,-0.514,0.769,-0.381,-0.589,0.589,-0.553,-0.463,0.693,-0.553,-0.514,0.769,-0.381,-0.514,0.769,-0.381,-0.463,0.693,-0.553,-0.354,0.854,-0.381,-0.463,0.693,-0.553,-0.319,0.77,-0.553,-0.354,0.854,-0.381,-0.354,0.854,-0.381,-0.319,0.77,-0.553,-0.18,0.907,-0.381,-0.319,0.77,-0.553,-0.163,0.817,-0.553,-0.18,0.907,-0.381,-0.18,0.907,-0.381,-0.163,0.817,-0.553,-0.0,0.925,-0.381,-0.163,0.817,-0.553,-0.0,0.833,-0.553,-0.0,0.925,-0.381,-0.0,0.981,-0.194,-0.0,0.925,-0.381,0.191,0.962,-0.194,-0.0,0.925,-0.381,0.18,0.907,-0.381,0.191,0.962,-0.194,0.191,0.962,-0.194,0.18,0.907,-0.381,0.375,0.906,-0.194,0.18,0.907,-0.381,0.354,0.854,-0.381,0.375,0.906,-0.194,0.375,0.906,-0.194,0.354,0.854,-0.381,0.545,0.816,-0.194,0.354,0.854,-0.381,0.514,0.769,-0.381,0.545,0.816,-0.194,0.545,0.816,-0.194,0.514,0.769,-0.381,0.694,0.694,-0.194,0.514,0.769,-0.381,0.654,0.654,-0.381,0.694,0.694,-0.194,0.694,0.694,-0.194,0.654,0.654,-0.381,0.816,0.545,-0.194,0.654,0.654,-0.381,0.769,0.514,-0.381,0.816,0.545,-0.194,0.816,0.545,-0.194,0.769,0.514,-0.381,0.906,0.375,-0.194,0.769,0.514,-0.381,0.854,0.354,-0.381,0.906,0.375,-0.194,0.906,0.375,-0.194,0.854,0.354,-0.381,0.962,0.191,-0.194,0.854,0.354,-0.381,0.907,0.18,-0.381,0.962,0.191,-0.194,0.962,0.191,-0.194,0.907,0.18,-0.381,0.981,0.0,-0.194,0.907,0.18,-0.381,0.925,0.0,-0.381,0.981,0.0,-0.194,0.981,0.0,-0.194,0.925,0.0,-0.381,0.962,-0.191,-0.194,0.925,0.0,-0.381,0.907,-0.18,-0.381,0.962,-0.191,-0.194,0.962,-0.191,-0.194,0.907,-0.18,-0.381,0.906,-0.375,-0.194,0.907,-0.18,-0.381,0.854,-0.354,-0.381,0.906,-0.375,-0.194,0.906,-0.375,-0.194,0.854,-0.354,-0.381,0.816,-0.545,-0.194,0.854,-0.354,-0.381,0.769,-0.514,-0.381,0.816,-0.545,-0.194,0.816,-0.545,-0.194,0.769,-0.514,-0.381,0.694,-0.694,-0.194,0.769,-0.514,-0.381,0.654,-0.654,-0.381,0.694,-0.694,-0.194,0.694,-0.694,-0.194,0.654,-0.654,-0.381,0.545,-0.816,-0.194,0.654,-0.654,-0.381,0.514,-0.769,-0.381,0.545,-0.816,-0.194,0.545,-0.816,-0.194,0.514,-0.769,-0.381,0.375,-0.906,-0.194,0.514,-0.769,-0.381,0.354,-0.854,-0.381,0.375,-0.906,-0.194,0.375,-0.906,-0.194,0.354,-0.854,-0.381,0.191,-0.962,-0.194,0.354,-0.854,-0.381,0.18,-0.907,-0.381,0.191,-0.962,-0.194,0.191,-0.962,-0.194,0.18,-0.907,-0.381,0.0,-0.981,-0.194,0.18,-0.907,-0.381,0.0,-0.925,-0.381,0.0,-0.981,-0.194,0.0,-0.981,-0.194,0.0,-0.925,-0.381,-0.191,-0.962,-0.194,0.0,-0.925,-0.381,-0.18,-0.907,-0.381,-0.191,-0.962,-0.194,-0.191,-0.962,-0.194,-0.18,-0.907,-0.381,-0.375,-0.906,-0.194,-0.18,-0.907,-0.381,-0.354,-0.854,-0.381,-0.375,-0.906,-0.194,-0.375,-0.906,-0.194,-0.354,-0.854,-0.381,-0.545,-0.816,-0.194,-0.354,-0.854,-0.381,-0.514,-0.769,-0.381,-0.545,-0.816,-0.194,-0.545,-0.816,-0.194,-0.514,-0.769,-0.381,-0.694,-0.694,-0.194,-0.514,-0.769,-0.381,-0.654,-0.654,-0.381,-0.694,-0.694,-0.194,-0.694,-0.694,-0.194,-0.654,-0.654,-0.381,-0.816,-0.545,-0.194,-0.654,-0.654,-0.381,-0.769,-0.514,-0.381,-0.816,-0.545,-0.194,-0.816,-0.545,-0.194,-0.769,-0.514,-0.381,-0.906,-0.375,-0.194,-0.769,-0.514,-0.381,-0.854,-0.354,-0.381,-0.906,-0.375,-0.194,-0.906,-0.375,-0.194,-0.854,-0.354,-0.381,-0.962,-0.191,-0.194,-0.854,-0.354,-0.381,-0.907,-0.18,-0.381,-0.962,-0.191,-0.194,-0.962,-0.191,-0.194,-0.907,-0.18,-0.381,-0.981,-0.0,-0.194,-0.907,-0.18,-0.381,-0.925,-0.0,-0.381,-0.981,-0.0,-0.194,-0.981,-0.0,-0.194,-0.925,-0.0,-0.381,-0.962,0.191,-0.194,-0.925,-0.0,-0.381,-0.907,0.18,-0.381,-0.962,0.191,-0.194,-0.962,0.191,-0.194,-0.907,0.18,-0.381,-0.906,0.375,-0.194,-0.907,0.18,-0.381,-0.854,0.354,-0.381,-0.906,0.375,-0.194,-0.906,0.375,-0.194,-0.854,0.354,-0.381,-0.816,0.545,-0.194,-0.854,0.354,-0.381,-0.769,0.514,-0.381,-0.816,0.545,-0.194,-0.816,0.545,-0.194,-0.769,0.514,-0.381,-0.694,0.694,-0.194,-0.769,0.514,-0.381,-0.654,0.654,-0.381,-0.694,0.694,-0.194,-0.694,0.694,-0.194,-0.654,0.654,-0.381,-0.545,0.816,-0.194,-0.654,0.654,-0.381,-0.514,0.769,-0.381,-0.545,0.816,-0.194,-0.545,0.816,-0.194,-0.514,0.769,-0.381,-0.375,0.906,-0.194,-0.514,0.769,-0.381,-0.354,0.854,-0.381,-0.375,0.906,-0.194,-0.375,0.906,-0.194,-0.354,0.854,-0.381,-0.191,0.962,-0.194,-0.354,0.854,-0.381,-0.18,0.907,-0.381,-0.191,0.962,-0.194,-0.191,0.962,-0.194,-0.18,0.907,-0.381,-0.0,0.981,-0.194,-0.18,0.907,-0.381,-0.0,0.925,-0.381,-0.0,0.981,-0.194,-0.0,1.0,0.0,-0.0,0.981,-0.194,0.195,0.981,0.0,-0.0,0.981,-0.194,0.191,0.962,-0.194,0.195,0.981,0.0,0.195,0.981,0.0,0.191,0.962,-0.194,0.383,0.924,0.0,0.191,0.962,-0.194,0.375,0.906,-0.194,0.383,0.924,0.0,0.383,0.924,0.0,0.375,0.906,-0.194,0.556,0.831,0.0,0.375,0.906,-0.194,0.545,0.816,-0.194,0.556,0.831,0.0,0.556,0.831,0.0,0.545,0.816,-0.194,0.707,0.707,0.0,0.545,0.816,-0.194,0.694,0.694,-0.194,0.707,0.707,0.0,0.707,0.707,0.0,0.694,0.694,-0.194,0.831,0.556,0.0,0.694,0.694,-0.194,0.816,0.545,-0.194,0.831,0.556,0.0,0.831,0.556,0.0,0.816,0.545,-0.194,0.924,0.383,0.0,0.816,0.545,-0.194,0.906,0.375,-0.194,0.924,0.383,0.0,0.924,0.383,0.0,0.906,0.375,-0.194,0.981,0.195,0.0,0.906,0.375,-0.194,0.962,0.191,-0.194,0.981,0.195,0.0,0.981,0.195,0.0,0.962,0.191,-0.194,1.0,0.0,0.0,0.962,0.191,-0.194,0.981,0.0,-0.194,1.0,0.0,0.0,1.0,0.0,0.0,0.981,0.0,-0.194,0.981,-0.195,0.0,0.981,0.0,-0.194,0.962,-0.191,-0.194,0.981,-0.195,0.0,0.981,-0.195,0.0,0.962,-0.191,-0.194,0.924,-0.383,0.0,0.962,-0.191,-0.194,0.906,-0.375,-0.194,0.924,-0.383,0.0,0.924,-0.383,0.0,0.906,-0.375,-0.194,0.831,-0.556,0.0,0.906,-0.375,-0.194,0.816,-0.545,-0.194,0.831,-0.556,0.0,0.831,-0.556,0.0,0.816,-0.545,-0.194,0.707,-0.707,0.0,0.816,-0.545,-0.194,0.694,-0.694,-0.194,0.707,-0.707,0.0,0.707,-0.707,0.0,0.694,-0.694,-0.194,0.556,-0.831,0.0,0.694,-0.694,-0.194,0.545,-0.816,-0.194,0.556,-0.831,0.0,0.556,-0.831,0.0,0.545,-0.816,-0.194,0.383,-0.924,0.0,0.545,-0.816,-0.194,0.375,-0.906,-0.194,0.383,-0.924,0.0,0.383,-0.924,0.0,0.375,-0.906,-0.194,0.195,-0.981,0.0,0.375,-0.906,-0.194,0.191,-0.962,-0.194,0.195,-0.981,0.0,0.195,-0.981,0.0,0.191,-0.962,-0.194,0.0,-1.0,0.0,0.191,-0.962,-0.194,0.0,-0.981,-0.194,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-0.981,-0.194,-0.195,-0.981,0.0,0.0,-0.981,-0.194,-0.191,-0.962,-0.194,-0.195,-0.981,0.0,-0.195,-0.981,0.0,-0.191,-0.962,-0.194,-0.383,-0.924,0.0,-0.191,-0.962,-0.194,-0.375,-0.906,-0.194,-0.383,-0.924,0.0,-0.383,-0.924,0.0,-0.375,-0.906,-0.194,-0.556,-0.831,0.0,-0.375,-0.906,-0.194,-0.545,-0.816,-0.194,-0.556,-0.831,0.0,-0.556,-0.831,0.0,-0.545,-0.816,-0.194,-0.707,-0.707,0.0,-0.545,-0.816,-0.194,-0.694,-0.694,-0.194,-0.707,-0.707,0.0,-0.707,-0.707,0.0,-0.694,-0.694,-0.194,-0.831,-0.556,0.0,-0.694,-0.694,-0.194,-0.816,-0.545,-0.194,-0.831,-0.556,0.0,-0.831,-0.556,0.0,-0.816,-0.545,-0.194,-0.924,-0.383,0.0,-0.816,-0.545,-0.194,-0.906,-0.375,-0.194,-0.924,-0.383,0.0,-0.924,-0.383,0.0,-0.906,-0.375,-0.194,-0.981,-0.195,0.0,-0.906,-0.375,-0.194,-0.962,-0.191,-0.194,-0.981,-0.195,0.0,-0.981,-0.195,0.0,-0.962,-0.191,-0.194,-1.0,-0.0,0.0,-0.962,-0.191,-0.194,-0.981,-0.0,-0.194,-1.0,-0.0,0.0,-1.0,-0.0,0.0,-0.981,-0.0,-0.194,-0.981,0.195,0.0,-0.981,-0.0,-0.194,-0.962,0.191,-0.194,-0.981,0.195,0.0,-0.981,0.195,0.0,-0.962,0.191,-0.194,-0.924,0.383,0.0,-0.962,0.191,-0.194,-0.906,0.375,-0.194,-0.924,0.383,0.0,-0.924,0.383,0.0,-0.906,0.375,-0.194,-0.831,0.556,0.0,-0.906,0.375,-0.194,-0.816,0.545,-0.194,-0.831,0.556,0.0,-0.831,0.556,0.0,-0.816,0.545,-0.194,-0.707,0.707,0.0,-0.816,0.545,-0.194,-0.694,0.694,-0.194,-0.707,0.707,0.0,-0.707,0.707,0.0,-0.694,0.694,-0.194,-0.556,0.831,0.0,-0.694,0.694,-0.194,-0.545,0.816,-0.194,-0.556,0.831,0.0,-0.556,0.831,0.0,-0.545,0.816,-0.194,-0.383,0.924,0.0,-0.545,0.816,-0.194,-0.375,0.906,-0.194,-0.383,0.924,0.0,-0.383,0.924,0.0,-0.375,0.906,-0.194,-0.195,0.981,0.0,-0.375,0.906,-0.194,-0.191,0.962,-0.194,-0.195,0.981,0.0,-0.195,0.981,0.0,-0.191,0.962,-0.194,-0.0,1.0,0.0,-0.191,0.962,-0.194,-0.0,0.981,-0.194,-0.0,1.0,0.0,-0.0,0.981,0.194,-0.0,1.0,0.0,0.191,0.962,0.194,-0.0,1.0,0.0,0.195,0.981,0.0,0.191,0.962,0.194,0.191,0.962,0.194,0.195,0.981,0.0,0.375,0.906,0.194,0.195,0.981,0.0,0.383,0.924,0.0,0.375,0.906,0.194,0.375,0.906,0.194,0.383,0.924,0.0,0.545,0.816,0.194,0.383,0.924,0.0,0.556,0.831,0.0,0.545,0.816,0.194,0.545,0.816,0.194,0.556,0.831,0.0,0.694,0.694,0.194,0.556,0.831,0.0,0.707,0.707,0.0,0.694,0.694,0.194,0.694,0.694,0.194,0.707,0.707,0.0,0.816,0.545,0.194,0.707,0.707,0.0,0.831,0.556,0.0,0.816,0.545,0.194,0.816,0.545,0.194,0.831,0.556,0.0,0.906,0.375,0.194,0.831,0.556,0.0,0.924,0.383,0.0,0.906,0.375,0.194,0.906,0.375,0.194,0.924,0.383,0.0,0.962,0.191,0.194,0.924,0.383,0.0,0.981,0.195,0.0,0.962,0.191,0.194,0.962,0.191,0.194,0.981,0.195,0.0,0.981,0.0,0.194,0.981,0.195,0.0,1.0,0.0,0.0,0.981,0.0,0.194,0.981,0.0,0.194,1.0,0.0,0.0,0.962,-0.191,0.194,1.0,0.0,0.0,0.981,-0.195,0.0,0.962,-0.191,0.194,0.962,-0.191,0.194,0.981,-0.195,0.0,0.906,-0.375,0.194,0.981,-0.195,0.0,0.924,-0.383,0.0,0.906,-0.375,0.194,0.906,-0.375,0.194,0.924,-0.383,0.0,0.816,-0.545,0.194,0.924,-0.383,0.0,0.831,-0.556,0.0,0.816,-0.545,0.194,0.816,-0.545,0.194,0.831,-0.556,0.0,0.694,-0.694,0.194,0.831,-0.556,0.0,0.707,-0.707,0.0,0.694,-0.694,0.194,0.694,-0.694,0.194,0.707,-0.707,0.0,0.545,-0.816,0.194,0.707,-0.707,0.0,0.556,-0.831,0.0,0.545,-0.816,0.194,0.545,-0.816,0.194,0.556,-0.831,0.0,0.375,-0.906,0.194,0.556,-0.831,0.0,0.383,-0.924,0.0,0.375,-0.906,0.194,0.375,-0.906,0.194,0.383,-0.924,0.0,0.191,-0.962,0.194,0.383,-0.924,0.0,0.195,-0.981,0.0,0.191,-0.962,0.194,0.191,-0.962,0.194,0.195,-0.981,0.0,0.0,-0.981,0.194,0.195,-0.981,0.0,0.0,-1.0,0.0,0.0,-0.981,0.194,0.0,-0.981,0.194,0.0,-1.0,0.0,-0.191,-0.962,0.194,0.0,-1.0,0.0,-0.195,-0.981,0.0,-0.191,-0.962,0.194,-0.191,-0.962,0.194,-0.195,-0.981,0.0,-0.375,-0.906,0.194,-0.195,-0.981,0.0,-0.383,-0.924,0.0,-0.375,-0.906,0.194,-0.375,-0.906,0.194,-0.383,-0.924,0.0,-0.545,-0.816,0.194,-0.383,-0.924,0.0,-0.556,-0.831,0.0,-0.545,-0.816,0.194,-0.545,-0.816,0.194,-0.556,-0.831,0.0,-0.694,-0.694,0.194,-0.556,-0.831,0.0,-0.707,-0.707,0.0,-0.694,-0.694,0.194,-0.694,-0.694,0.194,-0.707,-0.707,0.0,-0.816,-0.545,0.194,-0.707,-0.707,0.0,-0.831,-0.556,0.0,-0.816,-0.545,0.194,-0.816,-0.545,0.194,-0.831,-0.556,0.0,-0.906,-0.375,0.194,-0.831,-0.556,0.0,-0.924,-0.383,0.0,-0.906,-0.375,0.194,-0.906,-0.375,0.194,-0.924,-0.383,0.0,-0.962,-0.191,0.194,-0.924,-0.383,0.0,-0.981,-0.195,0.0,-0.962,-0.191,0.194,-0.962,-0.191,0.194,-0.981,-0.195,0.0,-0.981,-0.0,0.194,-0.981,-0.195,0.0,-1.0,-0.0,0.0,-0.981,-0.0,0.194,-0.981,-0.0,0.194,-1.0,-0.0,0.0,-0.962,0.191,0.194,-1.0,-0.0,0.0,-0.981,0.195,0.0,-0.962,0.191,0.194,-0.962,0.191,0.194,-0.981,0.195,0.0,-0.906,0.375,0.194,-0.981,0.195,0.0,-0.924,0.383,0.0,-0.906,0.375,0.194,-0.906,0.375,0.194,-0.924,0.383,0.0,-0.816,0.545,0.194,-0.924,0.383,0.0,-0.831,0.556,0.0,-0.816,0.545,0.194,-0.816,0.545,0.194,-0.831,0.556,0.0,-0.694,0.694,0.194,-0.831,0.556,0.0,-0.707,0.707,0.0,-0.694,0.694,0.194,-0.694,0.694,0.194,-0.707,0.707,0.0,-0.545,0.816,0.194,-0.707,0.707,0.0,-0.556,0.831,0.0,-0.545,0.816,0.194,-0.545,0.816,0.194,-0.556,0.831,0.0,-0.375,0.906,0.194,-0.556,0.831,0.0,-0.383,0.924,0.0,-0.375,0.906,0.194,-0.375,0.906,0.194,-0.383,0.924,0.0,-0.191,0.962,0.194,-0.383,0.924,0.0,-0.195,0.981,0.0,-0.191,0.962,0.194,-0.191,0.962,0.194,-0.195,0.981,0.0,-0.0,0.981,0.194,-0.195,0.981,0.0,-0.0,1.0,0.0,-0.0,0.981,0.194,-0.0,0.925,0.381,-0.0,0.981,0.194,0.18,0.907,0.381,-0.0,0.981,0.194,0.191,0.962,0.194,0.18,0.907,0.381,0.18,0.907,0.381,0.191,0.962,0.194,0.354,0.854,0.381,0.191,0.962,0.194,0.375,0.906,0.194,0.354,0.854,0.381,0.354,0.854,0.381,0.375,0.906,0.194,0.514,0.769,0.381,0.375,0.906,0.194,0.545,0.816,0.194,0.514,0.769,0.381,0.514,0.769,0.381,0.545,0.816,0.194,0.654,0.654,0.381,0.545,0.816,0.194,0.694,0.694,0.194,0.654,0.654,0.381,0.654,0.654,0.381,0.694,0.694,0.194,0.769,0.514,0.381,0.694,0.694,0.194,0.816,0.545,0.194,0.769,0.514,0.381,0.769,0.514,0.381,0.816,0.545,0.194,0.854,0.354,0.381,0.816,0.545,0.194,0.906,0.375,0.194,0.854,0.354,0.381,0.854,0.354,0.381,0.906,0.375,0.194,0.907,0.18,0.381,0.906,0.375,0.194,0.962,0.191,0.194,0.907,0.18,0.381,0.907,0.18,0.381,0.962,0.191,0.194,0.925,0.0,0.381,0.962,0.191,0.194,0.981,0.0,0.194,0.925,0.0,0.381,0.925,0.0,0.381,0.981,0.0,0.194,0.907,-0.18,0.381,0.981,0.0,0.194,0.962,-0.191,0.194,0.907,-0.18,0.381,0.907,-0.18,0.381,0.962,-0.191,0.194,0.854,-0.354,0.381,0.962,-0.191,0.194,0.906,-0.375,0.194,0.854,-0.354,0.381,0.854,-0.354,0.381,0.906,-0.375,0.194,0.769,-0.514,0.381,0.906,-0.375,0.194,0.816,-0.545,0.194,0.769,-0.514,0.381,0.769,-0.514,0.381,0.816,-0.545,0.194,0.654,-0.654,0.381,0.816,-0.545,0.194,0.694,-0.694,0.194,0.654,-0.654,0.381,0.654,-0.654,0.381,0.694,-0.694,0.194,0.514,-0.769,0.381,0.694,-0.694,0.194,0.545,-0.816,0.194,0.514,-0.769,0.381,0.514,-0.769,0.381,0.545,-0.816,0.194,0.354,-0.854,0.381,0.545,-0.816,0.194,0.375,-0.906,0.194,0.354,-0.854,0.381,0.354,-0.854,0.381,0.375,-0.906,0.194,0.18,-0.907,0.381,0.375,-0.906,0.194,0.191,-0.962,0.194,0.18,-0.907,0.381,0.18,-0.907,0.381,0.191,-0.962,0.194,0.0,-0.925,0.381,0.191,-0.962,0.194,0.0,-0.981,0.194,0.0,-0.925,0.381,0.0,-0.925,0.381,0.0,-0.981,0.194,-0.18,-0.907,0.381,0.0,-0.981,0.194,-0.191,-0.962,0.194,-0.18,-0.907,0.381,-0.18,-0.907,0.381,-0.191,-0.962,0.194,-0.354,-0.854,0.381,-0.191,-0.962,0.194,-0.375,-0.906,0.194,-0.354,-0.854,0.381,-0.354,-0.854,0.381,-0.375,-0.906,0.194,-0.514,-0.769,0.381,-0.375,-0.906,0.194,-0.545,-0.816,0.194,-0.514,-0.769,0.381,-0.514,-0.769,0.381,-0.545,-0.816,0.194,-0.654,-0.654,0.381,-0.545,-0.816,0.194,-0.694,-0.694,0.194,-0.654,-0.654,0.381,-0.654,-0.654,0.381,-0.694,-0.694,0.194,-0.769,-0.514,0.381,-0.694,-0.694,0.194,-0.816,-0.545,0.194,-0.769,-0.514,0.381,-0.769,-0.514,0.381,-0.816,-0.545,0.194,-0.854,-0.354,0.381,-0.816,-0.545,0.194,-0.906,-0.375,0.194,-0.854,-0.354,0.381,-0.854,-0.354,0.381,-0.906,-0.375,0.194,-0.907,-0.18,0.381,-0.906,-0.375,0.194,-0.962,-0.191,0.194,-0.907,-0.18,0.381,-0.907,-0.18,0.381,-0.962,-0.191,0.194,-0.925,-0.0,0.381,-0.962,-0.191,0.194,-0.981,-0.0,0.194,-0.925,-0.0,0.381,-0.925,-0.0,0.381,-0.981,-0.0,0.194,-0.907,0.18,0.381,-0.981,-0.0,0.194,-0.962,0.191,0.194,-0.907,0.18,0.381,-0.907,0.18,0.381,-0.962,0.191,0.194,-0.854,0.354,0.381,-0.962,0.191,0.194,-0.906,0.375,0.194,-0.854,0.354,0.381,-0.854,0.354,0.381,-0.906,0.375,0.194,-0.769,0.514,0.381,-0.906,0.375,0.194,-0.816,0.545,0.194,-0.769,0.514,0.381,-0.769,0.514,0.381,-0.816,0.545,0.194,-0.654,0.654,0.381,-0.816,0.545,0.194,-0.694,0.694,0.194,-0.654,0.654,0.381,-0.654,0.654,0.381,-0.694,0.694,0.194,-0.514,0.769,0.381,-0.694,0.694,0.194,-0.545,0.816,0.194,-0.514,0.769,0.381,-0.514,0.769,0.381,-0.545,0.816,0.194,-0.354,0.854,0.381,-0.545,0.816,0.194,-0.375,0.906,0.194,-0.354,0.854,0.381,-0.354,0.854,0.381,-0.375,0.906,0.194,-0.18,0.907,0.381,-0.375,0.906,0.194,-0.191,0.962,0.194,-0.18,0.907,0.381,-0.18,0.907,0.381,-0.191,0.962,0.194,-0.0,0.925,0.381,-0.191,0.962,0.194,-0.0,0.981,0.194,-0.0,0.925,0.381,-0.0,0.833,0.553,-0.0,0.925,0.381,0.163,0.817,0.553,-0.0,0.925,0.381,0.18,0.907,0.381,0.163,0.817,0.553,0.163,0.817,0.553,0.18,0.907,0.381,0.319,0.77,0.553,0.18,0.907,0.381,0.354,0.854,0.381,0.319,0.77,0.553,0.319,0.77,0.553,0.354,0.854,0.381,0.463,0.693,0.553,0.354,0.854,0.381,0.514,0.769,0.381,0.463,0.693,0.553,0.463,0.693,0.553,0.514,0.769,0.381,0.589,0.589,0.553,0.514,0.769,0.381,0.654,0.654,0.381,0.589,0.589,0.553,0.589,0.589,0.553,0.654,0.654,0.381,0.693,0.463,0.553,0.654,0.654,0.381,0.769,0.514,0.381,0.693,0.463,0.553,0.693,0.463,0.553,0.769,0.514,0.381,0.77,0.319,0.553,0.769,0.514,0.381,0.854,0.354,0.381,0.77,0.319,0.553,0.77,0.319,0.553,0.854,0.354,0.381,0.817,0.163,0.553,0.854,0.354,0.381,0.907,0.18,0.381,0.817,0.163,0.553,0.817,0.163,0.553,0.907,0.18,0.381,0.833,0.0,0.553,0.907,0.18,0.381,0.925,0.0,0.381,0.833,0.0,0.553,0.833,0.0,0.553,0.925,0.0,0.381,0.817,-0.163,0.553,0.925,0.0,0.381,0.907,-0.18,0.381,0.817,-0.163,0.553,0.817,-0.163,0.553,0.907,-0.18,0.381,0.77,-0.319,0.553,0.907,-0.18,0.381,0.854,-0.354,0.381,0.77,-0.319,0.553,0.77,-0.319,0.553,0.854,-0.354,0.381,0.693,-0.463,0.553,0.854,-0.354,0.381,0.769,-0.514,0.381,0.693,-0.463,0.553,0.693,-0.463,0.553,0.769,-0.514,0.381,0.589,-0.589,0.553,0.769,-0.514,0.381,0.654,-0.654,0.381,0.589,-0.589,0.553,0.589,-0.589,0.553,0.654,-0.654,0.381,0.463,-0.693,0.553,0.654,-0.654,0.381,0.514,-0.769,0.381,0.463,-0.693,0.553,0.463,-0.693,0.553,0.514,-0.769,0.381,0.319,-0.77,0.553,0.514,-0.769,0.381,0.354,-0.854,0.381,0.319,-0.77,0.553,0.319,-0.77,0.553,0.354,-0.854,0.381,0.163,-0.817,0.553,0.354,-0.854,0.381,0.18,-0.907,0.381,0.163,-0.817,0.553,0.163,-0.817,0.553,0.18,-0.907,0.381,0.0,-0.833,0.553,0.18,-0.907,0.381,0.0,-0.925,0.381,0.0,-0.833,0.553,0.0,-0.833,0.553,0.0,-0.925,0.381,-0.163,-0.817,0.553,0.0,-0.925,0.381,-0.18,-0.907,0.381,-0.163,-0.817,0.553,-0.163,-0.817,0.553,-0.18,-0.907,0.381,-0.319,-0.77,0.553,-0.18,-0.907,0.381,-0.354,-0.854,0.381,-0.319,-0.77,0.553,-0.319,-0.77,0.553,-0.354,-0.854,0.381,-0.463,-0.693,0.553,-0.354,-0.854,0.381,-0.514,-0.769,0.381,-0.463,-0.693,0.553,-0.463,-0.693,0.553,-0.514,-0.769,0.381,-0.589,-0.589,0.553,-0.514,-0.769,0.381,-0.654,-0.654,0.381,-0.589,-0.589,0.553,-0.589,-0.589,0.553,-0.654,-0.654,0.381,-0.693,-0.463,0.553,-0.654,-0.654,0.381,-0.769,-0.514,0.381,-0.693,-0.463,0.553,-0.693,-0.463,0.553,-0.769,-0.514,0.381,-0.77,-0.319,0.553,-0.769,-0.514,0.381,-0.854,-0.354,0.381,-0.77,-0.319,0.553,-0.77,-0.319,0.553,-0.854,-0.354,0.381,-0.817,-0.163,0.553,-0.854,-0.354,0.381,-0.907,-0.18,0.381,-0.817,-0.163,0.553,-0.817,-0.163,0.553,-0.907,-0.18,0.381,-0.833,-0.0,0.553,-0.907,-0.18,0.381,-0.925,-0.0,0.381,-0.833,-0.0,0.553,-0.833,-0.0,0.553,-0.925,-0.0,0.381,-0.817,0.163,0.553,-0.925,-0.0,0.381,-0.907,0.18,0.381,-0.817,0.163,0.553,-0.817,0.163,0.553,-0.907,0.18,0.381,-0.77,0.319,0.553,-0.907,0.18,0.381,-0.854,0.354,0.381,-0.77,0.319,0.553,-0.77,0.319,0.553,-0.854,0.354,0.381,-0.693,0.463,0.553,-0.854,0.354,0.381,-0.769,0.514,0.381,-0.693,0.463,0.553,-0.693,0.463,0.553,-0.769,0.514,0.381,-0.589,0.589,0.553,-0.769,0.514,0.381,-0.654,0.654,0.381,-0.589,0.589,0.553,-0.589,0.589,0.553,-0.654,0.654,0.381,-0.463,0.693,0.553,-0.654,0.654,0.381,-0.514,0.769,0.381,-0.463,0.693,0.553,-0.463,0.693,0.553,-0.514,0.769,0.381,-0.319,0.77,0.553,-0.514,0.769,0.381,-0.354,0.854,0.381,-0.319,0.77,0.553,-0.319,0.77,0.553,-0.354,0.854,0.381,-0.163,0.817,0.553,-0.354,0.854,0.381,-0.18,0.907,0.381,-0.163,0.817,0.553,-0.163,0.817,0.553,-0.18,0.907,0.381,-0.0,0.833,0.553,-0.18,0.907,0.381,-0.0,0.925,0.381,-0.0,0.833,0.553,-0.0,0.71,0.704,-0.0,0.833,0.553,0.139,0.696,0.704,-0.0,0.833,0.553,0.163,0.817,0.553,0.139,0.696,0.704,0.139,0.696,0.704,0.163,0.817,0.553,0.272,0.656,0.704,0.163,0.817,0.553,0.319,0.77,0.553,0.272,0.656,0.704,0.272,0.656,0.704,0.319,0.77,0.553,0.395,0.59,0.704,0.319,0.77,0.553,0.463,0.693,0.553,0.395,0.59,0.704,0.395,0.59,0.704,0.463,0.693,0.553,0.502,0.502,0.704,0.463,0.693,0.553,0.589,0.589,0.553,0.502,0.502,0.704,0.502,0.502,0.704,0.589,0.589,0.553,0.59,0.395,0.704,0.589,0.589,0.553,0.693,0.463,0.553,0.59,0.395,0.704,0.59,0.395,0.704,0.693,0.463,0.553,0.656,0.272,0.704,0.693,0.463,0.553,0.77,0.319,0.553,0.656,0.272,0.704,0.656,0.272,0.704,0.77,0.319,0.553,0.696,0.139,0.704,0.77,0.319,0.553,0.817,0.163,0.553,0.696,0.139,0.704,0.696,0.139,0.704,0.817,0.163,0.553,0.71,0.0,0.704,0.817,0.163,0.553,0.833,0.0,0.553,0.71,0.0,0.704,0.71,0.0,0.704,0.833,0.0,0.553,0.696,-0.139,0.704,0.833,0.0,0.553,0.817,-0.163,0.553,0.696,-0.139,0.704,0.696,-0.139,0.704,0.817,-0.163,0.553,0.656,-0.272,0.704,0.817,-0.163,0.553,0.77,-0.319,0.553,0.656,-0.272,0.704,0.656,-0.272,0.704,0.77,-0.319,0.553,0.59,-0.395,0.704,0.77,-0.319,0.553,0.693,-0.463,0.553,0.59,-0.395,0.704,0.59,-0.395,0.704,0.693,-0.463,0.553,0.502,-0.502,0.704,0.693,-0.463,0.553,0.589,-0.589,0.553,0.502,-0.502,0.704,0.502,-0.502,0.704,0.589,-0.589,0.553,0.395,-0.59,0.704,0.589,-0.589,0.553,0.463,-0.693,0.553,0.395,-0.59,0.704,0.395,-0.59,0.704,0.463,-0.693,0.553,0.272,-0.656,0.704,0.463,-0.693,0.553,0.319,-0.77,0.553,0.272,-0.656,0.704,0.272,-0.656,0.704,0.319,-0.77,0.553,0.139,-0.696,0.704,0.319,-0.77,0.553,0.163,-0.817,0.553,0.139,-0.696,0.704,0.139,-0.696,0.704,0.163,-0.817,0.553,0.0,-0.71,0.704,0.163,-0.817,0.553,0.0,-0.833,0.553,0.0,-0.71,0.704,0.0,-0.71,0.704,0.0,-0.833,0.553,-0.139,-0.696,0.704,0.0,-0.833,0.553,-0.163,-0.817,0.553,-0.139,-0.696,0.704,-0.139,-0.696,0.704,-0.163,-0.817,0.553,-0.272,-0.656,0.704,-0.163,-0.817,0.553,-0.319,-0.77,0.553,-0.272,-0.656,0.704,-0.272,-0.656,0.704,-0.319,-0.77,0.553,-0.395,-0.59,0.704,-0.319,-0.77,0.553,-0.463,-0.693,0.553,-0.395,-0.59,0.704,-0.395,-0.59,0.704,-0.463,-0.693,0.553,-0.502,-0.502,0.704,-0.463,-0.693,0.553,-0.589,-0.589,0.553,-0.502,-0.502,0.704,-0.502,-0.502,0.704,-0.589,-0.589,0.553,-0.59,-0.395,0.704,-0.589,-0.589,0.553,-0.693,-0.463,0.553,-0.59,-0.395,0.704,-0.59,-0.395,0.704,-0.693,-0.463,0.553,-0.656,-0.272,0.704,-0.693,-0.463,0.553,-0.77,-0.319,0.553,-0.656,-0.272,0.704,-0.656,-0.272,0.704,-0.77,-0.319,0.553,-0.696,-0.139,0.704,-0.77,-0.319,0.553,-0.817,-0.163,0.553,-0.696,-0.139,0.704,-0.696,-0.139,0.704,-0.817,-0.163,0.553,-0.71,-0.0,0.704,-0.817,-0.163,0.553,-0.833,-0.0,0.553,-0.71,-0.0,0.704,-0.71,-0.0,0.704,-0.833,-0.0,0.553,-0.696,0.139,0.704,-0.833,-0.0,0.553,-0.817,0.163,0.553,-0.696,0.139,0.704,-0.696,0.139,0.704,-0.817,0.163,0.553,-0.656,0.272,0.704,-0.817,0.163,0.553,-0.77,0.319,0.553,-0.656,0.272,0.704,-0.656,0.272,0.704,-0.77,0.319,0.553,-0.59,0.395,0.704,-0.77,0.319,0.553,-0.693,0.463,0.553,-0.59,0.395,0.704,-0.59,0.395,0.704,-0.693,0.463,0.553,-0.502,0.502,0.704,-0.693,0.463,0.553,-0.589,0.589,0.553,-0.502,0.502,0.704,-0.502,0.502,0.704,-0.589,0.589,0.553,-0.395,0.59,0.704,-0.589,0.589,0.553,-0.463,0.693,0.553,-0.395,0.59,0.704,-0.395,0.59,0.704,-0.463,0.693,0.553,-0.272,0.656,0.704,-0.463,0.693,0.553,-0.319,0.77,0.553,-0.272,0.656,0.704,-0.272,0.656,0.704,-0.319,0.77,0.553,-0.139,0.696,0.704,-0.319,0.77,0.553,-0.163,0.817,0.553,-0.139,0.696,0.704,-0.139,0.696,0.704,-0.163,0.817,0.553,-0.0,0.71,0.704,-0.163,0.817,0.553,-0.0,0.833,0.553,-0.0,0.71,0.704,-0.0,0.56,0.829,-0.0,0.71,0.704,0.109,0.549,0.829,-0.0,0.71,0.704,0.139,0.696,0.704,0.109,0.549,0.829,0.109,0.549,0.829,0.139,0.696,0.704,0.214,0.517,0.829,0.139,0.696,0.704,0.272,0.656,0.704,0.214,0.517,0.829,0.214,0.517,0.829,0.272,0.656,0.704,0.311,0.465,0.829,0.272,0.656,0.704,0.395,0.59,0.704,0.311,0.465,0.829,0.311,0.465,0.829,0.395,0.59,0.704,0.396,0.396,0.829,0.395,0.59,0.704,0.502,0.502,0.704,0.396,0.396,0.829,0.396,0.396,0.829,0.502,0.502,0.704,0.465,0.311,0.829,0.502,0.502,0.704,0.59,0.395,0.704,0.465,0.311,0.829,0.465,0.311,0.829,0.59,0.395,0.704,0.517,0.214,0.829,0.59,0.395,0.704,0.656,0.272,0.704,0.517,0.214,0.829,0.517,0.214,0.829,0.656,0.272,0.704,0.549,0.109,0.829,0.656,0.272,0.704,0.696,0.139,0.704,0.549,0.109,0.829,0.549,0.109,0.829,0.696,0.139,0.704,0.56,0.0,0.829,0.696,0.139,0.704,0.71,0.0,0.704,0.56,0.0,0.829,0.56,0.0,0.829,0.71,0.0,0.704,0.549,-0.109,0.829,0.71,0.0,0.704,0.696,-0.139,0.704,0.549,-0.109,0.829,0.549,-0.109,0.829,0.696,-0.139,0.704,0.517,-0.214,0.829,0.696,-0.139,0.704,0.656,-0.272,0.704,0.517,-0.214,0.829,0.517,-0.214,0.829,0.656,-0.272,0.704,0.465,-0.311,0.829,0.656,-0.272,0.704,0.59,-0.395,0.704,0.465,-0.311,0.829,0.465,-0.311,0.829,0.59,-0.395,0.704,0.396,-0.396,0.829,0.59,-0.395,0.704,0.502,-0.502,0.704,0.396,-0.396,0.829,0.396,-0.396,0.829,0.502,-0.502,0.704,0.311,-0.465,0.829,0.502,-0.502,0.704,0.395,-0.59,0.704,0.311,-0.465,0.829,0.311,-0.465,0.829,0.395,-0.59,0.704,0.214,-0.517,0.829,0.395,-0.59,0.704,0.272,-0.656,0.704,0.214,-0.517,0.829,0.214,-0.517,0.829,0.272,-0.656,0.704,0.109,-0.549,0.829,0.272,-0.656,0.704,0.139,-0.696,0.704,0.109,-0.549,0.829,0.109,-0.549,0.829,0.139,-0.696,0.704,0.0,-0.56,0.829,0.139,-0.696,0.704,0.0,-0.71,0.704,0.0,-0.56,0.829,0.0,-0.56,0.829,0.0,-0.71,0.704,-0.109,-0.549,0.829,0.0,-0.71,0.704,-0.139,-0.696,0.704,-0.109,-0.549,0.829,-0.109,-0.549,0.829,-0.139,-0.696,0.704,-0.214,-0.517,0.829,-0.139,-0.696,0.704,-0.272,-0.656,0.704,-0.214,-0.517,0.829,-0.214,-0.517,0.829,-0.272,-0.656,0.704,-0.311,-0.465,0.829,-0.272,-0.656,0.704,-0.395,-0.59,0.704,-0.311,-0.465,0.829,-0.311,-0.465,0.829,-0.395,-0.59,0.704,-0.396,-0.396,0.829,-0.395,-0.59,0.704,-0.502,-0.502,0.704,-0.396,-0.396,0.829,-0.396,-0.396,0.829,-0.502,-0.502,0.704,-0.465,-0.311,0.829,-0.502,-0.502,0.704,-0.59,-0.395,0.704,-0.465,-0.311,0.829,-0.465,-0.311,0.829,-0.59,-0.395,0.704,-0.517,-0.214,0.829,-0.59,-0.395,0.704,-0.656,-0.272,0.704,-0.517,-0.214,0.829,-0.517,-0.214,0.829,-0.656,-0.272,0.704,-0.549,-0.109,0.829,-0.656,-0.272,0.704,-0.696,-0.139,0.704,-0.549,-0.109,0.829,-0.549,-0.109,0.829,-0.696,-0.139,0.704,-0.56,-0.0,0.829,-0.696,-0.139,0.704,-0.71,-0.0,0.704,-0.56,-0.0,0.829,-0.56,-0.0,0.829,-0.71,-0.0,0.704,-0.549,0.109,0.829,-0.71,-0.0,0.704,-0.696,0.139,0.704,-0.549,0.109,0.829,-0.549,0.109,0.829,-0.696,0.139,0.704,-0.517,0.214,0.829,-0.696,0.139,0.704,-0.656,0.272,0.704,-0.517,0.214,0.829,-0.517,0.214,0.829,-0.656,0.272,0.704,-0.465,0.311,0.829,-0.656,0.272,0.704,-0.59,0.395,0.704,-0.465,0.311,0.829,-0.465,0.311,0.829,-0.59,0.395,0.704,-0.396,0.396,0.829,-0.59,0.395,0.704,-0.502,0.502,0.704,-0.396,0.396,0.829,-0.396,0.396,0.829,-0.502,0.502,0.704,-0.311,0.465,0.829,-0.502,0.502,0.704,-0.395,0.59,0.704,-0.311,0.465,0.829,-0.311,0.465,0.829,-0.395,0.59,0.704,-0.214,0.517,0.829,-0.395,0.59,0.704,-0.272,0.656,0.704,-0.214,0.517,0.829,-0.214,0.517,0.829,-0.272,0.656,0.704,-0.109,0.549,0.829,-0.272,0.656,0.704,-0.139,0.696,0.704,-0.109,0.549,0.829,-0.109,0.549,0.829,-0.139,0.696,0.704,-0.0,0.56,0.829,-0.139,0.696,0.704,-0.0,0.71,0.704,-0.0,0.56,0.829,-0.0,0.388,0.922,-0.0,0.56,0.829,0.076,0.38,0.922,-0.0,0.56,0.829,0.109,0.549,0.829,0.076,0.38,0.922,0.076,0.38,0.922,0.109,0.549,0.829,0.148,0.358,0.922,0.109,0.549,0.829,0.214,0.517,0.829,0.148,0.358,0.922,0.148,0.358,0.922,0.214,0.517,0.829,0.215,0.323,0.922,0.214,0.517,0.829,0.311,0.465,0.829,0.215,0.323,0.922,0.215,0.323,0.922,0.311,0.465,0.829,0.274,0.274,0.922,0.311,0.465,0.829,0.396,0.396,0.829,0.274,0.274,0.922,0.274,0.274,0.922,0.396,0.396,0.829,0.323,0.215,0.922,0.396,0.396,0.829,0.465,0.311,0.829,0.323,0.215,0.922,0.323,0.215,0.922,0.465,0.311,0.829,0.358,0.148,0.922,0.465,0.311,0.829,0.517,0.214,0.829,0.358,0.148,0.922,0.358,0.148,0.922,0.517,0.214,0.829,0.38,0.076,0.922,0.517,0.214,0.829,0.549,0.109,0.829,0.38,0.076,0.922,0.38,0.076,0.922,0.549,0.109,0.829,0.388,0.0,0.922,0.549,0.109,0.829,0.56,0.0,0.829,0.388,0.0,0.922,0.388,0.0,0.922,0.56,0.0,0.829,0.38,-0.076,0.922,0.56,0.0,0.829,0.549,-0.109,0.829,0.38,-0.076,0.922,0.38,-0.076,0.922,0.549,-0.109,0.829,0.358,-0.148,0.922,0.549,-0.109,0.829,0.517,-0.214,0.829,0.358,-0.148,0.922,0.358,-0.148,0.922,0.517,-0.214,0.829,0.323,-0.215,0.922,0.517,-0.214,0.829,0.465,-0.311,0.829,0.323,-0.215,0.922,0.323,-0.215,0.922,0.465,-0.311,0.829,0.274,-0.274,0.922,0.465,-0.311,0.829,0.396,-0.396,0.829,0.274,-0.274,0.922,0.274,-0.274,0.922,0.396,-0.396,0.829,0.215,-0.323,0.922,0.396,-0.396,0.829,0.311,-0.465,0.829,0.215,-0.323,0.922,0.215,-0.323,0.922,0.311,-0.465,0.829,0.148,-0.358,0.922,0.311,-0.465,0.829,0.214,-0.517,0.829,0.148,-0.358,0.922,0.148,-0.358,0.922,0.214,-0.517,0.829,0.076,-0.38,0.922,0.214,-0.517,0.829,0.109,-0.549,0.829,0.076,-0.38,0.922,0.076,-0.38,0.922,0.109,-0.549,0.829,0.0,-0.388,0.922,0.109,-0.549,0.829,0.0,-0.56,0.829,0.0,-0.388,0.922,0.0,-0.388,0.922,0.0,-0.56,0.829,-0.076,-0.38,0.922,0.0,-0.56,0.829,-0.109,-0.549,0.829,-0.076,-0.38,0.922,-0.076,-0.38,0.922,-0.109,-0.549,0.829,-0.148,-0.358,0.922,-0.109,-0.549,0.829,-0.214,-0.517,0.829,-0.148,-0.358,0.922,-0.148,-0.358,0.922,-0.214,-0.517,0.829,-0.215,-0.323,0.922,-0.214,-0.517,0.829,-0.311,-0.465,0.829,-0.215,-0.323,0.922,-0.215,-0.323,0.922,-0.311,-0.465,0.829,-0.274,-0.274,0.922,-0.311,-0.465,0.829,-0.396,-0.396,0.829,-0.274,-0.274,0.922,-0.274,-0.274,0.922,-0.396,-0.396,0.829,-0.323,-0.215,0.922,-0.396,-0.396,0.829,-0.465,-0.311,0.829,-0.323,-0.215,0.922,-0.323,-0.215,0.922,-0.465,-0.311,0.829,-0.358,-0.148,0.922,-0.465,-0.311,0.829,-0.517,-0.214,0.829,-0.358,-0.148,0.922,-0.358,-0.148,0.922,-0.517,-0.214,0.829,-0.38,-0.076,0.922,-0.517,-0.214,0.829,-0.549,-0.109,0.829,-0.38,-0.076,0.922,-0.38,-0.076,0.922,-0.549,-0.109,0.829,-0.388,-0.0,0.922,-0.549,-0.109,0.829,-0.56,-0.0,0.829,-0.388,-0.0,0.922,-0.388,-0.0,0.922,-0.56,-0.0,0.829,-0.38,0.076,0.922,-0.56,-0.0,0.829,-0.549,0.109,0.829,-0.38,0.076,0.922,-0.38,0.076,0.922,-0.549,0.109,0.829,-0.358,0.148,0.922,-0.549,0.109,0.829,-0.517,0.214,0.829,-0.358,0.148,0.922,-0.358,0.148,0.922,-0.517,0.214,0.829,-0.323,0.215,0.922,-0.517,0.214,0.829,-0.465,0.311,0.829,-0.323,0.215,0.922,-0.323,0.215,0.922,-0.465,0.311,0.829,-0.274,0.274,0.922,-0.465,0.311,0.829,-0.396,0.396,0.829,-0.274,0.274,0.922,-0.274,0.274,0.922,-0.396,0.396,0.829,-0.215,0.323,0.922,-0.396,0.396,0.829,-0.311,0.465,0.829,-0.215,0.323,0.922,-0.215,0.323,0.922,-0.311,0.465,0.829,-0.148,0.358,0.922,-0.311,0.465,0.829,-0.214,0.517,0.829,-0.148,0.358,0.922,-0.148,0.358,0.922,-0.214,0.517,0.829,-0.076,0.38,0.922,-0.214,0.517,0.829,-0.109,0.549,0.829,-0.076,0.38,0.922,-0.076,0.38,0.922,-0.109,0.549,0.829,-0.0,0.388,0.922,-0.109,0.549,0.829,-0.0,0.56,0.829,-0.0,0.388,0.922,-0.0,0.201,0.98,-0.0,0.388,0.922,0.039,0.197,0.98,-0.0,0.388,0.922,0.076,0.38,0.922,0.039,0.197,0.98,0.039,0.197,0.98,0.076,0.38,0.922,0.077,0.186,0.98,0.076,0.38,0.922,0.148,0.358,0.922,0.077,0.186,0.98,0.077,0.186,0.98,0.148,0.358,0.922,0.112,0.167,0.98,0.148,0.358,0.922,0.215,0.323,0.922,0.112,0.167,0.98,0.112,0.167,0.98,0.215,0.323,0.922,0.142,0.142,0.98,0.215,0.323,0.922,0.274,0.274,0.922,0.142,0.142,0.98,0.142,0.142,0.98,0.274,0.274,0.922,0.167,0.112,0.98,0.274,0.274,0.922,0.323,0.215,0.922,0.167,0.112,0.98,0.167,0.112,0.98,0.323,0.215,0.922,0.186,0.077,0.98,0.323,0.215,0.922,0.358,0.148,0.922,0.186,0.077,0.98,0.186,0.077,0.98,0.358,0.148,0.922,0.197,0.039,0.98,0.358,0.148,0.922,0.38,0.076,0.922,0.197,0.039,0.98,0.197,0.039,0.98,0.38,0.076,0.922,0.201,0.0,0.98,0.38,0.076,0.922,0.388,0.0,0.922,0.201,0.0,0.98,0.201,0.0,0.98,0.388,0.0,0.922,0.197,-0.039,0.98,0.388,0.0,0.922,0.38,-0.076,0.922,0.197,-0.039,0.98,0.197,-0.039,0.98,0.38,-0.076,0.922,0.186,-0.077,0.98,0.38,-0.076,0.922,0.358,-0.148,0.922,0.186,-0.077,0.98,0.186,-0.077,0.98,0.358,-0.148,0.922,0.167,-0.112,0.98,0.358,-0.148,0.922,0.323,-0.215,0.922,0.167,-0.112,0.98,0.167,-0.112,0.98,0.323,-0.215,0.922,0.142,-0.142,0.98,0.323,-0.215,0.922,0.274,-0.274,0.922,0.142,-0.142,0.98,0.142,-0.142,0.98,0.274,-0.274,0.922,0.112,-0.167,0.98,0.274,-0.274,0.922,0.215,-0.323,0.922,0.112,-0.167,0.98,0.112,-0.167,0.98,0.215,-0.323,0.922,0.077,-0.186,0.98,0.215,-0.323,0.922,0.148,-0.358,0.922,0.077,-0.186,0.98,0.077,-0.186,0.98,0.148,-0.358,0.922,0.039,-0.197,0.98,0.148,-0.358,0.922,0.076,-0.38,0.922,0.039,-0.197,0.98,0.039,-0.197,0.98,0.076,-0.38,0.922,0.0,-0.201,0.98,0.076,-0.38,0.922,0.0,-0.388,0.922,0.0,-0.201,0.98,0.0,-0.201,0.98,0.0,-0.388,0.922,-0.039,-0.197,0.98,0.0,-0.388,0.922,-0.076,-0.38,0.922,-0.039,-0.197,0.98,-0.039,-0.197,0.98,-0.076,-0.38,0.922,-0.077,-0.186,0.98,-0.076,-0.38,0.922,-0.148,-0.358,0.922,-0.077,-0.186,0.98,-0.077,-0.186,0.98,-0.148,-0.358,0.922,-0.112,-0.167,0.98,-0.148,-0.358,0.922,-0.215,-0.323,0.922,-0.112,-0.167,0.98,-0.112,-0.167,0.98,-0.215,-0.323,0.922,-0.142,-0.142,0.98,-0.215,-0.323,0.922,-0.274,-0.274,0.922,-0.142,-0.142,0.98,-0.142,-0.142,0.98,-0.274,-0.274,0.922,-0.167,-0.112,0.98,-0.274,-0.274,0.922,-0.323,-0.215,0.922,-0.167,-0.112,0.98,-0.167,-0.112,0.98,-0.323,-0.215,0.922,-0.186,-0.077,0.98,-0.323,-0.215,0.922,-0.358,-0.148,0.922,-0.186,-0.077,0.98,-0.186,-0.077,0.98,-0.358,-0.148,0.922,-0.197,-0.039,0.98,-0.358,-0.148,0.922,-0.38,-0.076,0.922,-0.197,-0.039,0.98,-0.197,-0.039,0.98,-0.38,-0.076,0.922,-0.201,-0.0,0.98,-0.38,-0.076,0.922,-0.388,-0.0,0.922,-0.201,-0.0,0.98,-0.201,-0.0,0.98,-0.388,-0.0,0.922,-0.197,0.039,0.98,-0.388,-0.0,0.922,-0.38,0.076,0.922,-0.197,0.039,0.98,-0.197,0.039,0.98,-0.38,0.076,0.922,-0.186,0.077,0.98,-0.38,0.076,0.922,-0.358,0.148,0.922,-0.186,0.077,0.98,-0.186,0.077,0.98,-0.358,0.148,0.922,-0.167,0.112,0.98,-0.358,0.148,0.922,-0.323,0.215,0.922,-0.167,0.112,0.98,-0.167,0.112,0.98,-0.323,0.215,0.922,-0.142,0.142,0.98,-0.323,0.215,0.922,-0.274,0.274,0.922,-0.142,0.142,0.98,-0.142,0.142,0.98,-0.274,0.274,0.922,-0.112,0.167,0.98,-0.274,0.274,0.922,-0.215,0.323,0.922,-0.112,0.167,0.98,-0.112,0.167,0.98,-0.215,0.323,0.922,-0.077,0.186,0.98,-0.215,0.323,0.922,-0.148,0.358,0.922,-0.077,0.186,0.98,-0.077,0.186,0.98,-0.148,0.358,0.922,-0.039,0.197,0.98,-0.148,0.358,0.922,-0.076,0.38,0.922,-0.039,0.197,0.98,-0.039,0.197,0.98,-0.076,0.38,0.922,-0.0,0.201,0.98,-0.076,0.38,0.922,-0.0,0.388,0.922,-0.0,0.201,0.98,-0.0,0.201,0.98,0.039,0.197,0.98,0.0,0.0,1.0,0.039,0.197,0.98,0.077,0.186,0.98,0.0,0.0,1.0,0.077,0.186,0.98,0.112,0.167,0.98,0.0,0.0,1.0,0.112,0.167,0.98,0.142,0.142,0.98,0.0,0.0,1.0,0.142,0.142,0.98,0.167,0.112,0.98,0.0,0.0,1.0,0.167,0.112,0.98,0.186,0.077,0.98,0.0,0.0,1.0,0.186,0.077,0.98,0.197,0.039,0.98,0.0,0.0,1.0,0.197,0.039,0.98,0.201,0.0,0.98,0.0,0.0,1.0,0.201,0.0,0.98,0.197,-0.039,0.98,0.0,0.0,1.0,0.197,-0.039,0.98,0.186,-0.077,0.98,0.0,0.0,1.0,0.186,-0.077,0.98,0.167,-0.112,0.98,0.0,0.0,1.0,0.167,-0.112,0.98,0.142,-0.142,0.98,0.0,0.0,1.0,0.142,-0.142,0.98,0.112,-0.167,0.98,0.0,0.0,1.0,0.112,-0.167,0.98,0.077,-0.186,0.98,0.0,0.0,1.0,0.077,-0.186,0.98,0.039,-0.197,0.98,0.0,0.0,1.0,0.039,-0.197,0.98,0.0,-0.201,0.98,0.0,0.0,1.0,0.0,-0.201,0.98,-0.039,-0.197,0.98,0.0,0.0,1.0,-0.039,-0.197,0.98,-0.077,-0.186,0.98,0.0,0.0,1.0,-0.077,-0.186,0.98,-0.112,-0.167,0.98,0.0,0.0,1.0,-0.112,-0.167,0.98,-0.142,-0.142,0.98,0.0,0.0,1.0,-0.142,-0.142,0.98,-0.167,-0.112,0.98,0.0,0.0,1.0,-0.167,-0.112,0.98,-0.186,-0.077,0.98,0.0,0.0,1.0,-0.186,-0.077,0.98,-0.197,-0.039,0.98,0.0,0.0,1.0,-0.197,-0.039,0.98,-0.201,-0.0,0.98,0.0,0.0,1.0,-0.201,-0.0,0.98,-0.197,0.039,0.98,0.0,0.0,1.0,-0.197,0.039,0.98,-0.186,0.077,0.98,0.0,0.0,1.0,-0.186,0.077,0.98,-0.167,0.112,0.98,0.0,0.0,1.0,-0.167,0.112,0.98,-0.142,0.142,0.98,0.0,0.0,1.0,-0.142,0.142,0.98,-0.112,0.167,0.98,0.0,0.0,1.0,-0.112,0.167,0.98,-0.077,0.186,0.98,0.0,0.0,1.0,-0.077,0.186,0.98,-0.039,0.197,0.98,0.0,0.0,1.0,-0.039,0.197,0.98,-0.0,0.201,0.98,0.0,0.0,1.0]},\\"uv0\\":{\\"valueType\\":\\"Float32\\",\\"valuesPerElement\\":2,\\"values\\":[0.9688,0.9375,1.0,0.9375,0.9688,1.0,0.9375,0.9375,0.9688,0.9375,0.9375,1.0,0.9063,0.9375,0.9375,0.9375,0.9063,1.0,0.875,0.9375,0.9063,0.9375,0.875,1.0,0.8438,0.9375,0.875,0.9375,0.8438,1.0,0.8125,0.9375,0.8438,0.9375,0.8125,1.0,0.7813,0.9375,0.8125,0.9375,0.7813,1.0,0.75,0.9375,0.7813,0.9375,0.75,1.0,0.7188,0.9375,0.75,0.9375,0.7188,1.0,0.6875,0.9375,0.7188,0.9375,0.6875,1.0,0.6563,0.9375,0.6875,0.9375,0.6563,1.0,0.625,0.9375,0.6563,0.9375,0.625,1.0,0.5938,0.9375,0.625,0.9375,0.5938,1.0,0.5625,0.9375,0.5938,0.9375,0.5625,1.0,0.5313,0.9375,0.5625,0.9375,0.5313,1.0,0.5,0.9375,0.5313,0.9375,0.5,1.0,0.4688,0.9375,0.5,0.9375,0.4688,1.0,0.4375,0.9375,0.4688,0.9375,0.4375,1.0,0.4063,0.9375,0.4375,0.9375,0.4063,1.0,0.375,0.9375,0.4063,0.9375,0.375,1.0,0.3438,0.9375,0.375,0.9375,0.3438,1.0,0.3125,0.9375,0.3438,0.9375,0.3125,1.0,0.2813,0.9375,0.3125,0.9375,0.2813,1.0,0.25,0.9375,0.2813,0.9375,0.25,1.0,0.2188,0.9375,0.25,0.9375,0.2188,1.0,0.1875,0.9375,0.2188,0.9375,0.1875,1.0,0.1563,0.9375,0.1875,0.9375,0.1563,1.0,0.125,0.9375,0.1563,0.9375,0.125,1.0,0.0938,0.9375,0.125,0.9375,0.0938,1.0,0.0625,0.9375,0.0938,0.9375,0.0625,1.0,0.0313,0.9375,0.0625,0.9375,0.0313,1.0,0.0,0.9375,0.0313,0.9375,0.0,1.0,1.0,0.875,1.0,0.9375,0.9688,0.875,1.0,0.9375,0.9688,0.9375,0.9688,0.875,0.9688,0.875,0.9688,0.9375,0.9375,0.875,0.9688,0.9375,0.9375,0.9375,0.9375,0.875,0.9375,0.875,0.9375,0.9375,0.9063,0.875,0.9375,0.9375,0.9063,0.9375,0.9063,0.875,0.9063,0.875,0.9063,0.9375,0.875,0.875,0.9063,0.9375,0.875,0.9375,0.875,0.875,0.875,0.875,0.875,0.9375,0.8438,0.875,0.875,0.9375,0.8438,0.9375,0.8438,0.875,0.8438,0.875,0.8438,0.9375,0.8125,0.875,0.8438,0.9375,0.8125,0.9375,0.8125,0.875,0.8125,0.875,0.8125,0.9375,0.7813,0.875,0.8125,0.9375,0.7813,0.9375,0.7813,0.875,0.7813,0.875,0.7813,0.9375,0.75,0.875,0.7813,0.9375,0.75,0.9375,0.75,0.875,0.75,0.875,0.75,0.9375,0.7188,0.875,0.75,0.9375,0.7188,0.9375,0.7188,0.875,0.7188,0.875,0.7188,0.9375,0.6875,0.875,0.7188,0.9375,0.6875,0.9375,0.6875,0.875,0.6875,0.875,0.6875,0.9375,0.6563,0.875,0.6875,0.9375,0.6563,0.9375,0.6563,0.875,0.6563,0.875,0.6563,0.9375,0.625,0.875,0.6563,0.9375,0.625,0.9375,0.625,0.875,0.625,0.875,0.625,0.9375,0.5938,0.875,0.625,0.9375,0.5938,0.9375,0.5938,0.875,0.5938,0.875,0.5938,0.9375,0.5625,0.875,0.5938,0.9375,0.5625,0.9375,0.5625,0.875,0.5625,0.875,0.5625,0.9375,0.5313,0.875,0.5625,0.9375,0.5313,0.9375,0.5313,0.875,0.5313,0.875,0.5313,0.9375,0.5,0.875,0.5313,0.9375,0.5,0.9375,0.5,0.875,0.5,0.875,0.5,0.9375,0.4688,0.875,0.5,0.9375,0.4688,0.9375,0.4688,0.875,0.4688,0.875,0.4688,0.9375,0.4375,0.875,0.4688,0.9375,0.4375,0.9375,0.4375,0.875,0.4375,0.875,0.4375,0.9375,0.4063,0.875,0.4375,0.9375,0.4063,0.9375,0.4063,0.875,0.4063,0.875,0.4063,0.9375,0.375,0.875,0.4063,0.9375,0.375,0.9375,0.375,0.875,0.375,0.875,0.375,0.9375,0.3438,0.875,0.375,0.9375,0.3438,0.9375,0.3438,0.875,0.3438,0.875,0.3438,0.9375,0.3125,0.875,0.3438,0.9375,0.3125,0.9375,0.3125,0.875,0.3125,0.875,0.3125,0.9375,0.2813,0.875,0.3125,0.9375,0.2813,0.9375,0.2813,0.875,0.2813,0.875,0.2813,0.9375,0.25,0.875,0.2813,0.9375,0.25,0.9375,0.25,0.875,0.25,0.875,0.25,0.9375,0.2188,0.875,0.25,0.9375,0.2188,0.9375,0.2188,0.875,0.2188,0.875,0.2188,0.9375,0.1875,0.875,0.2188,0.9375,0.1875,0.9375,0.1875,0.875,0.1875,0.875,0.1875,0.9375,0.1563,0.875,0.1875,0.9375,0.1563,0.9375,0.1563,0.875,0.1563,0.875,0.1563,0.9375,0.125,0.875,0.1563,0.9375,0.125,0.9375,0.125,0.875,0.125,0.875,0.125,0.9375,0.0938,0.875,0.125,0.9375,0.0938,0.9375,0.0938,0.875,0.0938,0.875,0.0938,0.9375,0.0625,0.875,0.0938,0.9375,0.0625,0.9375,0.0625,0.875,0.0625,0.875,0.0625,0.9375,0.0313,0.875,0.0625,0.9375,0.0313,0.9375,0.0313,0.875,0.0313,0.875,0.0313,0.9375,0.0,0.875,0.0313,0.9375,0.0,0.9375,0.0,0.875,1.0,0.8125,1.0,0.875,0.9688,0.8125,1.0,0.875,0.9688,0.875,0.9688,0.8125,0.9688,0.8125,0.9688,0.875,0.9375,0.8125,0.9688,0.875,0.9375,0.875,0.9375,0.8125,0.9375,0.8125,0.9375,0.875,0.9063,0.8125,0.9375,0.875,0.9063,0.875,0.9063,0.8125,0.9063,0.8125,0.9063,0.875,0.875,0.8125,0.9063,0.875,0.875,0.875,0.875,0.8125,0.875,0.8125,0.875,0.875,0.8438,0.8125,0.875,0.875,0.8438,0.875,0.8438,0.8125,0.8438,0.8125,0.8438,0.875,0.8125,0.8125,0.8438,0.875,0.8125,0.875,0.8125,0.8125,0.8125,0.8125,0.8125,0.875,0.7813,0.8125,0.8125,0.875,0.7813,0.875,0.7813,0.8125,0.7813,0.8125,0.7813,0.875,0.75,0.8125,0.7813,0.875,0.75,0.875,0.75,0.8125,0.75,0.8125,0.75,0.875,0.7188,0.8125,0.75,0.875,0.7188,0.875,0.7188,0.8125,0.7188,0.8125,0.7188,0.875,0.6875,0.8125,0.7188,0.875,0.6875,0.875,0.6875,0.8125,0.6875,0.8125,0.6875,0.875,0.6563,0.8125,0.6875,0.875,0.6563,0.875,0.6563,0.8125,0.6563,0.8125,0.6563,0.875,0.625,0.8125,0.6563,0.875,0.625,0.875,0.625,0.8125,0.625,0.8125,0.625,0.875,0.5938,0.8125,0.625,0.875,0.5938,0.875,0.5938,0.8125,0.5938,0.8125,0.5938,0.875,0.5625,0.8125,0.5938,0.875,0.5625,0.875,0.5625,0.8125,0.5625,0.8125,0.5625,0.875,0.5313,0.8125,0.5625,0.875,0.5313,0.875,0.5313,0.8125,0.5313,0.8125,0.5313,0.875,0.5,0.8125,0.5313,0.875,0.5,0.875,0.5,0.8125,0.5,0.8125,0.5,0.875,0.4688,0.8125,0.5,0.875,0.4688,0.875,0.4688,0.8125,0.4688,0.8125,0.4688,0.875,0.4375,0.8125,0.4688,0.875,0.4375,0.875,0.4375,0.8125,0.4375,0.8125,0.4375,0.875,0.4063,0.8125,0.4375,0.875,0.4063,0.875,0.4063,0.8125,0.4063,0.8125,0.4063,0.875,0.375,0.8125,0.4063,0.875,0.375,0.875,0.375,0.8125,0.375,0.8125,0.375,0.875,0.3438,0.8125,0.375,0.875,0.3438,0.875,0.3438,0.8125,0.3438,0.8125,0.3438,0.875,0.3125,0.8125,0.3438,0.875,0.3125,0.875,0.3125,0.8125,0.3125,0.8125,0.3125,0.875,0.2813,0.8125,0.3125,0.875,0.2813,0.875,0.2813,0.8125,0.2813,0.8125,0.2813,0.875,0.25,0.8125,0.2813,0.875,0.25,0.875,0.25,0.8125,0.25,0.8125,0.25,0.875,0.2188,0.8125,0.25,0.875,0.2188,0.875,0.2188,0.8125,0.2188,0.8125,0.2188,0.875,0.1875,0.8125,0.2188,0.875,0.1875,0.875,0.1875,0.8125,0.1875,0.8125,0.1875,0.875,0.1563,0.8125,0.1875,0.875,0.1563,0.875,0.1563,0.8125,0.1563,0.8125,0.1563,0.875,0.125,0.8125,0.1563,0.875,0.125,0.875,0.125,0.8125,0.125,0.8125,0.125,0.875,0.0938,0.8125,0.125,0.875,0.0938,0.875,0.0938,0.8125,0.0938,0.8125,0.0938,0.875,0.0625,0.8125,0.0938,0.875,0.0625,0.875,0.0625,0.8125,0.0625,0.8125,0.0625,0.875,0.0313,0.8125,0.0625,0.875,0.0313,0.875,0.0313,0.8125,0.0313,0.8125,0.0313,0.875,0.0,0.8125,0.0313,0.875,0.0,0.875,0.0,0.8125,1.0,0.75,1.0,0.8125,0.9688,0.75,1.0,0.8125,0.9688,0.8125,0.9688,0.75,0.9688,0.75,0.9688,0.8125,0.9375,0.75,0.9688,0.8125,0.9375,0.8125,0.9375,0.75,0.9375,0.75,0.9375,0.8125,0.9063,0.75,0.9375,0.8125,0.9063,0.8125,0.9063,0.75,0.9063,0.75,0.9063,0.8125,0.875,0.75,0.9063,0.8125,0.875,0.8125,0.875,0.75,0.875,0.75,0.875,0.8125,0.8438,0.75,0.875,0.8125,0.8438,0.8125,0.8438,0.75,0.8438,0.75,0.8438,0.8125,0.8125,0.75,0.8438,0.8125,0.8125,0.8125,0.8125,0.75,0.8125,0.75,0.8125,0.8125,0.7813,0.75,0.8125,0.8125,0.7813,0.8125,0.7813,0.75,0.7813,0.75,0.7813,0.8125,0.75,0.75,0.7813,0.8125,0.75,0.8125,0.75,0.75,0.75,0.75,0.75,0.8125,0.7188,0.75,0.75,0.8125,0.7188,0.8125,0.7188,0.75,0.7188,0.75,0.7188,0.8125,0.6875,0.75,0.7188,0.8125,0.6875,0.8125,0.6875,0.75,0.6875,0.75,0.6875,0.8125,0.6563,0.75,0.6875,0.8125,0.6563,0.8125,0.6563,0.75,0.6563,0.75,0.6563,0.8125,0.625,0.75,0.6563,0.8125,0.625,0.8125,0.625,0.75,0.625,0.75,0.625,0.8125,0.5938,0.75,0.625,0.8125,0.5938,0.8125,0.5938,0.75,0.5938,0.75,0.5938,0.8125,0.5625,0.75,0.5938,0.8125,0.5625,0.8125,0.5625,0.75,0.5625,0.75,0.5625,0.8125,0.5313,0.75,0.5625,0.8125,0.5313,0.8125,0.5313,0.75,0.5313,0.75,0.5313,0.8125,0.5,0.75,0.5313,0.8125,0.5,0.8125,0.5,0.75,0.5,0.75,0.5,0.8125,0.4688,0.75,0.5,0.8125,0.4688,0.8125,0.4688,0.75,0.4688,0.75,0.4688,0.8125,0.4375,0.75,0.4688,0.8125,0.4375,0.8125,0.4375,0.75,0.4375,0.75,0.4375,0.8125,0.4063,0.75,0.4375,0.8125,0.4063,0.8125,0.4063,0.75,0.4063,0.75,0.4063,0.8125,0.375,0.75,0.4063,0.8125,0.375,0.8125,0.375,0.75,0.375,0.75,0.375,0.8125,0.3438,0.75,0.375,0.8125,0.3438,0.8125,0.3438,0.75,0.3438,0.75,0.3438,0.8125,0.3125,0.75,0.3438,0.8125,0.3125,0.8125,0.3125,0.75,0.3125,0.75,0.3125,0.8125,0.2813,0.75,0.3125,0.8125,0.2813,0.8125,0.2813,0.75,0.2813,0.75,0.2813,0.8125,0.25,0.75,0.2813,0.8125,0.25,0.8125,0.25,0.75,0.25,0.75,0.25,0.8125,0.2188,0.75,0.25,0.8125,0.2188,0.8125,0.2188,0.75,0.2188,0.75,0.2188,0.8125,0.1875,0.75,0.2188,0.8125,0.1875,0.8125,0.1875,0.75,0.1875,0.75,0.1875,0.8125,0.1563,0.75,0.1875,0.8125,0.1563,0.8125,0.1563,0.75,0.1563,0.75,0.1563,0.8125,0.125,0.75,0.1563,0.8125,0.125,0.8125,0.125,0.75,0.125,0.75,0.125,0.8125,0.0938,0.75,0.125,0.8125,0.0938,0.8125,0.0938,0.75,0.0938,0.75,0.0938,0.8125,0.0625,0.75,0.0938,0.8125,0.0625,0.8125,0.0625,0.75,0.0625,0.75,0.0625,0.8125,0.0313,0.75,0.0625,0.8125,0.0313,0.8125,0.0313,0.75,0.0313,0.75,0.0313,0.8125,0.0,0.75,0.0313,0.8125,0.0,0.8125,0.0,0.75,1.0,0.6875,1.0,0.75,0.9688,0.6875,1.0,0.75,0.9688,0.75,0.9688,0.6875,0.9688,0.6875,0.9688,0.75,0.9375,0.6875,0.9688,0.75,0.9375,0.75,0.9375,0.6875,0.9375,0.6875,0.9375,0.75,0.9063,0.6875,0.9375,0.75,0.9063,0.75,0.9063,0.6875,0.9063,0.6875,0.9063,0.75,0.875,0.6875,0.9063,0.75,0.875,0.75,0.875,0.6875,0.875,0.6875,0.875,0.75,0.8438,0.6875,0.875,0.75,0.8438,0.75,0.8438,0.6875,0.8438,0.6875,0.8438,0.75,0.8125,0.6875,0.8438,0.75,0.8125,0.75,0.8125,0.6875,0.8125,0.6875,0.8125,0.75,0.7813,0.6875,0.8125,0.75,0.7813,0.75,0.7813,0.6875,0.7813,0.6875,0.7813,0.75,0.75,0.6875,0.7813,0.75,0.75,0.75,0.75,0.6875,0.75,0.6875,0.75,0.75,0.7188,0.6875,0.75,0.75,0.7188,0.75,0.7188,0.6875,0.7188,0.6875,0.7188,0.75,0.6875,0.6875,0.7188,0.75,0.6875,0.75,0.6875,0.6875,0.6875,0.6875,0.6875,0.75,0.6563,0.6875,0.6875,0.75,0.6563,0.75,0.6563,0.6875,0.6563,0.6875,0.6563,0.75,0.625,0.6875,0.6563,0.75,0.625,0.75,0.625,0.6875,0.625,0.6875,0.625,0.75,0.5938,0.6875,0.625,0.75,0.5938,0.75,0.5938,0.6875,0.5938,0.6875,0.5938,0.75,0.5625,0.6875,0.5938,0.75,0.5625,0.75,0.5625,0.6875,0.5625,0.6875,0.5625,0.75,0.5313,0.6875,0.5625,0.75,0.5313,0.75,0.5313,0.6875,0.5313,0.6875,0.5313,0.75,0.5,0.6875,0.5313,0.75,0.5,0.75,0.5,0.6875,0.5,0.6875,0.5,0.75,0.4688,0.6875,0.5,0.75,0.4688,0.75,0.4688,0.6875,0.4688,0.6875,0.4688,0.75,0.4375,0.6875,0.4688,0.75,0.4375,0.75,0.4375,0.6875,0.4375,0.6875,0.4375,0.75,0.4063,0.6875,0.4375,0.75,0.4063,0.75,0.4063,0.6875,0.4063,0.6875,0.4063,0.75,0.375,0.6875,0.4063,0.75,0.375,0.75,0.375,0.6875,0.375,0.6875,0.375,0.75,0.3438,0.6875,0.375,0.75,0.3438,0.75,0.3438,0.6875,0.3438,0.6875,0.3438,0.75,0.3125,0.6875,0.3438,0.75,0.3125,0.75,0.3125,0.6875,0.3125,0.6875,0.3125,0.75,0.2813,0.6875,0.3125,0.75,0.2813,0.75,0.2813,0.6875,0.2813,0.6875,0.2813,0.75,0.25,0.6875,0.2813,0.75,0.25,0.75,0.25,0.6875,0.25,0.6875,0.25,0.75,0.2188,0.6875,0.25,0.75,0.2188,0.75,0.2188,0.6875,0.2188,0.6875,0.2188,0.75,0.1875,0.6875,0.2188,0.75,0.1875,0.75,0.1875,0.6875,0.1875,0.6875,0.1875,0.75,0.1563,0.6875,0.1875,0.75,0.1563,0.75,0.1563,0.6875,0.1563,0.6875,0.1563,0.75,0.125,0.6875,0.1563,0.75,0.125,0.75,0.125,0.6875,0.125,0.6875,0.125,0.75,0.0938,0.6875,0.125,0.75,0.0938,0.75,0.0938,0.6875,0.0938,0.6875,0.0938,0.75,0.0625,0.6875,0.0938,0.75,0.0625,0.75,0.0625,0.6875,0.0625,0.6875,0.0625,0.75,0.0313,0.6875,0.0625,0.75,0.0313,0.75,0.0313,0.6875,0.0313,0.6875,0.0313,0.75,0.0,0.6875,0.0313,0.75,0.0,0.75,0.0,0.6875,1.0,0.625,1.0,0.6875,0.9688,0.625,1.0,0.6875,0.9688,0.6875,0.9688,0.625,0.9688,0.625,0.9688,0.6875,0.9375,0.625,0.9688,0.6875,0.9375,0.6875,0.9375,0.625,0.9375,0.625,0.9375,0.6875,0.9063,0.625,0.9375,0.6875,0.9063,0.6875,0.9063,0.625,0.9063,0.625,0.9063,0.6875,0.875,0.625,0.9063,0.6875,0.875,0.6875,0.875,0.625,0.875,0.625,0.875,0.6875,0.8438,0.625,0.875,0.6875,0.8438,0.6875,0.8438,0.625,0.8438,0.625,0.8438,0.6875,0.8125,0.625,0.8438,0.6875,0.8125,0.6875,0.8125,0.625,0.8125,0.625,0.8125,0.6875,0.7813,0.625,0.8125,0.6875,0.7813,0.6875,0.7813,0.625,0.7813,0.625,0.7813,0.6875,0.75,0.625,0.7813,0.6875,0.75,0.6875,0.75,0.625,0.75,0.625,0.75,0.6875,0.7188,0.625,0.75,0.6875,0.7188,0.6875,0.7188,0.625,0.7188,0.625,0.7188,0.6875,0.6875,0.625,0.7188,0.6875,0.6875,0.6875,0.6875,0.625,0.6875,0.625,0.6875,0.6875,0.6563,0.625,0.6875,0.6875,0.6563,0.6875,0.6563,0.625,0.6563,0.625,0.6563,0.6875,0.625,0.625,0.6563,0.6875,0.625,0.6875,0.625,0.625,0.625,0.625,0.625,0.6875,0.5938,0.625,0.625,0.6875,0.5938,0.6875,0.5938,0.625,0.5938,0.625,0.5938,0.6875,0.5625,0.625,0.5938,0.6875,0.5625,0.6875,0.5625,0.625,0.5625,0.625,0.5625,0.6875,0.5313,0.625,0.5625,0.6875,0.5313,0.6875,0.5313,0.625,0.5313,0.625,0.5313,0.6875,0.5,0.625,0.5313,0.6875,0.5,0.6875,0.5,0.625,0.5,0.625,0.5,0.6875,0.4688,0.625,0.5,0.6875,0.4688,0.6875,0.4688,0.625,0.4688,0.625,0.4688,0.6875,0.4375,0.625,0.4688,0.6875,0.4375,0.6875,0.4375,0.625,0.4375,0.625,0.4375,0.6875,0.4063,0.625,0.4375,0.6875,0.4063,0.6875,0.4063,0.625,0.4063,0.625,0.4063,0.6875,0.375,0.625,0.4063,0.6875,0.375,0.6875,0.375,0.625,0.375,0.625,0.375,0.6875,0.3438,0.625,0.375,0.6875,0.3438,0.6875,0.3438,0.625,0.3438,0.625,0.3438,0.6875,0.3125,0.625,0.3438,0.6875,0.3125,0.6875,0.3125,0.625,0.3125,0.625,0.3125,0.6875,0.2813,0.625,0.3125,0.6875,0.2813,0.6875,0.2813,0.625,0.2813,0.625,0.2813,0.6875,0.25,0.625,0.2813,0.6875,0.25,0.6875,0.25,0.625,0.25,0.625,0.25,0.6875,0.2188,0.625,0.25,0.6875,0.2188,0.6875,0.2188,0.625,0.2188,0.625,0.2188,0.6875,0.1875,0.625,0.2188,0.6875,0.1875,0.6875,0.1875,0.625,0.1875,0.625,0.1875,0.6875,0.1563,0.625,0.1875,0.6875,0.1563,0.6875,0.1563,0.625,0.1563,0.625,0.1563,0.6875,0.125,0.625,0.1563,0.6875,0.125,0.6875,0.125,0.625,0.125,0.625,0.125,0.6875,0.0938,0.625,0.125,0.6875,0.0938,0.6875,0.0938,0.625,0.0938,0.625,0.0938,0.6875,0.0625,0.625,0.0938,0.6875,0.0625,0.6875,0.0625,0.625,0.0625,0.625,0.0625,0.6875,0.0313,0.625,0.0625,0.6875,0.0313,0.6875,0.0313,0.625,0.0313,0.625,0.0313,0.6875,0.0,0.625,0.0313,0.6875,0.0,0.6875,0.0,0.625,1.0,0.5625,1.0,0.625,0.9688,0.5625,1.0,0.625,0.9688,0.625,0.9688,0.5625,0.9688,0.5625,0.9688,0.625,0.9375,0.5625,0.9688,0.625,0.9375,0.625,0.9375,0.5625,0.9375,0.5625,0.9375,0.625,0.9063,0.5625,0.9375,0.625,0.9063,0.625,0.9063,0.5625,0.9063,0.5625,0.9063,0.625,0.875,0.5625,0.9063,0.625,0.875,0.625,0.875,0.5625,0.875,0.5625,0.875,0.625,0.8438,0.5625,0.875,0.625,0.8438,0.625,0.8438,0.5625,0.8438,0.5625,0.8438,0.625,0.8125,0.5625,0.8438,0.625,0.8125,0.625,0.8125,0.5625,0.8125,0.5625,0.8125,0.625,0.7813,0.5625,0.8125,0.625,0.7813,0.625,0.7813,0.5625,0.7813,0.5625,0.7813,0.625,0.75,0.5625,0.7813,0.625,0.75,0.625,0.75,0.5625,0.75,0.5625,0.75,0.625,0.7188,0.5625,0.75,0.625,0.7188,0.625,0.7188,0.5625,0.7188,0.5625,0.7188,0.625,0.6875,0.5625,0.7188,0.625,0.6875,0.625,0.6875,0.5625,0.6875,0.5625,0.6875,0.625,0.6563,0.5625,0.6875,0.625,0.6563,0.625,0.6563,0.5625,0.6563,0.5625,0.6563,0.625,0.625,0.5625,0.6563,0.625,0.625,0.625,0.625,0.5625,0.625,0.5625,0.625,0.625,0.5938,0.5625,0.625,0.625,0.5938,0.625,0.5938,0.5625,0.5938,0.5625,0.5938,0.625,0.5625,0.5625,0.5938,0.625,0.5625,0.625,0.5625,0.5625,0.5625,0.5625,0.5625,0.625,0.5313,0.5625,0.5625,0.625,0.5313,0.625,0.5313,0.5625,0.5313,0.5625,0.5313,0.625,0.5,0.5625,0.5313,0.625,0.5,0.625,0.5,0.5625,0.5,0.5625,0.5,0.625,0.4688,0.5625,0.5,0.625,0.4688,0.625,0.4688,0.5625,0.4688,0.5625,0.4688,0.625,0.4375,0.5625,0.4688,0.625,0.4375,0.625,0.4375,0.5625,0.4375,0.5625,0.4375,0.625,0.4063,0.5625,0.4375,0.625,0.4063,0.625,0.4063,0.5625,0.4063,0.5625,0.4063,0.625,0.375,0.5625,0.4063,0.625,0.375,0.625,0.375,0.5625,0.375,0.5625,0.375,0.625,0.3438,0.5625,0.375,0.625,0.3438,0.625,0.3438,0.5625,0.3438,0.5625,0.3438,0.625,0.3125,0.5625,0.3438,0.625,0.3125,0.625,0.3125,0.5625,0.3125,0.5625,0.3125,0.625,0.2813,0.5625,0.3125,0.625,0.2813,0.625,0.2813,0.5625,0.2813,0.5625,0.2813,0.625,0.25,0.5625,0.2813,0.625,0.25,0.625,0.25,0.5625,0.25,0.5625,0.25,0.625,0.2188,0.5625,0.25,0.625,0.2188,0.625,0.2188,0.5625,0.2188,0.5625,0.2188,0.625,0.1875,0.5625,0.2188,0.625,0.1875,0.625,0.1875,0.5625,0.1875,0.5625,0.1875,0.625,0.1563,0.5625,0.1875,0.625,0.1563,0.625,0.1563,0.5625,0.1563,0.5625,0.1563,0.625,0.125,0.5625,0.1563,0.625,0.125,0.625,0.125,0.5625,0.125,0.5625,0.125,0.625,0.0938,0.5625,0.125,0.625,0.0938,0.625,0.0938,0.5625,0.0938,0.5625,0.0938,0.625,0.0625,0.5625,0.0938,0.625,0.0625,0.625,0.0625,0.5625,0.0625,0.5625,0.0625,0.625,0.0313,0.5625,0.0625,0.625,0.0313,0.625,0.0313,0.5625,0.0313,0.5625,0.0313,0.625,0.0,0.5625,0.0313,0.625,0.0,0.625,0.0,0.5625,1.0,0.5,1.0,0.5625,0.9688,0.5,1.0,0.5625,0.9688,0.5625,0.9688,0.5,0.9688,0.5,0.9688,0.5625,0.9375,0.5,0.9688,0.5625,0.9375,0.5625,0.9375,0.5,0.9375,0.5,0.9375,0.5625,0.9063,0.5,0.9375,0.5625,0.9063,0.5625,0.9063,0.5,0.9063,0.5,0.9063,0.5625,0.875,0.5,0.9063,0.5625,0.875,0.5625,0.875,0.5,0.875,0.5,0.875,0.5625,0.8438,0.5,0.875,0.5625,0.8438,0.5625,0.8438,0.5,0.8438,0.5,0.8438,0.5625,0.8125,0.5,0.8438,0.5625,0.8125,0.5625,0.8125,0.5,0.8125,0.5,0.8125,0.5625,0.7813,0.5,0.8125,0.5625,0.7813,0.5625,0.7813,0.5,0.7813,0.5,0.7813,0.5625,0.75,0.5,0.7813,0.5625,0.75,0.5625,0.75,0.5,0.75,0.5,0.75,0.5625,0.7188,0.5,0.75,0.5625,0.7188,0.5625,0.7188,0.5,0.7188,0.5,0.7188,0.5625,0.6875,0.5,0.7188,0.5625,0.6875,0.5625,0.6875,0.5,0.6875,0.5,0.6875,0.5625,0.6563,0.5,0.6875,0.5625,0.6563,0.5625,0.6563,0.5,0.6563,0.5,0.6563,0.5625,0.625,0.5,0.6563,0.5625,0.625,0.5625,0.625,0.5,0.625,0.5,0.625,0.5625,0.5938,0.5,0.625,0.5625,0.5938,0.5625,0.5938,0.5,0.5938,0.5,0.5938,0.5625,0.5625,0.5,0.5938,0.5625,0.5625,0.5625,0.5625,0.5,0.5625,0.5,0.5625,0.5625,0.5313,0.5,0.5625,0.5625,0.5313,0.5625,0.5313,0.5,0.5313,0.5,0.5313,0.5625,0.5,0.5,0.5313,0.5625,0.5,0.5625,0.5,0.5,0.5,0.5,0.5,0.5625,0.4688,0.5,0.5,0.5625,0.4688,0.5625,0.4688,0.5,0.4688,0.5,0.4688,0.5625,0.4375,0.5,0.4688,0.5625,0.4375,0.5625,0.4375,0.5,0.4375,0.5,0.4375,0.5625,0.4063,0.5,0.4375,0.5625,0.4063,0.5625,0.4063,0.5,0.4063,0.5,0.4063,0.5625,0.375,0.5,0.4063,0.5625,0.375,0.5625,0.375,0.5,0.375,0.5,0.375,0.5625,0.3438,0.5,0.375,0.5625,0.3438,0.5625,0.3438,0.5,0.3438,0.5,0.3438,0.5625,0.3125,0.5,0.3438,0.5625,0.3125,0.5625,0.3125,0.5,0.3125,0.5,0.3125,0.5625,0.2813,0.5,0.3125,0.5625,0.2813,0.5625,0.2813,0.5,0.2813,0.5,0.2813,0.5625,0.25,0.5,0.2813,0.5625,0.25,0.5625,0.25,0.5,0.25,0.5,0.25,0.5625,0.2188,0.5,0.25,0.5625,0.2188,0.5625,0.2188,0.5,0.2188,0.5,0.2188,0.5625,0.1875,0.5,0.2188,0.5625,0.1875,0.5625,0.1875,0.5,0.1875,0.5,0.1875,0.5625,0.1563,0.5,0.1875,0.5625,0.1563,0.5625,0.1563,0.5,0.1563,0.5,0.1563,0.5625,0.125,0.5,0.1563,0.5625,0.125,0.5625,0.125,0.5,0.125,0.5,0.125,0.5625,0.0938,0.5,0.125,0.5625,0.0938,0.5625,0.0938,0.5,0.0938,0.5,0.0938,0.5625,0.0625,0.5,0.0938,0.5625,0.0625,0.5625,0.0625,0.5,0.0625,0.5,0.0625,0.5625,0.0313,0.5,0.0625,0.5625,0.0313,0.5625,0.0313,0.5,0.0313,0.5,0.0313,0.5625,0.0,0.5,0.0313,0.5625,0.0,0.5625,0.0,0.5,1.0,0.4375,1.0,0.5,0.9688,0.4375,1.0,0.5,0.9688,0.5,0.9688,0.4375,0.9688,0.4375,0.9688,0.5,0.9375,0.4375,0.9688,0.5,0.9375,0.5,0.9375,0.4375,0.9375,0.4375,0.9375,0.5,0.9063,0.4375,0.9375,0.5,0.9063,0.5,0.9063,0.4375,0.9063,0.4375,0.9063,0.5,0.875,0.4375,0.9063,0.5,0.875,0.5,0.875,0.4375,0.875,0.4375,0.875,0.5,0.8438,0.4375,0.875,0.5,0.8438,0.5,0.8438,0.4375,0.8438,0.4375,0.8438,0.5,0.8125,0.4375,0.8438,0.5,0.8125,0.5,0.8125,0.4375,0.8125,0.4375,0.8125,0.5,0.7813,0.4375,0.8125,0.5,0.7813,0.5,0.7813,0.4375,0.7813,0.4375,0.7813,0.5,0.75,0.4375,0.7813,0.5,0.75,0.5,0.75,0.4375,0.75,0.4375,0.75,0.5,0.7188,0.4375,0.75,0.5,0.7188,0.5,0.7188,0.4375,0.7188,0.4375,0.7188,0.5,0.6875,0.4375,0.7188,0.5,0.6875,0.5,0.6875,0.4375,0.6875,0.4375,0.6875,0.5,0.6563,0.4375,0.6875,0.5,0.6563,0.5,0.6563,0.4375,0.6563,0.4375,0.6563,0.5,0.625,0.4375,0.6563,0.5,0.625,0.5,0.625,0.4375,0.625,0.4375,0.625,0.5,0.5938,0.4375,0.625,0.5,0.5938,0.5,0.5938,0.4375,0.5938,0.4375,0.5938,0.5,0.5625,0.4375,0.5938,0.5,0.5625,0.5,0.5625,0.4375,0.5625,0.4375,0.5625,0.5,0.5313,0.4375,0.5625,0.5,0.5313,0.5,0.5313,0.4375,0.5313,0.4375,0.5313,0.5,0.5,0.4375,0.5313,0.5,0.5,0.5,0.5,0.4375,0.5,0.4375,0.5,0.5,0.4688,0.4375,0.5,0.5,0.4688,0.5,0.4688,0.4375,0.4688,0.4375,0.4688,0.5,0.4375,0.4375,0.4688,0.5,0.4375,0.5,0.4375,0.4375,0.4375,0.4375,0.4375,0.5,0.4063,0.4375,0.4375,0.5,0.4063,0.5,0.4063,0.4375,0.4063,0.4375,0.4063,0.5,0.375,0.4375,0.4063,0.5,0.375,0.5,0.375,0.4375,0.375,0.4375,0.375,0.5,0.3438,0.4375,0.375,0.5,0.3438,0.5,0.3438,0.4375,0.3438,0.4375,0.3438,0.5,0.3125,0.4375,0.3438,0.5,0.3125,0.5,0.3125,0.4375,0.3125,0.4375,0.3125,0.5,0.2813,0.4375,0.3125,0.5,0.2813,0.5,0.2813,0.4375,0.2813,0.4375,0.2813,0.5,0.25,0.4375,0.2813,0.5,0.25,0.5,0.25,0.4375,0.25,0.4375,0.25,0.5,0.2188,0.4375,0.25,0.5,0.2188,0.5,0.2188,0.4375,0.2188,0.4375,0.2188,0.5,0.1875,0.4375,0.2188,0.5,0.1875,0.5,0.1875,0.4375,0.1875,0.4375,0.1875,0.5,0.1563,0.4375,0.1875,0.5,0.1563,0.5,0.1563,0.4375,0.1563,0.4375,0.1563,0.5,0.125,0.4375,0.1563,0.5,0.125,0.5,0.125,0.4375,0.125,0.4375,0.125,0.5,0.0938,0.4375,0.125,0.5,0.0938,0.5,0.0938,0.4375,0.0938,0.4375,0.0938,0.5,0.0625,0.4375,0.0938,0.5,0.0625,0.5,0.0625,0.4375,0.0625,0.4375,0.0625,0.5,0.0313,0.4375,0.0625,0.5,0.0313,0.5,0.0313,0.4375,0.0313,0.4375,0.0313,0.5,0.0,0.4375,0.0313,0.5,0.0,0.5,0.0,0.4375,1.0,0.375,1.0,0.4375,0.9688,0.375,1.0,0.4375,0.9688,0.4375,0.9688,0.375,0.9688,0.375,0.9688,0.4375,0.9375,0.375,0.9688,0.4375,0.9375,0.4375,0.9375,0.375,0.9375,0.375,0.9375,0.4375,0.9063,0.375,0.9375,0.4375,0.9063,0.4375,0.9063,0.375,0.9063,0.375,0.9063,0.4375,0.875,0.375,0.9063,0.4375,0.875,0.4375,0.875,0.375,0.875,0.375,0.875,0.4375,0.8438,0.375,0.875,0.4375,0.8438,0.4375,0.8438,0.375,0.8438,0.375,0.8438,0.4375,0.8125,0.375,0.8438,0.4375,0.8125,0.4375,0.8125,0.375,0.8125,0.375,0.8125,0.4375,0.7813,0.375,0.8125,0.4375,0.7813,0.4375,0.7813,0.375,0.7813,0.375,0.7813,0.4375,0.75,0.375,0.7813,0.4375,0.75,0.4375,0.75,0.375,0.75,0.375,0.75,0.4375,0.7188,0.375,0.75,0.4375,0.7188,0.4375,0.7188,0.375,0.7188,0.375,0.7188,0.4375,0.6875,0.375,0.7188,0.4375,0.6875,0.4375,0.6875,0.375,0.6875,0.375,0.6875,0.4375,0.6563,0.375,0.6875,0.4375,0.6563,0.4375,0.6563,0.375,0.6563,0.375,0.6563,0.4375,0.625,0.375,0.6563,0.4375,0.625,0.4375,0.625,0.375,0.625,0.375,0.625,0.4375,0.5938,0.375,0.625,0.4375,0.5938,0.4375,0.5938,0.375,0.5938,0.375,0.5938,0.4375,0.5625,0.375,0.5938,0.4375,0.5625,0.4375,0.5625,0.375,0.5625,0.375,0.5625,0.4375,0.5313,0.375,0.5625,0.4375,0.5313,0.4375,0.5313,0.375,0.5313,0.375,0.5313,0.4375,0.5,0.375,0.5313,0.4375,0.5,0.4375,0.5,0.375,0.5,0.375,0.5,0.4375,0.4688,0.375,0.5,0.4375,0.4688,0.4375,0.4688,0.375,0.4688,0.375,0.4688,0.4375,0.4375,0.375,0.4688,0.4375,0.4375,0.4375,0.4375,0.375,0.4375,0.375,0.4375,0.4375,0.4063,0.375,0.4375,0.4375,0.4063,0.4375,0.4063,0.375,0.4063,0.375,0.4063,0.4375,0.375,0.375,0.4063,0.4375,0.375,0.4375,0.375,0.375,0.375,0.375,0.375,0.4375,0.3438,0.375,0.375,0.4375,0.3438,0.4375,0.3438,0.375,0.3438,0.375,0.3438,0.4375,0.3125,0.375,0.3438,0.4375,0.3125,0.4375,0.3125,0.375,0.3125,0.375,0.3125,0.4375,0.2813,0.375,0.3125,0.4375,0.2813,0.4375,0.2813,0.375,0.2813,0.375,0.2813,0.4375,0.25,0.375,0.2813,0.4375,0.25,0.4375,0.25,0.375,0.25,0.375,0.25,0.4375,0.2188,0.375,0.25,0.4375,0.2188,0.4375,0.2188,0.375,0.2188,0.375,0.2188,0.4375,0.1875,0.375,0.2188,0.4375,0.1875,0.4375,0.1875,0.375,0.1875,0.375,0.1875,0.4375,0.1563,0.375,0.1875,0.4375,0.1563,0.4375,0.1563,0.375,0.1563,0.375,0.1563,0.4375,0.125,0.375,0.1563,0.4375,0.125,0.4375,0.125,0.375,0.125,0.375,0.125,0.4375,0.0938,0.375,0.125,0.4375,0.0938,0.4375,0.0938,0.375,0.0938,0.375,0.0938,0.4375,0.0625,0.375,0.0938,0.4375,0.0625,0.4375,0.0625,0.375,0.0625,0.375,0.0625,0.4375,0.0313,0.375,0.0625,0.4375,0.0313,0.4375,0.0313,0.375,0.0313,0.375,0.0313,0.4375,0.0,0.375,0.0313,0.4375,0.0,0.4375,0.0,0.375,1.0,0.3125,1.0,0.375,0.9688,0.3125,1.0,0.375,0.9688,0.375,0.9688,0.3125,0.9688,0.3125,0.9688,0.375,0.9375,0.3125,0.9688,0.375,0.9375,0.375,0.9375,0.3125,0.9375,0.3125,0.9375,0.375,0.9063,0.3125,0.9375,0.375,0.9063,0.375,0.9063,0.3125,0.9063,0.3125,0.9063,0.375,0.875,0.3125,0.9063,0.375,0.875,0.375,0.875,0.3125,0.875,0.3125,0.875,0.375,0.8438,0.3125,0.875,0.375,0.8438,0.375,0.8438,0.3125,0.8438,0.3125,0.8438,0.375,0.8125,0.3125,0.8438,0.375,0.8125,0.375,0.8125,0.3125,0.8125,0.3125,0.8125,0.375,0.7813,0.3125,0.8125,0.375,0.7813,0.375,0.7813,0.3125,0.7813,0.3125,0.7813,0.375,0.75,0.3125,0.7813,0.375,0.75,0.375,0.75,0.3125,0.75,0.3125,0.75,0.375,0.7188,0.3125,0.75,0.375,0.7188,0.375,0.7188,0.3125,0.7188,0.3125,0.7188,0.375,0.6875,0.3125,0.7188,0.375,0.6875,0.375,0.6875,0.3125,0.6875,0.3125,0.6875,0.375,0.6563,0.3125,0.6875,0.375,0.6563,0.375,0.6563,0.3125,0.6563,0.3125,0.6563,0.375,0.625,0.3125,0.6563,0.375,0.625,0.375,0.625,0.3125,0.625,0.3125,0.625,0.375,0.5938,0.3125,0.625,0.375,0.5938,0.375,0.5938,0.3125,0.5938,0.3125,0.5938,0.375,0.5625,0.3125,0.5938,0.375,0.5625,0.375,0.5625,0.3125,0.5625,0.3125,0.5625,0.375,0.5313,0.3125,0.5625,0.375,0.5313,0.375,0.5313,0.3125,0.5313,0.3125,0.5313,0.375,0.5,0.3125,0.5313,0.375,0.5,0.375,0.5,0.3125,0.5,0.3125,0.5,0.375,0.4688,0.3125,0.5,0.375,0.4688,0.375,0.4688,0.3125,0.4688,0.3125,0.4688,0.375,0.4375,0.3125,0.4688,0.375,0.4375,0.375,0.4375,0.3125,0.4375,0.3125,0.4375,0.375,0.4063,0.3125,0.4375,0.375,0.4063,0.375,0.4063,0.3125,0.4063,0.3125,0.4063,0.375,0.375,0.3125,0.4063,0.375,0.375,0.375,0.375,0.3125,0.375,0.3125,0.375,0.375,0.3438,0.3125,0.375,0.375,0.3438,0.375,0.3438,0.3125,0.3438,0.3125,0.3438,0.375,0.3125,0.3125,0.3438,0.375,0.3125,0.375,0.3125,0.3125,0.3125,0.3125,0.3125,0.375,0.2813,0.3125,0.3125,0.375,0.2813,0.375,0.2813,0.3125,0.2813,0.3125,0.2813,0.375,0.25,0.3125,0.2813,0.375,0.25,0.375,0.25,0.3125,0.25,0.3125,0.25,0.375,0.2188,0.3125,0.25,0.375,0.2188,0.375,0.2188,0.3125,0.2188,0.3125,0.2188,0.375,0.1875,0.3125,0.2188,0.375,0.1875,0.375,0.1875,0.3125,0.1875,0.3125,0.1875,0.375,0.1563,0.3125,0.1875,0.375,0.1563,0.375,0.1563,0.3125,0.1563,0.3125,0.1563,0.375,0.125,0.3125,0.1563,0.375,0.125,0.375,0.125,0.3125,0.125,0.3125,0.125,0.375,0.0938,0.3125,0.125,0.375,0.0938,0.375,0.0938,0.3125,0.0938,0.3125,0.0938,0.375,0.0625,0.3125,0.0938,0.375,0.0625,0.375,0.0625,0.3125,0.0625,0.3125,0.0625,0.375,0.0313,0.3125,0.0625,0.375,0.0313,0.375,0.0313,0.3125,0.0313,0.3125,0.0313,0.375,0.0,0.3125,0.0313,0.375,0.0,0.375,0.0,0.3125,1.0,0.25,1.0,0.3125,0.9688,0.25,1.0,0.3125,0.9688,0.3125,0.9688,0.25,0.9688,0.25,0.9688,0.3125,0.9375,0.25,0.9688,0.3125,0.9375,0.3125,0.9375,0.25,0.9375,0.25,0.9375,0.3125,0.9063,0.25,0.9375,0.3125,0.9063,0.3125,0.9063,0.25,0.9063,0.25,0.9063,0.3125,0.875,0.25,0.9063,0.3125,0.875,0.3125,0.875,0.25,0.875,0.25,0.875,0.3125,0.8438,0.25,0.875,0.3125,0.8438,0.3125,0.8438,0.25,0.8438,0.25,0.8438,0.3125,0.8125,0.25,0.8438,0.3125,0.8125,0.3125,0.8125,0.25,0.8125,0.25,0.8125,0.3125,0.7813,0.25,0.8125,0.3125,0.7813,0.3125,0.7813,0.25,0.7813,0.25,0.7813,0.3125,0.75,0.25,0.7813,0.3125,0.75,0.3125,0.75,0.25,0.75,0.25,0.75,0.3125,0.7188,0.25,0.75,0.3125,0.7188,0.3125,0.7188,0.25,0.7188,0.25,0.7188,0.3125,0.6875,0.25,0.7188,0.3125,0.6875,0.3125,0.6875,0.25,0.6875,0.25,0.6875,0.3125,0.6563,0.25,0.6875,0.3125,0.6563,0.3125,0.6563,0.25,0.6563,0.25,0.6563,0.3125,0.625,0.25,0.6563,0.3125,0.625,0.3125,0.625,0.25,0.625,0.25,0.625,0.3125,0.5938,0.25,0.625,0.3125,0.5938,0.3125,0.5938,0.25,0.5938,0.25,0.5938,0.3125,0.5625,0.25,0.5938,0.3125,0.5625,0.3125,0.5625,0.25,0.5625,0.25,0.5625,0.3125,0.5313,0.25,0.5625,0.3125,0.5313,0.3125,0.5313,0.25,0.5313,0.25,0.5313,0.3125,0.5,0.25,0.5313,0.3125,0.5,0.3125,0.5,0.25,0.5,0.25,0.5,0.3125,0.4688,0.25,0.5,0.3125,0.4688,0.3125,0.4688,0.25,0.4688,0.25,0.4688,0.3125,0.4375,0.25,0.4688,0.3125,0.4375,0.3125,0.4375,0.25,0.4375,0.25,0.4375,0.3125,0.4063,0.25,0.4375,0.3125,0.4063,0.3125,0.4063,0.25,0.4063,0.25,0.4063,0.3125,0.375,0.25,0.4063,0.3125,0.375,0.3125,0.375,0.25,0.375,0.25,0.375,0.3125,0.3438,0.25,0.375,0.3125,0.3438,0.3125,0.3438,0.25,0.3438,0.25,0.3438,0.3125,0.3125,0.25,0.3438,0.3125,0.3125,0.3125,0.3125,0.25,0.3125,0.25,0.3125,0.3125,0.2813,0.25,0.3125,0.3125,0.2813,0.3125,0.2813,0.25,0.2813,0.25,0.2813,0.3125,0.25,0.25,0.2813,0.3125,0.25,0.3125,0.25,0.25,0.25,0.25,0.25,0.3125,0.2188,0.25,0.25,0.3125,0.2188,0.3125,0.2188,0.25,0.2188,0.25,0.2188,0.3125,0.1875,0.25,0.2188,0.3125,0.1875,0.3125,0.1875,0.25,0.1875,0.25,0.1875,0.3125,0.1563,0.25,0.1875,0.3125,0.1563,0.3125,0.1563,0.25,0.1563,0.25,0.1563,0.3125,0.125,0.25,0.1563,0.3125,0.125,0.3125,0.125,0.25,0.125,0.25,0.125,0.3125,0.0938,0.25,0.125,0.3125,0.0938,0.3125,0.0938,0.25,0.0938,0.25,0.0938,0.3125,0.0625,0.25,0.0938,0.3125,0.0625,0.3125,0.0625,0.25,0.0625,0.25,0.0625,0.3125,0.0313,0.25,0.0625,0.3125,0.0313,0.3125,0.0313,0.25,0.0313,0.25,0.0313,0.3125,0.0,0.25,0.0313,0.3125,0.0,0.3125,0.0,0.25,1.0,0.1875,1.0,0.25,0.9688,0.1875,1.0,0.25,0.9688,0.25,0.9688,0.1875,0.9688,0.1875,0.9688,0.25,0.9375,0.1875,0.9688,0.25,0.9375,0.25,0.9375,0.1875,0.9375,0.1875,0.9375,0.25,0.9063,0.1875,0.9375,0.25,0.9063,0.25,0.9063,0.1875,0.9063,0.1875,0.9063,0.25,0.875,0.1875,0.9063,0.25,0.875,0.25,0.875,0.1875,0.875,0.1875,0.875,0.25,0.8438,0.1875,0.875,0.25,0.8438,0.25,0.8438,0.1875,0.8438,0.1875,0.8438,0.25,0.8125,0.1875,0.8438,0.25,0.8125,0.25,0.8125,0.1875,0.8125,0.1875,0.8125,0.25,0.7813,0.1875,0.8125,0.25,0.7813,0.25,0.7813,0.1875,0.7813,0.1875,0.7813,0.25,0.75,0.1875,0.7813,0.25,0.75,0.25,0.75,0.1875,0.75,0.1875,0.75,0.25,0.7188,0.1875,0.75,0.25,0.7188,0.25,0.7188,0.1875,0.7188,0.1875,0.7188,0.25,0.6875,0.1875,0.7188,0.25,0.6875,0.25,0.6875,0.1875,0.6875,0.1875,0.6875,0.25,0.6563,0.1875,0.6875,0.25,0.6563,0.25,0.6563,0.1875,0.6563,0.1875,0.6563,0.25,0.625,0.1875,0.6563,0.25,0.625,0.25,0.625,0.1875,0.625,0.1875,0.625,0.25,0.5938,0.1875,0.625,0.25,0.5938,0.25,0.5938,0.1875,0.5938,0.1875,0.5938,0.25,0.5625,0.1875,0.5938,0.25,0.5625,0.25,0.5625,0.1875,0.5625,0.1875,0.5625,0.25,0.5313,0.1875,0.5625,0.25,0.5313,0.25,0.5313,0.1875,0.5313,0.1875,0.5313,0.25,0.5,0.1875,0.5313,0.25,0.5,0.25,0.5,0.1875,0.5,0.1875,0.5,0.25,0.4688,0.1875,0.5,0.25,0.4688,0.25,0.4688,0.1875,0.4688,0.1875,0.4688,0.25,0.4375,0.1875,0.4688,0.25,0.4375,0.25,0.4375,0.1875,0.4375,0.1875,0.4375,0.25,0.4063,0.1875,0.4375,0.25,0.4063,0.25,0.4063,0.1875,0.4063,0.1875,0.4063,0.25,0.375,0.1875,0.4063,0.25,0.375,0.25,0.375,0.1875,0.375,0.1875,0.375,0.25,0.3438,0.1875,0.375,0.25,0.3438,0.25,0.3438,0.1875,0.3438,0.1875,0.3438,0.25,0.3125,0.1875,0.3438,0.25,0.3125,0.25,0.3125,0.1875,0.3125,0.1875,0.3125,0.25,0.2813,0.1875,0.3125,0.25,0.2813,0.25,0.2813,0.1875,0.2813,0.1875,0.2813,0.25,0.25,0.1875,0.2813,0.25,0.25,0.25,0.25,0.1875,0.25,0.1875,0.25,0.25,0.2188,0.1875,0.25,0.25,0.2188,0.25,0.2188,0.1875,0.2188,0.1875,0.2188,0.25,0.1875,0.1875,0.2188,0.25,0.1875,0.25,0.1875,0.1875,0.1875,0.1875,0.1875,0.25,0.1563,0.1875,0.1875,0.25,0.1563,0.25,0.1563,0.1875,0.1563,0.1875,0.1563,0.25,0.125,0.1875,0.1563,0.25,0.125,0.25,0.125,0.1875,0.125,0.1875,0.125,0.25,0.0938,0.1875,0.125,0.25,0.0938,0.25,0.0938,0.1875,0.0938,0.1875,0.0938,0.25,0.0625,0.1875,0.0938,0.25,0.0625,0.25,0.0625,0.1875,0.0625,0.1875,0.0625,0.25,0.0313,0.1875,0.0625,0.25,0.0313,0.25,0.0313,0.1875,0.0313,0.1875,0.0313,0.25,0.0,0.1875,0.0313,0.25,0.0,0.25,0.0,0.1875,1.0,0.125,1.0,0.1875,0.9688,0.125,1.0,0.1875,0.9688,0.1875,0.9688,0.125,0.9688,0.125,0.9688,0.1875,0.9375,0.125,0.9688,0.1875,0.9375,0.1875,0.9375,0.125,0.9375,0.125,0.9375,0.1875,0.9063,0.125,0.9375,0.1875,0.9063,0.1875,0.9063,0.125,0.9063,0.125,0.9063,0.1875,0.875,0.125,0.9063,0.1875,0.875,0.1875,0.875,0.125,0.875,0.125,0.875,0.1875,0.8438,0.125,0.875,0.1875,0.8438,0.1875,0.8438,0.125,0.8438,0.125,0.8438,0.1875,0.8125,0.125,0.8438,0.1875,0.8125,0.1875,0.8125,0.125,0.8125,0.125,0.8125,0.1875,0.7813,0.125,0.8125,0.1875,0.7813,0.1875,0.7813,0.125,0.7813,0.125,0.7813,0.1875,0.75,0.125,0.7813,0.1875,0.75,0.1875,0.75,0.125,0.75,0.125,0.75,0.1875,0.7188,0.125,0.75,0.1875,0.7188,0.1875,0.7188,0.125,0.7188,0.125,0.7188,0.1875,0.6875,0.125,0.7188,0.1875,0.6875,0.1875,0.6875,0.125,0.6875,0.125,0.6875,0.1875,0.6563,0.125,0.6875,0.1875,0.6563,0.1875,0.6563,0.125,0.6563,0.125,0.6563,0.1875,0.625,0.125,0.6563,0.1875,0.625,0.1875,0.625,0.125,0.625,0.125,0.625,0.1875,0.5938,0.125,0.625,0.1875,0.5938,0.1875,0.5938,0.125,0.5938,0.125,0.5938,0.1875,0.5625,0.125,0.5938,0.1875,0.5625,0.1875,0.5625,0.125,0.5625,0.125,0.5625,0.1875,0.5313,0.125,0.5625,0.1875,0.5313,0.1875,0.5313,0.125,0.5313,0.125,0.5313,0.1875,0.5,0.125,0.5313,0.1875,0.5,0.1875,0.5,0.125,0.5,0.125,0.5,0.1875,0.4688,0.125,0.5,0.1875,0.4688,0.1875,0.4688,0.125,0.4688,0.125,0.4688,0.1875,0.4375,0.125,0.4688,0.1875,0.4375,0.1875,0.4375,0.125,0.4375,0.125,0.4375,0.1875,0.4063,0.125,0.4375,0.1875,0.4063,0.1875,0.4063,0.125,0.4063,0.125,0.4063,0.1875,0.375,0.125,0.4063,0.1875,0.375,0.1875,0.375,0.125,0.375,0.125,0.375,0.1875,0.3438,0.125,0.375,0.1875,0.3438,0.1875,0.3438,0.125,0.3438,0.125,0.3438,0.1875,0.3125,0.125,0.3438,0.1875,0.3125,0.1875,0.3125,0.125,0.3125,0.125,0.3125,0.1875,0.2813,0.125,0.3125,0.1875,0.2813,0.1875,0.2813,0.125,0.2813,0.125,0.2813,0.1875,0.25,0.125,0.2813,0.1875,0.25,0.1875,0.25,0.125,0.25,0.125,0.25,0.1875,0.2188,0.125,0.25,0.1875,0.2188,0.1875,0.2188,0.125,0.2188,0.125,0.2188,0.1875,0.1875,0.125,0.2188,0.1875,0.1875,0.1875,0.1875,0.125,0.1875,0.125,0.1875,0.1875,0.1563,0.125,0.1875,0.1875,0.1563,0.1875,0.1563,0.125,0.1563,0.125,0.1563,0.1875,0.125,0.125,0.1563,0.1875,0.125,0.1875,0.125,0.125,0.125,0.125,0.125,0.1875,0.0938,0.125,0.125,0.1875,0.0938,0.1875,0.0938,0.125,0.0938,0.125,0.0938,0.1875,0.0625,0.125,0.0938,0.1875,0.0625,0.1875,0.0625,0.125,0.0625,0.125,0.0625,0.1875,0.0313,0.125,0.0625,0.1875,0.0313,0.1875,0.0313,0.125,0.0313,0.125,0.0313,0.1875,0.0,0.125,0.0313,0.1875,0.0,0.1875,0.0,0.125,1.0,0.0625,1.0,0.125,0.9688,0.0625,1.0,0.125,0.9688,0.125,0.9688,0.0625,0.9688,0.0625,0.9688,0.125,0.9375,0.0625,0.9688,0.125,0.9375,0.125,0.9375,0.0625,0.9375,0.0625,0.9375,0.125,0.9063,0.0625,0.9375,0.125,0.9063,0.125,0.9063,0.0625,0.9063,0.0625,0.9063,0.125,0.875,0.0625,0.9063,0.125,0.875,0.125,0.875,0.0625,0.875,0.0625,0.875,0.125,0.8438,0.0625,0.875,0.125,0.8438,0.125,0.8438,0.0625,0.8438,0.0625,0.8438,0.125,0.8125,0.0625,0.8438,0.125,0.8125,0.125,0.8125,0.0625,0.8125,0.0625,0.8125,0.125,0.7813,0.0625,0.8125,0.125,0.7813,0.125,0.7813,0.0625,0.7813,0.0625,0.7813,0.125,0.75,0.0625,0.7813,0.125,0.75,0.125,0.75,0.0625,0.75,0.0625,0.75,0.125,0.7188,0.0625,0.75,0.125,0.7188,0.125,0.7188,0.0625,0.7188,0.0625,0.7188,0.125,0.6875,0.0625,0.7188,0.125,0.6875,0.125,0.6875,0.0625,0.6875,0.0625,0.6875,0.125,0.6563,0.0625,0.6875,0.125,0.6563,0.125,0.6563,0.0625,0.6563,0.0625,0.6563,0.125,0.625,0.0625,0.6563,0.125,0.625,0.125,0.625,0.0625,0.625,0.0625,0.625,0.125,0.5938,0.0625,0.625,0.125,0.5938,0.125,0.5938,0.0625,0.5938,0.0625,0.5938,0.125,0.5625,0.0625,0.5938,0.125,0.5625,0.125,0.5625,0.0625,0.5625,0.0625,0.5625,0.125,0.5313,0.0625,0.5625,0.125,0.5313,0.125,0.5313,0.0625,0.5313,0.0625,0.5313,0.125,0.5,0.0625,0.5313,0.125,0.5,0.125,0.5,0.0625,0.5,0.0625,0.5,0.125,0.4688,0.0625,0.5,0.125,0.4688,0.125,0.4688,0.0625,0.4688,0.0625,0.4688,0.125,0.4375,0.0625,0.4688,0.125,0.4375,0.125,0.4375,0.0625,0.4375,0.0625,0.4375,0.125,0.4063,0.0625,0.4375,0.125,0.4063,0.125,0.4063,0.0625,0.4063,0.0625,0.4063,0.125,0.375,0.0625,0.4063,0.125,0.375,0.125,0.375,0.0625,0.375,0.0625,0.375,0.125,0.3438,0.0625,0.375,0.125,0.3438,0.125,0.3438,0.0625,0.3438,0.0625,0.3438,0.125,0.3125,0.0625,0.3438,0.125,0.3125,0.125,0.3125,0.0625,0.3125,0.0625,0.3125,0.125,0.2813,0.0625,0.3125,0.125,0.2813,0.125,0.2813,0.0625,0.2813,0.0625,0.2813,0.125,0.25,0.0625,0.2813,0.125,0.25,0.125,0.25,0.0625,0.25,0.0625,0.25,0.125,0.2188,0.0625,0.25,0.125,0.2188,0.125,0.2188,0.0625,0.2188,0.0625,0.2188,0.125,0.1875,0.0625,0.2188,0.125,0.1875,0.125,0.1875,0.0625,0.1875,0.0625,0.1875,0.125,0.1563,0.0625,0.1875,0.125,0.1563,0.125,0.1563,0.0625,0.1563,0.0625,0.1563,0.125,0.125,0.0625,0.1563,0.125,0.125,0.125,0.125,0.0625,0.125,0.0625,0.125,0.125,0.0938,0.0625,0.125,0.125,0.0938,0.125,0.0938,0.0625,0.0938,0.0625,0.0938,0.125,0.0625,0.0625,0.0938,0.125,0.0625,0.125,0.0625,0.0625,0.0625,0.0625,0.0625,0.125,0.0313,0.0625,0.0625,0.125,0.0313,0.125,0.0313,0.0625,0.0313,0.0625,0.0313,0.125,0.0,0.0625,0.0313,0.125,0.0,0.125,0.0,0.0625,1.0,0.0625,0.9688,0.0625,0.9688,0.0,0.9688,0.0625,0.9375,0.0625,0.9375,0.0,0.9375,0.0625,0.9063,0.0625,0.9063,0.0,0.9063,0.0625,0.875,0.0625,0.875,0.0,0.875,0.0625,0.8438,0.0625,0.8438,0.0,0.8438,0.0625,0.8125,0.0625,0.8125,0.0,0.8125,0.0625,0.7813,0.0625,0.7813,0.0,0.7813,0.0625,0.75,0.0625,0.75,0.0,0.75,0.0625,0.7188,0.0625,0.7188,0.0,0.7188,0.0625,0.6875,0.0625,0.6875,0.0,0.6875,0.0625,0.6563,0.0625,0.6563,0.0,0.6563,0.0625,0.625,0.0625,0.625,0.0,0.625,0.0625,0.5938,0.0625,0.5938,0.0,0.5938,0.0625,0.5625,0.0625,0.5625,0.0,0.5625,0.0625,0.5313,0.0625,0.5313,0.0,0.5313,0.0625,0.5,0.0625,0.5,0.0,0.5,0.0625,0.4688,0.0625,0.4688,0.0,0.4688,0.0625,0.4375,0.0625,0.4375,0.0,0.4375,0.0625,0.4063,0.0625,0.4063,0.0,0.4063,0.0625,0.375,0.0625,0.375,0.0,0.375,0.0625,0.3438,0.0625,0.3438,0.0,0.3438,0.0625,0.3125,0.0625,0.3125,0.0,0.3125,0.0625,0.2813,0.0625,0.2813,0.0,0.2813,0.0625,0.25,0.0625,0.25,0.0,0.25,0.0625,0.2188,0.0625,0.2188,0.0,0.2188,0.0625,0.1875,0.0625,0.1875,0.0,0.1875,0.0625,0.1563,0.0625,0.1563,0.0,0.1563,0.0625,0.125,0.0625,0.125,0.0,0.125,0.0625,0.0938,0.0625,0.0938,0.0,0.0938,0.0625,0.0625,0.0625,0.0625,0.0,0.0625,0.0625,0.0313,0.0625,0.0313,0.0,0.0313,0.0625,0.0,0.0625,0.0,0.0]}}}}],\\"name\\":\\"ModelFromPro\\",\\"mbb\\":[-1,-1,-1,1,1,1],\\"pivotOffset\\":[0,0,0]},\\"materialDefinitions\\":{\\"0\\":{\\"type\\":\\"standard\\",\\"params\\":{\\"transparency\\":0,\\"diffuse\\":[1.0,1.0,1.0],\\"externalColorMixMode\\":\\"tint\\"}}},\\"textureDefinitions\\":{\\"20fd5397feabb33012b0b3b27f683e3e.dat\\":{\\"encoding\\":\\"data:image/jpeg\\",\\"channels\\":\\"rgb\\",\\"alphaChannelUsage\\":\\"transparency\\",\\"images\\":[{\\"size\\":128,\\"data\\":\\"/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCACAAIADASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwDjaKKK+ZP3AKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooA//Z\\"}]}}}"
    },
    {
      "type" : "CIMBinaryReference",
      "uRI" : "CIMPATH=ObjectMarker3D/74713a22678849d97fa38d013785ecfb.dat",
      "data" : "{\\"version\\":\\"1.2\\",\\"authoringApp\\":\\"ArcGIS Pro\\",\\"authoringAppVersion\\":\\"2.6.0\\",\\"lods\\":[{\\"metrics\\":{\\"faceCount\\":1136}}],\\"model\\":{\\"geometries\\":[{\\"type\\":\\"Embedded\\",\\"transformation\\":[1.0,0.0,0.0,0.0,0.0,1.0,0.0,0.0,0.0,0.0,1.0,0.0,0.0,0.0,0.0,1.0],\\"params\\":{\\"topology\\":\\"PerAttributeArray\\",\\"material\\":\\"/materialDefinitions/0\\",\\"texture\\":\\"/textureDefinitions/20fd5397feabb33012b0b3b27f683e3e.dat\\",\\"vertexAttributes\\":{\\"position\\":{\\"valueType\\":\\"Float32\\",\\"valuesPerElement\\":3,\\"values\\":[0.0380598,0.1913416,-0.9807853,-0.0,0.1950901,-0.9807853,0.0,0.0,-1.0,0.0746574,0.1802399,-0.9807853,0.0380598,0.1913416,-0.9807853,0.0,0.0,-1.0,0.1083859,0.1622117,-0.9807853,0.0746574,0.1802399,-0.9807853,0.0,0.0,-1.0,0.1379493,0.1379498,-0.9807853,0.1083859,0.1622117,-0.9807853,0.0,0.0,-1.0,0.1622113,0.1083865,-0.9807853,0.1379493,0.1379498,-0.9807853,0.0,0.0,-1.0,0.1802396,0.0746581,-0.9807853,0.1622113,0.1083865,-0.9807853,0.0,0.0,-1.0,0.1913414,0.0380605,-0.9807853,0.1802396,0.0746581,-0.9807853,0.0,0.0,-1.0,0.1950901,0.0000003,-0.9807853,0.1913414,0.0380605,-0.9807853,0.0,0.0,-1.0,0.1913415,-0.0380599,-0.9807853,0.1950901,0.0000003,-0.9807853,0.0,0.0,-1.0,0.1802398,-0.0746575,-0.9807853,0.1913415,-0.0380599,-0.9807853,0.0,0.0,-1.0,0.1622116,-0.108386,-0.9807853,0.1802398,-0.0746575,-0.9807853,0.0,0.0,-1.0,0.1379497,-0.1379493,-0.9807853,0.1622116,-0.108386,-0.9807853,0.0,0.0,-1.0,0.1083864,-0.1622113,-0.9807853,0.1379497,-0.1379493,-0.9807853,0.0,0.0,-1.0,0.0746579,-0.1802396,-0.9807853,0.1083864,-0.1622113,-0.9807853,0.0,0.0,-1.0,0.0380604,-0.1913414,-0.9807853,0.0746579,-0.1802396,-0.9807853,0.0,0.0,-1.0,0.0000002,-0.1950901,-0.9807853,0.0380604,-0.1913414,-0.9807853,0.0,0.0,-1.0,-0.03806,-0.1913415,-0.9807853,0.0000002,-0.1950901,-0.9807853,0.0,0.0,-1.0,-0.0746576,-0.1802398,-0.9807853,-0.03806,-0.1913415,-0.9807853,0.0,0.0,-1.0,-0.1083861,-0.1622116,-0.9807853,-0.0746576,-0.1802398,-0.9807853,0.0,0.0,-1.0,-0.1379494,-0.1379496,-0.9807853,-0.1083861,-0.1622116,-0.9807853,0.0,0.0,-1.0,-0.1622114,-0.1083863,-0.9807853,-0.1379494,-0.1379496,-0.9807853,0.0,0.0,-1.0,-0.1802397,-0.0746578,-0.9807853,-0.1622114,-0.1083863,-0.9807853,0.0,0.0,-1.0,-0.1913415,-0.0380603,-0.9807853,-0.1802397,-0.0746578,-0.9807853,0.0,0.0,-1.0,-0.1950901,-0.0000001,-0.9807853,-0.1913415,-0.0380603,-0.9807853,0.0,0.0,-1.0,-0.1913415,0.0380601,-0.9807853,-0.1950901,-0.0000001,-0.9807853,0.0,0.0,-1.0,-0.1802398,0.0746577,-0.9807853,-0.1913415,0.0380601,-0.9807853,0.0,0.0,-1.0,-0.1622115,0.1083862,-0.9807853,-0.1802398,0.0746577,-0.9807853,0.0,0.0,-1.0,-0.1379495,0.1379495,-0.9807853,-0.1622115,0.1083862,-0.9807853,0.0,0.0,-1.0,-0.1083862,0.1622115,-0.9807853,-0.1379495,0.1379495,-0.9807853,0.0,0.0,-1.0,-0.0746577,0.1802397,-0.9807853,-0.1083862,0.1622115,-0.9807853,0.0,0.0,-1.0,-0.0380602,0.1913415,-0.9807853,-0.0746577,0.1802397,-0.9807853,0.0,0.0,-1.0,-0.0,0.1950901,-0.9807853,-0.0380602,0.1913415,-0.9807853,0.0,0.0,-1.0,-0.0,0.3826833,-0.9238796,-0.0,0.1950901,-0.9807853,0.074657,0.3753303,-0.9238796,-0.0,0.1950901,-0.9807853,0.0380598,0.1913416,-0.9807853,0.074657,0.3753303,-0.9238796,0.074657,0.3753303,-0.9238796,0.0380598,0.1913416,-0.9807853,0.1464458,0.3535536,-0.9238796,0.0380598,0.1913416,-0.9807853,0.0746574,0.1802399,-0.9807853,0.1464458,0.3535536,-0.9238796,0.1464458,0.3535536,-0.9238796,0.0746574,0.1802399,-0.9807853,0.2126068,0.3181899,-0.9238796,0.0746574,0.1802399,-0.9807853,0.1083859,0.1622117,-0.9807853,0.2126068,0.3181899,-0.9238796,0.2126068,0.3181899,-0.9238796,0.1083859,0.1622117,-0.9807853,0.2705974,0.2705985,-0.9238796,0.1083859,0.1622117,-0.9807853,0.1379493,0.1379498,-0.9807853,0.2705974,0.2705985,-0.9238796,0.2705974,0.2705985,-0.9238796,0.1379493,0.1379498,-0.9807853,0.3181891,0.212608,-0.9238796,0.1379493,0.1379498,-0.9807853,0.1622113,0.1083865,-0.9807853,0.3181891,0.212608,-0.9238796,0.3181891,0.212608,-0.9238796,0.1622113,0.1083865,-0.9807853,0.353553,0.1464472,-0.9238796,0.1622113,0.1083865,-0.9807853,0.1802396,0.0746581,-0.9807853,0.353553,0.1464472,-0.9238796,0.353553,0.1464472,-0.9238796,0.1802396,0.0746581,-0.9807853,0.37533,0.0746584,-0.9238796,0.1802396,0.0746581,-0.9807853,0.1913414,0.0380605,-0.9807853,0.37533,0.0746584,-0.9238796,0.37533,0.0746584,-0.9238796,0.1913414,0.0380605,-0.9807853,0.3826833,0.0000006,-0.9238796,0.1913414,0.0380605,-0.9807853,0.1950901,0.0000003,-0.9807853,0.3826833,0.0000006,-0.9238796,0.3826833,0.0000006,-0.9238796,0.1950901,0.0000003,-0.9807853,0.3753302,-0.0746572,-0.9238796,0.1950901,0.0000003,-0.9807853,0.1913415,-0.0380599,-0.9807853,0.3753302,-0.0746572,-0.9238796,0.3753302,-0.0746572,-0.9238796,0.1913415,-0.0380599,-0.9807853,0.3535535,-0.146446,-0.9238796,0.1913415,-0.0380599,-0.9807853,0.1802398,-0.0746575,-0.9807853,0.3535535,-0.146446,-0.9238796,0.3535535,-0.146446,-0.9238796,0.1802398,-0.0746575,-0.9807853,0.3181898,-0.212607,-0.9238796,0.1802398,-0.0746575,-0.9807853,0.1622116,-0.108386,-0.9807853,0.3181898,-0.212607,-0.9238796,0.3181898,-0.212607,-0.9238796,0.1622116,-0.108386,-0.9807853,0.2705983,-0.2705976,-0.9238796,0.1622116,-0.108386,-0.9807853,0.1379497,-0.1379493,-0.9807853,0.2705983,-0.2705976,-0.9238796,0.2705983,-0.2705976,-0.9238796,0.1379497,-0.1379493,-0.9807853,0.2126078,-0.3181893,-0.9238796,0.1379497,-0.1379493,-0.9807853,0.1083864,-0.1622113,-0.9807853,0.2126078,-0.3181893,-0.9238796,0.2126078,-0.3181893,-0.9238796,0.1083864,-0.1622113,-0.9807853,0.1464469,-0.3535531,-0.9238796,0.1083864,-0.1622113,-0.9807853,0.0746579,-0.1802396,-0.9807853,0.1464469,-0.3535531,-0.9238796,0.1464469,-0.3535531,-0.9238796,0.0746579,-0.1802396,-0.9807853,0.0746582,-0.37533,-0.9238796,0.0746579,-0.1802396,-0.9807853,0.0380604,-0.1913414,-0.9807853,0.0746582,-0.37533,-0.9238796,0.0746582,-0.37533,-0.9238796,0.0380604,-0.1913414,-0.9807853,0.0000004,-0.3826833,-0.9238796,0.0380604,-0.1913414,-0.9807853,0.0000002,-0.1950901,-0.9807853,0.0000004,-0.3826833,-0.9238796,0.0000004,-0.3826833,-0.9238796,0.0000002,-0.1950901,-0.9807853,-0.0746575,-0.3753302,-0.9238796,0.0000002,-0.1950901,-0.9807853,-0.03806,-0.1913415,-0.9807853,-0.0746575,-0.3753302,-0.9238796,-0.0746575,-0.3753302,-0.9238796,-0.03806,-0.1913415,-0.9807853,-0.1464463,-0.3535534,-0.9238796,-0.03806,-0.1913415,-0.9807853,-0.0746576,-0.1802398,-0.9807853,-0.1464463,-0.3535534,-0.9238796,-0.1464463,-0.3535534,-0.9238796,-0.0746576,-0.1802398,-0.9807853,-0.2126072,-0.3181897,-0.9238796,-0.0746576,-0.1802398,-0.9807853,-0.1083861,-0.1622116,-0.9807853,-0.2126072,-0.3181897,-0.9238796,-0.2126072,-0.3181897,-0.9238796,-0.1083861,-0.1622116,-0.9807853,-0.2705978,-0.2705981,-0.9238796,-0.1083861,-0.1622116,-0.9807853,-0.1379494,-0.1379496,-0.9807853,-0.2705978,-0.2705981,-0.9238796,-0.2705978,-0.2705981,-0.9238796,-0.1379494,-0.1379496,-0.9807853,-0.3181894,-0.2126076,-0.9238796,-0.1379494,-0.1379496,-0.9807853,-0.1622114,-0.1083863,-0.9807853,-0.3181894,-0.2126076,-0.9238796,-0.3181894,-0.2126076,-0.9238796,-0.1622114,-0.1083863,-0.9807853,-0.3535532,-0.1464467,-0.9238796,-0.1622114,-0.1083863,-0.9807853,-0.1802397,-0.0746578,-0.9807853,-0.3535532,-0.1464467,-0.9238796,-0.3535532,-0.1464467,-0.9238796,-0.1802397,-0.0746578,-0.9807853,-0.3753301,-0.074658,-0.9238796,-0.1802397,-0.0746578,-0.9807853,-0.1913415,-0.0380603,-0.9807853,-0.3753301,-0.074658,-0.9238796,-0.3753301,-0.074658,-0.9238796,-0.1913415,-0.0380603,-0.9807853,-0.3826833,-0.0000001,-0.9238796,-0.1913415,-0.0380603,-0.9807853,-0.1950901,-0.0000001,-0.9807853,-0.3826833,-0.0000001,-0.9238796,-0.3826833,-0.0000001,-0.9238796,-0.1950901,-0.0000001,-0.9807853,-0.3753302,0.0746577,-0.9238796,-0.1950901,-0.0000001,-0.9807853,-0.1913415,0.0380601,-0.9807853,-0.3753302,0.0746577,-0.9238796,-0.3753302,0.0746577,-0.9238796,-0.1913415,0.0380601,-0.9807853,-0.3535533,0.1464465,-0.9238796,-0.1913415,0.0380601,-0.9807853,-0.1802398,0.0746577,-0.9807853,-0.3535533,0.1464465,-0.9238796,-0.3535533,0.1464465,-0.9238796,-0.1802398,0.0746577,-0.9807853,-0.3181895,0.2126074,-0.9238796,-0.1802398,0.0746577,-0.9807853,-0.1622115,0.1083862,-0.9807853,-0.3181895,0.2126074,-0.9238796,-0.3181895,0.2126074,-0.9238796,-0.1622115,0.1083862,-0.9807853,-0.2705979,0.2705979,-0.9238796,-0.1622115,0.1083862,-0.9807853,-0.1379495,0.1379495,-0.9807853,-0.2705979,0.2705979,-0.9238796,-0.2705979,0.2705979,-0.9238796,-0.1379495,0.1379495,-0.9807853,-0.2126074,0.3181895,-0.9238796,-0.1379495,0.1379495,-0.9807853,-0.1083862,0.1622115,-0.9807853,-0.2126074,0.3181895,-0.9238796,-0.2126074,0.3181895,-0.9238796,-0.1083862,0.1622115,-0.9807853,-0.1464465,0.3535533,-0.9238796,-0.1083862,0.1622115,-0.9807853,-0.0746577,0.1802397,-0.9807853,-0.1464465,0.3535533,-0.9238796,-0.1464465,0.3535533,-0.9238796,-0.0746577,0.1802397,-0.9807853,-0.0746578,0.3753301,-0.9238796,-0.0746577,0.1802397,-0.9807853,-0.0380602,0.1913415,-0.9807853,-0.0746578,0.3753301,-0.9238796,-0.0746578,0.3753301,-0.9238796,-0.0380602,0.1913415,-0.9807853,-0.0,0.3826833,-0.9238796,-0.0380602,0.1913415,-0.9807853,-0.0,0.1950901,-0.9807853,-0.0,0.3826833,-0.9238796,-0.0,0.5555702,-0.8314697,-0.0,0.3826833,-0.9238796,0.1083852,0.5448953,-0.8314697,-0.0,0.3826833,-0.9238796,0.074657,0.3753303,-0.9238796,0.1083852,0.5448953,-0.8314697,0.1083852,0.5448953,-0.8314697,0.074657,0.3753303,-0.9238796,0.2126064,0.5132804,-0.8314697,0.074657,0.3753303,-0.9238796,0.1464458,0.3535536,-0.9238796,0.2126064,0.5132804,-0.8314697,0.2126064,0.5132804,-0.8314697,0.1464458,0.3535536,-0.9238796,0.3086573,0.4619403,-0.8314697,0.1464458,0.3535536,-0.9238796,0.2126068,0.3181899,-0.9238796,0.3086573,0.4619403,-0.8314697,0.3086573,0.4619403,-0.8314697,0.2126068,0.3181899,-0.9238796,0.3928467,0.3928482,-0.8314697,0.2126068,0.3181899,-0.9238796,0.2705974,0.2705985,-0.9238796,0.3928467,0.3928482,-0.8314697,0.3928467,0.3928482,-0.8314697,0.2705974,0.2705985,-0.9238796,0.4619392,0.3086591,-0.8314697,0.2705974,0.2705985,-0.9238796,0.3181891,0.212608,-0.9238796,0.4619392,0.3086591,-0.8314697,0.4619392,0.3086591,-0.8314697,0.3181891,0.212608,-0.9238796,0.5132796,0.2126084,-0.8314697,0.3181891,0.212608,-0.9238796,0.353553,0.1464472,-0.9238796,0.5132796,0.2126084,-0.8314697,0.5132796,0.2126084,-0.8314697,0.353553,0.1464472,-0.9238796,0.5448949,0.1083873,-0.8314697,0.353553,0.1464472,-0.9238796,0.37533,0.0746584,-0.9238796,0.5448949,0.1083873,-0.8314697,0.5448949,0.1083873,-0.8314697,0.37533,0.0746584,-0.9238796,0.5555702,0.0000009,-0.8314697,0.37533,0.0746584,-0.9238796,0.3826833,0.0000006,-0.9238796,0.5555702,0.0000009,-0.8314697,0.5555702,0.0000009,-0.8314697,0.3826833,0.0000006,-0.9238796,0.5448952,-0.1083855,-0.8314697,0.3826833,0.0000006,-0.9238796,0.3753302,-0.0746572,-0.9238796,0.5448952,-0.1083855,-0.8314697,0.5448952,-0.1083855,-0.8314697,0.3753302,-0.0746572,-0.9238796,0.5132802,-0.2126068,-0.8314697,0.3753302,-0.0746572,-0.9238796,0.3535535,-0.146446,-0.9238796,0.5132802,-0.2126068,-0.8314697,0.5132802,-0.2126068,-0.8314697,0.3535535,-0.146446,-0.9238796,0.4619401,-0.3086576,-0.8314697,0.3535535,-0.146446,-0.9238796,0.3181898,-0.212607,-0.9238796,0.4619401,-0.3086576,-0.8314697,0.4619401,-0.3086576,-0.8314697,0.3181898,-0.212607,-0.9238796,0.3928479,-0.3928469,-0.8314697,0.3181898,-0.212607,-0.9238796,0.2705983,-0.2705976,-0.9238796,0.3928479,-0.3928469,-0.8314697,0.3928479,-0.3928469,-0.8314697,0.2705983,-0.2705976,-0.9238796,0.3086588,-0.4619394,-0.8314697,0.2705983,-0.2705976,-0.9238796,0.2126078,-0.3181893,-0.9238796,0.3086588,-0.4619394,-0.8314697,0.3086588,-0.4619394,-0.8314697,0.2126078,-0.3181893,-0.9238796,0.2126081,-0.5132797,-0.8314697,0.2126078,-0.3181893,-0.9238796,0.1464469,-0.3535531,-0.9238796,0.2126081,-0.5132797,-0.8314697,0.2126081,-0.5132797,-0.8314697,0.1464469,-0.3535531,-0.9238796,0.1083869,-0.5448949,-0.8314697,0.1464469,-0.3535531,-0.9238796,0.0746582,-0.37533,-0.9238796,0.1083869,-0.5448949,-0.8314697,0.1083869,-0.5448949,-0.8314697,0.0746582,-0.37533,-0.9238796,0.0000005,-0.5555702,-0.8314697,0.0746582,-0.37533,-0.9238796,0.0000004,-0.3826833,-0.9238796,0.0000005,-0.5555702,-0.8314697,0.0000005,-0.5555702,-0.8314697,0.0000004,-0.3826833,-0.9238796,-0.1083859,-0.5448952,-0.8314697,0.0000004,-0.3826833,-0.9238796,-0.0746575,-0.3753302,-0.9238796,-0.1083859,-0.5448952,-0.8314697,-0.1083859,-0.5448952,-0.8314697,-0.0746575,-0.3753302,-0.9238796,-0.2126071,-0.5132801,-0.8314697,-0.0746575,-0.3753302,-0.9238796,-0.1464463,-0.3535534,-0.9238796,-0.2126071,-0.5132801,-0.8314697,-0.2126071,-0.5132801,-0.8314697,-0.1464463,-0.3535534,-0.9238796,-0.3086579,-0.46194,-0.8314697,-0.1464463,-0.3535534,-0.9238796,-0.2126072,-0.3181897,-0.9238796,-0.3086579,-0.46194,-0.8314697,-0.3086579,-0.46194,-0.8314697,-0.2126072,-0.3181897,-0.9238796,-0.3928472,-0.3928477,-0.8314697,-0.2126072,-0.3181897,-0.9238796,-0.2705978,-0.2705981,-0.9238796,-0.3928472,-0.3928477,-0.8314697,-0.3928472,-0.3928477,-0.8314697,-0.2705978,-0.2705981,-0.9238796,-0.4619395,-0.3086585,-0.8314697,-0.2705978,-0.2705981,-0.9238796,-0.3181894,-0.2126076,-0.9238796,-0.4619395,-0.3086585,-0.8314697,-0.4619395,-0.3086585,-0.8314697,-0.3181894,-0.2126076,-0.9238796,-0.5132798,-0.2126078,-0.8314697,-0.3181894,-0.2126076,-0.9238796,-0.3535532,-0.1464467,-0.9238796,-0.5132798,-0.2126078,-0.8314697,-0.5132798,-0.2126078,-0.8314697,-0.3535532,-0.1464467,-0.9238796,-0.544895,-0.1083866,-0.8314697,-0.3535532,-0.1464467,-0.9238796,-0.3753301,-0.074658,-0.9238796,-0.544895,-0.1083866,-0.8314697,-0.544895,-0.1083866,-0.8314697,-0.3753301,-0.074658,-0.9238796,-0.5555702,-0.0000002,-0.8314697,-0.3753301,-0.074658,-0.9238796,-0.3826833,-0.0000001,-0.9238796,-0.5555702,-0.0000002,-0.8314697,-0.5555702,-0.0000002,-0.8314697,-0.3826833,-0.0000001,-0.9238796,-0.5448951,0.1083862,-0.8314697,-0.3826833,-0.0000001,-0.9238796,-0.3753302,0.0746577,-0.9238796,-0.5448951,0.1083862,-0.8314697,-0.5448951,0.1083862,-0.8314697,-0.3753302,0.0746577,-0.9238796,-0.51328,0.2126074,-0.8314697,-0.3753302,0.0746577,-0.9238796,-0.3535533,0.1464465,-0.9238796,-0.51328,0.2126074,-0.8314697,-0.51328,0.2126074,-0.8314697,-0.3535533,0.1464465,-0.9238796,-0.4619398,0.3086582,-0.8314697,-0.3535533,0.1464465,-0.9238796,-0.3181895,0.2126074,-0.9238796,-0.4619398,0.3086582,-0.8314697,-0.4619398,0.3086582,-0.8314697,-0.3181895,0.2126074,-0.9238796,-0.3928474,0.3928474,-0.8314697,-0.3181895,0.2126074,-0.9238796,-0.2705979,0.2705979,-0.9238796,-0.3928474,0.3928474,-0.8314697,-0.3928474,0.3928474,-0.8314697,-0.2705979,0.2705979,-0.9238796,-0.3086582,0.4619398,-0.8314697,-0.2705979,0.2705979,-0.9238796,-0.2126074,0.3181895,-0.9238796,-0.3086582,0.4619398,-0.8314697,-0.3086582,0.4619398,-0.8314697,-0.2126074,0.3181895,-0.9238796,-0.2126075,0.5132799,-0.8314697,-0.2126074,0.3181895,-0.9238796,-0.1464465,0.3535533,-0.9238796,-0.2126075,0.5132799,-0.8314697,-0.2126075,0.5132799,-0.8314697,-0.1464465,0.3535533,-0.9238796,-0.1083864,0.5448951,-0.8314697,-0.1464465,0.3535533,-0.9238796,-0.0746578,0.3753301,-0.9238796,-0.1083864,0.5448951,-0.8314697,-0.1083864,0.5448951,-0.8314697,-0.0746578,0.3753301,-0.9238796,-0.0,0.5555702,-0.8314697,-0.0746578,0.3753301,-0.9238796,-0.0,0.3826833,-0.9238796,-0.0,0.5555702,-0.8314697,-0.0,0.7071068,-0.7071068,-0.0,0.5555702,-0.8314697,0.1379482,0.6935202,-0.7071068,-0.0,0.5555702,-0.8314697,0.1083852,0.5448953,-0.8314697,0.1379482,0.6935202,-0.7071068,0.1379482,0.6935202,-0.7071068,0.1083852,0.5448953,-0.8314697,0.2705967,0.653282,-0.7071068,0.1083852,0.5448953,-0.8314697,0.2126064,0.5132804,-0.8314697,0.2705967,0.653282,-0.7071068,0.2705967,0.653282,-0.7071068,0.2126064,0.5132804,-0.8314697,0.3928463,0.5879386,-0.7071068,0.2126064,0.5132804,-0.8314697,0.3086573,0.4619403,-0.8314697,0.3928463,0.5879386,-0.7071068,0.3928463,0.5879386,-0.7071068,0.3086573,0.4619403,-0.8314697,0.499999,0.500001,-0.7071068,0.3086573,0.4619403,-0.8314697,0.3928467,0.3928482,-0.8314697,0.499999,0.500001,-0.7071068,0.499999,0.500001,-0.7071068,0.3928467,0.3928482,-0.8314697,0.5879371,0.3928486,-0.7071068,0.3928467,0.3928482,-0.8314697,0.4619392,0.3086591,-0.8314697,0.5879371,0.3928486,-0.7071068,0.5879371,0.3928486,-0.7071068,0.4619392,0.3086591,-0.8314697,0.653281,0.2705992,-0.7071068,0.4619392,0.3086591,-0.8314697,0.5132796,0.2126084,-0.8314697,0.653281,0.2705992,-0.7071068,0.653281,0.2705992,-0.7071068,0.5132796,0.2126084,-0.8314697,0.6935197,0.1379509,-0.7071068,0.5132796,0.2126084,-0.8314697,0.5448949,0.1083873,-0.8314697,0.6935197,0.1379509,-0.7071068,0.6935197,0.1379509,-0.7071068,0.5448949,0.1083873,-0.8314697,0.7071068,0.0000011,-0.7071068,0.5448949,0.1083873,-0.8314697,0.5555702,0.0000009,-0.8314697,0.7071068,0.0000011,-0.7071068,0.7071068,0.0000011,-0.7071068,0.5555702,0.0000009,-0.8314697,0.6935201,-0.1379486,-0.7071068,0.5555702,0.0000009,-0.8314697,0.5448952,-0.1083855,-0.8314697,0.6935201,-0.1379486,-0.7071068,0.6935201,-0.1379486,-0.7071068,0.5448952,-0.1083855,-0.8314697,0.6532819,-0.2705971,-0.7071068,0.5448952,-0.1083855,-0.8314697,0.5132802,-0.2126068,-0.8314697,0.6532819,-0.2705971,-0.7071068,0.6532819,-0.2705971,-0.7071068,0.5132802,-0.2126068,-0.8314697,0.5879383,-0.3928467,-0.7071068,0.5132802,-0.2126068,-0.8314697,0.4619401,-0.3086576,-0.8314697,0.5879383,-0.3928467,-0.7071068,0.5879383,-0.3928467,-0.7071068,0.4619401,-0.3086576,-0.8314697,0.5000006,-0.4999993,-0.7071068,0.4619401,-0.3086576,-0.8314697,0.3928479,-0.3928469,-0.8314697,0.5000006,-0.4999993,-0.7071068,0.5000006,-0.4999993,-0.7071068,0.3928479,-0.3928469,-0.8314697,0.3928482,-0.5879373,-0.7071068,0.3928479,-0.3928469,-0.8314697,0.3086588,-0.4619394,-0.8314697,0.3928482,-0.5879373,-0.7071068,0.3928482,-0.5879373,-0.7071068,0.3086588,-0.4619394,-0.8314697,0.2705988,-0.6532812,-0.7071068,0.3086588,-0.4619394,-0.8314697,0.2126081,-0.5132797,-0.8314697,0.2705988,-0.6532812,-0.7071068,0.2705988,-0.6532812,-0.7071068,0.2126081,-0.5132797,-0.8314697,0.1379504,-0.6935198,-0.7071068,0.2126081,-0.5132797,-0.8314697,0.1083869,-0.5448949,-0.8314697,0.1379504,-0.6935198,-0.7071068,0.1379504,-0.6935198,-0.7071068,0.1083869,-0.5448949,-0.8314697,0.0000007,-0.7071068,-0.7071068,0.1083869,-0.5448949,-0.8314697,0.0000005,-0.5555702,-0.8314697,0.0000007,-0.7071068,-0.7071068,0.0000007,-0.7071068,-0.7071068,0.0000005,-0.5555702,-0.8314697,-0.1379491,-0.69352,-0.7071068,0.0000005,-0.5555702,-0.8314697,-0.1083859,-0.5448952,-0.8314697,-0.1379491,-0.69352,-0.7071068,-0.1379491,-0.69352,-0.7071068,-0.1083859,-0.5448952,-0.8314697,-0.2705975,-0.6532817,-0.7071068,-0.1083859,-0.5448952,-0.8314697,-0.2126071,-0.5132801,-0.8314697,-0.2705975,-0.6532817,-0.7071068,-0.2705975,-0.6532817,-0.7071068,-0.2126071,-0.5132801,-0.8314697,-0.3928471,-0.5879381,-0.7071068,-0.2126071,-0.5132801,-0.8314697,-0.3086579,-0.46194,-0.8314697,-0.3928471,-0.5879381,-0.7071068,-0.3928471,-0.5879381,-0.7071068,-0.3086579,-0.46194,-0.8314697,-0.4999997,-0.5000003,-0.7071068,-0.3086579,-0.46194,-0.8314697,-0.3928472,-0.3928477,-0.8314697,-0.4999997,-0.5000003,-0.7071068,-0.4999997,-0.5000003,-0.7071068,-0.3928472,-0.3928477,-0.8314697,-0.5879376,-0.3928478,-0.7071068,-0.3928472,-0.3928477,-0.8314697,-0.4619395,-0.3086585,-0.8314697,-0.5879376,-0.3928478,-0.7071068,-0.5879376,-0.3928478,-0.7071068,-0.4619395,-0.3086585,-0.8314697,-0.6532813,-0.2705984,-0.7071068,-0.4619395,-0.3086585,-0.8314697,-0.5132798,-0.2126078,-0.8314697,-0.6532813,-0.2705984,-0.7071068,-0.6532813,-0.2705984,-0.7071068,-0.5132798,-0.2126078,-0.8314697,-0.6935198,-0.13795,-0.7071068,-0.5132798,-0.2126078,-0.8314697,-0.544895,-0.1083866,-0.8314697,-0.6935198,-0.13795,-0.7071068,-0.6935198,-0.13795,-0.7071068,-0.544895,-0.1083866,-0.8314697,-0.7071068,-0.0000002,-0.7071068,-0.544895,-0.1083866,-0.8314697,-0.5555702,-0.0000002,-0.8314697,-0.7071068,-0.0000002,-0.7071068,-0.7071068,-0.0000002,-0.7071068,-0.5555702,-0.0000002,-0.8314697,-0.6935199,0.1379495,-0.7071068,-0.5555702,-0.0000002,-0.8314697,-0.5448951,0.1083862,-0.8314697,-0.6935199,0.1379495,-0.7071068,-0.6935199,0.1379495,-0.7071068,-0.5448951,0.1083862,-0.8314697,-0.6532815,0.2705979,-0.7071068,-0.5448951,0.1083862,-0.8314697,-0.51328,0.2126074,-0.8314697,-0.6532815,0.2705979,-0.7071068,-0.6532815,0.2705979,-0.7071068,-0.51328,0.2126074,-0.8314697,-0.5879378,0.3928474,-0.7071068,-0.51328,0.2126074,-0.8314697,-0.4619398,0.3086582,-0.8314697,-0.5879378,0.3928474,-0.7071068,-0.5879378,0.3928474,-0.7071068,-0.4619398,0.3086582,-0.8314697,-0.5,0.5,-0.7071068,-0.4619398,0.3086582,-0.8314697,-0.3928474,0.3928474,-0.8314697,-0.5,0.5,-0.7071068,-0.5,0.5,-0.7071068,-0.3928474,0.3928474,-0.8314697,-0.3928474,0.5879378,-0.7071068,-0.3928474,0.3928474,-0.8314697,-0.3086582,0.4619398,-0.8314697,-0.3928474,0.5879378,-0.7071068,-0.3928474,0.5879378,-0.7071068,-0.3086582,0.4619398,-0.8314697,-0.270598,0.6532815,-0.7071068,-0.3086582,0.4619398,-0.8314697,-0.2126075,0.5132799,-0.8314697,-0.270598,0.6532815,-0.7071068,-0.270598,0.6532815,-0.7071068,-0.2126075,0.5132799,-0.8314697,-0.1379497,0.6935199,-0.7071068,-0.2126075,0.5132799,-0.8314697,-0.1083864,0.5448951,-0.8314697,-0.1379497,0.6935199,-0.7071068,-0.1379497,0.6935199,-0.7071068,-0.1083864,0.5448951,-0.8314697,-0.0,0.7071068,-0.7071068,-0.1083864,0.5448951,-0.8314697,-0.0,0.5555702,-0.8314697,-0.0,0.7071068,-0.7071068,-0.0,0.8314697,-0.5555702,-0.0,0.7071068,-0.7071068,0.1622099,0.8154936,-0.5555702,-0.0,0.7071068,-0.7071068,0.1379482,0.6935202,-0.7071068,0.1622099,0.8154936,-0.5555702,0.1622099,0.8154936,-0.5555702,0.1379482,0.6935202,-0.7071068,0.3181881,0.7681785,-0.5555702,0.1379482,0.6935202,-0.7071068,0.2705967,0.653282,-0.7071068,0.3181881,0.7681785,-0.5555702,0.3181881,0.7681785,-0.5555702,0.2705967,0.653282,-0.7071068,0.4619384,0.6913427,-0.5555702,0.2705967,0.653282,-0.7071068,0.3928463,0.5879386,-0.7071068,0.4619384,0.6913427,-0.5555702,0.4619384,0.6913427,-0.5555702,0.3928463,0.5879386,-0.7071068,0.5879367,0.587939,-0.5555702,0.3928463,0.5879386,-0.7071068,0.499999,0.500001,-0.7071068,0.5879367,0.587939,-0.5555702,0.5879367,0.587939,-0.5555702,0.499999,0.500001,-0.7071068,0.6913409,0.4619411,-0.5555702,0.499999,0.500001,-0.7071068,0.5879371,0.3928486,-0.7071068,0.6913409,0.4619411,-0.5555702,0.6913409,0.4619411,-0.5555702,0.5879371,0.3928486,-0.7071068,0.7681772,0.318191,-0.5555702,0.5879371,0.3928486,-0.7071068,0.653281,0.2705992,-0.7071068,0.7681772,0.318191,-0.5555702,0.7681772,0.318191,-0.5555702,0.653281,0.2705992,-0.7071068,0.8154929,0.1622131,-0.5555702,0.653281,0.2705992,-0.7071068,0.6935197,0.1379509,-0.7071068,0.8154929,0.1622131,-0.5555702,0.8154929,0.1622131,-0.5555702,0.6935197,0.1379509,-0.7071068,0.8314697,0.0000013,-0.5555702,0.6935197,0.1379509,-0.7071068,0.7071068,0.0000011,-0.7071068,0.8314697,0.0000013,-0.5555702,0.8314697,0.0000013,-0.5555702,0.7071068,0.0000011,-0.7071068,0.8154934,-0.1622104,-0.5555702,0.7071068,0.0000011,-0.7071068,0.6935201,-0.1379486,-0.7071068,0.8154934,-0.1622104,-0.5555702,0.8154934,-0.1622104,-0.5555702,0.6935201,-0.1379486,-0.7071068,0.7681783,-0.3181885,-0.5555702,0.6935201,-0.1379486,-0.7071068,0.6532819,-0.2705971,-0.7071068,0.7681783,-0.3181885,-0.5555702,0.7681783,-0.3181885,-0.5555702,0.6532819,-0.2705971,-0.7071068,0.6913424,-0.4619389,-0.5555702,0.6532819,-0.2705971,-0.7071068,0.5879383,-0.3928467,-0.7071068,0.6913424,-0.4619389,-0.5555702,0.6913424,-0.4619389,-0.5555702,0.5879383,-0.3928467,-0.7071068,0.5879385,-0.5879371,-0.5555702,0.5879383,-0.3928467,-0.7071068,0.5000006,-0.4999993,-0.7071068,0.5879385,-0.5879371,-0.5555702,0.5879385,-0.5879371,-0.5555702,0.5000006,-0.4999993,-0.7071068,0.4619406,-0.6913412,-0.5555702,0.5000006,-0.4999993,-0.7071068,0.3928482,-0.5879373,-0.7071068,0.4619406,-0.6913412,-0.5555702,0.4619406,-0.6913412,-0.5555702,0.3928482,-0.5879373,-0.7071068,0.3181905,-0.7681774,-0.5555702,0.3928482,-0.5879373,-0.7071068,0.2705988,-0.6532812,-0.7071068,0.3181905,-0.7681774,-0.5555702,0.3181905,-0.7681774,-0.5555702,0.2705988,-0.6532812,-0.7071068,0.1622125,-0.815493,-0.5555702,0.2705988,-0.6532812,-0.7071068,0.1379504,-0.6935198,-0.7071068,0.1622125,-0.815493,-0.5555702,0.1622125,-0.815493,-0.5555702,0.1379504,-0.6935198,-0.7071068,0.0000008,-0.8314697,-0.5555702,0.1379504,-0.6935198,-0.7071068,0.0000007,-0.7071068,-0.7071068,0.0000008,-0.8314697,-0.5555702,0.0000008,-0.8314697,-0.5555702,0.0000007,-0.7071068,-0.7071068,-0.162211,-0.8154933,-0.5555702,0.0000007,-0.7071068,-0.7071068,-0.1379491,-0.69352,-0.7071068,-0.162211,-0.8154933,-0.5555702,-0.162211,-0.8154933,-0.5555702,-0.1379491,-0.69352,-0.7071068,-0.318189,-0.768178,-0.5555702,-0.1379491,-0.69352,-0.7071068,-0.2705975,-0.6532817,-0.7071068,-0.318189,-0.768178,-0.5555702,-0.318189,-0.768178,-0.5555702,-0.2705975,-0.6532817,-0.7071068,-0.4619393,-0.6913421,-0.5555702,-0.2705975,-0.6532817,-0.7071068,-0.3928471,-0.5879381,-0.7071068,-0.4619393,-0.6913421,-0.5555702,-0.4619393,-0.6913421,-0.5555702,-0.3928471,-0.5879381,-0.7071068,-0.5879375,-0.5879382,-0.5555702,-0.3928471,-0.5879381,-0.7071068,-0.4999997,-0.5000003,-0.7071068,-0.5879375,-0.5879382,-0.5555702,-0.5879375,-0.5879382,-0.5555702,-0.4999997,-0.5000003,-0.7071068,-0.6913415,-0.4619402,-0.5555702,-0.4999997,-0.5000003,-0.7071068,-0.5879376,-0.3928478,-0.7071068,-0.6913415,-0.4619402,-0.5555702,-0.6913415,-0.4619402,-0.5555702,-0.5879376,-0.3928478,-0.7071068,-0.7681776,-0.31819,-0.5555702,-0.5879376,-0.3928478,-0.7071068,-0.6532813,-0.2705984,-0.7071068,-0.7681776,-0.31819,-0.5555702,-0.7681776,-0.31819,-0.5555702,-0.6532813,-0.2705984,-0.7071068,-0.8154931,-0.162212,-0.5555702,-0.6532813,-0.2705984,-0.7071068,-0.6935198,-0.13795,-0.7071068,-0.8154931,-0.162212,-0.5555702,-0.8154931,-0.162212,-0.5555702,-0.6935198,-0.13795,-0.7071068,-0.8314697,-0.0000003,-0.5555702,-0.6935198,-0.13795,-0.7071068,-0.7071068,-0.0000002,-0.7071068,-0.8314697,-0.0000003,-0.5555702,-0.8314697,-0.0000003,-0.5555702,-0.7071068,-0.0000002,-0.7071068,-0.8154932,0.1622115,-0.5555702,-0.7071068,-0.0000002,-0.7071068,-0.6935199,0.1379495,-0.7071068,-0.8154932,0.1622115,-0.5555702,-0.8154932,0.1622115,-0.5555702,-0.6935199,0.1379495,-0.7071068,-0.7681779,0.3181895,-0.5555702,-0.6935199,0.1379495,-0.7071068,-0.6532815,0.2705979,-0.7071068,-0.7681779,0.3181895,-0.5555702,-0.7681779,0.3181895,-0.5555702,-0.6532815,0.2705979,-0.7071068,-0.6913418,0.4619398,-0.5555702,-0.6532815,0.2705979,-0.7071068,-0.5879378,0.3928474,-0.7071068,-0.6913418,0.4619398,-0.5555702,-0.6913418,0.4619398,-0.5555702,-0.5879378,0.3928474,-0.7071068,-0.5879378,0.5879378,-0.5555702,-0.5879378,0.3928474,-0.7071068,-0.5,0.5,-0.7071068,-0.5879378,0.5879378,-0.5555702,-0.5879378,0.5879378,-0.5555702,-0.5,0.5,-0.7071068,-0.4619398,0.6913418,-0.5555702,-0.5,0.5,-0.7071068,-0.3928474,0.5879378,-0.7071068,-0.4619398,0.6913418,-0.5555702,-0.4619398,0.6913418,-0.5555702,-0.3928474,0.5879378,-0.7071068,-0.3181896,0.7681778,-0.5555702,-0.3928474,0.5879378,-0.7071068,-0.270598,0.6532815,-0.7071068,-0.3181896,0.7681778,-0.5555702,-0.3181896,0.7681778,-0.5555702,-0.270598,0.6532815,-0.7071068,-0.1622117,0.8154932,-0.5555702,-0.270598,0.6532815,-0.7071068,-0.1379497,0.6935199,-0.7071068,-0.1622117,0.8154932,-0.5555702,-0.1622117,0.8154932,-0.5555702,-0.1379497,0.6935199,-0.7071068,-0.0,0.8314697,-0.5555702,-0.1379497,0.6935199,-0.7071068,-0.0,0.7071068,-0.7071068,-0.0,0.8314697,-0.5555702,-0.0,0.9238796,-0.3826833,-0.0,0.8314697,-0.5555702,0.180238,0.9061279,-0.3826833,-0.0,0.8314697,-0.5555702,0.1622099,0.8154936,-0.5555702,0.180238,0.9061279,-0.3826833,0.180238,0.9061279,-0.3826833,0.1622099,0.8154936,-0.5555702,0.3535516,0.8535542,-0.3826833,0.1622099,0.8154936,-0.5555702,0.3181881,0.7681785,-0.5555702,0.3535516,0.8535542,-0.3826833,0.3535516,0.8535542,-0.3826833,0.3181881,0.7681785,-0.5555702,0.5132784,0.7681788,-0.3826833,0.3181881,0.7681785,-0.5555702,0.4619384,0.6913427,-0.5555702,0.5132784,0.7681788,-0.3826833,0.5132784,0.7681788,-0.3826833,0.4619384,0.6913427,-0.5555702,0.6532802,0.6532828,-0.3826833,0.4619384,0.6913427,-0.5555702,0.5879367,0.587939,-0.5555702,0.6532802,0.6532828,-0.3826833,0.6532802,0.6532828,-0.3826833,0.5879367,0.587939,-0.5555702,0.7681769,0.5132814,-0.3826833,0.5879367,0.587939,-0.5555702,0.6913409,0.4619411,-0.5555702,0.7681769,0.5132814,-0.3826833,0.7681769,0.5132814,-0.3826833,0.6913409,0.4619411,-0.5555702,0.8535528,0.3535549,-0.3826833,0.6913409,0.4619411,-0.5555702,0.7681772,0.318191,-0.5555702,0.8535528,0.3535549,-0.3826833,0.8535528,0.3535549,-0.3826833,0.7681772,0.318191,-0.5555702,0.9061272,0.1802415,-0.3826833,0.7681772,0.318191,-0.5555702,0.8154929,0.1622131,-0.5555702,0.9061272,0.1802415,-0.3826833,0.9061272,0.1802415,-0.3826833,0.8154929,0.1622131,-0.5555702,0.9238796,0.0000015,-0.3826833,0.8154929,0.1622131,-0.5555702,0.8314697,0.0000013,-0.5555702,0.9238796,0.0000015,-0.3826833,0.9238796,0.0000015,-0.3826833,0.8314697,0.0000013,-0.5555702,0.9061278,-0.1802386,-0.3826833,0.8314697,0.0000013,-0.5555702,0.8154934,-0.1622104,-0.5555702,0.9061278,-0.1802386,-0.3826833,0.9061278,-0.1802386,-0.3826833,0.8154934,-0.1622104,-0.5555702,0.853554,-0.3535522,-0.3826833,0.8154934,-0.1622104,-0.5555702,0.7681783,-0.3181885,-0.5555702,0.853554,-0.3535522,-0.3826833,0.853554,-0.3535522,-0.3826833,0.7681783,-0.3181885,-0.5555702,0.7681785,-0.513279,-0.3826833,0.7681783,-0.3181885,-0.5555702,0.6913424,-0.4619389,-0.5555702,0.7681785,-0.513279,-0.3826833,0.7681785,-0.513279,-0.3826833,0.6913424,-0.4619389,-0.5555702,0.6532823,-0.6532807,-0.3826833,0.6913424,-0.4619389,-0.5555702,0.5879385,-0.5879371,-0.5555702,0.6532823,-0.6532807,-0.3826833,0.6532823,-0.6532807,-0.3826833,0.5879385,-0.5879371,-0.5555702,0.5132809,-0.7681772,-0.3826833,0.5879385,-0.5879371,-0.5555702,0.4619406,-0.6913412,-0.5555702,0.5132809,-0.7681772,-0.3826833,0.5132809,-0.7681772,-0.3826833,0.4619406,-0.6913412,-0.5555702,0.3535544,-0.853553,-0.3826833,0.4619406,-0.6913412,-0.5555702,0.3181905,-0.7681774,-0.5555702,0.3535544,-0.853553,-0.3826833,0.3535544,-0.853553,-0.3826833,0.3181905,-0.7681774,-0.5555702,0.1802409,-0.9061273,-0.3826833,0.3181905,-0.7681774,-0.5555702,0.1622125,-0.815493,-0.5555702,0.1802409,-0.9061273,-0.3826833,0.1802409,-0.9061273,-0.3826833,0.1622125,-0.815493,-0.5555702,0.0000009,-0.9238796,-0.3826833,0.1622125,-0.815493,-0.5555702,0.0000008,-0.8314697,-0.5555702,0.0000009,-0.9238796,-0.3826833,0.0000009,-0.9238796,-0.3826833,0.0000008,-0.8314697,-0.5555702,-0.1802392,-0.9061276,-0.3826833,0.0000008,-0.8314697,-0.5555702,-0.162211,-0.8154933,-0.5555702,-0.1802392,-0.9061276,-0.3826833,-0.1802392,-0.9061276,-0.3826833,-0.162211,-0.8154933,-0.5555702,-0.3535527,-0.8535537,-0.3826833,-0.162211,-0.8154933,-0.5555702,-0.318189,-0.768178,-0.5555702,-0.3535527,-0.8535537,-0.3826833,-0.3535527,-0.8535537,-0.3826833,-0.318189,-0.768178,-0.5555702,-0.5132794,-0.7681782,-0.3826833,-0.318189,-0.768178,-0.5555702,-0.4619393,-0.6913421,-0.5555702,-0.5132794,-0.7681782,-0.3826833,-0.5132794,-0.7681782,-0.3826833,-0.4619393,-0.6913421,-0.5555702,-0.6532811,-0.6532819,-0.3826833,-0.4619393,-0.6913421,-0.5555702,-0.5879375,-0.5879382,-0.5555702,-0.6532811,-0.6532819,-0.3826833,-0.6532811,-0.6532819,-0.3826833,-0.5879375,-0.5879382,-0.5555702,-0.7681775,-0.5132805,-0.3826833,-0.5879375,-0.5879382,-0.5555702,-0.6913415,-0.4619402,-0.5555702,-0.7681775,-0.5132805,-0.3826833,-0.7681775,-0.5132805,-0.3826833,-0.6913415,-0.4619402,-0.5555702,-0.8535532,-0.3535538,-0.3826833,-0.6913415,-0.4619402,-0.5555702,-0.7681776,-0.31819,-0.5555702,-0.8535532,-0.3535538,-0.3826833,-0.8535532,-0.3535538,-0.3826833,-0.7681776,-0.31819,-0.5555702,-0.9061274,-0.1802403,-0.3826833,-0.7681776,-0.31819,-0.5555702,-0.8154931,-0.162212,-0.5555702,-0.9061274,-0.1802403,-0.3826833,-0.9061274,-0.1802403,-0.3826833,-0.8154931,-0.162212,-0.5555702,-0.9238796,-0.0000003,-0.3826833,-0.8154931,-0.162212,-0.5555702,-0.8314697,-0.0000003,-0.5555702,-0.9238796,-0.0000003,-0.3826833,-0.9238796,-0.0000003,-0.3826833,-0.8314697,-0.0000003,-0.5555702,-0.9061275,0.1802397,-0.3826833,-0.8314697,-0.0000003,-0.5555702,-0.8154932,0.1622115,-0.5555702,-0.9061275,0.1802397,-0.3826833,-0.9061275,0.1802397,-0.3826833,-0.8154932,0.1622115,-0.5555702,-0.8535535,0.3535533,-0.3826833,-0.8154932,0.1622115,-0.5555702,-0.7681779,0.3181895,-0.5555702,-0.8535535,0.3535533,-0.3826833,-0.8535535,0.3535533,-0.3826833,-0.7681779,0.3181895,-0.5555702,-0.7681778,0.5132799,-0.3826833,-0.7681779,0.3181895,-0.5555702,-0.6913418,0.4619398,-0.5555702,-0.7681778,0.5132799,-0.3826833,-0.7681778,0.5132799,-0.3826833,-0.6913418,0.4619398,-0.5555702,-0.6532815,0.6532815,-0.3826833,-0.6913418,0.4619398,-0.5555702,-0.5879378,0.5879378,-0.5555702,-0.6532815,0.6532815,-0.3826833,-0.6532815,0.6532815,-0.3826833,-0.5879378,0.5879378,-0.5555702,-0.5132799,0.7681778,-0.3826833,-0.5879378,0.5879378,-0.5555702,-0.4619398,0.6913418,-0.5555702,-0.5132799,0.7681778,-0.3826833,-0.5132799,0.7681778,-0.3826833,-0.4619398,0.6913418,-0.5555702,-0.3535534,0.8535535,-0.3826833,-0.4619398,0.6913418,-0.5555702,-0.3181896,0.7681778,-0.5555702,-0.3535534,0.8535535,-0.3826833,-0.3535534,0.8535535,-0.3826833,-0.3181896,0.7681778,-0.5555702,-0.18024,0.9061275,-0.3826833,-0.3181896,0.7681778,-0.5555702,-0.1622117,0.8154932,-0.5555702,-0.18024,0.9061275,-0.3826833,-0.18024,0.9061275,-0.3826833,-0.1622117,0.8154932,-0.5555702,-0.0,0.9238796,-0.3826833,-0.1622117,0.8154932,-0.5555702,-0.0,0.8314697,-0.5555702,-0.0,0.9238796,-0.3826833,-0.0,0.9807853,-0.1950902,-0.0,0.9238796,-0.3826833,0.1913396,0.9619402,-0.1950902,-0.0,0.9238796,-0.3826833,0.180238,0.9061279,-0.3826833,0.1913396,0.9619402,-0.1950902,0.1913396,0.9619402,-0.1950902,0.180238,0.9061279,-0.3826833,0.3753284,0.9061283,-0.1950902,0.180238,0.9061279,-0.3826833,0.3535516,0.8535542,-0.3826833,0.3753284,0.9061283,-0.1950902,0.3753284,0.9061283,-0.1950902,0.3535516,0.8535542,-0.3826833,0.5448935,0.8154943,-0.1950902,0.3535516,0.8535542,-0.3826833,0.5132784,0.7681788,-0.3826833,0.5448935,0.8154943,-0.1950902,0.5448935,0.8154943,-0.1950902,0.5132784,0.7681788,-0.3826833,0.6935186,0.6935213,-0.1950902,0.5132784,0.7681788,-0.3826833,0.6532802,0.6532828,-0.3826833,0.6935186,0.6935213,-0.1950902,0.6935186,0.6935213,-0.1950902,0.6532802,0.6532828,-0.3826833,0.8154922,0.5448967,-0.1950902,0.6532802,0.6532828,-0.3826833,0.7681769,0.5132814,-0.3826833,0.8154922,0.5448967,-0.1950902,0.8154922,0.5448967,-0.1950902,0.7681769,0.5132814,-0.3826833,0.9061268,0.3753319,-0.1950902,0.7681769,0.5132814,-0.3826833,0.8535528,0.3535549,-0.3826833,0.9061268,0.3753319,-0.1950902,0.9061268,0.3753319,-0.1950902,0.8535528,0.3535549,-0.3826833,0.9619395,0.1913434,-0.1950902,0.8535528,0.3535549,-0.3826833,0.9061272,0.1802415,-0.3826833,0.9619395,0.1913434,-0.1950902,0.9619395,0.1913434,-0.1950902,0.9061272,0.1802415,-0.3826833,0.9807853,0.0000016,-0.1950902,0.9061272,0.1802415,-0.3826833,0.9238796,0.0000015,-0.3826833,0.9807853,0.0000016,-0.1950902,0.9807853,0.0000016,-0.1950902,0.9238796,0.0000015,-0.3826833,0.9619401,-0.1913403,-0.1950902,0.9238796,0.0000015,-0.3826833,0.9061278,-0.1802386,-0.3826833,0.9619401,-0.1913403,-0.1950902,0.9619401,-0.1913403,-0.1950902,0.9061278,-0.1802386,-0.3826833,0.906128,-0.375329,-0.1950902,0.9061278,-0.1802386,-0.3826833,0.853554,-0.3535522,-0.3826833,0.906128,-0.375329,-0.1950902,0.906128,-0.375329,-0.1950902,0.853554,-0.3535522,-0.3826833,0.8154939,-0.544894,-0.1950902,0.853554,-0.3535522,-0.3826833,0.7681785,-0.513279,-0.3826833,0.8154939,-0.544894,-0.1950902,0.8154939,-0.544894,-0.1950902,0.7681785,-0.513279,-0.3826833,0.6935208,-0.6935191,-0.1950902,0.7681785,-0.513279,-0.3826833,0.6532823,-0.6532807,-0.3826833,0.6935208,-0.6935191,-0.1950902,0.6935208,-0.6935191,-0.1950902,0.6532823,-0.6532807,-0.3826833,0.5448961,-0.8154925,-0.1950902,0.6532823,-0.6532807,-0.3826833,0.5132809,-0.7681772,-0.3826833,0.5448961,-0.8154925,-0.1950902,0.5448961,-0.8154925,-0.1950902,0.5132809,-0.7681772,-0.3826833,0.3753313,-0.906127,-0.1950902,0.5132809,-0.7681772,-0.3826833,0.3535544,-0.853553,-0.3826833,0.3753313,-0.906127,-0.1950902,0.3753313,-0.906127,-0.1950902,0.3535544,-0.853553,-0.3826833,0.1913427,-0.9619396,-0.1950902,0.3535544,-0.853553,-0.3826833,0.1802409,-0.9061273,-0.3826833,0.1913427,-0.9619396,-0.1950902,0.1913427,-0.9619396,-0.1950902,0.1802409,-0.9061273,-0.3826833,0.0000009,-0.9807853,-0.1950902,0.1802409,-0.9061273,-0.3826833,0.0000009,-0.9238796,-0.3826833,0.0000009,-0.9807853,-0.1950902,0.0000009,-0.9807853,-0.1950902,0.0000009,-0.9238796,-0.3826833,-0.1913409,-0.9619399,-0.1950902,0.0000009,-0.9238796,-0.3826833,-0.1802392,-0.9061276,-0.3826833,-0.1913409,-0.9619399,-0.1950902,-0.1913409,-0.9619399,-0.1950902,-0.1802392,-0.9061276,-0.3826833,-0.3753296,-0.9061278,-0.1950902,-0.1802392,-0.9061276,-0.3826833,-0.3535527,-0.8535537,-0.3826833,-0.3753296,-0.9061278,-0.1950902,-0.3753296,-0.9061278,-0.1950902,-0.3535527,-0.8535537,-0.3826833,-0.5448946,-0.8154936,-0.1950902,-0.3535527,-0.8535537,-0.3826833,-0.5132794,-0.7681782,-0.3826833,-0.5448946,-0.8154936,-0.1950902,-0.5448946,-0.8154936,-0.1950902,-0.5132794,-0.7681782,-0.3826833,-0.6935195,-0.6935204,-0.1950902,-0.5132794,-0.7681782,-0.3826833,-0.6532811,-0.6532819,-0.3826833,-0.6935195,-0.6935204,-0.1950902,-0.6935195,-0.6935204,-0.1950902,-0.6532811,-0.6532819,-0.3826833,-0.8154929,-0.5448956,-0.1950902,-0.6532811,-0.6532819,-0.3826833,-0.7681775,-0.5132805,-0.3826833,-0.8154929,-0.5448956,-0.1950902,-0.8154929,-0.5448956,-0.1950902,-0.7681775,-0.5132805,-0.3826833,-0.9061273,-0.3753307,-0.1950902,-0.7681775,-0.5132805,-0.3826833,-0.8535532,-0.3535538,-0.3826833,-0.9061273,-0.3753307,-0.1950902,-0.9061273,-0.3753307,-0.1950902,-0.8535532,-0.3535538,-0.3826833,-0.9619397,-0.1913421,-0.1950902,-0.8535532,-0.3535538,-0.3826833,-0.9061274,-0.1802403,-0.3826833,-0.9619397,-0.1913421,-0.1950902,-0.9619397,-0.1913421,-0.1950902,-0.9061274,-0.1802403,-0.3826833,-0.9807853,-0.0000003,-0.1950902,-0.9061274,-0.1802403,-0.3826833,-0.9238796,-0.0000003,-0.3826833,-0.9807853,-0.0000003,-0.1950902,-0.9807853,-0.0000003,-0.1950902,-0.9238796,-0.0000003,-0.3826833,-0.9619398,0.1913415,-0.1950902,-0.9238796,-0.0000003,-0.3826833,-0.9061275,0.1802397,-0.3826833,-0.9619398,0.1913415,-0.1950902,-0.9619398,0.1913415,-0.1950902,-0.9061275,0.1802397,-0.3826833,-0.9061276,0.3753302,-0.1950902,-0.9061275,0.1802397,-0.3826833,-0.8535535,0.3535533,-0.3826833,-0.9061276,0.3753302,-0.1950902,-0.9061276,0.3753302,-0.1950902,-0.8535535,0.3535533,-0.3826833,-0.8154932,0.5448951,-0.1950902,-0.8535535,0.3535533,-0.3826833,-0.7681778,0.5132799,-0.3826833,-0.8154932,0.5448951,-0.1950902,-0.8154932,0.5448951,-0.1950902,-0.7681778,0.5132799,-0.3826833,-0.6935199,0.6935199,-0.1950902,-0.7681778,0.5132799,-0.3826833,-0.6532815,0.6532815,-0.3826833,-0.6935199,0.6935199,-0.1950902,-0.6935199,0.6935199,-0.1950902,-0.6532815,0.6532815,-0.3826833,-0.5448951,0.8154932,-0.1950902,-0.6532815,0.6532815,-0.3826833,-0.5132799,0.7681778,-0.3826833,-0.5448951,0.8154932,-0.1950902,-0.5448951,0.8154932,-0.1950902,-0.5132799,0.7681778,-0.3826833,-0.3753302,0.9061275,-0.1950902,-0.5132799,0.7681778,-0.3826833,-0.3535534,0.8535535,-0.3826833,-0.3753302,0.9061275,-0.1950902,-0.3753302,0.9061275,-0.1950902,-0.3535534,0.8535535,-0.3826833,-0.1913417,0.9619398,-0.1950902,-0.3535534,0.8535535,-0.3826833,-0.18024,0.9061275,-0.3826833,-0.1913417,0.9619398,-0.1950902,-0.1913417,0.9619398,-0.1950902,-0.18024,0.9061275,-0.3826833,-0.0,0.9807853,-0.1950902,-0.18024,0.9061275,-0.3826833,-0.0,0.9238796,-0.3826833,-0.0,0.9807853,-0.1950902,-0.0,1.0,0.0000001,-0.0,0.9807853,-0.1950902,0.1950882,0.9807857,0.0000001,-0.0,0.9807853,-0.1950902,0.1913396,0.9619402,-0.1950902,0.1950882,0.9807857,0.0000001,0.1950882,0.9807857,0.0000001,0.1913396,0.9619402,-0.1950902,0.3826815,0.9238803,0.0000001,0.1913396,0.9619402,-0.1950902,0.3753284,0.9061283,-0.1950902,0.3826815,0.9238803,0.0000001,0.3826815,0.9238803,0.0000001,0.3753284,0.9061283,-0.1950902,0.5555686,0.8314707,0.0000001,0.3753284,0.9061283,-0.1950902,0.5448935,0.8154943,-0.1950902,0.5555686,0.8314707,0.0000001,0.5555686,0.8314707,0.0000001,0.5448935,0.8154943,-0.1950902,0.7071054,0.7071081,0.0000001,0.5448935,0.8154943,-0.1950902,0.6935186,0.6935213,-0.1950902,0.7071054,0.7071081,0.0000001,0.7071054,0.7071081,0.0000001,0.6935186,0.6935213,-0.1950902,0.8314686,0.5555718,0.0000001,0.6935186,0.6935213,-0.1950902,0.8154922,0.5448967,-0.1950902,0.8314686,0.5555718,0.0000001,0.8314686,0.5555718,0.0000001,0.8154922,0.5448967,-0.1950902,0.9238788,0.3826851,0.0000001,0.8154922,0.5448967,-0.1950902,0.9061268,0.3753319,-0.1950902,0.9238788,0.3826851,0.0000001,0.9238788,0.3826851,0.0000001,0.9061268,0.3753319,-0.1950902,0.980785,0.195092,0.0000001,0.9061268,0.3753319,-0.1950902,0.9619395,0.1913434,-0.1950902,0.980785,0.195092,0.0000001,0.980785,0.195092,0.0000001,0.9619395,0.1913434,-0.1950902,1.0,0.0000016,0.0000001,0.9619395,0.1913434,-0.1950902,0.9807853,0.0000016,-0.1950902,1.0,0.0000016,0.0000001,1.0,0.0000016,0.0000001,0.9807853,0.0000016,-0.1950902,0.9807855,-0.1950888,0.0000001,0.9807853,0.0000016,-0.1950902,0.9619401,-0.1913403,-0.1950902,0.9807855,-0.1950888,0.0000001,0.9807855,-0.1950888,0.0000001,0.9619401,-0.1913403,-0.1950902,0.9238801,-0.3826821,0.0000001,0.9619401,-0.1913403,-0.1950902,0.906128,-0.375329,-0.1950902,0.9238801,-0.3826821,0.0000001,0.9238801,-0.3826821,0.0000001,0.906128,-0.375329,-0.1950902,0.8314704,-0.5555691,0.0000001,0.906128,-0.375329,-0.1950902,0.8154939,-0.544894,-0.1950902,0.8314704,-0.5555691,0.0000001,0.8314704,-0.5555691,0.0000001,0.8154939,-0.544894,-0.1950902,0.7071077,-0.7071059,0.0000001,0.8154939,-0.544894,-0.1950902,0.6935208,-0.6935191,-0.1950902,0.7071077,-0.7071059,0.0000001,0.7071077,-0.7071059,0.0000001,0.6935208,-0.6935191,-0.1950902,0.5555713,-0.8314689,0.0000001,0.6935208,-0.6935191,-0.1950902,0.5448961,-0.8154925,-0.1950902,0.5555713,-0.8314689,0.0000001,0.5555713,-0.8314689,0.0000001,0.5448961,-0.8154925,-0.1950902,0.3826845,-0.9238791,0.0000001,0.5448961,-0.8154925,-0.1950902,0.3753313,-0.906127,-0.1950902,0.3826845,-0.9238791,0.0000001,0.3826845,-0.9238791,0.0000001,0.3753313,-0.906127,-0.1950902,0.1950914,-0.9807851,0.0000001,0.3753313,-0.906127,-0.1950902,0.1913427,-0.9619396,-0.1950902,0.1950914,-0.9807851,0.0000001,0.1950914,-0.9807851,0.0000001,0.1913427,-0.9619396,-0.1950902,0.000001,-1.0,0.0000001,0.1913427,-0.9619396,-0.1950902,0.0000009,-0.9807853,-0.1950902,0.000001,-1.0,0.0000001,0.000001,-1.0,0.0000001,0.0000009,-0.9807853,-0.1950902,-0.1950895,-0.9807854,0.0000001,0.0000009,-0.9807853,-0.1950902,-0.1913409,-0.9619399,-0.1950902,-0.1950895,-0.9807854,0.0000001,-0.1950895,-0.9807854,0.0000001,-0.1913409,-0.9619399,-0.1950902,-0.3826827,-0.9238799,0.0000001,-0.1913409,-0.9619399,-0.1950902,-0.3753296,-0.9061278,-0.1950902,-0.3826827,-0.9238799,0.0000001,-0.3826827,-0.9238799,0.0000001,-0.3753296,-0.9061278,-0.1950902,-0.5555696,-0.83147,0.0000001,-0.3753296,-0.9061278,-0.1950902,-0.5448946,-0.8154936,-0.1950902,-0.5555696,-0.83147,0.0000001,-0.5555696,-0.83147,0.0000001,-0.5448946,-0.8154936,-0.1950902,-0.7071064,-0.7071072,0.0000001,-0.5448946,-0.8154936,-0.1950902,-0.6935195,-0.6935204,-0.1950902,-0.7071064,-0.7071072,0.0000001,-0.7071064,-0.7071072,0.0000001,-0.6935195,-0.6935204,-0.1950902,-0.8314693,-0.5555707,0.0000001,-0.6935195,-0.6935204,-0.1950902,-0.8154929,-0.5448956,-0.1950902,-0.8314693,-0.5555707,0.0000001,-0.8314693,-0.5555707,0.0000001,-0.8154929,-0.5448956,-0.1950902,-0.9238793,-0.3826839,0.0000001,-0.8154929,-0.5448956,-0.1950902,-0.9061273,-0.3753307,-0.1950902,-0.9238793,-0.3826839,0.0000001,-0.9238793,-0.3826839,0.0000001,-0.9061273,-0.3753307,-0.1950902,-0.9807852,-0.1950907,0.0000001,-0.9061273,-0.3753307,-0.1950902,-0.9619397,-0.1913421,-0.1950902,-0.9807852,-0.1950907,0.0000001,-0.9807852,-0.1950907,0.0000001,-0.9619397,-0.1913421,-0.1950902,-1.0,-0.0000003,0.0000001,-0.9619397,-0.1913421,-0.1950902,-0.9807853,-0.0000003,-0.1950902,-1.0,-0.0000003,0.0000001,-1.0,-0.0000003,0.0000001,-0.9807853,-0.0000003,-0.1950902,-0.9807853,0.1950901,0.0000001,-0.9807853,-0.0000003,-0.1950902,-0.9619398,0.1913415,-0.1950902,-0.9807853,0.1950901,0.0000001,-0.9807853,0.1950901,0.0000001,-0.9619398,0.1913415,-0.1950902,-0.9238796,0.3826833,0.0000001,-0.9619398,0.1913415,-0.1950902,-0.9061276,0.3753302,-0.1950902,-0.9238796,0.3826833,0.0000001,-0.9238796,0.3826833,0.0000001,-0.9061276,0.3753302,-0.1950902,-0.8314697,0.5555702,0.0000001,-0.9061276,0.3753302,-0.1950902,-0.8154932,0.5448951,-0.1950902,-0.8314697,0.5555702,0.0000001,-0.8314697,0.5555702,0.0000001,-0.8154932,0.5448951,-0.1950902,-0.7071068,0.7071068,0.0000001,-0.8154932,0.5448951,-0.1950902,-0.6935199,0.6935199,-0.1950902,-0.7071068,0.7071068,0.0000001,-0.7071068,0.7071068,0.0000001,-0.6935199,0.6935199,-0.1950902,-0.5555702,0.8314697,0.0000001,-0.6935199,0.6935199,-0.1950902,-0.5448951,0.8154932,-0.1950902,-0.5555702,0.8314697,0.0000001,-0.5555702,0.8314697,0.0000001,-0.5448951,0.8154932,-0.1950902,-0.3826834,0.9238796,0.0000001,-0.5448951,0.8154932,-0.1950902,-0.3753302,0.9061275,-0.1950902,-0.3826834,0.9238796,0.0000001,-0.3826834,0.9238796,0.0000001,-0.3753302,0.9061275,-0.1950902,-0.1950903,0.9807853,0.0000001,-0.3753302,0.9061275,-0.1950902,-0.1913417,0.9619398,-0.1950902,-0.1950903,0.9807853,0.0000001,-0.1950903,0.9807853,0.0000001,-0.1913417,0.9619398,-0.1950902,-0.0,1.0,0.0000001,-0.1913417,0.9619398,-0.1950902,-0.0,0.9807853,-0.1950902,-0.0,1.0,0.0000001,-0.0,0.9807853,0.1950904,-0.0,1.0,0.0000001,0.1913396,0.9619402,0.1950904,-0.0,1.0,0.0000001,0.1950882,0.9807857,0.0000001,0.1913396,0.9619402,0.1950904,0.1913396,0.9619402,0.1950904,0.1950882,0.9807857,0.0000001,0.3753284,0.9061282,0.1950904,0.1950882,0.9807857,0.0000001,0.3826815,0.9238803,0.0000001,0.3753284,0.9061282,0.1950904,0.3753284,0.9061282,0.1950904,0.3826815,0.9238803,0.0000001,0.5448934,0.8154942,0.1950904,0.3826815,0.9238803,0.0000001,0.5555686,0.8314707,0.0000001,0.5448934,0.8154942,0.1950904,0.5448934,0.8154942,0.1950904,0.5555686,0.8314707,0.0000001,0.6935185,0.6935213,0.1950904,0.5555686,0.8314707,0.0000001,0.7071054,0.7071081,0.0000001,0.6935185,0.6935213,0.1950904,0.6935185,0.6935213,0.1950904,0.7071054,0.7071081,0.0000001,0.8154921,0.5448966,0.1950904,0.7071054,0.7071081,0.0000001,0.8314686,0.5555718,0.0000001,0.8154921,0.5448966,0.1950904,0.8154921,0.5448966,0.1950904,0.8314686,0.5555718,0.0000001,0.9061267,0.3753319,0.1950904,0.8314686,0.5555718,0.0000001,0.9238788,0.3826851,0.0000001,0.9061267,0.3753319,0.1950904,0.9061267,0.3753319,0.1950904,0.9238788,0.3826851,0.0000001,0.9619394,0.1913433,0.1950904,0.9238788,0.3826851,0.0000001,0.980785,0.195092,0.0000001,0.9619394,0.1913433,0.1950904,0.9619394,0.1913433,0.1950904,0.980785,0.195092,0.0000001,0.9807853,0.0000016,0.1950904,0.980785,0.195092,0.0000001,1.0,0.0000016,0.0000001,0.9807853,0.0000016,0.1950904,0.9807853,0.0000016,0.1950904,1.0,0.0000016,0.0000001,0.96194,-0.1913402,0.1950904,1.0,0.0000016,0.0000001,0.9807855,-0.1950888,0.0000001,0.96194,-0.1913402,0.1950904,0.96194,-0.1913402,0.1950904,0.9807855,-0.1950888,0.0000001,0.906128,-0.375329,0.1950904,0.9807855,-0.1950888,0.0000001,0.9238801,-0.3826821,0.0000001,0.906128,-0.375329,0.1950904,0.906128,-0.375329,0.1950904,0.9238801,-0.3826821,0.0000001,0.8154939,-0.544894,0.1950904,0.9238801,-0.3826821,0.0000001,0.8314704,-0.5555691,0.0000001,0.8154939,-0.544894,0.1950904,0.8154939,-0.544894,0.1950904,0.8314704,-0.5555691,0.0000001,0.6935208,-0.693519,0.1950904,0.8314704,-0.5555691,0.0000001,0.7071077,-0.7071059,0.0000001,0.6935208,-0.693519,0.1950904,0.6935208,-0.693519,0.1950904,0.7071077,-0.7071059,0.0000001,0.5448961,-0.8154925,0.1950904,0.7071077,-0.7071059,0.0000001,0.5555713,-0.8314689,0.0000001,0.5448961,-0.8154925,0.1950904,0.5448961,-0.8154925,0.1950904,0.5555713,-0.8314689,0.0000001,0.3753313,-0.906127,0.1950904,0.5555713,-0.8314689,0.0000001,0.3826845,-0.9238791,0.0000001,0.3753313,-0.906127,0.1950904,0.3753313,-0.906127,0.1950904,0.3826845,-0.9238791,0.0000001,0.1913427,-0.9619395,0.1950904,0.3826845,-0.9238791,0.0000001,0.1950914,-0.9807851,0.0000001,0.1913427,-0.9619395,0.1950904,0.1913427,-0.9619395,0.1950904,0.1950914,-0.9807851,0.0000001,0.0000009,-0.9807853,0.1950904,0.1950914,-0.9807851,0.0000001,0.000001,-1.0,0.0000001,0.0000009,-0.9807853,0.1950904,0.0000009,-0.9807853,0.1950904,0.000001,-1.0,0.0000001,-0.1913409,-0.9619399,0.1950904,0.000001,-1.0,0.0000001,-0.1950895,-0.9807854,0.0000001,-0.1913409,-0.9619399,0.1950904,-0.1913409,-0.9619399,0.1950904,-0.1950895,-0.9807854,0.0000001,-0.3753295,-0.9061278,0.1950904,-0.1950895,-0.9807854,0.0000001,-0.3826827,-0.9238799,0.0000001,-0.3753295,-0.9061278,0.1950904,-0.3753295,-0.9061278,0.1950904,-0.3826827,-0.9238799,0.0000001,-0.5448945,-0.8154935,0.1950904,-0.3826827,-0.9238799,0.0000001,-0.5555696,-0.83147,0.0000001,-0.5448945,-0.8154935,0.1950904,-0.5448945,-0.8154935,0.1950904,-0.5555696,-0.83147,0.0000001,-0.6935195,-0.6935204,0.1950904,-0.5555696,-0.83147,0.0000001,-0.7071064,-0.7071072,0.0000001,-0.6935195,-0.6935204,0.1950904,-0.6935195,-0.6935204,0.1950904,-0.7071064,-0.7071072,0.0000001,-0.8154928,-0.5448956,0.1950904,-0.7071064,-0.7071072,0.0000001,-0.8314693,-0.5555707,0.0000001,-0.8154928,-0.5448956,0.1950904,-0.8154928,-0.5448956,0.1950904,-0.8314693,-0.5555707,0.0000001,-0.9061272,-0.3753307,0.1950904,-0.8314693,-0.5555707,0.0000001,-0.9238793,-0.3826839,0.0000001,-0.9061272,-0.3753307,0.1950904,-0.9061272,-0.3753307,0.1950904,-0.9238793,-0.3826839,0.0000001,-0.9619396,-0.1913421,0.1950904,-0.9238793,-0.3826839,0.0000001,-0.9807852,-0.1950907,0.0000001,-0.9619396,-0.1913421,0.1950904,-0.9619396,-0.1913421,0.1950904,-0.9807852,-0.1950907,0.0000001,-0.9807853,-0.0000003,0.1950904,-0.9807852,-0.1950907,0.0000001,-1.0,-0.0000003,0.0000001,-0.9807853,-0.0000003,0.1950904,-0.9807853,-0.0000003,0.1950904,-1.0,-0.0000003,0.0000001,-0.9619398,0.1913415,0.1950904,-1.0,-0.0000003,0.0000001,-0.9807853,0.1950901,0.0000001,-0.9619398,0.1913415,0.1950904,-0.9619398,0.1913415,0.1950904,-0.9807853,0.1950901,0.0000001,-0.9061275,0.3753301,0.1950904,-0.9807853,0.1950901,0.0000001,-0.9238796,0.3826833,0.0000001,-0.9061275,0.3753301,0.1950904,-0.9061275,0.3753301,0.1950904,-0.9238796,0.3826833,0.0000001,-0.8154932,0.5448951,0.1950904,-0.9238796,0.3826833,0.0000001,-0.8314697,0.5555702,0.0000001,-0.8154932,0.5448951,0.1950904,-0.8154932,0.5448951,0.1950904,-0.8314697,0.5555702,0.0000001,-0.6935199,0.6935199,0.1950904,-0.8314697,0.5555702,0.0000001,-0.7071068,0.7071068,0.0000001,-0.6935199,0.6935199,0.1950904,-0.6935199,0.6935199,0.1950904,-0.7071068,0.7071068,0.0000001,-0.5448951,0.8154932,0.1950904,-0.7071068,0.7071068,0.0000001,-0.5555702,0.8314697,0.0000001,-0.5448951,0.8154932,0.1950904,-0.5448951,0.8154932,0.1950904,-0.5555702,0.8314697,0.0000001,-0.3753302,0.9061275,0.1950904,-0.5555702,0.8314697,0.0000001,-0.3826834,0.9238796,0.0000001,-0.3753302,0.9061275,0.1950904,-0.3753302,0.9061275,0.1950904,-0.3826834,0.9238796,0.0000001,-0.1913417,0.9619397,0.1950904,-0.3826834,0.9238796,0.0000001,-0.1950903,0.9807853,0.0000001,-0.1913417,0.9619397,0.1950904,-0.1913417,0.9619397,0.1950904,-0.1950903,0.9807853,0.0000001,-0.0,0.9807853,0.1950904,-0.1950903,0.9807853,0.0000001,-0.0,1.0,0.0000001,-0.0,0.9807853,0.1950904,-0.0,0.9238795,0.3826834,-0.0,0.9807853,0.1950904,0.180238,0.9061278,0.3826834,-0.0,0.9807853,0.1950904,0.1913396,0.9619402,0.1950904,0.180238,0.9061278,0.3826834,0.180238,0.9061278,0.3826834,0.1913396,0.9619402,0.1950904,0.3535516,0.8535541,0.3826834,0.1913396,0.9619402,0.1950904,0.3753284,0.9061282,0.1950904,0.3535516,0.8535541,0.3826834,0.3535516,0.8535541,0.3826834,0.3753284,0.9061282,0.1950904,0.5132784,0.7681788,0.3826834,0.3753284,0.9061282,0.1950904,0.5448934,0.8154942,0.1950904,0.5132784,0.7681788,0.3826834,0.5132784,0.7681788,0.3826834,0.5448934,0.8154942,0.1950904,0.6532802,0.6532827,0.3826834,0.5448934,0.8154942,0.1950904,0.6935185,0.6935213,0.1950904,0.6532802,0.6532827,0.3826834,0.6532802,0.6532827,0.3826834,0.6935185,0.6935213,0.1950904,0.7681768,0.5132814,0.3826834,0.6935185,0.6935213,0.1950904,0.8154921,0.5448966,0.1950904,0.7681768,0.5132814,0.3826834,0.7681768,0.5132814,0.3826834,0.8154921,0.5448966,0.1950904,0.8535528,0.3535549,0.3826834,0.8154921,0.5448966,0.1950904,0.9061267,0.3753319,0.1950904,0.8535528,0.3535549,0.3826834,0.8535528,0.3535549,0.3826834,0.9061267,0.3753319,0.1950904,0.9061271,0.1802415,0.3826834,0.9061267,0.3753319,0.1950904,0.9619394,0.1913433,0.1950904,0.9061271,0.1802415,0.3826834,0.9061271,0.1802415,0.3826834,0.9619394,0.1913433,0.1950904,0.9238795,0.0000015,0.3826834,0.9619394,0.1913433,0.1950904,0.9807853,0.0000016,0.1950904,0.9238795,0.0000015,0.3826834,0.9238795,0.0000015,0.3826834,0.9807853,0.0000016,0.1950904,0.9061277,-0.1802386,0.3826834,0.9807853,0.0000016,0.1950904,0.96194,-0.1913402,0.1950904,0.9061277,-0.1802386,0.3826834,0.9061277,-0.1802386,0.3826834,0.96194,-0.1913402,0.1950904,0.8535539,-0.3535521,0.3826834,0.96194,-0.1913402,0.1950904,0.906128,-0.375329,0.1950904,0.8535539,-0.3535521,0.3826834,0.8535539,-0.3535521,0.3826834,0.906128,-0.375329,0.1950904,0.7681785,-0.5132789,0.3826834,0.906128,-0.375329,0.1950904,0.8154939,-0.544894,0.1950904,0.7681785,-0.5132789,0.3826834,0.7681785,-0.5132789,0.3826834,0.8154939,-0.544894,0.1950904,0.6532823,-0.6532806,0.3826834,0.8154939,-0.544894,0.1950904,0.6935208,-0.693519,0.1950904,0.6532823,-0.6532806,0.3826834,0.6532823,-0.6532806,0.3826834,0.6935208,-0.693519,0.1950904,0.5132809,-0.7681771,0.3826834,0.6935208,-0.693519,0.1950904,0.5448961,-0.8154925,0.1950904,0.5132809,-0.7681771,0.3826834,0.5132809,-0.7681771,0.3826834,0.5448961,-0.8154925,0.1950904,0.3535543,-0.8535529,0.3826834,0.5448961,-0.8154925,0.1950904,0.3753313,-0.906127,0.1950904,0.3535543,-0.8535529,0.3826834,0.3535543,-0.8535529,0.3826834,0.3753313,-0.906127,0.1950904,0.1802409,-0.9061272,0.3826834,0.3753313,-0.906127,0.1950904,0.1913427,-0.9619395,0.1950904,0.1802409,-0.9061272,0.3826834,0.1802409,-0.9061272,0.3826834,0.1913427,-0.9619395,0.1950904,0.0000009,-0.9238795,0.3826834,0.1913427,-0.9619395,0.1950904,0.0000009,-0.9807853,0.1950904,0.0000009,-0.9238795,0.3826834,0.0000009,-0.9238795,0.3826834,0.0000009,-0.9807853,0.1950904,-0.1802392,-0.9061276,0.3826834,0.0000009,-0.9807853,0.1950904,-0.1913409,-0.9619399,0.1950904,-0.1802392,-0.9061276,0.3826834,-0.1802392,-0.9061276,0.3826834,-0.1913409,-0.9619399,0.1950904,-0.3535527,-0.8535537,0.3826834,-0.1913409,-0.9619399,0.1950904,-0.3753295,-0.9061278,0.1950904,-0.3535527,-0.8535537,0.3826834,-0.3535527,-0.8535537,0.3826834,-0.3753295,-0.9061278,0.1950904,-0.5132794,-0.7681781,0.3826834,-0.3753295,-0.9061278,0.1950904,-0.5448945,-0.8154935,0.1950904,-0.5132794,-0.7681781,0.3826834,-0.5132794,-0.7681781,0.3826834,-0.5448945,-0.8154935,0.1950904,-0.6532811,-0.6532819,0.3826834,-0.5448945,-0.8154935,0.1950904,-0.6935195,-0.6935204,0.1950904,-0.6532811,-0.6532819,0.3826834,-0.6532811,-0.6532819,0.3826834,-0.6935195,-0.6935204,0.1950904,-0.7681774,-0.5132804,0.3826834,-0.6935195,-0.6935204,0.1950904,-0.8154928,-0.5448956,0.1950904,-0.7681774,-0.5132804,0.3826834,-0.7681774,-0.5132804,0.3826834,-0.8154928,-0.5448956,0.1950904,-0.8535532,-0.3535538,0.3826834,-0.8154928,-0.5448956,0.1950904,-0.9061272,-0.3753307,0.1950904,-0.8535532,-0.3535538,0.3826834,-0.8535532,-0.3535538,0.3826834,-0.9061272,-0.3753307,0.1950904,-0.9061273,-0.1802403,0.3826834,-0.9061272,-0.3753307,0.1950904,-0.9619396,-0.1913421,0.1950904,-0.9061273,-0.1802403,0.3826834,-0.9061273,-0.1802403,0.3826834,-0.9619396,-0.1913421,0.1950904,-0.9238795,-0.0000003,0.3826834,-0.9619396,-0.1913421,0.1950904,-0.9807853,-0.0000003,0.1950904,-0.9238795,-0.0000003,0.3826834,-0.9238795,-0.0000003,0.3826834,-0.9807853,-0.0000003,0.1950904,-0.9061275,0.1802397,0.3826834,-0.9807853,-0.0000003,0.1950904,-0.9619398,0.1913415,0.1950904,-0.9061275,0.1802397,0.3826834,-0.9061275,0.1802397,0.3826834,-0.9619398,0.1913415,0.1950904,-0.8535535,0.3535532,0.3826834,-0.9619398,0.1913415,0.1950904,-0.9061275,0.3753301,0.1950904,-0.8535535,0.3535532,0.3826834,-0.8535535,0.3535532,0.3826834,-0.9061275,0.3753301,0.1950904,-0.7681777,0.5132799,0.3826834,-0.9061275,0.3753301,0.1950904,-0.8154932,0.5448951,0.1950904,-0.7681777,0.5132799,0.3826834,-0.7681777,0.5132799,0.3826834,-0.8154932,0.5448951,0.1950904,-0.6532815,0.6532815,0.3826834,-0.8154932,0.5448951,0.1950904,-0.6935199,0.6935199,0.1950904,-0.6532815,0.6532815,0.3826834,-0.6532815,0.6532815,0.3826834,-0.6935199,0.6935199,0.1950904,-0.5132799,0.7681777,0.3826834,-0.6935199,0.6935199,0.1950904,-0.5448951,0.8154932,0.1950904,-0.5132799,0.7681777,0.3826834,-0.5132799,0.7681777,0.3826834,-0.5448951,0.8154932,0.1950904,-0.3535534,0.8535534,0.3826834,-0.5448951,0.8154932,0.1950904,-0.3753302,0.9061275,0.1950904,-0.3535534,0.8535534,0.3826834,-0.3535534,0.8535534,0.3826834,-0.3753302,0.9061275,0.1950904,-0.1802399,0.9061274,0.3826834,-0.3753302,0.9061275,0.1950904,-0.1913417,0.9619397,0.1950904,-0.1802399,0.9061274,0.3826834,-0.1802399,0.9061274,0.3826834,-0.1913417,0.9619397,0.1950904,-0.0,0.9238795,0.3826834,-0.1913417,0.9619397,0.1950904,-0.0,0.9807853,0.1950904,-0.0,0.9238795,0.3826834,-0.0,0.8314697,0.5555702,-0.0,0.9238795,0.3826834,0.1622099,0.8154936,0.5555702,-0.0,0.9238795,0.3826834,0.180238,0.9061278,0.3826834,0.1622099,0.8154936,0.5555702,0.1622099,0.8154936,0.5555702,0.180238,0.9061278,0.3826834,0.3181881,0.7681785,0.5555702,0.180238,0.9061278,0.3826834,0.3535516,0.8535541,0.3826834,0.3181881,0.7681785,0.5555702,0.3181881,0.7681785,0.5555702,0.3535516,0.8535541,0.3826834,0.4619384,0.6913427,0.5555702,0.3535516,0.8535541,0.3826834,0.5132784,0.7681788,0.3826834,0.4619384,0.6913427,0.5555702,0.4619384,0.6913427,0.5555702,0.5132784,0.7681788,0.3826834,0.5879367,0.587939,0.5555702,0.5132784,0.7681788,0.3826834,0.6532802,0.6532827,0.3826834,0.5879367,0.587939,0.5555702,0.5879367,0.587939,0.5555702,0.6532802,0.6532827,0.3826834,0.6913409,0.4619411,0.5555702,0.6532802,0.6532827,0.3826834,0.7681768,0.5132814,0.3826834,0.6913409,0.4619411,0.5555702,0.6913409,0.4619411,0.5555702,0.7681768,0.5132814,0.3826834,0.7681772,0.318191,0.5555702,0.7681768,0.5132814,0.3826834,0.8535528,0.3535549,0.3826834,0.7681772,0.318191,0.5555702,0.7681772,0.318191,0.5555702,0.8535528,0.3535549,0.3826834,0.8154929,0.1622131,0.5555702,0.8535528,0.3535549,0.3826834,0.9061271,0.1802415,0.3826834,0.8154929,0.1622131,0.5555702,0.8154929,0.1622131,0.5555702,0.9061271,0.1802415,0.3826834,0.8314697,0.0000013,0.5555702,0.9061271,0.1802415,0.3826834,0.9238795,0.0000015,0.3826834,0.8314697,0.0000013,0.5555702,0.8314697,0.0000013,0.5555702,0.9238795,0.0000015,0.3826834,0.8154934,-0.1622104,0.5555702,0.9238795,0.0000015,0.3826834,0.9061277,-0.1802386,0.3826834,0.8154934,-0.1622104,0.5555702,0.8154934,-0.1622104,0.5555702,0.9061277,-0.1802386,0.3826834,0.7681783,-0.3181885,0.5555702,0.9061277,-0.1802386,0.3826834,0.8535539,-0.3535521,0.3826834,0.7681783,-0.3181885,0.5555702,0.7681783,-0.3181885,0.5555702,0.8535539,-0.3535521,0.3826834,0.6913424,-0.4619389,0.5555702,0.8535539,-0.3535521,0.3826834,0.7681785,-0.5132789,0.3826834,0.6913424,-0.4619389,0.5555702,0.6913424,-0.4619389,0.5555702,0.7681785,-0.5132789,0.3826834,0.5879385,-0.5879371,0.5555702,0.7681785,-0.5132789,0.3826834,0.6532823,-0.6532806,0.3826834,0.5879385,-0.5879371,0.5555702,0.5879385,-0.5879371,0.5555702,0.6532823,-0.6532806,0.3826834,0.4619406,-0.6913412,0.5555702,0.6532823,-0.6532806,0.3826834,0.5132809,-0.7681771,0.3826834,0.4619406,-0.6913412,0.5555702,0.4619406,-0.6913412,0.5555702,0.5132809,-0.7681771,0.3826834,0.3181905,-0.7681774,0.5555702,0.5132809,-0.7681771,0.3826834,0.3535543,-0.8535529,0.3826834,0.3181905,-0.7681774,0.5555702,0.3181905,-0.7681774,0.5555702,0.3535543,-0.8535529,0.3826834,0.1622125,-0.815493,0.5555702,0.3535543,-0.8535529,0.3826834,0.1802409,-0.9061272,0.3826834,0.1622125,-0.815493,0.5555702,0.1622125,-0.815493,0.5555702,0.1802409,-0.9061272,0.3826834,0.0000008,-0.8314697,0.5555702,0.1802409,-0.9061272,0.3826834,0.0000009,-0.9238795,0.3826834,0.0000008,-0.8314697,0.5555702,0.0000008,-0.8314697,0.5555702,0.0000009,-0.9238795,0.3826834,-0.162211,-0.8154933,0.5555702,0.0000009,-0.9238795,0.3826834,-0.1802392,-0.9061276,0.3826834,-0.162211,-0.8154933,0.5555702,-0.162211,-0.8154933,0.5555702,-0.1802392,-0.9061276,0.3826834,-0.318189,-0.768178,0.5555702,-0.1802392,-0.9061276,0.3826834,-0.3535527,-0.8535537,0.3826834,-0.318189,-0.768178,0.5555702,-0.318189,-0.768178,0.5555702,-0.3535527,-0.8535537,0.3826834,-0.4619393,-0.6913421,0.5555702,-0.3535527,-0.8535537,0.3826834,-0.5132794,-0.7681781,0.3826834,-0.4619393,-0.6913421,0.5555702,-0.4619393,-0.6913421,0.5555702,-0.5132794,-0.7681781,0.3826834,-0.5879375,-0.5879382,0.5555702,-0.5132794,-0.7681781,0.3826834,-0.6532811,-0.6532819,0.3826834,-0.5879375,-0.5879382,0.5555702,-0.5879375,-0.5879382,0.5555702,-0.6532811,-0.6532819,0.3826834,-0.6913415,-0.4619402,0.5555702,-0.6532811,-0.6532819,0.3826834,-0.7681774,-0.5132804,0.3826834,-0.6913415,-0.4619402,0.5555702,-0.6913415,-0.4619402,0.5555702,-0.7681774,-0.5132804,0.3826834,-0.7681776,-0.31819,0.5555702,-0.7681774,-0.5132804,0.3826834,-0.8535532,-0.3535538,0.3826834,-0.7681776,-0.31819,0.5555702,-0.7681776,-0.31819,0.5555702,-0.8535532,-0.3535538,0.3826834,-0.8154931,-0.162212,0.5555702,-0.8535532,-0.3535538,0.3826834,-0.9061273,-0.1802403,0.3826834,-0.8154931,-0.162212,0.5555702,-0.8154931,-0.162212,0.5555702,-0.9061273,-0.1802403,0.3826834,-0.8314697,-0.0000003,0.5555702,-0.9061273,-0.1802403,0.3826834,-0.9238795,-0.0000003,0.3826834,-0.8314697,-0.0000003,0.5555702,-0.8314697,-0.0000003,0.5555702,-0.9238795,-0.0000003,0.3826834,-0.8154932,0.1622115,0.5555702,-0.9238795,-0.0000003,0.3826834,-0.9061275,0.1802397,0.3826834,-0.8154932,0.1622115,0.5555702,-0.8154932,0.1622115,0.5555702,-0.9061275,0.1802397,0.3826834,-0.7681779,0.3181895,0.5555702,-0.9061275,0.1802397,0.3826834,-0.8535535,0.3535532,0.3826834,-0.7681779,0.3181895,0.5555702,-0.7681779,0.3181895,0.5555702,-0.8535535,0.3535532,0.3826834,-0.6913418,0.4619398,0.5555702,-0.8535535,0.3535532,0.3826834,-0.7681777,0.5132799,0.3826834,-0.6913418,0.4619398,0.5555702,-0.6913418,0.4619398,0.5555702,-0.7681777,0.5132799,0.3826834,-0.5879378,0.5879378,0.5555702,-0.7681777,0.5132799,0.3826834,-0.6532815,0.6532815,0.3826834,-0.5879378,0.5879378,0.5555702,-0.5879378,0.5879378,0.5555702,-0.6532815,0.6532815,0.3826834,-0.4619398,0.6913418,0.5555702,-0.6532815,0.6532815,0.3826834,-0.5132799,0.7681777,0.3826834,-0.4619398,0.6913418,0.5555702,-0.4619398,0.6913418,0.5555702,-0.5132799,0.7681777,0.3826834,-0.3181896,0.7681778,0.5555702,-0.5132799,0.7681777,0.3826834,-0.3535534,0.8535534,0.3826834,-0.3181896,0.7681778,0.5555702,-0.3181896,0.7681778,0.5555702,-0.3535534,0.8535534,0.3826834,-0.1622117,0.8154932,0.5555702,-0.3535534,0.8535534,0.3826834,-0.1802399,0.9061274,0.3826834,-0.1622117,0.8154932,0.5555702,-0.1622117,0.8154932,0.5555702,-0.1802399,0.9061274,0.3826834,-0.0,0.8314697,0.5555702,-0.1802399,0.9061274,0.3826834,-0.0,0.9238795,0.3826834,-0.0,0.8314697,0.5555702,-0.0,0.7071068,0.7071068,-0.0,0.8314697,0.5555702,0.1379482,0.6935202,0.7071068,-0.0,0.8314697,0.5555702,0.1622099,0.8154936,0.5555702,0.1379482,0.6935202,0.7071068,0.1379482,0.6935202,0.7071068,0.1622099,0.8154936,0.5555702,0.2705967,0.653282,0.7071068,0.1622099,0.8154936,0.5555702,0.3181881,0.7681785,0.5555702,0.2705967,0.653282,0.7071068,0.2705967,0.653282,0.7071068,0.3181881,0.7681785,0.5555702,0.3928463,0.5879386,0.7071068,0.3181881,0.7681785,0.5555702,0.4619384,0.6913427,0.5555702,0.3928463,0.5879386,0.7071068,0.3928463,0.5879386,0.7071068,0.4619384,0.6913427,0.5555702,0.499999,0.500001,0.7071068,0.4619384,0.6913427,0.5555702,0.5879367,0.587939,0.5555702,0.499999,0.500001,0.7071068,0.499999,0.500001,0.7071068,0.5879367,0.587939,0.5555702,0.5879371,0.3928486,0.7071068,0.5879367,0.587939,0.5555702,0.6913409,0.4619411,0.5555702,0.5879371,0.3928486,0.7071068,0.5879371,0.3928486,0.7071068,0.6913409,0.4619411,0.5555702,0.653281,0.2705992,0.7071068,0.6913409,0.4619411,0.5555702,0.7681772,0.318191,0.5555702,0.653281,0.2705992,0.7071068,0.653281,0.2705992,0.7071068,0.7681772,0.318191,0.5555702,0.6935197,0.1379509,0.7071068,0.7681772,0.318191,0.5555702,0.8154929,0.1622131,0.5555702,0.6935197,0.1379509,0.7071068,0.6935197,0.1379509,0.7071068,0.8154929,0.1622131,0.5555702,0.7071068,0.0000011,0.7071068,0.8154929,0.1622131,0.5555702,0.8314697,0.0000013,0.5555702,0.7071068,0.0000011,0.7071068,0.7071068,0.0000011,0.7071068,0.8314697,0.0000013,0.5555702,0.6935201,-0.1379486,0.7071068,0.8314697,0.0000013,0.5555702,0.8154934,-0.1622104,0.5555702,0.6935201,-0.1379486,0.7071068,0.6935201,-0.1379486,0.7071068,0.8154934,-0.1622104,0.5555702,0.6532819,-0.2705971,0.7071068,0.8154934,-0.1622104,0.5555702,0.7681783,-0.3181885,0.5555702,0.6532819,-0.2705971,0.7071068,0.6532819,-0.2705971,0.7071068,0.7681783,-0.3181885,0.5555702,0.5879383,-0.3928467,0.7071068,0.7681783,-0.3181885,0.5555702,0.6913424,-0.4619389,0.5555702,0.5879383,-0.3928467,0.7071068,0.5879383,-0.3928467,0.7071068,0.6913424,-0.4619389,0.5555702,0.5000006,-0.4999993,0.7071068,0.6913424,-0.4619389,0.5555702,0.5879385,-0.5879371,0.5555702,0.5000006,-0.4999993,0.7071068,0.5000006,-0.4999993,0.7071068,0.5879385,-0.5879371,0.5555702,0.3928482,-0.5879373,0.7071068,0.5879385,-0.5879371,0.5555702,0.4619406,-0.6913412,0.5555702,0.3928482,-0.5879373,0.7071068,0.3928482,-0.5879373,0.7071068,0.4619406,-0.6913412,0.5555702,0.2705988,-0.6532812,0.7071068,0.4619406,-0.6913412,0.5555702,0.3181905,-0.7681774,0.5555702,0.2705988,-0.6532812,0.7071068,0.2705988,-0.6532812,0.7071068,0.3181905,-0.7681774,0.5555702,0.1379504,-0.6935198,0.7071068,0.3181905,-0.7681774,0.5555702,0.1622125,-0.815493,0.5555702,0.1379504,-0.6935198,0.7071068,0.1379504,-0.6935198,0.7071068,0.1622125,-0.815493,0.5555702,0.0000007,-0.7071068,0.7071068,0.1622125,-0.815493,0.5555702,0.0000008,-0.8314697,0.5555702,0.0000007,-0.7071068,0.7071068,0.0000007,-0.7071068,0.7071068,0.0000008,-0.8314697,0.5555702,-0.1379491,-0.69352,0.7071068,0.0000008,-0.8314697,0.5555702,-0.162211,-0.8154933,0.5555702,-0.1379491,-0.69352,0.7071068,-0.1379491,-0.69352,0.7071068,-0.162211,-0.8154933,0.5555702,-0.2705975,-0.6532817,0.7071068,-0.162211,-0.8154933,0.5555702,-0.318189,-0.768178,0.5555702,-0.2705975,-0.6532817,0.7071068,-0.2705975,-0.6532817,0.7071068,-0.318189,-0.768178,0.5555702,-0.3928471,-0.5879381,0.7071068,-0.318189,-0.768178,0.5555702,-0.4619393,-0.6913421,0.5555702,-0.3928471,-0.5879381,0.7071068,-0.3928471,-0.5879381,0.7071068,-0.4619393,-0.6913421,0.5555702,-0.4999997,-0.5000003,0.7071068,-0.4619393,-0.6913421,0.5555702,-0.5879375,-0.5879382,0.5555702,-0.4999997,-0.5000003,0.7071068,-0.4999997,-0.5000003,0.7071068,-0.5879375,-0.5879382,0.5555702,-0.5879376,-0.3928478,0.7071068,-0.5879375,-0.5879382,0.5555702,-0.6913415,-0.4619402,0.5555702,-0.5879376,-0.3928478,0.7071068,-0.5879376,-0.3928478,0.7071068,-0.6913415,-0.4619402,0.5555702,-0.6532813,-0.2705984,0.7071068,-0.6913415,-0.4619402,0.5555702,-0.7681776,-0.31819,0.5555702,-0.6532813,-0.2705984,0.7071068,-0.6532813,-0.2705984,0.7071068,-0.7681776,-0.31819,0.5555702,-0.6935198,-0.13795,0.7071068,-0.7681776,-0.31819,0.5555702,-0.8154931,-0.162212,0.5555702,-0.6935198,-0.13795,0.7071068,-0.6935198,-0.13795,0.7071068,-0.8154931,-0.162212,0.5555702,-0.7071068,-0.0000002,0.7071068,-0.8154931,-0.162212,0.5555702,-0.8314697,-0.0000003,0.5555702,-0.7071068,-0.0000002,0.7071068,-0.7071068,-0.0000002,0.7071068,-0.8314697,-0.0000003,0.5555702,-0.6935199,0.1379495,0.7071068,-0.8314697,-0.0000003,0.5555702,-0.8154932,0.1622115,0.5555702,-0.6935199,0.1379495,0.7071068,-0.6935199,0.1379495,0.7071068,-0.8154932,0.1622115,0.5555702,-0.6532815,0.2705979,0.7071068,-0.8154932,0.1622115,0.5555702,-0.7681779,0.3181895,0.5555702,-0.6532815,0.2705979,0.7071068,-0.6532815,0.2705979,0.7071068,-0.7681779,0.3181895,0.5555702,-0.5879378,0.3928474,0.7071068,-0.7681779,0.3181895,0.5555702,-0.6913418,0.4619398,0.5555702,-0.5879378,0.3928474,0.7071068,-0.5879378,0.3928474,0.7071068,-0.6913418,0.4619398,0.5555702,-0.5,0.5,0.7071068,-0.6913418,0.4619398,0.5555702,-0.5879378,0.5879378,0.5555702,-0.5,0.5,0.7071068,-0.5,0.5,0.7071068,-0.5879378,0.5879378,0.5555702,-0.3928474,0.5879378,0.7071068,-0.5879378,0.5879378,0.5555702,-0.4619398,0.6913418,0.5555702,-0.3928474,0.5879378,0.7071068,-0.3928474,0.5879378,0.7071068,-0.4619398,0.6913418,0.5555702,-0.270598,0.6532815,0.7071068,-0.4619398,0.6913418,0.5555702,-0.3181896,0.7681778,0.5555702,-0.270598,0.6532815,0.7071068,-0.270598,0.6532815,0.7071068,-0.3181896,0.7681778,0.5555702,-0.1379497,0.6935199,0.7071068,-0.3181896,0.7681778,0.5555702,-0.1622117,0.8154932,0.5555702,-0.1379497,0.6935199,0.7071068,-0.1379497,0.6935199,0.7071068,-0.1622117,0.8154932,0.5555702,-0.0,0.7071068,0.7071068,-0.1622117,0.8154932,0.5555702,-0.0,0.8314697,0.5555702,-0.0,0.7071068,0.7071068,-0.0,0.5555702,0.8314696,-0.0,0.7071068,0.7071068,0.1083852,0.5448954,0.8314696,-0.0,0.7071068,0.7071068,0.1379482,0.6935202,0.7071068,0.1083852,0.5448954,0.8314696,0.1083852,0.5448954,0.8314696,0.1379482,0.6935202,0.7071068,0.2126065,0.5132805,0.8314696,0.1379482,0.6935202,0.7071068,0.2705967,0.653282,0.7071068,0.2126065,0.5132805,0.8314696,0.2126065,0.5132805,0.8314696,0.2705967,0.653282,0.7071068,0.3086574,0.4619404,0.8314696,0.2705967,0.653282,0.7071068,0.3928463,0.5879386,0.7071068,0.3086574,0.4619404,0.8314696,0.3086574,0.4619404,0.8314696,0.3928463,0.5879386,0.7071068,0.3928467,0.3928483,0.8314696,0.3928463,0.5879386,0.7071068,0.499999,0.500001,0.7071068,0.3928467,0.3928483,0.8314696,0.3928467,0.3928483,0.8314696,0.499999,0.500001,0.7071068,0.4619392,0.3086592,0.8314696,0.499999,0.500001,0.7071068,0.5879371,0.3928486,0.7071068,0.4619392,0.3086592,0.8314696,0.4619392,0.3086592,0.8314696,0.5879371,0.3928486,0.7071068,0.5132796,0.2126084,0.8314696,0.5879371,0.3928486,0.7071068,0.653281,0.2705992,0.7071068,0.5132796,0.2126084,0.8314696,0.5132796,0.2126084,0.8314696,0.653281,0.2705992,0.7071068,0.5448949,0.1083873,0.8314696,0.653281,0.2705992,0.7071068,0.6935197,0.1379509,0.7071068,0.5448949,0.1083873,0.8314696,0.5448949,0.1083873,0.8314696,0.6935197,0.1379509,0.7071068,0.5555702,0.0000009,0.8314696,0.6935197,0.1379509,0.7071068,0.7071068,0.0000011,0.7071068,0.5555702,0.0000009,0.8314696,0.5555702,0.0000009,0.8314696,0.7071068,0.0000011,0.7071068,0.5448953,-0.1083855,0.8314696,0.7071068,0.0000011,0.7071068,0.6935201,-0.1379486,0.7071068,0.5448953,-0.1083855,0.8314696,0.5448953,-0.1083855,0.8314696,0.6935201,-0.1379486,0.7071068,0.5132803,-0.2126068,0.8314696,0.6935201,-0.1379486,0.7071068,0.6532819,-0.2705971,0.7071068,0.5132803,-0.2126068,0.8314696,0.5132803,-0.2126068,0.8314696,0.6532819,-0.2705971,0.7071068,0.4619402,-0.3086577,0.8314696,0.6532819,-0.2705971,0.7071068,0.5879383,-0.3928467,0.7071068,0.4619402,-0.3086577,0.8314696,0.4619402,-0.3086577,0.8314696,0.5879383,-0.3928467,0.7071068,0.392848,-0.392847,0.8314696,0.5879383,-0.3928467,0.7071068,0.5000006,-0.4999993,0.7071068,0.392848,-0.392847,0.8314696,0.392848,-0.392847,0.8314696,0.5000006,-0.4999993,0.7071068,0.3086589,-0.4619394,0.8314696,0.5000006,-0.4999993,0.7071068,0.3928482,-0.5879373,0.7071068,0.3086589,-0.4619394,0.8314696,0.3086589,-0.4619394,0.8314696,0.3928482,-0.5879373,0.7071068,0.2126081,-0.5132797,0.8314696,0.3928482,-0.5879373,0.7071068,0.2705988,-0.6532812,0.7071068,0.2126081,-0.5132797,0.8314696,0.2126081,-0.5132797,0.8314696,0.2705988,-0.6532812,0.7071068,0.1083869,-0.544895,0.8314696,0.2705988,-0.6532812,0.7071068,0.1379504,-0.6935198,0.7071068,0.1083869,-0.544895,0.8314696,0.1083869,-0.544895,0.8314696,0.1379504,-0.6935198,0.7071068,0.0000005,-0.5555702,0.8314696,0.1379504,-0.6935198,0.7071068,0.0000007,-0.7071068,0.7071068,0.0000005,-0.5555702,0.8314696,0.0000005,-0.5555702,0.8314696,0.0000007,-0.7071068,0.7071068,-0.1083859,-0.5448952,0.8314696,0.0000007,-0.7071068,0.7071068,-0.1379491,-0.69352,0.7071068,-0.1083859,-0.5448952,0.8314696,-0.1083859,-0.5448952,0.8314696,-0.1379491,-0.69352,0.7071068,-0.2126071,-0.5132802,0.8314696,-0.1379491,-0.69352,0.7071068,-0.2705975,-0.6532817,0.7071068,-0.2126071,-0.5132802,0.8314696,-0.2126071,-0.5132802,0.8314696,-0.2705975,-0.6532817,0.7071068,-0.308658,-0.46194,0.8314696,-0.2705975,-0.6532817,0.7071068,-0.3928471,-0.5879381,0.7071068,-0.308658,-0.46194,0.8314696,-0.308658,-0.46194,0.8314696,-0.3928471,-0.5879381,0.7071068,-0.3928472,-0.3928477,0.8314696,-0.3928471,-0.5879381,0.7071068,-0.4999997,-0.5000003,0.7071068,-0.3928472,-0.3928477,0.8314696,-0.3928472,-0.3928477,0.8314696,-0.4999997,-0.5000003,0.7071068,-0.4619396,-0.3086586,0.8314696,-0.4999997,-0.5000003,0.7071068,-0.5879376,-0.3928478,0.7071068,-0.4619396,-0.3086586,0.8314696,-0.4619396,-0.3086586,0.8314696,-0.5879376,-0.3928478,0.7071068,-0.5132799,-0.2126078,0.8314696,-0.5879376,-0.3928478,0.7071068,-0.6532813,-0.2705984,0.7071068,-0.5132799,-0.2126078,0.8314696,-0.5132799,-0.2126078,0.8314696,-0.6532813,-0.2705984,0.7071068,-0.5448951,-0.1083866,0.8314696,-0.6532813,-0.2705984,0.7071068,-0.6935198,-0.13795,0.7071068,-0.5448951,-0.1083866,0.8314696,-0.5448951,-0.1083866,0.8314696,-0.6935198,-0.13795,0.7071068,-0.5555702,-0.0000002,0.8314696,-0.6935198,-0.13795,0.7071068,-0.7071068,-0.0000002,0.7071068,-0.5555702,-0.0000002,0.8314696,-0.5555702,-0.0000002,0.8314696,-0.7071068,-0.0000002,0.7071068,-0.5448951,0.1083862,0.8314696,-0.7071068,-0.0000002,0.7071068,-0.6935199,0.1379495,0.7071068,-0.5448951,0.1083862,0.8314696,-0.5448951,0.1083862,0.8314696,-0.6935199,0.1379495,0.7071068,-0.51328,0.2126074,0.8314696,-0.6935199,0.1379495,0.7071068,-0.6532815,0.2705979,0.7071068,-0.51328,0.2126074,0.8314696,-0.51328,0.2126074,0.8314696,-0.6532815,0.2705979,0.7071068,-0.4619398,0.3086583,0.8314696,-0.6532815,0.2705979,0.7071068,-0.5879378,0.3928474,0.7071068,-0.4619398,0.3086583,0.8314696,-0.4619398,0.3086583,0.8314696,-0.5879378,0.3928474,0.7071068,-0.3928475,0.3928475,0.8314696,-0.5879378,0.3928474,0.7071068,-0.5,0.5,0.7071068,-0.3928475,0.3928475,0.8314696,-0.3928475,0.3928475,0.8314696,-0.5,0.5,0.7071068,-0.3086583,0.4619398,0.8314696,-0.5,0.5,0.7071068,-0.3928474,0.5879378,0.7071068,-0.3086583,0.4619398,0.8314696,-0.3086583,0.4619398,0.8314696,-0.3928474,0.5879378,0.7071068,-0.2126075,0.51328,0.8314696,-0.3928474,0.5879378,0.7071068,-0.270598,0.6532815,0.7071068,-0.2126075,0.51328,0.8314696,-0.2126075,0.51328,0.8314696,-0.270598,0.6532815,0.7071068,-0.1083864,0.5448951,0.8314696,-0.270598,0.6532815,0.7071068,-0.1379497,0.6935199,0.7071068,-0.1083864,0.5448951,0.8314696,-0.1083864,0.5448951,0.8314696,-0.1379497,0.6935199,0.7071068,-0.0,0.5555702,0.8314696,-0.1379497,0.6935199,0.7071068,-0.0,0.7071068,0.7071068,-0.0,0.5555702,0.8314696,-0.0,0.3826835,0.9238795,-0.0,0.5555702,0.8314696,0.074657,0.3753305,0.9238795,-0.0,0.5555702,0.8314696,0.1083852,0.5448954,0.8314696,0.074657,0.3753305,0.9238795,0.074657,0.3753305,0.9238795,0.1083852,0.5448954,0.8314696,0.1464459,0.3535537,0.9238795,0.1083852,0.5448954,0.8314696,0.2126065,0.5132805,0.8314696,0.1464459,0.3535537,0.9238795,0.1464459,0.3535537,0.9238795,0.2126065,0.5132805,0.8314696,0.2126069,0.3181901,0.9238795,0.2126065,0.5132805,0.8314696,0.3086574,0.4619404,0.8314696,0.2126069,0.3181901,0.9238795,0.2126069,0.3181901,0.9238795,0.3086574,0.4619404,0.8314696,0.2705975,0.2705986,0.9238795,0.3086574,0.4619404,0.8314696,0.3928467,0.3928483,0.8314696,0.2705975,0.2705986,0.9238795,0.2705975,0.2705986,0.9238795,0.3928467,0.3928483,0.8314696,0.3181893,0.2126081,0.9238795,0.3928467,0.3928483,0.8314696,0.4619392,0.3086592,0.8314696,0.3181893,0.2126081,0.9238795,0.3181893,0.2126081,0.9238795,0.4619392,0.3086592,0.8314696,0.3535531,0.1464472,0.9238795,0.4619392,0.3086592,0.8314696,0.5132796,0.2126084,0.8314696,0.3535531,0.1464472,0.9238795,0.3535531,0.1464472,0.9238795,0.5132796,0.2126084,0.8314696,0.3753302,0.0746585,0.9238795,0.5132796,0.2126084,0.8314696,0.5448949,0.1083873,0.8314696,0.3753302,0.0746585,0.9238795,0.3753302,0.0746585,0.9238795,0.5448949,0.1083873,0.8314696,0.3826835,0.0000006,0.9238795,0.5448949,0.1083873,0.8314696,0.5555702,0.0000009,0.8314696,0.3826835,0.0000006,0.9238795,0.3826835,0.0000006,0.9238795,0.5555702,0.0000009,0.8314696,0.3753304,-0.0746573,0.9238795,0.5555702,0.0000009,0.8314696,0.5448953,-0.1083855,0.8314696,0.3753304,-0.0746573,0.9238795,0.3753304,-0.0746573,0.9238795,0.5448953,-0.1083855,0.8314696,0.3535536,-0.1464461,0.9238795,0.5448953,-0.1083855,0.8314696,0.5132803,-0.2126068,0.8314696,0.3535536,-0.1464461,0.9238795,0.3535536,-0.1464461,0.9238795,0.5132803,-0.2126068,0.8314696,0.3181899,-0.2126071,0.9238795,0.5132803,-0.2126068,0.8314696,0.4619402,-0.3086577,0.8314696,0.3181899,-0.2126071,0.9238795,0.3181899,-0.2126071,0.9238795,0.4619402,-0.3086577,0.8314696,0.2705984,-0.2705977,0.9238795,0.4619402,-0.3086577,0.8314696,0.392848,-0.392847,0.8314696,0.2705984,-0.2705977,0.9238795,0.2705984,-0.2705977,0.9238795,0.392848,-0.392847,0.8314696,0.2126079,-0.3181894,0.9238795,0.392848,-0.392847,0.8314696,0.3086589,-0.4619394,0.8314696,0.2126079,-0.3181894,0.9238795,0.2126079,-0.3181894,0.9238795,0.3086589,-0.4619394,0.8314696,0.146447,-0.3535532,0.9238795,0.3086589,-0.4619394,0.8314696,0.2126081,-0.5132797,0.8314696,0.146447,-0.3535532,0.9238795,0.146447,-0.3535532,0.9238795,0.2126081,-0.5132797,0.8314696,0.0746582,-0.3753302,0.9238795,0.2126081,-0.5132797,0.8314696,0.1083869,-0.544895,0.8314696,0.0746582,-0.3753302,0.9238795,0.0746582,-0.3753302,0.9238795,0.1083869,-0.544895,0.8314696,0.0000004,-0.3826835,0.9238795,0.1083869,-0.544895,0.8314696,0.0000005,-0.5555702,0.8314696,0.0000004,-0.3826835,0.9238795,0.0000004,-0.3826835,0.9238795,0.0000005,-0.5555702,0.8314696,-0.0746575,-0.3753304,0.9238795,0.0000005,-0.5555702,0.8314696,-0.1083859,-0.5448952,0.8314696,-0.0746575,-0.3753304,0.9238795,-0.0746575,-0.3753304,0.9238795,-0.1083859,-0.5448952,0.8314696,-0.1464463,-0.3535535,0.9238795,-0.1083859,-0.5448952,0.8314696,-0.2126071,-0.5132802,0.8314696,-0.1464463,-0.3535535,0.9238795,-0.1464463,-0.3535535,0.9238795,-0.2126071,-0.5132802,0.8314696,-0.2126073,-0.3181898,0.9238795,-0.2126071,-0.5132802,0.8314696,-0.308658,-0.46194,0.8314696,-0.2126073,-0.3181898,0.9238795,-0.2126073,-0.3181898,0.9238795,-0.308658,-0.46194,0.8314696,-0.2705979,-0.2705982,0.9238795,-0.308658,-0.46194,0.8314696,-0.3928472,-0.3928477,0.8314696,-0.2705979,-0.2705982,0.9238795,-0.2705979,-0.2705982,0.9238795,-0.3928472,-0.3928477,0.8314696,-0.3181895,-0.2126077,0.9238795,-0.3928472,-0.3928477,0.8314696,-0.4619396,-0.3086586,0.8314696,-0.3181895,-0.2126077,0.9238795,-0.3181895,-0.2126077,0.9238795,-0.4619396,-0.3086586,0.8314696,-0.3535533,-0.1464468,0.9238795,-0.4619396,-0.3086586,0.8314696,-0.5132799,-0.2126078,0.8314696,-0.3535533,-0.1464468,0.9238795,-0.3535533,-0.1464468,0.9238795,-0.5132799,-0.2126078,0.8314696,-0.3753303,-0.074658,0.9238795,-0.5132799,-0.2126078,0.8314696,-0.5448951,-0.1083866,0.8314696,-0.3753303,-0.074658,0.9238795,-0.3753303,-0.074658,0.9238795,-0.5448951,-0.1083866,0.8314696,-0.3826835,-0.0000001,0.9238795,-0.5448951,-0.1083866,0.8314696,-0.5555702,-0.0000002,0.8314696,-0.3826835,-0.0000001,0.9238795,-0.3826835,-0.0000001,0.9238795,-0.5555702,-0.0000002,0.8314696,-0.3753303,0.0746578,0.9238795,-0.5555702,-0.0000002,0.8314696,-0.5448951,0.1083862,0.8314696,-0.3753303,0.0746578,0.9238795,-0.3753303,0.0746578,0.9238795,-0.5448951,0.1083862,0.8314696,-0.3535534,0.1464466,0.9238795,-0.5448951,0.1083862,0.8314696,-0.51328,0.2126074,0.8314696,-0.3535534,0.1464466,0.9238795,-0.3535534,0.1464466,0.9238795,-0.51328,0.2126074,0.8314696,-0.3181897,0.2126075,0.9238795,-0.51328,0.2126074,0.8314696,-0.4619398,0.3086583,0.8314696,-0.3181897,0.2126075,0.9238795,-0.3181897,0.2126075,0.9238795,-0.4619398,0.3086583,0.8314696,-0.2705981,0.2705981,0.9238795,-0.4619398,0.3086583,0.8314696,-0.3928475,0.3928475,0.8314696,-0.2705981,0.2705981,0.9238795,-0.2705981,0.2705981,0.9238795,-0.3928475,0.3928475,0.8314696,-0.2126075,0.3181897,0.9238795,-0.3928475,0.3928475,0.8314696,-0.3086583,0.4619398,0.8314696,-0.2126075,0.3181897,0.9238795,-0.2126075,0.3181897,0.9238795,-0.3086583,0.4619398,0.8314696,-0.1464466,0.3535534,0.9238795,-0.3086583,0.4619398,0.8314696,-0.2126075,0.51328,0.8314696,-0.1464466,0.3535534,0.9238795,-0.1464466,0.3535534,0.9238795,-0.2126075,0.51328,0.8314696,-0.0746578,0.3753303,0.9238795,-0.2126075,0.51328,0.8314696,-0.1083864,0.5448951,0.8314696,-0.0746578,0.3753303,0.9238795,-0.0746578,0.3753303,0.9238795,-0.1083864,0.5448951,0.8314696,-0.0,0.3826835,0.9238795,-0.1083864,0.5448951,0.8314696,-0.0,0.5555702,0.8314696,-0.0,0.3826835,0.9238795,-0.0,0.1950903,0.9807853,-0.0,0.3826835,0.9238795,0.0380598,0.1913418,0.9807853,-0.0,0.3826835,0.9238795,0.074657,0.3753305,0.9238795,0.0380598,0.1913418,0.9807853,0.0380598,0.1913418,0.9807853,0.074657,0.3753305,0.9238795,0.0746575,0.1802401,0.9807853,0.074657,0.3753305,0.9238795,0.1464459,0.3535537,0.9238795,0.0746575,0.1802401,0.9807853,0.0746575,0.1802401,0.9807853,0.1464459,0.3535537,0.9238795,0.1083861,0.1622119,0.9807853,0.1464459,0.3535537,0.9238795,0.2126069,0.3181901,0.9238795,0.1083861,0.1622119,0.9807853,0.1083861,0.1622119,0.9807853,0.2126069,0.3181901,0.9238795,0.1379494,0.13795,0.9807853,0.2126069,0.3181901,0.9238795,0.2705975,0.2705986,0.9238795,0.1379494,0.13795,0.9807853,0.1379494,0.13795,0.9807853,0.2705975,0.2705986,0.9238795,0.1622115,0.1083867,0.9807853,0.2705975,0.2705986,0.9238795,0.3181893,0.2126081,0.9238795,0.1622115,0.1083867,0.9807853,0.1622115,0.1083867,0.9807853,0.3181893,0.2126081,0.9238795,0.1802398,0.0746582,0.9807853,0.3181893,0.2126081,0.9238795,0.3535531,0.1464472,0.9238795,0.1802398,0.0746582,0.9807853,0.1802398,0.0746582,0.9807853,0.3535531,0.1464472,0.9238795,0.1913417,0.0380606,0.9807853,0.3535531,0.1464472,0.9238795,0.3753302,0.0746585,0.9238795,0.1913417,0.0380606,0.9807853,0.1913417,0.0380606,0.9807853,0.3753302,0.0746585,0.9238795,0.1950903,0.0000003,0.9807853,0.3753302,0.0746585,0.9238795,0.3826835,0.0000006,0.9238795,0.1950903,0.0000003,0.9807853,0.1950903,0.0000003,0.9807853,0.3826835,0.0000006,0.9238795,0.1913418,-0.0380599,0.9807853,0.3826835,0.0000006,0.9238795,0.3753304,-0.0746573,0.9238795,0.1913418,-0.0380599,0.9807853,0.1913418,-0.0380599,0.9807853,0.3753304,-0.0746573,0.9238795,0.1802401,-0.0746576,0.9807853,0.3753304,-0.0746573,0.9238795,0.3535536,-0.1464461,0.9238795,0.1802401,-0.0746576,0.9807853,0.1802401,-0.0746576,0.9807853,0.3535536,-0.1464461,0.9238795,0.1622118,-0.1083862,0.9807853,0.3535536,-0.1464461,0.9238795,0.3181899,-0.2126071,0.9238795,0.1622118,-0.1083862,0.9807853,0.1622118,-0.1083862,0.9807853,0.3181899,-0.2126071,0.9238795,0.1379499,-0.1379495,0.9807853,0.3181899,-0.2126071,0.9238795,0.2705984,-0.2705977,0.9238795,0.1379499,-0.1379495,0.9807853,0.1379499,-0.1379495,0.9807853,0.2705984,-0.2705977,0.9238795,0.1083866,-0.1622115,0.9807853,0.2705984,-0.2705977,0.9238795,0.2126079,-0.3181894,0.9238795,0.1083866,-0.1622115,0.9807853,0.1083866,-0.1622115,0.9807853,0.2126079,-0.3181894,0.9238795,0.074658,-0.1802399,0.9807853,0.2126079,-0.3181894,0.9238795,0.146447,-0.3535532,0.9238795,0.074658,-0.1802399,0.9807853,0.074658,-0.1802399,0.9807853,0.146447,-0.3535532,0.9238795,0.0380604,-0.1913417,0.9807853,0.146447,-0.3535532,0.9238795,0.0746582,-0.3753302,0.9238795,0.0380604,-0.1913417,0.9807853,0.0380604,-0.1913417,0.9807853,0.0746582,-0.3753302,0.9238795,0.0000002,-0.1950903,0.9807853,0.0746582,-0.3753302,0.9238795,0.0000004,-0.3826835,0.9238795,0.0000002,-0.1950903,0.9807853,0.0000002,-0.1950903,0.9807853,0.0000004,-0.3826835,0.9238795,-0.0380601,-0.1913417,0.9807853,0.0000004,-0.3826835,0.9238795,-0.0746575,-0.3753304,0.9238795,-0.0380601,-0.1913417,0.9807853,-0.0380601,-0.1913417,0.9807853,-0.0746575,-0.3753304,0.9238795,-0.0746577,-0.18024,0.9807853,-0.0746575,-0.3753304,0.9238795,-0.1464463,-0.3535535,0.9238795,-0.0746577,-0.18024,0.9807853,-0.0746577,-0.18024,0.9807853,-0.1464463,-0.3535535,0.9238795,-0.1083863,-0.1622118,0.9807853,-0.1464463,-0.3535535,0.9238795,-0.2126073,-0.3181898,0.9238795,-0.1083863,-0.1622118,0.9807853,-0.1083863,-0.1622118,0.9807853,-0.2126073,-0.3181898,0.9238795,-0.1379496,-0.1379498,0.9807853,-0.2126073,-0.3181898,0.9238795,-0.2705979,-0.2705982,0.9238795,-0.1379496,-0.1379498,0.9807853,-0.1379496,-0.1379498,0.9807853,-0.2705979,-0.2705982,0.9238795,-0.1622116,-0.1083865,0.9807853,-0.2705979,-0.2705982,0.9238795,-0.3181895,-0.2126077,0.9238795,-0.1622116,-0.1083865,0.9807853,-0.1622116,-0.1083865,0.9807853,-0.3181895,-0.2126077,0.9238795,-0.1802399,-0.0746579,0.9807853,-0.3181895,-0.2126077,0.9238795,-0.3535533,-0.1464468,0.9238795,-0.1802399,-0.0746579,0.9807853,-0.1802399,-0.0746579,0.9807853,-0.3535533,-0.1464468,0.9238795,-0.1913417,-0.0380603,0.9807853,-0.3535533,-0.1464468,0.9238795,-0.3753303,-0.074658,0.9238795,-0.1913417,-0.0380603,0.9807853,-0.1913417,-0.0380603,0.9807853,-0.3753303,-0.074658,0.9238795,-0.1950903,-0.0000001,0.9807853,-0.3753303,-0.074658,0.9238795,-0.3826835,-0.0000001,0.9238795,-0.1950903,-0.0000001,0.9807853,-0.1950903,-0.0000001,0.9807853,-0.3826835,-0.0000001,0.9238795,-0.1913417,0.0380602,0.9807853,-0.3826835,-0.0000001,0.9238795,-0.3753303,0.0746578,0.9238795,-0.1913417,0.0380602,0.9807853,-0.1913417,0.0380602,0.9807853,-0.3753303,0.0746578,0.9238795,-0.18024,0.0746578,0.9807853,-0.3753303,0.0746578,0.9238795,-0.3535534,0.1464466,0.9238795,-0.18024,0.0746578,0.9807853,-0.18024,0.0746578,0.9807853,-0.3535534,0.1464466,0.9238795,-0.1622117,0.1083864,0.9807853,-0.3535534,0.1464466,0.9238795,-0.3181897,0.2126075,0.9238795,-0.1622117,0.1083864,0.9807853,-0.1622117,0.1083864,0.9807853,-0.3181897,0.2126075,0.9238795,-0.1379497,0.1379497,0.9807853,-0.3181897,0.2126075,0.9238795,-0.2705981,0.2705981,0.9238795,-0.1379497,0.1379497,0.9807853,-0.1379497,0.1379497,0.9807853,-0.2705981,0.2705981,0.9238795,-0.1083864,0.1622117,0.9807853,-0.2705981,0.2705981,0.9238795,-0.2126075,0.3181897,0.9238795,-0.1083864,0.1622117,0.9807853,-0.1083864,0.1622117,0.9807853,-0.2126075,0.3181897,0.9238795,-0.0746578,0.18024,0.9807853,-0.2126075,0.3181897,0.9238795,-0.1464466,0.3535534,0.9238795,-0.0746578,0.18024,0.9807853,-0.0746578,0.18024,0.9807853,-0.1464466,0.3535534,0.9238795,-0.0380602,0.1913417,0.9807853,-0.1464466,0.3535534,0.9238795,-0.0746578,0.3753303,0.9238795,-0.0380602,0.1913417,0.9807853,-0.0380602,0.1913417,0.9807853,-0.0746578,0.3753303,0.9238795,-0.0,0.1950903,0.9807853,-0.0746578,0.3753303,0.9238795,-0.0,0.3826835,0.9238795,-0.0,0.1950903,0.9807853,-0.0,0.1950903,0.9807853,0.0380598,0.1913418,0.9807853,0.0,0.0,1.0,0.0380598,0.1913418,0.9807853,0.0746575,0.1802401,0.9807853,0.0,0.0,1.0,0.0746575,0.1802401,0.9807853,0.1083861,0.1622119,0.9807853,0.0,0.0,1.0,0.1083861,0.1622119,0.9807853,0.1379494,0.13795,0.9807853,0.0,0.0,1.0,0.1379494,0.13795,0.9807853,0.1622115,0.1083867,0.9807853,0.0,0.0,1.0,0.1622115,0.1083867,0.9807853,0.1802398,0.0746582,0.9807853,0.0,0.0,1.0,0.1802398,0.0746582,0.9807853,0.1913417,0.0380606,0.9807853,0.0,0.0,1.0,0.1913417,0.0380606,0.9807853,0.1950903,0.0000003,0.9807853,0.0,0.0,1.0,0.1950903,0.0000003,0.9807853,0.1913418,-0.0380599,0.9807853,0.0,0.0,1.0,0.1913418,-0.0380599,0.9807853,0.1802401,-0.0746576,0.9807853,0.0,0.0,1.0,0.1802401,-0.0746576,0.9807853,0.1622118,-0.1083862,0.9807853,0.0,0.0,1.0,0.1622118,-0.1083862,0.9807853,0.1379499,-0.1379495,0.9807853,0.0,0.0,1.0,0.1379499,-0.1379495,0.9807853,0.1083866,-0.1622115,0.9807853,0.0,0.0,1.0,0.1083866,-0.1622115,0.9807853,0.074658,-0.1802399,0.9807853,0.0,0.0,1.0,0.074658,-0.1802399,0.9807853,0.0380604,-0.1913417,0.9807853,0.0,0.0,1.0,0.0380604,-0.1913417,0.9807853,0.0000002,-0.1950903,0.9807853,0.0,0.0,1.0,0.0000002,-0.1950903,0.9807853,-0.0380601,-0.1913417,0.9807853,0.0,0.0,1.0,-0.0380601,-0.1913417,0.9807853,-0.0746577,-0.18024,0.9807853,0.0,0.0,1.0,-0.0746577,-0.18024,0.9807853,-0.1083863,-0.1622118,0.9807853,0.0,0.0,1.0,-0.1083863,-0.1622118,0.9807853,-0.1379496,-0.1379498,0.9807853,0.0,0.0,1.0,-0.1379496,-0.1379498,0.9807853,-0.1622116,-0.1083865,0.9807853,0.0,0.0,1.0,-0.1622116,-0.1083865,0.9807853,-0.1802399,-0.0746579,0.9807853,0.0,0.0,1.0,-0.1802399,-0.0746579,0.9807853,-0.1913417,-0.0380603,0.9807853,0.0,0.0,1.0,-0.1913417,-0.0380603,0.9807853,-0.1950903,-0.0000001,0.9807853,0.0,0.0,1.0,-0.1950903,-0.0000001,0.9807853,-0.1913417,0.0380602,0.9807853,0.0,0.0,1.0,-0.1913417,0.0380602,0.9807853,-0.18024,0.0746578,0.9807853,0.0,0.0,1.0,-0.18024,0.0746578,0.9807853,-0.1622117,0.1083864,0.9807853,0.0,0.0,1.0,-0.1622117,0.1083864,0.9807853,-0.1379497,0.1379497,0.9807853,0.0,0.0,1.0,-0.1379497,0.1379497,0.9807853,-0.1083864,0.1622117,0.9807853,0.0,0.0,1.0,-0.1083864,0.1622117,0.9807853,-0.0746578,0.18024,0.9807853,0.0,0.0,1.0,-0.0746578,0.18024,0.9807853,-0.0380602,0.1913417,0.9807853,0.0,0.0,1.0,-0.0380602,0.1913417,0.9807853,-0.0,0.1950903,0.9807853,0.0,0.0,1.0,0.9396924,-0.3420208,1.966467,1.12763,-0.4104286,1.966467,1.2,-0.0000078,1.966467,1.2,-0.0000078,1.966467,0.9999999,-0.0,1.966467,0.9396924,-0.3420208,1.966467,0.766044,-0.6427881,1.966467,0.9192528,-0.7713457,1.966467,1.12763,-0.4104286,1.966467,1.12763,-0.4104286,1.966467,0.9396924,-0.3420208,1.966467,0.766044,-0.6427881,1.966467,0.4999995,-0.8660257,1.966467,0.5999994,-1.0392309,1.966467,0.9192528,-0.7713457,1.966467,0.9192528,-0.7713457,1.966467,0.766044,-0.6427881,1.966467,0.4999995,-0.8660257,1.966467,0.1736477,-0.9848078,1.966467,0.2083772,-1.181769,1.966467,0.5999994,-1.0392309,1.966467,0.5999994,-1.0392309,1.966467,0.4999995,-0.8660257,1.966467,0.1736477,-0.9848078,1.966467,-0.1736486,-0.9848077,1.966467,-0.2083783,-1.181769,1.966467,0.2083772,-1.181769,1.966467,0.2083772,-1.181769,1.966467,0.1736477,-0.9848078,1.966467,-0.1736486,-0.9848077,1.966467,-0.5000003,-0.8660252,1.966467,-0.6000004,-1.03923,1.966467,-0.2083783,-1.181769,1.966467,-0.2083783,-1.181769,1.966467,-0.1736486,-0.9848077,1.966467,-0.5000003,-0.8660252,1.966467,-0.7660446,-0.6427874,1.966467,-0.9192535,-0.7713449,1.966467,-0.6000004,-1.03923,1.966467,-0.6000004,-1.03923,1.966467,-0.5000003,-0.8660252,1.966467,-0.7660446,-0.6427874,1.966467,-0.9396927,-0.3420199,1.966467,-1.1276309,-0.4104239,1.966467,-0.9192535,-0.7713449,1.966467,-0.9192535,-0.7713449,1.966467,-0.7660446,-0.6427874,1.966467,-0.9396927,-0.3420199,1.966467,-1.0,0.0000001,1.966467,-1.2,-0.0000037,1.966467,-1.1276309,-0.4104239,1.966467,-1.1276309,-0.4104239,1.966467,-0.9396927,-0.3420199,1.966467,-1.0,0.0000001,1.966467,-0.9396926,0.3420202,1.966467,-1.127632,0.4104206,1.966467,-1.2,-0.0000037,1.966467,-1.2,-0.0000037,1.966467,-1.0,0.0000001,1.966467,-0.9396926,0.3420202,1.966467,-0.7660444,0.6427876,1.966467,-0.9192533,0.7713451,1.966467,-1.127632,0.4104206,1.966467,-1.127632,0.4104206,1.966467,-0.9396926,0.3420202,1.966467,-0.7660444,0.6427876,1.966467,-0.5000001,0.8660254,1.966467,-0.6000001,1.03923,1.966467,-0.9192533,0.7713451,1.966467,-0.9192533,0.7713451,1.966467,-0.7660444,0.6427876,1.966467,-0.5000001,0.8660254,1.966467,-0.1736483,0.9848077,1.966467,-0.208378,1.181769,1.966467,-0.6000001,1.03923,1.966467,-0.6000001,1.03923,1.966467,-0.5000001,0.8660254,1.966467,-0.1736483,0.9848077,1.966467,0.1736481,0.9848078,1.966467,0.2083777,1.181769,1.966467,-0.208378,1.181769,1.966467,-0.208378,1.181769,1.966467,-0.1736483,0.9848077,1.966467,0.1736481,0.9848078,1.966467,0.4999999,0.8660254,1.966467,0.6,1.0392309,1.966467,0.2083777,1.181769,1.966467,0.2083777,1.181769,1.966467,0.1736481,0.9848078,1.966467,0.4999999,0.8660254,1.966467,0.7660444,0.6427876,1.966467,0.9192533,0.7713451,1.966467,0.6,1.0392309,1.966467,0.6,1.0392309,1.966467,0.4999999,0.8660254,1.966467,0.7660444,0.6427876,1.966467,0.9396926,0.3420202,1.966467,1.127632,0.4104205,1.966467,0.9192533,0.7713451,1.966467,0.9192533,0.7713451,1.966467,0.7660444,0.6427876,1.966467,0.9396926,0.3420202,1.966467,0.9999999,-0.0,1.966467,1.2,-0.0000078,1.966467,1.127632,0.4104205,1.966467,1.127632,0.4104205,1.966467,0.9396926,0.3420202,1.966467,0.9999999,-0.0,1.966467,0.7660444,0.6427876,2.223201,0.9192533,0.7713451,2.223201,1.12763,0.4104278,2.223201,1.12763,0.4104278,2.223201,0.9396926,0.3420202,2.223201,0.7660444,0.6427876,2.223201,0.4999999,0.8660254,2.223201,0.6,1.0392309,2.223201,0.9192533,0.7713451,2.223201,0.9192533,0.7713451,2.223201,0.7660444,0.6427876,2.223201,0.4999999,0.8660254,2.223201,0.1736481,0.9848078,2.223201,0.2083777,1.181769,2.223201,0.6,1.0392309,2.223201,0.6,1.0392309,2.223201,0.4999999,0.8660254,2.223201,0.1736481,0.9848078,2.223201,-0.1736483,0.9848077,2.223201,-0.208378,1.181769,2.223201,0.2083777,1.181769,2.223201,0.2083777,1.181769,2.223201,0.1736481,0.9848078,2.223201,-0.1736483,0.9848077,2.223201,-0.5000001,0.8660254,2.223201,-0.6000001,1.03923,2.223201,-0.208378,1.181769,2.223201,-0.208378,1.181769,2.223201,-0.1736483,0.9848077,2.223201,-0.5000001,0.8660254,2.223201,-0.7660444,0.6427876,2.223201,-0.9192533,0.7713451,2.223201,-0.6000001,1.03923,2.223201,-0.6000001,1.03923,2.223201,-0.5000001,0.8660254,2.223201,-0.7660444,0.6427876,2.223201,-0.9396926,0.3420202,2.223201,-1.12763,0.4104279,2.223201,-0.9192533,0.7713451,2.223201,-0.9192533,0.7713451,2.223201,-0.7660444,0.6427876,2.223201,-0.9396926,0.3420202,2.223201,-1.0,0.0000002,2.223201,-1.2,0.0000041,2.223201,-1.12763,0.4104279,2.223201,-1.12763,0.4104279,2.223201,-0.9396926,0.3420202,2.223201,-1.0,0.0000002,2.223201,-0.9396927,-0.3420199,2.223201,-1.1276309,-0.4104239,2.223201,-1.2,0.0000041,2.223201,-1.2,0.0000041,2.223201,-1.0,0.0000002,2.223201,-0.9396927,-0.3420199,2.223201,-0.7660446,-0.6427874,2.223201,-0.9192535,-0.7713449,2.223201,-1.1276309,-0.4104239,2.223201,-1.1276309,-0.4104239,2.223201,-0.9396927,-0.3420199,2.223201,-0.7660446,-0.6427874,2.223201,-0.5000003,-0.8660252,2.223201,-0.6000004,-1.03923,2.223201,-0.9192535,-0.7713449,2.223201,-0.9192535,-0.7713449,2.223201,-0.7660446,-0.6427874,2.223201,-0.5000003,-0.8660252,2.223201,-0.1736486,-0.9848077,2.223201,-0.2083783,-1.181769,2.223201,-0.6000004,-1.03923,2.223201,-0.6000004,-1.03923,2.223201,-0.5000003,-0.8660252,2.223201,-0.1736486,-0.9848077,2.223201,0.1736477,-0.9848078,2.223201,0.2083772,-1.181769,2.223201,-0.2083783,-1.181769,2.223201,-0.2083783,-1.181769,2.223201,-0.1736486,-0.9848077,2.223201,0.1736477,-0.9848078,2.223201,0.4999995,-0.8660257,2.223201,0.5999994,-1.0392309,2.223201,0.2083772,-1.181769,2.223201,0.2083772,-1.181769,2.223201,0.1736477,-0.9848078,2.223201,0.4999995,-0.8660257,2.223201,0.766044,-0.6427881,2.223201,0.9192528,-0.7713457,2.223201,0.5999994,-1.0392309,2.223201,0.5999994,-1.0392309,2.223201,0.4999995,-0.8660257,2.223201,0.766044,-0.6427881,2.223201,0.9396924,-0.3420208,2.223201,1.127632,-0.4104213,2.223201,0.9192528,-0.7713457,2.223201,0.9192528,-0.7713457,2.223201,0.766044,-0.6427881,2.223201,0.9396924,-0.3420208,2.223201,0.9999999,0.0,2.223201,1.2,0.0000077,2.223201,1.127632,-0.4104213,2.223201,1.127632,-0.4104213,2.223201,0.9396924,-0.3420208,2.223201,0.9999999,0.0,2.223201,0.9396926,0.3420202,2.223201,1.12763,0.4104278,2.223201,1.2,0.0000077,2.223201,1.2,0.0000077,2.223201,0.9999999,0.0,2.223201,0.9396926,0.3420202,2.223201,1.2,0.0000077,2.223201,1.2,-0.0000078,1.966467,1.12763,-0.4104286,1.966467,1.12763,-0.4104286,1.966467,1.127632,-0.4104213,2.223201,1.2,0.0000077,2.223201,1.127632,-0.4104213,2.223201,1.12763,-0.4104286,1.966467,0.9192528,-0.7713457,1.966467,0.9192528,-0.7713457,1.966467,0.9192528,-0.7713457,2.223201,1.127632,-0.4104213,2.223201,0.9192528,-0.7713457,2.223201,0.9192528,-0.7713457,1.966467,0.5999994,-1.0392309,1.966467,0.5999994,-1.0392309,1.966467,0.5999994,-1.0392309,2.223201,0.9192528,-0.7713457,2.223201,0.5999994,-1.0392309,2.223201,0.5999994,-1.0392309,1.966467,0.2083772,-1.181769,1.966467,0.2083772,-1.181769,1.966467,0.2083772,-1.181769,2.223201,0.5999994,-1.0392309,2.223201,0.2083772,-1.181769,2.223201,0.2083772,-1.181769,1.966467,-0.2083783,-1.181769,1.966467,-0.2083783,-1.181769,1.966467,-0.2083783,-1.181769,2.223201,0.2083772,-1.181769,2.223201,-0.2083783,-1.181769,2.223201,-0.2083783,-1.181769,1.966467,-0.6000004,-1.03923,1.966467,-0.6000004,-1.03923,1.966467,-0.6000004,-1.03923,2.223201,-0.2083783,-1.181769,2.223201,-0.6000004,-1.03923,2.223201,-0.6000004,-1.03923,1.966467,-0.9192535,-0.7713449,1.966467,-0.9192535,-0.7713449,1.966467,-0.9192535,-0.7713449,2.223201,-0.6000004,-1.03923,2.223201,-0.9192535,-0.7713449,2.223201,-0.9192535,-0.7713449,1.966467,-1.1276309,-0.4104239,1.966467,-1.1276309,-0.4104239,1.966467,-1.1276309,-0.4104239,2.223201,-0.9192535,-0.7713449,2.223201,-1.1276309,-0.4104239,2.223201,-1.1276309,-0.4104239,1.966467,-1.2,-0.0000037,1.966467,-1.2,-0.0000037,1.966467,-1.2,0.0000041,2.223201,-1.1276309,-0.4104239,2.223201,-1.2,0.0000041,2.223201,-1.2,-0.0000037,1.966467,-1.127632,0.4104206,1.966467,-1.127632,0.4104206,1.966467,-1.12763,0.4104279,2.223201,-1.2,0.0000041,2.223201,-1.12763,0.4104279,2.223201,-1.127632,0.4104206,1.966467,-0.9192533,0.7713451,1.966467,-0.9192533,0.7713451,1.966467,-0.9192533,0.7713451,2.223201,-1.12763,0.4104279,2.223201,-0.9192533,0.7713451,2.223201,-0.9192533,0.7713451,1.966467,-0.6000001,1.03923,1.966467,-0.6000001,1.03923,1.966467,-0.6000001,1.03923,2.223201,-0.9192533,0.7713451,2.223201,-0.6000001,1.03923,2.223201,-0.6000001,1.03923,1.966467,-0.208378,1.181769,1.966467,-0.208378,1.181769,1.966467,-0.208378,1.181769,2.223201,-0.6000001,1.03923,2.223201,-0.208378,1.181769,2.223201,-0.208378,1.181769,1.966467,0.2083777,1.181769,1.966467,0.2083777,1.181769,1.966467,0.2083777,1.181769,2.223201,-0.208378,1.181769,2.223201,0.2083777,1.181769,2.223201,0.2083777,1.181769,1.966467,0.6,1.0392309,1.966467,0.6,1.0392309,1.966467,0.6,1.0392309,2.223201,0.2083777,1.181769,2.223201,0.6,1.0392309,2.223201,0.6,1.0392309,1.966467,0.9192533,0.7713451,1.966467,0.9192533,0.7713451,1.966467,0.9192533,0.7713451,2.223201,0.6,1.0392309,2.223201,0.9192533,0.7713451,2.223201,0.9192533,0.7713451,1.966467,1.127632,0.4104205,1.966467,1.127632,0.4104205,1.966467,1.12763,0.4104278,2.223201,0.9192533,0.7713451,2.223201,1.12763,0.4104278,2.223201,1.127632,0.4104205,1.966467,1.2,-0.0000078,1.966467,1.2,-0.0000078,1.966467,1.2,0.0000077,2.223201,1.12763,0.4104278,2.223201,0.9396924,-0.3420208,2.223201,0.9396926,0.3420202,2.223201,0.9999999,0.0,2.223201,0.9396924,-0.3420208,2.223201,0.4999999,0.8660254,2.223201,0.9396926,0.3420202,2.223201,0.9396924,-0.3420208,2.223201,0.4999995,-0.8660257,2.223201,0.4999999,0.8660254,2.223201,0.9396924,-0.3420208,2.223201,0.766044,-0.6427881,2.223201,0.4999995,-0.8660257,2.223201,0.4999995,-0.8660257,2.223201,-0.7660444,0.6427876,2.223201,0.4999999,0.8660254,2.223201,0.4999995,-0.8660257,2.223201,-0.7660446,-0.6427874,2.223201,-0.7660444,0.6427876,2.223201,0.4999995,-0.8660257,2.223201,-0.1736486,-0.9848077,2.223201,-0.7660446,-0.6427874,2.223201,0.4999995,-0.8660257,2.223201,0.1736477,-0.9848078,2.223201,-0.1736486,-0.9848077,2.223201,-0.1736486,-0.9848077,2.223201,-0.5000003,-0.8660252,2.223201,-0.7660446,-0.6427874,2.223201,-0.7660446,-0.6427874,2.223201,-1.0,0.0000002,2.223201,-0.7660444,0.6427876,2.223201,-0.7660446,-0.6427874,2.223201,-0.9396927,-0.3420199,2.223201,-1.0,0.0000002,2.223201,-1.0,0.0000002,2.223201,-0.9396926,0.3420202,2.223201,-0.7660444,0.6427876,2.223201,-0.7660444,0.6427876,2.223201,-0.1736483,0.9848077,2.223201,0.4999999,0.8660254,2.223201,-0.7660444,0.6427876,2.223201,-0.5000001,0.8660254,2.223201,-0.1736483,0.9848077,2.223201,-0.1736483,0.9848077,2.223201,0.1736481,0.9848078,2.223201,0.4999999,0.8660254,2.223201,0.4999999,0.8660254,2.223201,0.7660444,0.6427876,2.223201,0.9396926,0.3420202,2.223201,1.0,0.0,0.0,0.766044,-0.6427881,0.0,0.9396924,-0.3420208,0.0,1.0,0.0,0.0,0.1736477,-0.9848078,0.0,0.766044,-0.6427881,0.0,1.0,0.0,0.0,0.7660444,0.6427876,0.0,0.1736477,-0.9848078,0.0,1.0,0.0,0.0,0.9396926,0.3420202,0.0,0.7660444,0.6427876,0.0,0.7660444,0.6427876,0.0,-0.9396927,-0.3420199,0.0,0.1736477,-0.9848078,0.0,0.7660444,0.6427876,0.0,-0.5000001,0.8660254,0.0,-0.9396927,-0.3420199,0.0,0.7660444,0.6427876,0.0,0.1736481,0.9848078,0.0,-0.5000001,0.8660254,0.0,0.7660444,0.6427876,0.0,0.5,0.8660254,0.0,0.1736481,0.9848078,0.0,0.1736481,0.9848078,0.0,-0.1736483,0.9848077,0.0,-0.5000001,0.8660254,0.0,-0.5000001,0.8660254,0.0,-0.9396926,0.3420202,0.0,-0.9396927,-0.3420199,0.0,-0.5000001,0.8660254,0.0,-0.7660444,0.6427876,0.0,-0.9396926,0.3420202,0.0,-0.9396926,0.3420202,0.0,-1.0,0.0000002,0.0,-0.9396927,-0.3420199,0.0,-0.9396927,-0.3420199,0.0,-0.5000003,-0.8660252,0.0,0.1736477,-0.9848078,0.0,-0.9396927,-0.3420199,0.0,-0.7660446,-0.6427874,0.0,-0.5000003,-0.8660252,0.0,-0.5000003,-0.8660252,0.0,-0.1736486,-0.9848077,0.0,0.1736477,-0.9848078,0.0,0.1736477,-0.9848078,0.0,0.4999995,-0.8660257,0.0,0.766044,-0.6427881,0.0,1.0,0.0,0.0,0.9396924,-0.3420208,0.0,0.9396924,-0.3420208,1.966467,0.9396924,-0.3420208,1.966467,0.9999999,-0.0,1.966467,1.0,0.0,0.0,0.9396924,-0.3420208,0.0,0.766044,-0.6427881,0.0,0.766044,-0.6427881,1.966467,0.766044,-0.6427881,1.966467,0.9396924,-0.3420208,1.966467,0.9396924,-0.3420208,0.0,0.766044,-0.6427881,0.0,0.4999995,-0.8660257,0.0,0.4999995,-0.8660257,1.966467,0.4999995,-0.8660257,1.966467,0.766044,-0.6427881,1.966467,0.766044,-0.6427881,0.0,0.4999995,-0.8660257,0.0,0.1736477,-0.9848078,0.0,0.1736477,-0.9848078,1.966467,0.1736477,-0.9848078,1.966467,0.4999995,-0.8660257,1.966467,0.4999995,-0.8660257,0.0,0.1736477,-0.9848078,0.0,-0.1736486,-0.9848077,0.0,-0.1736486,-0.9848077,1.966467,-0.1736486,-0.9848077,1.966467,0.1736477,-0.9848078,1.966467,0.1736477,-0.9848078,0.0,-0.1736486,-0.9848077,0.0,-0.5000003,-0.8660252,0.0,-0.5000003,-0.8660252,1.966467,-0.5000003,-0.8660252,1.966467,-0.1736486,-0.9848077,1.966467,-0.1736486,-0.9848077,0.0,-0.5000003,-0.8660252,0.0,-0.7660446,-0.6427874,0.0,-0.7660446,-0.6427874,1.966467,-0.7660446,-0.6427874,1.966467,-0.5000003,-0.8660252,1.966467,-0.5000003,-0.8660252,0.0,-0.7660446,-0.6427874,0.0,-0.9396927,-0.3420199,0.0,-0.9396927,-0.3420199,1.966467,-0.9396927,-0.3420199,1.966467,-0.7660446,-0.6427874,1.966467,-0.7660446,-0.6427874,0.0,-0.9396927,-0.3420199,0.0,-1.0,0.0000002,0.0,-1.0,0.0000001,1.966467,-1.0,0.0000001,1.966467,-0.9396927,-0.3420199,1.966467,-0.9396927,-0.3420199,0.0,-1.0,0.0000002,0.0,-0.9396926,0.3420202,0.0,-0.9396926,0.3420202,1.966467,-0.9396926,0.3420202,1.966467,-1.0,0.0000001,1.966467,-1.0,0.0000002,0.0,-0.9396926,0.3420202,0.0,-0.7660444,0.6427876,0.0,-0.7660444,0.6427876,1.966467,-0.7660444,0.6427876,1.966467,-0.9396926,0.3420202,1.966467,-0.9396926,0.3420202,0.0,-0.7660444,0.6427876,0.0,-0.5000001,0.8660254,0.0,-0.5000001,0.8660254,1.966467,-0.5000001,0.8660254,1.966467,-0.7660444,0.6427876,1.966467,-0.7660444,0.6427876,0.0,-0.5000001,0.8660254,0.0,-0.1736483,0.9848077,0.0,-0.1736483,0.9848077,1.966467,-0.1736483,0.9848077,1.966467,-0.5000001,0.8660254,1.966467,-0.5000001,0.8660254,0.0,-0.1736483,0.9848077,0.0,0.1736481,0.9848078,0.0,0.1736481,0.9848078,1.966467,0.1736481,0.9848078,1.966467,-0.1736483,0.9848077,1.966467,-0.1736483,0.9848077,0.0,0.1736481,0.9848078,0.0,0.5,0.8660254,0.0,0.4999999,0.8660254,1.966467,0.4999999,0.8660254,1.966467,0.1736481,0.9848078,1.966467,0.1736481,0.9848078,0.0,0.5,0.8660254,0.0,0.7660444,0.6427876,0.0,0.7660444,0.6427876,1.966467,0.7660444,0.6427876,1.966467,0.4999999,0.8660254,1.966467,0.5,0.8660254,0.0,0.7660444,0.6427876,0.0,0.9396926,0.3420202,0.0,0.9396926,0.3420202,1.966467,0.9396926,0.3420202,1.966467,0.7660444,0.6427876,1.966467,0.7660444,0.6427876,0.0,0.9396926,0.3420202,0.0,1.0,0.0,0.0,0.9999999,-0.0,1.966467,0.9999999,-0.0,1.966467,0.9396926,0.3420202,1.966467,0.9396926,0.3420202,0.0]},\\"normal\\":{\\"valueType\\":\\"Float32\\",\\"valuesPerElement\\":3,\\"values\\":[0.039,0.197,-0.98,-0.0,0.201,-0.98,0.0,-0.0,-1.0,0.077,0.186,-0.98,0.039,0.197,-0.98,0.0,-0.0,-1.0,0.112,0.167,-0.98,0.077,0.186,-0.98,0.0,-0.0,-1.0,0.142,0.142,-0.98,0.112,0.167,-0.98,0.0,-0.0,-1.0,0.167,0.112,-0.98,0.142,0.142,-0.98,0.0,-0.0,-1.0,0.186,0.077,-0.98,0.167,0.112,-0.98,0.0,-0.0,-1.0,0.197,0.039,-0.98,0.186,0.077,-0.98,0.0,-0.0,-1.0,0.201,0.0,-0.98,0.197,0.039,-0.98,0.0,-0.0,-1.0,0.197,-0.039,-0.98,0.201,0.0,-0.98,0.0,-0.0,-1.0,0.186,-0.077,-0.98,0.197,-0.039,-0.98,0.0,-0.0,-1.0,0.167,-0.112,-0.98,0.186,-0.077,-0.98,0.0,-0.0,-1.0,0.142,-0.142,-0.98,0.167,-0.112,-0.98,0.0,-0.0,-1.0,0.112,-0.167,-0.98,0.142,-0.142,-0.98,0.0,-0.0,-1.0,0.077,-0.186,-0.98,0.112,-0.167,-0.98,0.0,-0.0,-1.0,0.039,-0.197,-0.98,0.077,-0.186,-0.98,0.0,-0.0,-1.0,0.0,-0.201,-0.98,0.039,-0.197,-0.98,0.0,-0.0,-1.0,-0.039,-0.197,-0.98,0.0,-0.201,-0.98,0.0,-0.0,-1.0,-0.077,-0.186,-0.98,-0.039,-0.197,-0.98,0.0,-0.0,-1.0,-0.112,-0.167,-0.98,-0.077,-0.186,-0.98,0.0,-0.0,-1.0,-0.142,-0.142,-0.98,-0.112,-0.167,-0.98,0.0,-0.0,-1.0,-0.167,-0.112,-0.98,-0.142,-0.142,-0.98,0.0,-0.0,-1.0,-0.186,-0.077,-0.98,-0.167,-0.112,-0.98,0.0,-0.0,-1.0,-0.197,-0.039,-0.98,-0.186,-0.077,-0.98,0.0,-0.0,-1.0,-0.201,-0.0,-0.98,-0.197,-0.039,-0.98,0.0,-0.0,-1.0,-0.197,0.039,-0.98,-0.201,-0.0,-0.98,0.0,-0.0,-1.0,-0.186,0.077,-0.98,-0.197,0.039,-0.98,0.0,-0.0,-1.0,-0.167,0.112,-0.98,-0.186,0.077,-0.98,0.0,-0.0,-1.0,-0.142,0.142,-0.98,-0.167,0.112,-0.98,0.0,-0.0,-1.0,-0.112,0.167,-0.98,-0.142,0.142,-0.98,0.0,-0.0,-1.0,-0.077,0.186,-0.98,-0.112,0.167,-0.98,0.0,-0.0,-1.0,-0.039,0.197,-0.98,-0.077,0.186,-0.98,0.0,-0.0,-1.0,-0.0,0.201,-0.98,-0.039,0.197,-0.98,0.0,-0.0,-1.0,-0.0,0.388,-0.922,-0.0,0.201,-0.98,0.076,0.38,-0.922,-0.0,0.201,-0.98,0.039,0.197,-0.98,0.076,0.38,-0.922,0.076,0.38,-0.922,0.039,0.197,-0.98,0.148,0.358,-0.922,0.039,0.197,-0.98,0.077,0.186,-0.98,0.148,0.358,-0.922,0.148,0.358,-0.922,0.077,0.186,-0.98,0.215,0.323,-0.922,0.077,0.186,-0.98,0.112,0.167,-0.98,0.215,0.323,-0.922,0.215,0.323,-0.922,0.112,0.167,-0.98,0.274,0.274,-0.922,0.112,0.167,-0.98,0.142,0.142,-0.98,0.274,0.274,-0.922,0.274,0.274,-0.922,0.142,0.142,-0.98,0.323,0.215,-0.922,0.142,0.142,-0.98,0.167,0.112,-0.98,0.323,0.215,-0.922,0.323,0.215,-0.922,0.167,0.112,-0.98,0.358,0.148,-0.922,0.167,0.112,-0.98,0.186,0.077,-0.98,0.358,0.148,-0.922,0.358,0.148,-0.922,0.186,0.077,-0.98,0.38,0.076,-0.922,0.186,0.077,-0.98,0.197,0.039,-0.98,0.38,0.076,-0.922,0.38,0.076,-0.922,0.197,0.039,-0.98,0.388,0.0,-0.922,0.197,0.039,-0.98,0.201,0.0,-0.98,0.388,0.0,-0.922,0.388,0.0,-0.922,0.201,0.0,-0.98,0.38,-0.076,-0.922,0.201,0.0,-0.98,0.197,-0.039,-0.98,0.38,-0.076,-0.922,0.38,-0.076,-0.922,0.197,-0.039,-0.98,0.358,-0.148,-0.922,0.197,-0.039,-0.98,0.186,-0.077,-0.98,0.358,-0.148,-0.922,0.358,-0.148,-0.922,0.186,-0.077,-0.98,0.323,-0.215,-0.922,0.186,-0.077,-0.98,0.167,-0.112,-0.98,0.323,-0.215,-0.922,0.323,-0.215,-0.922,0.167,-0.112,-0.98,0.274,-0.274,-0.922,0.167,-0.112,-0.98,0.142,-0.142,-0.98,0.274,-0.274,-0.922,0.274,-0.274,-0.922,0.142,-0.142,-0.98,0.215,-0.323,-0.922,0.142,-0.142,-0.98,0.112,-0.167,-0.98,0.215,-0.323,-0.922,0.215,-0.323,-0.922,0.112,-0.167,-0.98,0.148,-0.358,-0.922,0.112,-0.167,-0.98,0.077,-0.186,-0.98,0.148,-0.358,-0.922,0.148,-0.358,-0.922,0.077,-0.186,-0.98,0.076,-0.38,-0.922,0.077,-0.186,-0.98,0.039,-0.197,-0.98,0.076,-0.38,-0.922,0.076,-0.38,-0.922,0.039,-0.197,-0.98,0.0,-0.388,-0.922,0.039,-0.197,-0.98,0.0,-0.201,-0.98,0.0,-0.388,-0.922,0.0,-0.388,-0.922,0.0,-0.201,-0.98,-0.076,-0.38,-0.922,0.0,-0.201,-0.98,-0.039,-0.197,-0.98,-0.076,-0.38,-0.922,-0.076,-0.38,-0.922,-0.039,-0.197,-0.98,-0.148,-0.358,-0.922,-0.039,-0.197,-0.98,-0.077,-0.186,-0.98,-0.148,-0.358,-0.922,-0.148,-0.358,-0.922,-0.077,-0.186,-0.98,-0.215,-0.323,-0.922,-0.077,-0.186,-0.98,-0.112,-0.167,-0.98,-0.215,-0.323,-0.922,-0.215,-0.323,-0.922,-0.112,-0.167,-0.98,-0.274,-0.274,-0.922,-0.112,-0.167,-0.98,-0.142,-0.142,-0.98,-0.274,-0.274,-0.922,-0.274,-0.274,-0.922,-0.142,-0.142,-0.98,-0.323,-0.215,-0.922,-0.142,-0.142,-0.98,-0.167,-0.112,-0.98,-0.323,-0.215,-0.922,-0.323,-0.215,-0.922,-0.167,-0.112,-0.98,-0.358,-0.148,-0.922,-0.167,-0.112,-0.98,-0.186,-0.077,-0.98,-0.358,-0.148,-0.922,-0.358,-0.148,-0.922,-0.186,-0.077,-0.98,-0.38,-0.076,-0.922,-0.186,-0.077,-0.98,-0.197,-0.039,-0.98,-0.38,-0.076,-0.922,-0.38,-0.076,-0.922,-0.197,-0.039,-0.98,-0.388,-0.0,-0.922,-0.197,-0.039,-0.98,-0.201,-0.0,-0.98,-0.388,-0.0,-0.922,-0.388,-0.0,-0.922,-0.201,-0.0,-0.98,-0.38,0.076,-0.922,-0.201,-0.0,-0.98,-0.197,0.039,-0.98,-0.38,0.076,-0.922,-0.38,0.076,-0.922,-0.197,0.039,-0.98,-0.358,0.148,-0.922,-0.197,0.039,-0.98,-0.186,0.077,-0.98,-0.358,0.148,-0.922,-0.358,0.148,-0.922,-0.186,0.077,-0.98,-0.323,0.215,-0.922,-0.186,0.077,-0.98,-0.167,0.112,-0.98,-0.323,0.215,-0.922,-0.323,0.215,-0.922,-0.167,0.112,-0.98,-0.274,0.274,-0.922,-0.167,0.112,-0.98,-0.142,0.142,-0.98,-0.274,0.274,-0.922,-0.274,0.274,-0.922,-0.142,0.142,-0.98,-0.215,0.323,-0.922,-0.142,0.142,-0.98,-0.112,0.167,-0.98,-0.215,0.323,-0.922,-0.215,0.323,-0.922,-0.112,0.167,-0.98,-0.148,0.358,-0.922,-0.112,0.167,-0.98,-0.077,0.186,-0.98,-0.148,0.358,-0.922,-0.148,0.358,-0.922,-0.077,0.186,-0.98,-0.076,0.38,-0.922,-0.077,0.186,-0.98,-0.039,0.197,-0.98,-0.076,0.38,-0.922,-0.076,0.38,-0.922,-0.039,0.197,-0.98,-0.0,0.388,-0.922,-0.039,0.197,-0.98,-0.0,0.201,-0.98,-0.0,0.388,-0.922,-0.0,0.56,-0.829,-0.0,0.388,-0.922,0.109,0.549,-0.829,-0.0,0.388,-0.922,0.076,0.38,-0.922,0.109,0.549,-0.829,0.109,0.549,-0.829,0.076,0.38,-0.922,0.214,0.517,-0.829,0.076,0.38,-0.922,0.148,0.358,-0.922,0.214,0.517,-0.829,0.214,0.517,-0.829,0.148,0.358,-0.922,0.311,0.465,-0.829,0.148,0.358,-0.922,0.215,0.323,-0.922,0.311,0.465,-0.829,0.311,0.465,-0.829,0.215,0.323,-0.922,0.396,0.396,-0.829,0.215,0.323,-0.922,0.274,0.274,-0.922,0.396,0.396,-0.829,0.396,0.396,-0.829,0.274,0.274,-0.922,0.465,0.311,-0.829,0.274,0.274,-0.922,0.323,0.215,-0.922,0.465,0.311,-0.829,0.465,0.311,-0.829,0.323,0.215,-0.922,0.517,0.214,-0.829,0.323,0.215,-0.922,0.358,0.148,-0.922,0.517,0.214,-0.829,0.517,0.214,-0.829,0.358,0.148,-0.922,0.549,0.109,-0.829,0.358,0.148,-0.922,0.38,0.076,-0.922,0.549,0.109,-0.829,0.549,0.109,-0.829,0.38,0.076,-0.922,0.56,0.0,-0.829,0.38,0.076,-0.922,0.388,0.0,-0.922,0.56,0.0,-0.829,0.56,0.0,-0.829,0.388,0.0,-0.922,0.549,-0.109,-0.829,0.388,0.0,-0.922,0.38,-0.076,-0.922,0.549,-0.109,-0.829,0.549,-0.109,-0.829,0.38,-0.076,-0.922,0.517,-0.214,-0.829,0.38,-0.076,-0.922,0.358,-0.148,-0.922,0.517,-0.214,-0.829,0.517,-0.214,-0.829,0.358,-0.148,-0.922,0.465,-0.311,-0.829,0.358,-0.148,-0.922,0.323,-0.215,-0.922,0.465,-0.311,-0.829,0.465,-0.311,-0.829,0.323,-0.215,-0.922,0.396,-0.396,-0.829,0.323,-0.215,-0.922,0.274,-0.274,-0.922,0.396,-0.396,-0.829,0.396,-0.396,-0.829,0.274,-0.274,-0.922,0.311,-0.465,-0.829,0.274,-0.274,-0.922,0.215,-0.323,-0.922,0.311,-0.465,-0.829,0.311,-0.465,-0.829,0.215,-0.323,-0.922,0.214,-0.517,-0.829,0.215,-0.323,-0.922,0.148,-0.358,-0.922,0.214,-0.517,-0.829,0.214,-0.517,-0.829,0.148,-0.358,-0.922,0.109,-0.549,-0.829,0.148,-0.358,-0.922,0.076,-0.38,-0.922,0.109,-0.549,-0.829,0.109,-0.549,-0.829,0.076,-0.38,-0.922,0.0,-0.56,-0.829,0.076,-0.38,-0.922,0.0,-0.388,-0.922,0.0,-0.56,-0.829,0.0,-0.56,-0.829,0.0,-0.388,-0.922,-0.109,-0.549,-0.829,0.0,-0.388,-0.922,-0.076,-0.38,-0.922,-0.109,-0.549,-0.829,-0.109,-0.549,-0.829,-0.076,-0.38,-0.922,-0.214,-0.517,-0.829,-0.076,-0.38,-0.922,-0.148,-0.358,-0.922,-0.214,-0.517,-0.829,-0.214,-0.517,-0.829,-0.148,-0.358,-0.922,-0.311,-0.465,-0.829,-0.148,-0.358,-0.922,-0.215,-0.323,-0.922,-0.311,-0.465,-0.829,-0.311,-0.465,-0.829,-0.215,-0.323,-0.922,-0.396,-0.396,-0.829,-0.215,-0.323,-0.922,-0.274,-0.274,-0.922,-0.396,-0.396,-0.829,-0.396,-0.396,-0.829,-0.274,-0.274,-0.922,-0.465,-0.311,-0.829,-0.274,-0.274,-0.922,-0.323,-0.215,-0.922,-0.465,-0.311,-0.829,-0.465,-0.311,-0.829,-0.323,-0.215,-0.922,-0.517,-0.214,-0.829,-0.323,-0.215,-0.922,-0.358,-0.148,-0.922,-0.517,-0.214,-0.829,-0.517,-0.214,-0.829,-0.358,-0.148,-0.922,-0.549,-0.109,-0.829,-0.358,-0.148,-0.922,-0.38,-0.076,-0.922,-0.549,-0.109,-0.829,-0.549,-0.109,-0.829,-0.38,-0.076,-0.922,-0.56,-0.0,-0.829,-0.38,-0.076,-0.922,-0.388,-0.0,-0.922,-0.56,-0.0,-0.829,-0.56,-0.0,-0.829,-0.388,-0.0,-0.922,-0.549,0.109,-0.829,-0.388,-0.0,-0.922,-0.38,0.076,-0.922,-0.549,0.109,-0.829,-0.549,0.109,-0.829,-0.38,0.076,-0.922,-0.517,0.214,-0.829,-0.38,0.076,-0.922,-0.358,0.148,-0.922,-0.517,0.214,-0.829,-0.517,0.214,-0.829,-0.358,0.148,-0.922,-0.465,0.311,-0.829,-0.358,0.148,-0.922,-0.323,0.215,-0.922,-0.465,0.311,-0.829,-0.465,0.311,-0.829,-0.323,0.215,-0.922,-0.396,0.396,-0.829,-0.323,0.215,-0.922,-0.274,0.274,-0.922,-0.396,0.396,-0.829,-0.396,0.396,-0.829,-0.274,0.274,-0.922,-0.311,0.465,-0.829,-0.274,0.274,-0.922,-0.215,0.323,-0.922,-0.311,0.465,-0.829,-0.311,0.465,-0.829,-0.215,0.323,-0.922,-0.214,0.517,-0.829,-0.215,0.323,-0.922,-0.148,0.358,-0.922,-0.214,0.517,-0.829,-0.214,0.517,-0.829,-0.148,0.358,-0.922,-0.109,0.549,-0.829,-0.148,0.358,-0.922,-0.076,0.38,-0.922,-0.109,0.549,-0.829,-0.109,0.549,-0.829,-0.076,0.38,-0.922,-0.0,0.56,-0.829,-0.076,0.38,-0.922,-0.0,0.388,-0.922,-0.0,0.56,-0.829,-0.0,0.71,-0.704,-0.0,0.56,-0.829,0.139,0.696,-0.704,-0.0,0.56,-0.829,0.109,0.549,-0.829,0.139,0.696,-0.704,0.139,0.696,-0.704,0.109,0.549,-0.829,0.272,0.656,-0.704,0.109,0.549,-0.829,0.214,0.517,-0.829,0.272,0.656,-0.704,0.272,0.656,-0.704,0.214,0.517,-0.829,0.395,0.59,-0.704,0.214,0.517,-0.829,0.311,0.465,-0.829,0.395,0.59,-0.704,0.395,0.59,-0.704,0.311,0.465,-0.829,0.502,0.502,-0.704,0.311,0.465,-0.829,0.396,0.396,-0.829,0.502,0.502,-0.704,0.502,0.502,-0.704,0.396,0.396,-0.829,0.59,0.395,-0.704,0.396,0.396,-0.829,0.465,0.311,-0.829,0.59,0.395,-0.704,0.59,0.395,-0.704,0.465,0.311,-0.829,0.656,0.272,-0.704,0.465,0.311,-0.829,0.517,0.214,-0.829,0.656,0.272,-0.704,0.656,0.272,-0.704,0.517,0.214,-0.829,0.696,0.139,-0.704,0.517,0.214,-0.829,0.549,0.109,-0.829,0.696,0.139,-0.704,0.696,0.139,-0.704,0.549,0.109,-0.829,0.71,0.0,-0.704,0.549,0.109,-0.829,0.56,0.0,-0.829,0.71,0.0,-0.704,0.71,0.0,-0.704,0.56,0.0,-0.829,0.696,-0.139,-0.704,0.56,0.0,-0.829,0.549,-0.109,-0.829,0.696,-0.139,-0.704,0.696,-0.139,-0.704,0.549,-0.109,-0.829,0.656,-0.272,-0.704,0.549,-0.109,-0.829,0.517,-0.214,-0.829,0.656,-0.272,-0.704,0.656,-0.272,-0.704,0.517,-0.214,-0.829,0.59,-0.395,-0.704,0.517,-0.214,-0.829,0.465,-0.311,-0.829,0.59,-0.395,-0.704,0.59,-0.395,-0.704,0.465,-0.311,-0.829,0.502,-0.502,-0.704,0.465,-0.311,-0.829,0.396,-0.396,-0.829,0.502,-0.502,-0.704,0.502,-0.502,-0.704,0.396,-0.396,-0.829,0.395,-0.59,-0.704,0.396,-0.396,-0.829,0.311,-0.465,-0.829,0.395,-0.59,-0.704,0.395,-0.59,-0.704,0.311,-0.465,-0.829,0.272,-0.656,-0.704,0.311,-0.465,-0.829,0.214,-0.517,-0.829,0.272,-0.656,-0.704,0.272,-0.656,-0.704,0.214,-0.517,-0.829,0.139,-0.696,-0.704,0.214,-0.517,-0.829,0.109,-0.549,-0.829,0.139,-0.696,-0.704,0.139,-0.696,-0.704,0.109,-0.549,-0.829,0.0,-0.71,-0.704,0.109,-0.549,-0.829,0.0,-0.56,-0.829,0.0,-0.71,-0.704,0.0,-0.71,-0.704,0.0,-0.56,-0.829,-0.139,-0.696,-0.704,0.0,-0.56,-0.829,-0.109,-0.549,-0.829,-0.139,-0.696,-0.704,-0.139,-0.696,-0.704,-0.109,-0.549,-0.829,-0.272,-0.656,-0.704,-0.109,-0.549,-0.829,-0.214,-0.517,-0.829,-0.272,-0.656,-0.704,-0.272,-0.656,-0.704,-0.214,-0.517,-0.829,-0.395,-0.59,-0.704,-0.214,-0.517,-0.829,-0.311,-0.465,-0.829,-0.395,-0.59,-0.704,-0.395,-0.59,-0.704,-0.311,-0.465,-0.829,-0.502,-0.502,-0.704,-0.311,-0.465,-0.829,-0.396,-0.396,-0.829,-0.502,-0.502,-0.704,-0.502,-0.502,-0.704,-0.396,-0.396,-0.829,-0.59,-0.395,-0.704,-0.396,-0.396,-0.829,-0.465,-0.311,-0.829,-0.59,-0.395,-0.704,-0.59,-0.395,-0.704,-0.465,-0.311,-0.829,-0.656,-0.272,-0.704,-0.465,-0.311,-0.829,-0.517,-0.214,-0.829,-0.656,-0.272,-0.704,-0.656,-0.272,-0.704,-0.517,-0.214,-0.829,-0.696,-0.139,-0.704,-0.517,-0.214,-0.829,-0.549,-0.109,-0.829,-0.696,-0.139,-0.704,-0.696,-0.139,-0.704,-0.549,-0.109,-0.829,-0.71,-0.0,-0.704,-0.549,-0.109,-0.829,-0.56,-0.0,-0.829,-0.71,-0.0,-0.704,-0.71,-0.0,-0.704,-0.56,-0.0,-0.829,-0.696,0.139,-0.704,-0.56,-0.0,-0.829,-0.549,0.109,-0.829,-0.696,0.139,-0.704,-0.696,0.139,-0.704,-0.549,0.109,-0.829,-0.656,0.272,-0.704,-0.549,0.109,-0.829,-0.517,0.214,-0.829,-0.656,0.272,-0.704,-0.656,0.272,-0.704,-0.517,0.214,-0.829,-0.59,0.395,-0.704,-0.517,0.214,-0.829,-0.465,0.311,-0.829,-0.59,0.395,-0.704,-0.59,0.395,-0.704,-0.465,0.311,-0.829,-0.502,0.502,-0.704,-0.465,0.311,-0.829,-0.396,0.396,-0.829,-0.502,0.502,-0.704,-0.502,0.502,-0.704,-0.396,0.396,-0.829,-0.395,0.59,-0.704,-0.396,0.396,-0.829,-0.311,0.465,-0.829,-0.395,0.59,-0.704,-0.395,0.59,-0.704,-0.311,0.465,-0.829,-0.272,0.656,-0.704,-0.311,0.465,-0.829,-0.214,0.517,-0.829,-0.272,0.656,-0.704,-0.272,0.656,-0.704,-0.214,0.517,-0.829,-0.139,0.696,-0.704,-0.214,0.517,-0.829,-0.109,0.549,-0.829,-0.139,0.696,-0.704,-0.139,0.696,-0.704,-0.109,0.549,-0.829,-0.0,0.71,-0.704,-0.109,0.549,-0.829,-0.0,0.56,-0.829,-0.0,0.71,-0.704,-0.0,0.833,-0.553,-0.0,0.71,-0.704,0.163,0.817,-0.553,-0.0,0.71,-0.704,0.139,0.696,-0.704,0.163,0.817,-0.553,0.163,0.817,-0.553,0.139,0.696,-0.704,0.319,0.77,-0.553,0.139,0.696,-0.704,0.272,0.656,-0.704,0.319,0.77,-0.553,0.319,0.77,-0.553,0.272,0.656,-0.704,0.463,0.693,-0.553,0.272,0.656,-0.704,0.395,0.59,-0.704,0.463,0.693,-0.553,0.463,0.693,-0.553,0.395,0.59,-0.704,0.589,0.589,-0.553,0.395,0.59,-0.704,0.502,0.502,-0.704,0.589,0.589,-0.553,0.589,0.589,-0.553,0.502,0.502,-0.704,0.693,0.463,-0.553,0.502,0.502,-0.704,0.59,0.395,-0.704,0.693,0.463,-0.553,0.693,0.463,-0.553,0.59,0.395,-0.704,0.77,0.319,-0.553,0.59,0.395,-0.704,0.656,0.272,-0.704,0.77,0.319,-0.553,0.77,0.319,-0.553,0.656,0.272,-0.704,0.817,0.163,-0.553,0.656,0.272,-0.704,0.696,0.139,-0.704,0.817,0.163,-0.553,0.817,0.163,-0.553,0.696,0.139,-0.704,0.833,0.0,-0.553,0.696,0.139,-0.704,0.71,0.0,-0.704,0.833,0.0,-0.553,0.833,0.0,-0.553,0.71,0.0,-0.704,0.817,-0.163,-0.553,0.71,0.0,-0.704,0.696,-0.139,-0.704,0.817,-0.163,-0.553,0.817,-0.163,-0.553,0.696,-0.139,-0.704,0.77,-0.319,-0.553,0.696,-0.139,-0.704,0.656,-0.272,-0.704,0.77,-0.319,-0.553,0.77,-0.319,-0.553,0.656,-0.272,-0.704,0.693,-0.463,-0.553,0.656,-0.272,-0.704,0.59,-0.395,-0.704,0.693,-0.463,-0.553,0.693,-0.463,-0.553,0.59,-0.395,-0.704,0.589,-0.589,-0.553,0.59,-0.395,-0.704,0.502,-0.502,-0.704,0.589,-0.589,-0.553,0.589,-0.589,-0.553,0.502,-0.502,-0.704,0.463,-0.693,-0.553,0.502,-0.502,-0.704,0.395,-0.59,-0.704,0.463,-0.693,-0.553,0.463,-0.693,-0.553,0.395,-0.59,-0.704,0.319,-0.77,-0.553,0.395,-0.59,-0.704,0.272,-0.656,-0.704,0.319,-0.77,-0.553,0.319,-0.77,-0.553,0.272,-0.656,-0.704,0.163,-0.817,-0.553,0.272,-0.656,-0.704,0.139,-0.696,-0.704,0.163,-0.817,-0.553,0.163,-0.817,-0.553,0.139,-0.696,-0.704,0.0,-0.833,-0.553,0.139,-0.696,-0.704,0.0,-0.71,-0.704,0.0,-0.833,-0.553,0.0,-0.833,-0.553,0.0,-0.71,-0.704,-0.163,-0.817,-0.553,0.0,-0.71,-0.704,-0.139,-0.696,-0.704,-0.163,-0.817,-0.553,-0.163,-0.817,-0.553,-0.139,-0.696,-0.704,-0.319,-0.77,-0.553,-0.139,-0.696,-0.704,-0.272,-0.656,-0.704,-0.319,-0.77,-0.553,-0.319,-0.77,-0.553,-0.272,-0.656,-0.704,-0.463,-0.693,-0.553,-0.272,-0.656,-0.704,-0.395,-0.59,-0.704,-0.463,-0.693,-0.553,-0.463,-0.693,-0.553,-0.395,-0.59,-0.704,-0.589,-0.589,-0.553,-0.395,-0.59,-0.704,-0.502,-0.502,-0.704,-0.589,-0.589,-0.553,-0.589,-0.589,-0.553,-0.502,-0.502,-0.704,-0.693,-0.463,-0.553,-0.502,-0.502,-0.704,-0.59,-0.395,-0.704,-0.693,-0.463,-0.553,-0.693,-0.463,-0.553,-0.59,-0.395,-0.704,-0.77,-0.319,-0.553,-0.59,-0.395,-0.704,-0.656,-0.272,-0.704,-0.77,-0.319,-0.553,-0.77,-0.319,-0.553,-0.656,-0.272,-0.704,-0.817,-0.163,-0.553,-0.656,-0.272,-0.704,-0.696,-0.139,-0.704,-0.817,-0.163,-0.553,-0.817,-0.163,-0.553,-0.696,-0.139,-0.704,-0.833,-0.0,-0.553,-0.696,-0.139,-0.704,-0.71,-0.0,-0.704,-0.833,-0.0,-0.553,-0.833,-0.0,-0.553,-0.71,-0.0,-0.704,-0.817,0.163,-0.553,-0.71,-0.0,-0.704,-0.696,0.139,-0.704,-0.817,0.163,-0.553,-0.817,0.163,-0.553,-0.696,0.139,-0.704,-0.77,0.319,-0.553,-0.696,0.139,-0.704,-0.656,0.272,-0.704,-0.77,0.319,-0.553,-0.77,0.319,-0.553,-0.656,0.272,-0.704,-0.693,0.463,-0.553,-0.656,0.272,-0.704,-0.59,0.395,-0.704,-0.693,0.463,-0.553,-0.693,0.463,-0.553,-0.59,0.395,-0.704,-0.589,0.589,-0.553,-0.59,0.395,-0.704,-0.502,0.502,-0.704,-0.589,0.589,-0.553,-0.589,0.589,-0.553,-0.502,0.502,-0.704,-0.463,0.693,-0.553,-0.502,0.502,-0.704,-0.395,0.59,-0.704,-0.463,0.693,-0.553,-0.463,0.693,-0.553,-0.395,0.59,-0.704,-0.319,0.77,-0.553,-0.395,0.59,-0.704,-0.272,0.656,-0.704,-0.319,0.77,-0.553,-0.319,0.77,-0.553,-0.272,0.656,-0.704,-0.163,0.817,-0.553,-0.272,0.656,-0.704,-0.139,0.696,-0.704,-0.163,0.817,-0.553,-0.163,0.817,-0.553,-0.139,0.696,-0.704,-0.0,0.833,-0.553,-0.139,0.696,-0.704,-0.0,0.71,-0.704,-0.0,0.833,-0.553,-0.0,0.925,-0.381,-0.0,0.833,-0.553,0.18,0.907,-0.381,-0.0,0.833,-0.553,0.163,0.817,-0.553,0.18,0.907,-0.381,0.18,0.907,-0.381,0.163,0.817,-0.553,0.354,0.854,-0.381,0.163,0.817,-0.553,0.319,0.77,-0.553,0.354,0.854,-0.381,0.354,0.854,-0.381,0.319,0.77,-0.553,0.514,0.769,-0.381,0.319,0.77,-0.553,0.463,0.693,-0.553,0.514,0.769,-0.381,0.514,0.769,-0.381,0.463,0.693,-0.553,0.654,0.654,-0.381,0.463,0.693,-0.553,0.589,0.589,-0.553,0.654,0.654,-0.381,0.654,0.654,-0.381,0.589,0.589,-0.553,0.769,0.514,-0.381,0.589,0.589,-0.553,0.693,0.463,-0.553,0.769,0.514,-0.381,0.769,0.514,-0.381,0.693,0.463,-0.553,0.854,0.354,-0.381,0.693,0.463,-0.553,0.77,0.319,-0.553,0.854,0.354,-0.381,0.854,0.354,-0.381,0.77,0.319,-0.553,0.907,0.18,-0.381,0.77,0.319,-0.553,0.817,0.163,-0.553,0.907,0.18,-0.381,0.907,0.18,-0.381,0.817,0.163,-0.553,0.925,0.0,-0.381,0.817,0.163,-0.553,0.833,0.0,-0.553,0.925,0.0,-0.381,0.925,0.0,-0.381,0.833,0.0,-0.553,0.907,-0.18,-0.381,0.833,0.0,-0.553,0.817,-0.163,-0.553,0.907,-0.18,-0.381,0.907,-0.18,-0.381,0.817,-0.163,-0.553,0.854,-0.354,-0.381,0.817,-0.163,-0.553,0.77,-0.319,-0.553,0.854,-0.354,-0.381,0.854,-0.354,-0.381,0.77,-0.319,-0.553,0.769,-0.514,-0.381,0.77,-0.319,-0.553,0.693,-0.463,-0.553,0.769,-0.514,-0.381,0.769,-0.514,-0.381,0.693,-0.463,-0.553,0.654,-0.654,-0.381,0.693,-0.463,-0.553,0.589,-0.589,-0.553,0.654,-0.654,-0.381,0.654,-0.654,-0.381,0.589,-0.589,-0.553,0.514,-0.769,-0.381,0.589,-0.589,-0.553,0.463,-0.693,-0.553,0.514,-0.769,-0.381,0.514,-0.769,-0.381,0.463,-0.693,-0.553,0.354,-0.854,-0.381,0.463,-0.693,-0.553,0.319,-0.77,-0.553,0.354,-0.854,-0.381,0.354,-0.854,-0.381,0.319,-0.77,-0.553,0.18,-0.907,-0.381,0.319,-0.77,-0.553,0.163,-0.817,-0.553,0.18,-0.907,-0.381,0.18,-0.907,-0.381,0.163,-0.817,-0.553,0.0,-0.925,-0.381,0.163,-0.817,-0.553,0.0,-0.833,-0.553,0.0,-0.925,-0.381,0.0,-0.925,-0.381,0.0,-0.833,-0.553,-0.18,-0.907,-0.381,0.0,-0.833,-0.553,-0.163,-0.817,-0.553,-0.18,-0.907,-0.381,-0.18,-0.907,-0.381,-0.163,-0.817,-0.553,-0.354,-0.854,-0.381,-0.163,-0.817,-0.553,-0.319,-0.77,-0.553,-0.354,-0.854,-0.381,-0.354,-0.854,-0.381,-0.319,-0.77,-0.553,-0.514,-0.769,-0.381,-0.319,-0.77,-0.553,-0.463,-0.693,-0.553,-0.514,-0.769,-0.381,-0.514,-0.769,-0.381,-0.463,-0.693,-0.553,-0.654,-0.654,-0.381,-0.463,-0.693,-0.553,-0.589,-0.589,-0.553,-0.654,-0.654,-0.381,-0.654,-0.654,-0.381,-0.589,-0.589,-0.553,-0.769,-0.514,-0.381,-0.589,-0.589,-0.553,-0.693,-0.463,-0.553,-0.769,-0.514,-0.381,-0.769,-0.514,-0.381,-0.693,-0.463,-0.553,-0.854,-0.354,-0.381,-0.693,-0.463,-0.553,-0.77,-0.319,-0.553,-0.854,-0.354,-0.381,-0.854,-0.354,-0.381,-0.77,-0.319,-0.553,-0.907,-0.18,-0.381,-0.77,-0.319,-0.553,-0.817,-0.163,-0.553,-0.907,-0.18,-0.381,-0.907,-0.18,-0.381,-0.817,-0.163,-0.553,-0.925,-0.0,-0.381,-0.817,-0.163,-0.553,-0.833,-0.0,-0.553,-0.925,-0.0,-0.381,-0.925,-0.0,-0.381,-0.833,-0.0,-0.553,-0.907,0.18,-0.381,-0.833,-0.0,-0.553,-0.817,0.163,-0.553,-0.907,0.18,-0.381,-0.907,0.18,-0.381,-0.817,0.163,-0.553,-0.854,0.354,-0.381,-0.817,0.163,-0.553,-0.77,0.319,-0.553,-0.854,0.354,-0.381,-0.854,0.354,-0.381,-0.77,0.319,-0.553,-0.769,0.514,-0.381,-0.77,0.319,-0.553,-0.693,0.463,-0.553,-0.769,0.514,-0.381,-0.769,0.514,-0.381,-0.693,0.463,-0.553,-0.654,0.654,-0.381,-0.693,0.463,-0.553,-0.589,0.589,-0.553,-0.654,0.654,-0.381,-0.654,0.654,-0.381,-0.589,0.589,-0.553,-0.514,0.769,-0.381,-0.589,0.589,-0.553,-0.463,0.693,-0.553,-0.514,0.769,-0.381,-0.514,0.769,-0.381,-0.463,0.693,-0.553,-0.354,0.854,-0.381,-0.463,0.693,-0.553,-0.319,0.77,-0.553,-0.354,0.854,-0.381,-0.354,0.854,-0.381,-0.319,0.77,-0.553,-0.18,0.907,-0.381,-0.319,0.77,-0.553,-0.163,0.817,-0.553,-0.18,0.907,-0.381,-0.18,0.907,-0.381,-0.163,0.817,-0.553,-0.0,0.925,-0.381,-0.163,0.817,-0.553,-0.0,0.833,-0.553,-0.0,0.925,-0.381,-0.0,0.981,-0.194,-0.0,0.925,-0.381,0.191,0.962,-0.194,-0.0,0.925,-0.381,0.18,0.907,-0.381,0.191,0.962,-0.194,0.191,0.962,-0.194,0.18,0.907,-0.381,0.375,0.906,-0.194,0.18,0.907,-0.381,0.354,0.854,-0.381,0.375,0.906,-0.194,0.375,0.906,-0.194,0.354,0.854,-0.381,0.545,0.816,-0.194,0.354,0.854,-0.381,0.514,0.769,-0.381,0.545,0.816,-0.194,0.545,0.816,-0.194,0.514,0.769,-0.381,0.694,0.694,-0.194,0.514,0.769,-0.381,0.654,0.654,-0.381,0.694,0.694,-0.194,0.694,0.694,-0.194,0.654,0.654,-0.381,0.816,0.545,-0.194,0.654,0.654,-0.381,0.769,0.514,-0.381,0.816,0.545,-0.194,0.816,0.545,-0.194,0.769,0.514,-0.381,0.906,0.375,-0.194,0.769,0.514,-0.381,0.854,0.354,-0.381,0.906,0.375,-0.194,0.906,0.375,-0.194,0.854,0.354,-0.381,0.962,0.191,-0.194,0.854,0.354,-0.381,0.907,0.18,-0.381,0.962,0.191,-0.194,0.962,0.191,-0.194,0.907,0.18,-0.381,0.981,0.0,-0.194,0.907,0.18,-0.381,0.925,0.0,-0.381,0.981,0.0,-0.194,0.981,0.0,-0.194,0.925,0.0,-0.381,0.962,-0.191,-0.194,0.925,0.0,-0.381,0.907,-0.18,-0.381,0.962,-0.191,-0.194,0.962,-0.191,-0.194,0.907,-0.18,-0.381,0.906,-0.375,-0.194,0.907,-0.18,-0.381,0.854,-0.354,-0.381,0.906,-0.375,-0.194,0.906,-0.375,-0.194,0.854,-0.354,-0.381,0.816,-0.545,-0.194,0.854,-0.354,-0.381,0.769,-0.514,-0.381,0.816,-0.545,-0.194,0.816,-0.545,-0.194,0.769,-0.514,-0.381,0.694,-0.694,-0.194,0.769,-0.514,-0.381,0.654,-0.654,-0.381,0.694,-0.694,-0.194,0.694,-0.694,-0.194,0.654,-0.654,-0.381,0.545,-0.816,-0.194,0.654,-0.654,-0.381,0.514,-0.769,-0.381,0.545,-0.816,-0.194,0.545,-0.816,-0.194,0.514,-0.769,-0.381,0.375,-0.906,-0.194,0.514,-0.769,-0.381,0.354,-0.854,-0.381,0.375,-0.906,-0.194,0.375,-0.906,-0.194,0.354,-0.854,-0.381,0.191,-0.962,-0.194,0.354,-0.854,-0.381,0.18,-0.907,-0.381,0.191,-0.962,-0.194,0.191,-0.962,-0.194,0.18,-0.907,-0.381,0.0,-0.981,-0.194,0.18,-0.907,-0.381,0.0,-0.925,-0.381,0.0,-0.981,-0.194,0.0,-0.981,-0.194,0.0,-0.925,-0.381,-0.191,-0.962,-0.194,0.0,-0.925,-0.381,-0.18,-0.907,-0.381,-0.191,-0.962,-0.194,-0.191,-0.962,-0.194,-0.18,-0.907,-0.381,-0.375,-0.906,-0.194,-0.18,-0.907,-0.381,-0.354,-0.854,-0.381,-0.375,-0.906,-0.194,-0.375,-0.906,-0.194,-0.354,-0.854,-0.381,-0.545,-0.816,-0.194,-0.354,-0.854,-0.381,-0.514,-0.769,-0.381,-0.545,-0.816,-0.194,-0.545,-0.816,-0.194,-0.514,-0.769,-0.381,-0.694,-0.694,-0.194,-0.514,-0.769,-0.381,-0.654,-0.654,-0.381,-0.694,-0.694,-0.194,-0.694,-0.694,-0.194,-0.654,-0.654,-0.381,-0.816,-0.545,-0.194,-0.654,-0.654,-0.381,-0.769,-0.514,-0.381,-0.816,-0.545,-0.194,-0.816,-0.545,-0.194,-0.769,-0.514,-0.381,-0.906,-0.375,-0.194,-0.769,-0.514,-0.381,-0.854,-0.354,-0.381,-0.906,-0.375,-0.194,-0.906,-0.375,-0.194,-0.854,-0.354,-0.381,-0.962,-0.191,-0.194,-0.854,-0.354,-0.381,-0.907,-0.18,-0.381,-0.962,-0.191,-0.194,-0.962,-0.191,-0.194,-0.907,-0.18,-0.381,-0.981,-0.0,-0.194,-0.907,-0.18,-0.381,-0.925,-0.0,-0.381,-0.981,-0.0,-0.194,-0.981,-0.0,-0.194,-0.925,-0.0,-0.381,-0.962,0.191,-0.194,-0.925,-0.0,-0.381,-0.907,0.18,-0.381,-0.962,0.191,-0.194,-0.962,0.191,-0.194,-0.907,0.18,-0.381,-0.906,0.375,-0.194,-0.907,0.18,-0.381,-0.854,0.354,-0.381,-0.906,0.375,-0.194,-0.906,0.375,-0.194,-0.854,0.354,-0.381,-0.816,0.545,-0.194,-0.854,0.354,-0.381,-0.769,0.514,-0.381,-0.816,0.545,-0.194,-0.816,0.545,-0.194,-0.769,0.514,-0.381,-0.694,0.694,-0.194,-0.769,0.514,-0.381,-0.654,0.654,-0.381,-0.694,0.694,-0.194,-0.694,0.694,-0.194,-0.654,0.654,-0.381,-0.545,0.816,-0.194,-0.654,0.654,-0.381,-0.514,0.769,-0.381,-0.545,0.816,-0.194,-0.545,0.816,-0.194,-0.514,0.769,-0.381,-0.375,0.906,-0.194,-0.514,0.769,-0.381,-0.354,0.854,-0.381,-0.375,0.906,-0.194,-0.375,0.906,-0.194,-0.354,0.854,-0.381,-0.191,0.962,-0.194,-0.354,0.854,-0.381,-0.18,0.907,-0.381,-0.191,0.962,-0.194,-0.191,0.962,-0.194,-0.18,0.907,-0.381,-0.0,0.981,-0.194,-0.18,0.907,-0.381,-0.0,0.925,-0.381,-0.0,0.981,-0.194,-0.0,1.0,0.0,-0.0,0.981,-0.194,0.195,0.981,0.0,-0.0,0.981,-0.194,0.191,0.962,-0.194,0.195,0.981,0.0,0.195,0.981,0.0,0.191,0.962,-0.194,0.383,0.924,0.0,0.191,0.962,-0.194,0.375,0.906,-0.194,0.383,0.924,0.0,0.383,0.924,0.0,0.375,0.906,-0.194,0.556,0.831,0.0,0.375,0.906,-0.194,0.545,0.816,-0.194,0.556,0.831,0.0,0.556,0.831,0.0,0.545,0.816,-0.194,0.707,0.707,0.0,0.545,0.816,-0.194,0.694,0.694,-0.194,0.707,0.707,0.0,0.707,0.707,0.0,0.694,0.694,-0.194,0.831,0.556,0.0,0.694,0.694,-0.194,0.816,0.545,-0.194,0.831,0.556,0.0,0.831,0.556,0.0,0.816,0.545,-0.194,0.924,0.383,0.0,0.816,0.545,-0.194,0.906,0.375,-0.194,0.924,0.383,0.0,0.924,0.383,0.0,0.906,0.375,-0.194,0.981,0.195,0.0,0.906,0.375,-0.194,0.962,0.191,-0.194,0.981,0.195,0.0,0.981,0.195,0.0,0.962,0.191,-0.194,1.0,0.0,0.0,0.962,0.191,-0.194,0.981,0.0,-0.194,1.0,0.0,0.0,1.0,0.0,0.0,0.981,0.0,-0.194,0.981,-0.195,0.0,0.981,0.0,-0.194,0.962,-0.191,-0.194,0.981,-0.195,0.0,0.981,-0.195,0.0,0.962,-0.191,-0.194,0.924,-0.383,0.0,0.962,-0.191,-0.194,0.906,-0.375,-0.194,0.924,-0.383,0.0,0.924,-0.383,0.0,0.906,-0.375,-0.194,0.831,-0.556,0.0,0.906,-0.375,-0.194,0.816,-0.545,-0.194,0.831,-0.556,0.0,0.831,-0.556,0.0,0.816,-0.545,-0.194,0.707,-0.707,0.0,0.816,-0.545,-0.194,0.694,-0.694,-0.194,0.707,-0.707,0.0,0.707,-0.707,0.0,0.694,-0.694,-0.194,0.556,-0.831,0.0,0.694,-0.694,-0.194,0.545,-0.816,-0.194,0.556,-0.831,0.0,0.556,-0.831,0.0,0.545,-0.816,-0.194,0.383,-0.924,0.0,0.545,-0.816,-0.194,0.375,-0.906,-0.194,0.383,-0.924,0.0,0.383,-0.924,0.0,0.375,-0.906,-0.194,0.195,-0.981,0.0,0.375,-0.906,-0.194,0.191,-0.962,-0.194,0.195,-0.981,0.0,0.195,-0.981,0.0,0.191,-0.962,-0.194,0.0,-1.0,0.0,0.191,-0.962,-0.194,0.0,-0.981,-0.194,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-0.981,-0.194,-0.195,-0.981,0.0,0.0,-0.981,-0.194,-0.191,-0.962,-0.194,-0.195,-0.981,0.0,-0.195,-0.981,0.0,-0.191,-0.962,-0.194,-0.383,-0.924,0.0,-0.191,-0.962,-0.194,-0.375,-0.906,-0.194,-0.383,-0.924,0.0,-0.383,-0.924,0.0,-0.375,-0.906,-0.194,-0.556,-0.831,0.0,-0.375,-0.906,-0.194,-0.545,-0.816,-0.194,-0.556,-0.831,0.0,-0.556,-0.831,0.0,-0.545,-0.816,-0.194,-0.707,-0.707,0.0,-0.545,-0.816,-0.194,-0.694,-0.694,-0.194,-0.707,-0.707,0.0,-0.707,-0.707,0.0,-0.694,-0.694,-0.194,-0.831,-0.556,0.0,-0.694,-0.694,-0.194,-0.816,-0.545,-0.194,-0.831,-0.556,0.0,-0.831,-0.556,0.0,-0.816,-0.545,-0.194,-0.924,-0.383,0.0,-0.816,-0.545,-0.194,-0.906,-0.375,-0.194,-0.924,-0.383,0.0,-0.924,-0.383,0.0,-0.906,-0.375,-0.194,-0.981,-0.195,0.0,-0.906,-0.375,-0.194,-0.962,-0.191,-0.194,-0.981,-0.195,0.0,-0.981,-0.195,0.0,-0.962,-0.191,-0.194,-1.0,-0.0,0.0,-0.962,-0.191,-0.194,-0.981,-0.0,-0.194,-1.0,-0.0,0.0,-1.0,-0.0,0.0,-0.981,-0.0,-0.194,-0.981,0.195,0.0,-0.981,-0.0,-0.194,-0.962,0.191,-0.194,-0.981,0.195,0.0,-0.981,0.195,0.0,-0.962,0.191,-0.194,-0.924,0.383,0.0,-0.962,0.191,-0.194,-0.906,0.375,-0.194,-0.924,0.383,0.0,-0.924,0.383,0.0,-0.906,0.375,-0.194,-0.831,0.556,0.0,-0.906,0.375,-0.194,-0.816,0.545,-0.194,-0.831,0.556,0.0,-0.831,0.556,0.0,-0.816,0.545,-0.194,-0.707,0.707,0.0,-0.816,0.545,-0.194,-0.694,0.694,-0.194,-0.707,0.707,0.0,-0.707,0.707,0.0,-0.694,0.694,-0.194,-0.556,0.831,0.0,-0.694,0.694,-0.194,-0.545,0.816,-0.194,-0.556,0.831,0.0,-0.556,0.831,0.0,-0.545,0.816,-0.194,-0.383,0.924,0.0,-0.545,0.816,-0.194,-0.375,0.906,-0.194,-0.383,0.924,0.0,-0.383,0.924,0.0,-0.375,0.906,-0.194,-0.195,0.981,0.0,-0.375,0.906,-0.194,-0.191,0.962,-0.194,-0.195,0.981,0.0,-0.195,0.981,0.0,-0.191,0.962,-0.194,-0.0,1.0,0.0,-0.191,0.962,-0.194,-0.0,0.981,-0.194,-0.0,1.0,0.0,-0.0,0.981,0.194,-0.0,1.0,0.0,0.191,0.962,0.194,-0.0,1.0,0.0,0.195,0.981,0.0,0.191,0.962,0.194,0.191,0.962,0.194,0.195,0.981,0.0,0.375,0.906,0.194,0.195,0.981,0.0,0.383,0.924,0.0,0.375,0.906,0.194,0.375,0.906,0.194,0.383,0.924,0.0,0.545,0.816,0.194,0.383,0.924,0.0,0.556,0.831,0.0,0.545,0.816,0.194,0.545,0.816,0.194,0.556,0.831,0.0,0.694,0.694,0.194,0.556,0.831,0.0,0.707,0.707,0.0,0.694,0.694,0.194,0.694,0.694,0.194,0.707,0.707,0.0,0.816,0.545,0.194,0.707,0.707,0.0,0.831,0.556,0.0,0.816,0.545,0.194,0.816,0.545,0.194,0.831,0.556,0.0,0.906,0.375,0.194,0.831,0.556,0.0,0.924,0.383,0.0,0.906,0.375,0.194,0.906,0.375,0.194,0.924,0.383,0.0,0.962,0.191,0.194,0.924,0.383,0.0,0.981,0.195,0.0,0.962,0.191,0.194,0.962,0.191,0.194,0.981,0.195,0.0,0.981,0.0,0.194,0.981,0.195,0.0,1.0,0.0,0.0,0.981,0.0,0.194,0.981,0.0,0.194,1.0,0.0,0.0,0.962,-0.191,0.194,1.0,0.0,0.0,0.981,-0.195,0.0,0.962,-0.191,0.194,0.962,-0.191,0.194,0.981,-0.195,0.0,0.906,-0.375,0.194,0.981,-0.195,0.0,0.924,-0.383,0.0,0.906,-0.375,0.194,0.906,-0.375,0.194,0.924,-0.383,0.0,0.816,-0.545,0.194,0.924,-0.383,0.0,0.831,-0.556,0.0,0.816,-0.545,0.194,0.816,-0.545,0.194,0.831,-0.556,0.0,0.694,-0.694,0.194,0.831,-0.556,0.0,0.707,-0.707,0.0,0.694,-0.694,0.194,0.694,-0.694,0.194,0.707,-0.707,0.0,0.545,-0.816,0.194,0.707,-0.707,0.0,0.556,-0.831,0.0,0.545,-0.816,0.194,0.545,-0.816,0.194,0.556,-0.831,0.0,0.375,-0.906,0.194,0.556,-0.831,0.0,0.383,-0.924,0.0,0.375,-0.906,0.194,0.375,-0.906,0.194,0.383,-0.924,0.0,0.191,-0.962,0.194,0.383,-0.924,0.0,0.195,-0.981,0.0,0.191,-0.962,0.194,0.191,-0.962,0.194,0.195,-0.981,0.0,0.0,-0.981,0.194,0.195,-0.981,0.0,0.0,-1.0,0.0,0.0,-0.981,0.194,0.0,-0.981,0.194,0.0,-1.0,0.0,-0.191,-0.962,0.194,0.0,-1.0,0.0,-0.195,-0.981,0.0,-0.191,-0.962,0.194,-0.191,-0.962,0.194,-0.195,-0.981,0.0,-0.375,-0.906,0.194,-0.195,-0.981,0.0,-0.383,-0.924,0.0,-0.375,-0.906,0.194,-0.375,-0.906,0.194,-0.383,-0.924,0.0,-0.545,-0.816,0.194,-0.383,-0.924,0.0,-0.556,-0.831,0.0,-0.545,-0.816,0.194,-0.545,-0.816,0.194,-0.556,-0.831,0.0,-0.694,-0.694,0.194,-0.556,-0.831,0.0,-0.707,-0.707,0.0,-0.694,-0.694,0.194,-0.694,-0.694,0.194,-0.707,-0.707,0.0,-0.816,-0.545,0.194,-0.707,-0.707,0.0,-0.831,-0.556,0.0,-0.816,-0.545,0.194,-0.816,-0.545,0.194,-0.831,-0.556,0.0,-0.906,-0.375,0.194,-0.831,-0.556,0.0,-0.924,-0.383,0.0,-0.906,-0.375,0.194,-0.906,-0.375,0.194,-0.924,-0.383,0.0,-0.962,-0.191,0.194,-0.924,-0.383,0.0,-0.981,-0.195,0.0,-0.962,-0.191,0.194,-0.962,-0.191,0.194,-0.981,-0.195,0.0,-0.981,-0.0,0.194,-0.981,-0.195,0.0,-1.0,-0.0,0.0,-0.981,-0.0,0.194,-0.981,-0.0,0.194,-1.0,-0.0,0.0,-0.962,0.191,0.194,-1.0,-0.0,0.0,-0.981,0.195,0.0,-0.962,0.191,0.194,-0.962,0.191,0.194,-0.981,0.195,0.0,-0.906,0.375,0.194,-0.981,0.195,0.0,-0.924,0.383,0.0,-0.906,0.375,0.194,-0.906,0.375,0.194,-0.924,0.383,0.0,-0.816,0.545,0.194,-0.924,0.383,0.0,-0.831,0.556,0.0,-0.816,0.545,0.194,-0.816,0.545,0.194,-0.831,0.556,0.0,-0.694,0.694,0.194,-0.831,0.556,0.0,-0.707,0.707,0.0,-0.694,0.694,0.194,-0.694,0.694,0.194,-0.707,0.707,0.0,-0.545,0.816,0.194,-0.707,0.707,0.0,-0.556,0.831,0.0,-0.545,0.816,0.194,-0.545,0.816,0.194,-0.556,0.831,0.0,-0.375,0.906,0.194,-0.556,0.831,0.0,-0.383,0.924,0.0,-0.375,0.906,0.194,-0.375,0.906,0.194,-0.383,0.924,0.0,-0.191,0.962,0.194,-0.383,0.924,0.0,-0.195,0.981,0.0,-0.191,0.962,0.194,-0.191,0.962,0.194,-0.195,0.981,0.0,-0.0,0.981,0.194,-0.195,0.981,0.0,-0.0,1.0,0.0,-0.0,0.981,0.194,-0.0,0.925,0.381,-0.0,0.981,0.194,0.18,0.907,0.381,-0.0,0.981,0.194,0.191,0.962,0.194,0.18,0.907,0.381,0.18,0.907,0.381,0.191,0.962,0.194,0.354,0.854,0.381,0.191,0.962,0.194,0.375,0.906,0.194,0.354,0.854,0.381,0.354,0.854,0.381,0.375,0.906,0.194,0.514,0.769,0.381,0.375,0.906,0.194,0.545,0.816,0.194,0.514,0.769,0.381,0.514,0.769,0.381,0.545,0.816,0.194,0.654,0.654,0.381,0.545,0.816,0.194,0.694,0.694,0.194,0.654,0.654,0.381,0.654,0.654,0.381,0.694,0.694,0.194,0.769,0.514,0.381,0.694,0.694,0.194,0.816,0.545,0.194,0.769,0.514,0.381,0.769,0.514,0.381,0.816,0.545,0.194,0.854,0.354,0.381,0.816,0.545,0.194,0.906,0.375,0.194,0.854,0.354,0.381,0.854,0.354,0.381,0.906,0.375,0.194,0.907,0.18,0.381,0.906,0.375,0.194,0.962,0.191,0.194,0.907,0.18,0.381,0.907,0.18,0.381,0.962,0.191,0.194,0.925,0.0,0.381,0.962,0.191,0.194,0.981,0.0,0.194,0.925,0.0,0.381,0.925,0.0,0.381,0.981,0.0,0.194,0.907,-0.18,0.381,0.981,0.0,0.194,0.962,-0.191,0.194,0.907,-0.18,0.381,0.907,-0.18,0.381,0.962,-0.191,0.194,0.854,-0.354,0.381,0.962,-0.191,0.194,0.906,-0.375,0.194,0.854,-0.354,0.381,0.854,-0.354,0.381,0.906,-0.375,0.194,0.769,-0.514,0.381,0.906,-0.375,0.194,0.816,-0.545,0.194,0.769,-0.514,0.381,0.769,-0.514,0.381,0.816,-0.545,0.194,0.654,-0.654,0.381,0.816,-0.545,0.194,0.694,-0.694,0.194,0.654,-0.654,0.381,0.654,-0.654,0.381,0.694,-0.694,0.194,0.514,-0.769,0.381,0.694,-0.694,0.194,0.545,-0.816,0.194,0.514,-0.769,0.381,0.514,-0.769,0.381,0.545,-0.816,0.194,0.354,-0.854,0.381,0.545,-0.816,0.194,0.375,-0.906,0.194,0.354,-0.854,0.381,0.354,-0.854,0.381,0.375,-0.906,0.194,0.18,-0.907,0.381,0.375,-0.906,0.194,0.191,-0.962,0.194,0.18,-0.907,0.381,0.18,-0.907,0.381,0.191,-0.962,0.194,0.0,-0.925,0.381,0.191,-0.962,0.194,0.0,-0.981,0.194,0.0,-0.925,0.381,0.0,-0.925,0.381,0.0,-0.981,0.194,-0.18,-0.907,0.381,0.0,-0.981,0.194,-0.191,-0.962,0.194,-0.18,-0.907,0.381,-0.18,-0.907,0.381,-0.191,-0.962,0.194,-0.354,-0.854,0.381,-0.191,-0.962,0.194,-0.375,-0.906,0.194,-0.354,-0.854,0.381,-0.354,-0.854,0.381,-0.375,-0.906,0.194,-0.514,-0.769,0.381,-0.375,-0.906,0.194,-0.545,-0.816,0.194,-0.514,-0.769,0.381,-0.514,-0.769,0.381,-0.545,-0.816,0.194,-0.654,-0.654,0.381,-0.545,-0.816,0.194,-0.694,-0.694,0.194,-0.654,-0.654,0.381,-0.654,-0.654,0.381,-0.694,-0.694,0.194,-0.769,-0.514,0.381,-0.694,-0.694,0.194,-0.816,-0.545,0.194,-0.769,-0.514,0.381,-0.769,-0.514,0.381,-0.816,-0.545,0.194,-0.854,-0.354,0.381,-0.816,-0.545,0.194,-0.906,-0.375,0.194,-0.854,-0.354,0.381,-0.854,-0.354,0.381,-0.906,-0.375,0.194,-0.907,-0.18,0.381,-0.906,-0.375,0.194,-0.962,-0.191,0.194,-0.907,-0.18,0.381,-0.907,-0.18,0.381,-0.962,-0.191,0.194,-0.925,-0.0,0.381,-0.962,-0.191,0.194,-0.981,-0.0,0.194,-0.925,-0.0,0.381,-0.925,-0.0,0.381,-0.981,-0.0,0.194,-0.907,0.18,0.381,-0.981,-0.0,0.194,-0.962,0.191,0.194,-0.907,0.18,0.381,-0.907,0.18,0.381,-0.962,0.191,0.194,-0.854,0.354,0.381,-0.962,0.191,0.194,-0.906,0.375,0.194,-0.854,0.354,0.381,-0.854,0.354,0.381,-0.906,0.375,0.194,-0.769,0.514,0.381,-0.906,0.375,0.194,-0.816,0.545,0.194,-0.769,0.514,0.381,-0.769,0.514,0.381,-0.816,0.545,0.194,-0.654,0.654,0.381,-0.816,0.545,0.194,-0.694,0.694,0.194,-0.654,0.654,0.381,-0.654,0.654,0.381,-0.694,0.694,0.194,-0.514,0.769,0.381,-0.694,0.694,0.194,-0.545,0.816,0.194,-0.514,0.769,0.381,-0.514,0.769,0.381,-0.545,0.816,0.194,-0.354,0.854,0.381,-0.545,0.816,0.194,-0.375,0.906,0.194,-0.354,0.854,0.381,-0.354,0.854,0.381,-0.375,0.906,0.194,-0.18,0.907,0.381,-0.375,0.906,0.194,-0.191,0.962,0.194,-0.18,0.907,0.381,-0.18,0.907,0.381,-0.191,0.962,0.194,-0.0,0.925,0.381,-0.191,0.962,0.194,-0.0,0.981,0.194,-0.0,0.925,0.381,-0.0,0.833,0.553,-0.0,0.925,0.381,0.163,0.817,0.553,-0.0,0.925,0.381,0.18,0.907,0.381,0.163,0.817,0.553,0.163,0.817,0.553,0.18,0.907,0.381,0.319,0.77,0.553,0.18,0.907,0.381,0.354,0.854,0.381,0.319,0.77,0.553,0.319,0.77,0.553,0.354,0.854,0.381,0.463,0.693,0.553,0.354,0.854,0.381,0.514,0.769,0.381,0.463,0.693,0.553,0.463,0.693,0.553,0.514,0.769,0.381,0.589,0.589,0.553,0.514,0.769,0.381,0.654,0.654,0.381,0.589,0.589,0.553,0.589,0.589,0.553,0.654,0.654,0.381,0.693,0.463,0.553,0.654,0.654,0.381,0.769,0.514,0.381,0.693,0.463,0.553,0.693,0.463,0.553,0.769,0.514,0.381,0.77,0.319,0.553,0.769,0.514,0.381,0.854,0.354,0.381,0.77,0.319,0.553,0.77,0.319,0.553,0.854,0.354,0.381,0.817,0.163,0.553,0.854,0.354,0.381,0.907,0.18,0.381,0.817,0.163,0.553,0.817,0.163,0.553,0.907,0.18,0.381,0.833,0.0,0.553,0.907,0.18,0.381,0.925,0.0,0.381,0.833,0.0,0.553,0.833,0.0,0.553,0.925,0.0,0.381,0.817,-0.163,0.553,0.925,0.0,0.381,0.907,-0.18,0.381,0.817,-0.163,0.553,0.817,-0.163,0.553,0.907,-0.18,0.381,0.77,-0.319,0.553,0.907,-0.18,0.381,0.854,-0.354,0.381,0.77,-0.319,0.553,0.77,-0.319,0.553,0.854,-0.354,0.381,0.693,-0.463,0.553,0.854,-0.354,0.381,0.769,-0.514,0.381,0.693,-0.463,0.553,0.693,-0.463,0.553,0.769,-0.514,0.381,0.589,-0.589,0.553,0.769,-0.514,0.381,0.654,-0.654,0.381,0.589,-0.589,0.553,0.589,-0.589,0.553,0.654,-0.654,0.381,0.463,-0.693,0.553,0.654,-0.654,0.381,0.514,-0.769,0.381,0.463,-0.693,0.553,0.463,-0.693,0.553,0.514,-0.769,0.381,0.319,-0.77,0.553,0.514,-0.769,0.381,0.354,-0.854,0.381,0.319,-0.77,0.553,0.319,-0.77,0.553,0.354,-0.854,0.381,0.163,-0.817,0.553,0.354,-0.854,0.381,0.18,-0.907,0.381,0.163,-0.817,0.553,0.163,-0.817,0.553,0.18,-0.907,0.381,0.0,-0.833,0.553,0.18,-0.907,0.381,0.0,-0.925,0.381,0.0,-0.833,0.553,0.0,-0.833,0.553,0.0,-0.925,0.381,-0.163,-0.817,0.553,0.0,-0.925,0.381,-0.18,-0.907,0.381,-0.163,-0.817,0.553,-0.163,-0.817,0.553,-0.18,-0.907,0.381,-0.319,-0.77,0.553,-0.18,-0.907,0.381,-0.354,-0.854,0.381,-0.319,-0.77,0.553,-0.319,-0.77,0.553,-0.354,-0.854,0.381,-0.463,-0.693,0.553,-0.354,-0.854,0.381,-0.514,-0.769,0.381,-0.463,-0.693,0.553,-0.463,-0.693,0.553,-0.514,-0.769,0.381,-0.589,-0.589,0.553,-0.514,-0.769,0.381,-0.654,-0.654,0.381,-0.589,-0.589,0.553,-0.589,-0.589,0.553,-0.654,-0.654,0.381,-0.693,-0.463,0.553,-0.654,-0.654,0.381,-0.769,-0.514,0.381,-0.693,-0.463,0.553,-0.693,-0.463,0.553,-0.769,-0.514,0.381,-0.77,-0.319,0.553,-0.769,-0.514,0.381,-0.854,-0.354,0.381,-0.77,-0.319,0.553,-0.77,-0.319,0.553,-0.854,-0.354,0.381,-0.817,-0.163,0.553,-0.854,-0.354,0.381,-0.907,-0.18,0.381,-0.817,-0.163,0.553,-0.817,-0.163,0.553,-0.907,-0.18,0.381,-0.833,-0.0,0.553,-0.907,-0.18,0.381,-0.925,-0.0,0.381,-0.833,-0.0,0.553,-0.833,-0.0,0.553,-0.925,-0.0,0.381,-0.817,0.163,0.553,-0.925,-0.0,0.381,-0.907,0.18,0.381,-0.817,0.163,0.553,-0.817,0.163,0.553,-0.907,0.18,0.381,-0.77,0.319,0.553,-0.907,0.18,0.381,-0.854,0.354,0.381,-0.77,0.319,0.553,-0.77,0.319,0.553,-0.854,0.354,0.381,-0.693,0.463,0.553,-0.854,0.354,0.381,-0.769,0.514,0.381,-0.693,0.463,0.553,-0.693,0.463,0.553,-0.769,0.514,0.381,-0.589,0.589,0.553,-0.769,0.514,0.381,-0.654,0.654,0.381,-0.589,0.589,0.553,-0.589,0.589,0.553,-0.654,0.654,0.381,-0.463,0.693,0.553,-0.654,0.654,0.381,-0.514,0.769,0.381,-0.463,0.693,0.553,-0.463,0.693,0.553,-0.514,0.769,0.381,-0.319,0.77,0.553,-0.514,0.769,0.381,-0.354,0.854,0.381,-0.319,0.77,0.553,-0.319,0.77,0.553,-0.354,0.854,0.381,-0.163,0.817,0.553,-0.354,0.854,0.381,-0.18,0.907,0.381,-0.163,0.817,0.553,-0.163,0.817,0.553,-0.18,0.907,0.381,-0.0,0.833,0.553,-0.18,0.907,0.381,-0.0,0.925,0.381,-0.0,0.833,0.553,-0.0,0.71,0.704,-0.0,0.833,0.553,0.139,0.696,0.704,-0.0,0.833,0.553,0.163,0.817,0.553,0.139,0.696,0.704,0.139,0.696,0.704,0.163,0.817,0.553,0.272,0.656,0.704,0.163,0.817,0.553,0.319,0.77,0.553,0.272,0.656,0.704,0.272,0.656,0.704,0.319,0.77,0.553,0.395,0.59,0.704,0.319,0.77,0.553,0.463,0.693,0.553,0.395,0.59,0.704,0.395,0.59,0.704,0.463,0.693,0.553,0.502,0.502,0.704,0.463,0.693,0.553,0.589,0.589,0.553,0.502,0.502,0.704,0.502,0.502,0.704,0.589,0.589,0.553,0.59,0.395,0.704,0.589,0.589,0.553,0.693,0.463,0.553,0.59,0.395,0.704,0.59,0.395,0.704,0.693,0.463,0.553,0.656,0.272,0.704,0.693,0.463,0.553,0.77,0.319,0.553,0.656,0.272,0.704,0.656,0.272,0.704,0.77,0.319,0.553,0.696,0.139,0.704,0.77,0.319,0.553,0.817,0.163,0.553,0.696,0.139,0.704,0.696,0.139,0.704,0.817,0.163,0.553,0.71,0.0,0.704,0.817,0.163,0.553,0.833,0.0,0.553,0.71,0.0,0.704,0.71,0.0,0.704,0.833,0.0,0.553,0.696,-0.139,0.704,0.833,0.0,0.553,0.817,-0.163,0.553,0.696,-0.139,0.704,0.696,-0.139,0.704,0.817,-0.163,0.553,0.656,-0.272,0.704,0.817,-0.163,0.553,0.77,-0.319,0.553,0.656,-0.272,0.704,0.656,-0.272,0.704,0.77,-0.319,0.553,0.59,-0.395,0.704,0.77,-0.319,0.553,0.693,-0.463,0.553,0.59,-0.395,0.704,0.59,-0.395,0.704,0.693,-0.463,0.553,0.502,-0.502,0.704,0.693,-0.463,0.553,0.589,-0.589,0.553,0.502,-0.502,0.704,0.502,-0.502,0.704,0.589,-0.589,0.553,0.395,-0.59,0.704,0.589,-0.589,0.553,0.463,-0.693,0.553,0.395,-0.59,0.704,0.395,-0.59,0.704,0.463,-0.693,0.553,0.272,-0.656,0.704,0.463,-0.693,0.553,0.319,-0.77,0.553,0.272,-0.656,0.704,0.272,-0.656,0.704,0.319,-0.77,0.553,0.139,-0.696,0.704,0.319,-0.77,0.553,0.163,-0.817,0.553,0.139,-0.696,0.704,0.139,-0.696,0.704,0.163,-0.817,0.553,0.0,-0.71,0.704,0.163,-0.817,0.553,0.0,-0.833,0.553,0.0,-0.71,0.704,0.0,-0.71,0.704,0.0,-0.833,0.553,-0.139,-0.696,0.704,0.0,-0.833,0.553,-0.163,-0.817,0.553,-0.139,-0.696,0.704,-0.139,-0.696,0.704,-0.163,-0.817,0.553,-0.272,-0.656,0.704,-0.163,-0.817,0.553,-0.319,-0.77,0.553,-0.272,-0.656,0.704,-0.272,-0.656,0.704,-0.319,-0.77,0.553,-0.395,-0.59,0.704,-0.319,-0.77,0.553,-0.463,-0.693,0.553,-0.395,-0.59,0.704,-0.395,-0.59,0.704,-0.463,-0.693,0.553,-0.502,-0.502,0.704,-0.463,-0.693,0.553,-0.589,-0.589,0.553,-0.502,-0.502,0.704,-0.502,-0.502,0.704,-0.589,-0.589,0.553,-0.59,-0.395,0.704,-0.589,-0.589,0.553,-0.693,-0.463,0.553,-0.59,-0.395,0.704,-0.59,-0.395,0.704,-0.693,-0.463,0.553,-0.656,-0.272,0.704,-0.693,-0.463,0.553,-0.77,-0.319,0.553,-0.656,-0.272,0.704,-0.656,-0.272,0.704,-0.77,-0.319,0.553,-0.696,-0.139,0.704,-0.77,-0.319,0.553,-0.817,-0.163,0.553,-0.696,-0.139,0.704,-0.696,-0.139,0.704,-0.817,-0.163,0.553,-0.71,-0.0,0.704,-0.817,-0.163,0.553,-0.833,-0.0,0.553,-0.71,-0.0,0.704,-0.71,-0.0,0.704,-0.833,-0.0,0.553,-0.696,0.139,0.704,-0.833,-0.0,0.553,-0.817,0.163,0.553,-0.696,0.139,0.704,-0.696,0.139,0.704,-0.817,0.163,0.553,-0.656,0.272,0.704,-0.817,0.163,0.553,-0.77,0.319,0.553,-0.656,0.272,0.704,-0.656,0.272,0.704,-0.77,0.319,0.553,-0.59,0.395,0.704,-0.77,0.319,0.553,-0.693,0.463,0.553,-0.59,0.395,0.704,-0.59,0.395,0.704,-0.693,0.463,0.553,-0.502,0.502,0.704,-0.693,0.463,0.553,-0.589,0.589,0.553,-0.502,0.502,0.704,-0.502,0.502,0.704,-0.589,0.589,0.553,-0.395,0.59,0.704,-0.589,0.589,0.553,-0.463,0.693,0.553,-0.395,0.59,0.704,-0.395,0.59,0.704,-0.463,0.693,0.553,-0.272,0.656,0.704,-0.463,0.693,0.553,-0.319,0.77,0.553,-0.272,0.656,0.704,-0.272,0.656,0.704,-0.319,0.77,0.553,-0.139,0.696,0.704,-0.319,0.77,0.553,-0.163,0.817,0.553,-0.139,0.696,0.704,-0.139,0.696,0.704,-0.163,0.817,0.553,-0.0,0.71,0.704,-0.163,0.817,0.553,-0.0,0.833,0.553,-0.0,0.71,0.704,-0.0,0.56,0.829,-0.0,0.71,0.704,0.109,0.549,0.829,-0.0,0.71,0.704,0.139,0.696,0.704,0.109,0.549,0.829,0.109,0.549,0.829,0.139,0.696,0.704,0.214,0.517,0.829,0.139,0.696,0.704,0.272,0.656,0.704,0.214,0.517,0.829,0.214,0.517,0.829,0.272,0.656,0.704,0.311,0.465,0.829,0.272,0.656,0.704,0.395,0.59,0.704,0.311,0.465,0.829,0.311,0.465,0.829,0.395,0.59,0.704,0.396,0.396,0.829,0.395,0.59,0.704,0.502,0.502,0.704,0.396,0.396,0.829,0.396,0.396,0.829,0.502,0.502,0.704,0.465,0.311,0.829,0.502,0.502,0.704,0.59,0.395,0.704,0.465,0.311,0.829,0.465,0.311,0.829,0.59,0.395,0.704,0.517,0.214,0.829,0.59,0.395,0.704,0.656,0.272,0.704,0.517,0.214,0.829,0.517,0.214,0.829,0.656,0.272,0.704,0.549,0.109,0.829,0.656,0.272,0.704,0.696,0.139,0.704,0.549,0.109,0.829,0.549,0.109,0.829,0.696,0.139,0.704,0.56,0.0,0.829,0.696,0.139,0.704,0.71,0.0,0.704,0.56,0.0,0.829,0.56,0.0,0.829,0.71,0.0,0.704,0.549,-0.109,0.829,0.71,0.0,0.704,0.696,-0.139,0.704,0.549,-0.109,0.829,0.549,-0.109,0.829,0.696,-0.139,0.704,0.517,-0.214,0.829,0.696,-0.139,0.704,0.656,-0.272,0.704,0.517,-0.214,0.829,0.517,-0.214,0.829,0.656,-0.272,0.704,0.465,-0.311,0.829,0.656,-0.272,0.704,0.59,-0.395,0.704,0.465,-0.311,0.829,0.465,-0.311,0.829,0.59,-0.395,0.704,0.396,-0.396,0.829,0.59,-0.395,0.704,0.502,-0.502,0.704,0.396,-0.396,0.829,0.396,-0.396,0.829,0.502,-0.502,0.704,0.311,-0.465,0.829,0.502,-0.502,0.704,0.395,-0.59,0.704,0.311,-0.465,0.829,0.311,-0.465,0.829,0.395,-0.59,0.704,0.214,-0.517,0.829,0.395,-0.59,0.704,0.272,-0.656,0.704,0.214,-0.517,0.829,0.214,-0.517,0.829,0.272,-0.656,0.704,0.109,-0.549,0.829,0.272,-0.656,0.704,0.139,-0.696,0.704,0.109,-0.549,0.829,0.109,-0.549,0.829,0.139,-0.696,0.704,0.0,-0.56,0.829,0.139,-0.696,0.704,0.0,-0.71,0.704,0.0,-0.56,0.829,0.0,-0.56,0.829,0.0,-0.71,0.704,-0.109,-0.549,0.829,0.0,-0.71,0.704,-0.139,-0.696,0.704,-0.109,-0.549,0.829,-0.109,-0.549,0.829,-0.139,-0.696,0.704,-0.214,-0.517,0.829,-0.139,-0.696,0.704,-0.272,-0.656,0.704,-0.214,-0.517,0.829,-0.214,-0.517,0.829,-0.272,-0.656,0.704,-0.311,-0.465,0.829,-0.272,-0.656,0.704,-0.395,-0.59,0.704,-0.311,-0.465,0.829,-0.311,-0.465,0.829,-0.395,-0.59,0.704,-0.396,-0.396,0.829,-0.395,-0.59,0.704,-0.502,-0.502,0.704,-0.396,-0.396,0.829,-0.396,-0.396,0.829,-0.502,-0.502,0.704,-0.465,-0.311,0.829,-0.502,-0.502,0.704,-0.59,-0.395,0.704,-0.465,-0.311,0.829,-0.465,-0.311,0.829,-0.59,-0.395,0.704,-0.517,-0.214,0.829,-0.59,-0.395,0.704,-0.656,-0.272,0.704,-0.517,-0.214,0.829,-0.517,-0.214,0.829,-0.656,-0.272,0.704,-0.549,-0.109,0.829,-0.656,-0.272,0.704,-0.696,-0.139,0.704,-0.549,-0.109,0.829,-0.549,-0.109,0.829,-0.696,-0.139,0.704,-0.56,-0.0,0.829,-0.696,-0.139,0.704,-0.71,-0.0,0.704,-0.56,-0.0,0.829,-0.56,-0.0,0.829,-0.71,-0.0,0.704,-0.549,0.109,0.829,-0.71,-0.0,0.704,-0.696,0.139,0.704,-0.549,0.109,0.829,-0.549,0.109,0.829,-0.696,0.139,0.704,-0.517,0.214,0.829,-0.696,0.139,0.704,-0.656,0.272,0.704,-0.517,0.214,0.829,-0.517,0.214,0.829,-0.656,0.272,0.704,-0.465,0.311,0.829,-0.656,0.272,0.704,-0.59,0.395,0.704,-0.465,0.311,0.829,-0.465,0.311,0.829,-0.59,0.395,0.704,-0.396,0.396,0.829,-0.59,0.395,0.704,-0.502,0.502,0.704,-0.396,0.396,0.829,-0.396,0.396,0.829,-0.502,0.502,0.704,-0.311,0.465,0.829,-0.502,0.502,0.704,-0.395,0.59,0.704,-0.311,0.465,0.829,-0.311,0.465,0.829,-0.395,0.59,0.704,-0.214,0.517,0.829,-0.395,0.59,0.704,-0.272,0.656,0.704,-0.214,0.517,0.829,-0.214,0.517,0.829,-0.272,0.656,0.704,-0.109,0.549,0.829,-0.272,0.656,0.704,-0.139,0.696,0.704,-0.109,0.549,0.829,-0.109,0.549,0.829,-0.139,0.696,0.704,-0.0,0.56,0.829,-0.139,0.696,0.704,-0.0,0.71,0.704,-0.0,0.56,0.829,-0.0,0.388,0.922,-0.0,0.56,0.829,0.076,0.38,0.922,-0.0,0.56,0.829,0.109,0.549,0.829,0.076,0.38,0.922,0.076,0.38,0.922,0.109,0.549,0.829,0.148,0.358,0.922,0.109,0.549,0.829,0.214,0.517,0.829,0.148,0.358,0.922,0.148,0.358,0.922,0.214,0.517,0.829,0.215,0.323,0.922,0.214,0.517,0.829,0.311,0.465,0.829,0.215,0.323,0.922,0.215,0.323,0.922,0.311,0.465,0.829,0.274,0.274,0.922,0.311,0.465,0.829,0.396,0.396,0.829,0.274,0.274,0.922,0.274,0.274,0.922,0.396,0.396,0.829,0.323,0.215,0.922,0.396,0.396,0.829,0.465,0.311,0.829,0.323,0.215,0.922,0.323,0.215,0.922,0.465,0.311,0.829,0.358,0.148,0.922,0.465,0.311,0.829,0.517,0.214,0.829,0.358,0.148,0.922,0.358,0.148,0.922,0.517,0.214,0.829,0.38,0.076,0.922,0.517,0.214,0.829,0.549,0.109,0.829,0.38,0.076,0.922,0.38,0.076,0.922,0.549,0.109,0.829,0.388,0.0,0.922,0.549,0.109,0.829,0.56,0.0,0.829,0.388,0.0,0.922,0.388,0.0,0.922,0.56,0.0,0.829,0.38,-0.076,0.922,0.56,0.0,0.829,0.549,-0.109,0.829,0.38,-0.076,0.922,0.38,-0.076,0.922,0.549,-0.109,0.829,0.358,-0.148,0.922,0.549,-0.109,0.829,0.517,-0.214,0.829,0.358,-0.148,0.922,0.358,-0.148,0.922,0.517,-0.214,0.829,0.323,-0.215,0.922,0.517,-0.214,0.829,0.465,-0.311,0.829,0.323,-0.215,0.922,0.323,-0.215,0.922,0.465,-0.311,0.829,0.274,-0.274,0.922,0.465,-0.311,0.829,0.396,-0.396,0.829,0.274,-0.274,0.922,0.274,-0.274,0.922,0.396,-0.396,0.829,0.215,-0.323,0.922,0.396,-0.396,0.829,0.311,-0.465,0.829,0.215,-0.323,0.922,0.215,-0.323,0.922,0.311,-0.465,0.829,0.148,-0.358,0.922,0.311,-0.465,0.829,0.214,-0.517,0.829,0.148,-0.358,0.922,0.148,-0.358,0.922,0.214,-0.517,0.829,0.076,-0.38,0.922,0.214,-0.517,0.829,0.109,-0.549,0.829,0.076,-0.38,0.922,0.076,-0.38,0.922,0.109,-0.549,0.829,0.0,-0.388,0.922,0.109,-0.549,0.829,0.0,-0.56,0.829,0.0,-0.388,0.922,0.0,-0.388,0.922,0.0,-0.56,0.829,-0.076,-0.38,0.922,0.0,-0.56,0.829,-0.109,-0.549,0.829,-0.076,-0.38,0.922,-0.076,-0.38,0.922,-0.109,-0.549,0.829,-0.148,-0.358,0.922,-0.109,-0.549,0.829,-0.214,-0.517,0.829,-0.148,-0.358,0.922,-0.148,-0.358,0.922,-0.214,-0.517,0.829,-0.215,-0.323,0.922,-0.214,-0.517,0.829,-0.311,-0.465,0.829,-0.215,-0.323,0.922,-0.215,-0.323,0.922,-0.311,-0.465,0.829,-0.274,-0.274,0.922,-0.311,-0.465,0.829,-0.396,-0.396,0.829,-0.274,-0.274,0.922,-0.274,-0.274,0.922,-0.396,-0.396,0.829,-0.323,-0.215,0.922,-0.396,-0.396,0.829,-0.465,-0.311,0.829,-0.323,-0.215,0.922,-0.323,-0.215,0.922,-0.465,-0.311,0.829,-0.358,-0.148,0.922,-0.465,-0.311,0.829,-0.517,-0.214,0.829,-0.358,-0.148,0.922,-0.358,-0.148,0.922,-0.517,-0.214,0.829,-0.38,-0.076,0.922,-0.517,-0.214,0.829,-0.549,-0.109,0.829,-0.38,-0.076,0.922,-0.38,-0.076,0.922,-0.549,-0.109,0.829,-0.388,-0.0,0.922,-0.549,-0.109,0.829,-0.56,-0.0,0.829,-0.388,-0.0,0.922,-0.388,-0.0,0.922,-0.56,-0.0,0.829,-0.38,0.076,0.922,-0.56,-0.0,0.829,-0.549,0.109,0.829,-0.38,0.076,0.922,-0.38,0.076,0.922,-0.549,0.109,0.829,-0.358,0.148,0.922,-0.549,0.109,0.829,-0.517,0.214,0.829,-0.358,0.148,0.922,-0.358,0.148,0.922,-0.517,0.214,0.829,-0.323,0.215,0.922,-0.517,0.214,0.829,-0.465,0.311,0.829,-0.323,0.215,0.922,-0.323,0.215,0.922,-0.465,0.311,0.829,-0.274,0.274,0.922,-0.465,0.311,0.829,-0.396,0.396,0.829,-0.274,0.274,0.922,-0.274,0.274,0.922,-0.396,0.396,0.829,-0.215,0.323,0.922,-0.396,0.396,0.829,-0.311,0.465,0.829,-0.215,0.323,0.922,-0.215,0.323,0.922,-0.311,0.465,0.829,-0.148,0.358,0.922,-0.311,0.465,0.829,-0.214,0.517,0.829,-0.148,0.358,0.922,-0.148,0.358,0.922,-0.214,0.517,0.829,-0.076,0.38,0.922,-0.214,0.517,0.829,-0.109,0.549,0.829,-0.076,0.38,0.922,-0.076,0.38,0.922,-0.109,0.549,0.829,-0.0,0.388,0.922,-0.109,0.549,0.829,-0.0,0.56,0.829,-0.0,0.388,0.922,-0.0,0.201,0.98,-0.0,0.388,0.922,0.039,0.197,0.98,-0.0,0.388,0.922,0.076,0.38,0.922,0.039,0.197,0.98,0.039,0.197,0.98,0.076,0.38,0.922,0.077,0.186,0.98,0.076,0.38,0.922,0.148,0.358,0.922,0.077,0.186,0.98,0.077,0.186,0.98,0.148,0.358,0.922,0.112,0.167,0.98,0.148,0.358,0.922,0.215,0.323,0.922,0.112,0.167,0.98,0.112,0.167,0.98,0.215,0.323,0.922,0.142,0.142,0.98,0.215,0.323,0.922,0.274,0.274,0.922,0.142,0.142,0.98,0.142,0.142,0.98,0.274,0.274,0.922,0.167,0.112,0.98,0.274,0.274,0.922,0.323,0.215,0.922,0.167,0.112,0.98,0.167,0.112,0.98,0.323,0.215,0.922,0.186,0.077,0.98,0.323,0.215,0.922,0.358,0.148,0.922,0.186,0.077,0.98,0.186,0.077,0.98,0.358,0.148,0.922,0.197,0.039,0.98,0.358,0.148,0.922,0.38,0.076,0.922,0.197,0.039,0.98,0.197,0.039,0.98,0.38,0.076,0.922,0.201,0.0,0.98,0.38,0.076,0.922,0.388,0.0,0.922,0.201,0.0,0.98,0.201,0.0,0.98,0.388,0.0,0.922,0.197,-0.039,0.98,0.388,0.0,0.922,0.38,-0.076,0.922,0.197,-0.039,0.98,0.197,-0.039,0.98,0.38,-0.076,0.922,0.186,-0.077,0.98,0.38,-0.076,0.922,0.358,-0.148,0.922,0.186,-0.077,0.98,0.186,-0.077,0.98,0.358,-0.148,0.922,0.167,-0.112,0.98,0.358,-0.148,0.922,0.323,-0.215,0.922,0.167,-0.112,0.98,0.167,-0.112,0.98,0.323,-0.215,0.922,0.142,-0.142,0.98,0.323,-0.215,0.922,0.274,-0.274,0.922,0.142,-0.142,0.98,0.142,-0.142,0.98,0.274,-0.274,0.922,0.112,-0.167,0.98,0.274,-0.274,0.922,0.215,-0.323,0.922,0.112,-0.167,0.98,0.112,-0.167,0.98,0.215,-0.323,0.922,0.077,-0.186,0.98,0.215,-0.323,0.922,0.148,-0.358,0.922,0.077,-0.186,0.98,0.077,-0.186,0.98,0.148,-0.358,0.922,0.039,-0.197,0.98,0.148,-0.358,0.922,0.076,-0.38,0.922,0.039,-0.197,0.98,0.039,-0.197,0.98,0.076,-0.38,0.922,0.0,-0.201,0.98,0.076,-0.38,0.922,0.0,-0.388,0.922,0.0,-0.201,0.98,0.0,-0.201,0.98,0.0,-0.388,0.922,-0.039,-0.197,0.98,0.0,-0.388,0.922,-0.076,-0.38,0.922,-0.039,-0.197,0.98,-0.039,-0.197,0.98,-0.076,-0.38,0.922,-0.077,-0.186,0.98,-0.076,-0.38,0.922,-0.148,-0.358,0.922,-0.077,-0.186,0.98,-0.077,-0.186,0.98,-0.148,-0.358,0.922,-0.112,-0.167,0.98,-0.148,-0.358,0.922,-0.215,-0.323,0.922,-0.112,-0.167,0.98,-0.112,-0.167,0.98,-0.215,-0.323,0.922,-0.142,-0.142,0.98,-0.215,-0.323,0.922,-0.274,-0.274,0.922,-0.142,-0.142,0.98,-0.142,-0.142,0.98,-0.274,-0.274,0.922,-0.167,-0.112,0.98,-0.274,-0.274,0.922,-0.323,-0.215,0.922,-0.167,-0.112,0.98,-0.167,-0.112,0.98,-0.323,-0.215,0.922,-0.186,-0.077,0.98,-0.323,-0.215,0.922,-0.358,-0.148,0.922,-0.186,-0.077,0.98,-0.186,-0.077,0.98,-0.358,-0.148,0.922,-0.197,-0.039,0.98,-0.358,-0.148,0.922,-0.38,-0.076,0.922,-0.197,-0.039,0.98,-0.197,-0.039,0.98,-0.38,-0.076,0.922,-0.201,-0.0,0.98,-0.38,-0.076,0.922,-0.388,-0.0,0.922,-0.201,-0.0,0.98,-0.201,-0.0,0.98,-0.388,-0.0,0.922,-0.197,0.039,0.98,-0.388,-0.0,0.922,-0.38,0.076,0.922,-0.197,0.039,0.98,-0.197,0.039,0.98,-0.38,0.076,0.922,-0.186,0.077,0.98,-0.38,0.076,0.922,-0.358,0.148,0.922,-0.186,0.077,0.98,-0.186,0.077,0.98,-0.358,0.148,0.922,-0.167,0.112,0.98,-0.358,0.148,0.922,-0.323,0.215,0.922,-0.167,0.112,0.98,-0.167,0.112,0.98,-0.323,0.215,0.922,-0.142,0.142,0.98,-0.323,0.215,0.922,-0.274,0.274,0.922,-0.142,0.142,0.98,-0.142,0.142,0.98,-0.274,0.274,0.922,-0.112,0.167,0.98,-0.274,0.274,0.922,-0.215,0.323,0.922,-0.112,0.167,0.98,-0.112,0.167,0.98,-0.215,0.323,0.922,-0.077,0.186,0.98,-0.215,0.323,0.922,-0.148,0.358,0.922,-0.077,0.186,0.98,-0.077,0.186,0.98,-0.148,0.358,0.922,-0.039,0.197,0.98,-0.148,0.358,0.922,-0.076,0.38,0.922,-0.039,0.197,0.98,-0.039,0.197,0.98,-0.076,0.38,0.922,-0.0,0.201,0.98,-0.076,0.38,0.922,-0.0,0.388,0.922,-0.0,0.201,0.98,-0.0,0.201,0.98,0.039,0.197,0.98,0.0,0.0,1.0,0.039,0.197,0.98,0.077,0.186,0.98,0.0,0.0,1.0,0.077,0.186,0.98,0.112,0.167,0.98,0.0,0.0,1.0,0.112,0.167,0.98,0.142,0.142,0.98,0.0,0.0,1.0,0.142,0.142,0.98,0.167,0.112,0.98,0.0,0.0,1.0,0.167,0.112,0.98,0.186,0.077,0.98,0.0,0.0,1.0,0.186,0.077,0.98,0.197,0.039,0.98,0.0,0.0,1.0,0.197,0.039,0.98,0.201,0.0,0.98,0.0,0.0,1.0,0.201,0.0,0.98,0.197,-0.039,0.98,0.0,0.0,1.0,0.197,-0.039,0.98,0.186,-0.077,0.98,0.0,0.0,1.0,0.186,-0.077,0.98,0.167,-0.112,0.98,0.0,0.0,1.0,0.167,-0.112,0.98,0.142,-0.142,0.98,0.0,0.0,1.0,0.142,-0.142,0.98,0.112,-0.167,0.98,0.0,0.0,1.0,0.112,-0.167,0.98,0.077,-0.186,0.98,0.0,0.0,1.0,0.077,-0.186,0.98,0.039,-0.197,0.98,0.0,0.0,1.0,0.039,-0.197,0.98,0.0,-0.201,0.98,0.0,0.0,1.0,0.0,-0.201,0.98,-0.039,-0.197,0.98,0.0,0.0,1.0,-0.039,-0.197,0.98,-0.077,-0.186,0.98,0.0,0.0,1.0,-0.077,-0.186,0.98,-0.112,-0.167,0.98,0.0,0.0,1.0,-0.112,-0.167,0.98,-0.142,-0.142,0.98,0.0,0.0,1.0,-0.142,-0.142,0.98,-0.167,-0.112,0.98,0.0,0.0,1.0,-0.167,-0.112,0.98,-0.186,-0.077,0.98,0.0,0.0,1.0,-0.186,-0.077,0.98,-0.197,-0.039,0.98,0.0,0.0,1.0,-0.197,-0.039,0.98,-0.201,-0.0,0.98,0.0,0.0,1.0,-0.201,-0.0,0.98,-0.197,0.039,0.98,0.0,0.0,1.0,-0.197,0.039,0.98,-0.186,0.077,0.98,0.0,0.0,1.0,-0.186,0.077,0.98,-0.167,0.112,0.98,0.0,0.0,1.0,-0.167,0.112,0.98,-0.142,0.142,0.98,0.0,0.0,1.0,-0.142,0.142,0.98,-0.112,0.167,0.98,0.0,0.0,1.0,-0.112,0.167,0.98,-0.077,0.186,0.98,0.0,0.0,1.0,-0.077,0.186,0.98,-0.039,0.197,0.98,0.0,0.0,1.0,-0.039,0.197,0.98,-0.0,0.201,0.98,0.0,0.0,1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,1.0,0.0,-0.0,1.0,-0.0,0.0,0.94,-0.342,0.0,0.94,-0.342,0.0,0.94,-0.342,-0.0,1.0,0.0,-0.0,0.94,-0.342,-0.0,0.94,-0.342,0.0,0.766,-0.643,0.0,0.766,-0.643,0.0,0.766,-0.643,0.0,0.94,-0.342,-0.0,0.766,-0.643,0.0,0.766,-0.643,0.0,0.5,-0.866,0.0,0.5,-0.866,0.0,0.5,-0.866,0.0,0.766,-0.643,0.0,0.5,-0.866,0.0,0.5,-0.866,0.0,0.174,-0.985,0.0,0.174,-0.985,0.0,0.174,-0.985,0.0,0.5,-0.866,0.0,0.174,-0.985,0.0,0.174,-0.985,0.0,-0.174,-0.985,0.0,-0.174,-0.985,0.0,-0.174,-0.985,0.0,0.174,-0.985,0.0,-0.174,-0.985,0.0,-0.174,-0.985,0.0,-0.5,-0.866,0.0,-0.5,-0.866,0.0,-0.5,-0.866,0.0,-0.174,-0.985,0.0,-0.5,-0.866,0.0,-0.5,-0.866,0.0,-0.766,-0.643,0.0,-0.766,-0.643,0.0,-0.766,-0.643,0.0,-0.5,-0.866,0.0,-0.766,-0.643,0.0,-0.766,-0.643,0.0,-0.94,-0.342,0.0,-0.94,-0.342,0.0,-0.94,-0.342,0.0,-0.766,-0.643,0.0,-0.94,-0.342,0.0,-0.94,-0.342,0.0,-1.0,-0.0,-0.0,-1.0,-0.0,-0.0,-1.0,0.0,0.0,-0.94,-0.342,0.0,-1.0,0.0,0.0,-1.0,-0.0,-0.0,-0.94,0.342,-0.0,-0.94,0.342,-0.0,-0.94,0.342,0.0,-1.0,0.0,0.0,-0.94,0.342,0.0,-0.94,0.342,-0.0,-0.766,0.643,-0.0,-0.766,0.643,-0.0,-0.766,0.643,0.0,-0.94,0.342,0.0,-0.766,0.643,0.0,-0.766,0.643,-0.0,-0.5,0.866,0.0,-0.5,0.866,0.0,-0.5,0.866,0.0,-0.766,0.643,0.0,-0.5,0.866,0.0,-0.5,0.866,0.0,-0.174,0.985,0.0,-0.174,0.985,0.0,-0.174,0.985,0.0,-0.5,0.866,0.0,-0.174,0.985,0.0,-0.174,0.985,0.0,0.174,0.985,0.0,0.174,0.985,0.0,0.174,0.985,0.0,-0.174,0.985,0.0,0.174,0.985,0.0,0.174,0.985,0.0,0.5,0.866,0.0,0.5,0.866,0.0,0.5,0.866,0.0,0.174,0.985,0.0,0.5,0.866,0.0,0.5,0.866,0.0,0.766,0.643,0.0,0.766,0.643,0.0,0.766,0.643,-0.0,0.5,0.866,0.0,0.766,0.643,-0.0,0.766,0.643,0.0,0.94,0.342,0.0,0.94,0.342,0.0,0.94,0.342,-0.0,0.766,0.643,-0.0,0.94,0.342,-0.0,0.94,0.342,0.0,1.0,-0.0,0.0,1.0,-0.0,0.0,1.0,0.0,-0.0,0.94,0.342,-0.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-1.0,1.0,-0.0,0.0,0.94,-0.342,0.0,0.94,-0.342,0.0,0.94,-0.342,0.0,1.0,-0.0,0.0,1.0,-0.0,0.0,0.94,-0.342,0.0,0.766,-0.643,0.0,0.766,-0.643,0.0,0.766,-0.643,0.0,0.94,-0.342,0.0,0.94,-0.342,0.0,0.766,-0.643,0.0,0.5,-0.866,0.0,0.5,-0.866,0.0,0.5,-0.866,0.0,0.766,-0.643,0.0,0.766,-0.643,0.0,0.5,-0.866,0.0,0.174,-0.985,0.0,0.174,-0.985,0.0,0.174,-0.985,0.0,0.5,-0.866,0.0,0.5,-0.866,0.0,0.174,-0.985,0.0,-0.174,-0.985,0.0,-0.174,-0.985,0.0,-0.174,-0.985,0.0,0.174,-0.985,0.0,0.174,-0.985,0.0,-0.174,-0.985,0.0,-0.5,-0.866,0.0,-0.5,-0.866,0.0,-0.5,-0.866,0.0,-0.174,-0.985,0.0,-0.174,-0.985,0.0,-0.5,-0.866,0.0,-0.766,-0.643,0.0,-0.766,-0.643,0.0,-0.766,-0.643,0.0,-0.5,-0.866,0.0,-0.5,-0.866,0.0,-0.766,-0.643,0.0,-0.94,-0.342,-0.0,-0.94,-0.342,0.0,-0.94,-0.342,0.0,-0.766,-0.643,0.0,-0.766,-0.643,0.0,-0.94,-0.342,-0.0,-1.0,0.0,-0.0,-1.0,0.0,0.0,-1.0,0.0,0.0,-0.94,-0.342,0.0,-0.94,-0.342,-0.0,-1.0,0.0,-0.0,-0.94,0.342,0.0,-0.94,0.342,0.0,-0.94,0.342,0.0,-1.0,0.0,0.0,-1.0,0.0,-0.0,-0.94,0.342,0.0,-0.766,0.643,0.0,-0.766,0.643,0.0,-0.766,0.643,0.0,-0.94,0.342,0.0,-0.94,0.342,0.0,-0.766,0.643,0.0,-0.5,0.866,0.0,-0.5,0.866,0.0,-0.5,0.866,0.0,-0.766,0.643,0.0,-0.766,0.643,0.0,-0.5,0.866,0.0,-0.174,0.985,0.0,-0.174,0.985,0.0,-0.174,0.985,0.0,-0.5,0.866,0.0,-0.5,0.866,0.0,-0.174,0.985,0.0,0.174,0.985,0.0,0.174,0.985,0.0,0.174,0.985,0.0,-0.174,0.985,0.0,-0.174,0.985,0.0,0.174,0.985,0.0,0.5,0.866,0.0,0.5,0.866,0.0,0.5,0.866,0.0,0.174,0.985,0.0,0.174,0.985,0.0,0.5,0.866,0.0,0.766,0.643,0.0,0.766,0.643,0.0,0.766,0.643,0.0,0.5,0.866,0.0,0.5,0.866,0.0,0.766,0.643,0.0,0.94,0.342,0.0,0.94,0.342,0.0,0.94,0.342,0.0,0.766,0.643,0.0,0.766,0.643,0.0,0.94,0.342,0.0,1.0,-0.0,0.0,1.0,-0.0,0.0,1.0,-0.0,0.0,0.94,0.342,0.0,0.94,0.342,0.0]},\\"uv0\\":{\\"valueType\\":\\"Float32\\",\\"valuesPerElement\\":2,\\"values\\":[0.9688,0.9375,1.0,0.9375,0.9688,1.0,0.9375,0.9375,0.9688,0.9375,0.9375,1.0,0.9063,0.9375,0.9375,0.9375,0.9063,1.0,0.875,0.9375,0.9063,0.9375,0.875,1.0,0.8438,0.9375,0.875,0.9375,0.8438,1.0,0.8125,0.9375,0.8438,0.9375,0.8125,1.0,0.7813,0.9375,0.8125,0.9375,0.7813,1.0,0.75,0.9375,0.7813,0.9375,0.75,1.0,0.7188,0.9375,0.75,0.9375,0.7188,1.0,0.6875,0.9375,0.7188,0.9375,0.6875,1.0,0.6563,0.9375,0.6875,0.9375,0.6563,1.0,0.625,0.9375,0.6563,0.9375,0.625,1.0,0.5938,0.9375,0.625,0.9375,0.5938,1.0,0.5625,0.9375,0.5938,0.9375,0.5625,1.0,0.5313,0.9375,0.5625,0.9375,0.5313,1.0,0.5,0.9375,0.5313,0.9375,0.5,1.0,0.4688,0.9375,0.5,0.9375,0.4688,1.0,0.4375,0.9375,0.4688,0.9375,0.4375,1.0,0.4063,0.9375,0.4375,0.9375,0.4063,1.0,0.375,0.9375,0.4063,0.9375,0.375,1.0,0.3438,0.9375,0.375,0.9375,0.3438,1.0,0.3125,0.9375,0.3438,0.9375,0.3125,1.0,0.2813,0.9375,0.3125,0.9375,0.2813,1.0,0.25,0.9375,0.2813,0.9375,0.25,1.0,0.2188,0.9375,0.25,0.9375,0.2188,1.0,0.1875,0.9375,0.2188,0.9375,0.1875,1.0,0.1563,0.9375,0.1875,0.9375,0.1563,1.0,0.125,0.9375,0.1563,0.9375,0.125,1.0,0.0938,0.9375,0.125,0.9375,0.0938,1.0,0.0625,0.9375,0.0938,0.9375,0.0625,1.0,0.0313,0.9375,0.0625,0.9375,0.0313,1.0,0.0,0.9375,0.0313,0.9375,0.0,1.0,1.0,0.875,1.0,0.9375,0.9688,0.875,1.0,0.9375,0.9688,0.9375,0.9688,0.875,0.9688,0.875,0.9688,0.9375,0.9375,0.875,0.9688,0.9375,0.9375,0.9375,0.9375,0.875,0.9375,0.875,0.9375,0.9375,0.9063,0.875,0.9375,0.9375,0.9063,0.9375,0.9063,0.875,0.9063,0.875,0.9063,0.9375,0.875,0.875,0.9063,0.9375,0.875,0.9375,0.875,0.875,0.875,0.875,0.875,0.9375,0.8438,0.875,0.875,0.9375,0.8438,0.9375,0.8438,0.875,0.8438,0.875,0.8438,0.9375,0.8125,0.875,0.8438,0.9375,0.8125,0.9375,0.8125,0.875,0.8125,0.875,0.8125,0.9375,0.7813,0.875,0.8125,0.9375,0.7813,0.9375,0.7813,0.875,0.7813,0.875,0.7813,0.9375,0.75,0.875,0.7813,0.9375,0.75,0.9375,0.75,0.875,0.75,0.875,0.75,0.9375,0.7188,0.875,0.75,0.9375,0.7188,0.9375,0.7188,0.875,0.7188,0.875,0.7188,0.9375,0.6875,0.875,0.7188,0.9375,0.6875,0.9375,0.6875,0.875,0.6875,0.875,0.6875,0.9375,0.6563,0.875,0.6875,0.9375,0.6563,0.9375,0.6563,0.875,0.6563,0.875,0.6563,0.9375,0.625,0.875,0.6563,0.9375,0.625,0.9375,0.625,0.875,0.625,0.875,0.625,0.9375,0.5938,0.875,0.625,0.9375,0.5938,0.9375,0.5938,0.875,0.5938,0.875,0.5938,0.9375,0.5625,0.875,0.5938,0.9375,0.5625,0.9375,0.5625,0.875,0.5625,0.875,0.5625,0.9375,0.5313,0.875,0.5625,0.9375,0.5313,0.9375,0.5313,0.875,0.5313,0.875,0.5313,0.9375,0.5,0.875,0.5313,0.9375,0.5,0.9375,0.5,0.875,0.5,0.875,0.5,0.9375,0.4688,0.875,0.5,0.9375,0.4688,0.9375,0.4688,0.875,0.4688,0.875,0.4688,0.9375,0.4375,0.875,0.4688,0.9375,0.4375,0.9375,0.4375,0.875,0.4375,0.875,0.4375,0.9375,0.4063,0.875,0.4375,0.9375,0.4063,0.9375,0.4063,0.875,0.4063,0.875,0.4063,0.9375,0.375,0.875,0.4063,0.9375,0.375,0.9375,0.375,0.875,0.375,0.875,0.375,0.9375,0.3438,0.875,0.375,0.9375,0.3438,0.9375,0.3438,0.875,0.3438,0.875,0.3438,0.9375,0.3125,0.875,0.3438,0.9375,0.3125,0.9375,0.3125,0.875,0.3125,0.875,0.3125,0.9375,0.2813,0.875,0.3125,0.9375,0.2813,0.9375,0.2813,0.875,0.2813,0.875,0.2813,0.9375,0.25,0.875,0.2813,0.9375,0.25,0.9375,0.25,0.875,0.25,0.875,0.25,0.9375,0.2188,0.875,0.25,0.9375,0.2188,0.9375,0.2188,0.875,0.2188,0.875,0.2188,0.9375,0.1875,0.875,0.2188,0.9375,0.1875,0.9375,0.1875,0.875,0.1875,0.875,0.1875,0.9375,0.1563,0.875,0.1875,0.9375,0.1563,0.9375,0.1563,0.875,0.1563,0.875,0.1563,0.9375,0.125,0.875,0.1563,0.9375,0.125,0.9375,0.125,0.875,0.125,0.875,0.125,0.9375,0.0938,0.875,0.125,0.9375,0.0938,0.9375,0.0938,0.875,0.0938,0.875,0.0938,0.9375,0.0625,0.875,0.0938,0.9375,0.0625,0.9375,0.0625,0.875,0.0625,0.875,0.0625,0.9375,0.0313,0.875,0.0625,0.9375,0.0313,0.9375,0.0313,0.875,0.0313,0.875,0.0313,0.9375,0.0,0.875,0.0313,0.9375,0.0,0.9375,0.0,0.875,1.0,0.8125,1.0,0.875,0.9688,0.8125,1.0,0.875,0.9688,0.875,0.9688,0.8125,0.9688,0.8125,0.9688,0.875,0.9375,0.8125,0.9688,0.875,0.9375,0.875,0.9375,0.8125,0.9375,0.8125,0.9375,0.875,0.9063,0.8125,0.9375,0.875,0.9063,0.875,0.9063,0.8125,0.9063,0.8125,0.9063,0.875,0.875,0.8125,0.9063,0.875,0.875,0.875,0.875,0.8125,0.875,0.8125,0.875,0.875,0.8438,0.8125,0.875,0.875,0.8438,0.875,0.8438,0.8125,0.8438,0.8125,0.8438,0.875,0.8125,0.8125,0.8438,0.875,0.8125,0.875,0.8125,0.8125,0.8125,0.8125,0.8125,0.875,0.7813,0.8125,0.8125,0.875,0.7813,0.875,0.7813,0.8125,0.7813,0.8125,0.7813,0.875,0.75,0.8125,0.7813,0.875,0.75,0.875,0.75,0.8125,0.75,0.8125,0.75,0.875,0.7188,0.8125,0.75,0.875,0.7188,0.875,0.7188,0.8125,0.7188,0.8125,0.7188,0.875,0.6875,0.8125,0.7188,0.875,0.6875,0.875,0.6875,0.8125,0.6875,0.8125,0.6875,0.875,0.6563,0.8125,0.6875,0.875,0.6563,0.875,0.6563,0.8125,0.6563,0.8125,0.6563,0.875,0.625,0.8125,0.6563,0.875,0.625,0.875,0.625,0.8125,0.625,0.8125,0.625,0.875,0.5938,0.8125,0.625,0.875,0.5938,0.875,0.5938,0.8125,0.5938,0.8125,0.5938,0.875,0.5625,0.8125,0.5938,0.875,0.5625,0.875,0.5625,0.8125,0.5625,0.8125,0.5625,0.875,0.5313,0.8125,0.5625,0.875,0.5313,0.875,0.5313,0.8125,0.5313,0.8125,0.5313,0.875,0.5,0.8125,0.5313,0.875,0.5,0.875,0.5,0.8125,0.5,0.8125,0.5,0.875,0.4688,0.8125,0.5,0.875,0.4688,0.875,0.4688,0.8125,0.4688,0.8125,0.4688,0.875,0.4375,0.8125,0.4688,0.875,0.4375,0.875,0.4375,0.8125,0.4375,0.8125,0.4375,0.875,0.4063,0.8125,0.4375,0.875,0.4063,0.875,0.4063,0.8125,0.4063,0.8125,0.4063,0.875,0.375,0.8125,0.4063,0.875,0.375,0.875,0.375,0.8125,0.375,0.8125,0.375,0.875,0.3438,0.8125,0.375,0.875,0.3438,0.875,0.3438,0.8125,0.3438,0.8125,0.3438,0.875,0.3125,0.8125,0.3438,0.875,0.3125,0.875,0.3125,0.8125,0.3125,0.8125,0.3125,0.875,0.2813,0.8125,0.3125,0.875,0.2813,0.875,0.2813,0.8125,0.2813,0.8125,0.2813,0.875,0.25,0.8125,0.2813,0.875,0.25,0.875,0.25,0.8125,0.25,0.8125,0.25,0.875,0.2188,0.8125,0.25,0.875,0.2188,0.875,0.2188,0.8125,0.2188,0.8125,0.2188,0.875,0.1875,0.8125,0.2188,0.875,0.1875,0.875,0.1875,0.8125,0.1875,0.8125,0.1875,0.875,0.1563,0.8125,0.1875,0.875,0.1563,0.875,0.1563,0.8125,0.1563,0.8125,0.1563,0.875,0.125,0.8125,0.1563,0.875,0.125,0.875,0.125,0.8125,0.125,0.8125,0.125,0.875,0.0938,0.8125,0.125,0.875,0.0938,0.875,0.0938,0.8125,0.0938,0.8125,0.0938,0.875,0.0625,0.8125,0.0938,0.875,0.0625,0.875,0.0625,0.8125,0.0625,0.8125,0.0625,0.875,0.0313,0.8125,0.0625,0.875,0.0313,0.875,0.0313,0.8125,0.0313,0.8125,0.0313,0.875,0.0,0.8125,0.0313,0.875,0.0,0.875,0.0,0.8125,1.0,0.75,1.0,0.8125,0.9688,0.75,1.0,0.8125,0.9688,0.8125,0.9688,0.75,0.9688,0.75,0.9688,0.8125,0.9375,0.75,0.9688,0.8125,0.9375,0.8125,0.9375,0.75,0.9375,0.75,0.9375,0.8125,0.9063,0.75,0.9375,0.8125,0.9063,0.8125,0.9063,0.75,0.9063,0.75,0.9063,0.8125,0.875,0.75,0.9063,0.8125,0.875,0.8125,0.875,0.75,0.875,0.75,0.875,0.8125,0.8438,0.75,0.875,0.8125,0.8438,0.8125,0.8438,0.75,0.8438,0.75,0.8438,0.8125,0.8125,0.75,0.8438,0.8125,0.8125,0.8125,0.8125,0.75,0.8125,0.75,0.8125,0.8125,0.7813,0.75,0.8125,0.8125,0.7813,0.8125,0.7813,0.75,0.7813,0.75,0.7813,0.8125,0.75,0.75,0.7813,0.8125,0.75,0.8125,0.75,0.75,0.75,0.75,0.75,0.8125,0.7188,0.75,0.75,0.8125,0.7188,0.8125,0.7188,0.75,0.7188,0.75,0.7188,0.8125,0.6875,0.75,0.7188,0.8125,0.6875,0.8125,0.6875,0.75,0.6875,0.75,0.6875,0.8125,0.6563,0.75,0.6875,0.8125,0.6563,0.8125,0.6563,0.75,0.6563,0.75,0.6563,0.8125,0.625,0.75,0.6563,0.8125,0.625,0.8125,0.625,0.75,0.625,0.75,0.625,0.8125,0.5938,0.75,0.625,0.8125,0.5938,0.8125,0.5938,0.75,0.5938,0.75,0.5938,0.8125,0.5625,0.75,0.5938,0.8125,0.5625,0.8125,0.5625,0.75,0.5625,0.75,0.5625,0.8125,0.5313,0.75,0.5625,0.8125,0.5313,0.8125,0.5313,0.75,0.5313,0.75,0.5313,0.8125,0.5,0.75,0.5313,0.8125,0.5,0.8125,0.5,0.75,0.5,0.75,0.5,0.8125,0.4688,0.75,0.5,0.8125,0.4688,0.8125,0.4688,0.75,0.4688,0.75,0.4688,0.8125,0.4375,0.75,0.4688,0.8125,0.4375,0.8125,0.4375,0.75,0.4375,0.75,0.4375,0.8125,0.4063,0.75,0.4375,0.8125,0.4063,0.8125,0.4063,0.75,0.4063,0.75,0.4063,0.8125,0.375,0.75,0.4063,0.8125,0.375,0.8125,0.375,0.75,0.375,0.75,0.375,0.8125,0.3438,0.75,0.375,0.8125,0.3438,0.8125,0.3438,0.75,0.3438,0.75,0.3438,0.8125,0.3125,0.75,0.3438,0.8125,0.3125,0.8125,0.3125,0.75,0.3125,0.75,0.3125,0.8125,0.2813,0.75,0.3125,0.8125,0.2813,0.8125,0.2813,0.75,0.2813,0.75,0.2813,0.8125,0.25,0.75,0.2813,0.8125,0.25,0.8125,0.25,0.75,0.25,0.75,0.25,0.8125,0.2188,0.75,0.25,0.8125,0.2188,0.8125,0.2188,0.75,0.2188,0.75,0.2188,0.8125,0.1875,0.75,0.2188,0.8125,0.1875,0.8125,0.1875,0.75,0.1875,0.75,0.1875,0.8125,0.1563,0.75,0.1875,0.8125,0.1563,0.8125,0.1563,0.75,0.1563,0.75,0.1563,0.8125,0.125,0.75,0.1563,0.8125,0.125,0.8125,0.125,0.75,0.125,0.75,0.125,0.8125,0.0938,0.75,0.125,0.8125,0.0938,0.8125,0.0938,0.75,0.0938,0.75,0.0938,0.8125,0.0625,0.75,0.0938,0.8125,0.0625,0.8125,0.0625,0.75,0.0625,0.75,0.0625,0.8125,0.0313,0.75,0.0625,0.8125,0.0313,0.8125,0.0313,0.75,0.0313,0.75,0.0313,0.8125,0.0,0.75,0.0313,0.8125,0.0,0.8125,0.0,0.75,1.0,0.6875,1.0,0.75,0.9688,0.6875,1.0,0.75,0.9688,0.75,0.9688,0.6875,0.9688,0.6875,0.9688,0.75,0.9375,0.6875,0.9688,0.75,0.9375,0.75,0.9375,0.6875,0.9375,0.6875,0.9375,0.75,0.9063,0.6875,0.9375,0.75,0.9063,0.75,0.9063,0.6875,0.9063,0.6875,0.9063,0.75,0.875,0.6875,0.9063,0.75,0.875,0.75,0.875,0.6875,0.875,0.6875,0.875,0.75,0.8438,0.6875,0.875,0.75,0.8438,0.75,0.8438,0.6875,0.8438,0.6875,0.8438,0.75,0.8125,0.6875,0.8438,0.75,0.8125,0.75,0.8125,0.6875,0.8125,0.6875,0.8125,0.75,0.7813,0.6875,0.8125,0.75,0.7813,0.75,0.7813,0.6875,0.7813,0.6875,0.7813,0.75,0.75,0.6875,0.7813,0.75,0.75,0.75,0.75,0.6875,0.75,0.6875,0.75,0.75,0.7188,0.6875,0.75,0.75,0.7188,0.75,0.7188,0.6875,0.7188,0.6875,0.7188,0.75,0.6875,0.6875,0.7188,0.75,0.6875,0.75,0.6875,0.6875,0.6875,0.6875,0.6875,0.75,0.6563,0.6875,0.6875,0.75,0.6563,0.75,0.6563,0.6875,0.6563,0.6875,0.6563,0.75,0.625,0.6875,0.6563,0.75,0.625,0.75,0.625,0.6875,0.625,0.6875,0.625,0.75,0.5938,0.6875,0.625,0.75,0.5938,0.75,0.5938,0.6875,0.5938,0.6875,0.5938,0.75,0.5625,0.6875,0.5938,0.75,0.5625,0.75,0.5625,0.6875,0.5625,0.6875,0.5625,0.75,0.5313,0.6875,0.5625,0.75,0.5313,0.75,0.5313,0.6875,0.5313,0.6875,0.5313,0.75,0.5,0.6875,0.5313,0.75,0.5,0.75,0.5,0.6875,0.5,0.6875,0.5,0.75,0.4688,0.6875,0.5,0.75,0.4688,0.75,0.4688,0.6875,0.4688,0.6875,0.4688,0.75,0.4375,0.6875,0.4688,0.75,0.4375,0.75,0.4375,0.6875,0.4375,0.6875,0.4375,0.75,0.4063,0.6875,0.4375,0.75,0.4063,0.75,0.4063,0.6875,0.4063,0.6875,0.4063,0.75,0.375,0.6875,0.4063,0.75,0.375,0.75,0.375,0.6875,0.375,0.6875,0.375,0.75,0.3438,0.6875,0.375,0.75,0.3438,0.75,0.3438,0.6875,0.3438,0.6875,0.3438,0.75,0.3125,0.6875,0.3438,0.75,0.3125,0.75,0.3125,0.6875,0.3125,0.6875,0.3125,0.75,0.2813,0.6875,0.3125,0.75,0.2813,0.75,0.2813,0.6875,0.2813,0.6875,0.2813,0.75,0.25,0.6875,0.2813,0.75,0.25,0.75,0.25,0.6875,0.25,0.6875,0.25,0.75,0.2188,0.6875,0.25,0.75,0.2188,0.75,0.2188,0.6875,0.2188,0.6875,0.2188,0.75,0.1875,0.6875,0.2188,0.75,0.1875,0.75,0.1875,0.6875,0.1875,0.6875,0.1875,0.75,0.1563,0.6875,0.1875,0.75,0.1563,0.75,0.1563,0.6875,0.1563,0.6875,0.1563,0.75,0.125,0.6875,0.1563,0.75,0.125,0.75,0.125,0.6875,0.125,0.6875,0.125,0.75,0.0938,0.6875,0.125,0.75,0.0938,0.75,0.0938,0.6875,0.0938,0.6875,0.0938,0.75,0.0625,0.6875,0.0938,0.75,0.0625,0.75,0.0625,0.6875,0.0625,0.6875,0.0625,0.75,0.0313,0.6875,0.0625,0.75,0.0313,0.75,0.0313,0.6875,0.0313,0.6875,0.0313,0.75,0.0,0.6875,0.0313,0.75,0.0,0.75,0.0,0.6875,1.0,0.625,1.0,0.6875,0.9688,0.625,1.0,0.6875,0.9688,0.6875,0.9688,0.625,0.9688,0.625,0.9688,0.6875,0.9375,0.625,0.9688,0.6875,0.9375,0.6875,0.9375,0.625,0.9375,0.625,0.9375,0.6875,0.9063,0.625,0.9375,0.6875,0.9063,0.6875,0.9063,0.625,0.9063,0.625,0.9063,0.6875,0.875,0.625,0.9063,0.6875,0.875,0.6875,0.875,0.625,0.875,0.625,0.875,0.6875,0.8438,0.625,0.875,0.6875,0.8438,0.6875,0.8438,0.625,0.8438,0.625,0.8438,0.6875,0.8125,0.625,0.8438,0.6875,0.8125,0.6875,0.8125,0.625,0.8125,0.625,0.8125,0.6875,0.7813,0.625,0.8125,0.6875,0.7813,0.6875,0.7813,0.625,0.7813,0.625,0.7813,0.6875,0.75,0.625,0.7813,0.6875,0.75,0.6875,0.75,0.625,0.75,0.625,0.75,0.6875,0.7188,0.625,0.75,0.6875,0.7188,0.6875,0.7188,0.625,0.7188,0.625,0.7188,0.6875,0.6875,0.625,0.7188,0.6875,0.6875,0.6875,0.6875,0.625,0.6875,0.625,0.6875,0.6875,0.6563,0.625,0.6875,0.6875,0.6563,0.6875,0.6563,0.625,0.6563,0.625,0.6563,0.6875,0.625,0.625,0.6563,0.6875,0.625,0.6875,0.625,0.625,0.625,0.625,0.625,0.6875,0.5938,0.625,0.625,0.6875,0.5938,0.6875,0.5938,0.625,0.5938,0.625,0.5938,0.6875,0.5625,0.625,0.5938,0.6875,0.5625,0.6875,0.5625,0.625,0.5625,0.625,0.5625,0.6875,0.5313,0.625,0.5625,0.6875,0.5313,0.6875,0.5313,0.625,0.5313,0.625,0.5313,0.6875,0.5,0.625,0.5313,0.6875,0.5,0.6875,0.5,0.625,0.5,0.625,0.5,0.6875,0.4688,0.625,0.5,0.6875,0.4688,0.6875,0.4688,0.625,0.4688,0.625,0.4688,0.6875,0.4375,0.625,0.4688,0.6875,0.4375,0.6875,0.4375,0.625,0.4375,0.625,0.4375,0.6875,0.4063,0.625,0.4375,0.6875,0.4063,0.6875,0.4063,0.625,0.4063,0.625,0.4063,0.6875,0.375,0.625,0.4063,0.6875,0.375,0.6875,0.375,0.625,0.375,0.625,0.375,0.6875,0.3438,0.625,0.375,0.6875,0.3438,0.6875,0.3438,0.625,0.3438,0.625,0.3438,0.6875,0.3125,0.625,0.3438,0.6875,0.3125,0.6875,0.3125,0.625,0.3125,0.625,0.3125,0.6875,0.2813,0.625,0.3125,0.6875,0.2813,0.6875,0.2813,0.625,0.2813,0.625,0.2813,0.6875,0.25,0.625,0.2813,0.6875,0.25,0.6875,0.25,0.625,0.25,0.625,0.25,0.6875,0.2188,0.625,0.25,0.6875,0.2188,0.6875,0.2188,0.625,0.2188,0.625,0.2188,0.6875,0.1875,0.625,0.2188,0.6875,0.1875,0.6875,0.1875,0.625,0.1875,0.625,0.1875,0.6875,0.1563,0.625,0.1875,0.6875,0.1563,0.6875,0.1563,0.625,0.1563,0.625,0.1563,0.6875,0.125,0.625,0.1563,0.6875,0.125,0.6875,0.125,0.625,0.125,0.625,0.125,0.6875,0.0938,0.625,0.125,0.6875,0.0938,0.6875,0.0938,0.625,0.0938,0.625,0.0938,0.6875,0.0625,0.625,0.0938,0.6875,0.0625,0.6875,0.0625,0.625,0.0625,0.625,0.0625,0.6875,0.0313,0.625,0.0625,0.6875,0.0313,0.6875,0.0313,0.625,0.0313,0.625,0.0313,0.6875,0.0,0.625,0.0313,0.6875,0.0,0.6875,0.0,0.625,1.0,0.5625,1.0,0.625,0.9688,0.5625,1.0,0.625,0.9688,0.625,0.9688,0.5625,0.9688,0.5625,0.9688,0.625,0.9375,0.5625,0.9688,0.625,0.9375,0.625,0.9375,0.5625,0.9375,0.5625,0.9375,0.625,0.9063,0.5625,0.9375,0.625,0.9063,0.625,0.9063,0.5625,0.9063,0.5625,0.9063,0.625,0.875,0.5625,0.9063,0.625,0.875,0.625,0.875,0.5625,0.875,0.5625,0.875,0.625,0.8438,0.5625,0.875,0.625,0.8438,0.625,0.8438,0.5625,0.8438,0.5625,0.8438,0.625,0.8125,0.5625,0.8438,0.625,0.8125,0.625,0.8125,0.5625,0.8125,0.5625,0.8125,0.625,0.7813,0.5625,0.8125,0.625,0.7813,0.625,0.7813,0.5625,0.7813,0.5625,0.7813,0.625,0.75,0.5625,0.7813,0.625,0.75,0.625,0.75,0.5625,0.75,0.5625,0.75,0.625,0.7188,0.5625,0.75,0.625,0.7188,0.625,0.7188,0.5625,0.7188,0.5625,0.7188,0.625,0.6875,0.5625,0.7188,0.625,0.6875,0.625,0.6875,0.5625,0.6875,0.5625,0.6875,0.625,0.6563,0.5625,0.6875,0.625,0.6563,0.625,0.6563,0.5625,0.6563,0.5625,0.6563,0.625,0.625,0.5625,0.6563,0.625,0.625,0.625,0.625,0.5625,0.625,0.5625,0.625,0.625,0.5938,0.5625,0.625,0.625,0.5938,0.625,0.5938,0.5625,0.5938,0.5625,0.5938,0.625,0.5625,0.5625,0.5938,0.625,0.5625,0.625,0.5625,0.5625,0.5625,0.5625,0.5625,0.625,0.5313,0.5625,0.5625,0.625,0.5313,0.625,0.5313,0.5625,0.5313,0.5625,0.5313,0.625,0.5,0.5625,0.5313,0.625,0.5,0.625,0.5,0.5625,0.5,0.5625,0.5,0.625,0.4688,0.5625,0.5,0.625,0.4688,0.625,0.4688,0.5625,0.4688,0.5625,0.4688,0.625,0.4375,0.5625,0.4688,0.625,0.4375,0.625,0.4375,0.5625,0.4375,0.5625,0.4375,0.625,0.4063,0.5625,0.4375,0.625,0.4063,0.625,0.4063,0.5625,0.4063,0.5625,0.4063,0.625,0.375,0.5625,0.4063,0.625,0.375,0.625,0.375,0.5625,0.375,0.5625,0.375,0.625,0.3438,0.5625,0.375,0.625,0.3438,0.625,0.3438,0.5625,0.3438,0.5625,0.3438,0.625,0.3125,0.5625,0.3438,0.625,0.3125,0.625,0.3125,0.5625,0.3125,0.5625,0.3125,0.625,0.2813,0.5625,0.3125,0.625,0.2813,0.625,0.2813,0.5625,0.2813,0.5625,0.2813,0.625,0.25,0.5625,0.2813,0.625,0.25,0.625,0.25,0.5625,0.25,0.5625,0.25,0.625,0.2188,0.5625,0.25,0.625,0.2188,0.625,0.2188,0.5625,0.2188,0.5625,0.2188,0.625,0.1875,0.5625,0.2188,0.625,0.1875,0.625,0.1875,0.5625,0.1875,0.5625,0.1875,0.625,0.1563,0.5625,0.1875,0.625,0.1563,0.625,0.1563,0.5625,0.1563,0.5625,0.1563,0.625,0.125,0.5625,0.1563,0.625,0.125,0.625,0.125,0.5625,0.125,0.5625,0.125,0.625,0.0938,0.5625,0.125,0.625,0.0938,0.625,0.0938,0.5625,0.0938,0.5625,0.0938,0.625,0.0625,0.5625,0.0938,0.625,0.0625,0.625,0.0625,0.5625,0.0625,0.5625,0.0625,0.625,0.0313,0.5625,0.0625,0.625,0.0313,0.625,0.0313,0.5625,0.0313,0.5625,0.0313,0.625,0.0,0.5625,0.0313,0.625,0.0,0.625,0.0,0.5625,1.0,0.5,1.0,0.5625,0.9688,0.5,1.0,0.5625,0.9688,0.5625,0.9688,0.5,0.9688,0.5,0.9688,0.5625,0.9375,0.5,0.9688,0.5625,0.9375,0.5625,0.9375,0.5,0.9375,0.5,0.9375,0.5625,0.9063,0.5,0.9375,0.5625,0.9063,0.5625,0.9063,0.5,0.9063,0.5,0.9063,0.5625,0.875,0.5,0.9063,0.5625,0.875,0.5625,0.875,0.5,0.875,0.5,0.875,0.5625,0.8438,0.5,0.875,0.5625,0.8438,0.5625,0.8438,0.5,0.8438,0.5,0.8438,0.5625,0.8125,0.5,0.8438,0.5625,0.8125,0.5625,0.8125,0.5,0.8125,0.5,0.8125,0.5625,0.7813,0.5,0.8125,0.5625,0.7813,0.5625,0.7813,0.5,0.7813,0.5,0.7813,0.5625,0.75,0.5,0.7813,0.5625,0.75,0.5625,0.75,0.5,0.75,0.5,0.75,0.5625,0.7188,0.5,0.75,0.5625,0.7188,0.5625,0.7188,0.5,0.7188,0.5,0.7188,0.5625,0.6875,0.5,0.7188,0.5625,0.6875,0.5625,0.6875,0.5,0.6875,0.5,0.6875,0.5625,0.6563,0.5,0.6875,0.5625,0.6563,0.5625,0.6563,0.5,0.6563,0.5,0.6563,0.5625,0.625,0.5,0.6563,0.5625,0.625,0.5625,0.625,0.5,0.625,0.5,0.625,0.5625,0.5938,0.5,0.625,0.5625,0.5938,0.5625,0.5938,0.5,0.5938,0.5,0.5938,0.5625,0.5625,0.5,0.5938,0.5625,0.5625,0.5625,0.5625,0.5,0.5625,0.5,0.5625,0.5625,0.5313,0.5,0.5625,0.5625,0.5313,0.5625,0.5313,0.5,0.5313,0.5,0.5313,0.5625,0.5,0.5,0.5313,0.5625,0.5,0.5625,0.5,0.5,0.5,0.5,0.5,0.5625,0.4688,0.5,0.5,0.5625,0.4688,0.5625,0.4688,0.5,0.4688,0.5,0.4688,0.5625,0.4375,0.5,0.4688,0.5625,0.4375,0.5625,0.4375,0.5,0.4375,0.5,0.4375,0.5625,0.4063,0.5,0.4375,0.5625,0.4063,0.5625,0.4063,0.5,0.4063,0.5,0.4063,0.5625,0.375,0.5,0.4063,0.5625,0.375,0.5625,0.375,0.5,0.375,0.5,0.375,0.5625,0.3438,0.5,0.375,0.5625,0.3438,0.5625,0.3438,0.5,0.3438,0.5,0.3438,0.5625,0.3125,0.5,0.3438,0.5625,0.3125,0.5625,0.3125,0.5,0.3125,0.5,0.3125,0.5625,0.2813,0.5,0.3125,0.5625,0.2813,0.5625,0.2813,0.5,0.2813,0.5,0.2813,0.5625,0.25,0.5,0.2813,0.5625,0.25,0.5625,0.25,0.5,0.25,0.5,0.25,0.5625,0.2188,0.5,0.25,0.5625,0.2188,0.5625,0.2188,0.5,0.2188,0.5,0.2188,0.5625,0.1875,0.5,0.2188,0.5625,0.1875,0.5625,0.1875,0.5,0.1875,0.5,0.1875,0.5625,0.1563,0.5,0.1875,0.5625,0.1563,0.5625,0.1563,0.5,0.1563,0.5,0.1563,0.5625,0.125,0.5,0.1563,0.5625,0.125,0.5625,0.125,0.5,0.125,0.5,0.125,0.5625,0.0938,0.5,0.125,0.5625,0.0938,0.5625,0.0938,0.5,0.0938,0.5,0.0938,0.5625,0.0625,0.5,0.0938,0.5625,0.0625,0.5625,0.0625,0.5,0.0625,0.5,0.0625,0.5625,0.0313,0.5,0.0625,0.5625,0.0313,0.5625,0.0313,0.5,0.0313,0.5,0.0313,0.5625,0.0,0.5,0.0313,0.5625,0.0,0.5625,0.0,0.5,1.0,0.4375,1.0,0.5,0.9688,0.4375,1.0,0.5,0.9688,0.5,0.9688,0.4375,0.9688,0.4375,0.9688,0.5,0.9375,0.4375,0.9688,0.5,0.9375,0.5,0.9375,0.4375,0.9375,0.4375,0.9375,0.5,0.9063,0.4375,0.9375,0.5,0.9063,0.5,0.9063,0.4375,0.9063,0.4375,0.9063,0.5,0.875,0.4375,0.9063,0.5,0.875,0.5,0.875,0.4375,0.875,0.4375,0.875,0.5,0.8438,0.4375,0.875,0.5,0.8438,0.5,0.8438,0.4375,0.8438,0.4375,0.8438,0.5,0.8125,0.4375,0.8438,0.5,0.8125,0.5,0.8125,0.4375,0.8125,0.4375,0.8125,0.5,0.7813,0.4375,0.8125,0.5,0.7813,0.5,0.7813,0.4375,0.7813,0.4375,0.7813,0.5,0.75,0.4375,0.7813,0.5,0.75,0.5,0.75,0.4375,0.75,0.4375,0.75,0.5,0.7188,0.4375,0.75,0.5,0.7188,0.5,0.7188,0.4375,0.7188,0.4375,0.7188,0.5,0.6875,0.4375,0.7188,0.5,0.6875,0.5,0.6875,0.4375,0.6875,0.4375,0.6875,0.5,0.6563,0.4375,0.6875,0.5,0.6563,0.5,0.6563,0.4375,0.6563,0.4375,0.6563,0.5,0.625,0.4375,0.6563,0.5,0.625,0.5,0.625,0.4375,0.625,0.4375,0.625,0.5,0.5938,0.4375,0.625,0.5,0.5938,0.5,0.5938,0.4375,0.5938,0.4375,0.5938,0.5,0.5625,0.4375,0.5938,0.5,0.5625,0.5,0.5625,0.4375,0.5625,0.4375,0.5625,0.5,0.5313,0.4375,0.5625,0.5,0.5313,0.5,0.5313,0.4375,0.5313,0.4375,0.5313,0.5,0.5,0.4375,0.5313,0.5,0.5,0.5,0.5,0.4375,0.5,0.4375,0.5,0.5,0.4688,0.4375,0.5,0.5,0.4688,0.5,0.4688,0.4375,0.4688,0.4375,0.4688,0.5,0.4375,0.4375,0.4688,0.5,0.4375,0.5,0.4375,0.4375,0.4375,0.4375,0.4375,0.5,0.4063,0.4375,0.4375,0.5,0.4063,0.5,0.4063,0.4375,0.4063,0.4375,0.4063,0.5,0.375,0.4375,0.4063,0.5,0.375,0.5,0.375,0.4375,0.375,0.4375,0.375,0.5,0.3438,0.4375,0.375,0.5,0.3438,0.5,0.3438,0.4375,0.3438,0.4375,0.3438,0.5,0.3125,0.4375,0.3438,0.5,0.3125,0.5,0.3125,0.4375,0.3125,0.4375,0.3125,0.5,0.2813,0.4375,0.3125,0.5,0.2813,0.5,0.2813,0.4375,0.2813,0.4375,0.2813,0.5,0.25,0.4375,0.2813,0.5,0.25,0.5,0.25,0.4375,0.25,0.4375,0.25,0.5,0.2188,0.4375,0.25,0.5,0.2188,0.5,0.2188,0.4375,0.2188,0.4375,0.2188,0.5,0.1875,0.4375,0.2188,0.5,0.1875,0.5,0.1875,0.4375,0.1875,0.4375,0.1875,0.5,0.1563,0.4375,0.1875,0.5,0.1563,0.5,0.1563,0.4375,0.1563,0.4375,0.1563,0.5,0.125,0.4375,0.1563,0.5,0.125,0.5,0.125,0.4375,0.125,0.4375,0.125,0.5,0.0938,0.4375,0.125,0.5,0.0938,0.5,0.0938,0.4375,0.0938,0.4375,0.0938,0.5,0.0625,0.4375,0.0938,0.5,0.0625,0.5,0.0625,0.4375,0.0625,0.4375,0.0625,0.5,0.0313,0.4375,0.0625,0.5,0.0313,0.5,0.0313,0.4375,0.0313,0.4375,0.0313,0.5,0.0,0.4375,0.0313,0.5,0.0,0.5,0.0,0.4375,1.0,0.375,1.0,0.4375,0.9688,0.375,1.0,0.4375,0.9688,0.4375,0.9688,0.375,0.9688,0.375,0.9688,0.4375,0.9375,0.375,0.9688,0.4375,0.9375,0.4375,0.9375,0.375,0.9375,0.375,0.9375,0.4375,0.9063,0.375,0.9375,0.4375,0.9063,0.4375,0.9063,0.375,0.9063,0.375,0.9063,0.4375,0.875,0.375,0.9063,0.4375,0.875,0.4375,0.875,0.375,0.875,0.375,0.875,0.4375,0.8438,0.375,0.875,0.4375,0.8438,0.4375,0.8438,0.375,0.8438,0.375,0.8438,0.4375,0.8125,0.375,0.8438,0.4375,0.8125,0.4375,0.8125,0.375,0.8125,0.375,0.8125,0.4375,0.7813,0.375,0.8125,0.4375,0.7813,0.4375,0.7813,0.375,0.7813,0.375,0.7813,0.4375,0.75,0.375,0.7813,0.4375,0.75,0.4375,0.75,0.375,0.75,0.375,0.75,0.4375,0.7188,0.375,0.75,0.4375,0.7188,0.4375,0.7188,0.375,0.7188,0.375,0.7188,0.4375,0.6875,0.375,0.7188,0.4375,0.6875,0.4375,0.6875,0.375,0.6875,0.375,0.6875,0.4375,0.6563,0.375,0.6875,0.4375,0.6563,0.4375,0.6563,0.375,0.6563,0.375,0.6563,0.4375,0.625,0.375,0.6563,0.4375,0.625,0.4375,0.625,0.375,0.625,0.375,0.625,0.4375,0.5938,0.375,0.625,0.4375,0.5938,0.4375,0.5938,0.375,0.5938,0.375,0.5938,0.4375,0.5625,0.375,0.5938,0.4375,0.5625,0.4375,0.5625,0.375,0.5625,0.375,0.5625,0.4375,0.5313,0.375,0.5625,0.4375,0.5313,0.4375,0.5313,0.375,0.5313,0.375,0.5313,0.4375,0.5,0.375,0.5313,0.4375,0.5,0.4375,0.5,0.375,0.5,0.375,0.5,0.4375,0.4688,0.375,0.5,0.4375,0.4688,0.4375,0.4688,0.375,0.4688,0.375,0.4688,0.4375,0.4375,0.375,0.4688,0.4375,0.4375,0.4375,0.4375,0.375,0.4375,0.375,0.4375,0.4375,0.4063,0.375,0.4375,0.4375,0.4063,0.4375,0.4063,0.375,0.4063,0.375,0.4063,0.4375,0.375,0.375,0.4063,0.4375,0.375,0.4375,0.375,0.375,0.375,0.375,0.375,0.4375,0.3438,0.375,0.375,0.4375,0.3438,0.4375,0.3438,0.375,0.3438,0.375,0.3438,0.4375,0.3125,0.375,0.3438,0.4375,0.3125,0.4375,0.3125,0.375,0.3125,0.375,0.3125,0.4375,0.2813,0.375,0.3125,0.4375,0.2813,0.4375,0.2813,0.375,0.2813,0.375,0.2813,0.4375,0.25,0.375,0.2813,0.4375,0.25,0.4375,0.25,0.375,0.25,0.375,0.25,0.4375,0.2188,0.375,0.25,0.4375,0.2188,0.4375,0.2188,0.375,0.2188,0.375,0.2188,0.4375,0.1875,0.375,0.2188,0.4375,0.1875,0.4375,0.1875,0.375,0.1875,0.375,0.1875,0.4375,0.1563,0.375,0.1875,0.4375,0.1563,0.4375,0.1563,0.375,0.1563,0.375,0.1563,0.4375,0.125,0.375,0.1563,0.4375,0.125,0.4375,0.125,0.375,0.125,0.375,0.125,0.4375,0.0938,0.375,0.125,0.4375,0.0938,0.4375,0.0938,0.375,0.0938,0.375,0.0938,0.4375,0.0625,0.375,0.0938,0.4375,0.0625,0.4375,0.0625,0.375,0.0625,0.375,0.0625,0.4375,0.0313,0.375,0.0625,0.4375,0.0313,0.4375,0.0313,0.375,0.0313,0.375,0.0313,0.4375,0.0,0.375,0.0313,0.4375,0.0,0.4375,0.0,0.375,1.0,0.3125,1.0,0.375,0.9688,0.3125,1.0,0.375,0.9688,0.375,0.9688,0.3125,0.9688,0.3125,0.9688,0.375,0.9375,0.3125,0.9688,0.375,0.9375,0.375,0.9375,0.3125,0.9375,0.3125,0.9375,0.375,0.9063,0.3125,0.9375,0.375,0.9063,0.375,0.9063,0.3125,0.9063,0.3125,0.9063,0.375,0.875,0.3125,0.9063,0.375,0.875,0.375,0.875,0.3125,0.875,0.3125,0.875,0.375,0.8438,0.3125,0.875,0.375,0.8438,0.375,0.8438,0.3125,0.8438,0.3125,0.8438,0.375,0.8125,0.3125,0.8438,0.375,0.8125,0.375,0.8125,0.3125,0.8125,0.3125,0.8125,0.375,0.7813,0.3125,0.8125,0.375,0.7813,0.375,0.7813,0.3125,0.7813,0.3125,0.7813,0.375,0.75,0.3125,0.7813,0.375,0.75,0.375,0.75,0.3125,0.75,0.3125,0.75,0.375,0.7188,0.3125,0.75,0.375,0.7188,0.375,0.7188,0.3125,0.7188,0.3125,0.7188,0.375,0.6875,0.3125,0.7188,0.375,0.6875,0.375,0.6875,0.3125,0.6875,0.3125,0.6875,0.375,0.6563,0.3125,0.6875,0.375,0.6563,0.375,0.6563,0.3125,0.6563,0.3125,0.6563,0.375,0.625,0.3125,0.6563,0.375,0.625,0.375,0.625,0.3125,0.625,0.3125,0.625,0.375,0.5938,0.3125,0.625,0.375,0.5938,0.375,0.5938,0.3125,0.5938,0.3125,0.5938,0.375,0.5625,0.3125,0.5938,0.375,0.5625,0.375,0.5625,0.3125,0.5625,0.3125,0.5625,0.375,0.5313,0.3125,0.5625,0.375,0.5313,0.375,0.5313,0.3125,0.5313,0.3125,0.5313,0.375,0.5,0.3125,0.5313,0.375,0.5,0.375,0.5,0.3125,0.5,0.3125,0.5,0.375,0.4688,0.3125,0.5,0.375,0.4688,0.375,0.4688,0.3125,0.4688,0.3125,0.4688,0.375,0.4375,0.3125,0.4688,0.375,0.4375,0.375,0.4375,0.3125,0.4375,0.3125,0.4375,0.375,0.4063,0.3125,0.4375,0.375,0.4063,0.375,0.4063,0.3125,0.4063,0.3125,0.4063,0.375,0.375,0.3125,0.4063,0.375,0.375,0.375,0.375,0.3125,0.375,0.3125,0.375,0.375,0.3438,0.3125,0.375,0.375,0.3438,0.375,0.3438,0.3125,0.3438,0.3125,0.3438,0.375,0.3125,0.3125,0.3438,0.375,0.3125,0.375,0.3125,0.3125,0.3125,0.3125,0.3125,0.375,0.2813,0.3125,0.3125,0.375,0.2813,0.375,0.2813,0.3125,0.2813,0.3125,0.2813,0.375,0.25,0.3125,0.2813,0.375,0.25,0.375,0.25,0.3125,0.25,0.3125,0.25,0.375,0.2188,0.3125,0.25,0.375,0.2188,0.375,0.2188,0.3125,0.2188,0.3125,0.2188,0.375,0.1875,0.3125,0.2188,0.375,0.1875,0.375,0.1875,0.3125,0.1875,0.3125,0.1875,0.375,0.1563,0.3125,0.1875,0.375,0.1563,0.375,0.1563,0.3125,0.1563,0.3125,0.1563,0.375,0.125,0.3125,0.1563,0.375,0.125,0.375,0.125,0.3125,0.125,0.3125,0.125,0.375,0.0938,0.3125,0.125,0.375,0.0938,0.375,0.0938,0.3125,0.0938,0.3125,0.0938,0.375,0.0625,0.3125,0.0938,0.375,0.0625,0.375,0.0625,0.3125,0.0625,0.3125,0.0625,0.375,0.0313,0.3125,0.0625,0.375,0.0313,0.375,0.0313,0.3125,0.0313,0.3125,0.0313,0.375,0.0,0.3125,0.0313,0.375,0.0,0.375,0.0,0.3125,1.0,0.25,1.0,0.3125,0.9688,0.25,1.0,0.3125,0.9688,0.3125,0.9688,0.25,0.9688,0.25,0.9688,0.3125,0.9375,0.25,0.9688,0.3125,0.9375,0.3125,0.9375,0.25,0.9375,0.25,0.9375,0.3125,0.9063,0.25,0.9375,0.3125,0.9063,0.3125,0.9063,0.25,0.9063,0.25,0.9063,0.3125,0.875,0.25,0.9063,0.3125,0.875,0.3125,0.875,0.25,0.875,0.25,0.875,0.3125,0.8438,0.25,0.875,0.3125,0.8438,0.3125,0.8438,0.25,0.8438,0.25,0.8438,0.3125,0.8125,0.25,0.8438,0.3125,0.8125,0.3125,0.8125,0.25,0.8125,0.25,0.8125,0.3125,0.7813,0.25,0.8125,0.3125,0.7813,0.3125,0.7813,0.25,0.7813,0.25,0.7813,0.3125,0.75,0.25,0.7813,0.3125,0.75,0.3125,0.75,0.25,0.75,0.25,0.75,0.3125,0.7188,0.25,0.75,0.3125,0.7188,0.3125,0.7188,0.25,0.7188,0.25,0.7188,0.3125,0.6875,0.25,0.7188,0.3125,0.6875,0.3125,0.6875,0.25,0.6875,0.25,0.6875,0.3125,0.6563,0.25,0.6875,0.3125,0.6563,0.3125,0.6563,0.25,0.6563,0.25,0.6563,0.3125,0.625,0.25,0.6563,0.3125,0.625,0.3125,0.625,0.25,0.625,0.25,0.625,0.3125,0.5938,0.25,0.625,0.3125,0.5938,0.3125,0.5938,0.25,0.5938,0.25,0.5938,0.3125,0.5625,0.25,0.5938,0.3125,0.5625,0.3125,0.5625,0.25,0.5625,0.25,0.5625,0.3125,0.5313,0.25,0.5625,0.3125,0.5313,0.3125,0.5313,0.25,0.5313,0.25,0.5313,0.3125,0.5,0.25,0.5313,0.3125,0.5,0.3125,0.5,0.25,0.5,0.25,0.5,0.3125,0.4688,0.25,0.5,0.3125,0.4688,0.3125,0.4688,0.25,0.4688,0.25,0.4688,0.3125,0.4375,0.25,0.4688,0.3125,0.4375,0.3125,0.4375,0.25,0.4375,0.25,0.4375,0.3125,0.4063,0.25,0.4375,0.3125,0.4063,0.3125,0.4063,0.25,0.4063,0.25,0.4063,0.3125,0.375,0.25,0.4063,0.3125,0.375,0.3125,0.375,0.25,0.375,0.25,0.375,0.3125,0.3438,0.25,0.375,0.3125,0.3438,0.3125,0.3438,0.25,0.3438,0.25,0.3438,0.3125,0.3125,0.25,0.3438,0.3125,0.3125,0.3125,0.3125,0.25,0.3125,0.25,0.3125,0.3125,0.2813,0.25,0.3125,0.3125,0.2813,0.3125,0.2813,0.25,0.2813,0.25,0.2813,0.3125,0.25,0.25,0.2813,0.3125,0.25,0.3125,0.25,0.25,0.25,0.25,0.25,0.3125,0.2188,0.25,0.25,0.3125,0.2188,0.3125,0.2188,0.25,0.2188,0.25,0.2188,0.3125,0.1875,0.25,0.2188,0.3125,0.1875,0.3125,0.1875,0.25,0.1875,0.25,0.1875,0.3125,0.1563,0.25,0.1875,0.3125,0.1563,0.3125,0.1563,0.25,0.1563,0.25,0.1563,0.3125,0.125,0.25,0.1563,0.3125,0.125,0.3125,0.125,0.25,0.125,0.25,0.125,0.3125,0.0938,0.25,0.125,0.3125,0.0938,0.3125,0.0938,0.25,0.0938,0.25,0.0938,0.3125,0.0625,0.25,0.0938,0.3125,0.0625,0.3125,0.0625,0.25,0.0625,0.25,0.0625,0.3125,0.0313,0.25,0.0625,0.3125,0.0313,0.3125,0.0313,0.25,0.0313,0.25,0.0313,0.3125,0.0,0.25,0.0313,0.3125,0.0,0.3125,0.0,0.25,1.0,0.1875,1.0,0.25,0.9688,0.1875,1.0,0.25,0.9688,0.25,0.9688,0.1875,0.9688,0.1875,0.9688,0.25,0.9375,0.1875,0.9688,0.25,0.9375,0.25,0.9375,0.1875,0.9375,0.1875,0.9375,0.25,0.9063,0.1875,0.9375,0.25,0.9063,0.25,0.9063,0.1875,0.9063,0.1875,0.9063,0.25,0.875,0.1875,0.9063,0.25,0.875,0.25,0.875,0.1875,0.875,0.1875,0.875,0.25,0.8438,0.1875,0.875,0.25,0.8438,0.25,0.8438,0.1875,0.8438,0.1875,0.8438,0.25,0.8125,0.1875,0.8438,0.25,0.8125,0.25,0.8125,0.1875,0.8125,0.1875,0.8125,0.25,0.7813,0.1875,0.8125,0.25,0.7813,0.25,0.7813,0.1875,0.7813,0.1875,0.7813,0.25,0.75,0.1875,0.7813,0.25,0.75,0.25,0.75,0.1875,0.75,0.1875,0.75,0.25,0.7188,0.1875,0.75,0.25,0.7188,0.25,0.7188,0.1875,0.7188,0.1875,0.7188,0.25,0.6875,0.1875,0.7188,0.25,0.6875,0.25,0.6875,0.1875,0.6875,0.1875,0.6875,0.25,0.6563,0.1875,0.6875,0.25,0.6563,0.25,0.6563,0.1875,0.6563,0.1875,0.6563,0.25,0.625,0.1875,0.6563,0.25,0.625,0.25,0.625,0.1875,0.625,0.1875,0.625,0.25,0.5938,0.1875,0.625,0.25,0.5938,0.25,0.5938,0.1875,0.5938,0.1875,0.5938,0.25,0.5625,0.1875,0.5938,0.25,0.5625,0.25,0.5625,0.1875,0.5625,0.1875,0.5625,0.25,0.5313,0.1875,0.5625,0.25,0.5313,0.25,0.5313,0.1875,0.5313,0.1875,0.5313,0.25,0.5,0.1875,0.5313,0.25,0.5,0.25,0.5,0.1875,0.5,0.1875,0.5,0.25,0.4688,0.1875,0.5,0.25,0.4688,0.25,0.4688,0.1875,0.4688,0.1875,0.4688,0.25,0.4375,0.1875,0.4688,0.25,0.4375,0.25,0.4375,0.1875,0.4375,0.1875,0.4375,0.25,0.4063,0.1875,0.4375,0.25,0.4063,0.25,0.4063,0.1875,0.4063,0.1875,0.4063,0.25,0.375,0.1875,0.4063,0.25,0.375,0.25,0.375,0.1875,0.375,0.1875,0.375,0.25,0.3438,0.1875,0.375,0.25,0.3438,0.25,0.3438,0.1875,0.3438,0.1875,0.3438,0.25,0.3125,0.1875,0.3438,0.25,0.3125,0.25,0.3125,0.1875,0.3125,0.1875,0.3125,0.25,0.2813,0.1875,0.3125,0.25,0.2813,0.25,0.2813,0.1875,0.2813,0.1875,0.2813,0.25,0.25,0.1875,0.2813,0.25,0.25,0.25,0.25,0.1875,0.25,0.1875,0.25,0.25,0.2188,0.1875,0.25,0.25,0.2188,0.25,0.2188,0.1875,0.2188,0.1875,0.2188,0.25,0.1875,0.1875,0.2188,0.25,0.1875,0.25,0.1875,0.1875,0.1875,0.1875,0.1875,0.25,0.1563,0.1875,0.1875,0.25,0.1563,0.25,0.1563,0.1875,0.1563,0.1875,0.1563,0.25,0.125,0.1875,0.1563,0.25,0.125,0.25,0.125,0.1875,0.125,0.1875,0.125,0.25,0.0938,0.1875,0.125,0.25,0.0938,0.25,0.0938,0.1875,0.0938,0.1875,0.0938,0.25,0.0625,0.1875,0.0938,0.25,0.0625,0.25,0.0625,0.1875,0.0625,0.1875,0.0625,0.25,0.0313,0.1875,0.0625,0.25,0.0313,0.25,0.0313,0.1875,0.0313,0.1875,0.0313,0.25,0.0,0.1875,0.0313,0.25,0.0,0.25,0.0,0.1875,1.0,0.125,1.0,0.1875,0.9688,0.125,1.0,0.1875,0.9688,0.1875,0.9688,0.125,0.9688,0.125,0.9688,0.1875,0.9375,0.125,0.9688,0.1875,0.9375,0.1875,0.9375,0.125,0.9375,0.125,0.9375,0.1875,0.9063,0.125,0.9375,0.1875,0.9063,0.1875,0.9063,0.125,0.9063,0.125,0.9063,0.1875,0.875,0.125,0.9063,0.1875,0.875,0.1875,0.875,0.125,0.875,0.125,0.875,0.1875,0.8438,0.125,0.875,0.1875,0.8438,0.1875,0.8438,0.125,0.8438,0.125,0.8438,0.1875,0.8125,0.125,0.8438,0.1875,0.8125,0.1875,0.8125,0.125,0.8125,0.125,0.8125,0.1875,0.7813,0.125,0.8125,0.1875,0.7813,0.1875,0.7813,0.125,0.7813,0.125,0.7813,0.1875,0.75,0.125,0.7813,0.1875,0.75,0.1875,0.75,0.125,0.75,0.125,0.75,0.1875,0.7188,0.125,0.75,0.1875,0.7188,0.1875,0.7188,0.125,0.7188,0.125,0.7188,0.1875,0.6875,0.125,0.7188,0.1875,0.6875,0.1875,0.6875,0.125,0.6875,0.125,0.6875,0.1875,0.6563,0.125,0.6875,0.1875,0.6563,0.1875,0.6563,0.125,0.6563,0.125,0.6563,0.1875,0.625,0.125,0.6563,0.1875,0.625,0.1875,0.625,0.125,0.625,0.125,0.625,0.1875,0.5938,0.125,0.625,0.1875,0.5938,0.1875,0.5938,0.125,0.5938,0.125,0.5938,0.1875,0.5625,0.125,0.5938,0.1875,0.5625,0.1875,0.5625,0.125,0.5625,0.125,0.5625,0.1875,0.5313,0.125,0.5625,0.1875,0.5313,0.1875,0.5313,0.125,0.5313,0.125,0.5313,0.1875,0.5,0.125,0.5313,0.1875,0.5,0.1875,0.5,0.125,0.5,0.125,0.5,0.1875,0.4688,0.125,0.5,0.1875,0.4688,0.1875,0.4688,0.125,0.4688,0.125,0.4688,0.1875,0.4375,0.125,0.4688,0.1875,0.4375,0.1875,0.4375,0.125,0.4375,0.125,0.4375,0.1875,0.4063,0.125,0.4375,0.1875,0.4063,0.1875,0.4063,0.125,0.4063,0.125,0.4063,0.1875,0.375,0.125,0.4063,0.1875,0.375,0.1875,0.375,0.125,0.375,0.125,0.375,0.1875,0.3438,0.125,0.375,0.1875,0.3438,0.1875,0.3438,0.125,0.3438,0.125,0.3438,0.1875,0.3125,0.125,0.3438,0.1875,0.3125,0.1875,0.3125,0.125,0.3125,0.125,0.3125,0.1875,0.2813,0.125,0.3125,0.1875,0.2813,0.1875,0.2813,0.125,0.2813,0.125,0.2813,0.1875,0.25,0.125,0.2813,0.1875,0.25,0.1875,0.25,0.125,0.25,0.125,0.25,0.1875,0.2188,0.125,0.25,0.1875,0.2188,0.1875,0.2188,0.125,0.2188,0.125,0.2188,0.1875,0.1875,0.125,0.2188,0.1875,0.1875,0.1875,0.1875,0.125,0.1875,0.125,0.1875,0.1875,0.1563,0.125,0.1875,0.1875,0.1563,0.1875,0.1563,0.125,0.1563,0.125,0.1563,0.1875,0.125,0.125,0.1563,0.1875,0.125,0.1875,0.125,0.125,0.125,0.125,0.125,0.1875,0.0938,0.125,0.125,0.1875,0.0938,0.1875,0.0938,0.125,0.0938,0.125,0.0938,0.1875,0.0625,0.125,0.0938,0.1875,0.0625,0.1875,0.0625,0.125,0.0625,0.125,0.0625,0.1875,0.0313,0.125,0.0625,0.1875,0.0313,0.1875,0.0313,0.125,0.0313,0.125,0.0313,0.1875,0.0,0.125,0.0313,0.1875,0.0,0.1875,0.0,0.125,1.0,0.0625,1.0,0.125,0.9688,0.0625,1.0,0.125,0.9688,0.125,0.9688,0.0625,0.9688,0.0625,0.9688,0.125,0.9375,0.0625,0.9688,0.125,0.9375,0.125,0.9375,0.0625,0.9375,0.0625,0.9375,0.125,0.9063,0.0625,0.9375,0.125,0.9063,0.125,0.9063,0.0625,0.9063,0.0625,0.9063,0.125,0.875,0.0625,0.9063,0.125,0.875,0.125,0.875,0.0625,0.875,0.0625,0.875,0.125,0.8438,0.0625,0.875,0.125,0.8438,0.125,0.8438,0.0625,0.8438,0.0625,0.8438,0.125,0.8125,0.0625,0.8438,0.125,0.8125,0.125,0.8125,0.0625,0.8125,0.0625,0.8125,0.125,0.7813,0.0625,0.8125,0.125,0.7813,0.125,0.7813,0.0625,0.7813,0.0625,0.7813,0.125,0.75,0.0625,0.7813,0.125,0.75,0.125,0.75,0.0625,0.75,0.0625,0.75,0.125,0.7188,0.0625,0.75,0.125,0.7188,0.125,0.7188,0.0625,0.7188,0.0625,0.7188,0.125,0.6875,0.0625,0.7188,0.125,0.6875,0.125,0.6875,0.0625,0.6875,0.0625,0.6875,0.125,0.6563,0.0625,0.6875,0.125,0.6563,0.125,0.6563,0.0625,0.6563,0.0625,0.6563,0.125,0.625,0.0625,0.6563,0.125,0.625,0.125,0.625,0.0625,0.625,0.0625,0.625,0.125,0.5938,0.0625,0.625,0.125,0.5938,0.125,0.5938,0.0625,0.5938,0.0625,0.5938,0.125,0.5625,0.0625,0.5938,0.125,0.5625,0.125,0.5625,0.0625,0.5625,0.0625,0.5625,0.125,0.5313,0.0625,0.5625,0.125,0.5313,0.125,0.5313,0.0625,0.5313,0.0625,0.5313,0.125,0.5,0.0625,0.5313,0.125,0.5,0.125,0.5,0.0625,0.5,0.0625,0.5,0.125,0.4688,0.0625,0.5,0.125,0.4688,0.125,0.4688,0.0625,0.4688,0.0625,0.4688,0.125,0.4375,0.0625,0.4688,0.125,0.4375,0.125,0.4375,0.0625,0.4375,0.0625,0.4375,0.125,0.4063,0.0625,0.4375,0.125,0.4063,0.125,0.4063,0.0625,0.4063,0.0625,0.4063,0.125,0.375,0.0625,0.4063,0.125,0.375,0.125,0.375,0.0625,0.375,0.0625,0.375,0.125,0.3438,0.0625,0.375,0.125,0.3438,0.125,0.3438,0.0625,0.3438,0.0625,0.3438,0.125,0.3125,0.0625,0.3438,0.125,0.3125,0.125,0.3125,0.0625,0.3125,0.0625,0.3125,0.125,0.2813,0.0625,0.3125,0.125,0.2813,0.125,0.2813,0.0625,0.2813,0.0625,0.2813,0.125,0.25,0.0625,0.2813,0.125,0.25,0.125,0.25,0.0625,0.25,0.0625,0.25,0.125,0.2188,0.0625,0.25,0.125,0.2188,0.125,0.2188,0.0625,0.2188,0.0625,0.2188,0.125,0.1875,0.0625,0.2188,0.125,0.1875,0.125,0.1875,0.0625,0.1875,0.0625,0.1875,0.125,0.1563,0.0625,0.1875,0.125,0.1563,0.125,0.1563,0.0625,0.1563,0.0625,0.1563,0.125,0.125,0.0625,0.1563,0.125,0.125,0.125,0.125,0.0625,0.125,0.0625,0.125,0.125,0.0938,0.0625,0.125,0.125,0.0938,0.125,0.0938,0.0625,0.0938,0.0625,0.0938,0.125,0.0625,0.0625,0.0938,0.125,0.0625,0.125,0.0625,0.0625,0.0625,0.0625,0.0625,0.125,0.0313,0.0625,0.0625,0.125,0.0313,0.125,0.0313,0.0625,0.0313,0.0625,0.0313,0.125,0.0,0.0625,0.0313,0.125,0.0,0.125,0.0,0.0625,1.0,0.0625,0.9688,0.0625,0.9688,0.0,0.9688,0.0625,0.9375,0.0625,0.9375,0.0,0.9375,0.0625,0.9063,0.0625,0.9063,0.0,0.9063,0.0625,0.875,0.0625,0.875,0.0,0.875,0.0625,0.8438,0.0625,0.8438,0.0,0.8438,0.0625,0.8125,0.0625,0.8125,0.0,0.8125,0.0625,0.7813,0.0625,0.7813,0.0,0.7813,0.0625,0.75,0.0625,0.75,0.0,0.75,0.0625,0.7188,0.0625,0.7188,0.0,0.7188,0.0625,0.6875,0.0625,0.6875,0.0,0.6875,0.0625,0.6563,0.0625,0.6563,0.0,0.6563,0.0625,0.625,0.0625,0.625,0.0,0.625,0.0625,0.5938,0.0625,0.5938,0.0,0.5938,0.0625,0.5625,0.0625,0.5625,0.0,0.5625,0.0625,0.5313,0.0625,0.5313,0.0,0.5313,0.0625,0.5,0.0625,0.5,0.0,0.5,0.0625,0.4688,0.0625,0.4688,0.0,0.4688,0.0625,0.4375,0.0625,0.4375,0.0,0.4375,0.0625,0.4063,0.0625,0.4063,0.0,0.4063,0.0625,0.375,0.0625,0.375,0.0,0.375,0.0625,0.3438,0.0625,0.3438,0.0,0.3438,0.0625,0.3125,0.0625,0.3125,0.0,0.3125,0.0625,0.2813,0.0625,0.2813,0.0,0.2813,0.0625,0.25,0.0625,0.25,0.0,0.25,0.0625,0.2188,0.0625,0.2188,0.0,0.2188,0.0625,0.1875,0.0625,0.1875,0.0,0.1875,0.0625,0.1563,0.0625,0.1563,0.0,0.1563,0.0625,0.125,0.0625,0.125,0.0,0.125,0.0625,0.0938,0.0625,0.0938,0.0,0.0938,0.0625,0.0625,0.0625,0.0625,0.0,0.0625,0.0625,0.0313,0.0625,0.0313,0.0,0.0313,0.0625,0.0,0.0625,0.0,0.0,0.6944,0.5,0.6944,0.5,0.75,0.5,0.75,0.5,0.75,0.5,0.6944,0.5,0.6389,0.5,0.6389,0.5,0.6944,0.5,0.6944,0.5,0.6944,0.5,0.6389,0.5,0.5833,0.5,0.5833,0.5,0.6389,0.5,0.6389,0.5,0.6389,0.5,0.5833,0.5,0.5278,0.5,0.5278,0.5,0.5833,0.5,0.5833,0.5,0.5833,0.5,0.5278,0.5,0.4722,0.5,0.4722,0.5,0.5278,0.5,0.5278,0.5,0.5278,0.5,0.4722,0.5,0.4167,0.5,0.4167,0.5,0.4722,0.5,0.4722,0.5,0.4722,0.5,0.4167,0.5,0.3611,0.5,0.3611,0.5,0.4167,0.5,0.4167,0.5,0.4167,0.5,0.3611,0.5,0.3056,0.5,0.3056,0.5,0.3611,0.5,0.3611,0.5,0.3611,0.5,0.3056,0.5,0.25,0.5,0.25,0.5,0.3056,0.5,0.3056,0.5,0.3056,0.5,0.25,0.5,0.1944,0.5,0.1944,0.5,0.25,0.5,0.25,0.5,0.25,0.5,0.1944,0.5,0.1389,0.5,0.1389,0.5,0.1944,0.5,0.1944,0.5,0.1944,0.5,0.1389,0.5,0.0833,0.5,0.0833,0.5,0.1389,0.5,0.1389,0.5,0.1389,0.5,0.0833,0.5,0.0278,0.5,0.0278,0.5,0.0833,0.5,0.0833,0.5,0.0833,0.5,0.0278,0.5,0.9722,0.5,0.9722,0.5,1.0278,0.5,1.0278,0.5,1.0278,0.5,0.9722,0.5,0.9167,0.5,0.9167,0.5,0.9722,0.5,0.9722,0.5,0.9722,0.5,0.9167,0.5,0.8611,0.5,0.8611,0.5,0.9167,0.5,0.9167,0.5,0.9167,0.5,0.8611,0.5,0.8056,0.5,0.8056,0.5,0.8611,0.5,0.8611,0.5,0.8611,0.5,0.8056,0.5,0.75,0.5,0.75,0.5,0.8056,0.5,0.8056,0.5,0.8056,0.5,0.75,0.5,0.1786,0.117,0.8611,0.0,0.8056,0.0,0.8056,0.0,0.329,0.0302,0.1786,0.117,0.067,0.25,0.9167,0.0,0.8611,0.0,0.8611,0.0,0.1786,0.117,0.067,0.25,0.0076,0.4132,0.9722,0.0,0.9167,0.0,0.9167,0.0,0.067,0.25,0.0076,0.4132,0.0076,0.5868,1.0278,0.0,0.9722,0.0,0.9722,0.0,0.0076,0.4132,0.0076,0.5868,0.067,0.75,0.0833,0.0,0.0278,0.0,0.0278,0.0,0.0076,0.5868,0.067,0.75,0.1786,0.883,0.1389,0.0,0.0833,0.0,0.0833,0.0,0.067,0.75,0.1786,0.883,0.329,0.9698,0.1944,0.0,0.1389,0.0,0.1389,0.0,0.1786,0.883,0.329,0.9698,0.5,1.0,0.25,0.0,0.1944,0.0,0.1944,0.0,0.329,0.9698,0.5,1.0,0.671,0.9698,0.3056,0.0,0.25,0.0,0.25,0.0,0.5,1.0,0.671,0.9698,0.8214,0.883,0.3611,0.0,0.3056,0.0,0.3056,0.0,0.671,0.9698,0.8214,0.883,0.933,0.75,0.4167,0.0,0.3611,0.0,0.3611,0.0,0.8214,0.883,0.933,0.75,0.9924,0.5868,0.4722,0.0,0.4167,0.0,0.4167,0.0,0.933,0.75,0.9924,0.5868,0.9924,0.4132,0.5278,0.0,0.4722,0.0,0.4722,0.0,0.9924,0.5868,0.9924,0.4132,0.933,0.25,0.5833,0.0,0.5278,0.0,0.5278,0.0,0.9924,0.4132,0.933,0.25,0.8214,0.117,0.6389,0.0,0.5833,0.0,0.5833,0.0,0.933,0.25,0.8214,0.117,0.671,0.0302,0.6944,0.0,0.6389,0.0,0.6389,0.0,0.8214,0.117,0.671,0.0302,0.5,0.0,0.75,0.0,0.6944,0.0,0.6944,0.0,0.671,0.0302,0.5,0.0,0.329,0.0302,0.8056,0.0,0.75,0.0,0.75,0.0,0.5,0.0,0.329,0.0302,0.75,0.0,0.75,0.5,0.6944,0.5,0.6944,0.5,0.6944,0.0,0.75,0.0,0.6944,0.0,0.6944,0.5,0.6389,0.5,0.6389,0.5,0.6389,0.0,0.6944,0.0,0.6389,0.0,0.6389,0.5,0.5833,0.5,0.5833,0.5,0.5833,0.0,0.6389,0.0,0.5833,0.0,0.5833,0.5,0.5278,0.5,0.5278,0.5,0.5278,0.0,0.5833,0.0,0.5278,0.0,0.5278,0.5,0.4722,0.5,0.4722,0.5,0.4722,0.0,0.5278,0.0,0.4722,0.0,0.4722,0.5,0.4167,0.5,0.4167,0.5,0.4167,0.0,0.4722,0.0,0.4167,0.0,0.4167,0.5,0.3611,0.5,0.3611,0.5,0.3611,0.0,0.4167,0.0,0.3611,0.0,0.3611,0.5,0.3056,0.5,0.3056,0.5,0.3056,0.0,0.3611,0.0,0.3056,0.0,0.3056,0.5,0.25,0.5,0.25,0.5,0.25,0.0,0.3056,0.0,0.25,0.0,0.25,0.5,0.1944,0.5,0.1944,0.5,0.1944,0.0,0.25,0.0,0.1944,0.0,0.1944,0.5,0.1389,0.5,0.1389,0.5,0.1389,0.0,0.1944,0.0,0.1389,0.0,0.1389,0.5,0.0833,0.5,0.0833,0.5,0.0833,0.0,0.1389,0.0,0.0833,0.0,0.0833,0.5,0.0278,0.5,0.0278,0.5,0.0278,0.0,0.0833,0.0,1.0278,0.0,1.0278,0.5,0.9722,0.5,0.9722,0.5,0.9722,0.0,1.0278,0.0,0.9722,0.0,0.9722,0.5,0.9167,0.5,0.9167,0.5,0.9167,0.0,0.9722,0.0,0.9167,0.0,0.9167,0.5,0.8611,0.5,0.8611,0.5,0.8611,0.0,0.9167,0.0,0.8611,0.0,0.8611,0.5,0.8056,0.5,0.8056,0.5,0.8056,0.0,0.8611,0.0,0.8056,0.0,0.8056,0.5,0.75,0.5,0.75,0.5,0.75,0.0,0.8056,0.0,0.671,0.0302,0.329,0.0302,0.5,0.0,0.671,0.0302,0.067,0.25,0.329,0.0302,0.671,0.0302,0.933,0.25,0.067,0.25,0.671,0.0302,0.8214,0.117,0.933,0.25,0.933,0.25,0.1786,0.883,0.067,0.25,0.933,0.25,0.8214,0.883,0.1786,0.883,0.933,0.25,0.9924,0.5868,0.8214,0.883,0.933,0.25,0.9924,0.4132,0.9924,0.5868,0.9924,0.5868,0.933,0.75,0.8214,0.883,0.8214,0.883,0.5,1.0,0.1786,0.883,0.8214,0.883,0.671,0.9698,0.5,1.0,0.5,1.0,0.329,0.9698,0.1786,0.883,0.1786,0.883,0.0076,0.5868,0.067,0.25,0.1786,0.883,0.067,0.75,0.0076,0.5868,0.0076,0.5868,0.0076,0.4132,0.067,0.25,0.067,0.25,0.1786,0.117,0.329,0.0302,0.5,0.0,0.8214,0.117,0.671,0.0302,0.5,0.0,0.9924,0.4132,0.8214,0.117,0.5,0.0,0.1786,0.117,0.9924,0.4132,0.5,0.0,0.329,0.0302,0.1786,0.117,0.1786,0.117,0.671,0.9698,0.9924,0.4132,0.1786,0.117,0.067,0.75,0.671,0.9698,0.1786,0.117,0.0076,0.4132,0.067,0.75,0.1786,0.117,0.067,0.25,0.0076,0.4132,0.0076,0.4132,0.0076,0.5868,0.067,0.75,0.067,0.75,0.329,0.9698,0.671,0.9698,0.067,0.75,0.1786,0.883,0.329,0.9698,0.329,0.9698,0.5,1.0,0.671,0.9698,0.671,0.9698,0.933,0.75,0.9924,0.4132,0.671,0.9698,0.8214,0.883,0.933,0.75,0.933,0.75,0.9924,0.5868,0.9924,0.4132,0.9924,0.4132,0.933,0.25,0.8214,0.117,0.75,1.0,0.6944,1.0,0.6944,0.5,0.6944,0.5,0.75,0.5,0.75,1.0,0.6944,1.0,0.6389,1.0,0.6389,0.5,0.6389,0.5,0.6944,0.5,0.6944,1.0,0.6389,1.0,0.5833,1.0,0.5833,0.5,0.5833,0.5,0.6389,0.5,0.6389,1.0,0.5833,1.0,0.5278,1.0,0.5278,0.5,0.5278,0.5,0.5833,0.5,0.5833,1.0,0.5278,1.0,0.4722,1.0,0.4722,0.5,0.4722,0.5,0.5278,0.5,0.5278,1.0,0.4722,1.0,0.4167,1.0,0.4167,0.5,0.4167,0.5,0.4722,0.5,0.4722,1.0,0.4167,1.0,0.3611,1.0,0.3611,0.5,0.3611,0.5,0.4167,0.5,0.4167,1.0,0.3611,1.0,0.3056,1.0,0.3056,0.5,0.3056,0.5,0.3611,0.5,0.3611,1.0,0.3056,1.0,0.25,1.0,0.25,0.5,0.25,0.5,0.3056,0.5,0.3056,1.0,0.25,1.0,0.1944,1.0,0.1944,0.5,0.1944,0.5,0.25,0.5,0.25,1.0,0.1944,1.0,0.1389,1.0,0.1389,0.5,0.1389,0.5,0.1944,0.5,0.1944,1.0,0.1389,1.0,0.0833,1.0,0.0833,0.5,0.0833,0.5,0.1389,0.5,0.1389,1.0,0.0833,1.0,0.0278,1.0,0.0278,0.5,0.0278,0.5,0.0833,0.5,0.0833,1.0,1.0278,1.0,0.9722,1.0,0.9722,0.5,0.9722,0.5,1.0278,0.5,1.0278,1.0,0.9722,1.0,0.9167,1.0,0.9167,0.5,0.9167,0.5,0.9722,0.5,0.9722,1.0,0.9167,1.0,0.8611,1.0,0.8611,0.5,0.8611,0.5,0.9167,0.5,0.9167,1.0,0.8611,1.0,0.8056,1.0,0.8056,0.5,0.8056,0.5,0.8611,0.5,0.8611,1.0,0.8056,1.0,0.75,1.0,0.75,0.5,0.75,0.5,0.8056,0.5,0.8056,1.0]}}}}],\\"name\\":\\"ModelFromPro\\",\\"mbb\\":[-1.2000000476837158,-1.1817690134048462,-1,1.2000000476837158,1.1817690134048462,2.2232010364532471],\\"pivotOffset\\":[0,0,0]},\\"materialDefinitions\\":{\\"0\\":{\\"type\\":\\"standard\\",\\"params\\":{\\"transparency\\":0,\\"diffuse\\":[1.0,1.0,1.0],\\"externalColorMixMode\\":\\"tint\\"}}},\\"textureDefinitions\\":{\\"20fd5397feabb33012b0b3b27f683e3e.dat\\":{\\"encoding\\":\\"data:image/jpeg\\",\\"channels\\":\\"rgb\\",\\"alphaChannelUsage\\":\\"transparency\\",\\"images\\":[{\\"size\\":128,\\"data\\":\\"/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCACAAIADASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwDjaKKK+ZP3AKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooA//Z\\"}]}}}"
    }
  ],
  "rGBColorProfile" : "sRGB IEC61966-2-1 noBPC",
  "cMYKColorProfile" : "U.S. Web Coated (SWOP) v2"
}"""
    tbxDir = os.path.dirname(sys.argv[0])
    # tbxDir = r"E:\南京工具\工具交活整理_0916\测试数据\工具3_三通四通\res2"
    lyrDir = os.path.join(tbxDir, "lyrxDir")
    lyrTxtFile = os.path.join(lyrDir, "myLyr.txt")
    lyrLyrFile = os.path.join(lyrDir, "myLyr.lyrx")

    # create temp lyr file
    if not os.path.exists(lyrDir):
        os.makedirs(lyrDir)

    # clear already exists files
    if os.path.exists(lyrTxtFile):
        os.remove(lyrTxtFile)

    if os.path.exists(lyrLyrFile):
        os.remove(lyrLyrFile)

    # write lyr txt file
    with open(lyrTxtFile, "w", encoding="utf-8") as f:
        f.write(data)

    os.rename(lyrTxtFile, lyrLyrFile)

    return lyrLyrFile


@logIt
def addNewGJField(pntFC, gj, alpha):
    _addField(pntFC, "GJ_NEW", "DOUBLE")
    codes = f"""def f(PNT_TYPE, gj_gj_gj):
    gj = {gj}
    alpha = {alpha}
    PNT_TYPE = int(PNT_TYPE)
    gj_gj_gj = float(gj_gj_gj) / gj
    if PNT_TYPE == 0:
        return gj_gj_gj * alpha
    else:
        return gj_gj_gj"""
    arcpy.CalculateField_management(pntFC, "GJ_NEW", "f(!PNT_TYPE_!, !gj_gj_gj_!)", "PYTHON3", codes)

    return pntFC


def f(PNT_TYPE, gj_gj_gj):
    gj = 400
    alpha = 0.6
    PNT_TYPE = int(PNT_TYPE)
    gj_gj_gj = float(gj_gj_gj) / gj
    if PNT_TYPE == 0:
        return gj_gj_gj * alpha
    else:
        return gj_gj_gj


@logIt
def convert3D(trueResPnt, resLyr, outputGDB):
    global resDataList
    tbxDir = os.path.dirname(sys.argv[0])

    lyrDataDir = os.path.join(tbxDir, "lyrData")
    if not os.path.exists(lyrDataDir):
        os.makedirs(lyrDataDir)

    _addMessage(tbxDir)
    _addMessage(lyrDataDir)

    _addField(trueResPnt, "_unic_id_", "TEXT")
    arcpy.CalculateField_management(trueResPnt, "_unic_id_", "!unic_id_!", "PYTHON3")

    lyrDataFile = os.path.join(lyrDataDir, os.path.basename(trueResPnt))

    if os.path.exists(lyrDataFile):
        os.remove(lyrDataFile)

    res = arcpy.SaveToLayerFile_management(resLyr, lyrDataFile)
    mul = os.path.join(outputGDB, os.path.basename(trueResPnt) + "_MUL")
    resMul = arcpy.Layer3DToFeatureClass_3d(res, mul, "_unic_id_")

    # Modified in 20201127 add field join
    arcpy.env.qualifiedFieldNames = False
    resJoinField = arcpy.JoinField_management(resMul, "_unic_id_", trueResPnt, "_unic_id_")
    mulRes = os.path.join(outputGDB, os.path.basename(trueResPnt)[:-4] + "_RES")
    resMulNew = arcpy.CopyFeatures_management(resJoinField, mulRes)
    mulResName = os.path.basename(trueResPnt)[:-4] + "_RES"
    resDataList.append(mulResName)
    # arcpy.env.qualifiedFieldNames = True
    arcpy.ClearEnvironment("qualifiedFieldNames")

    return resMulNew


def delTempData(outputGDB, resDataList):
    arcpy.env.workspace = outputGDB
    fcList = [each for each in arcpy.ListFeatureClasses() if each not in resDataList]
    arcpy.Delete_management(fcList)



@getRunTime
@logIt
def main(pnt, ply, outputPath, outputGDB, sr, gj, alpha):
    global pntFieldList, plyFieldList, plyTotalExtent, resLyrList, resMulLyrList

    _addMessage(f" === Progress Step2 Initial Field === ")
    logging.info(f" === Progress Step2 Initial Field === ")
    # delete the field will be added later, make sure the field "unic_id_" is the last in fields
    _deleteFields(pnt, "point")
    _deleteFields(ply, "line")
    arcpy.SetProgressorPosition()

    _addMessage(f" === Progress Step3 Get Pipe Size === ")
    logging.info(f" === Progress Step3 Get Pipe Size === ")
    # add a gj_num_ field to ply, it will be include into lineObject
    generateGJNumField(ply)
    arcpy.SetProgressorPosition()

    _addMessage(f" === Progress Step4 Convert Data To Xlsx === ")
    logging.info(f" === Progress Step4 Convert Data To Xlsx === ")
    # convert ply and pnt table to xlsx
    pntXls = _convertToXlsx(pnt, outputPath, "point", convertAll=True)
    plyXls = _convertToXlsx(ply, outputPath, "line", convertAll=True)
    arcpy.SetProgressorPosition()

    _addMessage(f" === Progress Step5 Generate Line Object === ")
    logging.info(f" === Progress Step5 Generate Line Object === ")
    # generate line object with spatial index
    plgObjList = _generateLineObj(plyXls, plyFieldList, plyTotalExtent)
    arcpy.SetProgressorPosition()

    _addMessage(f" === Progress Step6 Get Line Directory To Point === ")
    logging.info(f" === Progress Step6 Get Line Directory To Point === ")
    # add directory data to pnt xlsx
    pntxlsx = _getDirectory(pntXls, plgObjList, plyTotalExtent)
    arcpy.SetProgressorPosition()

    _addMessage(f" === Progress Step7 Add X, Y, Z Rotation === ")
    logging.info(f" === Progress Step7 Add X, Y, Z Rotation === ")
    # copy all datas to make feature class
    pnt_res_xlsx = copyPointsInXlsx(pntxlsx)
    arcpy.SetProgressorPosition()

    _addMessage(f" === Progress Step8 Generate Point Feature Class === ")
    logging.info(f" === Progress Step8 Generate Point Feature Class === ")
    # generate point feature class from xlsx
    resPnt = xlsxToPointFC(pnt_res_xlsx, sr, outputGDB)
    arcpy.SetProgressorPosition()

    _addMessage(f" === Progress Step9 Add A New Pipe Size Field === ")
    logging.info(f" === Progress Step9 Add A New Pipe Size Field === ")
    # add a new gj field to adjust pipe size
    trueResPnt = addNewGJField(resPnt, gj, alpha)
    arcpy.SetProgressorPosition()

    _addMessage(f" === Progress Step10 Generate Multipatch Modeling === ")
    logging.info(f" === Progress Step10 Generate Multipatch Modeling === ")
    # create lyrx file
    lyrxFile = createSymbolFile()
    arcpy.SetProgressorPosition()

    _addMessage(f" === Progress Step11 Add Data To ArcGIS Pro === ")
    logging.info(f" === Progress Step11 Add Data To ArcGIS Pro === ")
    # add data to aprx
    pntLyr = arcpy.MakeFeatureLayer_management(trueResPnt, os.path.basename(trueResPnt))
    arcpy.SetProgressorPosition()

    _addMessage(f" === Progress Step12 Apply Multipatch Model === ")
    logging.info(f" === Progress Step12 Apply Multipatch Model === ")
    # apply multipatch model
    resLyr = arcpy.ApplySymbologyFromLayer_management(pntLyr, lyrxFile,
                                                      [["VALUE_FIELD", "PNT_TYPE_", "PNT_TYPE_"]], "MAINTAIN")
    arcpy.SetProgressorPosition()

    _addMessage(f" === Progress Step13 Convert To Multipatch === ")
    logging.info(f" === Progress Step13 Convert To Multipatch === ")
    # convert to multipatch
    resMul = convert3D(trueResPnt, resLyr, outputGDB)
    arcpy.SetProgressorPosition()

    resMulLyrList.append(resMul)
    resLyrList.append(resLyr)

    logging.info(f" === Progress Step14 Clear Temp Xlsx")
    try:
        os.remove(pntXls)
        os.remove(plyXls)
        os.remove(pntxlsx)
        # os.remove(pnt_res_xlsx)
    except:
        logging.info(f" +++ Delete Temp Xlsx Faild +++ ")


spatialIndex = (1, 1)
pntFieldList = []
plyFieldList = []
resLyrList = []
resMulLyrList = []
plyTotalExtent = None
tolerance = 0.1
resDataList = []

logging.info(f" === Common Para: \n == spatialIndex: {spatialIndex} \n"
             f" == tolerance: {tolerance} \n")

# inGDB = r"E:\南京工具\工具交活整理_0916\测试数据\工具3_三通四通\二维测试数据.gdb"
# outputGDB = r"E:\南京工具\工具交活整理_0916\测试数据\工具3_三通四通\res2\管点数据.gdb"
# outputPath = r"E:\南京工具\工具交活整理_0916\测试数据\工具3_三通四通\res2"
# sr = arcpy.SpatialReference(4549)
# tolerance = 0.1

inGDB = arcpy.GetParameterAsText(0)
outputGDB = arcpy.GetParameterAsText(1)
outputPath = arcpy.GetParameterAsText(2)
sr = arcpy.GetParameterAsText(3)
tolerance = float(arcpy.GetParameterAsText(4))
gj = float(arcpy.GetParameterAsText(5))
alpha = float(arcpy.GetParameterAsText(6))

_addMessage(f" === Input Para: \n == inGDB: {inGDB} \n == outputGDB: {outputGDB} \n"
            f" == outputPath: {outputPath} \n == sr: {sr} \n")
logging.info(f" === Input Para: \n == inGDB: {inGDB} \n == outputGDB: {outputGDB} \n"
             f" == outputPath: {outputPath} \n == sr: {sr} \n")

_addMessage(f" === Progress Step1 Select target data from origin gdb to target gdb ===")
logging.info(f" === Progress Step1 Select target data from origin gdb to target gdb ===")
resdataTupleList = selectTargetData(inGDB, outputGDB)

_addMessage(f" === Progress Step1 Finished, the data selected : {resdataTupleList} ===")
logging.info(f" === Progress Step1 Finished, the data selected : {resdataTupleList} ===")

totalNum = len(resdataTupleList)

arcpy.SetProgressor("step", "Data Progressing It Will Take A Long Time, Please Wait...",
                    1, (totalNum + 1) * 13 + 1)

_addMessage(" === Start Looping Data Selected ===")
logging.info(" === Start Looping Data Selected ===")
for eachtup in resdataTupleList:
    _addMessage(f"{resdataTupleList.index(eachtup) + 1}/{totalNum}")
    logging.info(f"Now Is Progressing data: {eachtup}, "
                 f"{resdataTupleList.index(eachtup) + 1}/{totalNum}")
    pntData, plyData = eachtup
    main(pntData, plyData, outputPath, outputGDB, sr, gj, alpha)

    arcpy.SetProgressorPosition()

_addMessage(f" === Progress Final Add Data To ArcGIS Pro === ")
logging.info(f" === Progress Final Add Data To ArcGIS Pro === ")
lyr = arcpy.SetParameter(7, resLyrList)
mulLyr = arcpy.SetParameter(8, resMulLyrList)

delTempData(outputGDB, resDataList)