import psycopg2
import openpyxl
import os


def getTabTitleMapping(xlsx):
    wb = openpyxl.load_workbook(xlsx)
    sht = wb.active

    maxCol = sht.max_column
    titleList = []
    for i in range(1, maxCol + 1):
        data = sht.cell(1, i).value
        titleList.append(data)

    # print(titleList)

    return titleList


def generateFieldTypeSQLExp(fieldList, tableFieldType):
    sqlExp = ""
    for i, eachField in enumerate(fieldList):
        if eachField == "SMID":
            sqlExp += "(SMID INTEGER PRIMARY KEY, "
        else:
            # the last field
            if i == len(fieldList) - 1:
                sqlExp += f"{eachField} {tableFieldType[eachField]});"
            else:
                sqlExp += f"{eachField} {tableFieldType[eachField]}, "

    print(sqlExp)
    return sqlExp


def createPGTable(host, port, username, password, dbname, tableName, fieldsTupString):
    # conncet to pg
    conn = psycopg2.connect(host=host, port=port, user=username, password=password, dbname=dbname)
    cur = conn.cursor()

    cur.execute(f"create table if not exists {tableName}{fieldsTupString};")

    conn.commit()
    conn.close()


def compareXlsxTitle(xlsxDir, tarHeaderList):
    # get all excel files
    xlsxFileList = [os.path.join(root, file) for root, dir, files in os.walk(xlsxDir) for file in files if
                    file[-5:] == ".xlsx" and file[:2] != "~$"]
    # print(xlsxFileList)

    for eachXlsx in xlsxFileList:
        xlsxHeader = getTabTitleMapping(eachXlsx)

        if len(xlsxHeader) == len(tarHeaderList):
            for each in tarHeaderList:
                xlsxHeader.remove(each)


def replaceXlsxHeader(xlsx, xlsxHeaderMapping, outputPath):
    wb = openpyxl.load_workbook(xlsx)
    sht = wb.active

    maxCol = sht.max_column
    for i in range(1, maxCol + 1):
        data = sht.cell(1, i).value
        print(data)

        # replace xlsx header to formatted name
        sht.cell(1, i).value = xlsxHeaderMapping[data]

    # get the name of xlsx
    outputName = os.path.basename(xlsx)

    # save the last directory of each excel file
    lastDir = os.path.basename(os.path.dirname(xlsx))

    if not os.path.exists(os.path.join(outputPath, lastDir)):
        os.makedirs(os.path.join(outputPath, lastDir))

    wb.save(os.path.join(outputPath, lastDir, outputName))


def repAllXlsxHeader(xlsxDir, xlsxHeaderMapping, outputPath):
    # get all excel files
    xlsxFileList = [os.path.join(root, file) for root, dir, files in os.walk(xlsxDir) for file in files if
                    file[-5:] == ".xlsx" and file[:2] != "~$"]
    # print(xlsxFileList)

    for eachXlsx in xlsxFileList:
        replaceXlsxHeader(eachXlsx, xlsxHeaderMapping, outputPath)


def createAndInsertValuesToTable(host, port, username, password, dbname, xlsxDir, fieldsTupString):
    # get excel files
    xlsxFileList = [os.path.join(root, file) for root, dir, files in os.walk(xlsxDir) for file in files if
                    file[-5:] == ".xlsx" and file[:2] != "~$"]

    # process each excel file
    for eachXlsx in xlsxFileList:
        # get the name of excel file
        tableName = os.path.splitext(os.path.basename(eachXlsx))[0]

        createPGTable(host, port, username, password, dbname, tableName, fieldsTupString)


        conn = psycopg2.connect(host=host, port=port, user=username, password=password, dbname=dbname)
        cur = conn.cursor()

        # read each row of excel and insert into pg table
        wb = openpyxl.load_workbook(eachXlsx)
        sht = wb.active

        for rowNum, eachRow in enumerate(sht.rows):
            if rowNum == 0:
                continue

            # format all type to str
            data = tuple([str(eachCell.value) for eachCell in eachRow])

            # insert into values
            cur.execute(f"insert into {tableName} values{data};")
            conn.commit()

        cur.execute(f"create unique index index_{tableName}_smid on {tableName}(smid);")
        conn.commit()

        conn.close()


def main(xlsxDir, xlsxHeaderMapping, outputPath, host, port, username, password, dbname):
    # format the header of each excel file
    repAllXlsxHeader(xlsxDir, xlsxHeaderMapping, outputPath)

    # generate the sql expression of table fields
    fieldSQLExp = generateFieldTypeSQLExp(fieldList, tableFieldType)
    # print(fieldSQLExp)

    # create table and insert values
    createAndInsertValuesToTable(host, port, username, password, dbname, xlsxDir, fieldSQLExp)


# tarHeaderList_bak = ['SmID', 'SmSdriW', 'SmSdriN', 'SmSdriE', 'SmSdriS', 'SmUserID', 'SmLibTileID', 'SmGeometrySize',
#                  'SmGeoPosition', 'Field_SmUserID', '模型名称', '图层名称', '父对象', '所在组', 'ID', '市', '区县', '街道']

# tarHeaderList = ['SmID', 'SmSdriW', 'SmSdriN', 'SmSdriE', 'SmSdriS', 'SmUserID', 'SmLibTileID', 'SmGeometrySize',
#                  'SmGeoPosition', 'Field_SmUserID', '模型名称', '图层名称', '父对象', '所在组', 'ID', '市', '区县', '街道']
#
# repHeaderList = ['SMID', 'SMSDRIW', 'SMSDRIN', 'SMSDRIE', 'SMSDRIS', 'SMUSERID', 'SMLIBTILEID', 'SMGEOMETRYSIZE',
#                  'SMGEOPOSITION', 'FIELD_SMUSERID', 'MODELNAME', 'LAYERNAME', 'FATHEROBJECT', 'GROUP', 'ID', 'XZQ_SQ',
#                  'XZQ_QX', 'XZQ_JD']
#
# fieldTypeList = ['INTEGER', 'NUMERIC', 'NUMERIC', 'NUMERIC', 'NUMERIC', 'INTEGER', 'INTEGER', 'INTEGER', 'BIGINT',
#                  'INTEGER', 'VARCHAR(30)', 'VARCHAR(50)', 'VARCHAR(50)', 'VARCHAR(50)', 'INTEGER', 'VARCHAR(20)',
#                  'VARCHAR(20)', 'VARCHAR(20)']

fieldList = ['SMID', 'SMSDRIW', 'SMSDRIN', 'SMSDRIE', 'SMSDRIS', 'SMUSERID', 'SMLIBTILEID', 'SMGEOMETRYSIZE',
             'SMGEOPOSITION', 'FIELD_SMUSERID', 'MODELNAME', 'LAYERNAME', 'FATHEROBJECT', 'GROUP_', 'ID', 'XZQ_SQ',
             'XZQ_QX', 'XZQ_JD']

xlsxHeaderMapping = {'SmID': 'SMID', 'SmSdriW': 'SMSDRIW', 'SmSdriN': 'SMSDRIN', 'SmSdriE': 'SMSDRIE',
                     'SmSdriS': 'SMSDRIS', 'SmUserID': 'SMUSERID', 'SmLibTileID': 'SMLIBTILEID',
                     'SmGeometrySize': 'SMGEOMETRYSIZE', 'SmGeoPosition': 'SMGEOPOSITION',
                     'Field_SmUserID': 'FIELD_SMUSERID', '模型名称': 'MODELNAME', '图层名称': 'LAYERNAME',
                     '父对象': 'FATHEROBJECT', '所在组': 'GROUP_', 'ID': 'ID', '市': 'XZQ_SQ', '区县': 'XZQ_QX',
                     '街道': 'XZQ_JD'}

tableFieldType = {'SMID': 'INTEGER', 'SMSDRIW': 'NUMERIC', 'SMSDRIN': 'NUMERIC', 'SMSDRIE': 'NUMERIC',
                  'SMSDRIS': 'NUMERIC', 'SMUSERID': 'INTEGER', 'SMLIBTILEID': 'INTEGER', 'SMGEOMETRYSIZE': 'INTEGER',
                  'SMGEOPOSITION': 'BIGINT', 'FIELD_SMUSERID': 'INTEGER', 'MODELNAME': 'VARCHAR(30)',
                  'LAYERNAME': 'VARCHAR(50)', 'FATHEROBJECT': 'VARCHAR(50)', 'GROUP_': 'VARCHAR(50)', 'ID': 'INTEGER',
                  'XZQ_SQ': 'VARCHAR(20)', 'XZQ_QX': 'VARCHAR(20)', 'XZQ_JD': 'VARCHAR(20)'}

# dir = r"F:\工作项目\项目_温州超图\数据库_excel入库\原始数据\20201229\分区"
# outputPath = r"F:\工作项目\项目_温州超图\数据库_excel入库\处理数据\20201229"

dir = r"F:\工作项目\项目_温州超图\数据库_excel入库\20200125鹿城区精模\鹿城区六要素属性表"
outputPath = r"F:\工作项目\项目_温州超图\数据库_excel入库\20200125鹿城区精模\鹿城区六要素属性表_表头格式化"
# host = "192.168.2.40"
# port = "5432"
# username = "postgres"
# password = "123456*a"
# dbname = "ZGJDB"

# # 本机虚拟机pg库
# host = "192.168.31.128"
# port = "5432"
# username = "postgres"
# password = "0403"
# dbname = "ZGJDB"

# 城安pg库
host = "192.168.10.150"
port = "5432"
username = "postgres"
password = "cadsj123456"
dbname = "ZGJDB"


main(dir, xlsxHeaderMapping, outputPath, host, port, username, password, dbname)
