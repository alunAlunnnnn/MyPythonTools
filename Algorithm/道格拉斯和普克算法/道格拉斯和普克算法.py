import math
import os
import arcpy
import openpyxl
import functools
import datetime
import sys
import sqlite3

arcpy.env.overwriteOutput = True
sys.setrecursionlimit(100000)


# ---------- line class ----------
# ---------- show message in toolbox ----------

def _addMessage(mes):
    print(mes)
    arcpy.AddMessage(mes)


def _addWarning(mes):
    print(mes)
    arcpy.AddWarning(mes)


def _addError(mes):
    print(mes)
    arcpy.AddError(mes)


# points type is not tuple
class pointError(Exception):
    pass


class calKError(Exception):
    pass


class GenerateSpatialIndexError(Exception):
    pass


class lineEquation:
    """
    usage: generate a line object
    :param args: accept as
     --- (fp_x, fp_y, fp_z), (lp_x, lp_y, lp_z), (ex_xin, ex_ymin, ex_xmax, ex_ymax)
    firstPoint --- fp
    lastPoint --- lp
    extent --- ex
    """

    def __init__(self, *args):
        # save all points
        self.points = []

        for each in args[:2]:
            if not isinstance(each, tuple):
                _addMessage('Point coord is not tuple type')
                raise pointError

            self.points.append((float(each[0]), float(each[1]), float(each[2])))

        self.extent_xmin = args[2][0]
        self.extent_ymin = args[2][1]
        self.extent_xmax = args[2][2]
        self.extent_ymax = args[2][3]
        self.extent = args[2]

        # get point number, start with 1
        self.pntNum = len(args)

        # init pipe size
        self.pipeSize = None

        # set coord of start point and finish point
        self.x1 = self.points[0][0]
        self.y1 = self.points[0][1]
        self.z1 = self.points[0][2]
        self.x2 = self.points[-1][0]
        self.y2 = self.points[-1][1]
        self.z2 = self.points[-1][2]

        # extent available detect
        self._extentAvailableDetect()

        self.spaindex = None
        self.spaindex_row = None
        self.spaindex_col = None
        self.spaind_totalext = None

        self.calculateK_xy()
        self.calculateB_xy()

        self.calculateK_yz()
        self.calculateB_yz()

        self.calculateK_xz()
        self.calculateB_xz()

        self.generateEquation()

    # calculate k --- ( y2 - y1 ) / ( x2 - x1 )
    def calculateK_xy(self):
        if self.x1 == self.x2:
            _addWarning('ERROR --- calculate k_xy faild,'
                        ' x1 is equal to x2, x1 is {}. x2 is {}'.format(self.x1, self.x2))
            # in math k is infinity / -infinity
            self.k_xy = -999
            return self
            # raise calKError
        k = (self.y2 - self.y1) / (self.x2 - self.x1)
        self.k_xy = k
        return self

    # calculate b --- y1 - k * x1
    def calculateB_xy(self):
        if self.k_xy == -999:
            self.b_xy = -999
        else:
            b = self.y1 - self.k_xy * self.x1
            self.b_xy = b
        return self

    # calculate k --- ( z2 - z1 ) / ( y2 - y1 )
    def calculateK_yz(self):
        if self.y1 == self.y2:
            _addWarning('ERROR --- calculate k_yz faild, y1 is equal to y2. y1 is %s' % self.y1)
            self.k_yz = -999
            return self
            # raise calKError
        k = (self.z2 - self.z1) / (self.y2 - self.y1)
        self.k_yz = k
        return self

    # calculate b --- z1 - k * y1
    def calculateB_yz(self):
        if self.k_yz == -999:
            self.b_yz = -999
        else:
            b = self.z1 - self.k_yz * self.y1
            self.b_yz = b
        return self

    # calculate k --- ( z2 - z1 ) / ( y2 - y1 )
    def calculateK_xz(self):
        if self.x1 == self.x2:
            _addWarning('ERROR --- calculate k_xz faild, x1 is equal to x2. x1 is %s' % self.x1)
            self.k_xz = -999
            return self
            # raise calKError
        k = (self.z2 - self.z1) / (self.x2 - self.x1)
        self.k_xz = k
        return self

    # calculate b --- z1 - k * y1
    def calculateB_xz(self):
        if self.k_xz == -999:
            self.b_xz = -999
        else:
            b = self.z1 - self.k_xz * self.x1
            self.b_xz = b
        return self

    # generate function equation
    def generateEquation(self):
        self.euqation_xy = '%s * x + %s' % (self.k_xy, self.b_xy)
        self.euqation_yz = '%s * x + %s' % (self.k_yz, self.b_yz)
        self.euqation_xz = '%s * x + %s' % (self.k_xz, self.b_xz)
        return self

    # calculate the intersect point
    def calculateIntersect(self, otherLineObj):
        if self.k_xy == otherLineObj.k_xy:
            self.intersect = 'false'
            otherLineObj.intersect = 'false'
            return None

        if self.b_xy == otherLineObj.b_xy:
            x = 0
            y = self.b_xy
        else:
            x = (otherLineObj.b_xy - self.b_xy) / (self.k_xy - otherLineObj.k_xy)
            y = self.k_xy * x + self.b_xy

        # detect whether the point in the rectangle of self line
        if x > self.extent_xmin and x < self.extent_xmax:
            if y > self.extent_ymin and y < self.extent_ymax:
                self.intersect = 'true'
            else:
                self.intersect = 'false'
        else:
            self.intersect = 'false'

        # detect whether the point in the rectangle of another line
        if x > otherLineObj.extent_xmin and x < otherLineObj.extent_xmax:
            if y > otherLineObj.extent_ymin and y < otherLineObj.extent_ymax:
                otherLineObj.intersect = 'true'
            else:
                otherLineObj.intersect = 'false'
        else:
            otherLineObj.intersect = 'false'

        return x, y

    # calculate z value in intersect x,y
    def calculateZCoord_yz(self, y):
        z = self.k_yz * y + self.b_yz
        return z

    # calculate z value in intersect x,y
    def calculateZCoord_xz(self, x):
        z = self.k_xz * x + self.b_xz
        return z

    # make sure the x_min, y_min is less or equal than x_max, y_max in extent
    def _extentAvailableDetect(self):
        assert int(self.extent_xmin * 10 ** 8) <= int(
            self.extent_xmax * 10 ** 8), "Error --- Extent of line object is not available"
        assert int(self.extent_ymin * 10 ** 8) <= int(
            self.extent_ymax * 10 ** 8), "Error --- Extent of line object is not available"

    # detect the point can touch the line with a tolerance or not
    def pointTouchDet(self, pnt, tolerance, totalExtent):
        """
        usage: used to detect that, the point is in the line with a tolerance num.
                before use this, please call the function "generateSpatialIndex()" yet
        :param pnt:(x, y)
        :param tolerance:
        :return:
        """
        pnt_x, pnt_y = pnt[0], pnt[1]
        pnt_index = None
        spatialIndex_x = 1
        spatialIndex_y = 1
        # generate the spatial index for point
        spaindex_step_x = round(float((totalExtent[2] - totalExtent[0]) / spatialIndex_x), 6) + 0.0001
        spaindex_step_y = round(float((totalExtent[3] - totalExtent[1]) / spatialIndex_y), 6) + 0.0001
        total_ext_ymax = totalExtent[3]
        total_ext_xmax = totalExtent[2]

        # detect whether is out of the ply's extent
        if (self.extent_xmax > totalExtent[2] + 0.0001 or self.extent_ymax > totalExtent[3] + 0.0001
                or self.extent_xmin < totalExtent[0] - 0.0001 or self.extent_ymin < totalExtent[1] - 0.0001):
            if self.extent_xmax > totalExtent[2] + 0.0001:
                print("pnt xxxx1xxxx pnt")
            if self.extent_ymax > totalExtent[3] + 0.0001:
                print("pnt xxxx2xxxx pnt")
            if self.extent_xmin < totalExtent[0] - 0.0001:
                print("pnt xxxx3xxxx pnt")
            if self.extent_ymin < totalExtent[1] - 0.0001:
                print("pnt xxxx4xxxx pnt")
            _addError("Error --- generate spatial index for point failed, "
                      "the point is ({}, {})".format(pnt_x, pnt_y))
            _addError("total extent is {}".format(totalExtent))
            raise GenerateSpatialIndexError

        find_key = False
        # for i in range(1, 11):
        for i in range(1, spatialIndex_y + 1):
            # if spatial index has finded, break the loop
            if find_key:
                break
            if pnt_y >= total_ext_ymax - i * spaindex_step_y:
                pnt_index_y = str(i)
                # for j in range(1, 11):
                for j in range(1, spatialIndex_x + 1):
                    if pnt_x >= total_ext_xmax - j * spaindex_step_x:
                        pnt_index_x = str(j)
                        pnt_index = (str(i) + "," + str(j))
                        find_key = True
                        break

        # no spatial index, no calculate
        assert pnt_index, "there are no spatial index in point"

        # match spatial index
        if pnt_index == self.spaindex:
            # point is in the extent of polyline
            if (self.extent_xmin - tolerance <= pnt_x <= self.extent_xmax + tolerance
                    and self.extent_ymin - tolerance <= pnt_y <= self.extent_ymax + tolerance):
                # point is in the extent of polyline
                # detect whether the point is on the line

                # vertical line
                if self.k_xy == -999:
                    if min(self.x1, self.x2) - tolerance <= pnt_x <= max(self.x1, self.x2) + tolerance:
                        if min(self.y1, self.y2) - tolerance <= pnt_y <= max(self.y1, self.y2) + tolerance:
                            return True
                        else:
                            return False
                    else:
                        return False
                # horizon line
                elif self.k_xy == 0:
                    if min(self.y1, self.y2) - tolerance <= pnt_y <= max(self.y1, self.y2) + tolerance:
                        if min(self.x1, self.x2) - tolerance <= pnt_x <= max(self.x1, self.x2) + tolerance:
                            return True
                        else:
                            return False
                    else:
                        return False
                # calculate the point in line with k_xy
                else:
                    det_y = self.k_xy * pnt_x + self.b_xy
                    if det_y - tolerance <= pnt_y <= det_y + tolerance:
                        return True
                    else:
                        return False
            # point is out of ply's extent
            else:
                return False
        # the spatial index between point and ply is not equal
        else:
            return False

    def generateSpatialIndex(self, totalExtent):
        """
        usage: generate spatial index, now is generate 10*10 grid index.

        index number as this:
         ----------
        |1,10|...|1,2|1,1|
         ----------
        |2,10|...|2,2|2,1|
         ----------
        |...........|
        |.          |
        |.          |
         -----------
        |10,10|...|10,2|10,1|

        :param totalExtent: (xmin, ymin, xmax, ymax)
        :return:
        """
        spatialIndex_x = 1
        spatialIndex_y = 1
        spaindex_step_x = round(float((totalExtent[2] - totalExtent[0]) / spatialIndex_x), 6) + 0.0001
        spaindex_step_y = round(float((totalExtent[3] - totalExtent[1]) / spatialIndex_y), 6) + 0.0001
        total_ext_ymax = totalExtent[3]
        total_ext_xmax = totalExtent[2]

        self.spaind_totalext = totalExtent

        if (self.extent_xmax > totalExtent[2] + 0.0001 or self.extent_ymax > totalExtent[3] + 0.0001
                or self.extent_xmin < totalExtent[0] - 0.0001 or self.extent_ymin < totalExtent[1] - 0.0001):
            if self.extent_xmax > totalExtent[2] + 0.0001:
                print("xxxx1xxxx")
            if self.extent_ymax > totalExtent[3] + 0.0001:
                print("xxxx2xxxx")
            if self.extent_xmin < totalExtent[0] - 0.0001:
                print("xxxx3xxxx")
            if self.extent_ymin < totalExtent[1] - 0.0001:
                print("xxxx4xxxx")
            _addError("Error --- generate spatial index failed, "
                      "line object's extent is not in total extent. "
                      "the line's first point is ({}, {})".format(self.x1, self.y1))
            _addError("total extent is {}".format(totalExtent))
            _addError("ply extent is {}".format(self.extent))
            raise GenerateSpatialIndexError

        find_key = False
        # for i in range(1, 11):
        for i in range(1, spatialIndex_y + 1):
            # if spatial index has finded, break the loop
            if find_key:
                break
            if self.extent_ymax >= total_ext_ymax - i * spaindex_step_y:
                self.spaindex_row = str(i)
                # for j in range(1, 11):
                for j in range(1, spatialIndex_y + 1):
                    if self.extent_xmax >= total_ext_xmax - j * spaindex_step_x:
                        self.spaindex_col = str(j)
                        self.spaindex = (str(i) + "," + str(j))
                        find_key = True
                        break
        return self

    def setPipeSize(self, pipesize):
        if isinstance(pipesize, int) or isinstance(pipesize, float):
            self.pipeSize = pipesize
        else:
            _addWarning("Warning --- pipe size is not a number type, pipe size init failed")
            self.pipeSize = None

    def calDisFromPnt(self, firstPoint):
        """
        usage: 使用第一个点坐标及 k 、 b 值来生成 line 对象
        :param firstPoint: Tuple --- 点坐标（x, y, z）
        :return: Double --- 距离值
        """
        x, y, z = firstPoint[0], firstPoint[1], firstPoint[2]

        # get k value
        originK = self.k_xy
        if originK == -999:
            k = 0
        elif -0.0001 <= originK <= 0.0001:
            k = -999
        else:
            k = 1 / originK

        # origin line is y = n
        if k == -999:
            # x = n (k is infinity)
            originY = self.k_xy * x + self.b_xy
            dis = abs(y - originY)
        # origin line is x = n (k is -999, b is -999)
        elif k == 0:
            dis = abs(x - self.extent_xmin)
        else:
            b = y - k * x
            calX = self.extent_xmin
            calY = k * calX + b

            ext = (min(x, calX), min(y, calY), max(x, calX), max(y, calY))

            newlineObj = lineEquation((x, y, z), (calX, calY, z), ext)
            interX, interY = self.calculateIntersect(newlineObj)
            dis = math.sqrt((y - interY) ** 2 + (x - interX) ** 2)

        return dis


def getRunTime(func):
    @functools.wraps(func)
    def _wrapper(*args, **kwargs):
        start = datetime.datetime.now()
        print(f"Start function '{func.__name__}' at : {start}")
        res = func(*args, **kwargs)
        end = datetime.datetime.now()
        print("*" * 30)
        print(f"Start function '{func.__name__}' at : {start}")
        print(f"Finish function '{func.__name__}' at : {end}")
        print(f"Function '{func.__name__}' total cost  at : {end - start}")
        print("*" * 30)
        return res

    return _wrapper


# 给点数据添加坐标信息
def addCoordField(inFC):
    fieldName = ["id_", "x_cen_", "y_cen_", "z_cen_"]
    fieldType = ["LONG", "DOUBLE", "DOUBLE", "DOUBLE"]
    fieldExp = ["f()", "!shape.centroid.X!", "!shape.centroid.Y!", "!shape.centroid.Z!"]
    codes = """a = -1
def f():
    global a
    a += 1
    return a"""

    z_index = fieldName.index("z_cen_")
    id_index = fieldName.index("id_")
    # 读取shp数据
    baseName = os.path.basename(inFC)
    dirName = os.path.dirname(inFC)
    if not arcpy.env.workspace:
        arcpy.env.workspace = dirName

    desc = arcpy.Describe(inFC)

    for i, eachField in enumerate(fieldName):
        print(eachField)
        try:
            arcpy.AddField_management(inFC, eachField, fieldType[i], field_is_nullable=True)
        except:
            arcpy.DeleteField_management(inFC, eachField)
            arcpy.AddField_management(inFC, eachField, fieldType[i], field_is_nullable=True)

        if i == id_index:
            arcpy.CalculateField_management(inFC, eachField, fieldExp[i], "PYTHON3", codes)
        else:
            if desc.hasZ:
                arcpy.CalculateField_management(inFC, eachField, fieldExp[i], "PYTHON3")
            else:
                if i != z_index:
                    arcpy.CalculateField_management(inFC, eachField, fieldExp[i], "PYTHON3")
                else:
                    arcpy.CalculateField_management(inFC, eachField, "0", "PYTHON3")


@getRunTime
def featureAttrToXlsx(inFC, outxlsx):
    # 添加坐标信息
    addCoordField(inFC)

    wb = openpyxl.Workbook()
    sht = wb.active

    # 读取shp数据
    baseName = os.path.basename(inFC)
    dirName = os.path.dirname(inFC)
    fieldList = [eachField.name for eachField in arcpy.ListFields(inFC) if eachField.type != "OID" and
                 eachField.name != "Shape_Leng" and eachField.name != "Shape_Area"
                 and eachField.name != "Shape"]
    print(fieldList)
    fieldLength = len(fieldList)

    for i in range(1, fieldLength + 1):
        sht.cell(1, i).value = fieldList[i - 1]

    with arcpy.da.SearchCursor(inFC, fieldList) as cur:
        rowNum = 1
        for row in cur:
            rowNum += 1
            for i in range(fieldLength):
                colNum = i + 1
                data = row[i]
                try:
                    sht.cell(rowNum, colNum).value = data
                except:
                    sht.cell(rowNum, colNum).value = str(data)
    wb.save(outxlsx)


def readDataFromXlsx(inxlsx):
    wb = openpyxl.load_workbook(inxlsx)
    sht = wb.active

    maxRow = sht.max_row
    maxCol = sht.max_column

    titleList = []
    for i in range(1, maxCol + 1):
        titleList.append(sht.cell(1, i).value)

    index_x, index_y, index_z = (titleList.index("x_cen_"),
                                 titleList.index("y_cen_"), titleList.index("z_cen_"))

    resList = []
    for i, eachRow in enumerate(sht.rows):
        i += 1
        if i == 1:
            continue

        x, y, z = eachRow[index_x].value, eachRow[index_y].value, eachRow[index_z].value
        resList.append([x, y, z])

    return resList


def writeDataToDB(pntList, db, table):
    conn = sqlite3.connect(db)
    c = conn.cursor()
    c.execute(f"DROP TABLE IF EXISTS {table}")
    c.execute(f"CREATE TABLE IF NOT EXISTS {table}(X real, Y real, Z real); ")
    c.executemany("INSERT INTO pntcoord VALUES(?, ?, ?);", pntList)

    conn.commit()
    conn.close()

    return db


def readDataFromDB(db, table):
    conn = sqlite3.connect(db)
    c = conn.cursor()
    data = c.execute(f"SELECT * FROM {table}").fetchall()
    print("db data is : ", data)

    return data


def DP(pntList, tolerance):
    print(pntList)
    """
    :param pntList: [(x1, y1, z1), (x2, y2, z2), (x3, y3, z3), ....]
    :return:
    """
    global resList
    if len(pntList) > 2:

        x_f, y_f, z_f = (pntList[0])
        x_l, y_l, z_l = (pntList[-1])
        ext = (min(x_f, x_l), min(y_f, y_l), max(x_f, x_l), max(y_f, y_l))

        # 实例化起点到终点的线
        line = lineEquation((x_f, y_f, z_f), (x_l, y_l, z_l), ext)

        # 获取每个点到线的距离
        pntNum = len(pntList)

        for eachPnt in pntList[1:-1]:
            x, y, z = eachPnt
            dis = line.calDisFromPnt((x, y, z))
            # 筛出点到线距离小于容差的点集
            if dis < tolerance:
                # 遍历结束后，将此列表中的所有点位全部清除
                pntList.remove(eachPnt)

        maxDis = 0
        maxDisPntIndex = 0
        for i, eachPnt in enumerate(pntList[1:-1]):
            x, y, z = eachPnt
            dis = line.calDisFromPnt((x, y, z))
            if dis > maxDis:
                maxDis = dis
                maxDisPntIndex = i

        # 找到距离最远的点，将线拆分为两条
        pntList1 = pntList[:maxDisPntIndex + 1]
        pntList2 = pntList[maxDisPntIndex:]
        if len(pntList1) > 2:
            DP(pntList1, tolerance)
        else:
            resList = resList + pntList1
            return resList

        if len(pntList2) > 2:
            DP(pntList2, tolerance)
        else:
            resList = resList + pntList2
            return resList
    else:
        resList = resList + pntList
        return resList


@getRunTime
def main(inFC, outxlsx, tolerance, outdb, table):
    featureAttrToXlsx(inFC, outxlsx)
    pntList = readDataFromXlsx(outxlsx)
    writeDataToDB(pntList, outdb, table)
    pntDataList = readDataFromDB(outdb, table)
    DP(pntDataList, tolerance)


data = r"E:\GIS算法\道格拉斯和普克算法\测试数据\路径点.shp"
outxlsx = r"E:\GIS算法\道格拉斯和普克算法\测试数据\测试excel.xlsx"
outdb = r"E:\GIS算法\道格拉斯和普克算法\测试数据\DPTest.db"
table = "pntcoord"
tolerance = 0.0000000000001
resList = []

main(data, outxlsx, tolerance, outdb, table)

print(resList)
