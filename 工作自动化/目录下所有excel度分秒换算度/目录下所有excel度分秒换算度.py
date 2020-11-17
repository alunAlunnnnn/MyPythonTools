import pandas as pd
import os
import openpyxl


def _convertToDegree(inDMS):
    data = inDMS.replace("°", "_").replace("′", "_").replace("″", "").split("_")
    return float(data[0]) + float(data[1]) / 60 + float(data[2]) / 3600


def _converDMSToDegree(inDir, outDir):
    files = [os.path.join(inDir, eachFile) for eachFile in os.listdir(inDir)]
    for eachFile in files:
        df = pd.read_excel(eachFile)
        xlsx = os.path.join(outDir, os.path.splitext(os.path.split(eachFile)[1])[0] + ".xlsx")
        df.to_excel(xlsx, index=False)

        wb = openpyxl.load_workbook(xlsx)
        sht = wb.active
        maxRow = sht.max_row
        maxCol = sht.max_column
        print(maxRow, maxCol)
        for i in range(2, maxRow + 1):
            lon, lat = sht.cell(i, 4).value, sht.cell(i, 5).value
            lon, lat = _convertToDegree(lon), _convertToDegree(lat)
            print(lon, lat)
            sht.cell(i, 4).value, sht.cell(i, 5).value = lon, lat
        wb.save(xlsx)

        df = pd.read_excel(xlsx)
        csv = os.path.join(outDir, os.path.splitext(os.path.split(eachFile)[1])[0] + ".txt")
        df.to_csv(csv, index=False)
        os.remove(xlsx)


dir = r"./data"
newDir = r"./newData"
_converDMSToDegree(dir, newDir)
