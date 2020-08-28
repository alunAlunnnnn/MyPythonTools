import openpyxl, os, requests, re
from lxml import etree
import time

datalist = ["超声波液位仪", "防爆型含氧量检测", "防爆型红外入侵", "防爆型摄像机", "防爆型声光报警", "防爆型温湿度检测",
            "含氧量检测", "红外入侵", "甲烷检测", "硫化氢检测", "摄像机", "声光报警", "温湿度检测", "自动液压井盖"]
newNameList = ["GL_SENSOR_CSBYWY", "GL_SENSOR_FBXHYLJC", "GL_SENSOR_FBXHWRQ", "GL_SENSOR_FBXSXJ",
               "GL_SENSOR_FBXSGBJ", "GL_SENSOR_FBXWSDJC", "GL_SENSOR_HYLJC", "GL_SENSOR_HWRQ",
               "GL_SENSOR_JWJC", "GL_SENSOR_LHQJC", "GL_SENSOR_SXJ", "GL_SENSOR_SGBJ",
               "GL_SENSOR_WSDJC", "GL_SENSOR_ZDYYJG"]

dir = r"E:\松江管廊\新数据0805\服务发布\对应关系表"

baseurl = "http://192.168.2.242/server/rest/services/GL_SJ/"
serverType = "/MapServer"
wb = openpyxl.Workbook()
sht = wb.active

shtTitle = ["服务名", "服务含义", "服务类型", "坐标系", "中心点", "最小包络矩形", "url", "发布人", "发布时间"]

# 表头
for i, each in enumerate(shtTitle):
    sht.cell(1, i + 1).value = each

# 获取服务信息
for i, eachService in enumerate(newNameList):
    m = i * 2 + 2
    n = m + 1
    mapurl = baseurl + eachService + serverType
    feaurl = baseurl + eachService + "/FeatureServer"
    time.sleep(1)
    print(mapurl)
    res = requests.get(mapurl)
    if res.status_code == 200:
        html = etree.HTML(res.text)

        # /html/body/div/ul[3]/text()[1]
        xmin = html.xpath("/html/body/div/ul[3]/text()[1]")
        ymin = html.xpath("/html/body/div/ul[3]/text()[2]")
        xmax = html.xpath("/html/body/div/ul[3]/text()[3]")
        ymax = html.xpath("/html/body/div/ul[3]/text()[4]")
        sr = html.xpath("/html/body/div/ul[3]/text()[5]")

        xmin = float(xmin[0].split(" ")[1])
        ymin = float(ymin[0].split(" ")[1])
        xmax = float(xmax[0].split(" ")[1])
        ymax = float(ymax[0].split(" ")[1])
        center = str(((xmin + xmax) / 2, (ymin + ymax) / 2))
        extent = str((xmin, ymin, xmax, ymax))
        sr = sr[0].replace("\r\n", "").split(" ")[2].split("\xa0")[0]
        publisher = "阿伦"
        publishTime = "2020-08-19"
        print(xmin)
        print(ymin)
        print(xmax)
        print(ymax)
        print(center)
        print(sr)

        # exit()

        rowData_1 = [eachService, datalist[i], "MapServer", sr, center, extent, mapurl, publisher, publishTime]
        rowData_2 = [eachService, datalist[i], "FeatureServer", sr, center, extent, feaurl, publisher, publishTime]
        for colNum, eachData in enumerate(shtTitle):
            listIndex = colNum
            colNum += 1
            sht.cell(m, colNum).value = rowData_1[listIndex]
            sht.cell(n, colNum).value = rowData_2[listIndex]
        mergeList = [1, 2, 4, 5, 6]
        for mergeCol in mergeList:
            sht.merge_cells(start_row=m, start_column=mergeCol, end_row=n, end_column=mergeCol)
            wb.save(os.path.join(dir, "松江管廊服务对应关系.xlsx"))


sht.merge_cells(start_row=2, start_column=8, end_row=n, end_column=8)
sht.merge_cells(start_row=2, start_column=9, end_row=n, end_column=9)
wb.save(os.path.join(dir, "松江管廊服务对应关系.xlsx"))
