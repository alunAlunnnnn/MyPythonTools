import arcpy
import os
import openpyxl
import pandas as pd

# import win32com.client as win32

arcpy.env.overwriteOutput = True


def addMessage(mes):
    print(mes)
    arcpy.AddMessage(mes)


def addWarning(mes):
    print(mes)
    arcpy.AddWarning(mes)


def addError(mes):
    print(mes)
    arcpy.AddError(mes)


# points type is not tuple
class pointError(Exception):
    pass


class calKError(Exception):
    pass


class lineEquation:
    def __init__(self, *args):
        """
        usage: 实例化线要素时，传入参数 (tuple1, tuple2, tuple3)
            tuple1 --- x_f, y_f, z_f, 起点的 x, y, z坐标值
            tuple2 --- x_l, y_l, z_l, 终点的 x, y, z坐标值
            tuple3 --- extent_xmin, extent_ymin, extent_xmax, extent_ymax
        :param args:
        """
        # save all points
        self.points = []

        for each in args[:2]:
            if not isinstance(each, tuple):
                addMessage('Point coord is not tuple type')
                raise pointError

            self.points.append((float(each[0]), float(each[1]), float(each[2])))

        self.extent_xmin = args[2][0]
        self.extent_ymin = args[2][1]
        self.extent_xmax = args[2][2]
        self.extent_ymax = args[2][3]

        # get point number, start with 1
        self.pntNum = len(args)

        # set coord of start point and finish point
        self.x1 = self.points[0][0]
        self.y1 = self.points[0][1]
        self.z1 = self.points[0][2]
        self.x2 = self.points[-1][0]
        self.y2 = self.points[-1][1]
        self.z2 = self.points[-1][2]

        self.calculateK_xy()
        self.calculateB_xy()

        self.calculateK_yz()
        self.calculateB_yz()

        self.calculateK_xz()
        self.calculateB_xz()

        self.generateEquation()

    # calculate k --- ( y2 - y1 ) / ( x2 - x1 )
    def calculateK_xy(self):
        if self.x1 == self.x2:
            addError('ERROR --- calculate k_xy faild, x1 is equal to x2')
            self.k_xy = 0
            return self
            # raise calKError
        k = (self.y2 - self.y1) / (self.x2 - self.x1)
        self.k_xy = k
        return self

    # calculate b --- y1 - k * x1
    def calculateB_xy(self):
        b = self.y1 - self.k_xy * self.x1
        self.b_xy = b
        return self

    # calculate k --- ( z2 - z1 ) / ( y2 - y1 )
    def calculateK_yz(self):
        if self.y1 == self.y2:
            # addError('ERROR --- calculate k_yz faild, y1 is equal to y2. y1 is %s' % self.y1)
            self.k_yz = 0
            return self
            # raise calKError
        k = (self.z2 - self.z1) / (self.y2 - self.y1)
        self.k_yz = k
        return self

    # calculate b --- z1 - k * y1
    def calculateB_yz(self):
        b = self.z1 - self.k_yz * self.y1
        self.b_yz = b
        return self

    # calculate k --- ( z2 - z1 ) / ( y2 - y1 )
    def calculateK_xz(self):
        if self.x1 == self.x2:
            # addError('ERROR --- calculate k_xz faild, x1 is equal to x2')
            self.k_xz = 0
            return self
            # raise calKError
        k = (self.z2 - self.z1) / (self.x2 - self.x1)
        self.k_xz = k
        return self

    # calculate b --- z1 - k * y1
    def calculateB_xz(self):
        b = self.z1 - self.k_xz * self.x1
        self.b_xz = b
        return self

    # generate function equation
    def generateEquation(self):
        self.euqation_xy = '%s * x + %s' % (self.k_xy, self.b_xy)
        self.euqation_yz = '%s * x + %s' % (self.k_yz, self.b_yz)
        self.euqation_xz = '%s * x + %s' % (self.k_xz, self.b_xz)
        return self

    # calculate the intersect point
    def calculateIntersect(self, otherLineObj):
        if self.k_xy == otherLineObj.k_xy:
            self.intersect = 'false'
            otherLineObj.intersect = 'false'
            return None

        if self.b_xy == otherLineObj.b_xy:
            x = 0
            y = self.b_xy
        else:
            x = (otherLineObj.b_xy - self.b_xy) / (self.k_xy - otherLineObj.k_xy)
            y = self.k_xy * x + self.b_xy

        # detect whether the point in the rectangle of self line
        if x > self.extent_xmin and x < self.extent_xmax:
            if y > self.extent_ymin and y < self.extent_ymax:
                self.intersect = 'true'
            else:
                self.intersect = 'false'
        else:
            self.intersect = 'false'

        # detect whether the point in the rectangle of another line
        if x > otherLineObj.extent_xmin and x < otherLineObj.extent_xmax:
            if y > otherLineObj.extent_ymin and y < otherLineObj.extent_ymax:
                otherLineObj.intersect = 'true'
            else:
                otherLineObj.intersect = 'false'
        else:
            otherLineObj.intersect = 'false'

        return x, y

    # calculate z value in intersect x,y
    def calculateZCoord_yz(self, y):
        z = self.k_yz * y + self.b_yz
        return z

    # calculate z value in intersect x,y
    def calculateZCoord_xz(self, x):
        z = self.k_xz * x + self.b_xz
        return z


# 给以桩号K-1+360为起始的线，通过横断面获取的中间节点 添加需要的字段，并计算其坐标值
def _addFieldsAndCalculateXYZ(data, fieldsName, fieldsType, fieldsAlias):
    for i, eachField in enumerate(fieldsName):
        try:
            arcpy.AddField_management(data, eachField, fieldsType[i], field_alias=fieldsAlias[i])
        except:
            arcpy.DeleteField_management(data, eachField)
            arcpy.AddField_management(data, eachField, fieldsType[i], field_alias=fieldsAlias[i])

        if eachField == "x":
            arcpy.CalculateField_management(data, eachField, "!shape.centroid.X!", "PYTHON3")
        elif eachField == "y":
            arcpy.CalculateField_management(data, eachField, "!shape.centroid.Y!", "PYTHON3")
        elif eachField == "z":
            arcpy.CalculateField_management(data, eachField, "!shape.centroid.Z!", "PYTHON3")


# 计算每个点到下个点的距离，确保每个点的距离都是以1米作为分割
def _calDistance(dataSorted):
    try:
        arcpy.AddField_management(dataSorted, "distance", "DOUBLE", field_alias="与后点距离")
    except:
        arcpy.DeleteField_management(dataSorted, "distance")
        arcpy.AddField_management(dataSorted, "distance", "DOUBLE", field_alias="与后点距离")

    codes = """import math
x = -99999
y = -99999
def f(a, b):
    a = float(a)
    b = float(b)
    global x, y
    if x == -99999 and y == -99999:
        x = a
        y = b
        dis = 0
    else:
        dis = math.sqrt((a - x) ** 2 + (b - y) ** 2)
        x = a
        y = b
    
    return dis"""
    arcpy.CalculateField_management(dataSorted, "distance", "f(!x!, !y!)", "PYTHON3", codes)


# 从k-1+360，即 -1360开始向后每个点增加1，作为点号
def _calPosition(dataSorted):
    try:
        arcpy.AddField_management(dataSorted, "position", "DOUBLE", field_alias="标号值")
    except:
        arcpy.DeleteField_management(dataSorted, "position")
        arcpy.AddField_management(dataSorted, "position", "DOUBLE", field_alias="标号值")

    # 此处传入的点是已经排过序的无重复的 线路的1m间隔点，只要从上到下递增即可。 线若是纵向的，则在排序和此处进行修改
    startNum = -1359
    step = 1
    codes = """a = {0}
def f():
    global a
    if a >= - 1999 and a <= -1359:
        a -= {1}
    if a >= 0:
        a += {1}
    if a == -2000:
        a = 0
    return a""".format(startNum, step)
    arcpy.CalculateField_management(dataSorted, "position", "f()", "PYTHON3", codes)

    try:
        arcpy.AddField_management(dataSorted, "zh_posi", "TEXT")
    except:
        arcpy.DeleteField_management(dataSorted, "zh_posi")
        arcpy.AddField_management(dataSorted, "zh_posi", "TEXT")

    codes = """def f(a):
    a = int(a)
    if a >= 0:
        a = str(a)
        if len(a) == 1:
            res = "K0+00" + a
        elif len(a) == 2:
            res = "K0+0" + a
        elif len(a) == 3:
            res = "K0+" + a
        elif len(a) == 4:
            res = "K" + a[0] + "+" + a[1:]
        else:
            res = "ERROR"
    else:
        a = str(abs(a))
        if len(a) == 2:
            res = "-K0+00" + a
        elif len(a) == 3:
            res = "-K0+0" + a
        elif len(a) == 4:
            res = "-K0+" + a
        elif len(a) == 5:
            res = "-K" + a[0] + "+" + a[1:]
        else:
            res = "ERROR"
    return res"""
    arcpy.CalculateField_management(dataSorted, "zh_posi", "f(!position!)", "PYTHON3", codes)


# 通过excel赋值到属性表
def _mergeExcel(xlsx):
    """
    :usage: 将excel文件中的多个sheet合并到一个新的excel中
    :param xlsx: 输入excel文件路径，string
    :return: 返回一个新的excel文件
    """

    sheetNameList = ["环境监测", "设备监控", "视频监控", "防入侵"]
    wb = openpyxl.load_workbook(xlsx)

    # 创建一个新的excel用来合并所有数据
    wb_new = openpyxl.Workbook()
    sht_new = wb_new.create_sheet(title="total")
    sht_new.append(["所属分区", "所属舱室", "桩号", "设备类型", "唯一ID"])

    for eachSheet in sheetNameList:
        sht = wb[eachSheet]
        for eachRow in sht.rows:
            eachRow = [eachCell.value for eachCell in eachRow]
            sht_new.append(eachRow)

    path = os.path.dirname(xlsx)
    print(os.path.join(path, "data_total.xlsx"))

    del wb_new["Sheet"]
    wb_new.save(os.path.join(path, "data_total.xlsx"))
    return os.path.join(path, "data_total.xlsx")


# arcmap dose not support the .xlsx file, so convert .xlsx to .xls
def _convertXlsxToXls(xlsx):
    data = pd.read_excel(xlsx)
    data.to_excel(xlsx.split(".xlsx")[0] + ".xls", sheet_name="total", encoding="utf-8")
    # data.to_csv(xlsx.split(".xlsx")[0] + ".csv", encoding="utf-8")
    return xlsx.split(".xlsx")[0] + ".xls"


# # use win32com to convert
# def _convertXlsxToXls_win32(xlsx):
#     excel = win32.gencache.EnsureDispatch('Excel.Application')
#     wb = excel.Workbooks.Open(xlsx)
#
#     # 51 --- .xlsx, 56 --- .xls
#     wb.SaveAs(xlsx.split(".xlsx")[0] + ".xls", FileFormat=56)
#     wb.close()
#     excel.Application.Quit()
#     return xlsx.split(".xlsx")[0] + ".xls"


# join feature class attribute to excel
def _joinFeatureAttrToExcel(xls, inFC, outTableName):
    # 构建excel表视图
    tv = arcpy.MakeTableView_management(xls + r"\total$", "tempTable")
    table = arcpy.TableToTable_conversion(tv, os.path.dirname(inFC), "tempTable")
    # table = arcpy.TableToTable_conversion(xls + r"\total$", os.path.dirname(inFC), "tempTable")
    newtv = arcpy.MakeTableView_management(table, "newTempTable")

    # 构建要素图层
    lyr = arcpy.MakeFeatureLayer_management(inFC, "tempLyr")

    # 将要素图层属性连接到表视图中
    attrTv = arcpy.AddJoin_management(newtv, "桩号", lyr, "zh_posi")

    # 导出以保存数据
    coordTable = arcpy.TableToTable_conversion(attrTv, os.path.dirname(inFC), outTableName)

    # 排序表并获取重复点位的值
    coTabSorted = arcpy.Sort_management(coordTable, os.path.join(os.path.dirname(inFC), "finTab_sorted"),
                                        [["x", "ASCENDING"], ["y", "ASCENDING"]])
    newtv = arcpy.MakeTableView_management(coTabSorted, "tabSorted")
    try:
        arcpy.AddField_management(newtv, "inner_id_", "LONG")
    except:
        arcpy.DeleteField_management(newtv, "inner_id_")
        arcpy.AddField_management(newtv, "inner_id_", "LONG")
    codes = """x = 0
y = 0
num = 1
def f(a, b):
    global x,y,num
    a = float(a)
    b = float(b)
    if a == x:
        if b == y:
            num += 1
        else:
            num = 1
    else:
        num = 1
    x = a
    y = b
    return num"""
    arcpy.CalculateField_management(newtv, "inner_id_", "f(!x!, !y!)", "PYTHON3", codes)

    # 汇总统计数据，目的是获得每个点位的最大值，这里其实可以省略掉上面计算有几点点的那步，直接用频数来就行
    tabSta = arcpy.Statistics_analysis(newtv, os.path.join(os.path.dirname(inFC), "finTab_sorted_sta"),
                                       [["inner_id_", "MAX"]],
                                       ["x", "y"])

    # 读取统计属性表，将结果写出为字典，之后通过字段计算器将最大值固定到字段中
    staDict = {}
    with arcpy.da.SearchCursor(tabSta, ["x", "y", "MAX_inner_id_"]) as cur:
        for row in cur:
            # 可能存在空值导致程序运行失败
            try:
                newKey = (round(float(row[0]), 7), round(float(row[1]), 7))
                newValue = int(row[2])
                staDict[newKey] = newValue
            except:
                pass
        del row

    try:
        arcpy.AddField_management(newtv, "pnt_id_", "TEXT")
    except:
        arcpy.DeleteField_management(newtv, "pnt_id_")
        arcpy.AddField_management(newtv, "pnt_id_", "TEXT")

    codes = """def f(a, b):
    return {}[(round(a, 7), round(b, 7))]""".format(staDict)
    print(staDict)
    arcpy.CalculateField_management(newtv, "pnt_id_", "f(!x!, !y!)", "PYTHON3", codes)

    try:
        arcpy.AddField_management(newtv, "new_x_", "DOUBLE")
    except:
        arcpy.DeleteField_management(newtv, "new_x_")
        arcpy.AddField_management(newtv, "new_x_", "DOUBLE")

    try:
        arcpy.AddField_management(newtv, "new_y_", "DOUBLE")
    except:
        arcpy.DeleteField_management(newtv, "new_y_")
        arcpy.AddField_management(newtv, "new_y_", "DOUBLE")

    # 环绕分布的点的半径
    r = 0.4
    # 第一个开始环绕的点的角度 —— 0 代表从正东方向开始，角度为0
    startDeg = 0

    codes1 = """import math
def f(x, pntNo, totalNo, startDeg, r):
    try:
        x, pntNo, totalNo, r, startDeg = float(x), float(pntNo), float(totalNo), float(r), float(startDeg)
        if totalNo > 1:
            deg = (360 / totalNo) + startDeg
            rad = math.radians(deg * pntNo)
            newx = r * round(math.cos(rad), 7) + x
            return newx
        else:
            return x
    except:
        return -999999"""
    arcpy.CalculateField_management(newtv, "new_x_", "f(!x!, !inner_id_!, !pnt_id_!, {}, {})".format(startDeg, r), "PYTHON3", codes1)

    codes1 = """import math
def f(y, pntNo, totalNo, startDeg, r):
    try:
        y, pntNo, totalNo, r, startDeg = float(y), float(pntNo), float(totalNo), float(r), float(startDeg)
        if totalNo > 1:
            deg = (360 / totalNo) + startDeg
            rad = math.radians(deg * pntNo)
            newy = r * round(math.sin(rad), 7) + y
            return newy
        else:
            return y
    except:
        return -999999"""
    arcpy.CalculateField_management(newtv, "new_y_", "f(!y!, !inner_id_!, !pnt_id_!, {}, {})".format(startDeg, r), "PYTHON3", codes1)


    # 提取坐标表，点位重复的地方，设两种模式处理
    # 一种将中心点空下来，所有点以圆环绕排布
    # 另一种在中心点生成一个点，其他点以圆环绕排布

    # 删除临时数据
    try:
        arcpy.Delete_management(table)
    except:
        pass


def main(data, dataSorted, fieldsName, fieldsType, fieldsAlias, xlsx):
    # 需要首先在cass里面先生成一米分割点，然后每条线上的所有点单独搞一个图层进来
    # 不要把多条线的点一次性都搞进来
    _addFieldsAndCalculateXYZ(data, fieldsName, fieldsType, fieldsAlias)

    arcpy.Sort_management(data, dataSorted, ["x", "y", "z"])

    _calPosition(dataSorted)

    _calDistance(dataSorted)

    resXlsx = _mergeExcel(xlsx)
    print("excel 合并完成")
    print(resXlsx)

    # ******* 这块暂时舍弃了 down
    # 安装了AccessDatabaseEngine.exe 32bit以后，可以直接连接xlsx
    # xls = _convertXlsxToXls(resXlsx)
    # xls = _convertXlsxToXls_win32(resXlsx)

    # 这里用excel连接点数据有点问题。---- 没啥问题arcgis犯病，按一下xlsx驱动就行了
    # _joinFeatureAttrToExcel(xls, data, "finTable")
    # ****** 舍弃 up

    _joinFeatureAttrToExcel(resXlsx, dataSorted, "finTable")


gdb = r"E:\松江管廊\新数据0805\道路中心线提取\数据处理.gdb"
arcpy.env.workspace = gdb

data = "道路中心线_1米分割点_桩点起始"
dataSorted = "道路中心线_1米分割点_桩点起始_sorted"
xlsx = r"E:\松江管廊\新数据0805\波汇设备清单_处理.xlsx"

fieldsName = ["unic_id", "ssfq", "sscs", "zh", "sblx", "x", "y", "z"]
fieldsType = ["LONG", "TEXT", "TEXT", "TEXT", "TEXT", "DOUBLE", "DOUBLE", "DOUBLE"]
fieldsAlias = ["唯一ID", "所属分区", "所属舱室", "桩号", "设备类型", "x", "y", "z"]

main(data, dataSorted, fieldsName, fieldsType, fieldsAlias, xlsx)

# TODO 目前是以管廊中心线为定位依据，没有考虑到管廊中舱室的区别（如综合舱、天然气舱、电力舱污水舱，目前有4个舱
#  、5个工作井、1个变电所、4个通风口）

