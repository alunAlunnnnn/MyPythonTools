import openpyxl, os

dir = r"E:\松江管廊\新数据0805\服务发布\对应关系表"
wb = openpyxl.load_workbook(os.path.join(dir, "松江管廊服务对应关系.xlsx"))
sht = wb.active
n = 29
sht.merge_cells(start_row=2, start_column=8, end_row=n, end_column=8)
sht.merge_cells(start_row=2, start_column=9, end_row=n, end_column=9)
wb.save(os.path.join(dir, "松江管廊服务对应关系_new.xlsx"))