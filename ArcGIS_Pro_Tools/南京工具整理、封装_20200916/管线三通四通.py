import sys
import os
import logging
import datetime

# get arcgis pro toolbox directory
tbxDir = os.path.dirname(sys.argv[0])

# create log directory
logDir = os.path.join(tbxDir, "tbx_log")
try:
    if not os.path.exists(logDir):
        os.makedirs(logDir)
except:
    pass

# make sure this script have rights to create file here
createFileRights = False
fileRightTest = os.path.join(logDir, "t_t_.txt")
logFile = os.path.join(logDir, "tool3_gxstst_log.txt")
try:
    with open(fileRightTest, "w", encoding="utf-8") as f:
        f.write("create file rights test")
    createFileRights = True
except:
    createFileRights = False

# init log set config
if createFileRights:
    logging.basicConfig(filename=logFile, filemode="w", level=logging.DEBUG,
                        format="\n\n *** \n %(asctime)s    %(levelname)s ==== %(message)s \n *** \n\n")
else:
    logging.basicConfig(level=logging.DEBUG,
                        format="\n\n *** \n %(asctime)s    %(levelname)s ==== %(message)s \n *** \n\n")

try:
    import arcpy
    import functools
    import inspect
    import openpyxl
    import math

    logging.debug("Module import success")
except BaseException as e:
    logging.error(str(e))

arcpy.env.overwriteOutput = True

logging.info(" ========== Progress Start Running ========== ")


def logIt(func):
    @functools.wraps(func)
    def _wrapper(*args, **kwargs):
        logging.info(f" ======================== Start Function {func.__name__} ========================")
        parameters = [each.strip() for each in str(inspect.signature(func))[1:-1].split(",")]
        keyPara = kwargs
        for eachKey in keyPara:
            eachKey = eachKey.strip()
            if eachKey in parameters:
                parameters.remove(eachKey)
        new_args = list(args)
        for eachValue in kwargs.values():
            new_args.append(eachValue)

        mes = ""
        for i, eachPara in enumerate(parameters):
            mes += f"\n {eachPara}: {new_args[i]} "

        for eachKey, eachValue in keyPara.items():
            mes += f"\n {eachKey}: {eachValue} "

        logging.debug(mes)
        res = func(*args, **kwargs)

        return res

    return _wrapper


class NotPointStyleField(Exception):
    pass


# ---------- show message in toolbox ----------

def _addMessage(mes):
    print(mes)
    arcpy.AddMessage(mes)


def _addWarning(mes):
    print(mes)
    arcpy.AddWarning(mes)


def _addError(mes):
    print(mes)
    arcpy.AddError(mes)


def getRunTime(func):
    @functools.wraps(func)
    def _wrapper(*args, **kwargs):
        start = datetime.datetime.now()
        print("start at: {} ".format(start))
        res = func(*args, **kwargs)
        finish = datetime.datetime.now()
        print("start at: {} ".format(start))
        print("finish at: {}".format(finish))
        cost = finish - start
        print("total cost: {}".format(cost))
        return res

    return _wrapper


# ---------- data fields process ----------

def _addField(inFeaShp, fieldName, fieldType):
    # arcpy.AddField_management(inFeaShp, fieldName, fieldType)
    try:
        arcpy.AddField_management(inFeaShp, fieldName, fieldType)
    except:
        arcpy.DeleteField_management(inFeaShp, fieldName)
        arcpy.AddField_management(inFeaShp, fieldName, fieldType)


# add a filed named "unic_id_" and calculate the unic id
def _addUnicIDField(inFeaShp):
    _addField(inFeaShp, "unic_id_", "LONG")
    codes = """a = 0
def calUnicField():
    global a
    a += 1
    return a"""
    arcpy.CalculateField_management(inFeaShp, "unic_id_", "calUnicField()", "PYTHON3", codes)


# add coord fields
def _addCoordFields(inFC, feaType):
    if feaType == "point":
        fieldList = ["x_cen_", "y_cen_", "z_cen_"]
        expList = ["!shape.centroid.X!", "!shape.centroid.Y!", "!shape.centroid.Z!"]
        for i, each in enumerate(fieldList):
            _addField(inFC, each, "DOUBLE")
            arcpy.CalculateField_management(inFC, each, expList[i], "PYTHON3")
    elif feaType == "line":
        fieldList = ["x_f_", "y_f_", "z_f_", "x_l_", "y_l_", "z_l_"]
        expList = ["!shape.firstPoint.X!", "!shape.firstPoint.Y!", "!shape.firstPoint.Z!",
                   "!shape.lastPoint.X!", "!shape.lastPoint.Y!", "!shape.lastPoint.Z!"]
        for i, each in enumerate(fieldList):
            _addField(inFC, each, "DOUBLE")
            arcpy.CalculateField_management(inFC, each, expList[i], "PYTHON3")


# ---------- line class ----------

# points type is not tuple
class pointError(Exception):
    pass


class calKError(Exception):
    pass


class GenerateSpatialIndexError(Exception):
    pass


class lineEquation:
    """
    usage: generate a line object
    :param args: accept as
     --- (fp_x, fp_y, fp_z), (lp_x, lp_y, lp_z), (ex_xin, ex_ymin, ex_xmax, ex_ymax)
    firstPoint --- fp
    lastPoint --- lp
    extent --- ex
    """

    def __init__(self, *args):
        # save all points
        self.points = []

        for each in args[:2]:
            if not isinstance(each, tuple):
                _addMessage('Point coord is not tuple type')
                raise pointError

            self.points.append((float(each[0]), float(each[1]), float(each[2])))

        self.extent_xmin = args[2][0]
        self.extent_ymin = args[2][1]
        self.extent_xmax = args[2][2]
        self.extent_ymax = args[2][3]
        self.extent = args[2]

        # get point number, start with 1
        self.pntNum = len(args)

        # init pipe size
        self.pipeSize = None

        # set coord of start point and finish point
        self.x1 = self.points[0][0]
        self.y1 = self.points[0][1]
        self.z1 = self.points[0][2]
        self.x2 = self.points[-1][0]
        self.y2 = self.points[-1][1]
        self.z2 = self.points[-1][2]

        # extent available detect
        self._extentAvailableDetect()

        self.spaindex = None
        self.spaindex_row = None
        self.spaindex_col = None
        self.spaind_totalext = None

        self.calculateK_xy()
        self.calculateB_xy()

        self.calculateK_yz()
        self.calculateB_yz()

        self.calculateK_xz()
        self.calculateB_xz()

        self.generateEquation()

    # calculate k --- ( y2 - y1 ) / ( x2 - x1 )
    def calculateK_xy(self):
        if self.x1 == self.x2:
            _addWarning('ERROR --- calculate k_xy faild,'
                        ' x1 is equal to x2, x1 is {}. x2 is {}'.format(self.x1, self.x2))
            self.k_xy = -999
            return self
            # raise calKError
        k = (self.y2 - self.y1) / (self.x2 - self.x1)
        self.k_xy = k
        return self

    # calculate b --- y1 - k * x1
    def calculateB_xy(self):
        if self.k_xy == -999:
            self.b_xy = -999
        else:
            b = self.y1 - self.k_xy * self.x1
            self.b_xy = b
        return self

    # calculate k --- ( z2 - z1 ) / ( y2 - y1 )
    def calculateK_yz(self):
        if self.y1 == self.y2:
            _addWarning('ERROR --- calculate k_yz faild, y1 is equal to y2. y1 is %s' % self.y1)
            self.k_yz = -999
            return self
            # raise calKError
        k = (self.z2 - self.z1) / (self.y2 - self.y1)
        self.k_yz = k
        return self

    # calculate b --- z1 - k * y1
    def calculateB_yz(self):
        if self.k_yz == -999:
            self.b_yz = -999
        else:
            b = self.z1 - self.k_yz * self.y1
            self.b_yz = b
        return self

    # calculate k --- ( z2 - z1 ) / ( y2 - y1 )
    def calculateK_xz(self):
        if self.x1 == self.x2:
            _addWarning('ERROR --- calculate k_xz faild, x1 is equal to x2. x1 is %s' % self.x1)
            self.k_xz = -999
            return self
            # raise calKError
        k = (self.z2 - self.z1) / (self.x2 - self.x1)
        self.k_xz = k
        return self

    # calculate b --- z1 - k * y1
    def calculateB_xz(self):
        if self.k_xz == -999:
            self.b_xz = -999
        else:
            b = self.z1 - self.k_xz * self.x1
            self.b_xz = b
        return self

    # generate function equation
    def generateEquation(self):
        self.euqation_xy = '%s * x + %s' % (self.k_xy, self.b_xy)
        self.euqation_yz = '%s * x + %s' % (self.k_yz, self.b_yz)
        self.euqation_xz = '%s * x + %s' % (self.k_xz, self.b_xz)
        return self

    # calculate the intersect point
    def calculateIntersect(self, otherLineObj):
        if self.k_xy == otherLineObj.k_xy:
            self.intersect = 'false'
            otherLineObj.intersect = 'false'
            return None

        if self.b_xy == otherLineObj.b_xy:
            x = 0
            y = self.b_xy
        else:
            x = (otherLineObj.b_xy - self.b_xy) / (self.k_xy - otherLineObj.k_xy)
            y = self.k_xy * x + self.b_xy

        # detect whether the point in the rectangle of self line
        if x > self.extent_xmin and x < self.extent_xmax:
            if y > self.extent_ymin and y < self.extent_ymax:
                self.intersect = 'true'
            else:
                self.intersect = 'false'
        else:
            self.intersect = 'false'

        # detect whether the point in the rectangle of another line
        if x > otherLineObj.extent_xmin and x < otherLineObj.extent_xmax:
            if y > otherLineObj.extent_ymin and y < otherLineObj.extent_ymax:
                otherLineObj.intersect = 'true'
            else:
                otherLineObj.intersect = 'false'
        else:
            otherLineObj.intersect = 'false'

        return x, y

    # calculate z value in intersect x,y
    def calculateZCoord_yz(self, y):
        z = self.k_yz * y + self.b_yz
        return z

    # calculate z value in intersect x,y
    def calculateZCoord_xz(self, x):
        z = self.k_xz * x + self.b_xz
        return z

    # make sure the x_min, y_min is less or equal than x_max, y_max in extent
    def _extentAvailableDetect(self):
        assert int(self.extent_xmin * 10 ** 8) <= int(
            self.extent_xmax * 10 ** 8), "Error --- Extent of line object is not available"
        assert int(self.extent_ymin * 10 ** 8) <= int(
            self.extent_ymax * 10 ** 8), "Error --- Extent of line object is not available"

    # detect the point can touch the line with a tolerance or not
    def pointTouchDet(self, pnt, tolerance, totalExtent):
        """
        usage: used to detect that, the point is in the line with a tolerance num.
                before use this, please call the function "generateSpatialIndex()" yet
        :param pnt:(x, y)
        :param tolerance:
        :return:
        """
        pnt_x, pnt_y = pnt[0], pnt[1]
        pnt_index = None
        spatialIndex_x = 1
        spatialIndex_y = 1
        # generate the spatial index for point
        spaindex_step_x = round(float((totalExtent[2] - totalExtent[0]) / spatialIndex_x), 6) + 0.0001
        spaindex_step_y = round(float((totalExtent[3] - totalExtent[1]) / spatialIndex_y), 6) + 0.0001
        total_ext_ymax = totalExtent[3]
        total_ext_xmax = totalExtent[2]

        # detect whether is out of the ply's extent
        if (self.extent_xmax > totalExtent[2] + 0.0001 or self.extent_ymax > totalExtent[3] + 0.0001
                or self.extent_xmin < totalExtent[0] - 0.0001 or self.extent_ymin < totalExtent[1] - 0.0001):
            if self.extent_xmax > totalExtent[2] + 0.0001:
                print("pnt xxxx1xxxx pnt")
            if self.extent_ymax > totalExtent[3] + 0.0001:
                print("pnt xxxx2xxxx pnt")
            if self.extent_xmin < totalExtent[0] - 0.0001:
                print("pnt xxxx3xxxx pnt")
            if self.extent_ymin < totalExtent[1] - 0.0001:
                print("pnt xxxx4xxxx pnt")
            _addError("Error --- generate spatial index for point failed, "
                      "the point is ({}, {})".format(pnt_x, pnt_y))
            _addError("total extent is {}".format(totalExtent))
            raise GenerateSpatialIndexError

        find_key = False
        # for i in range(1, 11):
        for i in range(1, spatialIndex_y + 1):
            # if spatial index has finded, break the loop
            if find_key:
                break
            if pnt_y >= total_ext_ymax - i * spaindex_step_y:
                pnt_index_y = str(i)
                # for j in range(1, 11):
                for j in range(1, spatialIndex_x + 1):
                    if pnt_x >= total_ext_xmax - j * spaindex_step_x:
                        pnt_index_x = str(j)
                        pnt_index = (str(i) + "," + str(j))
                        find_key = True
                        break

        # no spatial index, no calculate
        assert pnt_index, "there are no spatial index in point"

        # match spatial index
        if pnt_index == self.spaindex:
            # point is in the extent of polyline
            if (self.extent_xmin - tolerance <= pnt_x <= self.extent_xmax + tolerance
                    and self.extent_ymin - tolerance <= pnt_y <= self.extent_ymax + tolerance):
                # point is in the extent of polyline
                # detect whether the point is on the line

                # vertical line
                if self.k_xy == -999:
                    if min(self.x1, self.x2) - tolerance <= pnt_x <= max(self.x1, self.x2) + tolerance:
                        if min(self.y1, self.y2) - tolerance <= pnt_y <= max(self.y1, self.y2) + tolerance:
                            return True
                        else:
                            return False
                    else:
                        return False
                # horizon line
                elif self.k_xy == 0:
                    if min(self.y1, self.y2) - tolerance <= pnt_y <= max(self.y1, self.y2) + tolerance:
                        if min(self.x1, self.x2) - tolerance <= pnt_x <= max(self.x1, self.x2) + tolerance:
                            return True
                        else:
                            return False
                    else:
                        return False
                # calculate the point in line with k_xy
                else:
                    det_y = self.k_xy * pnt_x + self.b_xy
                    if det_y - tolerance <= pnt_y <= det_y + tolerance:
                        return True
                    else:
                        return False
            # point is out of ply's extent
            else:
                return False
        # the spatial index between point and ply is not equal
        else:
            return False

    def generateSpatialIndex(self, totalExtent):
        """
        usage: generate spatial index, now is generate 10*10 grid index.

        index number as this:
         ----------
        |1,10|...|1,2|1,1|
         ----------
        |2,10|...|2,2|2,1|
         ----------
        |...........|
        |.          |
        |.          |
         -----------
        |10,10|...|10,2|10,1|

        :param totalExtent: (xmin, ymin, xmax, ymax)
        :return:
        """
        spatialIndex_x = 1
        spatialIndex_y = 1
        spaindex_step_x = round(float((totalExtent[2] - totalExtent[0]) / spatialIndex_x), 6) + 0.0001
        spaindex_step_y = round(float((totalExtent[3] - totalExtent[1]) / spatialIndex_y), 6) + 0.0001
        total_ext_ymax = totalExtent[3]
        total_ext_xmax = totalExtent[2]

        self.spaind_totalext = totalExtent

        if (self.extent_xmax > totalExtent[2] + 0.0001 or self.extent_ymax > totalExtent[3] + 0.0001
                or self.extent_xmin < totalExtent[0] - 0.0001 or self.extent_ymin < totalExtent[1] - 0.0001):
            if self.extent_xmax > totalExtent[2] + 0.0001:
                print("xxxx1xxxx")
            if self.extent_ymax > totalExtent[3] + 0.0001:
                print("xxxx2xxxx")
            if self.extent_xmin < totalExtent[0] - 0.0001:
                print("xxxx3xxxx")
            if self.extent_ymin < totalExtent[1] - 0.0001:
                print("xxxx4xxxx")
            _addError("Error --- generate spatial index failed, "
                      "line object's extent is not in total extent. "
                      "the line's first point is ({}, {})".format(self.x1, self.y1))
            _addError("total extent is {}".format(totalExtent))
            _addError("ply extent is {}".format(self.extent))
            raise GenerateSpatialIndexError

        find_key = False
        # for i in range(1, 11):
        for i in range(1, spatialIndex_y + 1):
            # if spatial index has finded, break the loop
            if find_key:
                break
            if self.extent_ymax >= total_ext_ymax - i * spaindex_step_y:
                self.spaindex_row = str(i)
                # for j in range(1, 11):
                for j in range(1, spatialIndex_y + 1):
                    if self.extent_xmax >= total_ext_xmax - j * spaindex_step_x:
                        self.spaindex_col = str(j)
                        self.spaindex = (str(i) + "," + str(j))
                        find_key = True
                        break
        return self

    def setPipeSize(self, pipesize):
        if isinstance(pipesize, int) or isinstance(pipesize, float):
            self.pipeSize = pipesize
        else:
            _addWarning("Warning --- pipe size is not a number type, pipe size init failed")
            self.pipeSize = None


# convert feature class to excel
@logIt
def _convertToXlsx(FC, outputPath, feaType, convertAll=True):
    """
    usage: used to convert feature class's attributes to xlsx with the same name
           based on module --- arcpy， openpyxl
    :param FC: type --- String.  input the feature class in esri data type
    :param outputPath: type --- String.  define where the xlsx data will be saved
    :param convertAll: type --- Boolean.  True --> convert all attribute,
           False --> add a field named 'unic_id_' and only convert geometry coordinate and 'unic_id_'
    :return: type --- String.  the directory of output xlsx
    """
    global pntFieldList, plyFieldList, plyTotalExtent
    # create a xlsx file
    wb = openpyxl.Workbook()
    sht = wb.active

    # make sure that the environment of workspace is not None
    arcpy.env.workspace = os.path.dirname(FC) if not arcpy.env.workspace else arcpy.env.workspace

    # add coord field, feaType -- "line" or "point"
    _addCoordFields(FC, feaType)
    # add 'unic_id_' field and calculate the unic id for each row,
    # this must under other add field function, such as _addCoordFields
    _addUnicIDField(FC)

    # convert all attributes to xlsx
    if convertAll:
        fieldNameList = [each.name for each in arcpy.ListFields(FC)
                         if each.type != "OID" and each.type != "Geometry"
                         and each.name != "Shape_Length" and each.name != "Shape_Area"]
        fieldTypeList = [each.type for each in arcpy.ListFields(FC)
                         if each.type != "OID" and each.type != "Geometry"
                         and each.name != "Shape_Length" and each.name != "Shape_Area"]
        if feaType == "point":
            pntFieldList = fieldNameList
        elif feaType == "line":
            plyFieldList = fieldNameList
            ext = arcpy.Describe(FC).extent
            plyTotalExtent = (ext.XMin, ext.YMin, ext.XMax, ext.YMax)

        # add data title to xlsx
        sht["A1"] = "unic_id_"
        for i, eachFieldName in enumerate(fieldNameList[:-1]):
            i += 2
            sht.cell(1, i).value = eachFieldName

        # statistic the number of all fields without unic_id_, it will be add to the field column
        fieldLength = len(fieldNameList) - 1
        with arcpy.da.SearchCursor(FC, fieldNameList) as cur:
            for rowNum, row in enumerate(cur):
                unic_id_ = row[fieldLength]
                sht.cell(rowNum + 2, 1).value = unic_id_
                for i in range(fieldLength):
                    cellData = row[i]
                    sht.cell(rowNum + 2, i + 2).value = cellData

    else:
        pass

    wb.save(os.path.join(outputPath, os.path.basename(FC) + ".xlsx"))
    return os.path.join(outputPath, os.path.basename(FC) + ".xlsx")


# delete the field will be added later, make sure the field "unic_id_" is the last in fields
@logIt
def _deleteFields(inFC, fcType):
    _addMessage("Step1 --- clear fields start")
    fieldList = []
    if fcType == "point":
        fieldList = [each.name for each in arcpy.ListFields(inFC) if each.name == "unic_id_" or
                     each.name == "x_cen_" or each.name == "y_cen_" or each.name == "z_cen_"]
    elif fcType == "line":
        fieldList = [each.name for each in arcpy.ListFields(inFC) if each.name == "unic_id_" or
                     each.name == "x_f_" or each.name == "y_f_" or each.name == "z_f_" or
                     each.name == "x_l_" or each.name == "y_l_" or each.name == "z_l_" or
                     each.name == "gj_num_"]
    if len(fieldList) > 0:
        arcpy.DeleteField_management(inFC, fieldList)
        _addMessage("Step1 --- clear fields finish \n")
    else:
        _addMessage("Step1 --- field is clear, slip delete \n")


@logIt
def _generateLineObj(lineXlsx, plyFieldList, plyTotalExtent):
    wb = openpyxl.load_workbook(lineXlsx)
    sht = wb.active

    # line obj
    lineObjList = []

    # get the index of each point in xlsx
    (x_f_index, y_f_index, z_f_index, x_l_index, y_l_index, z_l_index) = (
        plyFieldList.index("x_f_") + 1, plyFieldList.index("y_f_") + 1, plyFieldList.index("z_f_") + 1,
        plyFieldList.index("x_l_") + 1, plyFieldList.index("y_l_") + 1, plyFieldList.index("z_l_") + 1)
    gl_size_index = plyFieldList.index("gj_num_") + 1

    # read xlsx, start at row 2
    for i, eachRowTup in enumerate(sht.rows):
        # skip the title row
        if i == 0:
            continue

        # get coord
        x_f, y_f, z_f, x_l, y_l, z_l = (eachRowTup[x_f_index].value, eachRowTup[y_f_index].value,
                                        eachRowTup[z_f_index].value, eachRowTup[x_l_index].value,
                                        eachRowTup[y_l_index].value, eachRowTup[z_l_index].value)

        # get the attribute of the pipe size, and save it into line object
        gl_size = eachRowTup[gl_size_index].value

        # generate line object
        firstPoint = (x_f, y_f, z_f)
        lastPoint = (x_l, y_l, z_l)
        extent = (min(x_f, x_l), min(y_f, y_l), max(x_f, x_l), max(y_f, y_l))
        lineObj = lineEquation(firstPoint, lastPoint, extent)
        lineObj.generateSpatialIndex(plyTotalExtent)
        lineObj.setPipeSize(gl_size)

        lineObjList.append(lineObj)

    return lineObjList


@logIt
def _getDirectory(pntXlsx, plyObjList, plyTotalExtent):
    global pntFieldList, tolerance
    # read each row in xlsx, get point type, x, y, z
    wb = openpyxl.load_workbook(pntXlsx)
    sht = wb.active

    # "STYLE_TYPE" is the field within santong, sitong
    maxCol = sht.max_column
    maxRow = sht.max_row
    pntStyleIndex = pntFieldList.index("STYLE_TYPE") + 2
    xCenIndex = pntFieldList.index("x_cen_") + 2
    yCenIndex = pntFieldList.index("y_cen_") + 2
    zCenIndex = pntFieldList.index("z_cen_") + 2

    if sht.cell(1, pntStyleIndex).value == "STYLE_TYPE":

        # set directory fields
        fieldTitle = ["z1_line_", "dir_1_", "gj_line1_", "roll_z_adjust1_", "roll_z_adjust1_y_",
                      "z2_line_", "dir_2_", "gj_line2_", "roll_z_adjust2_", "roll_z_adjust2_y_",
                      "z3_line_", "dir_3_", "gj_line3_", "roll_z_adjust3_", "roll_z_adjust3_y_",
                      "z4_line_", "dir_4_", "gj_line4_", "roll_z_adjust4_", "roll_z_adjust4_y_"]
        for i, eachTitle in enumerate(fieldTitle):
            i += 1
            sht.cell(1, maxCol + i).value = eachTitle
        for i, eachRow in enumerate(sht.rows):
            step = 1
            i += 1
            # skip title row
            if i == 1:
                continue

            pntType = eachRow[pntStyleIndex].value.strip()
            x_cord = float(sht.cell(i, xCenIndex).value)
            y_cord = float(sht.cell(i, yCenIndex).value)
            z_cord = float(sht.cell(i, zCenIndex).value)

            # whether the point is on the line
            for eachLineObj in plyObjList:

                if eachLineObj.pointTouchDet((x_cord, y_cord), tolerance, plyTotalExtent):
                    writeCol = maxCol + (5 * step - 3)
                    gjSize = maxCol + (5 * step - 2)
                    newzCol = maxCol + (5 * step - 4)
                    rollZAdjust = maxCol + (5 * step - 1)
                    rollZAdjust_y = maxCol + (5 * step)

                    # point is on the ply
                    dis_1 = math.sqrt((x_cord - eachLineObj.x1) ** 2 + (y_cord - eachLineObj.y1) ** 2)
                    dis_2 = math.sqrt((x_cord - eachLineObj.x2) ** 2 + (y_cord - eachLineObj.y2) ** 2)
                    # end point in ply is more far from point
                    if dis_1 < dis_2:
                        assert dis_2 > 0, "Error --- the max distance is not bigger 0"
                        # dir_pnt = math.degrees(math.asin((y_cord - eachLineObj.y2) / dis_2))
                        dir_pnt = math.degrees(
                            math.atan2((eachLineObj.y2 - eachLineObj.y1), (eachLineObj.x2 - eachLineObj.x1)))

                        roll_z_adjust = math.degrees(
                            math.atan2((eachLineObj.z2 - eachLineObj.z1), (eachLineObj.x2 - eachLineObj.x1)))

                        roll_z_adjust_y = math.degrees(
                            math.atan2((eachLineObj.z2 - eachLineObj.z1), (eachLineObj.y2 - eachLineObj.y1)))

                        # get the line z in point x, y as the new z of point
                        pnt_line_z = eachLineObj.z1

                    else:
                        assert dis_1 > 0, "Error --- the max distance is not bigger 0"
                        # dir_pnt = math.degrees(math.asin((y_cord - eachLineObj.y1) / dis_1))
                        dir_pnt = math.degrees(
                            math.atan2((eachLineObj.y1 - eachLineObj.y2), (eachLineObj.x1 - eachLineObj.x2)))

                        roll_z_adjust = math.degrees(
                            math.atan2((eachLineObj.z1 - eachLineObj.z2), (eachLineObj.x1 - eachLineObj.x2)))

                        roll_z_adjust_y = math.degrees(
                            math.atan2((eachLineObj.z1 - eachLineObj.z2), (eachLineObj.y1 - eachLineObj.y2)))
                        # get the line z in point x, y as the new z of point
                        pnt_line_z = eachLineObj.z2

                    # pnt size
                    pnt_gj = eachLineObj.pipeSize
                    sht.cell(i, newzCol).value = pnt_line_z
                    sht.cell(i, writeCol).value = dir_pnt
                    sht.cell(i, gjSize).value = pnt_gj
                    sht.cell(i, rollZAdjust).value = roll_z_adjust
                    sht.cell(i, rollZAdjust_y).value = roll_z_adjust_y
                    step += 1

    else:
        _addError("Error --- the index of field 'STYLE_TYPE' from fiedl list, is not same"
                  "in poing xlsx file. check it!")
        raise NotPointStyleField

    resxlsx = os.path.join(os.path.dirname(pntXlsx), os.path.basename(pntXlsx).split(".xlsx")[0] + "_new.xlsx")
    wb.save(resxlsx)
    return resxlsx


# calculate gj to a number field
@logIt
def generateGJNumField(ply):
    _addField(ply, "gj_num_", "SHORT")
    codes = """def f(a):
    a = a.strip()
    try:
        if "X" in a:
            dataList = a.split("X")
            x, y = int(dataList[0]), int(dataList[1])
            res = max(x, y)
        else:
            a = int(a)
            if a < 20:
                res = 20
            else:
                res = a
        return res
    except:
        return 20"""
    arcpy.CalculateField_management(ply, "gj_num_", "f(!GJ!)", "PYTHON3", codes)


# copy points records depend on the field "STYLE_TYPE"
@logIt
def copyPointsInXlsx(pntxlsx):
    """
    usage: copy the records in pnt xlsx file depend on the field pnt_style(santong, sitong,
    guaidian, wantou).
    :param pntxlsx:
    :return:
    """
    wb = openpyxl.load_workbook(pntxlsx)
    sht = wb.active

    # get max row number and copy all points records on the end
    maxRow = sht.max_row
    maxCol = sht.max_column

    # set type field
    sht.cell(1, maxCol + 1).value = "PNT_TYPE_"
    sht.cell(1, maxCol + 2).value = "ROLL_Z_"
    sht.cell(1, maxCol + 3).value = "ROLL_X_"
    sht.cell(1, maxCol + 4).value = "new_z_fin_"
    sht.cell(1, maxCol + 5).value = "gj_gj_gj_"
    sht.cell(1, maxCol + 6).value = "roll_z_adjust_"
    sht.cell(1, maxCol + 7).value = "roll_z_adjust_y"

    styleFiledIndex = 0
    dir_1_index = 0
    dir_2_index = 0
    dir_3_index = 0
    dir_4_index = 0
    # z1_line_
    z_1_index = 0
    z_2_index = 0
    z_3_index = 0
    z_4_index = 0

    gj_line1_index = 0
    gj_line2_index = 0
    gj_line3_index = 0
    gj_line4_index = 0

    roll_z_adjust1_ = 0
    roll_z_adjust2_ = 0
    roll_z_adjust3_ = 0
    roll_z_adjust4_ = 0

    roll_z_adjust1_y = 0
    roll_z_adjust2_y = 0
    roll_z_adjust3_y = 0
    roll_z_adjust4_y = 0

    z_origin = 0
    # get the field of santong, sitong
    for colNum in range(1, maxCol + 1):
        if sht.cell(1, colNum).value == "STYLE_TYPE":
            styleFiledIndex = colNum
        elif sht.cell(1, colNum).value == "dir_1_":
            dir_1_index = colNum
        elif sht.cell(1, colNum).value == "dir_2_":
            dir_2_index = colNum
        elif sht.cell(1, colNum).value == "dir_3_":
            dir_3_index = colNum
        elif sht.cell(1, colNum).value == "dir_4_":
            dir_4_index = colNum
        elif sht.cell(1, colNum).value == "z1_line_":
            z_1_index = colNum
        elif sht.cell(1, colNum).value == "z2_line_":
            z_2_index = colNum
        elif sht.cell(1, colNum).value == "z3_line_":
            z_3_index = colNum
        elif sht.cell(1, colNum).value == "z4_line_":
            z_4_index = colNum
        elif sht.cell(1, colNum).value == "z_cen_":
            z_origin = colNum
        elif sht.cell(1, colNum).value == "gj_line1_":
            gj_line1_index = colNum
        elif sht.cell(1, colNum).value == "gj_line2_":
            gj_line2_index = colNum
        elif sht.cell(1, colNum).value == "gj_line3_":
            gj_line3_index = colNum
        elif sht.cell(1, colNum).value == "gj_line4_":
            gj_line4_index = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust1_":
            roll_z_adjust1_ = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust2_":
            roll_z_adjust2_ = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust3_":
            roll_z_adjust3_ = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust4_":
            roll_z_adjust4_ = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust1_y_":
            roll_z_adjust1_y = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust2_y_":
            roll_z_adjust2_y = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust3_y_":
            roll_z_adjust3_y = colNum
        elif sht.cell(1, colNum).value == "roll_z_adjust4_y_":
            roll_z_adjust4_y = colNum

    #
    for i, eachRow in enumerate(sht.rows):
        i += 1
        # pass the title row
        if i == 1:
            continue

        # set common z to ball model
        z1, z2, z3, z4 = (sht.cell(i, z_1_index).value, sht.cell(i, z_2_index).value,
                          sht.cell(i, z_3_index).value, sht.cell(i, z_4_index).value)
        gj1, gj2, gj3, gj4 = (sht.cell(i, gj_line1_index).value, sht.cell(i, gj_line2_index).value,
                              sht.cell(i, gj_line3_index).value, sht.cell(i, gj_line4_index).value)
        z_a1, z_a2, z_a3, z_a4 = (sht.cell(i, roll_z_adjust1_).value, sht.cell(i, roll_z_adjust2_).value,
                                  sht.cell(i, roll_z_adjust3_).value, sht.cell(i, roll_z_adjust4_).value)
        if z4 is not None:
            z1, z2, z3, z4 = (float(z1), float(z2), float(z3), float(z4))
            gj1, gj2, gj3, gj4 = (float(gj1), float(gj2), float(gj3), float(gj4))
            sht.cell(i, maxCol + 4).value = max(z1, z2, z3, z4)
            sht.cell(i, maxCol + 5).value = max(gj1, gj2, gj3, gj4)
            sht.cell(i, maxCol + 6).value = 0
            sht.cell(i, maxCol + 7).value = 0
        else:
            if z3 is not None:
                z1, z2, z3 = (float(z1), float(z2), float(z3))
                gj1, gj2, gj3 = (float(gj1), float(gj2), float(gj3))
                sht.cell(i, maxCol + 4).value = max(z1, z2, z3)
                sht.cell(i, maxCol + 5).value = max(gj1, gj2, gj3)
                sht.cell(i, maxCol + 6).value = 0
                sht.cell(i, maxCol + 7).value = 0
            else:
                if z2 is not None:
                    z1, z2 = (float(z1), float(z2))
                    gj1, gj2 = (float(gj1), float(gj2))
                    sht.cell(i, maxCol + 4).value = max(z1, z2)
                    sht.cell(i, maxCol + 5).value = max(gj1, gj2)
                    sht.cell(i, maxCol + 6).value = 0
                    sht.cell(i, maxCol + 7).value = 0
                else:
                    if z1 is not None:
                        z1 = float(z1)
                        gj1 = float(gj1)
                        sht.cell(i, maxCol + 4).value = z1
                        sht.cell(i, maxCol + 5).value = gj1
                        sht.cell(i, maxCol + 6).value = 0
                        sht.cell(i, maxCol + 7).value = 0
                    else:
                        sht.cell(i, maxCol + 4).value = float(z_origin)
                        sht.cell(i, maxCol + 4).value = 100

        # add point type to exlce
        sht.cell(i, maxCol + 1).value = 0
        sht.cell(i, maxCol + 2).value = 0
        sht.cell(i, maxCol + 3).value = 0

        # get the santong, sitong field value
        pointStyle = sht.cell(i, styleFiledIndex).value.strip()

        # copy two new rows, total 3 rows
        if pointStyle == "弯头" or pointStyle == "拐点":
            dir_1_data = sht.cell(i, dir_1_index).value
            dir_2_data = sht.cell(i, dir_2_index).value
            z_1_data = sht.cell(i, z_1_index).value
            z_2_data = sht.cell(i, z_2_index).value
            gj_1_data = sht.cell(i, gj_line1_index).value
            gj_2_data = sht.cell(i, gj_line2_index).value
            roll_z_1_data = sht.cell(i, roll_z_adjust1_).value
            roll_z_2_data = sht.cell(i, roll_z_adjust2_).value
            roll_z_1_data_y = sht.cell(i, roll_z_adjust1_y).value
            roll_z_2_data_y = sht.cell(i, roll_z_adjust2_y).value

            # set the end field first
            if dir_2_data is not None:
                new_row1 = maxRow + 1
                new_row2 = maxRow + 2
                maxRow += 2
                # set new point type to excel
                sht.cell(new_row1, maxCol + 1).value = 1
                sht.cell(new_row2, maxCol + 1).value = 1

                # set point roll-z
                sht.cell(new_row1, maxCol + 2).value = dir_1_data
                sht.cell(new_row2, maxCol + 2).value = dir_2_data

                # set point roll-x
                sht.cell(new_row1, maxCol + 3).value = 270
                sht.cell(new_row2, maxCol + 3).value = 270

                # set new z of point
                sht.cell(new_row1, maxCol + 4).value = z_1_data
                sht.cell(new_row2, maxCol + 4).value = z_2_data

                # set gj field
                sht.cell(new_row1, maxCol + 5).value = gj_1_data
                sht.cell(new_row2, maxCol + 5).value = gj_2_data

                # set roll_z_adjust field
                sht.cell(new_row1, maxCol + 6).value = roll_z_1_data
                sht.cell(new_row2, maxCol + 6).value = roll_z_2_data

                sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                sht.cell(new_row2, maxCol + 7).value = roll_z_2_data_y
                copyControl = 2
            else:
                if dir_1_data is not None:
                    new_row1 = maxRow + 1
                    maxRow += 1
                    # set new point type to excel
                    sht.cell(new_row1, maxCol + 1).value = 1

                    # set point roll-z
                    sht.cell(new_row1, maxCol + 2).value = dir_1_data

                    # set point roll-x
                    sht.cell(new_row1, maxCol + 3).value = 270

                    # set new z of point
                    sht.cell(new_row1, maxCol + 4).value = z_1_data

                    # set gj field
                    sht.cell(new_row1, maxCol + 5).value = gj_1_data

                    sht.cell(new_row1, maxCol + 6).value = roll_z_1_data

                    sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                    copyControl = 1
                else:
                    copyControl = 0

            # set new row's value
            if copyControl > 0:
                for eachCol in range(1, maxCol + 1):
                    data = sht.cell(i, eachCol).value

                    # get origin attributes from xlsx
                    if copyControl == 2:
                        sht.cell(new_row1, eachCol).value = data
                        sht.cell(new_row2, eachCol).value = data
                    elif copyControl == 1:
                        sht.cell(new_row1, eachCol).value = data
                    else:
                        pass
            else:
                pass


        # copy three new rows, total 4 rows
        elif pointStyle == "三通":
            dir_1_data = sht.cell(i, dir_1_index).value
            dir_2_data = sht.cell(i, dir_2_index).value
            dir_3_data = sht.cell(i, dir_3_index).value
            z_1_data = sht.cell(i, z_1_index).value
            z_2_data = sht.cell(i, z_2_index).value
            z_3_data = sht.cell(i, z_3_index).value
            gj_1_data = sht.cell(i, gj_line1_index).value
            gj_2_data = sht.cell(i, gj_line2_index).value
            gj_3_data = sht.cell(i, gj_line3_index).value
            roll_z_1_data = sht.cell(i, roll_z_adjust1_).value
            roll_z_2_data = sht.cell(i, roll_z_adjust2_).value
            roll_z_3_data = sht.cell(i, roll_z_adjust3_).value

            roll_z_1_data_y = sht.cell(i, roll_z_adjust1_y).value
            roll_z_2_data_y = sht.cell(i, roll_z_adjust2_y).value
            roll_z_3_data_y = sht.cell(i, roll_z_adjust3_y).value

            # set the end field first
            if dir_3_data is not None:
                new_row1 = maxRow + 1
                new_row2 = maxRow + 2
                new_row3 = maxRow + 3
                maxRow += 3
                # set new point type to excel
                sht.cell(new_row1, maxCol + 1).value = 1
                sht.cell(new_row2, maxCol + 1).value = 1
                sht.cell(new_row3, maxCol + 1).value = 1

                # set point roll-z
                sht.cell(new_row1, maxCol + 2).value = dir_1_data
                sht.cell(new_row2, maxCol + 2).value = dir_2_data
                sht.cell(new_row3, maxCol + 2).value = dir_3_data

                # set point roll-x
                sht.cell(new_row1, maxCol + 3).value = 270
                sht.cell(new_row2, maxCol + 3).value = 270
                sht.cell(new_row3, maxCol + 3).value = 270

                # set new z of point
                sht.cell(new_row1, maxCol + 4).value = z_1_data
                sht.cell(new_row2, maxCol + 4).value = z_2_data
                sht.cell(new_row3, maxCol + 4).value = z_3_data

                # set gj field
                sht.cell(new_row1, maxCol + 5).value = gj_1_data
                sht.cell(new_row2, maxCol + 5).value = gj_2_data
                sht.cell(new_row3, maxCol + 5).value = gj_3_data

                sht.cell(new_row1, maxCol + 6).value = roll_z_1_data
                sht.cell(new_row2, maxCol + 6).value = roll_z_2_data
                sht.cell(new_row3, maxCol + 6).value = roll_z_3_data

                sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                sht.cell(new_row2, maxCol + 7).value = roll_z_2_data_y
                sht.cell(new_row3, maxCol + 7).value = roll_z_3_data_y
                copyControl = 3
            else:
                if dir_2_data is not None:
                    new_row1 = maxRow + 1
                    new_row2 = maxRow + 2
                    maxRow += 2
                    # set new point type to excel
                    sht.cell(new_row1, maxCol + 1).value = 1
                    sht.cell(new_row2, maxCol + 1).value = 1

                    # set point roll-z
                    sht.cell(new_row1, maxCol + 2).value = dir_1_data
                    sht.cell(new_row2, maxCol + 2).value = dir_2_data

                    # set point roll-x
                    sht.cell(new_row1, maxCol + 3).value = 270
                    sht.cell(new_row2, maxCol + 3).value = 270

                    # set new z of point
                    sht.cell(new_row1, maxCol + 4).value = z_1_data
                    sht.cell(new_row2, maxCol + 4).value = z_2_data

                    # set gj field
                    sht.cell(new_row1, maxCol + 5).value = gj_1_data
                    sht.cell(new_row2, maxCol + 5).value = gj_2_data

                    sht.cell(new_row1, maxCol + 6).value = roll_z_1_data
                    sht.cell(new_row2, maxCol + 6).value = roll_z_2_data

                    sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                    sht.cell(new_row2, maxCol + 7).value = roll_z_2_data_y
                    copyControl = 2
                else:
                    if dir_3_data is not None:
                        new_row1 = maxRow + 1
                        maxRow += 1
                        # set new point type to excel
                        sht.cell(new_row1, maxCol + 1).value = 1
                        # set point roll-z
                        sht.cell(new_row1, maxCol + 2).value = dir_1_data
                        # set point roll-x
                        sht.cell(new_row1, maxCol + 3).value = 270
                        # set new z of point
                        sht.cell(new_row1, maxCol + 4).value = z_1_data
                        # set gj field
                        sht.cell(new_row1, maxCol + 5).value = gj_1_data

                        sht.cell(new_row1, maxCol + 6).value = roll_z_1_data

                        sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                        copyControl = 1
                    else:
                        copyControl = 0

            if copyControl > 0:
                # set new row's value
                for eachCol in range(1, maxCol + 1):
                    data = sht.cell(i, eachCol).value

                    # get origin attributes from xlsx
                    if copyControl == 3:
                        sht.cell(new_row1, eachCol).value = data
                        sht.cell(new_row2, eachCol).value = data
                        sht.cell(new_row3, eachCol).value = data
                    elif copyControl == 2:
                        sht.cell(new_row1, eachCol).value = data
                        sht.cell(new_row2, eachCol).value = data
                    elif copyControl == 1:
                        sht.cell(new_row1, eachCol).value = data

        # copy four new rows, total 5 rows
        elif pointStyle == "四通":
            dir_1_data = sht.cell(i, dir_1_index).value
            dir_2_data = sht.cell(i, dir_2_index).value
            dir_3_data = sht.cell(i, dir_3_index).value
            dir_4_data = sht.cell(i, dir_4_index).value

            z_1_data = sht.cell(i, z_1_index).value
            z_2_data = sht.cell(i, z_2_index).value
            z_3_data = sht.cell(i, z_3_index).value
            z_4_data = sht.cell(i, z_4_index).value

            gj_1_data = sht.cell(i, gj_line1_index).value
            gj_2_data = sht.cell(i, gj_line2_index).value
            gj_3_data = sht.cell(i, gj_line3_index).value
            gj_4_data = sht.cell(i, gj_line4_index).value

            roll_z_1_data = sht.cell(i, roll_z_adjust1_).value
            roll_z_2_data = sht.cell(i, roll_z_adjust2_).value
            roll_z_3_data = sht.cell(i, roll_z_adjust3_).value
            roll_z_4_data = sht.cell(i, roll_z_adjust4_).value

            roll_z_1_data_y = sht.cell(i, roll_z_adjust1_y).value
            roll_z_2_data_y = sht.cell(i, roll_z_adjust2_y).value
            roll_z_3_data_y = sht.cell(i, roll_z_adjust3_y).value
            roll_z_4_data_y = sht.cell(i, roll_z_adjust4_y).value

            # set the end field first
            if dir_4_data is not None:
                new_row1 = maxRow + 1
                new_row2 = maxRow + 2
                new_row3 = maxRow + 3
                new_row4 = maxRow + 4
                maxRow += 4

                # set new point type to excel
                sht.cell(new_row1, maxCol + 1).value = 1
                sht.cell(new_row2, maxCol + 1).value = 1
                sht.cell(new_row3, maxCol + 1).value = 1
                sht.cell(new_row4, maxCol + 1).value = 1

                # set point roll-z
                sht.cell(new_row1, maxCol + 2).value = dir_1_data
                sht.cell(new_row2, maxCol + 2).value = dir_2_data
                sht.cell(new_row3, maxCol + 2).value = dir_3_data
                sht.cell(new_row4, maxCol + 2).value = dir_4_data

                # set point roll-x
                sht.cell(new_row1, maxCol + 3).value = 270
                sht.cell(new_row2, maxCol + 3).value = 270
                sht.cell(new_row3, maxCol + 3).value = 270
                sht.cell(new_row4, maxCol + 3).value = 270

                # set new z of point
                sht.cell(new_row1, maxCol + 4).value = z_1_data
                sht.cell(new_row2, maxCol + 4).value = z_2_data
                sht.cell(new_row3, maxCol + 4).value = z_3_data
                sht.cell(new_row4, maxCol + 4).value = z_4_data

                # set gj field
                sht.cell(new_row1, maxCol + 5).value = gj_1_data
                sht.cell(new_row2, maxCol + 5).value = gj_2_data
                sht.cell(new_row3, maxCol + 5).value = gj_3_data
                sht.cell(new_row4, maxCol + 5).value = gj_4_data

                sht.cell(new_row1, maxCol + 6).value = roll_z_1_data
                sht.cell(new_row2, maxCol + 6).value = roll_z_2_data
                sht.cell(new_row3, maxCol + 6).value = roll_z_3_data
                sht.cell(new_row4, maxCol + 6).value = roll_z_4_data

                sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                sht.cell(new_row2, maxCol + 7).value = roll_z_2_data_y
                sht.cell(new_row3, maxCol + 7).value = roll_z_3_data_y
                sht.cell(new_row4, maxCol + 7).value = roll_z_4_data_y
                copyControl = 4
            else:
                # set the end field first
                if dir_3_data is not None:
                    new_row1 = maxRow + 1
                    new_row2 = maxRow + 2
                    new_row3 = maxRow + 3
                    maxRow += 3
                    # set new point type to excel
                    sht.cell(new_row1, maxCol + 1).value = 1
                    sht.cell(new_row2, maxCol + 1).value = 1
                    sht.cell(new_row3, maxCol + 1).value = 1

                    # set point roll-z
                    sht.cell(new_row1, maxCol + 2).value = dir_1_data
                    sht.cell(new_row2, maxCol + 2).value = dir_2_data
                    sht.cell(new_row3, maxCol + 2).value = dir_3_data

                    # set point roll-x
                    sht.cell(new_row1, maxCol + 3).value = 270
                    sht.cell(new_row2, maxCol + 3).value = 270
                    sht.cell(new_row3, maxCol + 3).value = 270

                    # set new z of point
                    sht.cell(new_row1, maxCol + 4).value = z_1_data
                    sht.cell(new_row2, maxCol + 4).value = z_2_data
                    sht.cell(new_row3, maxCol + 4).value = z_3_data

                    # set gj field
                    sht.cell(new_row1, maxCol + 5).value = gj_1_data
                    sht.cell(new_row2, maxCol + 5).value = gj_2_data
                    sht.cell(new_row3, maxCol + 5).value = gj_3_data

                    sht.cell(new_row1, maxCol + 6).value = roll_z_1_data
                    sht.cell(new_row2, maxCol + 6).value = roll_z_2_data
                    sht.cell(new_row3, maxCol + 6).value = roll_z_3_data

                    sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                    sht.cell(new_row2, maxCol + 7).value = roll_z_2_data_y
                    sht.cell(new_row3, maxCol + 7).value = roll_z_3_data_y
                    copyControl = 3
                else:
                    if dir_2_data is not None:
                        new_row1 = maxRow + 1
                        new_row2 = maxRow + 2
                        maxRow += 2
                        # set new point type to excel
                        sht.cell(new_row1, maxCol + 1).value = 1
                        sht.cell(new_row2, maxCol + 1).value = 1

                        # set point roll-z
                        sht.cell(new_row1, maxCol + 2).value = dir_1_data
                        sht.cell(new_row2, maxCol + 2).value = dir_2_data

                        # set point roll-x
                        sht.cell(new_row1, maxCol + 3).value = 270
                        sht.cell(new_row2, maxCol + 3).value = 270

                        # set new z of point
                        sht.cell(new_row1, maxCol + 4).value = z_1_data
                        sht.cell(new_row2, maxCol + 4).value = z_2_data

                        # set gj field
                        sht.cell(new_row1, maxCol + 5).value = gj_1_data
                        sht.cell(new_row2, maxCol + 5).value = gj_2_data

                        sht.cell(new_row1, maxCol + 6).value = roll_z_1_data
                        sht.cell(new_row2, maxCol + 6).value = roll_z_2_data

                        sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                        sht.cell(new_row2, maxCol + 7).value = roll_z_2_data_y
                        copyControl = 2
                    else:
                        if dir_1_data is not None:
                            new_row1 = maxRow + 1
                            maxRow += 1
                            # set new point type to excel
                            sht.cell(new_row1, maxCol + 1).value = 1
                            # set point roll-z
                            sht.cell(new_row1, maxCol + 2).value = dir_1_data
                            # set point roll-x
                            sht.cell(new_row1, maxCol + 3).value = 270
                            # set new z of point
                            sht.cell(new_row1, maxCol + 4).value = z_1_data
                            # set gj field
                            sht.cell(new_row1, maxCol + 5).value = gj_1_data

                            sht.cell(new_row1, maxCol + 6).value = roll_z_1_data

                            sht.cell(new_row1, maxCol + 7).value = roll_z_1_data_y
                            copyControl = 1
                        else:
                            copyControl = 0

            if copyControl > 0:
                # set new row's value
                for eachCol in range(1, maxCol + 1):
                    data = sht.cell(i, eachCol).value

                    # get origin attributes from xlsx
                    if copyControl == 4:
                        sht.cell(new_row1, eachCol).value = data
                        sht.cell(new_row2, eachCol).value = data
                        sht.cell(new_row3, eachCol).value = data
                        sht.cell(new_row4, eachCol).value = data
                    elif copyControl == 3:
                        sht.cell(new_row1, eachCol).value = data
                        sht.cell(new_row2, eachCol).value = data
                        sht.cell(new_row3, eachCol).value = data
                    elif copyControl == 2:
                        sht.cell(new_row1, eachCol).value = data
                        sht.cell(new_row2, eachCol).value = data
                    elif copyControl == 1:
                        sht.cell(new_row1, eachCol).value = data

    delZCen, delZ1Line, delPntType, delRollZ, delRollZ1 = None, None, None, None, None
    # delete temp fields
    maxCol = sht.max_column
    for i in range(1, maxCol + 1):
        value = sht.cell(1, i).value
        if value == "z1_line_":
            delZ1Line = i
        elif value == "PNT_TYPE_":
            delPntType = i

    if delZ1Line and delPntType:
        sht.delete_cols(delZ1Line, (delPntType - delZ1Line))

    maxCol = sht.max_column
    for i in range(1, maxCol + 1):
        value = sht.cell(1, i).value
        if value == "z_cen_":
            delZCen = i

    if delZCen:
        sht.delete_cols(delZCen, 1)

    # maxCol = sht.max_column
    # for i in range(1, maxCol + 1):
    #     value = sht.cell(1, i).value
    #     if value == "roll_z_adjust_":
    #         delRollZ1 = i
    #
    # if delRollZ1:
    #     sht.delete_cols(delRollZ1, 1)
    #
    # maxCol = sht.max_column
    # for i in range(1, maxCol + 1):
    #     value = sht.cell(1, i).value
    #     if value == "roll_z_adjust_2":
    #         delRollZ1 = i
    #
    # if delRollZ1:
    #     sht.delete_cols(delRollZ1, 1)

    basedir = os.path.dirname(pntxlsx)
    basename = os.path.basename(pntxlsx)
    newdir = os.path.join(basedir, "res_xlsx")
    if not os.path.exists(newdir):
        os.makedirs(newdir)
    wb.save(os.path.join(newdir, basename))
    return os.path.join(newdir, basename)


# get all point data with the attribute of santong, sitong, wantou
def selectTargetData(inGDB, outputGDB):
    arcpy.env.workspace = inGDB

    outputdir = os.path.dirname(outputGDB)
    outputname = os.path.basename(outputGDB)
    # make sure the target GDB is exists
    if not arcpy.Exists(outputGDB):
        arcpy.CreateFileGDB_management(outputdir,
                                       outputname)

    dataList = arcpy.ListFeatureClasses()
    for each in dataList:
        # select data
        if each.split("_")[-1] == "POINT":
            outFC = arcpy.Select_analysis(each, os.path.join(outputGDB, each),
                                          "TZD LIKE '%三通%' Or TZD LIKE '%四通%' Or TZD LIKE '%弯头%' Or TZD LIKE '%拐点%'")
        elif each.split("_")[-1] == "LINE":
            arcpy.CopyFeatures_management(each, os.path.join(outputGDB, each))

    # reset out put gdb
    arcpy.env.workspace = outputGDB

    pntList = arcpy.ListFeatureClasses("*_POINT")
    plyList = arcpy.ListFeatureClasses("*_LINE")

    resList = []
    for eachpnt in pntList:
        basename = eachpnt[:-6]
        dataindex = plyList.index(basename + "_LINE")
        dataTup = (eachpnt, plyList[dataindex])
        resList.append(dataTup)

    return resList


def f(rollz, rollx, zx, zy):
    # dir - x
    if -45 <= rollz <= 45 or 135 <= rollz <= 180 or -180 <= rollz <= -135:
        if 0 <= zx <= 90:
            res = rollx + zx
        elif 90 < zx <= 180:
            res = 180 - zx + rollx
        elif 0 > zx >= -90:
            res = rollx + zx
        elif -180 <= zx < -90:
            res = rollx - (180 + zx)
    else:
        if 0 <= zy <= 90:
            res = rollx + zy
        elif 90 < zy <= 180:
            res = 180 - zy + rollx
        elif 0 > zy >= -90:
            res = rollx + zy
        elif -180 <= zy < -90:
            res = rollx - (180 + zy)
    return res


def generateNewRoll_X(pntxlsx):
    wb = openpyxl.load_workbook(pntxlsx)
    sht = wb.active

    roll_z_index = 0
    roll_x_index = 0
    roll_x_adju_x_index = 0
    roll_x_adju_y_index = 0

    maxCol = sht.max_column
    for i, eachCol in enumerate(sht.columns):
        i += 1
        if sht.cell(1, i).value == "ROLL_Z_":
            roll_z_index = i
        elif sht.cell(1, i).value == "ROLL_X_":
            roll_x_index = i
        elif sht.cell(1, i).value == "roll_z_adjust_":
            roll_x_adju_x_index = i
        elif sht.cell(1, i).value == "roll_z_adjust_y":
            roll_x_adju_y_index = i

    for i, eachRow in enumerate(sht.rows):
        i += 1
        if i == 1:
            pass

        roll_z_data = sht.cell(i, roll_z_index).value
        roll_x_data = sht.cell(i, roll_x_index).value
        roll_x_adju_x_data = sht.cell(i, roll_x_adju_x_index).value
        roll_x_adju_y_data = sht.cell(i, roll_x_adju_y_index).value
        # try:
        roll_z_data, roll_x_data, roll_x_adju_x_data, roll_x_adju_y_data = (
            float(roll_z_data), float(roll_x_data), float(roll_x_adju_x_data),
            float(roll_x_adju_y_data))
        res = f(roll_z_data, roll_x_data, roll_x_adju_x_data, roll_x_adju_y_data)
        sht.cell(i, maxCol + 1).value = res
        # except:
        #     sht.cell(i, maxCol + 1).value = "None"

    wb.save(os.path.join(os.path.dirname(pntxlsx),
                         os.path.basename(pntxlsx).strip(".xlsx") + "_res.xlsx"))

    return os.path.join(os.path.dirname(pntxlsx),
                        os.path.basename(pntxlsx).strip(".xlsx") + "_res.xlsx")


@logIt
def xlsxToPointFC(xlsx, sr, outputPath):
    resData = os.path.join(outputPath, os.path.basename(xlsx).strip(".xlsx"))
    tv = arcpy.MakeTableView_management(xlsx + "\Sheet$", "tv_tmp")
    xyLyr = arcpy.MakeXYEventLayer_management(tv, "x_cen_", "y_cen_",
                                              os.path.basename(xlsx).strip(".xlsx") + "_temp",
                                              sr, "new_z_fin_")
    arcpy.CopyFeatures_management(xyLyr, resData)
    addNewRollXField(resData)
    return resData


@logIt
def addNewRollXField(pntFC):
    _addField(pntFC, "ROLL_X_N_", "DOUBLE")
    codes = """def f(rollz, rollx, zx, zy):
    # dir - x
    if -45 <= rollz <= 45 or 135 <= rollz <= 180 or -180 <= rollz <= -135:
        if 0 <= zx <= 90:
            res = rollx + zx
        elif 90 < zx <= 180:
            res = 180 - zx + rollx
        elif 0 > zx >= -90:
            res = rollx + zx
        elif -180 <= zx < -90:
            res = rollx - (180 + zx)
    else:
        if 0 <= zy <= 90:
            res = rollx + zy
        elif 90 < zy <= 180:
            res = 180 - zy + rollx
        elif 0 > zy >= -90:
            res = rollx + zy
        elif -180 <= zy < -90:
            res = rollx - (180 + zy)
    return res"""
    arcpy.CalculateField_management(pntFC, "ROLL_X_N_", "f(!ROLL_Z_!, !ROLL_X_!, !roll_z_adjust_!, !roll_z_adjust_y!)",
                                    "PYTHON3", codes)



@getRunTime
@logIt
def main(pnt, ply, outputPath, outputGDB, sr):
    global pntFieldList, plyFieldList, plyTotalExtent

    logging.info(f" === Progress Step2 Initial Field === ")
    # delete the field will be added later, make sure the field "unic_id_" is the last in fields
    _deleteFields(pnt, "point")
    _deleteFields(ply, "line")

    logging.info(f" === Progress Step3 Get Pipe Size === ")
    # add a gj_num_ field to ply, it will be include into lineObject
    generateGJNumField(ply)

    logging.info(f" === Progress Step4 Convert Data To Xlsx === ")
    # convert ply and pnt table to xlsx
    pntXls = _convertToXlsx(pnt, outputPath, "point", convertAll=True)
    plyXls = _convertToXlsx(ply, outputPath, "line", convertAll=True)

    logging.info(f" === Progress Step5 Generate Line Object === ")
    # generate line object with spatial index
    plgObjList = _generateLineObj(plyXls, plyFieldList, plyTotalExtent)

    logging.info(f" === Progress Step6 Get Line Directory To Point === ")
    # add directory data to pnt xlsx
    pntxlsx = _getDirectory(pntXls, plgObjList, plyTotalExtent)

    logging.info(f" === Progress Step7 Add X, Y, Z Rotation === ")
    # copy all datas to make feature class
    pnt_res_xlsx = copyPointsInXlsx(pntxlsx)

    # # calculate new roll_x as the field of roll_z_adjust_y
    # # give up this way
    # final_pnt = generateNewRoll_X(pnt_res_xlsx)

    logging.info(f" === Progress Step8 Generate Point Feature Class === ")
    # generate point feature class from xlsx
    resPnt = xlsxToPointFC(pnt_res_xlsx, sr, outputGDB)

    # add data to aprx
    

    # todo add to arcgis pro and auto apply the symbol of .dae

    logging.info(f" === Progress Step9 Clear Temp Xlsx")
    try:
        os.remove(pntXls)
        os.remove(plyXls)
        os.remove(pntxlsx)
        # os.remove(pnt_res_xlsx)
    except:
        logging.info(f" +++ Delete Temp Xlsx Faild +++ ")


spatialIndex = (1, 1)
pntFieldList = []
plyFieldList = []
plyTotalExtent = None
tolerance = 0.1

logging.info(f" === Common Para: \n == spatialIndex: {spatialIndex} \n"
             f" == tolerance: {tolerance} \n")

inGDB = r"E:\南京工具\工具交活整理_0916\测试数据\工具3_三通四通\二维测试数据.gdb"
outputGDB = r"E:\南京工具\工具交活整理_0916\测试数据\工具3_三通四通\res2\管点数据.gdb"
outputPath = r"E:\南京工具\工具交活整理_0916\测试数据\工具3_三通四通\res2"
sr = arcpy.SpatialReference(4549)

logging.info(f" === Input Para: \n == inGDB: {inGDB} \n == outputGDB: {outputGDB} \n"
             f" == outputPath: {outputPath} \n == sr: {sr} \n")

logging.info(f" === Progress Step1 Select target data from origin gdb to target gdb ===")
resdataTupleList = selectTargetData(inGDB, outputGDB)

logging.info(f" === Progress Step1 Finished, the data selected : {resdataTupleList} ===")

totalNum = len(resdataTupleList)

arcpy.SetProgressor("step", "Data Progressing It Will Take A Long Time, Please Wait...",
                    1, totalNum)

logging.info(" === Start Looping Data Selected ===")
for eachtup in resdataTupleList:
    print(f"{resdataTupleList.index(eachtup) + 1}/{totalNum}")
    logging.info(f"Now Is Progressing data: {eachtup}, "
                 f"{resdataTupleList.index(eachtup) + 1}/{totalNum}")
    pntData, plyData = eachtup
    main(pntData, plyData, outputPath, outputGDB, sr)
    arcpy.SetProgressorPosition()

