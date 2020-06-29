import openpyxl
import arcpy

arcpy.env.overwriteOutput = True


# xlsx = r'E:\空管\数据制作\生成面\扇区.xlsx'
# resxlsx = r'E:\空管\数据制作\生成面\扇区_process.xlsx'
#
# wb = openpyxl.load_workbook(xlsx)
# print(wb.sheetnames)
# sttrSht = wb['attr']
# geoSht = wb['geo']
# print(sttrSht)
# print(geoSht)
# for i in range(1, sttrSht.max_row):
#     data = sttrSht['W%s' % i]
#     data1 = sttrSht['Y%s' % i]
#     if data.value == None:
#         data.value = 90000
#     if data1.value == None:
#         data1.value = 0
# wb.save(resxlsx)

resxlsx = r'E:\空管\数据制作\生成面\扇区_process.xlsx'

wb = openpyxl.load_workbook(resxlsx)
attrSht = wb['attr']
geoSht = wb['geo']


uniId = []
feature = []
plg = []
idset = set()
maxrow = geoSht.max_row
idLst = []
for i, eachRow in enumerate(geoSht.rows):
    if i == 0:
        continue
    id = eachRow[2].value
    lon = eachRow[3].value
    lat = eachRow[4].value
    sortid = eachRow[6].value
    idset.add(id)
    # print(id, sortid)
    if id in idLst:
        pass
    else:
        idLst.append(id)

    # convert dms to dd
    lons = str(lon)
    if '.' in lons:
        lonssecAfterPnt = lons.split('.')[1]
        lons = lons.split('.')[0]
    else:
        lonssecAfterPnt = 0
    lonssec = lons[-2:]
    lonsmin = lons[-4:-2]
    lonsdeg = lons[:-4]
    newLon = (float(lonssec) + float(lonssecAfterPnt)) / 3600 + float(lonsmin) / 60 + float(lonsdeg)

    lats = str(lat)
    if '.' in lats:
        latssecAfterPnt = lats.split('.')[1]
        lats = lats.split('.')[0]
    else:
        latssecAfterPnt = 0
    latssec = lats[-2:]
    latsmin = lats[-4:-2]
    latsdeg = lats[:-4]
    newLat = (float(latssec) + float(latssecAfterPnt)) / 3600 + float(latsmin) / 60 + float(latsdeg)

    coord = [newLon, newLat]

    feature.append(coord)

    # new polygon
    if sortid == 1:
        if i == 1:
            continue
        plg.append(feature[:-1])
        feature = []

    if i+1 >= maxrow:
        plg.append(feature)

    if id == 1229:
        print(feature)

print(plg)

plgShp = arcpy.CreateFeatureclass_management(r'E:\空管\数据制作\生成面\shp\gdb.gdb', 'plg', 'POLYGON')
arcpy.AddField_management(plgShp, 'CONTROL_SHAPE_ID', 'TEXT')

arcpy.Array([arcpy.Point(*coords) for coords in feature])

with arcpy.da.InsertCursor(plgShp, ['SHAPE@', 'CONTROL_SHAPE_ID']) as cur:
    for i, eachPlg in enumerate(plg):
        # idLst
        array = arcpy.Array([arcpy.Point(*coords) for coords in eachPlg])
        plgFea = arcpy.Polygon(array)
        cur.insertRow([plgFea, idLst[i]])

sr = arcpy.SpatialReference(4326)
arcpy.DefineProjection_management(plgShp, sr)