import os
import openpyxl
import json
import requests
from bs4 import BeautifulSoup
import bs4
import time
import datetime
import functools
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def getRunTime(func):
    @functools.wraps(func)
    def _wrapper(*args, **kwargs):
        start = datetime.datetime.now()
        print(f"Start at: {start}")
        res = func(*args, **kwargs)
        end = datetime.datetime.now()
        cost = end - start
        print(f"Start at: {start}")
        print(f"Finish at: {end}")
        print(f"Total cost at: {cost}")
        return res

    return _wrapper


def getTargetServicesUrl(serTarName: str, baseurl: str) -> dict:
    # target base url
    # tarUrl = "http://192.168.2.242"
    tarUrl = "http://" + baseurl.split("/")[2]

    # send get requests
    res = requests.get(baseurl)

    # make sure there are a vailable response
    res.raise_for_status()

    # parse html text
    soup = BeautifulSoup(res.text, "lxml")
    ul = soup.ul

    totalUrlJson = {}
    for li in ul.children:
        if isinstance(li, bs4.element.Tag):
            for subTag in li.children:
                # get <a> tag
                if isinstance(subTag, bs4.element.Tag):
                    tarTotalUrl = tarUrl + subTag.attrs["href"]
                    serviceName = subTag.attrs["href"].split("/")[-2]
                    serviceType = subTag.attrs["href"].split("/")[-1]
                    # filter data from tarSerName
                    for eachSerName in serTarName:
                        if eachSerName in serviceName:
                            unicSerName = subTag.attrs["href"].split("/")[-2] + "." + subTag.attrs["href"].split("/")[
                                -1]
                            totalUrlJson[unicSerName] = [serviceName, serviceType, tarTotalUrl]
    # print(totalUrlJson)
    return totalUrlJson


# enrich services messages
def enrichMessage(urlDict: dict) -> dict:
    assert isinstance(urlDict, dict), "the data input in method 'enrichMessage' is not dict"
    resDict = urlDict

    for key, value in urlDict.items():
        serviceType = value[1]

        if serviceType != "SceneServer":
            serUrl = value[2]

            try:
                # gentlemanly reptiles
                time.sleep(0.1)

                # send get requests
                res = requests.get(serUrl)
                res.raise_for_status()

                soup = BeautifulSoup(res.text, "lxml")
                ul = soup.findAll("ul")

                fullExtUl = ul[2]

                dataExtent = []
                sr = ""
                n = 1
                for sub in fullExtUl.children:
                    if isinstance(sub, bs4.element.NavigableString):
                        # print(sub)
                        if n < 5:
                            subStr = float(sub.replace("\n", "").split(":")[1].strip())
                            dataExtent.append(subStr)
                        if n == 5:
                            sr = int(sub.replace("\n", "").split(":")[1].split("\r")[0].strip())
                        n += 1
                resDict[key].append(dataExtent)
                center = [(dataExtent[0] + dataExtent[2]) / 2, (dataExtent[1] + dataExtent[3]) / 2]
                resDict[key].append(center)
                resDict[key].append(sr)
            except BaseException as e:
                print(e, key)

        # scene services
        else:
            serUrl = value[2]

            try:
                # gentlemanly reptiles
                time.sleep(0.1)

                # send get requests
                res = requests.get(serUrl)
                res.raise_for_status()

                data = json.loads(res.text)

                # the scene layer of 3d object
                try:
                    dataExtent = data["layers"][0]["store"]["extent"]
                    sr = data["layers"][0]["spatialReference"]["wkid"]
                    resDict[key].append(dataExtent)
                    center = [(dataExtent[0] + dataExtent[2]) / 2, (dataExtent[1] + dataExtent[3]) / 2]
                    resDict[key].append(center)
                    resDict[key].append(sr)
                except:
                    # the bim building scene layer
                    dataExtent = data["layers"][0]["fullExtent"]
                    sr = data["layers"][0]["fullExtent"]["spatialReference"]["wkid"]
                    # print(sr)
                    dataExtent = [dataExtent["xmin"], dataExtent["ymin"], dataExtent["xmax"], dataExtent["ymax"]]

                    resDict[key].append(dataExtent)
                    center = [(dataExtent[0] + dataExtent[2]) / 2, (dataExtent[1] + dataExtent[3]) / 2]
                    resDict[key].append(center)
                    resDict[key].append(sr)

            except BaseException as e:
                print(e, key)

    # print(resDict)
    return resDict


# write information to xlsx
def writeToXlsx(enrichedUrlDictData, serChinese_Key, serChinese_NameMappingFile, outputXlsx):
    global publisher, publishTime
    assert isinstance(enrichedUrlDictData, dict)

    wb = openpyxl.Workbook()
    sht = wb.active

    row1 = ["服务名", "服务含义", "服务类型", "坐标系", "中心点", "最小包络矩形", "url", "发布人", "发布时间"]

    # set xlsx title
    for i, eachCol in enumerate(row1):
        i += 1
        sht.cell(1, i).value = eachCol

    # add chinese serviecs name
    nameMappingData = {}
    if serChinese_Key:
        with open(serChinese_NameMappingFile, encoding="utf-8") as f:
            datas = f.read()
        nameMappingData = json.loads(datas)

    # add other datas
    unicName = ""
    for i, (key, value) in enumerate(enrichedUrlDictData.items()):
        try:
            i += 1
            # print(value)
            [serName, serCnName, serType, \
             serSr, serCen, serExt, serUrl] = (value[0], nameMappingData.get(value[0]), value[1],
                                               value[5], value[4], value[3], value[2])
            dataList = [serName, serCnName, serType, serSr, serCen, serExt, serUrl, publisher, publishTime]
            for j, eachCol in enumerate(row1):
                j += 1
                sht.cell(i + 1, j).value = str(dataList[j - 1])
        except:
            i -= 1
            pass

    wb.save(outputXlsx)
    return outputXlsx


def formatXlsx(xlsxData, outputXlsx):
    wb = openpyxl.load_workbook(xlsxData)
    sht = wb.active

    serName = ""
    mergeRows = []
    row_max = sht.max_row
    col_max = sht.max_column
    for i, eachRow in enumerate(sht.rows):
        i += 1
        if i == 1:
            continue

        serName_newRow = sht.cell(i, 1).value
        # statistics merge cells
        if serName == serName_newRow:
            mergeRows.append(i)

        else:
            serName = serName_newRow
            if len(mergeRows) > 0:
                minRow = min(mergeRows)
                maxRow = max(mergeRows)
                merCol = [1, 2, 4, 5, 6]
                # print(minRow, maxRow)
                for col in merCol:
                    sht.merge_cells(start_row=minRow, start_column=col, end_row=maxRow, end_column=col)
            mergeRows = [i]

        if i == row_max:
            serName = serName_newRow
            if len(mergeRows) > 0:
                minRow = min(mergeRows)
                maxRow = max(mergeRows)
                merCol = [1, 2, 4, 5, 6]
                # print(minRow, maxRow)
                for col in merCol:
                    sht.merge_cells(start_row=minRow, start_column=col, end_row=maxRow, end_column=col)
            mergeRows = [i]

    # merge the column of 8, 9
    sht.merge_cells(start_row=2, start_column=8, end_row=row_max, end_column=8)
    sht.merge_cells(start_row=2, start_column=9, end_row=row_max, end_column=9)

    # modify column width
    columnList = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
    columnWidthList = [21, 22, 18, 15, 21, 24, 85, 15, 15]
    for i, eachCol in enumerate(columnList):
        sht.column_dimensions[eachCol].width = columnWidthList[i]

    # modify row high
    for i in range(1, row_max + 1):
        sht.row_dimensions[i].height = 30

    # set word in center, word color and other styles
    font_green = Font(color="0099CC00")
    font_blue = Font(color="0099CCFF")
    font_orange = Font(color="00FFCC99")
    for eachRow in sht.rows:
        for eachCell in eachRow:
            eachCell.alignment = Alignment(horizontal="center", vertical="center")
            if eachCell.value == "FeatureServer":
                eachCell.font = font_green
            elif eachCell.value == "MapServer":
                eachCell.font = font_blue
            elif eachCell.value == "SceneServer":
                eachCell.font = font_orange

    pFill = PatternFill("solid", fgColor="00DCE6F1")
    pBorder = Border(top=Side(border_style='thin', color="00538DD5"),
                     left=Side(border_style='thin', color="00538DD5"),
                     bottom=Side(border_style='thin', color="00538DD5"),
                     right=Side(border_style='thin', color="00538DD5"))
    for i in range(1, col_max + 1):
        sht.cell(1, i).fill = pFill
        sht.cell(1, i).border = pBorder

    wb.save(outputXlsx)


@getRunTime
def main(serTarName, baseurl, serChinese_Key, outputXlsx, serChinese_NameMappingFile=None):
    urlDict = getTargetServicesUrl(serTarName, baseurl)

    enrichedDict = enrichMessage(urlDict)

    # add chinese name to dict data
    proXlsx = writeToXlsx(enrichedDict, serChinese_Key, serChinese_NameMappingFile, outputXlsx)

    # format the xlsx data
    reultXlsx = outputXlsx.split(".xlsx")[0] + "_res.xlsx"
    formatXlsx(proXlsx, reultXlsx)


# key word, used to filter the services by it's name
serTarName = [""]

# the url of services folder's name
baseurl = "http://192.168.2.33/server/rest/services/Hosted"
# baseurl = "http://192.168.2.242/server/rest/services/GL_SJ"

# the message will be write into excel
publisher = "苟芳"
publishTime = "2020-10-15"

# control whether the services chinese meaning will be write into the excel file
serChinese_Key = True
serChinese_NameMappingFile = r"./服务中英文对照.json"

# the result of services detected
outputXlsx = r"./三维.xlsx"

if __name__ == "__main__":
    if serChinese_Key:
        main(serTarName, baseurl, serChinese_Key, outputXlsx, serChinese_NameMappingFile)
    else:
        main(serTarName, baseurl, serChinese_Key, outputXlsx)
    print("程序运行完成")
